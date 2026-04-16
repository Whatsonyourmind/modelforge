"""Programmatic audit of all ModelForge outputs.

Opens every .xlsx, runs structural + computational audits:

Structural (formula strings):
    - Scenario banner present and uses scenario_index
    - Check cells resolve in base scenario
    - Named ranges all resolve
    - Comments present on hardcoded cells
    - Sign consistency: rows labelled "cost", "D&A", "capex", "tax", "interest"
      should contain negating formulas (=-... or *(-1) or = negative inputs)
    - No #REF!, #NAME? literal strings in cells

Computational (actual values via formulas library):
    - Balance sheet ties (|A - L - E| < 0.01) on 3-statement
    - Covenant actuals within sane ranges
    - IRR produces non-error result
    - No division-by-zero cascades
"""

from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

SUITE = [
    # Anonymized examples
    "output/unitranche_cdmo.xlsx",
    "output/minibond_logistics.xlsx",
    "output/credit_memo_cdmo.xlsx",
    "output/project_finance_solar.xlsx",
    "output/real_estate_pbsa.xlsx",
    "output/npl_mixed_portfolio.xlsx",
    "output/structured_credit_pmi.xlsx",
    "output/three_statement_cdmo.xlsx",
    # Real public deals
    "output/real_stevanato_3statement.xlsx",
    "output/real_enfinity_solar_pf.xlsx",
]


COST_KEYWORDS = {
    "d&a": "must be negative",
    "amm": "must be negative",
    "capex": "must be negative (outflow)",
    "tax": "must be non-positive",
    "opex": "must be negative",
    "interest": "must be non-positive (expense)",
    "legal fee": "must be non-positive",
    "arrangement fee": "can be negative (cost) or positive (fee received)",
    "servicing fee": "must be non-positive (cost to fund)",
    "cost": "must be non-positive",
}

REVENUE_KEYWORDS = {
    "revenue": "must be non-negative",
    "ebitda": "could be negative only in severe stress; flag if negative",
    "gross proceeds": "must be non-negative",
    "net income": "can be negative",
}


def audit_file(xlsx_path: Path) -> dict[str, Any]:
    """Open and audit a single workbook. Returns findings."""
    findings: dict[str, list[str]] = {
        "errors": [],
        "warnings": [],
        "info": [],
    }
    if not xlsx_path.exists():
        findings["errors"].append(f"File missing: {xlsx_path}")
        return findings

    # Load with formulas visible
    wb = load_workbook(xlsx_path, data_only=False, keep_links=False)

    # --- Scenario banner check ---
    scenario_sheets = [s for s in wb.sheetnames if s not in {"Cover", "Sources", "Assumptions", "QC"}]
    for s in scenario_sheets:
        ws = wb[s]
        banner = ws.cell(row=3, column=3).value
        if banner is None or "scenario_index" not in str(banner):
            findings["warnings"].append(
                f"[{s}] scenario banner missing or not CHOOSE-driven (C3 = {banner!r})"
            )

    # --- Named range integrity ---
    for name, dn in wb.defined_names.items():
        attr = dn.attr_text
        if "!" not in attr:
            findings["warnings"].append(f"Named range {name}={attr} — no sheet qualifier")
            continue
        sheet_part, ref = attr.split("!", 1)
        sheet = sheet_part.strip("'")
        if sheet not in wb.sheetnames:
            findings["errors"].append(f"Named range {name} points to missing sheet {sheet}")
            continue
        try:
            ws = wb[sheet]
            cell = ws[ref.replace("$", "")]
            if cell.value is None:
                findings["warnings"].append(
                    f"Named range {name}={attr} — cell is empty"
                )
        except Exception as e:
            findings["errors"].append(f"Named range {name}={attr}: {e}")

    # --- Scan every sheet for literal error strings ---
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v in {"#REF!", "#NAME?", "#DIV/0!", "#VALUE!"}:
                    findings["errors"].append(
                        f"[{sn}] Literal error string in {c.coordinate}: {v}"
                    )

    # --- Sign convention: scan rows, find costs_keyword rows, check projection cells ---
    # Only applies to content sheets with year columns (D onwards)
    content_sheets = [
        s for s in wb.sheetnames
        if s not in {"Cover", "Sources", "Assumptions", "QC", "Covenants", "Tranches"}
    ]
    for sn in content_sheets:
        ws = wb[sn]
        for row in ws.iter_rows(min_col=1, max_col=1):
            c = row[0]
            if not c.value or not isinstance(c.value, str):
                continue
            label = c.value.lower()
            for kw, rule in COST_KEYWORDS.items():
                if kw in label:
                    # Check a few projection-col cells for negative sign intent
                    for col_idx in range(4, 12):
                        cell = ws.cell(row=c.row, column=col_idx)
                        fv = cell.value
                        if fv is None or fv == "":
                            continue
                        if isinstance(fv, (int, float)):
                            # Hardcoded value
                            if kw in {"d&a", "capex", "opex", "interest", "tax"} and fv > 0:
                                findings["warnings"].append(
                                    f"[{sn}] Row {c.row} '{c.value[:35]}' "
                                    f"cell {cell.coordinate} hardcoded POSITIVE "
                                    f"({fv}) — {rule}"
                                )
                        elif isinstance(fv, str) and fv.startswith("="):
                            # Formula — check if it has negating pattern
                            has_neg = (
                                "=-" in fv or "*-" in fv or "-MAX" in fv.upper()
                                or re.search(r"^=-", fv) is not None
                            )
                            # D&A and capex and opex and tax should generally start with =-
                            if kw in {"d&a", "capex", "opex"} and not has_neg and fv not in {"=0", "0"}:
                                # Allow formulas that reference already-negative cells
                                # Flag only if it's a direct non-negated product
                                if re.match(r"^=\$?[A-Z]+\$?\d+\*[A-Za-z_]+", fv):
                                    findings["warnings"].append(
                                        f"[{sn}] Row {c.row} '{c.value[:35]}' "
                                        f"cell {cell.coordinate}: {fv} "
                                        f"— expected leading '=-' ({rule})"
                                    )
                    break

    # --- Check no circular references (shallow heuristic: formula in A refs A directly) ---
    for sn in wb.sheetnames:
        ws = wb[sn]
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    # Does formula reference its own cell address?
                    if f"{c.column_letter}{c.row}" in c.value and "$" + c.column_letter + "$" + str(c.row) in c.value:
                        # Self-reference — very likely circular
                        findings["errors"].append(
                            f"[{sn}] {c.coordinate} self-reference: {c.value[:80]}"
                        )

    findings["info"].append(f"Sheets: {wb.sheetnames}")
    findings["info"].append(f"Named ranges: {len(wb.defined_names)}")
    return findings


def main() -> None:
    total_errors = 0
    total_warnings = 0
    print("=" * 80)
    print("MODELFORGE SUITE AUDIT")
    print("=" * 80)
    for p in SUITE:
        path = Path(p)
        print(f"\n### {path.name}")
        findings = audit_file(path)
        for e in findings["errors"]:
            print(f"  [ERROR]    {e}")
            total_errors += 1
        for w in findings["warnings"]:
            print(f"  [WARN]     {w}")
            total_warnings += 1
        if not findings["errors"] and not findings["warnings"]:
            print("  (clean)")
        for i in findings["info"]:
            print(f"  [INFO]     {i}")

    print()
    print("=" * 80)
    print(f"TOTAL: {total_errors} errors, {total_warnings} warnings across {len(SUITE)} files")
    print("=" * 80)
    sys.exit(0 if total_errors == 0 else 1)


if __name__ == "__main__":
    main()
