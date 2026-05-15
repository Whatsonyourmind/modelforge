"""Cell-by-cell diff between two workbook builds.

Designed for "what changed since the last sign-off?" reviews. Compares
cells across all sheets present in both workbooks, ignoring known-noisy
fields (creation timestamp, modification timestamp).

Returns a WorkbookDiff dataclass with one CellChange per altered cell.
"""
from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


@dataclass(frozen=True)
class CellChange:
    sheet: str
    cell: str
    before: Any
    after: Any


@dataclass
class WorkbookDiff:
    """Diff result between two workbook builds."""
    a_path: str
    b_path: str
    cell_changes: list[CellChange] = field(default_factory=list)
    sheets_added: list[str] = field(default_factory=list)
    sheets_removed: list[str] = field(default_factory=list)
    cells_added: list[CellChange] = field(default_factory=list)
    cells_removed: list[CellChange] = field(default_factory=list)

    @property
    def total_changes(self) -> int:
        return (len(self.cell_changes) + len(self.sheets_added)
                + len(self.sheets_removed) + len(self.cells_added)
                + len(self.cells_removed))

    def summary(self) -> str:
        return (
            f"changed={len(self.cell_changes)} "
            f"added_cells={len(self.cells_added)} "
            f"removed_cells={len(self.cells_removed)} "
            f"added_sheets={len(self.sheets_added)} "
            f"removed_sheets={len(self.sheets_removed)}"
        )


def _get_value(ws, addr: str) -> Any:
    cell = ws[addr]
    if isinstance(cell, tuple):  # range — flatten first cell
        return cell[0][0].value if cell and cell[0] else None
    return cell.value


def _all_cell_coords(ws) -> set[str]:
    coords: set[str] = set()
    for row in ws.iter_rows(values_only=False):
        for c in row:
            if c.value is not None:
                coords.add(c.coordinate)
    return coords


def diff_workbooks(a_path: Path | str, b_path: Path | str,
                   *, ignore_sheets: Optional[set[str]] = None) -> WorkbookDiff:
    """Compute a WorkbookDiff between two .xlsx files.

    By default ignores no sheets. Pass ``ignore_sheets={"Reproducibility"}``
    to skip volatile metadata sheets that are expected to differ.
    """
    a = load_workbook(a_path, data_only=False, keep_links=True)
    b = load_workbook(b_path, data_only=False, keep_links=True)
    ignore = ignore_sheets or set()

    a_sheets = {s for s in a.sheetnames if s not in ignore}
    b_sheets = {s for s in b.sheetnames if s not in ignore}

    diff = WorkbookDiff(a_path=str(a_path), b_path=str(b_path))
    diff.sheets_added = sorted(b_sheets - a_sheets)
    diff.sheets_removed = sorted(a_sheets - b_sheets)

    common = sorted(a_sheets & b_sheets)
    for s in common:
        ws_a = a[s]
        ws_b = b[s]
        coords_a = _all_cell_coords(ws_a)
        coords_b = _all_cell_coords(ws_b)
        common_cells = coords_a & coords_b
        added_cells = coords_b - coords_a
        removed_cells = coords_a - coords_b
        for addr in sorted(common_cells):
            va = _get_value(ws_a, addr)
            vb = _get_value(ws_b, addr)
            if va != vb:
                diff.cell_changes.append(CellChange(s, addr, va, vb))
        for addr in sorted(added_cells):
            diff.cells_added.append(CellChange(s, addr, None, _get_value(ws_b, addr)))
        for addr in sorted(removed_cells):
            diff.cells_removed.append(CellChange(s, addr, _get_value(ws_a, addr), None))
    return diff
