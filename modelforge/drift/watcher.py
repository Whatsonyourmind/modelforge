"""Drift detection engine.

Mapping strategy: each driver name (e.g. `euribor_6m_rate`) is tied to
a feed accessor (e.g. `ECBFeed.euribor_6m`). When both are present,
drift = current − assumed. Drift items are flagged when:

    * |delta| ≥ threshold_bps (absolute bps, for rate-style drivers)
    * |delta / assumed| ≥ threshold_rel (relative, for value drivers)

Both thresholds configurable per call; sensible defaults (50bps /
10%) baked in.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, Optional

from openpyxl import load_workbook

from modelforge.feeds import DamodaranFeed, ECBFeed


# ── Driver → (feed accessor, kind) map ──────────────────────────────────────

# "kind" is informational: "rate" drivers are compared in bps; "value"
# drivers relative. Extend as new feeds land.

def _ecb_accessor(attr: str) -> Callable[[], float]:
    def _acc() -> float:
        return getattr(ECBFeed.load(), attr)
    return _acc


def _damodaran_accessor(iso2: str, method: str = "country_erp") -> Callable[[], float]:
    def _acc() -> float:
        dam = DamodaranFeed.load()
        return getattr(dam, method)(iso2)
    return _acc


DRIVER_FEED_MAP: dict[str, tuple[Callable[[], float], str, str]] = {
    # Italian debt rate drivers
    "euribor_3m_rate":      (_ecb_accessor("euribor_3m"), "rate", "ECB SDW"),
    "euribor_3m_pct":       (_ecb_accessor("euribor_3m"), "rate", "ECB SDW"),
    "euribor_6m_rate":      (_ecb_accessor("euribor_6m"), "rate", "ECB SDW"),
    "euribor_6m_pct":       (_ecb_accessor("euribor_6m"), "rate", "ECB SDW"),
    "euribor_12m_rate":     (_ecb_accessor("euribor_12m"), "rate", "ECB SDW"),
    "ecb_main_refi":        (_ecb_accessor("ecb_main_refi"), "rate", "ECB SDW"),
    "ecb_refi_rate":        (_ecb_accessor("ecb_main_refi"), "rate", "ECB SDW"),
    "eur_swap_10y":         (_ecb_accessor("euribor_12m"), "rate",
                             "ECB SDW (proxy — 12M EURIBOR)"),
    # Damodaran Italy-centric
    "equity_risk_premium":  (_damodaran_accessor("IT"), "rate",
                             "Damodaran country risk table (Italy)"),
    "risk_free_rate":       (_ecb_accessor("ecb_main_refi"), "rate",
                             "ECB SDW (proxy — main refi rate)"),
}


# ── Data classes ────────────────────────────────────────────────────────────


@dataclass
class DriftItem:
    driver_name: str
    assumed_value: float
    current_value: float
    delta_abs: float
    delta_bps: float         # 10_000 × delta (meaningful for rates)
    delta_rel: float         # delta / assumed (meaningful for non-rate values)
    source: str
    flagged: bool
    kind: str                # "rate" or "value"


@dataclass
class DriftReport:
    xlsx_path: Path
    checked_drivers: int
    missing_drivers: list[str] = field(default_factory=list)
    items: list[DriftItem] = field(default_factory=list)

    @property
    def flagged(self) -> list[DriftItem]:
        return [i for i in self.items if i.flagged]

    @property
    def n_flagged(self) -> int:
        return len(self.flagged)

    @property
    def clean(self) -> bool:
        return self.n_flagged == 0


# ── Core check ──────────────────────────────────────────────────────────────


def _read_named_range_value(wb, name: str) -> Optional[float]:
    """Read the Active (col I) cell behind a driver's named range.

    Named ranges in ModelForge point at 'Assumptions'!$I$N. We read
    the BASE value (col G of the same row) since that's the
    user-editable input — not the Active formula output.
    """
    if name not in wb.defined_names:
        return None
    dn = wb.defined_names[name]
    attr = dn.attr_text  # e.g. 'Assumptions'!$I$12
    if "!" not in attr:
        return None
    sheet_part, ref = attr.split("!", 1)
    sheet = sheet_part.strip("'")
    if sheet not in wb.sheetnames:
        return None
    # Strip $ signs, parse col + row
    ref_clean = ref.replace("$", "")
    # Extract col letter + row number
    col_letters = "".join(c for c in ref_clean if c.isalpha())
    row_part = "".join(c for c in ref_clean if c.isdigit())
    if not row_part:
        return None
    row = int(row_part)
    # Move from col I (active) to col G (base)
    ws = wb[sheet]
    base_cell = ws[f"G{row}"]
    v = base_cell.value
    return float(v) if isinstance(v, (int, float)) else None


def check_drift(
    xlsx_path: Path | str,
    threshold_bps: float = 50.0,
    threshold_rel: float = 0.10,
    feed_map: dict[str, tuple] | None = None,
) -> DriftReport:
    """Compare workbook drivers to current feed values.

    Parameters
    ----------
    threshold_bps : float
        Absolute bps delta that triggers a flag for rate-style drivers.
        Default 50bps (a half-percent move on any rate is "material").
    threshold_rel : float
        Relative delta that triggers for non-rate drivers. Default 10%.
    feed_map : Optional override
        Inject a custom driver→feed map. Defaults to DRIVER_FEED_MAP.
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    m = feed_map or DRIVER_FEED_MAP

    items: list[DriftItem] = []
    missing: list[str] = []

    for driver_name, (accessor, kind, source) in m.items():
        assumed = _read_named_range_value(wb, driver_name)
        if assumed is None:
            missing.append(driver_name)
            continue
        try:
            current = float(accessor())
        except Exception:
            missing.append(driver_name)
            continue
        delta_abs = current - assumed
        delta_bps = delta_abs * 10_000.0
        delta_rel = (delta_abs / assumed) if assumed != 0 else 0.0
        if kind == "rate":
            flagged = abs(delta_bps) >= threshold_bps
        else:
            flagged = abs(delta_rel) >= threshold_rel
        items.append(DriftItem(
            driver_name=driver_name,
            assumed_value=assumed,
            current_value=current,
            delta_abs=delta_abs,
            delta_bps=delta_bps,
            delta_rel=delta_rel,
            source=source,
            flagged=flagged,
            kind=kind,
        ))

    return DriftReport(
        xlsx_path=xlsx_path,
        checked_drivers=len(items),
        missing_drivers=missing,
        items=items,
    )


# ── Rendering ───────────────────────────────────────────────────────────────


# ── Portfolio-level sweep ───────────────────────────────────────────────────


@dataclass
class PortfolioDriftReport:
    folder: Path
    per_workbook: list[DriftReport] = field(default_factory=list)

    @property
    def n_workbooks(self) -> int:
        return len(self.per_workbook)

    @property
    def n_flagged_workbooks(self) -> int:
        return sum(1 for r in self.per_workbook if r.n_flagged > 0)

    @property
    def total_flags(self) -> int:
        return sum(r.n_flagged for r in self.per_workbook)

    @property
    def clean(self) -> bool:
        return self.total_flags == 0


def check_portfolio(
    folder: Path | str,
    threshold_bps: float = 50.0,
    threshold_rel: float = 0.10,
    glob_pattern: str = "*.xlsx",
) -> PortfolioDriftReport:
    """Run drift check on every workbook matching glob_pattern in folder.

    Recursive via `**/*.xlsx` supported.
    """
    folder = Path(folder)
    if not folder.is_dir():
        raise ValueError(f"{folder} is not a directory")
    rep = PortfolioDriftReport(folder=folder)
    for xlsx in sorted(folder.glob(glob_pattern)):
        # Skip dossier PDFs + autosave shadows
        if xlsx.suffix.lower() != ".xlsx":
            continue
        try:
            one = check_drift(xlsx, threshold_bps=threshold_bps,
                              threshold_rel=threshold_rel)
            rep.per_workbook.append(one)
        except Exception:
            # Skip unreadable workbooks gracefully
            continue
    return rep


def render_portfolio_markdown(rep: PortfolioDriftReport) -> str:
    out = [f"# Portfolio drift report — {rep.folder}", ""]
    out.append(f"**{rep.n_workbooks} workbook(s) scanned · "
               f"{rep.n_flagged_workbooks} with flagged drivers · "
               f"{rep.total_flags} total flags.**")
    out.append("")

    if not rep.per_workbook:
        out.append("_(no .xlsx files matched)_")
        return "\n".join(out)

    out.append("| Workbook | Checked | Flagged | Top driver drift (bps) |")
    out.append("|---|---|---|---|")
    for r in rep.per_workbook:
        top = max(r.flagged, key=lambda i: abs(i.delta_bps), default=None)
        top_str = (f"`{top.driver_name}` {top.delta_bps:+,.0f}bps"
                   if top else "—")
        out.append(
            f"| {r.xlsx_path.name} | {r.checked_drivers} | "
            f"{r.n_flagged} | {top_str} |"
        )
    out.append("")

    # Per-workbook detail (only the flagged rows)
    for r in rep.per_workbook:
        if not r.flagged:
            continue
        out.append(f"## {r.xlsx_path.name} — flagged drivers")
        out.append("")
        out.append("| Driver | Assumed | Current | Δ bps | Source |")
        out.append("|---|---|---|---|---|")
        for i in r.flagged:
            assumed_f = (f"{i.assumed_value:.4%}" if i.kind == "rate"
                         else f"{i.assumed_value:,.3f}")
            current_f = (f"{i.current_value:.4%}" if i.kind == "rate"
                         else f"{i.current_value:,.3f}")
            out.append(f"| `{i.driver_name}` | {assumed_f} | {current_f} "
                       f"| {i.delta_bps:+,.1f} | {i.source} |")
        out.append("")

    return "\n".join(out)


def render_markdown(rep: DriftReport) -> str:
    out = [f"# Drift report — {rep.xlsx_path.name}", ""]
    if rep.clean:
        out.append(f"**Clean — {rep.checked_drivers} drivers checked, "
                   f"none flagged above threshold.**")
    else:
        out.append(f"**{rep.n_flagged} of {rep.checked_drivers} drivers "
                   f"flagged.**")
    out.append("")

    if rep.items:
        out.append("## Checked drivers")
        out.append("")
        out.append("| Driver | Assumed | Current | Δ bps | Δ % | Source | Flag |")
        out.append("|---|---|---|---|---|---|---|")
        for it in rep.items:
            assumed_f = f"{it.assumed_value:.4%}" if it.kind == "rate" else f"{it.assumed_value:,.3f}"
            current_f = f"{it.current_value:.4%}" if it.kind == "rate" else f"{it.current_value:,.3f}"
            flag = "⚠ FLAG" if it.flagged else "OK"
            out.append(
                f"| `{it.driver_name}` | {assumed_f} | {current_f} "
                f"| {it.delta_bps:+,.1f} | {it.delta_rel:+.2%} "
                f"| {it.source} | {flag} |"
            )
        out.append("")

    return "\n".join(out)
