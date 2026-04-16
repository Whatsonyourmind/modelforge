"""Sources sheet.

Full provenance registry. Every hardcoded number in the workbook has a
cell comment with an S-id; this sheet is where those IDs resolve to
doc + page + publisher + URL + verification flag.

The layout:
    A  ID (S-001)
    B  Document
    C  Page
    D  Publisher
    E  Date
    F  URL (hyperlink)
    G  Verified
    H  Note

Every row is also added to the linkage graph as a SOURCE node with a
CITES edge to its DOC_PAGE.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles
from modelforge.builder.i18n import L
from modelforge.graph.schema import LinkageGraph, GraphNode, NodeKind
from modelforge.spec.unitranche import UnitrancheSpec


def build(ws: Worksheet, spec: UnitrancheSpec, graph: LinkageGraph) -> dict[str, int]:
    """Write Sources sheet. Returns {source_id: row_number} for cross-reference."""
    # Column widths
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 40

    # Title
    ws.cell(row=1, column=1, value="Sources").font = styles.font_title
    ws.cell(row=2, column=1, value="Fonti").font = styles.font_label_it
    ws.cell(
        row=3,
        column=1,
        value="Every hardcoded number in this model traces back to a row below. "
        "Comment on the cell shows the ID; look it up here.",
    ).font = styles.font_label_it

    # Header row
    header_row = 5
    headers = [
        ("source_id", 10),
        ("source_doc", 38),
        ("source_page", 8),
        ("source_publisher", 22),
        ("source_date", 12),
        ("source_url", 40),
        ("source_verified", 10),
        ("source_note", 40),
    ]
    for i, (key, _) in enumerate(headers):
        lbl = L(key)
        c = ws.cell(row=header_row, column=i + 1, value=f"{lbl.en}")
        styles.style_header(c)
        # IT label on row above
        it_cell = ws.cell(row=header_row - 1, column=i + 1, value=lbl.it)
        it_cell.font = styles.font_label_it
        it_cell.alignment = styles.align_center

    # Data rows
    source_rows: dict[str, int] = {}
    r = header_row + 1
    for s in spec.sources:
        ws.cell(row=r, column=1, value=s.id).font = styles.font_subheader
        ws.cell(row=r, column=2, value=s.doc).font = styles.font_label_en
        if s.page is not None:
            p = ws.cell(row=r, column=3, value=s.page)
            styles.style_input(p, number_format=styles.FMT_INTEGER)
        ws.cell(row=r, column=4, value=s.publisher).font = styles.font_label_en
        d = ws.cell(row=r, column=5, value=s.date.isoformat())
        styles.style_input(d, number_format=styles.FMT_DATE)
        if s.url:
            u = ws.cell(row=r, column=6, value=s.url)
            u.hyperlink = s.url
            u.font = styles.Font(
                name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
                color="0563C1", underline="single",
            )
        v_cell = ws.cell(row=r, column=7, value="✔" if s.verified else "")
        v_cell.alignment = styles.align_center
        v_cell.font = (
            styles.Font(name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
                        bold=True, color="008000")
            if s.verified
            else styles.Font(name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
                             color="C00000")
        )
        ws.cell(row=r, column=8, value=s.note).font = styles.font_label_en

        source_rows[s.id] = r

        # Graph: SOURCE node + DOC_PAGE node + CITES edge
        src_node_id = s.id  # "S-001"
        graph.add_node(
            GraphNode(
                id=src_node_id,
                kind=NodeKind.SOURCE,
                label=f"{s.doc} (p.{s.page or '-'})",
                payload={
                    "doc": s.doc, "page": s.page, "publisher": s.publisher,
                    "date": s.date.isoformat(), "url": s.url, "verified": s.verified,
                    "note": s.note,
                },
            )
        )
        if s.page is not None:
            dp_id = LinkageGraph.doc_page_id(s.doc, s.page)
            graph.add_node(
                GraphNode(
                    id=dp_id,
                    kind=NodeKind.DOC_PAGE,
                    label=f"{s.doc}#p{s.page}",
                    payload={"doc": s.doc, "page": s.page},
                )
            )
            graph.cite(src_node_id, dp_id)

        r += 1

    ws.freeze_panes = "A6"
    ws.print_title_rows = "4:5"
    return source_rows
