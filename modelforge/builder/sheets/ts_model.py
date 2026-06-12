"""3-Statement model sheet — P&L + BS + CFS integrated."""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.i18n import L


def _register_name(wb, name: str, addr: str) -> None:
    """Register (or replace) a workbook-level named range pointing at addr."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=addr)


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="3-Statement Model", title_it="Modello a tre prospetti",
        subtitle=f"P&L · BS · CFS integrated · {spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    # Year headers
    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        is_hist = i < h
        cc = ws.cell(row=yr_row, column=col_idx, value=f"{'A' if is_hist else 'E'} {yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    # ── P&L ──────────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, L("ts_pnl_header").en, L("ts_pnl_header").secondary)
    r += 1

    rows["revenue_growth"] = r
    layout.write_row_label(ws, r, L("revenue_growth").en, L("revenue_growth").secondary)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(p):
        a = spec.pl.revenue_growth_by_year[i]
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(cc, number_format=styles.FMT_PCT)
    r += 1

    rows["revenue"] = r
    layout.write_row_label(ws, r, L("revenue").en, L("revenue").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=spec.historical_revenue_eur_m[i])
        styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(p):
        col = layout.year_col(h + i); col_idx = ord(col) - ord("A") + 1
        prior_col = layout.year_col(h + i - 1)
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${prior_col}${r}*(1+${col}${rows['revenue_growth']})")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    # Historical growth formulas
    for i in range(1, h):
        col = layout.year_col(i); prior_col = layout.year_col(i - 1)
        col_idx = ord(col) - ord("A") + 1
        gc = ws.cell(row=rows["revenue_growth"], column=col_idx,
                     value=f"=IFERROR(${col}${rows['revenue']}/${prior_col}${rows['revenue']}-1,0)")
        styles.style_formula(gc, number_format=styles.FMT_PCT)
    r += 1

    rows["ebitda_margin"] = r
    layout.write_row_label(ws, r, L("ebitda_margin").en, L("ebitda_margin").secondary)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        hrev = spec.historical_revenue_eur_m[i]
        val = spec.historical_ebitda_eur_m[i] / hrev if hrev else 0
        cc = ws.cell(row=r, column=col_idx, value=val)
        styles.style_input(cc, number_format=styles.FMT_PCT)
    for i in range(p):
        a = spec.pl.ebitda_margin_by_year[i]
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(cc, number_format=styles.FMT_PCT)
    r += 1

    rows["ebitda"] = r
    layout.write_row_label(ws, r, L("ebitda").en, L("ebitda").secondary)
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=spec.historical_ebitda_eur_m[i])
        styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(p):
        col = layout.year_col(h + i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['revenue']}*${col}${rows['ebitda_margin']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["da"] = r
    layout.write_row_label(ws, r, L("da").en, L("da").secondary, indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"=-${col}${rows['revenue']}*da_pct_revenue")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ebit"] = r
    layout.write_row_label(ws, r, L("ebit").en, L("ebit").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['ebitda']}+${col}${rows['da']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Interest (cross-ref to BS debt row — which we build below)
    rows["interest"] = r
    layout.write_row_label(ws, r, L("interest_expense").en, L("interest_expense").secondary, indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    r += 1

    rows["ebt"] = r
    layout.write_row_label(ws, r, L("ebt").en, L("ebt").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['ebit']}+${col}${rows['interest']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["tax"] = r
    layout.write_row_label(ws, r, L("tax").en, L("tax").secondary, indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-MAX(${col}${rows['ebt']},0)*effective_tax_rate")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_income"] = r
    layout.write_row_label(ws, r, L("net_income").en, L("net_income").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['ebt']}+${col}${rows['tax']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # ── Balance Sheet ────────────────────────────────────────────────────
    layout.write_section_header(ws, r, L("ts_bs_header").en, L("ts_bs_header").secondary)
    r += 1

    # Assets
    rows["cash"] = r
    layout.write_row_label(ws, r, L("ts_cash").en, L("ts_cash").secondary, indent=True)
    # Cash is the plug — formula later in CFS section; here hardcode opening
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.cash_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["receivables"] = r
    layout.write_row_label(ws, r, L("ts_ar").en, L("ts_ar").secondary, indent=True)
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.receivables_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['revenue']}/365*receivables_days")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["inventory"] = r
    layout.write_row_label(ws, r, L("ts_inventory").en, L("ts_inventory").secondary, indent=True)
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.inventory_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['revenue']}/365*inventory_days")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_ppe"] = r
    layout.write_row_label(ws, r, L("ts_net_ppe").en, L("ts_net_ppe").secondary, indent=True)
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.net_ppe_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${prior}${r}+${col}${rows['revenue']}*capex_pct_revenue+${col}${rows['da']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_assets"] = r
    layout.write_row_label(ws, r, L("ts_total_assets").en, L("ts_total_assets").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['cash']}+${col}${rows['receivables']}+${col}${rows['inventory']}+${col}${rows['net_ppe']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Liabilities & Equity
    rows["payables"] = r
    layout.write_row_label(ws, r, L("ts_ap").en, L("ts_ap").secondary, indent=True)
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.payables_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['revenue']}/365*payables_days")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # v0.7: Debt roll-forward (US-075). BOP debt − scheduled repayment = EOP,
    # floored at 0. Opening balance from spec; repayment = spec
    # debt_annual_repayment_eur_m (default 0 → debt stays flat). The matching
    # cash outflow is booked in CFF below so the balance sheet stays balanced.
    rows["debt"] = r
    layout.write_row_label(ws, r, L("ts_debt").en, L("ts_debt").secondary, indent=True)
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.debt_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    annual_repay = getattr(spec.opening_bs, 'debt_annual_repayment_eur_m', 0.0) or 0.0
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        # Roll-forward: prior balance minus scheduled repayment (not below 0)
        cc = ws.cell(
            row=r, column=col_idx,
            value=f"=MAX(${prior}${r}-{annual_repay},0)",
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["equity"] = r
    layout.write_row_label(ws, r, L("ts_equity").en, L("ts_equity").secondary, indent=True)
    cc = ws.cell(row=r, column=4, value=spec.opening_bs.equity_eur_m)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        # Equity = prior equity + net income - dividends
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${prior}${r}+${col}${rows['net_income']}"
                           f"-MAX(${col}${rows['net_income']},0)*dividend_payout_ratio")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_le"] = r
    layout.write_row_label(ws, r, L("ts_total_le").en, L("ts_total_le").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['payables']}+${col}${rows['debt']}+${col}${rows['equity']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 1

    rows["bs_check"] = r
    layout.write_row_label(ws, r, L("ts_bs_check").en, L("ts_bs_check").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['total_assets']}-${col}${rows['total_le']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Now patch Interest row (=-debt * rate, on opening debt)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        opening_debt_ref = f"${prior}${rows['debt']}"
        int_cell = ws.cell(row=rows["interest"], column=col_idx)
        int_cell.value = f"=-{opening_debt_ref}*interest_on_debt_pct"
        styles.style_formula(int_cell, number_format=styles.FMT_EUR_M)

    # ── Cash Flow Statement ──────────────────────────────────────────────
    layout.write_section_header(ws, r, L("ts_cfs_header").en, L("ts_cfs_header").secondary)
    r += 1

    rows["cf_ni"] = r
    layout.write_row_label(ws, r, L("ts_ni_from_pl").en, L("ts_ni_from_pl").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['net_income']}")
        styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_da_addback"] = r
    layout.write_row_label(ws, r, "+ D&A (non-cash)", "+ Ammortamenti", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['da']}")  # flip sign since da is negative
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_nwc"] = r
    layout.write_row_label(ws, r, "- ΔNWC", "- ΔCCN", indent=True)
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        cc = ws.cell(
            row=r, column=col_idx,
            value=(
                f"=-(${col}${rows['receivables']}-${prior}${rows['receivables']}"
                f"+${col}${rows['inventory']}-${prior}${rows['inventory']}"
                f"-${col}${rows['payables']}+${prior}${rows['payables']})"
            ),
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_cfo"] = r
    layout.write_row_label(ws, r, L("ts_cfo").en, L("ts_cfo").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['cf_ni']}+${col}${rows['cf_da_addback']}+${col}${rows['cf_nwc']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    r += 1

    rows["cf_capex"] = r
    layout.write_row_label(ws, r, L("ts_capex").en, L("ts_capex").secondary, indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['revenue']}*capex_pct_revenue")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_cfi"] = r
    layout.write_row_label(ws, r, L("ts_cfi").en, L("ts_cfi").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"=${col}${rows['cf_capex']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_div"] = r
    layout.write_row_label(ws, r, "- Dividends paid", "- Dividendi", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-MAX(${col}${rows['net_income']},0)*dividend_payout_ratio")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Debt repayment (CFF outflow) = the reduction in the debt balance this
    # period. Booking it here makes the cash plug net the paydown so A = L + E
    # holds when debt_annual_repayment_eur_m > 0 (previously the BS debt fell
    # with no matching cash outflow, breaking the balance-sheet tie).
    rows["cf_debt_repay"] = r
    layout.write_row_label(ws, r, "Debt repayment", "Rimborso debito", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            prior = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-(${prior}${rows['debt']}-${col}${rows['debt']})")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_cff"] = r
    layout.write_row_label(ws, r, L("ts_cff").en, L("ts_cff").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['cf_div']}+${col}${rows['cf_debt_repay']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cf_net_change"] = r
    layout.write_row_label(ws, r, L("ts_net_change_cash").en, L("ts_net_change_cash").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['cf_cfo']}+${col}${rows['cf_cfi']}+${col}${rows['cf_cff']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 1

    # Patch Cash row with t>0 values from CFS
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        cash_cell = ws.cell(row=rows["cash"], column=col_idx)
        cash_cell.value = f"=${prior}${rows['cash']}+${col}${rows['cf_net_change']}"
        styles.style_formula(cash_cell, number_format=styles.FMT_EUR_M)

    r += 1
    # ── v0.8 Supplementary: NOL, DTA/DTL, SBC, Minority, Revolver ────────
    layout.write_section_header(
        ws, r, "Supplementary schedules (NOL, DTA/DTL, SBC, MI, Revolver)",
        "Schedules supplementari",
    )
    r += 1

    # v0.8 US-220: NOL schedule — Italian 5-year limit + 80% current-year
    # offset (post Legge Bilancio 2024). Opening → Generated → Used (cap
    # 80% of positive EBT) → Expired (>5y) → Closing.
    rows["nol_opening"] = r
    layout.write_row_label(ws, r, "NOL balance (opening)",
                           "Perdite fiscali (apertura)", indent=True)
    ws.cell(row=r, column=4, value=0)  # opens at zero (or spec override)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${prior}${r + 4}")  # forward ref to closing below
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["nol_generated"] = r
    layout.write_row_label(ws, r, "NOL generated (NI < 0)",
                           "Perdita generata", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=MAX(-${col}${rows['ebt']},0)")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["nol_used"] = r
    layout.write_row_label(ws, r, "NOL used (cap on positive EBT)",
                           "Perdite utilizzate (cap su EBT positivo)",
                           indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(
            row=r, column=col_idx,
            value=(f"=MIN(${col}${rows['nol_opening']},"
                   f"MAX(${col}${rows['ebt']},0)*nol_util_cap_pct)"),
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["nol_expired"] = r
    layout.write_row_label(ws, r, "NOL expired (>5 years)",
                           "Perdite scadute (>5 anni)", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    r += 1

    rows["nol_closing"] = r
    layout.write_row_label(ws, r, "NOL balance (closing)",
                           "Perdite fiscali (chiusura)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(
            row=r, column=col_idx,
            value=(f"=${col}${rows['nol_opening']}"
                   f"+${col}${rows['nol_generated']}"
                   f"-${col}${rows['nol_used']}"
                   f"-${col}${rows['nol_expired']}"),
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # v0.8 US-221: DTA from NOL + DTL on D&A book-tax difference
    rows["dta"] = r
    layout.write_row_label(ws, r, "DTA (NOL × tax rate)",
                           "Imposte differite attive", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['nol_closing']}*effective_tax_rate")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["dtl"] = r
    layout.write_row_label(ws, r, "DTL (accumulated on D&A timing diffs)",
                           "Imposte differite passive", indent=True)
    # Simplified: DTL accumulates at book_tax_diff_da_pct of |D&A| × tax rate
    # per year as a proxy for book-vs-tax D&A timing differences.
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        cc = ws.cell(
            row=r, column=col_idx,
            value=(f"=${prior}${r}+ABS(${col}${rows['da']})"
                   f"*book_tax_diff_da_pct*effective_tax_rate"),
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # v0.8 US-222: SBC — expense as % revenue (non-cash, CFS addback).
    rows["sbc_expense"] = r
    layout.write_row_label(ws, r, "Stock-based compensation (SBC) expense",
                           "Compensi in azioni (costo)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['revenue']}*sbc_pct_revenue")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "SBC expense = revenue × sbc_pct_revenue (default 1%, see Model "
        "parameters). Non-cash → added back on CFS. Override via "
        "spec.sbc_pct_revenue.",
        "ModelForge",
    )
    r += 1

    # v0.8 US-223: Minority interest — NI attributable to NCI (default 0)
    rows["minority_ni"] = r
    layout.write_row_label(ws, r, "(−) Minority interest in NI",
                           "(−) Interessenze di minoranza su utile",
                           indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)  # default 0 for wholly-owned
    r += 1

    rows["ni_to_parent"] = r
    layout.write_row_label(ws, r, "Net income to parent",
                           "Utile netto di gruppo (parent)")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=(f"=${col}${rows['net_income']}"
                            f"-${col}${rows['minority_ni']}"))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["minority_bs"] = r
    layout.write_row_label(ws, r, "Minority interest (BS equity)",
                           "PN di terzi (SP)", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    r += 1

    # v0.8 US-224: Revolver plug — auto-draw when ending cash < 0.
    # Simple proxy: revolver draw = MAX(0, −cash). Does not iterate; meant
    # as a structural placeholder until the revolver facility is sized.
    rows["revolver_plug"] = r
    layout.write_row_label(ws, r, "Revolver draw (plug when cash < 0)",
                           "Utilizzo revolver (se cassa < 0)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=MAX(-${col}${rows['cash']},0)")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["revolver_commit_fee"] = r
    layout.write_row_label(ws, r, "Revolver commitment fee (× undrawn)",
                           "Commissione impegno revolver", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        # capacity × fee × (1 − utilization). Capacity and fee are named
        # inputs (see Model parameters). Simplified.
        cc = ws.cell(
            row=r, column=col_idx,
            value=(f"=-MAX(revolver_capacity_eur_m-${col}${r-1},0)"
                   f"*revolver_commitment_fee_pct"),
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # ── Model parameters (overridable named inputs) ──────────────────────
    # Each value below was formerly hardcoded in a formula string. It is now
    # a visible, styled input cell in column D, registered as a workbook-level
    # named range, and referenced by the formulas above. Defaults reproduce
    # the prior behaviour exactly; override any of them via the spec.
    r += 1
    wb = ws.parent
    layout.write_section_header(ws, r, "Model parameters",
                                "Parametri del modello")
    r += 1

    params = [
        ("sbc_pct_revenue",
         "SBC (% of revenue)", "Compensi in azioni (% ricavi)",
         getattr(spec, "sbc_pct_revenue", None) if
         getattr(spec, "sbc_pct_revenue", None) is not None else 0.01,
         styles.FMT_PCT,
         "Stock-based compensation as % of revenue (non-cash; CFS add-back). "
         "Default 1%. Override via spec.sbc_pct_revenue."),
        ("revolver_capacity_eur_m",
         "Revolver capacity", "Capacità revolver",
         getattr(spec, "revolver_capacity_eur_m", None) if
         getattr(spec, "revolver_capacity_eur_m", None) is not None else 100.0,
         styles.FMT_EUR_M,
         "Revolver facility capacity. Commitment fee accrues on the undrawn "
         "portion. Default 100m. Override via spec.revolver_capacity_eur_m."),
        ("revolver_commitment_fee_pct",
         "Revolver commitment fee", "Commissione impegno revolver",
         getattr(spec, "revolver_commitment_fee_pct", None) if
         getattr(spec, "revolver_commitment_fee_pct", None) is not None
         else 0.005,
         styles.FMT_PCT,
         "Annual fee on undrawn revolver capacity. Default 0.5%. Override via "
         "spec.revolver_commitment_fee_pct."),
        ("book_tax_diff_da_pct",
         "Book-tax D&A diff (DTL)", "Differenza fiscale su amm.ti (DTL)",
         getattr(spec, "book_tax_diff_da_pct", None) if
         getattr(spec, "book_tax_diff_da_pct", None) is not None else 0.05,
         styles.FMT_PCT,
         "Book-vs-tax D&A timing difference as % of |D&A|; accrues the DTL "
         "each year × tax rate. Default 5%. Override via "
         "spec.book_tax_diff_da_pct."),
        ("nol_util_cap_pct",
         "NOL utilisation cap", "Limite utilizzo perdite (NOL)",
         getattr(spec, "nol_util_cap_pct", None) if
         getattr(spec, "nol_util_cap_pct", None) is not None else 0.80,
         styles.FMT_PCT,
         "NOL carryforward utilisation cap as % of positive EBT (IRC §172 / "
         "EU 80% limitation). Default 80%. Override via spec.nol_util_cap_pct."),
    ]
    for name, lbl_en, lbl_it, value, nfmt, note in params:
        layout.write_row_label(ws, r, lbl_en, lbl_it, indent=True)
        c = ws.cell(row=r, column=4, value=value)
        styles.style_input(c, number_format=nfmt)
        c.comment = Comment(note, "ModelForge")
        _register_name(wb, name, f"'{ws.title}'!$D${r}")
        rows[name] = r
        r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
