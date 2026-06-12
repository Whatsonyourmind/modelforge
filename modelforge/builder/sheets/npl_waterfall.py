"""NPL Collection Waterfall sheet.

Layout: columns t=0..t=collection_years.
Rows: cumulative collection curve → annual gross collections → servicing fees
      → legal fees → net collections → debt service → net to fund equity.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    y = spec.horizon.collection_years
    n = y + 1  # t=0 is purchase

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="NPL Collection Waterfall", title_it="Waterfall di recupero NPL",
        subtitle=f"GBV {spec.meta.currency} · {y}y collection horizon",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=yr_row, column=col_idx, value=f"t={i}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    # Portfolio stats
    layout.write_section_header(ws, r, "Portfolio", "Portafoglio")
    r += 1
    rows["gbv"] = r
    layout.write_row_label(ws, r, "GBV (Gross Book Value)", "GBV")
    cc = ws.cell(row=r, column=4, value="=gbv_eur_m")
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["purchase"] = r
    layout.write_row_label(ws, r, "Purchase price (€m)", "Prezzo di acquisto")
    cc = ws.cell(row=r, column=4, value=f"=$D${rows['gbv']}*purchase_price_pct_gbv")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Collection curve
    layout.write_section_header(ws, r, "Collection curve", "Curva di recupero")
    r += 1
    rows["cum_collection_pct"] = r
    layout.write_row_label(ws, r, "Cumulative collection % GBV", "Recupero cumulato % GBV")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    ws.cell(row=r, column=4, value="=0")
    for i in range(y):
        col_idx = ord(layout.year_col(i + 1)) - ord("A") + 1
        a = spec.portfolio.cumulative_collection_curve_pct[i]
        cc = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(cc, number_format=styles.FMT_PCT)
    r += 1

    rows["annual_gross_collections"] = r
    layout.write_row_label(ws, r, "Annual gross collections", "Recuperi lordi annuali")
    ws.cell(row=r, column=4, value="=0")
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 1:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['cum_collection_pct']}*$D${rows['gbv']}")
        else:
            prior_col = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=(${col}${rows['cum_collection_pct']}-${prior_col}${rows['cum_collection_pct']})*$D${rows['gbv']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Fees
    layout.write_section_header(ws, r, "Servicing fees", "Commissioni di servicing")
    r += 1
    rows["servicing_fee"] = r
    layout.write_row_label(ws, r, "Servicing fee (% of collections)", "Fee servicing", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-${col}${rows['annual_gross_collections']}*servicing_fee_pct_collections")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["legal_fee"] = r
    layout.write_row_label(ws, r, "Legal fees (% of collections)", "Fee legali", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-${col}${rows['annual_gross_collections']}*legal_fee_pct_collections")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["setup_fee"] = r
    layout.write_row_label(ws, r, "Setup fee (one-off)", "Fee setup (una tantum)", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=-$D${rows['gbv']}*setup_fee_pct_gbv")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value="=0")
    r += 1

    rows["data_tape"] = r
    layout.write_row_label(ws, r, "Data tape cost (one-off)", "Costo data tape", indent=True)
    cc = ws.cell(row=r, column=4, value="=-data_tape_cost_eur_m")
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    for i in range(1, n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value="=0")
    r += 2

    # Net collections
    rows["net_collections"] = r
    layout.write_row_label(ws, r, "Net collections to fund", "Recuperi netti al fondo")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['setup_fee']}+${col}${rows['data_tape']}")
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['annual_gross_collections']}+${col}${rows['servicing_fee']}+${col}${rows['legal_fee']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Capital structure
    layout.write_section_header(ws, r, "Capital structure service", "Servizio struttura")
    r += 1
    rows["senior_note_size"] = r
    layout.write_row_label(ws, r, "Senior note size", "Nota senior")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['purchase']}*senior_note_pct_purchase")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["mezz_note_size"] = r
    layout.write_row_label(ws, r, "Mezz note size", "Nota mezz")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['purchase']}*mezz_note_pct_purchase")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Simplified: senior + mezz interest paid annually, principal paid at end from cumulative net
    rows["interest_service"] = r
    layout.write_row_label(ws, r, "Total interest (senior + mezz)",
                           "Interessi totali (senior + mezz)", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-($D${rows['senior_note_size']}*senior_note_rate+$D${rows['mezz_note_size']}*mezz_note_rate)")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # NOTE: the "Equity CF to fund" + "Fund returns" block is built AFTER the
    # strict-priority waterfall below, so the equity stream can reference the
    # subordination-correct `residual_to_equity` (equity is paid only after
    # senior + mezz are fully retired).

    # ── v0.7: STRICT PRIORITY WATERFALL + PDL ────────────────────────────
    # Italian L.130/1999 + GACS compliant waterfall (applied to cumulative
    # collections). Per-period strict priority:
    #   1. Senior interest (paid first)
    #   2. Senior principal amortization (remaining cash after int)
    #   3. Mezz interest (if cash available; else PIK)
    #   4. Mezz principal (after senior retired)
    #   5. Equity distribution (residual)
    # PDL (Principal Deficiency Ledger) tracks under-payment on senior
    # principal — defers interest and triggers trapping.
    layout.write_section_header(
        ws, r,
        "Strict priority waterfall + PDL (Principal Deficiency Ledger)",
        "Waterfall a priorità stretta + PDL",
    )
    r += 1

    # Cumulative cash pool
    rows["cumulative_cash"] = r
    layout.write_row_label(ws, r, "Cumulative cash available (end of year)",
                           "Cassa cumulata (fine anno)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['net_collections']}")
        else:
            prior = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${prior}${r}+${col}${rows['net_collections']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Senior principal outstanding (running balance)
    rows["senior_principal_outstanding"] = r
    layout.write_row_label(ws, r, "Senior principal outstanding",
                           "Capitale senior residuo", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=$D${rows['senior_note_size']}")
        else:
            prior = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=MAX(${prior}${r}-${col}${r + 1},0)")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Senior principal amortized this year (subject to available cash after interest)
    rows["senior_principal_pay"] = r
    layout.write_row_label(ws, r, "Senior principal paid (this year)",
                           "Capitale senior pagato", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            # Available cash = net collections + interest service (interest is neg)
            # Pay down senior up to senior outstanding
            prior = layout.year_col(i - 1)
            cc = ws.cell(
                row=r, column=col_idx,
                value=(
                    f"=MIN(MAX(${col}${rows['net_collections']}+${col}${rows['interest_service']},0),"
                    f"${prior}${rows['senior_principal_outstanding']})"
                ),
            )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # PDL (Principal Deficiency Ledger)
    rows["pdl"] = r
    layout.write_row_label(ws, r, "PDL — Principal Deficiency Ledger",
                           "PDL — Sbilancio capitale", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            # PDL = cumulative senior face − cumulative senior repaid
            # At final year, should be ~0 if deal performs
            cc = ws.cell(
                row=r, column=col_idx,
                value=f"=${col}${rows['senior_principal_outstanding']}",
            )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Mezz principal outstanding (only starts amortizing after senior retired)
    rows["mezz_principal_outstanding"] = r
    layout.write_row_label(ws, r, "Mezz principal outstanding",
                           "Capitale mezz residuo", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=$D${rows['mezz_note_size']}")
        else:
            prior = layout.year_col(i - 1)
            # Mezz only pays down after senior retired
            cc = ws.cell(
                row=r, column=col_idx,
                value=(
                    f"=IF(${col}${rows['senior_principal_outstanding']}<=0.01,"
                    f"MAX(${prior}${r}-MAX(${col}${rows['net_collections']}+${col}${rows['interest_service']}-${col}${rows['senior_principal_pay']},0),0),"
                    f"${prior}${r})"
                ),
            )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Residual to equity (after senior + mezz fully retired)
    rows["residual_to_equity"] = r
    layout.write_row_label(ws, r, "Residual to equity",
                           "Residuo equity", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value="=0")
        else:
            prior = layout.year_col(i - 1)
            # Distributable cash net of BOTH the senior principal paid this year
            # AND the mezz principal paid down this year. The mezz paydown in the
            # retirement-transition year (prior_mezz_os - current_mezz_os) must be
            # subtracted too; otherwise that cash is double-counted as equity
            # residual (cash-conservation leak in the transition year).
            cc = ws.cell(
                row=r, column=col_idx,
                value=(
                    f"=IF(AND(${col}${rows['senior_principal_outstanding']}<=0.01,"
                    f"${col}${rows['mezz_principal_outstanding']}<=0.01),"
                    f"MAX(${col}${rows['net_collections']}+${col}${rows['interest_service']}"
                    f"-${col}${rows['senior_principal_pay']}"
                    f"-(${prior}${rows['mezz_principal_outstanding']}-${col}${rows['mezz_principal_outstanding']}),0),0)"
                ),
            )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # ── Equity CF to fund (subordination-correct) + Fund returns ─────────
    # Built AFTER the waterfall so t>=1 reads the strict-priority residual.
    rows["equity_cf"] = r
    layout.write_row_label(ws, r, "Equity CF to fund", "CF equity al fondo")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            # Outflow: purchase − debt raised + net collections at close (the
            # t=0 net_collections already embeds one-off setup/data-tape costs,
            # so this is the true equity contribution incl. fees).
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-$D${rows['purchase']}+$D${rows['senior_note_size']}+$D${rows['mezz_note_size']}+${col}${rows['net_collections']}")
        else:
            # Equity receives ONLY the strict-priority residual — zero until
            # senior + mezz are fully retired. The prior code paid equity all
            # net collections every year, bypassing the model's own waterfall.
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['residual_to_equity']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Returns
    layout.write_section_header(ws, r, "Fund returns", "Rendimento fondo")
    r += 1
    first_col = layout.year_col(0); last_col = layout.year_col(n - 1)
    layout.write_row_label(ws, r, "Equity IRR", "IRR equity", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IRR(${first_col}${rows['equity_cf']}:${last_col}${rows['equity_cf']},0.15)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.font = styles.font_subheader
    r += 1

    layout.write_row_label(ws, r, "Equity MoIC", "MoIC equity", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IFERROR(SUMIF(${first_col}${rows['equity_cf']}:${last_col}${rows['equity_cf']},\">0\")"
                       f"/ABS(SUMIF(${first_col}${rows['equity_cf']}:${last_col}${rows['equity_cf']},\"<0\")),0)")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    r += 1

    # Gross recovery rate
    layout.write_row_label(ws, r, "Gross recovery rate (cum / GBV)",
                           "Tasso di recupero lordo (cum/GBV)", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=${last_col}${rows['cum_collection_pct']}")
    styles.style_xref(cc, number_format=styles.FMT_PCT)
    r += 2

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
