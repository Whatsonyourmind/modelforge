"""Bank capital return — MDA-gated dividends and buybacks.

Distributions are sized off PRE-DISTRIBUTION CET1 = prior-period CET1 + current
retained earnings (NI attributable to CET1). Because the constraint uses the
PRIOR period's CET1 (a known value) plus current earnings — never the current
closing equity it ultimately feeds — there is no circular reference.

    floor_ratio        = MAX(target CET1, requirement + mgmt buffer)
    MDA cap            = MAX(pre-dist CET1 − floor_ratio × RWA, 0)
    dividend           = MIN(payout × max(NI,0), MDA cap)
    buyback            = MIN(MDA cap − dividend, buyback target)

so total distributions can never push the closing CET1 ratio below the floor.
The dividend and buyback rows are consumed (negated) by the BalanceSheet
common-equity walk via the orchestrator's patch-back.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, capital_refs: dict[str, str], pnl_refs: dict[str, str],
          capital_sheet: str, pnl_sheet: str) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p
    cur = spec.meta.currency

    cet1 = int(capital_refs["cet1_row"])
    rwa = int(capital_refs["rwa_row"])
    ni = int(pnl_refs["ni_row"])
    ni_to_cet1 = int(pnl_refs["ni_to_cet1_row"])

    layout.set_column_widths(ws, label_width=48, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Capital Return (MDA-gated)", title_it="Distribuzione capitale (vincolo MDA)",
        subtitle=f"Dividend + buyback throttled to the CET1 floor · "
                 f"{spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        cc = ws.cell(row=yr_row, column=ci, value=f"{'A' if i < h else 'E'} {yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    def _cap(col, row):
        return f"'{capital_sheet}'!${col}${row}"

    def _pnl(col, row):
        return f"'{pnl_sheet}'!${col}${row}"

    layout.write_section_header(ws, r, "Distribution capacity", "Capacità di distribuzione")
    r += 1

    # Pre-distribution CET1 = prior-period CET1 + current NI-to-CET1.
    rows["pre_dist_cet1"] = r
    layout.write_row_label(ws, r, "Pre-distribution CET1", "CET1 pre-distribuzione", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci,
                        value=f"={_cap(prior, cet1)}+{_pnl(col, ni_to_cet1)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Floor ratio = MAX(target, requirement + buffer)
    rows["floor_ratio"] = r
    layout.write_row_label(ws, r, "Distribution floor ratio", "Soglia minima distribuzione", indent=True)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value="=MAX(target_cet1_ratio,cet1_requirement_ratio+mda_buffer_pct)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    # Min CET1 to hold = floor_ratio × RWA
    rows["min_cet1_hold"] = r
    layout.write_row_label(ws, r, "Minimum CET1 to hold", "CET1 minimo da detenere", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['floor_ratio']}*{_cap(col, rwa)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # MDA cap = MAX(pre-dist CET1 − min hold, 0)
    rows["mda_cap"] = r
    layout.write_row_label(ws, r, "Maximum distributable amount (MDA cap)", "Importo massimo distribuibile")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=MAX(${col}${rows['pre_dist_cet1']}-${col}${rows['min_cet1_hold']},0)")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 2

    # Intended dividend = payout × max(NI, 0)
    rows["intended_dividend"] = r
    layout.write_row_label(ws, r, "Intended dividend (payout × NI)", "Dividendo teorico", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            c = ws.cell(row=r, column=ci,
                        value=f"=MAX({_pnl(col, ni)},0)*dividend_payout_ratio")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Dividend (constrained) = MIN(intended, MDA cap)
    rows["dividend"] = r
    layout.write_row_label(ws, r, "Dividend (MDA-constrained)", "Dividendo (vincolato MDA)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            c = ws.cell(row=r, column=ci,
                        value=f"=MIN(${col}${rows['intended_dividend']},${col}${rows['mda_cap']})")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 1

    # Residual capacity after dividend
    rows["residual_capacity"] = r
    layout.write_row_label(ws, r, "Residual capacity after dividend", "Capacità residua post-dividendo", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['mda_cap']}-${col}${rows['dividend']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Buyback = MIN(residual, target)
    rows["buyback"] = r
    layout.write_row_label(ws, r, "Buyback (residual capacity)", "Riacquisti (capacità residua)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            c = ws.cell(row=r, column=ci,
                        value=f"=MIN(${col}${rows['residual_capacity']},buyback_target_eur_m)")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 1

    # Total distributions
    rows["total_distributions"] = r
    layout.write_row_label(ws, r, "Total distributions", "Distribuzioni totali")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['dividend']}+${col}${rows['buyback']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # Resulting CET1 ratio (must be ≥ floor)
    rows["resulting_cet1_ratio"] = r
    layout.write_row_label(ws, r, "Resulting CET1 ratio (post-distribution)", "Coefficiente CET1 risultante")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value="")
        else:
            c = ws.cell(row=r, column=ci,
                        value=(f"=IFERROR((${col}${rows['pre_dist_cet1']}"
                               f"-${col}${rows['total_distributions']})/{_cap(col, rwa)},0)"))
            styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    # Payout achieved
    rows["payout_achieved"] = r
    layout.write_row_label(ws, r, "Payout achieved (dist / NI)", "Payout effettivo", indent=True)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value="")
        else:
            c = ws.cell(row=r, column=ci,
                        value=f"=IFERROR(${col}${rows['total_distributions']}/MAX({_pnl(col, ni)},0),0)")
            styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {f"{k}_row": str(v) for k, v in rows.items()}
