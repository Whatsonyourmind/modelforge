"""Liability waterfall sheet — Template 19.

Consumes the LoanTape pool cashflows through a sequential-pay (turbo) cash
waterfall:

    gross interest
      − servicing fee  (on performing BOP)         ┐ top of waterfall
      − trustee / senior fees                       ┘ (trustee gated to pool>0)
      = available interest, topped up by a reserve draw on a senior shortfall
      → senior coupon → mezz coupon → … (sequential, BOP-balance accrual,
        each floored at zero — a note can never pay a negative coupon)
      = residual interest (floored at zero)
    principal collected
      → senior → mezz → … (sequential, capped at each note's BOP balance)
      = principal to the residual certificate
    OC / IC triggers: a breach diverts residual interest as a TURBO paydown of
      the most-senior note STILL OUTSTANDING (senior → mezz → …), so excess
      spread de-levers the impaired debt stack instead of leaking to equity.
    reserve account: drawn on a senior interest shortfall; at maturity the
      remaining reserve first CURES any outstanding debt (senior → mezz → …)
      and only the leftover returns to the residual as return-of-capital.
    residual interest (net of turbo & SPV tax) + residual principal + the
      reserve return → the first-loss residual / equity certificate

Acyclic by construction: note interest accrues on the prior-period CLOSING
balance (BOP), sequential allocations read current-period collections plus
prior-period balances, and the turbo / reserve cures are funded from the
already-computed residual interest / reserve balance — no same-period cell
feeds back on itself. The full row map is allocated up front so the
note-balance roll-forwards (placed near the top for readability) can
forward-reference the paydown rows below them.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws, spec, lt_refs: dict[str, str], lt_sheet: str) -> dict[str, str]:
    periods = spec.horizon.periods
    n = periods + 1
    cur = spec.meta.currency

    notes = spec.notes
    debt = list(range(len(notes) - 1))   # indices of debt notes (senior → …)
    eq = len(notes) - 1                   # index of the residual / equity note
    senior = 0

    def lt(col, row_key):
        return f"'{lt_sheet}'!${col}${lt_refs[row_key]}"

    # ── Allocate the full row map up front ─────────────────────────────────
    R: dict[str, int] = {}
    headers: list[tuple[int, str, str]] = []
    r = 7

    def hdr(en, it):
        nonlocal r
        headers.append((r, en, it)); r += 1

    def alloc(key):
        nonlocal r
        R[key] = r; r += 1

    hdr("Available cash", "Cassa disponibile")
    alloc("gross_int"); alloc("servicing"); alloc("trustee")
    alloc("avail_int_after_fees"); alloc("principal_collected")
    r += 1

    hdr("Note balances — outstanding (EOP)", "Saldo note (chiusura)")
    for j in range(len(notes)):
        alloc(f"eop_{j}")
    r += 1

    hdr("Interest waterfall (sequential, BOP coupon)", "Waterfall interessi")
    for j in debt:
        alloc(f"int_due_{j}")
    alloc("reserve_draw")
    alloc("avail_int_total")
    for j in debt:
        alloc(f"int_paid_{j}")
    alloc("residual_int")
    r += 1

    hdr("Principal waterfall (sequential, from collections)", "Waterfall capitale")
    for j in debt:
        alloc(f"prin_paid_{j}")
    alloc("prin_to_residual")
    r += 1

    hdr("Credit enhancement — OC / IC triggers + sequential turbo",
        "Supporto credito — trigger OC/IC + turbo")
    alloc("oc_ratio"); alloc("oc_pass"); alloc("ic_ratio"); alloc("ic_pass")
    for j in debt:
        alloc(f"turbo_{j}")
    r += 1

    hdr("Reserve account (cures debt before returning to equity)", "Conto di riserva")
    alloc("reserve_eop"); alloc("reserve_release")
    for j in debt:
        alloc(f"reserve_cure_{j}")
    alloc("reserve_to_equity")
    r += 1

    hdr("Residual / equity distribution", "Distribuzione residuale / equity")
    alloc("residual_int_to_equity"); alloc("spv_tax"); alloc("cash_to_equity")
    r += 1

    hdr("Investor cashflows (per note)", "Flussi di cassa investitori")
    for j in range(len(notes)):
        alloc(f"inv_cf_{j}")
    last_row = r

    # ── Sheet chrome ───────────────────────────────────────────────────────
    layout.set_column_widths(ws, label_width=48, it_width=30, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Liability Waterfall (sequential-pay)",
        title_it="Waterfall passività (sequenziale)",
        subtitle=f"{cur} · {len(debt)} debt notes + residual · OC/IC turbo · reserve",
    )
    layout.write_scenario_banner(ws, row=3)
    yr_row = 5
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        styles.style_header(ws.cell(row=yr_row, column=ci, value=f"t={i}"))

    for hr, en, it in headers:
        layout.write_section_header(ws, hr, en, it)

    def _ci(i):
        return ord(layout.year_col(i)) - ord("A") + 1

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    def put(key, i, formula, fmt=styles.FMT_EUR_M, kind="formula", bold=False, top=False):
        ci = _ci(i)
        c = ws.cell(row=R[key], column=ci, value=formula)
        getattr(styles, f"style_{kind}")(c, number_format=fmt)
        if bold:
            c.font = styles.font_subheader
        if top:
            c.border = styles.BORDER_TOP_THIN
        return c

    initial_pool = lt("D", "initial_pool_row")

    def adv(j):  # initial balance of note j
        return f"{notes[j].advance_pct.name}*{initial_pool}"

    # ── Available cash ─────────────────────────────────────────────────────
    layout.write_row_label(ws, R["gross_int"], "Gross interest collected", "Interessi lordi")
    layout.write_row_label(ws, R["servicing"], "(−) Servicing fee", "(−) Fee servicing", indent=True)
    layout.write_row_label(ws, R["trustee"], "(−) Trustee / senior fees", "(−) Fee trustee", indent=True)
    layout.write_row_label(ws, R["avail_int_after_fees"], "Available interest (after fees)",
                           "Interessi disponibili (netti fee)")
    layout.write_row_label(ws, R["principal_collected"], "Principal collected", "Capitale incassato")
    for k in ("gross_int", "servicing", "trustee", "avail_int_after_fees", "principal_collected"):
        _unit(R[k], cur)
    for i in range(n):
        col = layout.year_col(i)
        put("gross_int", i, 0 if i == 0 else f"={lt(col, 'pool_interest_row')}", kind="xref")
        put("servicing", i, 0 if i == 0 else f"=-{lt(col, 'pool_bop_row')}*servicing_fee_pct")
        # Trustee / admin fees only accrue while the pool is performing — once the
        # collateral has fully amortized the deal is wound down, so no fixed fee
        # accrues (and none can be funded); avoids a phantom tail fee drag.
        put("trustee", i,
            0 if i == 0 else f"=-senior_fees_eur_m*IF({lt(col, 'pool_bop_row')}>0.0001,1,0)")
        put("avail_int_after_fees", i,
            0 if i == 0 else f"=${col}${R['gross_int']}+${col}${R['servicing']}+${col}${R['trustee']}",
            bold=True)
        put("principal_collected", i,
            0 if i == 0 else f"={lt(col, 'principal_collected_row')}", kind="xref", bold=True)

    # ── Note balances — outstanding (EOP) ──────────────────────────────────
    for j in range(len(notes)):
        layout.write_row_label(ws, R[f"eop_{j}"], f"Outstanding — {notes[j].name.en}",
                               f"In essere — {notes[j].name.secondary}", indent=True)
        _unit(R[f"eop_{j}"], cur)
    for i in range(n):
        col = layout.year_col(i); prior = layout.year_col(i - 1)
        for j in range(len(notes)):
            if i == 0:
                put(f"eop_{j}", i, f"={adv(j)}", bold=(j == senior))
                continue
            if j == eq:
                put(f"eop_{j}", i,
                    f"=MAX(${prior}${R[f'eop_{j}']}-${col}${R['prin_to_residual']},0)")
            else:
                # debt note: collections paydown + turbo + maturity reserve cure
                put(f"eop_{j}", i,
                    f"=MAX(${prior}${R[f'eop_{j}']}-${col}${R[f'prin_paid_{j}']}"
                    f"-${col}${R[f'turbo_{j}']}-${col}${R[f'reserve_cure_{j}']},0)",
                    bold=(j == senior))

    # ── Interest waterfall ─────────────────────────────────────────────────
    for j in debt:
        layout.write_row_label(ws, R[f"int_due_{j}"], f"Interest due — {notes[j].name.en}",
                               f"Interessi dovuti — {notes[j].name.secondary}", indent=True)
        _unit(R[f"int_due_{j}"], cur)
    layout.write_row_label(ws, R["reserve_draw"], "Reserve draw (senior shortfall)",
                           "Prelievo riserva", indent=True)
    layout.write_row_label(ws, R["avail_int_total"], "Available interest (incl. reserve draw)",
                           "Interessi disponibili (incl. riserva)")
    for j in debt:
        layout.write_row_label(ws, R[f"int_paid_{j}"], f"Interest paid — {notes[j].name.en}",
                               f"Interessi pagati — {notes[j].name.secondary}", indent=True)
        _unit(R[f"int_paid_{j}"], cur)
    layout.write_row_label(ws, R["residual_int"], "Residual interest (after coupons)",
                           "Interessi residui")
    for k in ("reserve_draw", "avail_int_total", "residual_int"):
        _unit(R[k], cur)
    for i in range(n):
        col = layout.year_col(i); prior = layout.year_col(i - 1)
        for j in debt:
            if i == 0:
                put(f"int_due_{j}", i, 0)
            else:
                put(f"int_due_{j}", i,
                    f"=${prior}${R[f'eop_{j}']}*{notes[j].coupon_pct.name}")
        if i == 0:
            put("reserve_draw", i, 0)
        else:
            put("reserve_draw", i,
                f"=MIN(MAX(${col}${R[f'int_due_{senior}']}-${col}${R['avail_int_after_fees']},0),"
                f"${prior}${R['reserve_eop']})")
        put("avail_int_total", i,
            0 if i == 0 else f"=${col}${R['avail_int_after_fees']}+${col}${R['reserve_draw']}",
            bold=True)
        # sequential interest payment — floored at zero (no negative coupon)
        for idx, j in enumerate(debt):
            if i == 0:
                put(f"int_paid_{j}", i, 0)
                continue
            higher = "".join(f"-${col}${R[f'int_paid_{debt[k]}']}" for k in range(idx))
            put(f"int_paid_{j}", i,
                f"=MIN(MAX(${col}${R['avail_int_total']}{higher},0),${col}${R[f'int_due_{j}']})")
        if i == 0:
            put("residual_int", i, 0, bold=True)
        else:
            paid = "".join(f"-${col}${R[f'int_paid_{j}']}" for j in debt)
            put("residual_int", i, f"=MAX(${col}${R['avail_int_total']}{paid},0)",
                bold=True, top=True)

    # ── Principal waterfall ────────────────────────────────────────────────
    for j in debt:
        layout.write_row_label(ws, R[f"prin_paid_{j}"], f"Principal paid — {notes[j].name.en}",
                               f"Capitale pagato — {notes[j].name.secondary}", indent=True)
        _unit(R[f"prin_paid_{j}"], cur)
    layout.write_row_label(ws, R["prin_to_residual"], "Principal to residual (excess)",
                           "Capitale al residuale")
    _unit(R["prin_to_residual"], cur)
    for i in range(n):
        col = layout.year_col(i); prior = layout.year_col(i - 1)
        for idx, j in enumerate(debt):
            if i == 0:
                put(f"prin_paid_{j}", i, 0)
                continue
            higher = "".join(f"-${col}${R[f'prin_paid_{debt[k]}']}" for k in range(idx))
            put(f"prin_paid_{j}", i,
                f"=MIN(MAX(${col}${R['principal_collected']}{higher},0),${prior}${R[f'eop_{j}']})")
        if i == 0:
            put("prin_to_residual", i, 0, bold=True)
        else:
            paid = "".join(f"-${col}${R[f'prin_paid_{j}']}" for j in debt)
            put("prin_to_residual", i,
                f"=MAX(${col}${R['principal_collected']}{paid},0)", bold=True, top=True)

    # ── OC / IC triggers + sequential turbo ────────────────────────────────
    layout.write_row_label(ws, R["oc_ratio"], "OC ratio (pool / debt notes, BOP)",
                           "Rapporto OC", indent=True)
    layout.write_row_label(ws, R["oc_pass"], "OC test pass (1=ok)", "Test OC", indent=True)
    layout.write_row_label(ws, R["ic_ratio"], "IC ratio (avail int / debt coupon)",
                           "Rapporto IC", indent=True)
    layout.write_row_label(ws, R["ic_pass"], "IC test pass (1=ok)", "Test IC", indent=True)
    for j in debt:
        layout.write_row_label(ws, R[f"turbo_{j}"], f"Turbo principal → {notes[j].name.en}",
                               f"Turbo capitale → {notes[j].name.secondary}", indent=True)
        _unit(R[f"turbo_{j}"], cur)
    ws.cell(row=R[f"turbo_{senior}"], column=4).comment = Comment(
        "On an OC or IC breach, residual interest is diverted to an accelerated "
        "paydown of the most-senior note STILL OUTSTANDING (senior → mezz → …), "
        "each capped at that note's balance remaining after the scheduled "
        "collections paydown — excess-spread capture that de-levers the impaired "
        "debt stack rather than leaking to the first-loss equity.", "ModelForge")
    for i in range(n):
        col = layout.year_col(i); prior = layout.year_col(i - 1)
        if i == 0:
            for k in ("oc_ratio", "ic_ratio"):
                put(k, i, 0, fmt=styles.FMT_MULTIPLE)
            for k in ("oc_pass", "ic_pass"):
                put(k, i, 1, fmt=styles.FMT_INTEGER)
            for j in debt:
                put(f"turbo_{j}", i, 0)
            continue
        debt_bop = "+".join(f"${prior}${R[f'eop_{j}']}" for j in debt)
        debt_due = "+".join(f"${col}${R[f'int_due_{j}']}" for j in debt)
        put("oc_ratio", i, f"=IFERROR({lt(col, 'pool_bop_row')}/({debt_bop}),0)",
            fmt=styles.FMT_MULTIPLE)
        put("oc_pass", i, f"=IF(${col}${R['oc_ratio']}>=oc_trigger_pct,1,0)", fmt=styles.FMT_INTEGER)
        put("ic_ratio", i, f"=IFERROR(${col}${R['avail_int_after_fees']}/({debt_due}),0)",
            fmt=styles.FMT_MULTIPLE)
        put("ic_pass", i, f"=IF(${col}${R['ic_ratio']}>=ic_trigger_pct,1,0)", fmt=styles.FMT_INTEGER)
        # Sequential turbo: senior-most outstanding note first, funded from the
        # residual interest remaining after higher notes' turbo.
        for idx, j in enumerate(debt):
            higher_turbo = "".join(f"-${col}${R[f'turbo_{debt[k]}']}" for k in range(idx))
            put(f"turbo_{j}", i,
                f"=IF(AND(${col}${R['oc_pass']}=1,${col}${R['ic_pass']}=1),0,"
                f"MIN(MAX(${col}${R['residual_int']}{higher_turbo},0),"
                f"MAX(${prior}${R[f'eop_{j}']}-${col}${R[f'prin_paid_{j}']},0)))")

    # ── Reserve account (cures debt before returning to equity) ────────────
    layout.write_row_label(ws, R["reserve_eop"], "Reserve balance (EOP)", "Saldo riserva", indent=True)
    layout.write_row_label(ws, R["reserve_release"], "Reserve released at maturity (gross)",
                           "Riserva rilasciata a scadenza", indent=True)
    for j in debt:
        layout.write_row_label(ws, R[f"reserve_cure_{j}"],
                               f"Reserve cure → {notes[j].name.en}",
                               f"Riserva a copertura → {notes[j].name.secondary}", indent=True)
        _unit(R[f"reserve_cure_{j}"], cur)
    layout.write_row_label(ws, R["reserve_to_equity"], "Reserve return to residual (capital)",
                           "Rientro riserva al residuale", indent=True)
    for k in ("reserve_eop", "reserve_release", "reserve_to_equity"):
        _unit(R[k], cur)
    for i in range(n):
        col = layout.year_col(i); prior = layout.year_col(i - 1)
        if i == 0:
            put("reserve_eop", i, f"=reserve_pct_initial*{initial_pool}")
            put("reserve_release", i, 0)
            for j in debt:
                put(f"reserve_cure_{j}", i, 0)
            put("reserve_to_equity", i, 0)
            continue
        if i < periods:
            put("reserve_eop", i, f"=${prior}${R['reserve_eop']}-${col}${R['reserve_draw']}")
            put("reserve_release", i, 0)
            for j in debt:
                put(f"reserve_cure_{j}", i, 0)
            put("reserve_to_equity", i, 0)
        else:  # final period: release the reserve, cure debt sequentially, return the rest
            put("reserve_eop", i, 0)
            put("reserve_release", i, f"=${prior}${R['reserve_eop']}-${col}${R['reserve_draw']}")
            for idx, j in enumerate(debt):
                higher_cure = "".join(f"-${col}${R[f'reserve_cure_{debt[k]}']}" for k in range(idx))
                put(f"reserve_cure_{j}", i,
                    f"=MIN(MAX(${col}${R['reserve_release']}{higher_cure},0),"
                    f"MAX(${prior}${R[f'eop_{j}']}-${col}${R[f'prin_paid_{j}']}-${col}${R[f'turbo_{j}']},0))")
            cured = "".join(f"-${col}${R[f'reserve_cure_{j}']}" for j in debt)
            put("reserve_to_equity", i, f"=MAX(${col}${R['reserve_release']}{cured},0)")

    # ── Residual / equity distribution ─────────────────────────────────────
    layout.write_row_label(ws, R["residual_int_to_equity"], "Residual interest to equity (pre-tax)",
                           "Interessi residui all'equity")
    layout.write_row_label(ws, R["spv_tax"], "(−) SPV tax", "(−) Imposta SPV", indent=True)
    layout.write_row_label(ws, R["cash_to_equity"], "Total cash to residual / equity",
                           "Cassa totale al residuale")
    for k in ("residual_int_to_equity", "spv_tax", "cash_to_equity"):
        _unit(R[k], cur)
    for i in range(n):
        col = layout.year_col(i)
        if i == 0:
            put("residual_int_to_equity", i, 0)
            put("spv_tax", i, 0)
            put("cash_to_equity", i, 0, bold=True)
            continue
        turbo_terms = "".join(f"-${col}${R[f'turbo_{j}']}" for j in debt)
        put("residual_int_to_equity", i, f"=${col}${R['residual_int']}{turbo_terms}")
        put("spv_tax", i,
            f"=-MAX(${col}${R['residual_int_to_equity']},0)*effective_tax_rate")
        put("cash_to_equity", i,
            f"=${col}${R['residual_int_to_equity']}+${col}${R['spv_tax']}"
            f"+${col}${R['prin_to_residual']}+${col}${R['reserve_to_equity']}",
            bold=True, top=True)

    # ── Investor cashflows ─────────────────────────────────────────────────
    for j in range(len(notes)):
        layout.write_row_label(ws, R[f"inv_cf_{j}"], f"Investor CF — {notes[j].name.en}",
                               f"CF investitore — {notes[j].name.secondary}", indent=True)
        _unit(R[f"inv_cf_{j}"], cur)
    for i in range(n):
        col = layout.year_col(i)
        for j in range(len(notes)):
            if i == 0:
                if j == eq:
                    # the residual holder funds its note AND the cash reserve at close
                    put(f"inv_cf_{j}", i, f"=-({adv(j)})-reserve_pct_initial*{initial_pool}",
                        bold=(j == senior))
                else:
                    put(f"inv_cf_{j}", i, f"=-({adv(j)})")
                continue
            if j == eq:
                put(f"inv_cf_{j}", i, f"=${col}${R['cash_to_equity']}")
            else:
                put(f"inv_cf_{j}", i,
                    f"=${col}${R[f'int_paid_{j}']}+${col}${R[f'prin_paid_{j}']}"
                    f"+${col}${R[f'turbo_{j}']}+${col}${R[f'reserve_cure_{j}']}")

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in R.items()}
    out["n_debt"] = str(len(debt))
    out["n_notes"] = str(len(notes))
    out["senior_idx"] = str(senior)
    out["eq_idx"] = str(eq)
    out["last_row"] = str(last_row)
    return out
