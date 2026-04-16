"""Assumptions sheet.

The single source of truth for every driver. Every other sheet reads from
here via named ranges. No other sheet holds a hardcoded number.

Layout (columns):
    A  ID (A-001 or S-001 for sourced)
    B  Driver (snake_case name; this is the named range)
    C  Description (EN)
    D  Description (IT)
    E  Unit
    F  Worst
    G  Base     ← the default
    H  Best
    I  Active   = CHOOSE(scenario_index, Worst, Base, Best)  ← other sheets read this
    J  Rationale
    K  Conf.
    L  Source

Every BASE cell gets a comment with the source/assumption id.
Every 'Active' cell becomes a workbook-level named range.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles
from modelforge.builder.formulas import scenario_pick
from modelforge.builder.i18n import L
from modelforge.graph.schema import LinkageGraph, GraphNode, NodeKind
from modelforge.spec.base import Assumption


def _unit_to_number_format(unit: str) -> str:
    return {
        "eur_m": styles.FMT_EUR_M,
        "eur_k": styles.FMT_EUR_K,
        "eur": styles.FMT_EUR_ACTUAL,
        "pct": styles.FMT_PCT,
        "x": styles.FMT_MULTIPLE,
        "years": styles.FMT_YEARS,
        "bps": styles.FMT_BPS,
        "count": styles.FMT_INTEGER,
        "ratio": styles.FMT_MULTIPLE,
    }.get(unit, "General")


def _register_name(wb, name: str, addr: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=addr)


def build(
    ws: Worksheet,
    spec,
    graph: LinkageGraph,
    source_rows: dict[str, int],
) -> dict[str, str]:
    """Emit Assumptions sheet. Returns {assumption_name: active_cell_ref_string}.

    Active cell refs are like "'Assumptions'!$I$12" — ready for other sheets.
    """
    # Column widths
    widths = {"A": 9, "B": 28, "C": 40, "D": 34, "E": 8,
              "F": 11, "G": 11, "H": 11, "I": 11, "J": 48, "K": 6, "L": 8}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Title block
    ws.cell(row=1, column=1, value="Assumptions").font = styles.font_title
    ws.cell(row=2, column=1, value="Ipotesi e driver").font = styles.font_label_it
    ws.cell(
        row=3, column=1,
        value="Single source of truth. Every formula in the workbook reads "
              "the 'Active' column via named range. Change numbers here — "
              "nowhere else.",
    ).font = styles.font_label_it

    # Header row
    hr = 5
    headers = [
        ("assumption_id", 1),
        ("assumption_name", 2),
        ("assumption_label", 3),   # EN desc
        # IT in col 4
        ("assumption_unit", 5),
        ("assumption_worst", 6),
        ("assumption_base", 7),
        ("assumption_best", 8),
        ("assumption_active", 9),
        ("assumption_rationale", 10),
        ("assumption_confidence", 11),
        ("assumption_source", 12),
    ]
    for key, col in headers:
        c = ws.cell(row=hr, column=col, value=L(key).en)
        styles.style_header(c)
    # IT description header in col 4
    c = ws.cell(row=hr, column=4, value=L("assumption_label").it)
    styles.style_header(c)

    # Rows
    driver_refs: dict[str, str] = {}
    r = hr + 1
    assumptions = spec.all_assumptions()
    seen_names: set[str] = set()
    for a in assumptions:
        if a.name in seen_names:
            # Avoid duplicate rows when the same default Assumption
            # object is reused (e.g. two tranches with the same default
            # commitment_fee_bps). Skip — first write wins.
            continue
        seen_names.add(a.name)
        _emit_row(ws, r, a, graph, source_rows)
        driver_refs[a.name] = f"'{ws.title}'!$I${r}"
        # Workbook-level named range points at the Active cell
        _register_name(ws.parent, a.name, driver_refs[a.name])
        r += 1

    ws.freeze_panes = "E6"
    ws.print_title_rows = f"{hr}:{hr}"
    return driver_refs


def _emit_row(
    ws: Worksheet,
    r: int,
    a: Assumption,
    graph: LinkageGraph,
    source_rows: dict[str, int],
) -> None:
    nfmt = _unit_to_number_format(a.unit)

    # Col A: ID
    id_cell = ws.cell(row=r, column=1, value=a.id)
    id_cell.font = styles.font_subheader

    # Col B: name (this is how formulas reference the driver)
    name_cell = ws.cell(row=r, column=2, value=a.name)
    name_cell.font = styles.Font(
        name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY, italic=True, color="555555",
    )

    # Cols C & D: EN / IT description
    en = ws.cell(row=r, column=3, value=a.label.en)
    styles.style_label_en(en)
    it = ws.cell(row=r, column=4, value=a.label.it or a.label.en)
    styles.style_label_it(it)

    # Col E: unit
    u = ws.cell(row=r, column=5, value=a.unit)
    u.font = styles.font_label_it
    u.alignment = styles.align_center

    # Cols F/G/H: Worst / Base / Best (hardcoded inputs)
    worst_val = a.worst if a.worst is not None else a.base
    best_val = a.best if a.best is not None else a.base
    w_cell = ws.cell(row=r, column=6, value=worst_val)
    b_cell = ws.cell(row=r, column=7, value=a.base)
    x_cell = ws.cell(row=r, column=8, value=best_val)
    for c in (w_cell, b_cell, x_cell):
        styles.style_input(c, number_format=nfmt)

    # Comment with source/assumption id on BASE cell
    comment_lines = [f"{a.id}"]
    if a.source_id:
        comment_lines.append(f"Source: {a.source_id}")
        if a.source_id in source_rows:
            comment_lines.append(
                f"(see Sources!A{source_rows[a.source_id]})"
            )
    comment_lines.append(f"Confidence: {a.confidence.value}")
    comment_lines.append(f"Rationale: {a.rationale}")
    b_cell.comment = Comment("\n".join(comment_lines), "ModelForge")
    b_cell.comment.width = 280
    b_cell.comment.height = 120

    # Col I: Active = CHOOSE(scenario_index, Worst, Base, Best)
    active_formula = f"=CHOOSE(scenario_index,F{r},G{r},H{r})"
    a_cell = ws.cell(row=r, column=9, value=active_formula)
    styles.style_formula(a_cell, number_format=nfmt)
    a_cell.font = styles.Font(
        name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY, bold=True,
    )

    # Col J: Rationale
    rat = ws.cell(row=r, column=10, value=a.rationale)
    rat.font = styles.font_label_it
    rat.alignment = styles.Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Col K: Confidence
    conf = ws.cell(row=r, column=11, value=a.confidence.value)
    conf.alignment = styles.align_center
    conf.font = styles.Font(
        name=styles.FONT_BASE,
        size=styles.FONT_SIZE_BODY,
        bold=True,
        color={"H": "006100", "M": "7F6000", "L": "9C0006"}[a.confidence.value],
    )

    # Col L: Source link
    if a.source_id:
        src = ws.cell(row=r, column=12, value=a.source_id)
        src.alignment = styles.align_center
        if a.source_id in source_rows:
            src.hyperlink = f"#'Sources'!A{source_rows[a.source_id]}"
            src.font = styles.Font(
                name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
                color="0563C1", underline="single",
            )

    # Graph: register the driver and its Active cell
    drv_id = LinkageGraph.driver_id(a.name)
    graph.add_node(
        GraphNode(
            id=drv_id,
            kind=NodeKind.DRIVER,
            label=a.name,
            payload={
                "unit": a.unit, "assumption_id": a.id,
                "source_id": a.source_id, "confidence": a.confidence.value,
            },
        )
    )
    base_cell_id = LinkageGraph.cell_id(ws.title, f"G{r}")
    active_cell_id = LinkageGraph.cell_id(ws.title, f"I{r}")
    graph.add_node(GraphNode(id=base_cell_id, kind=NodeKind.CELL,
                              label=f"{a.name} (base)"))
    graph.add_node(GraphNode(id=active_cell_id, kind=NodeKind.CELL,
                              label=f"{a.name} (active)"))
    graph.lives_on(drv_id, active_cell_id)

    # Source provenance
    if a.source_id:
        if a.source_id not in graph.nodes:
            graph.add_node(GraphNode(id=a.source_id, kind=NodeKind.SOURCE, label=a.source_id))
        graph.provides(a.source_id, drv_id)
    else:
        # Register as an ASSUMPTION node and link to driver
        if a.id not in graph.nodes:
            graph.add_node(
                GraphNode(
                    id=a.id, kind=NodeKind.ASSUMPTION, label=a.name,
                    payload={"rationale": a.rationale, "confidence": a.confidence.value},
                )
            )
        graph.provides(a.id, drv_id)
