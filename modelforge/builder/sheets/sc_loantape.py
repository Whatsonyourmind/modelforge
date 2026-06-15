"""Loan-tape (asset side) sheet — Template 19.

A stratified loan tape projected period-by-period. Each stratum amortizes on
its OWN WAM (straight-line to maturity with a clean-up sweep at the deal's
final period) and defaults on its OWN CDR; the pool is the sum of strata. Then
pool-level recoveries (lagged) and the principal-collected stream the liability
waterfall consumes.

Roll-forward safety: every stratum's beginning-of-period performing balance
references the PRIOR period's CLOSING row via a symbolically-reserved row +
an ``assert`` drift-guard (the debt.py / bank_bs precedent), so inserting a
row inside a stratum block can never silently mis-wire the roll-forward.

Within-period flow order (costs do not feed back on the same cell):
    surviving = opening − default        (defaults leave the performing pool)
    scheduled = surviving × amort_rate    (straight-line to remaining term)
    prepay    = (surviving − scheduled) × CPR
    closing   = opening − default − scheduled − prepay
    interest  = opening × WAC             (BOP-balance accrual)
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    periods = spec.horizon.periods
    lag = spec.horizon.recovery_lag_periods
    n = periods + 1  # t=0 (close) .. t=periods
    cur = spec.meta.currency

    layout.set_column_widths(ws, label_width=46, it_width=32, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Loan Tape (collateral cashflow)",
        title_it="Loan tape (cassa collaterale)",
        subtitle=f"{cur} · {len(spec.tape)} strata · {periods}p · "
                 f"CPR/CDR/recovery (lag {lag}p)",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=yr_row, column=ci, value=f"t={i}")
        styles.style_header(cc)

    def _ci(i):
        return ord(layout.year_col(i)) - ord("A") + 1

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    rows: dict[str, int] = {}
    r = 7

    # ── Tape summary (balance-weighted pool stats) ─────────────────────────
    layout.write_section_header(ws, r, "Stratified loan tape — pool summary",
                                "Loan tape stratificato — sintesi pool")
    r += 1

    rows["initial_pool"] = r
    layout.write_row_label(ws, r, "Initial pool balance (UPB)", "Saldo pool iniziale (UPB)")
    _unit(r, cur)
    upb_terms = "+".join(f"{s.balance_eur_m.name}" for s in spec.tape)
    c = ws.cell(row=r, column=4, value=f"={upb_terms}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    r += 1

    rows["pool_wac"] = r
    layout.write_row_label(ws, r, "Pool WAC (balance-weighted)", "WAC pool", indent=True)
    _unit(r, "%")
    wac_num = "+".join(f"{s.balance_eur_m.name}*{s.wac_pct.name}" for s in spec.tape)
    c = ws.cell(row=r, column=4, value=f"=({wac_num})/$D${rows['initial_pool']}")
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["pool_wam"] = r
    layout.write_row_label(ws, r, "Pool WAM (balance-weighted, years)", "WAM pool", indent=True)
    _unit(r, "y")
    wam_num = "+".join(f"{s.balance_eur_m.name}*{s.wam_years.name}" for s in spec.tape)
    c = ws.cell(row=r, column=4, value=f"=({wam_num})/$D${rows['initial_pool']}")
    styles.style_formula(c, number_format=styles.FMT_YEARS)
    r += 2

    # ── Per-stratum projection ─────────────────────────────────────────────
    strata_rows: list[dict] = []
    for s in spec.tape:
        layout.write_section_header(
            ws, r, f"Stratum: {s.name.en}", f"Strato: {s.name.secondary}")
        r += 1

        upb_row = r
        layout.write_row_label(ws, r, "UPB at close", "UPB alla chiusura", indent=True)
        _unit(r, cur)
        c = ws.cell(row=r, column=4, value=f"={s.balance_eur_m.name}")
        styles.style_xref(c, number_format=styles.FMT_EUR_M)
        r += 1

        # Memo: WAC / WAM / CDR (named ranges, shown for tape transparency)
        layout.write_row_label(ws, r, "WAC (gross coupon)", "WAC", indent=True)
        _unit(r, "%")
        c = ws.cell(row=r, column=4, value=f"={s.wac_pct.name}")
        styles.style_xref(c, number_format=styles.FMT_PCT_2DP)
        r += 1
        layout.write_row_label(ws, r, "WAM (remaining term, years)", "WAM", indent=True)
        _unit(r, "y")
        c = ws.cell(row=r, column=4, value=f"={s.wam_years.name}")
        styles.style_xref(c, number_format=styles.FMT_YEARS)
        r += 1
        layout.write_row_label(ws, r, "CDR (annual default rate)", "CDR", indent=True)
        _unit(r, "%")
        c = ws.cell(row=r, column=4, value=f"={s.cdr_pct.name}")
        styles.style_xref(c, number_format=styles.FMT_PCT_2DP)
        r += 1

        # Performing balance — opening (BOP). Reserve the closing row
        # symbolically: between opening and closing sit default, scheduled,
        # prepay (3 rows) ⇒ closing = opening + 4. An assert at the closing
        # write hard-fails if a row is ever inserted between them.
        open_row = r
        close_row = r + 4
        layout.write_row_label(ws, r, "Performing balance — opening (BOP)",
                               "Saldo in bonis — apertura", indent=True)
        _unit(r, cur)
        for i in range(n):
            col = layout.year_col(i); ci = _ci(i)
            if i == 0:
                c = ws.cell(row=r, column=ci, value=0)
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
            elif i == 1:
                c = ws.cell(row=r, column=ci, value=f"=$D${upb_row}")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
            else:
                prior = layout.year_col(i - 1)
                c = ws.cell(row=r, column=ci, value=f"=${prior}${close_row}")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        default_row = r
        layout.write_row_label(ws, r, "(−) Defaults (CDR × BOP)", "(−) Default", indent=True)
        _unit(r, cur)
        for i in range(1, n):
            col = layout.year_col(i); ci = _ci(i)
            c = ws.cell(row=r, column=ci,
                        value=f"=${col}${open_row}*{s.cdr_pct.name}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        sched_row = r
        layout.write_row_label(ws, r, "(−) Scheduled principal", "(−) Capitale programmato",
                               indent=True)
        _unit(r, cur)
        for i in range(1, n):
            col = layout.year_col(i); ci = _ci(i)
            surviving = f"(${col}${open_row}-${col}${default_row})"
            if i == periods:
                # Clean-up / call sweep at the deal's final period: redeem the
                # whole surviving balance as scheduled principal so the pool
                # fully amortizes regardless of WAM > horizon.
                c = ws.cell(row=r, column=ci, value=f"={surviving}")
            else:
                # Straight-line to remaining term: rate = MIN(1, 1/MAX(WAM−elapsed,0.5)).
                # elapsed = i−1 completed periods; rate → 1 (sweep) as term runs out.
                rate = f"MIN(1,1/MAX({s.wam_years.name}-{i-1},0.5))"
                c = ws.cell(row=r, column=ci, value=f"={surviving}*{rate}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        prepay_row = r
        layout.write_row_label(ws, r, "(−) Prepayments (CPR)", "(−) Prepagamenti", indent=True)
        _unit(r, cur)
        for i in range(1, n):
            col = layout.year_col(i); ci = _ci(i)
            # CPR on the balance remaining after defaults and scheduled amort.
            rem = f"(${col}${open_row}-${col}${default_row}-${col}${sched_row})"
            c = ws.cell(row=r, column=ci, value=f"={rem}*cpr_pct")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        assert r == close_row, (
            f"loan-tape stratum {s.name.en!r}: closing-row drift, expected "
            f"{close_row}, got {r}"
        )
        layout.write_row_label(ws, r, "Performing balance — closing (EOP)",
                               "Saldo in bonis — chiusura", indent=True)
        _unit(r, cur)
        for i in range(n):
            col = layout.year_col(i); ci = _ci(i)
            if i == 0:
                c = ws.cell(row=r, column=ci, value=f"=$D${upb_row}")
            else:
                c = ws.cell(row=r, column=ci,
                            value=(f"=${col}${open_row}-${col}${default_row}"
                                   f"-${col}${sched_row}-${col}${prepay_row}"))
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        interest_row = r
        layout.write_row_label(ws, r, "Interest accrued (WAC × BOP)", "Interessi maturati",
                               indent=True)
        _unit(r, cur)
        for i in range(n):
            col = layout.year_col(i); ci = _ci(i)
            if i == 0:
                c = ws.cell(row=r, column=ci, value=0)
            else:
                c = ws.cell(row=r, column=ci,
                            value=f"=${col}${open_row}*{s.wac_pct.name}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 2

        strata_rows.append({
            "open": open_row, "default": default_row, "sched": sched_row,
            "prepay": prepay_row, "close": close_row, "interest": interest_row,
        })

    # ── Pool totals (sum across strata) ────────────────────────────────────
    layout.write_section_header(ws, r, "Pool cashflow totals", "Totali cassa pool")
    r += 1

    def _sum_row(label_en, label_it, key, t_from=0, bold=False):
        nonlocal r
        rr = r
        layout.write_row_label(ws, rr, label_en, label_it, indent=not bold)
        _unit(rr, cur)
        for i in range(n):
            col = layout.year_col(i); ci = _ci(i)
            if i < t_from:
                c = ws.cell(row=rr, column=ci, value=0)
            else:
                terms = "+".join(f"${col}${st[key]}" for st in strata_rows)
                c = ws.cell(row=rr, column=ci, value=f"={terms}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
            if bold:
                c.font = styles.font_subheader
        r += 1
        return rr

    rows["pool_bop"] = _sum_row("Pool performing balance (BOP)", "Saldo pool (apertura)", "open")
    rows["pool_default"] = _sum_row("Pool defaults", "Default pool", "default", t_from=1)
    rows["pool_sched"] = _sum_row("Pool scheduled principal", "Capitale programmato pool", "sched", t_from=1)
    rows["pool_prepay"] = _sum_row("Pool prepayments", "Prepagamenti pool", "prepay", t_from=1)

    # Recoveries — pool defaults collected with a fixed `lag`-period delay,
    # at recovery_pct (1 − severity). The final period sweeps any not-yet-
    # collected vintages so total recoveries == recovery_pct × total defaults
    # exactly (no leak past maturity). See spec docstring for the derivation.
    rec_row = r
    rows["pool_recovery"] = rec_row
    layout.write_row_label(ws, rec_row, "Pool recoveries (lagged)", "Recuperi pool (ritardati)",
                           indent=True)
    _unit(rec_row, cur)
    pdef = rows["pool_default"]
    for i in range(n):
        col = layout.year_col(i); ci = _ci(i)
        if i == 0:
            c = ws.cell(row=rec_row, column=ci, value=0)
        elif i < periods:
            j = i - lag
            if j >= 1:
                jc = layout.year_col(j)
                c = ws.cell(row=rec_row, column=ci,
                            value=f"=recovery_pct*${jc}${pdef}")
            else:
                c = ws.cell(row=rec_row, column=ci, value=0)
        else:  # final period: sweep uncollected vintages (periods−lag … periods)
            start = max(periods - lag, 1)
            terms = "+".join(f"${layout.year_col(j)}${pdef}" for j in range(start, periods + 1))
            c = ws.cell(row=rec_row, column=ci, value=f"=recovery_pct*({terms})")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=rec_row, column=4).comment = Comment(
        f"Recoveries arrive {lag} period(s) after default at recovery_pct. The "
        "final period sweeps any vintage whose lagged collection would fall "
        "beyond maturity, so Σ recoveries = recovery_pct × Σ defaults exactly.",
        "ModelForge")
    r += 1

    rows["pool_interest"] = _sum_row("Pool interest accrued", "Interessi pool", "interest")

    rows["principal_collected"] = r
    layout.write_row_label(ws, r, "Principal collected (sched + prepay + recovery)",
                           "Capitale incassato")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = _ci(i)
        if i == 0:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            c = ws.cell(row=r, column=ci,
                        value=(f"=${col}${rows['pool_sched']}+${col}${rows['pool_prepay']}"
                               f"+${col}${rows['pool_recovery']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    rows["pool_eop"] = _sum_row("Pool performing balance (EOP)", "Saldo pool (chiusura)", "close")

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {f"{k}_row": str(v) for k, v in rows.items()}
