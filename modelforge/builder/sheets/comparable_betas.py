"""ComparableBetas sheet — Hamada unlever/relever.

Bulge-bracket practice for DCF / fairness opinions:

  1. Pull levered β for each public comparable
  2. Unlever each: β_U = β_L / [1 + (1 − t)(D/E)]
  3. Take MEDIAN of unlevered β (not mean — bulge standard)
  4. Relever to target capital structure:
     β_relevered = β_U × [1 + (1 − t_target)(D/E)_target]

Emits:
  - `unlevered_beta_median` named range (consumed by WACCBuild)
  - `relevered_beta` named range (consumed by WACCBuild)

References:
  - Hamada's Equation — Wikipedia
  - IB Interview Questions: Beta Unlever / Relever
  - Damodaran (Chapter 12 DCF derivations)
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def _define_name(wb, name: str, sheet: str, cell: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text=f"'{sheet}'!${cell[:1]}${cell[1:]}" if cell[0].isalpha() else cell,
    )


def build(ws: Worksheet, spec) -> dict[str, str]:
    """Build the ComparableBetas sheet. Returns refs for WACCBuild to consume."""
    wb = ws.parent
    comps = spec.wacc.comparable_betas

    if not comps:
        # Nothing to build — WACCBuild will fall back to beta_levered input
        ws.cell(row=1, column=1, value="Comparable Betas (none provided)").font = styles.font_title
        ws.cell(row=2, column=1, value="Add wacc.comparable_betas to spec to enable Hamada unlever/relever").font = styles.font_label_it
        return {}

    layout.set_column_widths(ws, label_width=40, it_width=20, year_width=14, unit_width=8)
    layout.write_title_block(
        ws, "Comparable Betas — Hamada unlever/relever",
        "Beta comparabili — unlever/relever Hamada",
        "Bulge-bracket beta derivation: unlever each comp, take median, relever to target",
    )

    hr = 5
    headers = ["Company", "Levered β", "D/E", "Marginal tax", "Unlevered β (Hamada)"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        styles.style_header(c)

    r0 = hr + 1
    for i, comp in enumerate(comps):
        r = r0 + i
        ws.cell(row=r, column=1, value=comp.name).font = styles.font_label_en
        c_beta = ws.cell(row=r, column=2, value=comp.beta_levered)
        styles.style_input(c_beta, number_format=styles.FMT_MULTIPLE)
        c_beta.comment = Comment(f"{comp.name} levered β (5Y weekly vs local benchmark)", "ModelForge")
        c_de = ws.cell(row=r, column=3, value=comp.debt_to_equity)
        styles.style_input(c_de, number_format=styles.FMT_MULTIPLE)
        c_tax = ws.cell(row=r, column=4, value=comp.marginal_tax_rate)
        styles.style_input(c_tax, number_format=styles.FMT_PCT)
        # Hamada unlever: β_U = β_L / (1 + (1 − t) × D/E)
        c_beta_u = ws.cell(
            row=r, column=5,
            value=f"=B{r}/(1+(1-D{r})*C{r})",
        )
        styles.style_formula(c_beta_u, number_format=styles.FMT_MULTIPLE)
    r_last = r0 + len(comps) - 1

    # Summary
    r_sum = r_last + 2
    ws.cell(row=r_sum, column=1, value="Median unlevered β").font = styles.font_subheader
    median_cell = ws.cell(row=r_sum, column=5,
                          value=f"=MEDIAN(E{r0}:E{r_last})")
    styles.style_formula(median_cell, number_format=styles.FMT_MULTIPLE)
    median_cell.font = styles.font_subheader
    median_cell.border = styles.BORDER_TOP_THIN
    _define_name(wb, "unlevered_beta_median", ws.title, f"E{r_sum}")
    r_sum += 1

    ws.cell(row=r_sum, column=1, value="Mean unlevered β (reference)").font = styles.font_label_it
    ws.cell(row=r_sum, column=5, value=f"=AVERAGE(E{r0}:E{r_last})").number_format = styles.FMT_MULTIPLE
    r_sum += 2

    # Relever to target
    ws.cell(row=r_sum, column=1, value="Target D/E (from target_debt_weight)").font = styles.font_label_en
    target_de = ws.cell(row=r_sum, column=5,
                        value="=target_debt_weight/(1-target_debt_weight)")
    styles.style_formula(target_de, number_format=styles.FMT_MULTIPLE)
    r_sum += 1

    ws.cell(row=r_sum, column=1, value="Target marginal tax").font = styles.font_label_en
    ws.cell(row=r_sum, column=5, value="=effective_tax_rate").number_format = styles.FMT_PCT
    r_sum += 1

    ws.cell(row=r_sum, column=1, value="Relevered β (target structure)").font = styles.font_subheader
    relev = ws.cell(
        row=r_sum, column=5,
        value=f"=E{r_sum-3}*(1+(1-E{r_sum-1})*E{r_sum-2})",
    )
    styles.style_formula(relev, number_format=styles.FMT_MULTIPLE)
    relev.font = styles.font_subheader
    relev.border = styles.BORDER_TOP_THIN
    relev.comment = Comment(
        "Relevered β = β_U × [1 + (1 − t_target) × (D/E)_target]\n"
        "Hamada's equation, ignoring debt beta.\n"
        "Reference: Damodaran Ch.12; IB Interview Questions.",
        "ModelForge",
    )
    _define_name(wb, "relevered_beta", ws.title, f"E{r_sum}")

    ws.freeze_panes = f"A{r0}"
    ws.print_title_rows = f"{hr}:{hr}"

    return {
        "unlevered_beta_median_row": str(r_last + 2),
        "relevered_beta_row": str(r_sum),
    }
