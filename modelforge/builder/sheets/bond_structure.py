"""Bond Structure sheet (Minibond Template 2).

Amortization schedule + coupon schedule for the minibond.
Patches IssuerFinancials interest row with total cash interest.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.i18n import L


def build(ws: Worksheet, spec, operating_refs: dict[str, str],
          operating_sheet_name: str) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Bond Structure", title_it="Struttura del bond",
        subtitle=f"Notional {spec.meta.currency} · "
                 f"{spec.bond.tenor_years}y · "
                 f"{spec.bond.amortization} · "
                 f"{'listed ExtraMOT Pro' if spec.bond.listed_extramot_pro else 'unlisted'}"
                 + (" · ElTIF eligible" if spec.bond.eltif_eligible else ""),
    )
    layout.write_scenario_banner(ws, row=3)

    # Year headers
    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        is_hist = i < h
        c = ws.cell(row=yr_row, column=col_idx, value=f"{'A' if is_hist else 'E'} {yr}")
        styles.style_header(c)

    r = 7
    rows: dict[str, int] = {}

    layout.write_section_header(ws, r, "Principal roll-forward", "Capitale in circolazione")
    r += 1

    rows["opening"] = r
    layout.write_row_label(ws, r, L("debt_opening").en, L("debt_opening").it)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            c = ws.cell(row=r, column=col_idx, value=0)
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        else:
            prior_col = layout.year_col(i - 1)
            c = ws.cell(row=r, column=col_idx, value=f"=${prior_col}${r + 3}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Drawdown (full notional at close)
    rows["drawdown"] = r
    layout.write_row_label(ws, r, L("debt_drawdown").en, L("debt_drawdown").it, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    notional = spec.bond.notional.name
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if i == h:
            c = ws.cell(row=r, column=col_idx, value=f"={notional}")
            styles.style_xref(c, number_format=styles.FMT_EUR_M)
        else:
            c = ws.cell(row=r, column=col_idx, value=0)
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Scheduled amortization
    rows["amort"] = r
    layout.write_row_label(ws, r, L("debt_repayment").en, L("debt_repayment").it, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    tenor = spec.bond.tenor_years
    maturity_year = h + tenor
    amort_start = h + spec.bond.amortization_start_year  # absolute col
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if spec.bond.amortization == "bullet":
            if i == maturity_year:
                c = ws.cell(row=r, column=col_idx, value=f"=-${col}${rows['opening']}")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
                continue
        elif spec.bond.amortization == "linear_from_year":
            # Amortize evenly from amort_start to maturity
            if amort_start <= i <= maturity_year:
                years_amort = maturity_year - amort_start + 1
                c = ws.cell(row=r, column=col_idx, value=f"=-{notional}/{years_amort}")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
                continue
        c = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Closing
    rows["closing"] = r
    layout.write_row_label(ws, r, L("debt_closing").en, L("debt_closing").it)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=${col}${rows['opening']}+${col}${rows['drawdown']}+${col}${rows['amort']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # Coupon
    layout.write_section_header(ws, r, "Coupon schedule", "Cedole")
    r += 1
    rows["avg_balance"] = r
    layout.write_row_label(ws, r, L("debt_average").en, L("debt_average").it, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=(${col}${rows['opening']}+${col}${rows['closing']})/2")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["all_in_rate"] = r
    layout.write_row_label(ws, r, L("all_in_rate").en, L("all_in_rate").it, indent=True)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    if spec.bond.coupon.kind == "fixed":
        rate_ref = spec.bond.coupon.fixed_rate.name
    else:
        ref = spec.bond.coupon.reference_rate_value.name
        margin = spec.bond.coupon.margin_bps.name
        floor = spec.bond.coupon.floor_pct.name if spec.bond.coupon.floor_pct else "0"
        rate_ref = f"MAX({ref},{floor})+({margin}/10000)"
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"={rate_ref}")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["interest"] = r
    layout.write_row_label(ws, r, L("cash_interest").en, L("cash_interest").it)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-${col}${rows['avg_balance']}*${col}${rows['all_in_rate']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 1

    # Patch IssuerFinancials interest
    interest_row_op = int(operating_refs["interest_row"])
    op_ws = ws.parent[operating_sheet_name]
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        op_cell = op_ws.cell(row=interest_row_op, column=col_idx)
        op_cell.value = f"='{ws.title}'!{col}{rows['interest']}"
        styles.style_xref(op_cell, number_format=styles.FMT_EUR_M)

    # Issuance proceeds section
    r += 1
    layout.write_section_header(ws, r, "Issuance proceeds (t=0)", "Proventi in emissione (t=0)")
    r += 1
    rows["gross_proceeds"] = r
    layout.write_row_label(ws, r, "Gross proceeds", "Proventi lordi")
    c = ws.cell(row=r, column=4, value=f"={notional}")
    styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["arrangement_fee"] = r
    layout.write_row_label(ws, r, "Arrangement fee", "Commissione di strutturazione", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=-{notional}*{spec.bond.arrangement_fee_pct.name}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["legal_fees"] = r
    layout.write_row_label(ws, r, "Legal fees", "Spese legali", indent=True)
    c = ws.cell(row=r, column=4, value=f"=-{spec.bond.legal_fees_eur_m.name}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["listing_fees"] = r
    layout.write_row_label(ws, r, "Listing fees", "Spese di quotazione", indent=True)
    c = ws.cell(row=r, column=4, value=f"=-{spec.bond.listing_fees_eur_m.name}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["rating_fees"] = r
    layout.write_row_label(ws, r, "Rating fees", "Spese di rating", indent=True)
    c = ws.cell(row=r, column=4, value=f"=-{spec.bond.rating_fees_eur_m.name}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_proceeds"] = r
    layout.write_row_label(ws, r, "Net proceeds to issuer", "Proventi netti all'emittente")
    c = ws.cell(row=r, column=4,
                value=f"=${r-5}$4+$D${rows['arrangement_fee']}+$D${rows['legal_fees']}+$D${rows['listing_fees']}+$D${rows['rating_fees']}".replace(f"${r-5}$4", f"$D${rows['gross_proceeds']}"))
    # Above construction is brittle; use direct form:
    c.value = (
        f"=$D${rows['gross_proceeds']}+$D${rows['arrangement_fee']}"
        f"+$D${rows['legal_fees']}+$D${rows['listing_fees']}"
        f"+$D${rows['rating_fees']}"
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.border = styles.BORDER_TOP_THIN
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
