"""Generic covenants sheet usable across Minibond, Credit Memo, PF templates.

Each covenant has an 'actual' formula dispatched by kind, a threshold
row (reads from Assumptions year-indexed), headroom, and breach flag.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.formulas import ebitda_multiple, interest_coverage, headroom
from modelforge.builder.i18n import L


def build(
    ws: Worksheet, spec,
    operating_refs: dict[str, str],
    debt_refs: dict[str, str],
    operating_sheet_name: str,
    debt_sheet_name: str,
    ebitda_row_key: str = "ebitda_row",
    closing_debt_row_key: str = "closing_row",
    interest_row_key: str = "interest_row",
    fcf_row_key: str = "fcf_row",
) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Covenants", title_it="Covenant",
        subtitle="Period-by-period headroom & breach monitor",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        is_hist = i < h
        c = ws.cell(row=yr_row, column=col_idx, value=f"{'A' if is_hist else 'E'} {yr}")
        styles.style_header(c)

    ebitda_row = int(operating_refs[ebitda_row_key])
    closing_row = int(debt_refs[closing_debt_row_key])
    interest_row = int(debt_refs[interest_row_key])
    fcf_row = int(operating_refs[fcf_row_key]) if fcf_row_key in operating_refs else None

    r = 7
    breach_rows: list[int] = []

    for cov in spec.covenants:
        layout.write_section_header(ws, r, cov.name.en, cov.name.secondary)
        r += 1

        actual_row = r
        if cov.kind == "leverage":
            layout.write_row_label(ws, r, L("leverage_ratio").en, L("leverage_ratio").secondary)
            ws.cell(row=r, column=3, value="x").font = styles.font_label_it
            for i in range(n):
                col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
                debt_ref = f"'{debt_sheet_name}'!{col}{closing_row}"
                ebitda_ref = f"'{operating_sheet_name}'!{col}{ebitda_row}"
                c = ws.cell(row=r, column=col_idx, value=ebitda_multiple(debt_ref, ebitda_ref))
                styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
        elif cov.kind == "icr":
            layout.write_row_label(ws, r, L("icr").en, L("icr").secondary)
            ws.cell(row=r, column=3, value="x").font = styles.font_label_it
            for i in range(n):
                col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
                interest_ref = f"'{debt_sheet_name}'!{col}{interest_row}"
                ebitda_ref = f"'{operating_sheet_name}'!{col}{ebitda_row}"
                c = ws.cell(row=r, column=col_idx,
                            value=interest_coverage(ebitda_ref, interest_ref))
                styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
        elif cov.kind == "dscr" and fcf_row is not None:
            layout.write_row_label(ws, r, "DSCR (FCF / Debt service)", "DSCR")
            ws.cell(row=r, column=3, value="x").font = styles.font_label_it
            for i in range(n):
                col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
                fcf_ref = f"'{operating_sheet_name}'!{col}{fcf_row}"
                interest_ref = f"'{debt_sheet_name}'!{col}{interest_row}"
                c = ws.cell(row=r, column=col_idx,
                            value=f"=IFERROR({fcf_ref}/ABS({interest_ref}),0)")
                styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
        elif cov.kind == "minimum_ebitda":
            layout.write_row_label(ws, r, "EBITDA absolute", "EBITDA assoluto")
            ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
            for i in range(n):
                col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
                ebitda_ref = f"'{operating_sheet_name}'!{col}{ebitda_row}"
                c = ws.cell(row=r, column=col_idx, value=f"={ebitda_ref}")
                styles.style_xref(c, number_format=styles.FMT_EUR_M)
        else:
            layout.write_row_label(ws, r, cov.name.en, cov.name.secondary)
        r += 1

        threshold_row = r
        layout.write_row_label(ws, r, f"{cov.name.en} — threshold",
                               f"{cov.name.secondary} — soglia", indent=True)
        ws.cell(row=r, column=3, value="x" if cov.kind != "minimum_ebitda" else spec.meta.currency).font = styles.font_label_it
        for i in range(n):
            col_idx = ord(layout.year_col(i)) - ord("A") + 1
            if i < h:
                ws.cell(row=r, column=col_idx, value="")
            else:
                proj_idx = i - h
                if proj_idx < len(cov.threshold_by_year):
                    a = cov.threshold_by_year[proj_idx]
                    c = ws.cell(row=r, column=col_idx, value=f"={a.name}")
                    nf = styles.FMT_EUR_M if cov.kind == "minimum_ebitda" else styles.FMT_MULTIPLE
                    styles.style_xref(c, number_format=nf)
        r += 1

        # v0.6: gate breach check on |interest| > 0 so drawdown and
        # post-maturity years with zero interest don't false-flag ICR.
        breach_row = r
        layout.write_row_label(ws, r, f"{cov.name.en} — breach",
                               f"{cov.name.secondary} — violazione", indent=True)
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < h:
                ws.cell(row=r, column=col_idx, value="")
            else:
                actual_ref = f"${col}${actual_row}"
                thr_ref = f"${col}${threshold_row}"
                direction = "max" if cov.kind == "leverage" else "min"
                if direction == "max":
                    core = f"{actual_ref}>{thr_ref}"
                else:
                    core = f"{actual_ref}<{thr_ref}"
                int_ref = f"'{debt_sheet_name}'!{col}{interest_row}"
                formula = f"=IF(AND({core},{int_ref}<-0.01),1,0)"
                c = ws.cell(row=r, column=col_idx, value=formula)
                styles.style_formula(c, number_format=styles.FMT_INTEGER)
                c.alignment = styles.align_center
        breach_rows.append(breach_row)
        r += 2

    # Aggregate
    layout.write_section_header(ws, r, "Aggregate breach counter", "Contatore violazioni")
    r += 1
    total_breach_row = r
    layout.write_row_label(ws, r, "Total breaches (projection)", "Violazioni totali")
    first_proj_col = layout.year_col(h); last_proj_col = layout.year_col(n - 1)
    parts = [f"SUM(${first_proj_col}${br}:${last_proj_col}${br})" for br in breach_rows]
    formula = "=" + "+".join(parts) if parts else "=0"
    c = ws.cell(row=r, column=3, value=formula)
    styles.style_formula(c, number_format=styles.FMT_INTEGER)
    c.font = styles.font_subheader

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {
        "total_breach_row": str(total_breach_row),
        "total_breach_cell": f"'{ws.title}'!$C${total_breach_row}",
    }
