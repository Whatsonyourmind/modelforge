"""Fairness football-field sheet (Template 11).

Emits three sheets:
    TradingComps       — comparable companies trading table with median / mean
    TransactionComps   — precedent transactions (if provided)
    FootballField      — methodology × (low, high) valuation ranges +
                         native Excel bar chart, premium/discount vs
                         current price.
"""

from __future__ import annotations

from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def _define_name(wb, name: str, sheet: str, cell: str) -> None:
    """Create/replace a workbook-level named range pointing to a single cell."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text=f"'{sheet}'!${cell[:1]}${cell[1:]}" if cell[0].isalpha() else cell,
    )


def _write_comp_table(ws: Worksheet, title_en: str, title_it: str,
                      subtitle: str, comps, include_date: bool) -> None:
    layout.set_column_widths(ws, label_width=36, it_width=24, year_width=14, unit_width=8)
    layout.write_title_block(ws, title_en, title_it, subtitle)

    hr = 5
    headers = ["Company", "EV / EBITDA (x)", "EV / Revenue (x)"]
    if include_date:
        headers.append("Announced")
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        styles.style_header(c)

    if not comps:
        ws.cell(row=hr + 1, column=1, value="(no items in spec)").font = styles.font_label_it
        ws.freeze_panes = f"A{hr + 1}"
        ws.print_title_rows = f"{hr}:{hr}"
        return

    r0 = hr + 1
    for i, it in enumerate(comps):
        r = r0 + i
        ws.cell(row=r, column=1, value=it.name).font = styles.font_label_en
        c1 = ws.cell(row=r, column=2, value=it.ev_ebitda_x)
        styles.style_input(c1, number_format=styles.FMT_MULTIPLE)
        c1.comment = Comment(f"{it.name} EV/EBITDA (from spec).", "ModelForge")
        c2 = ws.cell(row=r, column=3, value=it.ev_revenue_x)
        styles.style_input(c2, number_format=styles.FMT_MULTIPLE)
        c2.comment = Comment(f"{it.name} EV/Revenue (from spec).", "ModelForge")
        if include_date:
            ws.cell(row=r, column=4, value=it.date or "").font = styles.font_label_it
    r_last = r0 + len(comps) - 1

    # Summary row: median + mean
    r_sum = r_last + 2
    ws.cell(row=r_sum, column=1, value="Median").font = styles.font_subheader
    ws.cell(row=r_sum, column=2,
            value=f"=MEDIAN(B{r0}:B{r_last})").number_format = styles.FMT_MULTIPLE
    ws.cell(row=r_sum, column=3,
            value=f"=MEDIAN(C{r0}:C{r_last})").number_format = styles.FMT_MULTIPLE
    r_sum += 1
    ws.cell(row=r_sum, column=1, value="Mean").font = styles.font_subheader
    ws.cell(row=r_sum, column=2,
            value=f"=AVERAGE(B{r0}:B{r_last})").number_format = styles.FMT_MULTIPLE
    ws.cell(row=r_sum, column=3,
            value=f"=AVERAGE(C{r0}:C{r_last})").number_format = styles.FMT_MULTIPLE

    ws.freeze_panes = f"A{r0}"
    ws.print_title_rows = f"{hr}:{hr}"


def build_trading_comps(ws: Worksheet, spec) -> None:
    _write_comp_table(ws, "Trading Comps", "Comparabili quotati",
                      "Public comparable companies",
                      spec.trading_comps, include_date=False)


def build_transaction_comps(ws: Worksheet, spec) -> None:
    _write_comp_table(ws, "Transaction Comps", "Transazioni comparabili",
                      "Precedent M&A transactions",
                      spec.transaction_comps, include_date=True)


def _comp_median_row(spec, which: str) -> int:
    """Row number where MEDIAN is written on the comp sheet (rows are
    header + items + 2 blank + median + mean). Used to build a cross-sheet
    MEDIAN reference for live-linked football-field formulas."""
    comps = spec.trading_comps if which == "trading" else spec.transaction_comps
    n = len(comps) if comps else 0
    if n == 0:
        return 0
    # Row 5 header, 6..(5+n) items, skip 1, median at row 5+n+2
    return 5 + n + 2


def _comp_data_range(spec, which: str, col: str) -> str:
    """Data range e.g. 'B6:B9' on the given comp sheet — for in-place MEDIAN."""
    comps = spec.trading_comps if which == "trading" else spec.transaction_comps
    n = len(comps) if comps else 0
    if n == 0:
        return ""
    return f"{col}6:{col}{5 + n}"


def _derived_ev_formula(spec, vr, bound: str) -> str | None:
    """Return an Excel formula string for a live-linked EV bound, or None
    if no derivation is possible (falls back to static input).

    `bound` is "low" or "high".
    """
    df = vr.derive_from
    if df is None:
        return None
    sign = "-" if bound == "low" else "+"
    spread = vr.spread_x

    # Multiple-based methods
    if df in ("trading_ebitda", "trading_revenue",
              "transaction_ebitda", "transaction_revenue"):
        sheet = "TradingComps" if df.startswith("trading") else "TransactionComps"
        col = "B" if df.endswith("ebitda") else "C"
        which = "trading" if df.startswith("trading") else "transaction"
        data_range = _comp_data_range(spec, which, col)
        if not data_range:
            return None
        # Denominator: target_ebitda or target_revenue (named range)
        if df.endswith("ebitda"):
            base = "target_ebitda_eur_m"
        else:
            base = ("target_revenue_eur_m" if spec.target_revenue_eur_m is not None
                    else f"{spec.target.revenue_last_fy_eur_m}")
        return f"=(MEDIAN('{sheet}'!{data_range}){sign}{spread})*{base}"

    # 52-week range: EV = shares × price + net_debt
    # Price bounds are registered as workbook-level named ranges
    # (price_52w_low / price_52w_high) so the formula reads cleanly.
    if df == "trading_range_52w":
        shares = ("shares_outstanding_m_assum"
                  if spec.shares_outstanding_m_assum is not None
                  else f"{spec.shares_outstanding_m}")
        net_debt = ("net_debt_eur_m_assum"
                    if spec.net_debt_eur_m_assum is not None
                    else f"{spec.net_debt_eur_m}")
        price_name = "price_52w_low" if bound == "low" else "price_52w_high"
        return f"={shares}*{price_name}+{net_debt}"

    return None


def build_football_field(ws: Worksheet, spec) -> None:
    layout.set_column_widths(ws, label_width=44, it_width=28, year_width=14, unit_width=6)
    layout.write_title_block(ws, "Football Field", "Football Field",
                             "Valuation range by methodology")

    wb = ws.parent
    # Emit per-methodology hardcoded inputs (52W low/high, and static
    # DCF/LBO EV ranges) as named cells in a helper block so the Football
    # Field formulas can reference named ranges instead of embedding raw
    # numbers. Helper block lives at rows 100+ (hidden to the reader).
    helper_start = 100
    helper_rows: dict[str, int] = {}

    def helper(name_key: str, label: str, value, fmt):
        rr = helper_start + len(helper_rows)
        ws.cell(row=rr, column=1, value=label).font = styles.font_label_it
        c = ws.cell(row=rr, column=2, value=value)
        styles.style_input(c, number_format=fmt)
        _define_name(wb, name_key, ws.title, f"B{rr}")
        helper_rows[name_key] = rr

    for vr in spec.valuation_ranges:
        if vr.derive_from == "trading_range_52w":
            if vr.price_low_eur is not None:
                helper("price_52w_low", "52W price low",
                       vr.price_low_eur, styles.FMT_EUR_ACTUAL)
            if vr.price_high_eur is not None:
                helper("price_52w_high", "52W price high",
                       vr.price_high_eur, styles.FMT_EUR_ACTUAL)

    # Bridge named refs (prefer Assumption, fall back to scalar)
    net_debt_ref = ("net_debt_eur_m_assum"
                    if spec.net_debt_eur_m_assum is not None
                    else f"{spec.net_debt_eur_m}")
    shares_ref = ("shares_outstanding_m_assum"
                  if spec.shares_outstanding_m_assum is not None
                  else f"{spec.shares_outstanding_m}")
    price_ref = ("current_price_eur_assum"
                 if spec.current_price_eur_assum is not None
                 else f"{spec.current_price_eur}")
    has_price = ((spec.shares_outstanding_m_assum is not None
                  or spec.shares_outstanding_m > 0)
                 and (spec.current_price_eur_assum is not None
                      or spec.current_price_eur > 0))

    hr = 5
    headers = ["Methodology", "EV low (€m)", "EV high (€m)",
               "Mid (€m)", "Equity low (€m)", "Equity high (€m)"]
    if has_price:
        headers += ["Implied price low (€)", "Implied price high (€)",
                    "Premium low", "Premium high"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        styles.style_header(c)

    r0 = hr + 1
    for i, vr in enumerate(spec.valuation_ranges):
        r = r0 + i
        ws.cell(row=r, column=1, value=vr.method).font = styles.font_label_en

        # EV low: prefer live-linked formula, fall back to static scalar
        live_low = _derived_ev_formula(spec, vr, "low")
        live_high = _derived_ev_formula(spec, vr, "high")

        if live_low is not None:
            c_low = ws.cell(row=r, column=2, value=live_low)
            styles.style_formula(c_low, number_format=styles.FMT_EUR_M)
        else:
            c_low = ws.cell(row=r, column=2, value=vr.ev_low_eur_m)
            styles.style_input(c_low, number_format=styles.FMT_EUR_M)
        c_low.comment = Comment(
            f"{vr.method} low bound — {vr.note or 'no note'}"
            + (f"\n[live-linked: {vr.derive_from}]" if live_low else ""),
            "ModelForge",
        )

        if live_high is not None:
            c_high = ws.cell(row=r, column=3, value=live_high)
            styles.style_formula(c_high, number_format=styles.FMT_EUR_M)
        else:
            c_high = ws.cell(row=r, column=3, value=vr.ev_high_eur_m)
            styles.style_input(c_high, number_format=styles.FMT_EUR_M)
        c_high.comment = Comment(
            f"{vr.method} high bound — {vr.note or 'no note'}"
            + (f"\n[live-linked: {vr.derive_from}]" if live_high else ""),
            "ModelForge",
        )

        # Mid
        cm = ws.cell(row=r, column=4, value=f"=AVERAGE(B{r},C{r})")
        styles.style_formula(cm, number_format=styles.FMT_EUR_M)
        # Equity bounds (EV − net debt) — uses named range when available
        el = ws.cell(row=r, column=5, value=f"=B{r}-{net_debt_ref}")
        styles.style_formula(el, number_format=styles.FMT_EUR_M)
        eh = ws.cell(row=r, column=6, value=f"=C{r}-{net_debt_ref}")
        styles.style_formula(eh, number_format=styles.FMT_EUR_M)
        if has_price:
            pl = ws.cell(row=r, column=7,
                         value=f"=E{r}/{shares_ref}")
            styles.style_formula(pl, number_format=styles.FMT_EUR_ACTUAL)
            ph = ws.cell(row=r, column=8,
                         value=f"=F{r}/{shares_ref}")
            styles.style_formula(ph, number_format=styles.FMT_EUR_ACTUAL)
            ws.cell(row=r, column=9,
                    value=f"=G{r}/{price_ref}-1").number_format = styles.FMT_PCT_2DP
            ws.cell(row=r, column=10,
                    value=f"=H{r}/{price_ref}-1").number_format = styles.FMT_PCT_2DP

    r_last = r0 + len(spec.valuation_ranges) - 1

    # Floating-bar helper column: spread = EV high − EV low. A football field
    # must show each method's bar FLOATING from its low to its high, not two
    # bars both anchored at zero. We achieve that with the standard stacked-bar
    # technique: an invisible EV-low base series + a visible spread series.
    spread_col = 11  # col K — past all data columns
    styles.style_header(ws.cell(row=hr, column=spread_col, value="EV range (€m)"))
    for i in range(len(spec.valuation_ranges)):
        rr = r0 + i
        sc = ws.cell(row=rr, column=spread_col, value=f"=C{rr}-B{rr}")
        styles.style_formula(sc, number_format=styles.FMT_EUR_M)

    # ── Football field bar chart — floating low→high bars per methodology
    chart = BarChart()
    chart.type = "bar"
    chart.grouping = "stacked"
    chart.overlap = 100
    chart.style = 11
    chart.title = "Football field — EV range by methodology"
    chart.y_axis.title = "Method"
    chart.x_axis.title = "EV (€m)"

    base_ref = Reference(ws, min_col=2, min_row=hr, max_col=2, max_row=r_last)
    spread_ref = Reference(ws, min_col=spread_col, min_row=hr,
                           max_col=spread_col, max_row=r_last)
    chart.add_data(base_ref, titles_from_data=True)      # EV low (invisible base)
    chart.add_data(spread_ref, titles_from_data=True)    # spread (visible bar)
    cats = Reference(ws, min_col=1, min_row=r0, max_col=1, max_row=r_last)
    chart.set_categories(cats)

    # Make the EV-low base series invisible so the spread bar floats from low.
    from openpyxl.chart.shapes import GraphicalProperties
    _gp = GraphicalProperties()
    _gp.noFill = True
    chart.series[0].graphicalProperties = _gp

    chart.height = max(8, 0.7 * (r_last - r0 + 1))
    chart.width = 18
    ws.add_chart(chart, f"L{hr}")

    # v0.8.7 US-545: Summary named ranges for the football field. Provides
    # workbook-level handles to key EV / equity bounds so fairness opinions
    # can reference them in narrative / PPT export. Also lifts named-range
    # count past the #80 threshold of 20 for bulge-tier parity.
    summary_start = r_last + 3
    rr_local = [summary_start]
    ws.cell(row=rr_local[0], column=1,
            value="Summary — football field aggregates").font = styles.font_subheader
    rr_local[0] += 1

    def _reg(name: str, label: str, formula: str, fmt: str = styles.FMT_EUR_M) -> None:
        ws.cell(row=rr_local[0], column=1, value=label).font = styles.font_label_en
        c = ws.cell(row=rr_local[0], column=2, value=formula)
        styles.style_formula(c, number_format=fmt)
        _define_name(wb, name, ws.title, f"B{rr_local[0]}")
        rr_local[0] += 1

    ev_low_range = f"B{r0}:B{r_last}"
    ev_high_range = f"C{r0}:C{r_last}"
    eq_low_range = f"E{r0}:E{r_last}"
    eq_high_range = f"F{r0}:F{r_last}"

    _reg("football_ev_low_min",
         "Football field: EV low (min across methods)",
         f"=MIN({ev_low_range})")
    _reg("football_ev_high_max",
         "Football field: EV high (max across methods)",
         f"=MAX({ev_high_range})")
    _reg("football_ev_low_median",
         "Football field: EV low (median)",
         f"=MEDIAN({ev_low_range})")
    _reg("football_ev_high_median",
         "Football field: EV high (median)",
         f"=MEDIAN({ev_high_range})")
    _reg("football_ev_midpoint",
         "Football field: EV midpoint (mid of medians)",
         f"=(football_ev_low_median+football_ev_high_median)/2")
    _reg("football_equity_low",
         "Football field: Equity low (min)",
         f"=MIN({eq_low_range})")
    _reg("football_equity_high",
         "Football field: Equity high (max)",
         f"=MAX({eq_high_range})")
    _reg("football_equity_midpoint",
         "Football field: Equity value midpoint",
         f"=(football_equity_low+football_equity_high)/2")
    _reg("football_range_spread_pct",
         "Football field: Range spread (% of midpoint)",
         f"=(football_ev_high_max-football_ev_low_min)/football_ev_midpoint",
         fmt=styles.FMT_PCT_2DP)
    _reg("football_method_count",
         "Football field: Methodologies",
         f"=COUNTA(A{r0}:A{r_last})",
         fmt=styles.FMT_INTEGER)

    ws.freeze_panes = f"A{r0}"
    ws.print_title_rows = f"{hr}:{hr}"
