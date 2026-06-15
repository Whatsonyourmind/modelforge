"""Net interest income (average-balance convention).

Interest income on loans/securities and interest expense on deposits use the
AVERAGE of opening and closing balances (both volume-driven on the balance
sheet, so no circularity). Interest expense on WHOLESALE funding — the only
equity-dependent (plugged) balance — is charged on its BEGINNING-of-period
(prior-closing) balance, which breaks the funding→interest→earnings→capital→
funding cycle. Period 0 has no prior, so the average / BOP falls back to the
current column (the standard ``prior = year_col(i-1) if i>0 else col`` guard).

Reads the BalanceSheet gross-loans / securities / deposits / wholesale rows
(passed in via ``bs_refs``); written before P&L, which reads this sheet's NII
and average-gross-loans rows.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, bs_refs: dict[str, str], bs_sheet: str) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p
    cur = spec.meta.currency

    gl = int(bs_refs["gross_loans_row"])
    sec = int(bs_refs["securities_row"])
    dep = int(bs_refs["deposits_row"])
    whl = int(bs_refs["wholesale_row"])

    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Net Interest Income", title_it="Margine di interesse",
        subtitle=f"Average-balance convention · NIM · {spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        cc = ws.cell(row=yr_row, column=ci, value=f"{'A' if i < h else 'E'} {yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    def _bs(col, row):
        return f"'{bs_sheet}'!${col}${row}"

    # ── Interest-earning assets ───────────────────────────────────────────
    layout.write_section_header(ws, r, "Interest-earning assets", "Attività fruttifere")
    r += 1

    rows["avg_loans"] = r
    layout.write_row_label(ws, r, "Average gross loans", "Crediti lordi medi", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        c = ws.cell(row=r, column=ci, value=f"=({_bs(prior, gl)}+{_bs(col, gl)})/2")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["avg_securities"] = r
    layout.write_row_label(ws, r, "Average securities", "Titoli medi", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        c = ws.cell(row=r, column=ci, value=f"=({_bs(prior, sec)}+{_bs(col, sec)})/2")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ii_loans"] = r
    layout.write_row_label(ws, r, "Interest income — loans", "Interessi attivi — crediti")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=${col}${rows['avg_loans']}*loan_yield")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ii_securities"] = r
    layout.write_row_label(ws, r, "Interest income — securities", "Interessi attivi — titoli")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=${col}${rows['avg_securities']}*securities_yield")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_ii"] = r
    layout.write_row_label(ws, r, "Total interest income", "Totale interessi attivi")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['ii_loans']}+${col}${rows['ii_securities']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # ── Interest-bearing liabilities ──────────────────────────────────────
    layout.write_section_header(ws, r, "Interest-bearing liabilities", "Passività onerose")
    r += 1

    rows["avg_deposits"] = r
    layout.write_row_label(ws, r, "Average deposits", "Depositi medi", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        c = ws.cell(row=r, column=ci, value=f"=({_bs(prior, dep)}+{_bs(col, dep)})/2")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["bop_wholesale"] = r
    layout.write_row_label(ws, r, "Wholesale funding — BOP", "Funding wholesale — inizio periodo", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        c = ws.cell(row=r, column=ci, value=f"={_bs(prior, whl)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ie_deposits"] = r
    layout.write_row_label(ws, r, "Interest expense — deposits", "Interessi passivi — depositi")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=-${col}${rows['avg_deposits']}*deposit_cost")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ie_wholesale"] = r
    layout.write_row_label(ws, r, "Interest expense — wholesale", "Interessi passivi — wholesale")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=-${col}${rows['bop_wholesale']}*wholesale_cost")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_ie"] = r
    layout.write_row_label(ws, r, "Total interest expense", "Totale interessi passivi")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['ie_deposits']}+${col}${rows['ie_wholesale']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # ── NII + NIM ─────────────────────────────────────────────────────────
    rows["nii"] = r
    layout.write_row_label(ws, r, "Net interest income (NII)", "Margine di interesse (NII)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['total_ii']}+${col}${rows['total_ie']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    rows["avg_earning_assets"] = r
    layout.write_row_label(ws, r, "Average earning assets", "Attività fruttifere medie", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['avg_loans']}+${col}${rows['avg_securities']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["nim"] = r
    layout.write_row_label(ws, r, "Net interest margin (NIM)", "NIM")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=IFERROR(${col}${rows['nii']}/${col}${rows['avg_earning_assets']},0)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["risk_free_memo"] = r
    layout.write_row_label(ws, r, "Reference risk-free rate (memo)", "Tasso risk-free (memo)", indent=True)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value="=risk_free_rate")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {f"{k}_row": str(v) for k, v in rows.items()}
