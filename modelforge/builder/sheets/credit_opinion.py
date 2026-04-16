"""Credit Opinion sheet — for credit memo template.

Contains:
    - Credit strengths / weaknesses / mitigating factors (narrative)
    - Recovery waterfall at default (stress EV → senior → mezz → equity)
    - Expected loss = PD × LGD × EAD
    - Rating shadow mapping
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.i18n import L


def build(ws: Worksheet, spec, operating_refs: dict[str, str],
          debt_refs: dict[str, str], operating_sheet: str, debt_sheet: str) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=13, unit_width=6)
    layout.write_title_block(
        ws, title_en="Credit Opinion",
        title_it="Opinione di credito",
        subtitle=f"Rating shadow: internal {spec.rating.internal_rating} · "
                 f"Moody's {spec.rating.moodys_equivalent} · S&P {spec.rating.sp_equivalent}",
    )
    layout.write_scenario_banner(ws, row=3)

    r = 7

    # Narrative blocks
    layout.write_section_header(ws, r, "Credit strengths", "Punti di forza")
    r += 1
    for s in spec.credit_strengths:
        ws.cell(row=r, column=1, value=f"• {s}").font = styles.font_label_en
        r += 1
    r += 1

    layout.write_section_header(ws, r, "Credit weaknesses", "Punti di debolezza")
    r += 1
    for s in spec.credit_weaknesses:
        ws.cell(row=r, column=1, value=f"• {s}").font = styles.font_warning
        r += 1
    r += 1

    layout.write_section_header(ws, r, "Mitigating factors", "Fattori mitiganti")
    r += 1
    for s in spec.mitigating_factors:
        ws.cell(row=r, column=1, value=f"• {s}").font = styles.font_label_en
        r += 1
    r += 1

    # Recovery waterfall (at stress EBITDA × stress multiple)
    layout.write_section_header(ws, r, "Recovery waterfall (stress scenario)",
                                "Waterfall di recupero (stress)")
    r += 1

    # Use last projected year EBITDA as baseline
    last_col = layout.year_col(n - 1)
    ebitda_row = int(operating_refs["ebitda_row"])
    closing_row = int(debt_refs["total_closing_row"])

    # Stress EBITDA
    layout.write_row_label(ws, r, "Base EBITDA (at default year)", "EBITDA base (anno default)")
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    base_ebitda_row = r
    c = ws.cell(row=r, column=4, value=f"='{operating_sheet}'!{last_col}{ebitda_row}")
    styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "× (1 + stress %)", "× (1 + stress %)", indent=True)
    stressed_ebitda_row = r
    c = ws.cell(row=r, column=4, value=f"=$D${base_ebitda_row}*(1+ebitda_stress_pct)")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Stress exit multiple", "Multiplo di uscita in stress")
    c = ws.cell(row=r, column=4, value="=stress_ebitda_multiple")
    styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
    stress_mult_row = r
    r += 1

    layout.write_row_label(ws, r, "Enterprise Value (stress)", "Enterprise Value (stress)")
    ev_row = r
    c = ws.cell(row=r, column=4, value=f"=$D${stressed_ebitda_row}*$D${stress_mult_row}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    r += 1

    layout.write_row_label(ws, r, "× (1 − liquidation discount)", "× (1 − sconto liquidazione)", indent=True)
    ev_after_disc_row = r
    c = ws.cell(row=r, column=4, value=f"=$D${ev_row}*(1-liquidation_discount_pct)")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Recoverable to creditors", "Recuperabile ai creditori")
    recoverable_row = r
    c = ws.cell(row=r, column=4, value=f"=$D${ev_after_disc_row}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.border = styles.BORDER_TOP_THIN
    r += 2

    # Senior recovery
    layout.write_row_label(ws, r, "Senior debt outstanding (at default)", "Debito senior (al default)")
    senior_outstanding_row = r
    c = ws.cell(row=r, column=4, value=f"='{debt_sheet}'!{last_col}{closing_row}")
    styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Senior recovery (€m)", "Recupero senior (€m)")
    senior_recovery_row = r
    c = ws.cell(row=r, column=4,
                value=f"=MIN($D${recoverable_row},$D${senior_outstanding_row})")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Senior recovery %", "Recupero senior %")
    c = ws.cell(row=r, column=4,
                value=f"=IFERROR($D${senior_recovery_row}/$D${senior_outstanding_row},0)")
    styles.style_formula(c, number_format=styles.FMT_PCT)
    c.font = styles.font_subheader
    r += 1

    # LGD
    layout.write_row_label(ws, r, "LGD (1 − recovery %)", "LGD (1 − recupero %)")
    lgd_row = r
    c = ws.cell(row=r, column=4,
                value=f"=1-IFERROR($D${senior_recovery_row}/$D${senior_outstanding_row},0)")
    styles.style_formula(c, number_format=styles.FMT_PCT)
    r += 2

    # Expected loss
    layout.write_section_header(ws, r, "Expected loss estimate", "Stima perdita attesa")
    r += 1

    layout.write_row_label(ws, r, "PD (probability of default, life)", "PD (probabilità default, vita)")
    pd_row = r
    c = ws.cell(row=r, column=4, value="=probability_of_default_pct")
    styles.style_xref(c, number_format=styles.FMT_PCT)
    r += 1

    layout.write_row_label(ws, r, "LGD (Loss Given Default)", "LGD")
    c = ws.cell(row=r, column=4, value=f"=$D${lgd_row}")
    styles.style_xref(c, number_format=styles.FMT_PCT)
    r += 1

    layout.write_row_label(ws, r, "EAD (Exposure at Default)", "EAD")
    ead_row = r
    c = ws.cell(row=r, column=4, value=f"='{debt_sheet}'!{last_col}{closing_row}")
    styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Expected Loss (PD × LGD × EAD)", "Perdita attesa")
    c = ws.cell(row=r, column=4,
                value=f"=$D${pd_row}*$D${lgd_row}*$D${ead_row}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.border = styles.BORDER_TOP_THIN
    c.comment = Comment(
        "Expected Loss = PD × LGD × EAD. Basel-style credit metric. "
        "PD is life-of-instrument, LGD from recovery waterfall above.",
        "ModelForge",
    )
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {"recovery_pct_row": str(senior_recovery_row), "lgd_row": str(lgd_row)}
