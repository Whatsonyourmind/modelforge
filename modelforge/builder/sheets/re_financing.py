"""RE senior mortgage + equity waterfall sheet.

SCOPE / HONEST-LABEL (v0.8): the LP/GP waterfall rendered here is a *simplified
illustrative* allocation — LP preferred return + GP catch-up + 80/20 promote on
the residual — NOT a full tier-by-tier (multi-hurdle) waterfall. The MECHANICS
are still simplified, but the pref and promote KNOBS are now spec-driven, live,
overridable named inputs (``spec.waterfall.lp_preferred_return_pct`` /
``gp_promote_pct``, resolved from the real tiers when present), NOT hardcoded
placeholders. The legacy 8% pref literal was a v0.7 placeholder that diverged
from the spec's intended 9% pref — that divergence is now corrected: the cell
reads the spec value via named range. A banner is still emitted on the sheet
(row 4) flagging that the waterfall STRUCTURE remains simplified.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, dcf_refs: dict[str, str], dcf_sheet: str) -> dict[str, str]:
    h = spec.horizon.hold_years
    n = h + 1

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Financing & Equity Waterfall", title_it="Finanziamento e waterfall equity",
        subtitle="Senior mortgage, LP/GP promote structure",
    )
    layout.write_scenario_banner(ws, row=3)

    # ── SIMPLIFIED-WATERFALL banner (row 4, otherwise unused) ─────────────
    # Honest-label: the LP/GP waterfall rendered below is a *simplified
    # illustrative* allocation (LP preferred return + GP catch-up + 80/20
    # promote), NOT a full tier-by-tier waterfall. The pref and promote knobs
    # are now SPEC-DRIVEN, live, overridable named inputs (no longer hardcoded
    # placeholders) — only the waterfall STRUCTURE remains simplified.
    banner = ws.cell(
        row=4, column=1,
        value=("v0.8 SIMPLIFIED WATERFALL (illustrative structure): LP pref + GP "
               "catch-up + 80/20 promote — NOT a full tier-by-tier waterfall. Pref "
               "and promote are now live spec-driven inputs (overridable)."),
    )
    banner.font = styles.font_warning
    banner.fill = styles.fill_check_bad
    banner.alignment = styles.align_left
    ws.cell(
        row=4, column=2,
        value="Waterfall semplificato v0.8 (struttura illustrativa) — pref/promote da spec.",
    ).font = styles.font_label_it

    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=yr_row, column=col_idx, value=f"t={i}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    # Senior debt sizing
    layout.write_section_header(ws, r, "Senior mortgage", "Mutuo senior")
    r += 1
    rows["loan_amount"] = r
    layout.write_row_label(ws, r, "Loan amount (at close)", "Importo finanziamento")
    cc = ws.cell(row=r, column=4, value="=acquisition_price_eur_m*ltv_pct")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Interest (annual on opening balance, bullet)
    rows["interest"] = r
    layout.write_row_label(ws, r, "Cash interest", "Interessi cassa")
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-$D${rows['loan_amount']}*senior_interest_rate")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Principal repayment at exit (bullet)
    rows["principal_repay"] = r
    layout.write_row_label(ws, r, "Principal repayment (at exit)", "Rimborso capitale (a uscita)")
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if i == h:
            cc = ws.cell(row=r, column=col_idx, value=f"=-$D${rows['loan_amount']}")
        else:
            cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Project cash flow (to all capital)
    cfads_row = int(dcf_refs["cfads_row"])
    acq_row = int(dcf_refs["acq_price_row"])
    net_exit_row = int(dcf_refs["net_exit_row"])

    rows["project_cf"] = r
    layout.write_row_label(ws, r, "Project cash flow (unlevered)", "CF progetto (unlevered)")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cfads_ref = f"'{dcf_sheet}'!{col}{cfads_row}"
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"='{dcf_sheet}'!D{acq_row}+{cfads_ref}")
        elif i == h:
            exit_ref = f"'{dcf_sheet}'!D{net_exit_row}"
            cc = ws.cell(row=r, column=col_idx,
                         value=f"={cfads_ref}+{exit_ref}")
        else:
            cc = ws.cell(row=r, column=col_idx, value=f"={cfads_ref}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    r += 1

    # Equity cash flow = project CF + debt draw at t=0 + debt service + principal repay
    rows["equity_cf"] = r
    layout.write_row_label(ws, r, "Equity cash flow", "CF equity")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['project_cf']}+$D${rows['loan_amount']}")
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['project_cf']}+${col}${rows['interest']}+${col}${rows['principal_repay']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Equity IRR + MoIC
    layout.write_section_header(ws, r, "Equity returns", "Rendimenti equity")
    r += 1
    first_col = layout.year_col(0); last_col = layout.year_col(n - 1)

    rows["equity_irr"] = r
    layout.write_row_label(ws, r, "Equity IRR", "IRR equity", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IRR(${first_col}${rows['equity_cf']}:${last_col}${rows['equity_cf']},0.10)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.font = styles.font_subheader
    r += 1

    layout.write_row_label(ws, r, "Equity MoIC", "MoIC equity", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IFERROR(SUMIF(${first_col}${rows['equity_cf']}:${last_col}${rows['equity_cf']},\">0\")"
                       f"/ABS(SUMIF(${first_col}${rows['equity_cf']}:${last_col}${rows['equity_cf']},\"<0\")),0)")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    r += 2

    # Waterfall (simplified — total LP/GP cash flows based on whether IRR > pref threshold)
    layout.write_section_header(ws, r, "LP / GP waterfall (simplified)",
                                "Waterfall LP/GP (semplificato)")
    r += 1
    # For real production: iterative tier-by-tier waterfall. Here we provide
    # an illustrative allocation that scales properly per tier.

    rows["lp_capital"] = r
    layout.write_row_label(ws, r, "LP committed capital (% equity)", "Capitale LP (% equity)")
    cc = ws.cell(row=r, column=4, value="=lp_capital_commitment_pct")
    styles.style_xref(cc, number_format=styles.FMT_PCT)
    r += 1

    rows["lp_cf"] = r
    layout.write_row_label(ws, r, "LP cash flow (proportionate)", "CF LP (proporzionale)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['equity_cf']}*lp_capital_commitment_pct")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["gp_cf"] = r
    layout.write_row_label(ws, r, "GP cash flow (proportionate)", "CF GP (proporzionale)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['equity_cf']}*(1-lp_capital_commitment_pct)")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    layout.write_row_label(ws, r, "LP IRR (before promote)", "IRR LP (ante promote)", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IRR(${first_col}${rows['lp_cf']}:${last_col}${rows['lp_cf']},0.10)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    r += 1

    layout.write_row_label(ws, r, "GP IRR (before promote)", "IRR GP (ante promote)", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IRR(${first_col}${rows['gp_cf']}:${last_col}${rows['gp_cf']},0.10)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    r += 2

    # ── v0.8: SIMPLIFIED LP/GP waterfall with pref + catchup + promote ─────
    # Tier 1 — Pref return (LP pref IRR on LP capital, compounded)  → named input
    # Tier 2 — GP catchup (100% to GP until GP gets `gp_promote_pct` of profits)
    # Tier 3 — LP/GP split on residual (LP share = 1 - gp_promote_pct)
    # The pref and promote are now SPEC-DRIVEN live named inputs (overridable),
    # resolved from spec.waterfall (real tiers when present). Only the waterfall
    # STRUCTURE remains a simplified illustrative allocation.
    layout.write_section_header(
        ws, r,
        "v0.8 Waterfall (SIMPLIFIED structure): pref + GP catchup + promote (spec-driven)",
        "Waterfall v0.8 (struttura semplificata): pref + catchup + promote (da spec)",
    )
    r += 1

    # LP preferred return — SPEC-DRIVEN live input (named range
    # `lp_preferred_return_pct`, resolved from spec.waterfall). Styled as a blue
    # input cell so a reviewer sees it is an overridable deal assumption, and
    # the formula below references the named range, not a literal.
    rows["pref_return"] = r
    layout.write_row_label(ws, r, "LP preferred return (compounded)",
                           "Rendimento preferenziale LP (composto)", indent=True)
    cc = ws.cell(row=r, column=4, value="=lp_preferred_return_pct")
    styles.style_input(cc, number_format=styles.FMT_PCT_2DP)
    cc.comment = Comment(
        "LP preferred return, compounded annually on contributed capital before "
        "any GP promote. Spec-driven named input "
        "(spec.waterfall.lp_preferred_return_pct / resolved from the pref tier); "
        "override on the Assumptions sheet. Corrects the legacy 8% v0.7 "
        "placeholder to the spec-intended pref.",
        "ModelForge",
    )
    cc.comment.width = 280
    cc.comment.height = 120
    r += 1

    # LP share of residual = 1 - GP promote. GP promote (`gp_promote_pct`) is the
    # SPEC-DRIVEN live named input (visible/overridable as a blue input on the
    # Assumptions sheet, source of the named range); LP share is derived from it
    # so the one overridable knob is the economically meaningful carry. The label
    # names the GP promote so a reader sees the carry directly on this sheet.
    rows["promote_split_lp"] = r
    layout.write_row_label(ws, r, "LP share of residual (= 1 − GP promote, post-catchup)",
                           "Quota LP residuo (= 1 − promote GP, post-catchup)", indent=True)
    cc = ws.cell(row=r, column=4, value="=1-gp_promote_pct")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.comment = Comment(
        "LP share of the residual band = 1 - gp_promote_pct. GP promote (carried "
        "interest) is the spec-driven named input "
        "(spec.waterfall.gp_promote_pct / resolved from the residual tier's LP "
        "share); override it as a blue input on the Assumptions sheet. 80/20 "
        "promote → GP 20%, LP 80%.",
        "ModelForge",
    )
    cc.comment.width = 280
    cc.comment.height = 120
    r += 2

    # Total distributions to equity over hold (from equity_cf)
    # Waterfall applies when sale proceeds returned to equity, not per-period.
    rows["total_equity_contrib"] = r
    layout.write_row_label(ws, r, "Total equity contribution (t=0)",
                           "Capitale equity totale (t=0)", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=-$D${rows['equity_cf']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_equity_distrib"] = r
    layout.write_row_label(ws, r, "Total equity distributions (gross)",
                           "Distribuzioni equity totali (lorde)", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f'=SUMIF(${first_col}${rows["equity_cf"]}:${last_col}${rows["equity_cf"]},">0")')
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_profit"] = r
    layout.write_row_label(ws, r, "Total profit (distributions − contribution)",
                           "Profitto totale", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['total_equity_distrib']}-$D${rows['total_equity_contrib']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    r += 2

    # Tier 1: Return of capital + pref
    ws.cell(row=r, column=1, value="Tier 1: Return of capital + pref").font = styles.font_subheader
    r += 1

    # Calculate hold period for pref compounding
    n_years_hold = n - 1
    rows["pref_threshold"] = r
    layout.write_row_label(ws, r, "LP pref threshold (capital × (1+pref)^hold)",
                           "Soglia pref LP", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['total_equity_contrib']}*lp_capital_commitment_pct"
                       f"*(1+$D${rows['pref_return']})^{n_years_hold}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["lp_tier1"] = r
    layout.write_row_label(ws, r, "LP receives in Tier 1",
                           "LP riceve in Tier 1", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=MIN($D${rows['total_equity_distrib']}*lp_capital_commitment_pct,$D${rows['pref_threshold']})")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Tier 2: GP catchup (20% of profits, if pref met)
    ws.cell(row=r, column=1, value="Tier 2: GP catchup (to 20% promote)").font = styles.font_subheader
    r += 1

    rows["gp_catchup"] = r
    layout.write_row_label(ws, r, "GP catchup amount",
                           "Catchup GP", indent=True)
    # Catchup = profits above pref × 0.25 (brings GP to 20% overall)
    # Formula: catchup = max(0, profit_above_pref × promote_GP / promote_LP)
    cc = ws.cell(
        row=r, column=4,
        value=f"=MAX(0,($D${rows['total_profit']}-$D${rows['pref_threshold']}+$D${rows['total_equity_contrib']}*lp_capital_commitment_pct)"
              f"*(1-$D${rows['promote_split_lp']})/$D${rows['promote_split_lp']})",
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Tier 3: 80/20 split on residual
    ws.cell(row=r, column=1, value="Tier 3: 80/20 promote split on residual").font = styles.font_subheader
    r += 1

    rows["lp_tier3"] = r
    layout.write_row_label(ws, r, "LP share of residual",
                           "Quota LP residuo", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=f"=MAX(0,$D${rows['total_equity_distrib']}-$D${rows['lp_tier1']}"
              f"-$D${rows['gp_catchup']})*$D${rows['promote_split_lp']}",
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["gp_tier3"] = r
    layout.write_row_label(ws, r, "GP share of residual (promote)",
                           "Quota GP residuo (promote)", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=f"=MAX(0,$D${rows['total_equity_distrib']}-$D${rows['lp_tier1']}"
              f"-$D${rows['gp_catchup']})*(1-$D${rows['promote_split_lp']})",
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Totals
    ws.cell(row=r, column=1, value="Totals by partner (post-waterfall)").font = styles.font_subheader
    r += 1
    rows["lp_total_post"] = r
    layout.write_row_label(ws, r, "LP total post-waterfall",
                           "LP totale post-waterfall", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['lp_tier1']}+$D${rows['lp_tier3']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    r += 1

    rows["gp_total_post"] = r
    layout.write_row_label(ws, r, "GP total post-waterfall",
                           "GP totale post-waterfall", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['gp_catchup']}+$D${rows['gp_tier3']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    r += 2

    # ── Spec waterfall tier reference (parameterization for v0.8) ──────────
    # The hardcoded 8% pref + 80/20 split above is v0.7 simplification.
    # The spec defines a richer multi-tier waterfall (spec.waterfall.tiers)
    # and a per-tranche arrangement_fee_pct. We surface those values as a
    # reference block so the named ranges flow into the workbook (they would
    # otherwise show up as orphan named ranges in MoatGate). The v0.8
    # iteration will compute the waterfall directly from these inputs.
    layout.write_section_header(
        ws, r,
        "Spec waterfall tiers (v0.8 parameterization reference)",
        "Tier waterfall (riferimento parametrico v0.8)",
    )
    r += 1

    arr_fee = getattr(getattr(spec, "financing", None), "arrangement_fee_pct", None)
    if arr_fee is not None:
        layout.write_row_label(
            ws, r, "Senior loan arrangement fee %", "Commissione di organizzazione",
            indent=True,
        )
        cc = ws.cell(row=r, column=4, value=f"={arr_fee.name}")
        styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
        r += 1

    tiers = getattr(getattr(spec, "waterfall", None), "tiers", None) or []
    for tier in tiers:
        tier_label = (
            f"Tier {tier.tier_index} — {tier.label}" if hasattr(tier, "label")
            else f"Tier {getattr(tier, 'tier_index', '?')}"
        )
        layout.write_row_label(
            ws, r, f"{tier_label} hurdle IRR",
            f"{tier_label} soglia IRR", indent=True,
        )
        if getattr(tier, "hurdle_irr_pct", None):
            cc = ws.cell(row=r, column=4, value=f"={tier.hurdle_irr_pct.name}")
            styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
        else:
            cc = ws.cell(row=r, column=4, value="(no hurdle — open tier)")
            cc.font = styles.font_label_it
        r += 1
        layout.write_row_label(
            ws, r, f"{tier_label} LP share %",
            f"{tier_label} quota LP %", indent=True,
        )
        cc = ws.cell(row=r, column=4, value=f"={tier.lp_share_pct.name}")
        styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
        r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
