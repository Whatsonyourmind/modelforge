"""Bank P&L — NII + fees + trading − opex − provisions → PPOP/PBT/NI + ROE/ROTE.

Reads NII (and average gross loans for the impairment charge) from the NII
sheet, and the BalanceSheet common-equity / intangibles rows for the return
ratios. Historical fee income is the PLUG that ties total operating income to
the reported historical total income; projected fees grow at ``fee_income_growth``.
The historical reported net income is shown as a memo (the P&L itself is fully
driver-computed, costs-negative).
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, nii_refs: dict[str, str], bs_refs: dict[str, str],
          nii_sheet: str, bs_sheet: str) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p
    cur = spec.meta.currency

    nii_row = int(nii_refs["nii_row"])
    avg_loans_row = int(nii_refs["avg_loans_row"])
    eq_row = int(bs_refs["equity_closing_row"])
    intang_row = int(bs_refs["intangibles_row"])

    hist_income = spec.historical_total_income_eur_m or [0.0] * h
    hist_ni = spec.historical_net_income_eur_m or [0.0] * h

    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Income Statement", title_it="Conto economico",
        subtitle=f"NII · fees · provisions · PPOP/PBT/NI · ROE/ROTE · "
                 f"{spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        cc = ws.cell(row=yr_row, column=ci, value=f"{'A' if i < h else 'E'} {yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    def _nii(col, row):
        return f"'{nii_sheet}'!${col}${row}"

    def _bs(col, row):
        return f"'{bs_sheet}'!${col}${row}"

    layout.write_section_header(ws, r, "Income statement", "Conto economico")
    r += 1

    # Net interest income (xref)
    rows["nii"] = r
    layout.write_row_label(ws, r, "Net interest income", "Margine di interesse")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"={_nii(col, nii_row)}")
        styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Net trading income (level driver, all columns)
    rows["trading"] = r
    layout.write_row_label(ws, r, "Net trading income", "Risultato di negoziazione", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value="=trading_income_eur_m")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Net fee & commission income: historical = plug to reported total income;
    # projection = prior × (1+fee_income_growth).
    rows["fees"] = r
    layout.write_row_label(ws, r, "Net fee & commission income", "Commissioni nette")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            ti = hist_income[i] if i < len(hist_income) else 0.0
            c = ws.cell(row=r, column=ci,
                        value=(f"={ti}-${col}${rows['nii']}-${col}${rows['trading']}"))
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}*(1+fee_income_growth)")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Total operating income
    rows["total_income"] = r
    layout.write_row_label(ws, r, "Total operating income", "Margine di intermediazione")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=(f"=${col}${rows['nii']}+${col}${rows['fees']}"
                           f"+${col}${rows['trading']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # Operating expenses = -C/I × total income. MAX(total_income,0) keeps opex a
    # COST (≤ 0) even in a deep-loss period where total operating income turns
    # negative — without the floor, -(negative)×C/I would flip opex to a credit
    # (certify-clean but economically wrong, and a self-inflicted opex≤0 QC fail).
    # For a going-concern bank total income is strongly positive, so this is
    # byte-identical to the unguarded form in every realistic scenario.
    rows["opex"] = r
    layout.write_row_label(ws, r, "Operating expenses", "Costi operativi", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=-MAX(${col}${rows['total_income']},0)*cost_income_ratio")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # PPOP
    rows["ppop"] = r
    layout.write_row_label(ws, r, "Pre-provision operating profit (PPOP)", "Risultato lordo di gestione")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['total_income']}+${col}${rows['opex']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # Loan-loss provisions = -CoR/1e4 × avg gross loans (negative)
    rows["provisions"] = r
    layout.write_row_label(ws, r, "Loan-loss provisions (impairment)", "Rettifiche su crediti", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=-cost_of_risk_bps/10000*{_nii(col, avg_loans_row)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # PBT
    rows["pbt"] = r
    layout.write_row_label(ws, r, "Profit before tax (PBT)", "Utile ante imposte")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['ppop']}+${col}${rows['provisions']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # Tax = -MAX(PBT,0) × tax_rate (no tax credit on losses)
    rows["tax"] = r
    layout.write_row_label(ws, r, "Tax", "Imposte", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=-MAX(${col}${rows['pbt']},0)*tax_rate")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Net income
    rows["ni"] = r
    layout.write_row_label(ws, r, "Net income (NI)", "Utile netto")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['pbt']}+${col}${rows['tax']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # AT1 coupon (negative; post-NI deduction from what builds CET1)
    rows["at1_coupon"] = r
    layout.write_row_label(ws, r, "(−) AT1 coupon", "(−) Cedola AT1", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value="=-at1_coupon_eur_m")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # NI attributable to CET1 (= NI − AT1 coupon)
    rows["ni_to_cet1"] = r
    layout.write_row_label(ws, r, "NI attributable to CET1", "Utile attribuibile a CET1")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['ni']}+${col}${rows['at1_coupon']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 2

    # ── Returns ───────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "Returns", "Redditività")
    r += 1

    rows["avg_equity"] = r
    layout.write_row_label(ws, r, "Average common equity", "Patrimonio comune medio", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        c = ws.cell(row=r, column=ci, value=f"=({_bs(prior, eq_row)}+{_bs(col, eq_row)})/2")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["avg_tangible_equity"] = r
    layout.write_row_label(ws, r, "Average tangible common equity", "Patrimonio tangibile medio", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1) if i > 0 else col
        avg_intang = f"({_bs(prior, intang_row)}+{_bs(col, intang_row)})/2"
        c = ws.cell(row=r, column=ci, value=f"=${col}${rows['avg_equity']}-{avg_intang}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # ROE numerator is NI ATTRIBUTABLE TO COMMON (NI − AT1 coupon), matching the
    # ROTE numerator: the AT1 coupon is a distribution to AT1 holders, not common
    # shareholders, so return to common equity must net it out (equity-research
    # convention; keeps ROE and ROTE internally consistent).
    rows["roe"] = r
    layout.write_row_label(ws, r, "ROE (NI to common / avg common equity)", "ROE")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=IFERROR(${col}${rows['ni_to_cet1']}/${col}${rows['avg_equity']},0)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["rote"] = r
    layout.write_row_label(ws, r, "ROTE (NI to CET1 / avg tangible equity)", "ROTE")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=IFERROR(${col}${rows['ni_to_cet1']}/${col}${rows['avg_tangible_equity']},0)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 2

    # Reported actuals (memo, historical only)
    layout.write_section_header(ws, r, "Reported actuals (memo)", "Dati riportati (memo)")
    r += 1
    rows["reported_ni"] = r
    layout.write_row_label(ws, r, "Reported net income (memo)", "Utile netto riportato (memo)", indent=True)
    _unit(r, cur)
    for i in range(h):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        val = hist_ni[i] if i < len(hist_ni) else 0.0
        c = ws.cell(row=r, column=ci, value=val)
        styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {f"{k}_row": str(v) for k, v in rows.items()}
