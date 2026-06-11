"""Italian / EU regulatory compliance checks.

Adds a dedicated `ComplianceCheck` sheet to any credit / PF / NPL / SC
workbook. Encodes the quantitative tests a BaFin / ECB / Banca d'Italia
supervisor would run on AIFM activity or a Basel NPL / securitization
capital computation.

Sections:
  1. AIFMD II leverage & concentration tests (Art. 15a + Art. 17)
  2. Loan-originating AIF classification (>50% NAV test)
  3. IFRS 9 three-stage ECL recap (summary)
  4. Basel III/IV NPL calendar provisioning (Reg. EU 2019/630)
  5. GACS eligibility (NPL only)
  6. IRES + IRAP split (tax transparency)

References cited inline:
  - BCLP "AIFMD II Leverage Limits" (April 2024 transposition)
  - Linklaters "Loan Origination under AIFMD II"
  - Clifford Chance / Dechert guidance
  - BIS IFRS 9 FSI summary
  - Regulation (EU) 2019/630 (calendar provisioning)
  - KPMG / Jones Day GACS
  - PwC Italy IRES/IRAP tax summaries
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.spec.compliance import ComplianceContext


def build(ws: Worksheet, spec, context: dict | None = None) -> None:
    """Emit ComplianceCheck sheet.

    `context` may carry fund-level parameters (NAV, total_commitments) when
    applicable; absent fields render as informational placeholders.

    v0.12 — regulatory & tax constants lifted from hardcodes to spec-driven,
    overridable named-input cells. Resolution order for each value:
        1. spec.compliance.<field>  (a ComplianceContext block, if present)
        2. the legacy `context` dict key (back-compat for direct callers)
        3. the field's default (= the previously-hardcoded value)
    All defaults preserve byte-identical output.
    """
    context = context or {}

    # Resolve the compliance/tax context. Templates call build(ws, spec) with
    # no context dict, so `spec.compliance` (when present) is the source of
    # truth; otherwise every field falls back to its hardcoded-equal default.
    cc = getattr(spec, "compliance", None)
    if not isinstance(cc, ComplianceContext):
        cc = ComplianceContext()

    # Legacy `context` dict still wins for the three scenario inputs it used to
    # drive (keeps any direct caller that passes a context working unchanged).
    aif_type_default = context.get("aif_type", cc.aif_type)
    actual_leverage_default = context.get("actual_leverage", cc.actual_leverage)
    largest_borrower_default = context.get(
        "largest_single_borrower_pct_nav", cc.largest_single_borrower_pct_nav
    )

    wb = ws.parent

    def _reg(name: str, cell_addr: str) -> None:
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names[name] = DefinedName(
            name=name, attr_text=f"'{ws.title}'!{cell_addr}",
        )

    layout.set_column_widths(ws, label_width=54, it_width=32, year_width=14, unit_width=8)
    layout.write_title_block(
        ws, "Compliance Check", "Controllo di conformità",
        "AIFMD II / IFRS 9 / Basel III-IV / GACS — Italian & EU regulatory plumbing",
    )

    r = 5

    # ─── Section 1: AIFMD II leverage & concentration ──────────────────────
    layout.write_section_header(ws, r, "AIFMD II leverage & concentration",
                                 "AIFMD II leva & concentrazione")
    r += 1

    # AIFMD II leverage cap (commitment method)
    # Open-ended AIF: 175%; closed-ended: 300%
    ws.cell(row=r, column=1, value="AIF type (open/closed)").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Tipo AIF").font = styles.font_label_it
    aif_type_cell = ws.cell(row=r, column=4, value=aif_type_default)
    styles.style_input(aif_type_cell)
    aif_type_cell.comment = Comment(
        "AIF type (open / closed). Selects which AIFMD II leverage cap applies. "
        "Override via spec.compliance.aif_type. Default: closed.",
        "ModelForge",
    )
    aif_type_row = r
    r += 1

    # AIFMD II leverage caps — lifted from hardcoded 1.75 / 3.00 literals to
    # two visible named-input cells so the bracket is auditable & overridable.
    ws.cell(row=r, column=1, value="AIFMD II leverage cap — open-ended").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Limite leva AIFMD II — aperto").font = styles.font_label_it
    cap_open_cell = ws.cell(row=r, column=4, value=cc.aif_leverage_cap_open_pct)
    styles.style_input(cap_open_cell, number_format=styles.FMT_PCT_2DP)
    cap_open_cell.comment = Comment(
        "AIFMD II Article 15a: open-ended AIF max 175% leverage (commitment "
        "method). Override via spec.compliance.aif_leverage_cap_open_pct.",
        "ModelForge",
    )
    _reg("aif_leverage_cap_open", f"$D${r}")
    r += 1

    ws.cell(row=r, column=1, value="AIFMD II leverage cap — closed-ended").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Limite leva AIFMD II — chiuso").font = styles.font_label_it
    cap_closed_cell = ws.cell(row=r, column=4, value=cc.aif_leverage_cap_closed_pct)
    styles.style_input(cap_closed_cell, number_format=styles.FMT_PCT_2DP)
    cap_closed_cell.comment = Comment(
        "AIFMD II Article 15a: closed-ended AIF max 300% leverage (commitment "
        "method). Override via spec.compliance.aif_leverage_cap_closed_pct. "
        "Source: BCLP April 2024 transposition note.",
        "ModelForge",
    )
    _reg("aif_leverage_cap_closed", f"$D${r}")
    r += 1

    ws.cell(row=r, column=1, value="AIFMD II leverage cap (applied)").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Limite AIFMD II (applicato)").font = styles.font_label_it
    cap_cell = ws.cell(
        row=r, column=4,
        value=f'=IF($D${aif_type_row}="open",aif_leverage_cap_open,aif_leverage_cap_closed)',
    )
    styles.style_formula(cap_cell, number_format=styles.FMT_PCT_2DP)
    cap_cell.comment = Comment(
        "Applied cap = open-ended limit if AIF type is open, else closed-ended "
        "limit. Both bounds are the named inputs above.",
        "ModelForge",
    )
    cap_row = r
    r += 1

    ws.cell(row=r, column=1, value="Actual portfolio leverage").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Leva effettiva").font = styles.font_label_it
    actual_lev = ws.cell(row=r, column=4, value=actual_leverage_default)
    styles.style_input(actual_lev, number_format=styles.FMT_PCT_2DP)
    actual_lev.comment = Comment(
        "Actual portfolio leverage (commitment method) tested against the "
        "applied AIFMD II cap. Override via spec.compliance.actual_leverage.",
        "ModelForge",
    )
    _reg("aif_actual_leverage", f"$D${r}")
    r += 1

    ws.cell(row=r, column=1, value="AIFMD II leverage compliance").font = styles.font_subheader
    comp_cell = ws.cell(
        row=r, column=4,
        value=f'=IF(aif_actual_leverage<=$D${cap_row},"PASS","FAIL")',
    )
    styles.style_formula(comp_cell, number_format="General")
    comp_cell.font = styles.font_subheader
    r += 2

    # AIFMD II single-borrower cap (20% NAV)
    ws.cell(row=r, column=1, value="Largest single borrower % NAV").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Maggior debitore % NAV").font = styles.font_label_it
    sb_cell = ws.cell(row=r, column=4, value=largest_borrower_default)
    styles.style_input(sb_cell, number_format=styles.FMT_PCT_2DP)
    sb_cell.comment = Comment(
        "Largest single-borrower exposure as % of AIF NAV, tested against the "
        "concentration cap. Override via "
        "spec.compliance.largest_single_borrower_pct_nav.",
        "ModelForge",
    )
    _reg("aif_largest_single_borrower_pct_nav", f"$D${r}")
    r += 1

    # Single-borrower cap — lifted from hardcoded literal 0.20.
    ws.cell(row=r, column=1, value="AIFMD II single-borrower cap").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Limite singolo debitore").font = styles.font_label_it
    cap2 = ws.cell(row=r, column=4, value=cc.largest_borrower_cap_pct)
    styles.style_input(cap2, number_format=styles.FMT_PCT_2DP)
    cap2.comment = Comment(
        "AIFMD II Article 15a(4): loans to a single borrower capped at 20% of "
        "AIF NAV. Override via spec.compliance.largest_borrower_cap_pct. "
        "Source: Clifford Chance guidance.",
        "ModelForge",
    )
    _reg("aif_single_borrower_cap", f"$D${r}")
    r += 1

    ws.cell(row=r, column=1, value="Single-borrower compliance").font = styles.font_subheader
    comp2 = ws.cell(
        row=r, column=4,
        value='=IF(aif_largest_single_borrower_pct_nav<=aif_single_borrower_cap,"PASS","FAIL")',
    )
    styles.style_formula(comp2)
    comp2.font = styles.font_subheader
    comp2.comment = Comment(
        "AIFMD II Article 15a(4): loans to single borrower capped at 20% "
        "of AIF NAV. Source: Clifford Chance guidance.",
        "ModelForge",
    )
    r += 2

    # ─── Section 2: Loan-originating AIF status ────────────────────────────
    layout.write_section_header(ws, r, "Loan-originating AIF status",
                                 "Classificazione AIF con origination")
    r += 1

    ws.cell(row=r, column=1, value="Originated loans % NAV").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Prestiti originati % NAV").font = styles.font_label_it
    lorig = ws.cell(row=r, column=4,
                    value=context.get("originated_loans_pct_nav", 0.55))
    styles.style_input(lorig, number_format=styles.FMT_PCT_2DP)
    lorig_row = r
    r += 1

    ws.cell(row=r, column=1, value="Loan-originating AIF flag").font = styles.font_subheader
    flag = ws.cell(
        row=r, column=4,
        value=f'=IF($D${lorig_row}>=0.50,"YES — closed-ended required if >60%","NO")',
    )
    styles.style_formula(flag)
    flag.font = styles.font_subheader
    flag.comment = Comment(
        "AIFMD II: AIF deemed loan-originating if >50% NAV in "
        "originated loans. Must be closed-ended if origination >60% NAV. "
        "Source: Linklaters / Ropes & Gray.",
        "ModelForge",
    )
    r += 2

    # ─── Section 3: IFRS 9 three-stage ECL recap ───────────────────────────
    layout.write_section_header(ws, r, "IFRS 9 three-stage ECL",
                                 "IFRS 9 modello a tre stadi")
    r += 1

    # v0.8 US-260: Canonical ECL formula for auditor reference
    ws.cell(row=r, column=1,
            value="ECL formula (canonical): ECL = PD × LGD × EAD × DF").font = styles.font_subheader
    ws.cell(row=r, column=2,
            value="Formula ECL: PD × LGD × EAD × Fattore sconto").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1,
            value="  where PD = probability of default, LGD = loss given "
                  "default, EAD = exposure at default, DF = discount factor"
            ).font = styles.font_label_it
    r += 2

    stages = [
        ("Stage 1 (12-month ECL — no SICR)",
         "Stadio 1 (ECL 12m — no SICR)",
         "12m PD × LGD × EAD × DF"),
        ("Stage 2 (Lifetime ECL — SICR triggered)",
         "Stadio 2 (ECL lifetime — SICR)",
         "Lifetime PD × LGD × EAD × DF | 30+DPD backstop"),
        ("Stage 3 (Lifetime ECL — credit-impaired)",
         "Stadio 3 (ECL lifetime — impaired)",
         "PD 100% × LGD × EAD × DF | interest on NET"),
    ]
    for en, it, f in stages:
        ws.cell(row=r, column=1, value=en).font = styles.font_label_en
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        ws.cell(row=r, column=4, value=f).font = styles.font_label_it
        r += 1

    # SICR triggers
    r += 1
    ws.cell(row=r, column=1, value="SICR triggers (document bank policy)").font = styles.font_subheader
    r += 1
    triggers = [
        "Absolute PD threshold breach",
        "Relative PD doubling vs origination",
        "Rating downgrade by ≥2 notches",
        "Watchlist flag",
        "Forbearance granted",
        "30+ days past due (DPD) presumption",
    ]
    for t in triggers:
        ws.cell(row=r, column=1, value=f"  • {t}").font = styles.font_label_it
        r += 1
    r += 1

    # ─── Section 4: Basel NPL calendar provisioning ────────────────────────
    layout.write_section_header(ws, r, "Basel NPL calendar provisioning",
                                 "Calendar provisioning NPL")
    r += 1

    ws.cell(row=r, column=1, value="Regulation (EU) 2019/630 — NPL minimum coverage").font = styles.font_label_en
    r += 1

    bands = [
        ("Unsecured NPEs", "NPE non garantite", "0% / 35% / 100% (Y1-2 / Y3 / Y3+)"),
        ("NPEs secured by real estate", "NPE garantite da immobili", "0% / 25% / 70% / 100% (Y1-4 / Y5 / Y7 / Y9+)"),
        ("NPEs secured by other collateral", "NPE garantite altro", "0% / 35% / 70% / 100% (Y1-2 / Y3 / Y4 / Y7+)"),
    ]
    for en, it, f in bands:
        ws.cell(row=r, column=1, value=f"  • {en}").font = styles.font_label_en
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        ws.cell(row=r, column=4, value=f).font = styles.font_label_it
        r += 1
    ws.cell(row=r, column=1, value="").comment = Comment(
        "Minimum prudential backstop per Regulation (EU) 2019/630 "
        "applied to credit facilities originated from 26-Apr-2019.",
        "ModelForge",
    )
    r += 2

    # ─── Section 5: GACS eligibility (NPL) ─────────────────────────────────
    layout.write_section_header(ws, r, "GACS (Garanzia Cartolarizzazione Sofferenze)",
                                 "GACS — Garanzia stato su senior NPL")
    r += 1

    gacs_items = [
        ("Senior tranche rating ≥ BBB-", "Rating senior ≥ investment grade",
         context.get("senior_rating", "BBB")),
        ("Legge 130/1999 SPV (bankruptcy-remote)", "SPV legge 130/1999",
         "YES (required)"),
        ("Guarantee fee priced on IT financial CDS", "Fee garanzia su CDS finanziari italiani",
         "IT Italian financial-sector CDS basket (5Y)"),
        ("Servicer fit-and-proper (Banca d'Italia)", "Servicer requisiti",
         "Required"),
    ]
    for en, it, val in gacs_items:
        ws.cell(row=r, column=1, value=f"  • {en}").font = styles.font_label_en
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        ws.cell(row=r, column=4, value=val).font = styles.font_label_it
        r += 1
    r += 1

    # ─── Section 6: IRES + IRAP split (Italian tax transparency) ───────────
    layout.write_section_header(ws, r, "Italian tax — IRES + IRAP split",
                                 "Tassazione italiana — IRES + IRAP")
    r += 1

    # IRES rate — lifted from hardcoded literal 0.24.
    ws.cell(row=r, column=1, value="IRES rate").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Aliquota IRES").font = styles.font_label_it
    c = ws.cell(row=r, column=4, value=cc.tax.ires_rate_pct)
    styles.style_input(c, number_format=styles.FMT_PCT_2DP)
    c.comment = Comment(
        "IRES — Italian corporate income tax. Override via "
        "spec.compliance.tax.ires_rate_pct. Default 24% (Italy 2026).",
        "ModelForge",
    )
    _reg("tax_ires_rate", f"$D${r}")
    r += 1

    # IRAP rate — lifted from hardcoded literal 0.039.
    ws.cell(row=r, column=1, value="IRAP rate (national base + regional)").font = styles.font_label_en
    ws.cell(row=r, column=2, value="Aliquota IRAP").font = styles.font_label_it
    c = ws.cell(row=r, column=4, value=cc.tax.irap_rate_pct)
    styles.style_input(c, number_format=styles.FMT_PCT_2DP)
    c.comment = Comment(
        "IRAP — regional production tax (national base + regional add-on). "
        "Override via spec.compliance.tax.irap_rate_pct. Default 3.9% (Italy "
        "non-financial corporate base).",
        "ModelForge",
    )
    _reg("tax_irap_rate", f"$D${r}")
    r += 1

    ws.cell(row=r, column=1, value="IRES+IRAP combined (approximate)").font = styles.font_subheader
    ws.cell(row=r, column=2, value="IRES+IRAP combinate (stima)").font = styles.font_label_it
    c = ws.cell(row=r, column=4, value="=tax_ires_rate+tax_irap_rate")
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    c.font = styles.font_subheader
    c.comment = Comment(
        "Caveat: IRES and IRAP have DIFFERENT tax bases. IRES includes "
        "financial items; IRAP excludes them. The 'combined' rate is "
        "an APPROXIMATION for non-financial corporates. Banks / insurers "
        "pay IRAP at 4.65% + 2pp (Italian 2026 Budget Law). "
        "Source: PwC Italy Corporate Tax Summaries.",
        "ModelForge",
    )
    r += 2

    r += 1

    # ─── Section 7: Basel securitization capital framework (v0.8.7 US-510/512)
    layout.write_section_header(
        ws, r, "Basel III/IV securitization capital (SEC-SA / SEC-IRBA / SEC-ERBA)",
        "Capitale Basel su cartolarizzazioni",
    )
    r += 1

    ws.cell(row=r, column=1, value="Framework hierarchy (BCBS d374 / CRR Art. 254)").font = styles.font_subheader
    ws.cell(row=r, column=2, value="Gerarchia framework").font = styles.font_label_it
    r += 1

    bsec_items = [
        ("SEC-IRBA (internal ratings-based)",
         "Banca IRB con rating interni",
         "Priority 1 — required if bank has IRB approval for underlying"),
        ("SEC-ERBA (external ratings-based)",
         "Basato su rating esterni",
         "Priority 2 — used when ECAI rating available on tranche"),
        ("SEC-SA (standardised approach)",
         "Approccio standard",
         "Priority 3 — fallback; Kirb-based look-through formula"),
        ("Risk-weight floor (all approaches)",
         "Floor ponderazione",
         "15% minimum on senior; 100% on mezz; 1,250% on equity/residual"),
        ("Output floor (Basel IV)",
         "Output floor Basel IV",
         "72.5% of SA by 2028 (phased from 50% in 2022)"),
        ("STS qualifying (Reg. EU 2017/2402)",
         "STS conforme",
         "Simple / Transparent / Standardised — reduced RW"),
    ]
    for en, it, val in bsec_items:
        ws.cell(row=r, column=1, value=f"  • {en}").font = styles.font_label_en
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        ws.cell(row=r, column=4, value=val).font = styles.font_label_it
        r += 1

    ws.cell(row=r, column=1, value="").comment = Comment(
        "Basel III/IV securitization framework per BCBS d374 (December 2014) "
        "implemented in EU via CRR Art. 254 / Reg. EU 2017/2401 + 2402 (STS). "
        "Hierarchy: SEC-IRBA > SEC-ERBA > SEC-SA. Output floor 72.5% of SA "
        "applicable from 2028. Sources: BIS BCBS d374; EBA technical "
        "standards; Banca d'Italia Circolare 285 Title III.",
        "ModelForge",
    )
    r += 2

    ws.freeze_panes = "D5"
    ws.print_title_rows = "1:4"
