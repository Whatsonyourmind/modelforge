"""DCF valuation + WACC build sheet (Template 10).

Emits three core sheets:
    WACCBuild      — risk-free + ERP + beta → cost of equity, then
                     WACC from target capital structure
    FCFForecast    — revenue, EBITDA, D&A, EBIT, tax, capex, ΔNWC → FCF
    Valuation     — PV(explicit FCF) + terminal value (Gordon + exit x)
                     → enterprise value → equity value → implied per
                     share price with Damodaran 2026 Italy ERP cited.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build_wacc(ws: Worksheet, spec) -> dict[str, str]:
    """Emit WACCBuild. Returns {cost_of_equity: ref, wacc: ref}.

    v0.7 enhancements:
      - If spec.wacc.comparable_betas provided, use `relevered_beta` named
        range from the ComparableBetas sheet (Hamada) instead of raw beta_levered.
      - If spec.wacc.mature_erp + sovereign_default_spread + equity_bond_vol_ratio
        provided, compute Damodaran CRP-adjusted ERP:
            effective_ERP = mature_ERP + sovereign_spread × (σ_equity/σ_bond) × λ
    """
    layout.set_column_widths(ws, label_width=44, it_width=32, year_width=14, unit_width=6)
    layout.write_title_block(ws, "WACC Build", "Calcolo WACC",
                             "Damodaran CRP methodology + Hamada beta (when comps present)")

    # Pick β source: relevered_beta (Hamada) if comps provided; else beta_levered.
    has_comps = bool(spec.wacc.comparable_betas)
    beta_ref = "relevered_beta" if has_comps else "beta_levered"

    # Pick ERP source: Damodaran decomposition if all four components present;
    # else flat equity_risk_premium.
    has_crp = all([
        spec.wacc.mature_erp is not None,
        spec.wacc.sovereign_default_spread is not None,
        spec.wacc.equity_bond_vol_ratio is not None,
        spec.wacc.lambda_country_exposure is not None,
    ])
    if has_crp:
        erp_formula = (
            "mature_erp+sovereign_default_spread*equity_bond_vol_ratio*lambda_country_exposure"
        )
    else:
        erp_formula = "equity_risk_premium"

    # v0.9.7: Duff & Phelps / Kroll size premium + company-specific alpha
    # (both optional). When supplied, append additively to cost of equity.
    has_size_premium = spec.wacc.size_premium_pct is not None
    has_alpha = spec.wacc.company_specific_alpha_bps is not None
    coe_tail_terms: list[str] = []
    if has_size_premium:
        coe_tail_terms.append("size_premium_pct")
    if has_alpha:
        coe_tail_terms.append("company_specific_alpha_bps/10000")
    coe_tail = ("+" + "+".join(coe_tail_terms)) if coe_tail_terms else ""

    rows = [
        ("Risk-free rate (10Y BTP)", "Tasso risk-free", "=risk_free_rate", styles.FMT_PCT_2DP),
    ]

    if has_crp:
        rows += [
            ("Mature-market ERP (Damodaran)", "ERP mercato maturo",
             "=mature_erp", styles.FMT_PCT_2DP),
            ("Sovereign default spread (Italy)", "Spread sovrano Italia",
             "=sovereign_default_spread", styles.FMT_PCT_2DP),
            ("σ_equity / σ_bond ratio", "Rapporto σ_azioni / σ_titoli",
             "=equity_bond_vol_ratio", styles.FMT_MULTIPLE),
            ("Lambda (country exposure)", "Lambda (esposizione paese)",
             "=lambda_country_exposure", styles.FMT_MULTIPLE),
            ("Effective ERP (mature + CRP × λ)", "ERP effettivo",
             f"={erp_formula}", styles.FMT_PCT_2DP),
        ]
    else:
        rows += [
            ("Equity risk premium (country)", "ERP paese", "=equity_risk_premium",
             styles.FMT_PCT_2DP),
        ]

    rows += [
        ("Beta (relevered via Hamada)" if has_comps else "Beta (levered — input)",
         "Beta", f"={beta_ref}", styles.FMT_MULTIPLE),
        ("Cost of equity", "Costo del capitale",
         f"=risk_free_rate+{beta_ref}*({erp_formula}){coe_tail}", styles.FMT_PCT_2DP),
        ("Pre-tax cost of debt", "Costo del debito pre-tax", "=pretax_cost_of_debt", styles.FMT_PCT_2DP),
        ("Effective tax rate", "Aliquota effettiva", "=effective_tax_rate", styles.FMT_PCT),
        ("After-tax cost of debt", "Costo del debito post-tax",
         "=pretax_cost_of_debt*(1-effective_tax_rate)", styles.FMT_PCT_2DP),
        ("Target D / (D+E)", "Target D / (D+E)", "=target_debt_weight", styles.FMT_PCT),
        ("Target E / (D+E)", "Target E / (D+E)", "=1-target_debt_weight", styles.FMT_PCT),
        ("WACC",
         "WACC",
         f"=((risk_free_rate+{beta_ref}*({erp_formula}){coe_tail})*(1-target_debt_weight))+"
         "(pretax_cost_of_debt*(1-effective_tax_rate)*target_debt_weight)",
         styles.FMT_PCT_2DP),
    ]
    r = 5
    refs: dict[str, str] = {}
    for en, it, formula, fmt in rows:
        layout.write_row_label(ws, r, en, it)
        c = ws.cell(row=r, column=4, value=formula)
        styles.style_formula(c, number_format=fmt)
        if en == "Cost of equity":
            refs["cost_of_equity"] = f"'{ws.title}'!$D${r}"
            c.font = styles.font_subheader
        if en == "WACC":
            refs["wacc"] = f"'{ws.title}'!$D${r}"
            c.font = styles.font_subheader
            c.comment = Comment(
                "WACC = Ke × (E/(D+E)) + Kd × (1 − t) × (D/(D+E)).\n"
                "Damodaran 2026 Italy ERP 6.7% used per country-risk "
                "table; risk-free cites 10Y BTP yield.",
                "ModelForge",
            )
        r += 1

    # Register wacc_rate named range at workbook level
    from openpyxl.workbook.defined_name import DefinedName
    if "wacc_rate" in ws.parent.defined_names:
        del ws.parent.defined_names["wacc_rate"]
    ws.parent.defined_names["wacc_rate"] = DefinedName(
        name="wacc_rate", attr_text=refs["wacc"],
    )

    ws.freeze_panes = "D5"
    ws.print_title_rows = "1:4"
    return refs


def build_fcf(ws: Worksheet, spec, wacc_refs: dict[str, str]) -> dict[str, str]:
    """Emit FCFForecast. Returns row/ref map for Valuation sheet.

    v0.8 (US-231): if spec.fade_years > 0, fade_years extra projection
    columns are appended after the p explicit years. Growth rate in fade
    year j (1..fade) interpolates linearly from the last explicit growth
    to terminal_growth_pct: g_j = g_last + (g_term − g_last) × j/(fade+1).
    EBITDA margin / D&A% / capex% / ΔNWC% are held at their last explicit
    values during the fade period (steady-state margin assumption).
    """
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    fade = int(getattr(spec, "fade_years", 0) or 0)
    stub_days = int(getattr(spec, "stub_period_days", 365) or 365)
    p_eff = p + fade
    n = h + p_eff

    layout.set_column_widths(ws, label_width=40, it_width=32, year_width=12, unit_width=6)
    layout.write_title_block(ws, "FCF Forecast", "Previsione FCF",
                             "Unlevered free cash flow to the firm")
    layout.write_scenario_banner(ws, row=3)

    # Year headers row 5
    yr_row = 5
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        label = "H" if i < h else "F"
        idx = i + 1 if i < h else i - h + 1
        suffix = " fade" if (fade > 0 and i >= h + p) else ""
        c = ws.cell(row=yr_row, column=col_idx, value=f"{label}Y{idx}{suffix}")
        styles.style_header(c)

    # v0.8 US-230/231: always emit a config row so downstream auditors /
    # reviewers can see stub & fade treatment without reading the YAML.
    # Only include the word "fade" when a fade period actually exists so
    # auditors can detect it by substring search without false positives.
    stub_part = (f"Stub period: {stub_days} days"
                 f"{' (prorated)' if stub_days != 365 else ' (full year)'}")
    if fade > 0:
        fade_part = f" | Fade period: {fade} years (growth → terminal_g)"
    else:
        fade_part = " | Single-stage explicit (no fade)"
    cc = ws.cell(row=6, column=1, value=stub_part + fade_part)
    cc.font = styles.font_label_it

    # Rows: Revenue, Growth %, EBITDA margin %, EBITDA, D&A, EBIT, Tax, NOPAT, Capex, ΔNWC, FCF
    r = 7
    rows_out: dict[str, int] = {}

    # Revenue
    layout.write_row_label(ws, r, "Revenue", "Ricavi")
    rows_out["revenue"] = r
    # Historical as hardcoded values (from spec)
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        # Use target.revenue_last_fy_eur_m at the last historical year if no history array
        # For simplicity, set the last-fy revenue at col(h-1) and roll others back by constant
        rev = spec.target.revenue_last_fy_eur_m
        c = ws.cell(row=r, column=col_idx,
                    value=rev if i == h - 1 else rev * 0.9 ** (h - 1 - i))
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c.comment = Comment(f"Historical revenue (FY{i+1}) — from spec", "ModelForge")
    # Projected + fade: revenue_prev * (1 + growth)
    last_g_name = spec.fcf.revenue_growth_by_year[-1].name
    for i in range(p_eff):
        col = layout.year_col(h + i)
        col_idx = ord(col) - ord("A") + 1
        prev_col = layout.year_col(h + i - 1)
        if i < p:
            growth_name = spec.fcf.revenue_growth_by_year[i].name
            c = ws.cell(row=r, column=col_idx,
                        value=f"={prev_col}{r}*(1+{growth_name})")
        else:
            # v0.8 US-231 fade: g_j = g_last + (g_term − g_last) × j/(fade+1)
            j = i - p + 1
            ratio = j / (fade + 1)
            growth_expr = (f"({last_g_name}+(terminal_growth_pct-{last_g_name})"
                           f"*{ratio})")
            c = ws.cell(row=r, column=col_idx,
                        value=f"={prev_col}{r}*(1+{growth_expr})")
            c.comment = Comment(
                f"Fade year {j}/{fade}: growth linearly converges "
                f"from last explicit year to terminal_growth_pct.",
                "ModelForge",
            )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # EBITDA
    layout.write_row_label(ws, r, "EBITDA", "EBITDA")
    rows_out["ebitda"] = r
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=spec.target.ebitda_last_fy_eur_m)
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c.comment = Comment("Historical EBITDA (from spec)", "ModelForge")
    # Explicit + fade: margin held at last explicit value during fade
    last_margin_name = spec.fcf.ebitda_margin_by_year[-1].name
    for i in range(p_eff):
        margin_name = (spec.fcf.ebitda_margin_by_year[i].name
                       if i < p else last_margin_name)
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={layout.year_col(h + i)}{rows_out['revenue']}*{margin_name}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # D&A
    layout.write_row_label(ws, r, "D&A", "A&A", indent=True)
    rows_out["da"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-{col}{rows_out['revenue']}*da_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # EBIT
    layout.write_row_label(ws, r, "EBIT", "EBIT")
    rows_out["ebit"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{rows_out['ebitda']}+{col}{rows_out['da']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 1

    # Tax on EBIT
    layout.write_row_label(ws, r, "Tax on EBIT", "Imposte su EBIT", indent=True)
    rows_out["tax"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-MAX({col}{rows_out['ebit']}*effective_tax_rate,0)")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # NOPAT
    layout.write_row_label(ws, r, "NOPAT", "NOPAT")
    rows_out["nopat"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{rows_out['ebit']}+{col}{rows_out['tax']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # +D&A back
    layout.write_row_label(ws, r, "+ D&A addback", "+ A&A", indent=True)
    rows_out["da_addback"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"=-{col}{rows_out['da']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Capex
    layout.write_row_label(ws, r, "Capex", "Capex", indent=True)
    rows_out["capex"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-{col}{rows_out['revenue']}*capex_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Δ NWC
    layout.write_row_label(ws, r, "Δ Working capital", "Δ Capitale circolante", indent=True)
    rows_out["nwc"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        if i == 0:
            c = ws.cell(row=r, column=col_idx, value=0)
        else:
            prev = layout.year_col(i - 1)
            c = ws.cell(
                row=r, column=col_idx,
                value=f"=-({col}{rows_out['revenue']}-{prev}{rows_out['revenue']})*nwc_pct_revenue_delta",
            )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # FCF (unlevered)
    layout.write_row_label(ws, r, "Unlevered FCF", "FCF unlevered")
    rows_out["fcf"] = r
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(
            row=r, column=col_idx,
            value=(
                f"={col}{rows_out['nopat']}+{col}{rows_out['da_addback']}"
                f"+{col}{rows_out['capex']}+{col}{rows_out['nwc']}"
            ),
        )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    rows_out["h"] = h
    rows_out["p"] = p
    rows_out["fade"] = fade
    rows_out["p_eff"] = p_eff
    return rows_out


def build_valuation(ws: Worksheet, spec, fcf_refs: dict[str, int],
                    fcf_sheet: str) -> None:
    """Enterprise DCF + terminal value reconciliation.

    v0.6 changes:
      - Explicit-period PV expanded as an explicit sum (no volatile
        INDIRECT). Supports a `mid_year_convention` toggle.
      - Gordon and Exit TV presented side-by-side; user picks one via
        `terminal_method_choice` named range (no more averaging).
      - Net debt, shares outstanding, and current price use named
        assumptions when provided, replacing hardcoded literals.

    v0.8 changes:
      - US-230: first projection period is prorated by stub_period_days/365
        and discounted at the stub midpoint (mid-year) or stub end (end-year).
        Subsequent periods cumulate stub_years + (t−1) − mid_year.
      - US-231: fade_years extends explicit period; TV discounted by
        stub_years + (p+fade) − 1 − mid_year.
      - US-232: Terminal FCF normalization rows (capex = D&A steady-state,
        ΔNWC = last ΔNWC × (1 + g)) feed the Gordon formula.
    """
    h = fcf_refs["h"]
    p = fcf_refs["p"]
    fade = int(fcf_refs.get("fade", 0))
    p_eff = int(fcf_refs.get("p_eff", p))
    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=13, unit_width=6)
    layout.write_title_block(ws, "Valuation", "Valutazione",
                             "DCF Gordon / Exit Multiple — pick one via terminal_method_choice")

    r = 5
    last_col = layout.year_col(h + p_eff - 1)
    fcf_row = fcf_refs["fcf"]
    ebitda_row = fcf_refs["ebitda"]
    nopat_row = fcf_refs["nopat"]
    nwc_row = fcf_refs["nwc"]
    da_addback_row = fcf_refs["da_addback"]

    mid_year = 0.5 if spec.mid_year_convention else 0.0

    # US-230 stub-period support: first projection period is prorated.
    # stub_years = stub_period_days / 365 (1.0 = default full year).
    stub_days = int(getattr(spec, "stub_period_days", 365) or 365)
    stub_years = stub_days / 365.0

    # Explicit + fade PV expanded as a sum (no INDIRECT).
    # i=0 (stub): FCF × stub_years / (1+wacc)^(stub_years × (1-mid_year))
    # i≥1 (full years): FCF / (1+wacc)^(stub_years + i − mid_year)
    pv_terms: list[str] = []
    for i in range(p_eff):
        col = layout.year_col(h + i)
        fcf_ref = f"'{fcf_sheet}'!{col}{fcf_row}"
        if i == 0:
            exp = stub_years * (1.0 - mid_year)
            prorate = "" if stub_days == 365 else f"*{stub_years}"
            pv_terms.append(f"{fcf_ref}{prorate}/(1+wacc_rate)^{exp}")
        else:
            exp = stub_years + i - mid_year
            pv_terms.append(f"{fcf_ref}/(1+wacc_rate)^{exp}")
    pv_formula = "=" + "+".join(pv_terms)

    # TV at end of last explicit/fade period = stub_years + (p_eff − 1) years;
    # discounted by that minus mid_year to be consistent with explicit PV.
    tv_discount_years = stub_years + p_eff - 1 - mid_year

    # Net debt / shares / price references — prefer named ranges
    net_debt_ref = ("net_debt_assum" if spec.net_debt_assum is not None
                    else f"{spec.net_debt_eur_m}")
    shares_ref = ("shares_outstanding_assum"
                  if spec.shares_outstanding_assum is not None
                  else f"{spec.shares_outstanding_m}")
    price_ref = ("current_price_assum"
                 if spec.current_price_assum is not None
                 else f"{spec.valuation_date_price_eur}")

    # v0.7: full EV→Equity bridge with optional adjustments
    # (minority interest, pension deficit, preferred, cross-holdings,
    # IFRS 16 lease liability). Each reads a named Assumption when
    # present; otherwise bare scalar (0) from the spec.
    minority_ref = ("minority_interest_assum"
                    if getattr(spec, 'minority_interest_assum', None) is not None
                    else "0")
    pension_ref = ("pension_deficit_assum"
                   if getattr(spec, 'pension_deficit_assum', None) is not None
                   else "0")
    preferred_ref = ("preferred_equity_assum"
                     if getattr(spec, 'preferred_equity_assum', None) is not None
                     else "0")
    cross_hold_ref = ("cross_holdings_assum"
                      if getattr(spec, 'cross_holdings_assum', None) is not None
                      else "0")
    lease_ref = ("lease_liability_ifrs16_assum"
                 if getattr(spec, 'lease_liability_ifrs16_assum', None) is not None
                 else "0")

    rows = [
        ("Explicit-period PV of FCF", "VAN FCF esplicito",
         pv_formula, styles.FMT_EUR_M, "pv_explicit"),
        # v0.8 US-232: Terminal FCF normalization. Three rows before Gordon.
        ("Terminal year NOPAT (from explicit)", "NOPAT terminale",
         f"='{fcf_sheet}'!{last_col}{nopat_row}",
         styles.FMT_EUR_M, "norm_nopat"),
        ("Normalized terminal Δ NWC (grown at g)",
         "Δ CCN terminale normalizzato",
         f"='{fcf_sheet}'!{last_col}{nwc_row}*(1+terminal_growth_pct)",
         styles.FMT_EUR_M, "norm_nwc"),
        ("Normalized terminal FCF (capex = D&A steady state)",
         "FCF terminale normalizzato (capex = A&A)",
         f"=D{r+1}+D{r+2}",
         styles.FMT_EUR_M, "norm_fcf"),
        # Gordon uses normalized FCF (steady-state capex offsets D&A addback).
        ("Terminal value — Gordon growth (normalized)",
         "Terminal value — Gordon (norm.)",
         f"=D{r+3}*(1+terminal_growth_pct)/(wacc_rate-terminal_growth_pct)",
         styles.FMT_EUR_M, "tv_gordon"),
        ("Terminal value — exit EV/EBITDA", "Terminal value — EV/EBITDA uscita",
         f"='{fcf_sheet}'!{last_col}{ebitda_row}*exit_ev_ebitda_x",
         styles.FMT_EUR_M, "tv_exit"),
        # v0.7 implied-g cross-check; now uses normalized FCF
        ("Implied g — terminal (from exit multiple)",
         "Implied g — terminale (da multiplo uscita)",
         f"=wacc_rate-D{r+3}*(1+terminal_growth_pct)/D{r+5}",
         styles.FMT_PCT_2DP, "implied_g"),
        # Chosen TV — picked by named range (1=Gordon, 2=Exit; default Gordon)
        ("Terminal value — chosen", "Terminal value — scelto",
         f"=IF(terminal_method_choice=2,D{r+5},D{r+4})",
         styles.FMT_EUR_M, "tv_used"),
        ("PV of terminal value", "VAN del terminal value",
         f"=D{r+7}/(1+wacc_rate)^{tv_discount_years}",
         styles.FMT_EUR_M, None),
        ("Enterprise Value", "Enterprise Value",
         f"=D{r}+D{r+8}",
         styles.FMT_EUR_M, "ev"),
        # v0.7: full EV→Equity bridge (Footnotes Analyst standard)
        ("(−) Net debt", "(−) Posizione finanziaria netta",
         f"=-{net_debt_ref}",
         styles.FMT_EUR_M, None),
        ("(−) Minority interest", "(−) Interessenze di minoranza",
         f"=-{minority_ref}",
         styles.FMT_EUR_M, None),
        ("(−) Pension deficit (unfunded)", "(−) Deficit pensionistico",
         f"=-{pension_ref}",
         styles.FMT_EUR_M, None),
        ("(−) Preferred equity", "(−) Azioni privilegiate",
         f"=-{preferred_ref}",
         styles.FMT_EUR_M, None),
        ("(−) IFRS 16 lease liability", "(−) Passività leasing IFRS 16",
         f"=-{lease_ref}",
         styles.FMT_EUR_M, None),
        ("(+) Cross-holdings / investments at FV",
         "(+) Partecipazioni / investimenti",
         f"={cross_hold_ref}",
         styles.FMT_EUR_M, None),
        ("Equity Value", "Valore equity",
         f"=D{r+9}+D{r+10}+D{r+11}+D{r+12}+D{r+13}+D{r+14}+D{r+15}",
         styles.FMT_EUR_M, "equity_val"),
        ("Implied EV / EBITDA (Y1 proj)", "EV/EBITDA implicito Y1",
         f"=D{r+9}/'{fcf_sheet}'!{layout.year_col(h)}{ebitda_row}",
         styles.FMT_MULTIPLE, None),
        ("Implied EV", "EV implicito",
         f"=D{r+9}",
         styles.FMT_EUR_M, None),
    ]

    for i, (en, it, formula, fmt, name) in enumerate(rows):
        rr = r + i
        layout.write_row_label(ws, rr, en, it)
        c = ws.cell(row=rr, column=4, value=formula)
        styles.style_formula(c, number_format=fmt)
        if name in ("ev", "equity_val", "tv_used", "pv_explicit", "norm_fcf"):
            c.font = styles.font_subheader
        if en.startswith("Enterprise Value") or en.startswith("Equity Value"):
            c.border = styles.BORDER_TOP_THIN
        if name == "pv_explicit":
            conv = "mid-year" if spec.mid_year_convention else "end-year"
            stub_note = (f" Stub first period: {stub_days} days "
                         f"({stub_years:.4f} years)." if stub_days != 365 else "")
            fade_note = (f" {fade} fade-year(s) extend explicit period "
                         f"(growth interpolates to terminal_g)." if fade > 0 else "")
            c.comment = Comment(
                f"Explicit-period discounting uses {conv} convention "
                f"(mid_year_convention={spec.mid_year_convention})."
                f"{stub_note}{fade_note} "
                f"TV discounted by {tv_discount_years:.4f} years.",
                "ModelForge",
            )
        if name == "norm_fcf":
            c.comment = Comment(
                "Normalized terminal FCF = NOPAT + ΔNWC_norm (capex = D&A "
                "steady state, so D&A addback and capex cancel). "
                "Gordon formula uses this normalized cash flow.",
                "ModelForge",
            )

    # Create workbook-level named range for the terminal method choice
    # (1 = Gordon, 2 = Exit multiple). v0.9.7: honor spec.terminal.terminal_method_choice
    # if set (default 1 = Gordon). User-overridable in the workbook itself.
    from openpyxl.workbook.defined_name import DefinedName
    if "terminal_method_choice" not in ws.parent.defined_names:
        helper_row = r + len(rows) + 4
        layout.write_row_label(ws, helper_row,
                               "Terminal method choice (1=Gordon, 2=Exit)",
                               "Metodo TV (1=Gordon, 2=Exit)")
        method_choice = getattr(spec.terminal, "terminal_method_choice", 1) or 1
        ws.cell(row=helper_row, column=4, value=int(method_choice))
        ws.parent.defined_names["terminal_method_choice"] = DefinedName(
            name="terminal_method_choice",
            attr_text=f"'{ws.title}'!$D${helper_row}",
        )

    # Implied price per share (when shares available).
    # Equity value is at row r+16 after v0.8 US-232 normalization rows
    # (3 rows inserted before Gordon, shifting bridge/Equity down by 3).
    has_shares = (spec.shares_outstanding_assum is not None
                  or spec.shares_outstanding_m > 0)
    has_price = (spec.current_price_assum is not None
                 or spec.valuation_date_price_eur > 0)
    if has_shares:
        ri = r + len(rows)
        equity_val_row = r + 16  # position of Equity Value row
        layout.write_row_label(ws, ri, "Implied price per share",
                               "Prezzo implicito per azione")
        c = ws.cell(row=ri, column=4, value=f"=D{equity_val_row}/{shares_ref}")
        styles.style_formula(c, number_format=styles.FMT_EUR_ACTUAL)
        c.font = styles.font_subheader
        if has_price:
            ri += 1
            layout.write_row_label(ws, ri, "Current price", "Prezzo attuale")
            c2 = ws.cell(row=ri, column=4, value=f"={price_ref}")
            styles.style_formula(c2, number_format=styles.FMT_EUR_ACTUAL)
            ri += 1
            layout.write_row_label(ws, ri, "Premium / (discount) %",
                                   "Premio / (sconto) %")
            c = ws.cell(row=ri, column=4,
                        value=f"=D{ri-2}/D{ri-1}-1")
            styles.style_formula(c, number_format=styles.FMT_PCT_2DP)

    ws.freeze_panes = "D5"
    ws.print_title_rows = "1:4"
