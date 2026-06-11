"""Debt schedule.

Period-by-period debt roll-forward for each tranche. Reads drivers
(amount, margin, floor, reference rate, arrangement fee) as named ranges
from Assumptions. Computes opening → drawdown → amortization → closing,
average balance, all-in rate, cash interest.

Also writes back into the Operating Model's Interest Expense row via a
cross-sheet SUMIFS pattern (one tranche per period → summed).

Sign convention: cash interest is NEGATIVE on Operating Model (it's a
cost). Debt balances on this sheet are POSITIVE (they're liabilities
shown at face value).
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.formulas import all_in_rate, average_of, cash_sweep
from modelforge.builder.i18n import L
from modelforge.graph.schema import LinkageGraph, GraphNode, NodeKind
from modelforge.spec.unitranche import UnitrancheSpec


def build(
    ws: Worksheet,
    spec: UnitrancheSpec,
    graph: LinkageGraph,
    driver_refs: dict[str, str],
    operating_refs: dict[str, str],
    operating_sheet_name: str,
) -> dict[str, str]:
    horizon = spec.horizon
    h = horizon.historical_years
    p = horizon.projection_years
    n_years = h + p

    layout.set_column_widths(ws, label_width=40, it_width=32, year_width=12, unit_width=6)

    layout.write_title_block(
        ws,
        title_en="Debt Schedule",
        title_it="Piano del debito",
        subtitle=f"{spec.meta.currency} {spec.meta.unit_scale} · "
                 f"average-balance interest convention",
    )
    layout.write_scenario_banner(ws, row=3)

    # Year header row
    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        is_hist = i < h
        c = ws.cell(row=yr_row, column=col_idx,
                    value=f"{'A' if is_hist else 'E'} {yr}")
        styles.style_header(c)

    # For each tranche, write a block
    r = 7
    tranche_blocks: list[dict] = []
    for tr_idx, tr in enumerate(spec.debt.tranches):
        layout.write_section_header(ws, r, f"Tranche — {tr.name.en}", tr.name.secondary)
        r += 1

        block: dict[str, int] = {"name": tr.name.en}

        # Opening debt
        block["opening"] = r
        layout.write_row_label(ws, r, L("debt_opening").en, L("debt_opening").secondary)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i == 0:
                # First historical: 0 (we assume the deal closes at end of last historical)
                c = ws.cell(row=r, column=col_idx, value="=0")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
            else:
                # opening = prior closing
                prior_col = layout.year_col(i - 1)
                c = ws.cell(row=r, column=col_idx, value=f"=${prior_col}${r + 3}")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # Drawdown (only at close — the first projected year)
        block["drawdown"] = r
        layout.write_row_label(ws, r, L("debt_drawdown").en, L("debt_drawdown").secondary, indent=True)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i == h:  # first projected year = close
                c = ws.cell(row=r, column=col_idx, value=f"={tr.amount.name}")
                styles.style_xref(c, number_format=styles.FMT_EUR_M)
            else:
                c = ws.cell(row=r, column=col_idx, value="=0")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # Mandatory-amortization rate input (only for the mandatory_1pct
        # profile). Lifted from a hard-coded literal 1% into a VISIBLE,
        # styled, overridable named-input cell on the Debt sheet so the
        # assumption is traceable and can be driven from the spec
        # (amortization_rate_pct, default 0.01 == the historical 1%).
        amort_rate_name: str | None = None
        if tr.amortization == "mandatory_1pct":
            amort_rate_pct = getattr(tr, "amortization_rate_pct", 0.01)
            amort_rate_name = f"mandatory_amort_rate_t{tr_idx + 1}"
            layout.write_row_label(
                ws, r,
                "Mandatory amortization rate (% of orig.)",
                "Ammortamento obbligatorio (% capitale)",
                indent=True,
            )
            ws.cell(row=r, column=3, value="%").font = styles.font_label_it
            rate_cell = ws.cell(row=r, column=4, value=amort_rate_pct)
            styles.style_input(rate_cell, number_format=styles.FMT_PCT_2DP)
            rate_cell.comment = Comment(
                "Mandatory per-period amortization as a % of the tranche's "
                "original commitment (overridable; default 1.00%).",
                "ModelForge",
            )
            wb = ws.parent
            if amort_rate_name in wb.defined_names:
                del wb.defined_names[amort_rate_name]
            wb.defined_names[amort_rate_name] = DefinedName(
                name=amort_rate_name,
                attr_text=f"'{ws.title}'!$D${r}",
            )
            r += 1

        # Repayment / amortization
        block["repayment"] = r
        layout.write_row_label(ws, r, L("debt_repayment").en, L("debt_repayment").secondary, indent=True)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        # v0.6: corrected off-by-one. Drawdown is at column index h (Year 1
        # of the bond). A tenor_years=N bond lives through Year N, which is
        # column index h + N - 1. Previously `h + tr.tenor_years` pushed
        # maturity one column past its intended end, leaving residual
        # principal at the real maturity.
        maturity_year = h + tr.tenor_years - 1
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            amort_ref = 0
            if tr.amortization == "bullet":
                if i == maturity_year:
                    # Repay full opening at maturity
                    c = ws.cell(row=r, column=col_idx,
                                value=f"=-${col}${block['opening']}")
                    styles.style_formula(c, number_format=styles.FMT_EUR_M)
                    continue
            elif tr.amortization == "linear":
                if h <= i <= maturity_year and tr.tenor_years > 0:
                    c = ws.cell(
                        row=r, column=col_idx,
                        value=f"=-{tr.amount.name}/{tr.tenor_years}",
                    )
                    styles.style_formula(c, number_format=styles.FMT_EUR_M)
                    continue
            elif tr.amortization == "mandatory_1pct":
                if h <= i < maturity_year:
                    c = ws.cell(
                        row=r, column=col_idx,
                        value=f"=-{tr.amount.name}*{amort_rate_name}",
                    )
                    styles.style_formula(c, number_format=styles.FMT_EUR_M)
                    continue
                elif i == maturity_year:
                    c = ws.cell(
                        row=r, column=col_idx,
                        value=f"=-${col}${block['opening']}",
                    )
                    styles.style_formula(c, number_format=styles.FMT_EUR_M)
                    continue
            # Default: no amortization this period
            c = ws.cell(row=r, column=col_idx, value="=0")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # Closing debt = opening + drawdown + repayment (repayment negative)
        block["closing"] = r
        layout.write_row_label(ws, r, L("debt_closing").en, L("debt_closing").secondary)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            c = ws.cell(
                row=r, column=col_idx,
                value=(
                    f"=${col}${block['opening']}"
                    f"+${col}${block['drawdown']}"
                    f"+${col}${block['repayment']}"
                ),
            )
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
            c.font = styles.font_subheader
            c.border = styles.BORDER_TOP_THIN
        r += 1

        # Average balance
        block["average"] = r
        layout.write_row_label(ws, r, L("debt_average").en, L("debt_average").secondary, indent=True)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            opening_ref = f"${col}${block['opening']}"
            closing_ref = f"${col}${block['closing']}"
            c = ws.cell(row=r, column=col_idx,
                        value=f"=({opening_ref}+{closing_ref})/2")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # All-in rate (constant within period, floor-adjusted)
        block["all_in_rate"] = r
        layout.write_row_label(ws, r, L("all_in_rate").en, L("all_in_rate").secondary, indent=True)
        ws.cell(row=r, column=3, value="%").font = styles.font_label_it
        ref_rate_name = tr.reference_rate.rate_decimal.name
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            c = ws.cell(
                row=r, column=col_idx,
                value=all_in_rate(ref_rate_name, tr.margin_bps.name, tr.floor_pct.name),
            )
            styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        r += 1

        # Cash interest (negative)
        #
        # v0.6 DESIGN NOTE: interest computed on BEGINNING-OF-PERIOD debt,
        # not average. Average-balance interest creates a circular via
        #   sweep[t] ← FCF[t] ← tax[t] ← PBT[t] ← interest[t] ← avg[t] ← closing[t] ← sweep[t]
        # WSP / bulge-bracket convention: use BOP to avoid the circular.
        # Slight accuracy tradeoff but required when a cash-sweep is in play.
        #
        # UNI-1 FIX (v0.11): the loan is FUNDED at close (start of the
        # drawdown year, column index h), so a par lender earns a FULL first
        # coupon that period. BOP-on-opening makes opening[h]==prior closing
        # ==0, dropping the first coupon entirely (lender IRR < coupon — an
        # impossible result for a par loan with an upfront fee). For the
        # DRAWDOWN YEAR ONLY we therefore accrue on the CLOSING (funded)
        # balance; all later years stay BOP (no circular, since later closings
        # depend on the sweep). Scoped to model_type=="unitranche": the
        # lender-side direct-lending template. credit_memo and sponsor_lbo keep
        # the documented BOP-on-opening convention unchanged (their schedules
        # and downstream sweeps are byte-identical to before this fix).
        first_coupon_on_closing = getattr(spec, "model_type", "") == "unitranche"
        block["interest"] = r
        layout.write_row_label(ws, r, L("cash_interest").en, L("cash_interest").secondary)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            rate_ref = f"${col}${block['all_in_rate']}"
            if first_coupon_on_closing and i == h:
                # Drawdown year: accrue on the funded (closing) balance so the
                # first coupon is not lost. closing[h] depends only on the
                # constant drawdown (sweep is disabled in the drawdown year),
                # so this introduces no circular reference.
                bal_ref = f"${col}${block['closing']}"  # funded balance
            else:
                bal_ref = f"${col}${block['opening']}"  # BOP = opening balance
            c = ws.cell(row=r, column=col_idx, value=f"=-{bal_ref}*{rate_ref}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
            c.font = styles.font_subheader
        r += 1

        # Arrangement fee (paid at close only)
        block["arrangement_fee"] = r
        layout.write_row_label(ws, r, L("arrangement_fee").en, L("arrangement_fee").secondary, indent=True)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i == h:
                c = ws.cell(
                    row=r, column=col_idx,
                    value=f"={tr.amount.name}*{tr.arrangement_fee_pct.name}",
                )
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
            else:
                c = ws.cell(row=r, column=col_idx, value="=0")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 2

        tranche_blocks.append(block)

    # ─── Cash sweep block (if enabled) ─────────────────────────────────────
    #
    # DESIGN NOTE (v0.6): sweep uses PRIOR-period leverage to gate the
    # current-period sweep. This breaks the circular dependency
    #   closing[t] → amort[t] → sweep[t] → leverage[t] → closing[t]
    # by making sweep depend on leverage[t-1], which is already computed.
    # For the first operating year we compare against entry leverage
    # (drawdown / last-historical EBITDA).
    cash_sweep_row: int | None = None
    if spec.debt.cash_sweep.enabled and spec.debt.cash_sweep.sweep_pct and spec.debt.cash_sweep.trigger_leverage:
        layout.write_section_header(ws, r, L("debt_cash_sweep").en, L("debt_cash_sweep").secondary)
        r += 1

        # v0.8.9 US-582: for sponsor_lbo, register the 6 tier named ranges
        # before the sweep formula is emitted (so cash_sweep_tiered works).
        if getattr(spec, "model_type", "") == "sponsor_lbo":
            wb = ws.parent
            tier_defaults = [
                ("sweep_tier1_lev", 5.0, styles.FMT_MULTIPLE,
                 "Cash sweep tier 1 — leverage threshold (≥)"),
                ("sweep_tier1_pct", 1.0, styles.FMT_PCT,
                 "Cash sweep tier 1 — fraction of sweep_pct applied"),
                ("sweep_tier2_lev", 4.0, styles.FMT_MULTIPLE,
                 "Cash sweep tier 2 — leverage threshold (≥)"),
                ("sweep_tier2_pct", 0.75, styles.FMT_PCT,
                 "Cash sweep tier 2 — fraction of sweep_pct applied"),
                ("sweep_tier3_lev", 3.0, styles.FMT_MULTIPLE,
                 "Cash sweep tier 3 — leverage threshold (≥)"),
                ("sweep_tier3_pct", 0.50, styles.FMT_PCT,
                 "Cash sweep tier 3 — fraction of sweep_pct applied"),
            ]
            for name, default, fmt, label in tier_defaults:
                layout.write_row_label(ws, r, label, label, indent=True)
                c = ws.cell(row=r, column=4, value=default)
                styles.style_input(c, number_format=fmt)
                if name in wb.defined_names:
                    del wb.defined_names[name]
                wb.defined_names[name] = DefinedName(
                    name=name,
                    attr_text=f"'{ws.title}'!$D${r}",
                )
                r += 1
            r += 1  # blank separator

        # Interim leverage = (sum of pre-sweep closing) / EBITDA
        # Note: this row is informational only; the sweep itself uses the
        # PRIOR period's value of this row, so the cycle is broken.
        #
        # LBO-4 FIX (v0.11): for the CLOSE column (drawdown year, index h) the
        # debt is dated at close while the same-column projected EBITDA is the
        # FORWARD (Year-1) figure. Dividing close debt by forward EBITDA
        # understates entry leverage (e.g. 3.81x instead of the contractual
        # 4.13x = senior / trailing LTM EBITDA). Bulge-bracket convention quotes
        # opening/entry leverage on TRAILING (last-FY) EBITDA. So at the close
        # column we divide by the prior column's (entry/LTM) EBITDA; all later
        # columns keep same-period EBITDA (true interim leverage). No new
        # circular: entry EBITDA is a historical constant.
        interim_lev_row = r
        layout.write_row_label(
            ws, r, "Interim leverage (pre-sweep)", "Leva pre-sweep", indent=True,
        )
        ws.cell(row=r, column=3, value="x").font = styles.font_label_it
        ebitda_row = int(operating_refs["ebitda_row"])
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            # Sum of all tranche closings this period
            parts = [f"${col}${b['closing']}" for b in tranche_blocks]
            debt_sum = "(" + "+".join(parts) + ")"
            # Close column → trailing (entry/LTM) EBITDA; otherwise same-period.
            ebitda_col = layout.year_col(i - 1) if (i == h and h > 0) else col
            ebitda_ref = f"'{operating_sheet_name}'!{ebitda_col}{ebitda_row}"
            c = ws.cell(row=r, column=col_idx,
                        value=f"=IFERROR({debt_sum}/{ebitda_ref},0)")
            styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
        r += 1

        # Sweep amount — uses PRIOR period leverage (breaks circular)
        cash_sweep_row = r
        layout.write_row_label(ws, r, "Cash sweep (applied to senior)",
                               "Cash sweep (applicato al senior)", indent=True)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
        fcf_row = int(operating_refs["fcf_row"])
        sweep_pct_name = spec.debt.cash_sweep.sweep_pct.name
        trigger_name = spec.debt.cash_sweep.trigger_leverage.name
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i <= h:
                # No sweep in historical years or the drawdown year itself
                # (i == h is the drawdown year; industry convention is a
                # grace period — no amortization or sweep in Y1 of the
                # loan, and using same-period leverage here would re-open
                # the circular-reference cycle).
                c = ws.cell(row=r, column=col_idx, value="=0")
                styles.style_formula(c, number_format=styles.FMT_EUR_M)
                continue
            fcf_ref = f"'{operating_sheet_name}'!{col}{fcf_row}"
            # PRIOR-period leverage to break the circular reference:
            #   closing[t] → amort[t] → sweep[t] → leverage[t] → closing[t]
            # becomes
            #   closing[t] ← amort[t] ← sweep[t] ← leverage[t-1]  (no cycle).
            prior_col = layout.year_col(i - 1)
            lev_ref = f"${prior_col}${interim_lev_row}"
            # v0.8.7 US-542: Sponsor LBO uses TIERED sweep (step-down by
            # leverage: 100% ≥5x, 75% 4-5x, 50% 3-4x, 0% <3x). All other
            # templates (unitranche / credit_memo) use single-tier sweep.
            if getattr(spec, "model_type", "") == "sponsor_lbo":
                from modelforge.builder.formulas import cash_sweep_tiered
                c = ws.cell(
                    row=r, column=col_idx,
                    value=cash_sweep_tiered(fcf_ref, lev_ref, sweep_pct_name),
                )
            else:
                c = ws.cell(
                    row=r, column=col_idx,
                    value=cash_sweep(fcf_ref, lev_ref, trigger_name, sweep_pct_name),
                )
            styles.style_formula(c, number_format=styles.FMT_EUR_M)

            # Apply the sweep back to the SENIOR tranche's repayment row.
            # Senior = first tranche (by convention).
            if tranche_blocks:
                senior_repay_row = tranche_blocks[0]["repayment"]
                senior_repay_cell = ws.cell(row=senior_repay_row, column=col_idx)
                existing = senior_repay_cell.value
                if existing in (None, 0):
                    senior_repay_cell.value = f"=${col}${cash_sweep_row}"
                elif isinstance(existing, str) and existing.startswith("="):
                    # Append sweep to existing repayment formula
                    senior_repay_cell.value = existing + f"+${col}${cash_sweep_row}"
                styles.style_formula(senior_repay_cell, number_format=styles.FMT_EUR_M)
        r += 2

    # ─── v0.7: Sources & Uses table (bulge-bracket standard) ──────────────
    layout.write_section_header(ws, r, "Sources & Uses of Funds",
                                "Fonti e impieghi")
    r += 1

    # Uses
    ws.cell(row=r, column=1, value="USES").font = styles.font_subheader
    r += 1
    ws.cell(row=r, column=1, value="Purchase equity (offer × shares)").font = styles.font_label_en
    ws.cell(row=r, column=4, value="(See DealStructure if available)").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1, value="Refinance existing debt").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=0").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1, value="M&A advisory fees (expensed at close)").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=0").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1, value="Financing fees (capitalized)").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=0").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1, value="OID discount").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=0").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1, value="Minimum cash to BS").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=0").font = styles.font_label_it
    r += 2

    ws.cell(row=r, column=1, value="SOURCES").font = styles.font_subheader
    r += 1
    total_debt_row = r
    parts = [f"${layout.year_col(h)}${b['closing']}" for b in tranche_blocks]
    ws.cell(row=r, column=1, value="New debt raised (all tranches)").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=" + "+".join(parts)).number_format = styles.FMT_EUR_M
    r += 1
    ws.cell(row=r, column=1, value="Sponsor equity").font = styles.font_label_en
    ws.cell(row=r, column=4, value="(Balance plug)").font = styles.font_label_it
    r += 1
    ws.cell(row=r, column=1, value="Management rollover").font = styles.font_label_en
    ws.cell(row=r, column=4, value="=0").font = styles.font_label_it
    r += 2

    # Check: Sources must equal Uses — in full sponsor LBO (skip for unitranche lender view)
    ws.cell(row=r, column=1, value="Check: Sources = Uses").font = styles.font_subheader
    ws.cell(row=r, column=4, value="(Sponsor equity plugs the balance — full S&U in v0.8 SponsorLBO)").font = styles.font_label_it
    r += 2

    # ─── Additional bulge-tier rows (stubs for v0.8 full sponsor LBO) ───
    ws.cell(row=r, column=1, value="— Additional bulge-tier stubs —").font = styles.font_subheader
    r += 1
    stubs = [
        ("Purchase price build", "Offer premium × FD shares + existing net debt + fees"),
        ("Goodwill created (PPA)", "Equity price − BV equity − write-ups + DTL"),
        ("OID amortization schedule", "Straight-line over tenor, CFS addback"),
        ("PIK toggle (payment-in-kind)", "Accrues to principal if cash insufficient"),
        ("Revolver facility", "Auto-draw when min cash breached; commitment fee on undrawn"),
        ("Management rollover / MIP", "4-year vest, 8-12% post-close equity"),
        ("Dividend recap path", "Refinance to target leverage at year 3"),
        ("Earnout / CVR", "Contingent consideration at FV"),
        ("NWC closing adjustment", "Target peg + true-up mechanism"),
        ("Exit scenarios (3)", "Strategic + IPO + secondary LBO"),
        ("Hurdle (reverse-solve max price)", "At target 20%/25%/30% IRR"),
        ("Sponsor GP promote (fund-level)", "Pref 8% + catchup + 20% carry"),
    ]
    for en, desc in stubs:
        ws.cell(row=r, column=1, value=f"  • {en}").font = styles.font_label_en
        ws.cell(row=r, column=4, value=desc).font = styles.font_label_it
        r += 1
    ws.cell(row=r, column=1, value="→ Full sponsor-LBO template in v0.8 (SponsorLBOSpec)").font = styles.font_label_it
    r += 2

    # ─── Totals block ──────────────────────────────────────────────────────
    layout.write_section_header(ws, r, L("debt_totals_tranches").en, L("debt_totals_tranches").secondary)
    r += 1

    # Total closing debt
    total_closing_row = r
    layout.write_row_label(ws, r, L("debt_total_debt").en, L("debt_total_debt").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        parts = [f"${col}${b['closing']}" for b in tranche_blocks]
        formula = "=" + "+".join(parts)
        c = ws.cell(row=r, column=col_idx, value=formula)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 1

    # Total interest
    total_interest_row = r
    layout.write_row_label(ws, r, L("debt_total_interest").en, L("debt_total_interest").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        parts = [f"${col}${b['interest']}" for b in tranche_blocks]
        formula = "=" + "+".join(parts)
        c = ws.cell(row=r, column=col_idx, value=formula)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # ── Patch Operating Model's interest row ───────────────────────────────
    interest_row_op = int(operating_refs["interest_row"])
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        formula_target_cell = f"'{operating_sheet_name}'!{col}{interest_row_op}"
        # We can't mutate operating through its sheet object here (we'd need
        # the ws). Instead, we return the ref and the top-level builder will
        # wire it — but to keep things self-contained, we'll let operating's
        # interest remain 0 and add a note. Actually we need to do it:
        # fetch operating worksheet through workbook.
        wb = ws.parent
        op_ws = wb[operating_sheet_name]
        op_cell = op_ws.cell(row=interest_row_op, column=col_idx)
        op_cell.value = f"='{ws.title}'!{col}{total_interest_row}"
        styles.style_xref(op_cell, number_format=styles.FMT_EUR_M)

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {
        "total_closing_row": str(total_closing_row),
        "total_interest_row": str(total_interest_row),
        "tranche_blocks": tranche_blocks,  # type: ignore[dict-item]
    }
