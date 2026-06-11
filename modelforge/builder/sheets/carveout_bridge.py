"""Carve-out standalone-EBITDA bridge + carve-out EV — HGB carve-out template.

A carve-out's *reported* EBITDA (inside the seller's group) is not the EBITDA
the buyer acquires. This sheet renders the standard carve-out normalization
bridge and the resulting carve-out Enterprise Value:

    reported EBITDA
      (+) allocated corporate costs   add-back (seller over-allocation removed)
      (-) dis-synergies / stranded    recurring stand-alone cost the unit now bears
      (-) TSA cost                    transition-period only (excluded from run-rate)
      (-) one-time separation         one-off (excluded from run-rate)
      = standalone adjusted EBITDA    (during the TSA window)

      run-rate (steady-state) EBITDA  = standalone adjusted + TSA + one-time
                                        (transitory lines added back out)
      carve-out EV = run-rate standalone EBITDA × entry multiple

Bridge integrity (asserted by hardtest against an independent clean-room calc):
    * the four signed bridge components sum EXACTLY to (standalone − reported);
    * the steady-state line excludes TSA + one-time (= reported + alloc − dis);
    * EV == steady-state standalone EBITDA × entry_multiple.

Styling: every assumption magnitude is a blue ``style_input`` cell; every
computed running-total / EV cell is a black ``style_formula`` cell. Each carries
an explicit number_format and a naming comment so the deliverable is auditable
and the certify styling-gap check stays clean.

References: BIWS / Macabacus carve-out adjustments; Rosenbaum-Pearl ("Investment
Banking") on stand-alone vs as-reported EBITDA; IDW S 1 stand-alone valuation.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout

# Fixed value column for the single-column bridge (year columns start at D).
VAL_COL = "D"
VAL_COL_IDX = 4
AUTHOR = "ModelForge"

# Bold subtotal/EV font that KEEPS an explicit (black formula) colour so the
# certify "no font colour" styling-gap check stays clean. styles.font_subheader
# is bold but colourless, which would register as a styling gap on these
# computed total cells; this variant is bold AND coloured.
_font_total = Font(
    name=styles.FONT_BASE,
    size=styles.FONT_SIZE_BODY,
    bold=True,
    color=styles.COLOR_FORMULA,
)


def _def(wb, name: str, sheet: str, cell: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sheet}'!{cell}")


def build(ws: Worksheet, spec) -> dict[str, int]:
    """Emit the Carve-out Bridge + carve-out EV sheet.

    Returns a dict of {logical_name: row} so the template / QC can reference
    bridge cells (e.g. for a QC parts-sum-to-total check).
    """
    cb = spec.carveout_bridge
    assert cb is not None, "carveout_bridge.build called without spec.carveout_bridge"

    layout.set_column_widths(ws, label_width=52, it_width=40, year_width=16, unit_width=6)
    layout.write_title_block(
        ws,
        title_en="Carve-out Standalone-EBITDA Bridge + EV",
        title_it="Standalone-EBITDA-Brücke (Carve-out) + Unternehmenswert",
        subtitle=("Reported → standalone adjusted EBITDA → run-rate → "
                  "carve-out Enterprise Value (EV = run-rate EBITDA × entry x)"),
    )

    wb = ws.parent
    cur = spec.meta.currency
    refs: dict[str, int] = {}
    r = 5

    # ── SECTION 1: Standalone-EBITDA bridge ───────────────────────────────
    layout.write_section_header(
        ws, r,
        "Standalone-EBITDA bridge (carve-out perimeter)",
        "Standalone-EBITDA-Brücke (Carve-out-Perimeter)",
    )
    r += 1

    # Reported EBITDA (input)
    layout.write_row_label(
        ws, r, "Reported EBITDA (as in seller's group books)",
        "Berichtetes EBITDA (Konzernbücher Verkäufer)",
    )
    ws.cell(row=r, column=3, value=cur).font = styles.font_label_it
    c = ws.cell(row=r, column=VAL_COL_IDX, value=float(cb.reported_ebitda))
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "reported_ebitda: carve-out unit EBITDA as reported inside the "
        "seller's consolidation (carve-out perimeter, last FY).",
        AUTHOR,
    )
    refs["reported"] = r
    r += 1

    # (+) Allocated corporate costs add-back (input magnitude, positive sign)
    layout.write_row_label(
        ws, r, "(+) Allocated corporate costs removed (add-back)",
        "(+) Entfallende Konzernumlagen (Hinzurechnung)", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=float(cb.allocated_corporate_costs))
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "allocated_corporate_costs: parent overhead allocation that does NOT "
        "transfer with the carve-out. Positive add-back → raises standalone "
        "EBITDA. Bridge sign: +.",
        AUTHOR,
    )
    refs["alloc"] = r
    r += 1

    # (-) Dis-synergies / stranded costs (input magnitude, applied negative)
    layout.write_row_label(
        ws, r, "(-) Dis-synergies / stranded costs (recurring)",
        "(-) Dissynergien / Stranded Costs (laufend)", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=float(cb.dis_synergies))
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "dis_synergies: recurring stand-alone costs (own ERP/HR/finance, lost "
        "group purchasing scale). Entered positive; bridge sign: −. Recurring "
        "→ INCLUDED in the steady-state run-rate.",
        AUTHOR,
    )
    refs["dis"] = r
    r += 1

    # (-) TSA cost (input magnitude, applied negative, time-bounded)
    tsa_years = float(cb.tsa_period_years)
    layout.write_row_label(
        ws, r, f"(-) TSA cost (transition only, {tsa_years:g}y window)",
        f"(-) TSA-Kosten (nur Übergang, {tsa_years:g} J.)", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=float(cb.tsa_costs))
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "tsa_costs: Transition Service Agreement costs paid to the seller "
        f"during the {tsa_years:g}-year TSA window. Entered positive; bridge "
        "sign: −. Time-bounded → EXCLUDED from the steady-state run-rate.",
        AUTHOR,
    )
    refs["tsa"] = r
    r += 1

    # (-) One-time separation costs (input magnitude, applied negative)
    layout.write_row_label(
        ws, r, "(-) One-time separation / stand-up costs",
        "(-) Einmalige Separations- / Aufbaukosten", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=float(cb.one_time_separation_costs))
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "one_time_separation_costs: one-off carve-out stand-up costs (IT "
        "cloning, rebranding, entity setup). Entered positive; bridge sign: "
        "−. One-off → EXCLUDED from the steady-state run-rate.",
        AUTHOR,
    )
    refs["onetime"] = r
    r += 1

    # = Standalone adjusted EBITDA (during the TSA window) — running total.
    # NB: labels must NOT begin with "=" (Excel/openpyxl would treat the cell
    # text as a formula and emit #NAME?). Use a word prefix instead.
    layout.write_row_label(
        ws, r, "Standalone adjusted EBITDA (during TSA)",
        "Standalone bereinigtes EBITDA (während TSA)",
    )
    rep = refs["reported"]; al = refs["alloc"]; di = refs["dis"]
    ts = refs["tsa"]; ot = refs["onetime"]
    c = ws.cell(
        row=r, column=VAL_COL_IDX,
        value=(f"={VAL_COL}{rep}+{VAL_COL}{al}-{VAL_COL}{di}"
               f"-{VAL_COL}{ts}-{VAL_COL}{ot}"),
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = _font_total
    c.border = styles.BORDER_TOP_THIN
    c.comment = Comment(
        "standalone_adjusted_ebitda = reported + allocated_corp_costs "
        "− dis_synergies − tsa_costs − one_time_separation. This is the "
        "during-TSA standalone EBITDA (transition drag still in). The signed "
        "bridge components sum EXACTLY to (standalone − reported).",
        AUTHOR,
    )
    refs["standalone_adjusted"] = r
    _def(wb, "carveout_standalone_adjusted_ebitda", ws.title, f"${VAL_COL}${r}")
    r += 2

    # ── SECTION 2: Steady-state run-rate (transitory lines excluded) ──────
    layout.write_section_header(
        ws, r,
        "Steady-state run-rate EBITDA (TSA + one-time excluded)",
        "Run-Rate-EBITDA (ohne TSA / Einmalkosten)",
    )
    r += 1

    # + TSA add-back (transition only)
    layout.write_row_label(
        ws, r, "(+) Add back TSA cost (transition only)",
        "(+) TSA-Kosten zurückgerechnet (nur Übergang)", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=f"={VAL_COL}{ts}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "TSA cost added back to reach the steady-state run-rate: the TSA is "
        "time-bounded and rolls off after the transition window, so it is "
        "EXCLUDED from the perpetuity base for the EV multiple.",
        AUTHOR,
    )
    refs["addback_tsa"] = r
    r += 1

    # + one-time add-back
    layout.write_row_label(
        ws, r, "(+) Add back one-time separation (one-off)",
        "(+) Einmalkosten zurückgerechnet", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=f"={VAL_COL}{ot}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "One-time separation cost added back: a one-off cost is not part of "
        "the recurring run-rate, so it is EXCLUDED from the EV multiple base.",
        AUTHOR,
    )
    refs["addback_onetime"] = r
    r += 1

    # = Steady-state run-rate standalone EBITDA
    layout.write_row_label(
        ws, r, "Run-rate standalone EBITDA (steady-state)",
        "Run-Rate Standalone-EBITDA (eingeschwungen)",
    )
    sa = refs["standalone_adjusted"]
    at = refs["addback_tsa"]; ao = refs["addback_onetime"]
    c = ws.cell(
        row=r, column=VAL_COL_IDX,
        value=f"={VAL_COL}{sa}+{VAL_COL}{at}+{VAL_COL}{ao}",
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = _font_total
    c.border = styles.BORDER_TOP_THIN
    c.comment = Comment(
        "run_rate_standalone_ebitda = standalone_adjusted + tsa_costs "
        "+ one_time_separation = reported + allocated_corp_costs "
        "− dis_synergies. Transitory TSA + one-time lines are excluded; the "
        "recurring dis-synergy stays in. This run-rate is the base for the "
        "carve-out EV multiple.",
        AUTHOR,
    )
    refs["run_rate"] = r
    _def(wb, "carveout_run_rate_ebitda", ws.title, f"${VAL_COL}${r}")
    r += 2

    # ── SECTION 3: Carve-out Enterprise Value ─────────────────────────────
    layout.write_section_header(
        ws, r,
        "Carve-out Enterprise Value (EV = run-rate EBITDA × entry x)",
        "Carve-out-Unternehmenswert (EV = Run-Rate × Eintrittsmultiplikator)",
    )
    r += 1

    # Entry multiple (input)
    layout.write_row_label(
        ws, r, "Entry multiple (EV / standalone EBITDA)",
        "Eintrittsmultiplikator (EV / Standalone-EBITDA)", indent=True,
    )
    c = ws.cell(row=r, column=VAL_COL_IDX, value=float(cb.entry_multiple))
    styles.style_input(c, number_format=styles.FMT_MULTIPLE)
    c.comment = Comment(
        "entry_multiple: EV / standalone-EBITDA paid at entry, applied to the "
        "steady-state run-rate standalone EBITDA.",
        AUTHOR,
    )
    refs["entry_multiple"] = r
    rr = refs["run_rate"]; em = refs["entry_multiple"]
    r += 1

    # = Carve-out EV
    layout.write_row_label(
        ws, r, "Carve-out Enterprise Value (EV)",
        "Carve-out-Unternehmenswert (EV)",
    )
    c = ws.cell(
        row=r, column=VAL_COL_IDX,
        value=f"={VAL_COL}{rr}*{VAL_COL}{em}",
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = _font_total
    c.border = styles.BORDER_TOP_THICK
    c.comment = Comment(
        "carve_out_EV = run-rate standalone EBITDA × entry_multiple. The EV "
        "multiple is applied to the steady-state run-rate (transitory TSA + "
        "one-time costs excluded), the audit-defensible perpetuity base.",
        AUTHOR,
    )
    refs["ev"] = r
    _def(wb, "carveout_enterprise_value", ws.title, f"${VAL_COL}${r}")
    r += 2

    # ── SECTION 4: Bridge integrity checks (live, self-documenting) ───────
    layout.write_section_header(
        ws, r,
        "Bridge integrity checks (must equal 1)",
        "Brücken-Integritätsprüfungen (müssen 1 ergeben)",
    )
    r += 1

    # Check 1: signed bridge parts sum to (standalone − reported)
    layout.write_row_label(
        ws, r, "Bridge parts sum to (standalone − reported)",
        "Brücken-Komponenten = (Standalone − Berichtet)", indent=True,
    )
    parts = (f"{VAL_COL}{al}-{VAL_COL}{di}-{VAL_COL}{ts}-{VAL_COL}{ot}")
    c = ws.cell(
        row=r, column=VAL_COL_IDX,
        value=(f"=IF(ABS(({parts})-({VAL_COL}{sa}-{VAL_COL}{rep}))"
               f"<=0.0001,1,0)"),
    )
    styles.style_formula(c, number_format=styles.FMT_INTEGER)
    c.alignment = styles.align_center
    refs["chk_parts"] = r
    r += 1

    # Check 2: run-rate excludes TSA + one-time
    layout.write_row_label(
        ws, r, "Run-rate = reported + alloc − dis (TSA/one-time excluded)",
        "Run-Rate = Berichtet + Umlage − Dissynergien", indent=True,
    )
    c = ws.cell(
        row=r, column=VAL_COL_IDX,
        value=(f"=IF(ABS({VAL_COL}{rr}-"
               f"({VAL_COL}{rep}+{VAL_COL}{al}-{VAL_COL}{di}))<=0.0001,1,0)"),
    )
    styles.style_formula(c, number_format=styles.FMT_INTEGER)
    c.alignment = styles.align_center
    refs["chk_runrate"] = r
    r += 1

    # Check 3: EV == run-rate × multiple
    layout.write_row_label(
        ws, r, "EV = run-rate standalone EBITDA × entry multiple",
        "EV = Run-Rate-EBITDA × Multiplikator", indent=True,
    )
    ev = refs["ev"]
    c = ws.cell(
        row=r, column=VAL_COL_IDX,
        value=(f"=IF(ABS({VAL_COL}{ev}-{VAL_COL}{rr}*{VAL_COL}{em})"
               f"<=0.0001,1,0)"),
    )
    styles.style_formula(c, number_format=styles.FMT_INTEGER)
    c.alignment = styles.align_center
    refs["chk_ev"] = r
    r += 1

    ws.freeze_panes = f"{VAL_COL}5"
    ws.print_title_cols = "A:C"
    return refs
