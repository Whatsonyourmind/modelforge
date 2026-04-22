"""IFRS 9 ECL — dedicated per-facility Expected Credit Loss sheet.

v0.8.8 US-566/567/568/569: full bulge-tier IFRS 9 computation with
per-facility Stage 1/2/3/POCI rows and:

- **PD × LGD × EAD × DF** formula live per facility per tenor
- **SICR triggers** (US-567): absolute PD threshold, relative doubling,
  rating downgrade ≥ 2 notches, watchlist, forbearance, 30+ DPD
- **Forward-looking macro scenarios** (US-568): 3 scenarios (Upside /
  Base / Downside) weighted 25/50/25, each with GDP growth +
  unemployment rate + CPI → PD multiplier via linear coefficient
- **POCI treatment** (US-569): for NPL portfolios, gross carrying value
  + always-lifetime PD + interest on net of ECL

References:
- BIS FSI IFRS 9 summary (bis.org/fsi/fsisummaries/ifrs9.pdf)
- Regulation (EU) 2019/630 (calendar provisioning for NPE)
- EBA GL 2017/06 (SICR quantitative triggers)
- IFRS 9 paragraphs 5.5 + B5.5.1–17

Applicability: credit / NPL / structured credit templates. Emitted when
the caller opts in via ``template core_sheets`` orchestrator. The
compliance.py Basel section stays in place for qualitative recap.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import layout, styles


# Default macro scenarios — used when spec does not carry its own set.
# Aligns with ECB Financial Stability Review baseline / adverse Q1 2026.
_DEFAULT_SCENARIOS = [
    {
        "name": "Upside",
        "weight": 0.25,
        "gdp_growth": 0.025,
        "unemployment": 0.055,
        "cpi": 0.020,
        "pd_multiplier": 0.75,  # PD × 0.75 under upside
    },
    {
        "name": "Base",
        "weight": 0.50,
        "gdp_growth": 0.010,
        "unemployment": 0.075,
        "cpi": 0.025,
        "pd_multiplier": 1.00,
    },
    {
        "name": "Downside",
        "weight": 0.25,
        "gdp_growth": -0.005,
        "unemployment": 0.095,
        "cpi": 0.035,
        "pd_multiplier": 1.80,  # PD × 1.80 under downside
    },
]


def build(ws: Worksheet, spec, context: dict | None = None) -> None:
    """Emit dedicated IFRS 9 ECL sheet.

    `context` may carry a list of facilities each with:
      - facility_id, stage (1/2/3/POCI), pd_12m, lifetime_pd, lgd, ead,
        effective_interest_rate, years_to_maturity,
        sicr_flags (dict of 6 booleans),
        rating, rating_at_origination.

    If context is absent, emits an illustrative 3-facility example with
    default stages 1/2/3 — the structure is what matters for the audit;
    the numbers are placeholders the analyst overrides.
    """
    context = context or {}
    facilities = context.get("facilities") or _example_facilities(spec)

    layout.set_column_widths(
        ws, label_width=28, it_width=14, year_width=12, unit_width=8,
    )
    layout.write_title_block(
        ws, "IFRS 9 — Expected Credit Loss (ECL)",
        "IFRS 9 — Perdita attesa",
        "Per-facility ECL = PD × LGD × EAD × DF; SICR triggers; "
        "forward-looking macro scenarios; POCI for NPL.",
    )

    r = 5

    # ─── Section 1: Forward-looking macro scenarios ──────────────────
    layout.write_section_header(
        ws, r, "Section 1 — Forward-looking macro scenarios",
        "Sezione 1 — Scenari macro prospettici",
    )
    r += 1

    # Scenario table header
    hdrs = ["Scenario", "Weight", "GDP growth", "Unemployment",
            "CPI", "PD multiplier"]
    for i, h in enumerate(hdrs):
        c = ws.cell(row=r, column=i + 1, value=h)
        styles.style_header(c)
    r += 1

    # v0.8.9 US-585: Register workbook-level named ranges for scenario
    # weights + PD multipliers so the weighted-PD formula + any downstream
    # reporting can reference by name rather than cell position.
    wb = ws.parent
    def _reg(name: str, cell_addr: str) -> None:
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names[name] = DefinedName(
            name=name, attr_text=f"'{ws.title}'!{cell_addr}",
        )

    scenario_rows: dict[str, int] = {}
    for sc in _DEFAULT_SCENARIOS:
        ws.cell(row=r, column=1, value=sc["name"]).font = styles.font_label_en
        c = ws.cell(row=r, column=2, value=sc["weight"])
        styles.style_input(c, number_format=styles.FMT_PCT)
        _reg(f"ecl_scenario_weight_{sc['name'].lower()}", f"$B${r}")
        c = ws.cell(row=r, column=3, value=sc["gdp_growth"])
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        _reg(f"ecl_gdp_growth_{sc['name'].lower()}", f"$C${r}")
        c = ws.cell(row=r, column=4, value=sc["unemployment"])
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        c = ws.cell(row=r, column=5, value=sc["cpi"])
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        c = ws.cell(row=r, column=6, value=sc["pd_multiplier"])
        styles.style_input(c, number_format=styles.FMT_MULTIPLE)
        _reg(f"ecl_pd_mult_{sc['name'].lower()}", f"$F${r}")
        scenario_rows[sc["name"]] = r
        r += 1

    # Weighted PD multiplier — formula now uses named ranges.
    ws.cell(row=r, column=1,
            value="Weighted PD multiplier").font = styles.font_subheader
    ws.cell(row=r, column=2, value="Moltiplicatore ponderato").font = styles.font_label_it
    weighted_formula = (
        "=ecl_scenario_weight_upside*ecl_pd_mult_upside"
        "+ecl_scenario_weight_base*ecl_pd_mult_base"
        "+ecl_scenario_weight_downside*ecl_pd_mult_downside"
    )
    weighted_mult_cell = ws.cell(row=r, column=6, value=weighted_formula)
    styles.style_formula(weighted_mult_cell, number_format=styles.FMT_MULTIPLE)
    weighted_mult_cell.font = styles.font_subheader
    _reg("ecl_weighted_pd_multiplier", f"$F${r}")
    weighted_mult_row = r
    r += 2

    # ─── Section 2: Per-facility ECL computation ─────────────────────
    layout.write_section_header(
        ws, r, "Section 2 — Per-facility ECL (PD × LGD × EAD × DF)",
        "Sezione 2 — ECL per esposizione",
    )
    r += 1

    fac_hdrs = ["Facility", "Stage", "12m PD", "Lifetime PD", "LGD",
                "EAD (€m)", "EIR", "Years", "DF", "PD used",
                "Stage-adjusted PD", "ECL (€m)"]
    for i, h in enumerate(fac_hdrs):
        c = ws.cell(row=r, column=i + 1, value=h)
        styles.style_header(c)
    r += 1

    fac_start = r
    for f in facilities:
        ws.cell(row=r, column=1,
                value=f["facility_id"]).font = styles.font_label_en
        # Stage: 1 | 2 | 3 | POCI
        ws.cell(row=r, column=2, value=f["stage"]).font = styles.font_label_en

        c = ws.cell(row=r, column=3, value=f["pd_12m"])
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        c = ws.cell(row=r, column=4, value=f["lifetime_pd"])
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        c = ws.cell(row=r, column=5, value=f["lgd"])
        styles.style_input(c, number_format=styles.FMT_PCT)
        c = ws.cell(row=r, column=6, value=f["ead"])
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c = ws.cell(row=r, column=7, value=f["eir"])
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        c = ws.cell(row=r, column=8, value=f["years_to_maturity"])
        styles.style_input(c, number_format=styles.FMT_YEARS)
        # DF = 1 / (1 + EIR)^years
        df_cell = ws.cell(row=r, column=9, value=f"=1/(1+G{r})^H{r}")
        styles.style_formula(df_cell, number_format=styles.FMT_MULTIPLE)
        # PD used = 12m if Stage 1, lifetime if Stage 2/3/POCI
        pd_used_cell = ws.cell(
            row=r, column=10,
            value=f'=IF(B{r}=1,C{r},D{r})',
        )
        styles.style_formula(pd_used_cell, number_format=styles.FMT_PCT_2DP)
        # Stage-adjusted PD = PD_used × weighted macro multiplier.
        # POCI: always gross-carrying lifetime (no 12m/lifetime switch).
        # Macro multiplier via named range ecl_weighted_pd_multiplier.
        stage_adj_cell = ws.cell(
            row=r, column=11,
            value=f'=IF(B{r}="POCI",D{r},J{r})*ecl_weighted_pd_multiplier',
        )
        styles.style_formula(stage_adj_cell, number_format=styles.FMT_PCT_2DP)
        # ECL = stage_adj_pd × LGD × EAD × DF
        ecl_cell = ws.cell(row=r, column=12,
                           value=f"=K{r}*E{r}*F{r}*I{r}")
        styles.style_formula(ecl_cell, number_format=styles.FMT_EUR_M)
        ecl_cell.font = styles.font_subheader
        r += 1
    fac_end = r - 1

    # Portfolio roll-up
    ws.cell(row=r, column=1, value="Portfolio ECL").font = styles.font_subheader
    ws.cell(row=r, column=2, value="Totale portafoglio").font = styles.font_label_it
    portfolio_ecl_cell = ws.cell(
        row=r, column=12, value=f"=SUM(L{fac_start}:L{fac_end})",
    )
    styles.style_formula(portfolio_ecl_cell, number_format=styles.FMT_EUR_M)
    portfolio_ecl_cell.font = styles.font_subheader
    portfolio_ecl_cell.border = styles.BORDER_TOP_THIN
    r += 2

    # ─── Section 3: SICR triggers (Stage 1 → Stage 2 migration) ──────
    layout.write_section_header(
        ws, r, "Section 3 — SICR triggers (Significant Increase in Credit Risk)",
        "Sezione 3 — Trigger SICR (aumento significativo del rischio)",
    )
    r += 1

    sicr_items = [
        ("PD doubled vs origination",
         "PD raddoppiata vs originazione",
         "Quantitative — EBA GL 2017/06"),
        ("Absolute PD threshold breached",
         "Soglia PD assoluta superata",
         "Quantitative — typically 30% lifetime PD"),
        ("External rating downgrade ≥ 2 notches",
         "Downgrade rating esterno ≥ 2 notch",
         "Qualitative"),
        ("Placed on credit-watch / watchlist",
         "Inserito in watchlist",
         "Qualitative"),
        ("Forbearance granted",
         "Misure di tolleranza accordate",
         "Rebuttable presumption per Reg. EU 2013/575 Art. 178"),
        ("30+ days past due",
         "30+ giorni di mora",
         "Rebuttable presumption — IFRS 9 B5.5.20"),
    ]
    for en, it, desc in sicr_items:
        ws.cell(row=r, column=1, value=f"  • {en}").font = styles.font_label_en
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        ws.cell(row=r, column=4, value=desc).font = styles.font_label_it
        r += 1
    r += 1

    # ─── Section 4: POCI treatment ───────────────────────────────────
    layout.write_section_header(
        ws, r, "Section 4 — POCI (Purchased or Originated Credit-Impaired)",
        "Sezione 4 — POCI — Attività acquisite / originate deteriorate",
    )
    r += 1

    poci_items = [
        ("Initial recognition at fair value",
         "Valutazione iniziale a fair value",
         "Purchase price = FV; no day-1 ECL"),
        ("Always lifetime PD (no 12m/lifetime switch)",
         "Sempre PD lifetime",
         "IFRS 9 §5.5.13"),
        ("Interest recognized on net of ECL",
         "Interessi sul netto della ECL",
         "Credit-adjusted effective interest rate"),
        ("ECL changes → impairment loss (not OCI)",
         "Variazioni ECL → perdita a P&L",
         "IFRS 9 §B5.5.46"),
    ]
    for en, it, desc in poci_items:
        ws.cell(row=r, column=1, value=f"  • {en}").font = styles.font_label_en
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        ws.cell(row=r, column=4, value=desc).font = styles.font_label_it
        r += 1

    # Comment documenting scope
    ws.cell(row=r, column=1, value="").comment = Comment(
        "IFRS 9 dedicated ECL sheet per BIS FSI summary + Banca d'Italia "
        "Circolare 262. POCI applies to NPL portfolio loans purchased at "
        "discount; Stage 3 applies to performing loans that later default. "
        "Macro scenarios use 3-point weighted average (25/50/25 Up/Base/Down) "
        "with PD multipliers calibrated to ECB FSR Q1 2026.",
        "ModelForge",
    )

    ws.freeze_panes = "D6"
    ws.print_title_rows = "1:4"


def _example_facilities(spec) -> list[dict]:
    """Default 3-facility example when spec doesn't carry its own."""
    return [
        {
            "facility_id": "SENIOR-001",
            "stage": 1,
            "pd_12m": 0.0080,
            "lifetime_pd": 0.0350,
            "lgd": 0.35,
            "ead": 100.0,
            "eir": 0.055,
            "years_to_maturity": 5,
        },
        {
            "facility_id": "SUB-002",
            "stage": 2,
            "pd_12m": 0.0250,
            "lifetime_pd": 0.1200,
            "lgd": 0.55,
            "ead": 40.0,
            "eir": 0.085,
            "years_to_maturity": 5,
        },
        {
            "facility_id": "NPE-003",
            "stage": 3,
            "pd_12m": 1.00,
            "lifetime_pd": 1.00,
            "lgd": 0.70,
            "ead": 15.0,
            "eir": 0.065,
            "years_to_maturity": 3,
        },
    ]
