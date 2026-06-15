"""Bank capital — RWA, CET1, leverage ratio, MDA headroom.

RWA is driven by RISK-BEARING earning assets (gross loans + securities) via a
Standardised-approach density — NOT the cash-plugged total assets, which keeps
the capital block independent of the funding/cash plugs (no circularity).

CET1 is DERIVED from the BalanceSheet common-equity walk:
    CET1 = common equity − intangibles − regulatory adjustments
so there is no parallel CET1 roll-forward to drift; the single equity
roll-forward lives on the BalanceSheet (symbolic, drift-guarded). The QC sheet
verifies CET1 generation ties to retained earnings net of distributions.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, bs_refs: dict[str, str], bs_sheet: str) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p
    cur = spec.meta.currency

    gl = int(bs_refs["gross_loans_row"])
    sec = int(bs_refs["securities_row"])
    eq = int(bs_refs["equity_closing_row"])
    intang = int(bs_refs["intangibles_row"])
    ta = int(bs_refs["total_assets_row"])
    at1 = int(bs_refs["at1_row"])

    layout.set_column_widths(ws, label_width=48, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Capital (RWA, CET1, Leverage)", title_it="Capitale (RWA, CET1, Leva)",
        subtitle=f"Std-approach RWA density · CET1 = equity − intangibles − adj · "
                 f"{spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        cc = ws.cell(row=yr_row, column=ci, value=f"{'A' if i < h else 'E'} {yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    def _bs(col, row):
        return f"'{bs_sheet}'!${col}${row}"

    # ── RWA ───────────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "Risk-weighted assets", "Attività ponderate per il rischio")
    r += 1

    rows["risk_bearing"] = r
    layout.write_row_label(ws, r, "Risk-bearing assets (gross loans + securities)",
                           "Attività a rischio (crediti lordi + titoli)", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"={_bs(col, gl)}+{_bs(col, sec)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["rwa"] = r
    layout.write_row_label(ws, r, "Risk-weighted assets (RWA)", "RWA")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=${col}${rows['risk_bearing']}*rwa_density")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    ws.cell(row=r, column=4).comment = Comment(
        "RWA = density × risk-bearing earning assets (Standardised-approach "
        "proxy). NOT an A-IRB PD/LGD/M computation; density is the sell-side / "
        "equity-research convention.", "ModelForge")
    r += 2

    # ── CET1 (derived from the equity walk) ───────────────────────────────
    layout.write_section_header(ws, r, "Common Equity Tier 1 (CET1)", "Capitale primario di classe 1 (CET1)")
    r += 1

    rows["common_equity"] = r
    layout.write_row_label(ws, r, "Common equity (from balance sheet)", "Patrimonio comune (da SP)", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"={_bs(col, eq)}")
        styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["less_intangibles"] = r
    layout.write_row_label(ws, r, "(−) Goodwill & intangibles", "(−) Avviamento e immateriali", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=-{_bs(col, intang)}")
        styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["reg_adj"] = r
    reg_adj = float(spec.cet1_regulatory_adjustments_eur_m)
    layout.write_row_label(ws, r, "(−) Regulatory adjustments (CRR Art.26-36 proxy)",
                           "(−) Rettifiche regolamentari", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=-reg_adj)
        styles.style_input(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Single CRR Art. 26-36 proxy plug (DTAs, prudent valuation, IFRS 9 "
        "transitional, etc.). Replace with a COREP figure for a real bank. "
        "Negative = a deduction from CET1.", "ModelForge")
    r += 1

    rows["cet1"] = r
    layout.write_row_label(ws, r, "CET1 capital", "Capitale CET1")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=(f"=${col}${rows['common_equity']}+${col}${rows['less_intangibles']}"
                           f"+${col}${rows['reg_adj']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # ── Ratios ────────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "Capital ratios", "Coefficienti patrimoniali")
    r += 1

    rows["cet1_ratio"] = r
    layout.write_row_label(ws, r, "CET1 ratio", "Coefficiente CET1")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=IFERROR(${col}${rows['cet1']}/${col}${rows['rwa']},0)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        c.font = styles.font_subheader
    r += 1

    rows["cet1_requirement"] = r
    layout.write_row_label(ws, r, "CET1 requirement (OCR)", "Requisito CET1 (OCR)", indent=True)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value="=cet1_requirement_ratio")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["cet1_surplus"] = r
    layout.write_row_label(ws, r, "CET1 surplus over requirement (ratio pts)",
                           "Eccedenza CET1 sul requisito", indent=True)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['cet1_ratio']}-${col}${rows['cet1_requirement']}")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    # Tier 1 capital = CET1 + AT1
    rows["tier1"] = r
    layout.write_row_label(ws, r, "Tier 1 capital (CET1 + AT1)", "Capitale di classe 1", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=f"=${col}${rows['cet1']}+{_bs(col, at1)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["leverage_exposure"] = r
    layout.write_row_label(ws, r, "Leverage exposure (total assets − intangibles)",
                           "Esposizione leva", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        # CRR Art. 429a: the exposure measure deducts items deducted from Tier 1
        # capital (notably goodwill & intangibles), so numerator and denominator
        # are consistent. Proxy still omits off-balance-sheet add-ons (stated).
        c = ws.cell(row=r, column=ci, value=f"={_bs(col, ta)}-{_bs(col, intang)}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["leverage_ratio"] = r
    layout.write_row_label(ws, r, "Leverage ratio (Tier 1 / exposure)", "Coefficiente di leva")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=IFERROR(${col}${rows['tier1']}/${col}${rows['leverage_exposure']},0)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["mda_headroom"] = r
    layout.write_row_label(ws, r, "MDA / mgmt-buffer headroom (ratio pts)",
                           "Margine MDA / buffer gestionale")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=(f"=${col}${rows['cet1_ratio']}"
                           f"-(cet1_requirement_ratio+mda_buffer_pct)"))
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    ws.cell(row=r, column=4).comment = Comment(
        "CET1 ratio less (requirement + management buffer). Negative ⇒ inside "
        "the MDA restriction zone (distributions throttle on the Capital Return "
        "sheet).", "ModelForge")
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {f"{k}_row": str(v) for k, v in rows.items()}
