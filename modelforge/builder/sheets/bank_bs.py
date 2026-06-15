"""Bank balance sheet — volume roll-forwards, allowance stock, equity walk.

Built FIRST (before NII/P&L/Capital/CapitalReturn) because every other sheet
reads its volume balances. Three cross-sheet-dependent cells are written as
``=0`` placeholders and PATCHED by the template orchestrator once P&L and
CapitalReturn exist (the same patch-back pattern debt.py uses on the operating
sheet):

    * Impairment charge      ← P&L loan-loss provisions
    * Retained earnings       ← P&L net income attributable to CET1
    * Dividends / Buybacks    ← CapitalReturn

ACYCLICITY: the wholesale-funding line is the FUNDING PLUG (sized off the
asset side less the other L&E items, so it depends on closing equity); cash is
the ASSET PLUG (residual on the L&E side). RWA on the Capital sheet is driven
by risk-bearing assets (loans + securities), NOT this cash-plugged total, and
NII charges the wholesale plug on its prior-closing balance — so nothing here
feeds back into the same-period interest/earnings, and no iteration is needed.

ROLL-FORWARD DISCIPLINE: the two opening balances (allowance, common equity)
reference the prior period's CLOSING row SYMBOLICALLY (a reserved row variable
with an ``assert`` drift-guard), never a literal cursor offset — the off-by-one
class fixed across the debt schedules.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p
    ob = spec.opening_bs

    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Balance Sheet", title_it="Stato patrimoniale",
        subtitle=f"Volume roll-forwards · allowance stock · equity walk · "
                 f"{spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        cc = ws.cell(row=yr_row, column=ci, value=f"{'A' if i < h else 'E'} {yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7
    cur = spec.meta.currency

    def _unit(row, txt):
        ws.cell(row=row, column=3, value=txt).font = styles.font_label_it

    # ── ASSETS ────────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "Assets", "Attività")
    r += 1

    # Gross loans — volume roll-forward (prior × (1+loan_growth))
    rows["gross_loans"] = r
    layout.write_row_label(ws, r, "Gross customer loans", "Crediti vs clientela (lordi)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.gross_loans_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}*(1+loan_growth)")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Allowance roll-forward (opening → +charge → −writeoff → closing).
    # Reserve the closing row symbolically; assert at the write site.
    rows["allowance_opening"] = r
    allowance_close_row = r + 3
    layout.write_row_label(ws, r, "Loan-loss allowance — opening", "Fondo svalutazione — apertura", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${allowance_close_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # (+) Impairment charge — PATCHED from P&L provisions (negative; grows the stock).
    rows["allowance_charge"] = r
    layout.write_row_label(ws, r, "(+) Impairment charge", "(+) Accantonamenti", indent=True)
    _unit(r, cur)
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=0)  # PLACEHOLDER (patched ← P&L provisions)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Patched to reference the P&L loan-loss provisions row. Provisions are "
        "negative (costs-negative), so the charge increases the magnitude of the "
        "(negative, contra-asset) allowance stock.", "ModelForge")
    r += 1

    # (−) Write-offs / utilisation (positive; reduces stock magnitude).
    rows["allowance_writeoff"] = r
    layout.write_row_label(ws, r, "(−) Write-offs / utilisation", "(−) Stralci / utilizzo", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            # writeoff reduces the (negative) stock toward zero:
            #   = -opening_allowance * writeoff_pct   (opening ≤ 0 → result ≥ 0)
            c = ws.cell(row=r, column=ci,
                        value=f"=-${col}${rows['allowance_opening']}*writeoff_pct_opening_allowance")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Allowance — closing (= opening + charge + writeoff). Drift-guarded.
    assert r == allowance_close_row, (
        f"allowance closing-row drift: expected {allowance_close_row}, got {r}")
    rows["allowance_closing"] = r
    layout.write_row_label(ws, r, "Loan-loss allowance — closing", "Fondo svalutazione — chiusura", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.allowance_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            c = ws.cell(row=r, column=ci,
                        value=(f"=${col}${rows['allowance_opening']}"
                               f"+${col}${rows['allowance_charge']}"
                               f"+${col}${rows['allowance_writeoff']}"))
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Write-offs are NET-LOAN-NEUTRAL: a written-off loan leaves the gross book
    # AND releases the allowance that covered it, simultaneously. The gross-loans
    # roll-forward above grew purely on loan_growth; reduce it by the period
    # write-off so utilisation does not fabricate net loans / total assets
    # (correct double-entry). Patched here — after the writeoff row exists — so
    # the gross-loans formula references a symbolic row, never a literal forward
    # offset. writeoff[t] depends only on the prior-period allowance, so this
    # introduces no circular reference.
    for i in range(h, n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        prior = layout.year_col(i - 1)
        cc = ws.cell(row=rows["gross_loans"], column=ci,
                     value=(f"=${prior}${rows['gross_loans']}*(1+loan_growth)"
                            f"-${col}${rows['allowance_writeoff']}"))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)

    # Net loans = gross + allowance(closing) (allowance ≤ 0)
    rows["net_loans"] = r
    layout.write_row_label(ws, r, "Net customer loans", "Crediti vs clientela (netti)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['gross_loans']}+${col}${rows['allowance_closing']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Securities — volume roll-forward
    rows["securities"] = r
    layout.write_row_label(ws, r, "Securities (liquid book)", "Titoli (portafoglio liquido)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.securities_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}*(1+securities_growth)")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Goodwill & intangibles — flat
    rows["intangibles"] = r
    layout.write_row_label(ws, r, "Goodwill & intangibles", "Avviamento e immateriali")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.intangibles_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Other assets — flat
    rows["other_assets"] = r
    layout.write_row_label(ws, r, "Other assets", "Altre attività")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.other_assets_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Non-cash assets subtotal (used to size the wholesale plug + cash plug).
    rows["noncash_assets"] = r
    layout.write_row_label(ws, r, "Non-cash assets (loans+sec+intang+other)", "Attività non liquide", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=(f"=${col}${rows['net_loans']}+${col}${rows['securities']}"
                           f"+${col}${rows['intangibles']}+${col}${rows['other_assets']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Reserve forward rows we need for the cash plug (deposits/wholesale/other_liab/
    # equity_closing/at1 live below). We will fill cash AFTER they are placed, so
    # park its row now and write it at the end.
    rows["cash"] = r
    cash_row = r
    layout.write_row_label(ws, r, "Cash & central bank (plug)", "Cassa e banca centrale (plug)")
    _unit(r, cur)
    r += 1

    # Total assets = non-cash + cash
    rows["total_assets"] = r
    layout.write_row_label(ws, r, "Total assets", "Totale attività")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['noncash_assets']}+${col}${cash_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # ── LIABILITIES & EQUITY ──────────────────────────────────────────────
    layout.write_section_header(ws, r, "Liabilities & equity", "Passività e patrimonio")
    r += 1

    # Customer deposits — volume roll-forward
    rows["deposits"] = r
    layout.write_row_label(ws, r, "Customer deposits", "Depositi clientela")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.deposits_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}*(1+deposit_growth)")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Other liabilities — flat
    rows["other_liab"] = r
    layout.write_row_label(ws, r, "Other liabilities", "Altre passività")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.other_liabilities_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # AT1 — flat
    rows["at1"] = r
    layout.write_row_label(ws, r, "Additional Tier 1 (AT1)", "Tier 1 aggiuntivo (AT1)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.at1_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${r}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Common equity roll-forward (opening → +retained → −div → −buyback → closing).
    rows["equity_opening"] = r
    equity_close_row = r + 4
    layout.write_row_label(ws, r, "Common equity — opening", "Patrimonio comune — apertura", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=0)
        else:
            prior = layout.year_col(i - 1)
            c = ws.cell(row=r, column=ci, value=f"=${prior}${equity_close_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # (+) Retained earnings to equity — PATCHED ← P&L net income attributable to CET1.
    rows["equity_retained"] = r
    layout.write_row_label(ws, r, "(+) Retained earnings", "(+) Utili a riserva", indent=True)
    _unit(r, cur)
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=0)  # PLACEHOLDER (patched ← P&L NI to CET1)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # (−) Dividends — PATCHED ← CapitalReturn (negative).
    rows["equity_dividends"] = r
    layout.write_row_label(ws, r, "(−) Dividends", "(−) Dividendi", indent=True)
    _unit(r, cur)
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=0)  # PLACEHOLDER (patched ← CapitalReturn)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # (−) Buybacks — PATCHED ← CapitalReturn (negative).
    rows["equity_buybacks"] = r
    layout.write_row_label(ws, r, "(−) Share buybacks", "(−) Riacquisti azioni", indent=True)
    _unit(r, cur)
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=ci, value=0)  # PLACEHOLDER (patched ← CapitalReturn)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Common equity — closing. Drift-guarded.
    assert r == equity_close_row, (
        f"equity closing-row drift: expected {equity_close_row}, got {r}")
    rows["equity_closing"] = r
    layout.write_row_label(ws, r, "Common equity — closing", "Patrimonio comune — chiusura")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.common_equity)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            c = ws.cell(row=r, column=ci,
                        value=(f"=${col}${rows['equity_opening']}"
                               f"+${col}${rows['equity_retained']}"
                               f"+${col}${rows['equity_dividends']}"
                               f"+${col}${rows['equity_buybacks']}"))
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 1

    # Wholesale funding — FUNDING PLUG (sized off the asset side less the other
    # L&E items). MAX(.,0) so it never goes negative.
    rows["wholesale"] = r
    layout.write_row_label(ws, r, "Wholesale funding (plug)", "Funding wholesale (plug)")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=r, column=ci, value=ob.wholesale_funding_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            c = ws.cell(row=r, column=ci,
                        value=(f"=MAX(${col}${rows['noncash_assets']}"
                               f"-${col}${rows['deposits']}-${col}${rows['other_liab']}"
                               f"-${col}${rows['equity_closing']}-${col}${rows['at1']},0)"))
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Total liabilities & equity
    rows["total_le"] = r
    layout.write_row_label(ws, r, "Total liabilities & equity", "Totale passività e patrimonio")
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=(f"=${col}${rows['deposits']}+${col}${rows['wholesale']}"
                           f"+${col}${rows['other_liab']}+${col}${rows['equity_closing']}"
                           f"+${col}${rows['at1']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # Balance check (Assets − L&E)
    rows["bs_check"] = r
    layout.write_row_label(ws, r, "Balance check (A − L − E)", "Quadratura (A − P − PN)", indent=True)
    _unit(r, cur)
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=ci,
                    value=f"=${col}${rows['total_assets']}-${col}${rows['total_le']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Now fill the CASH plug row (asset residual on the L&E side):
    #   historical → input; projection → total_le − non-cash assets (guarantees A = L&E).
    for i in range(n):
        col = layout.year_col(i); ci = ord(col) - ord("A") + 1
        if i < h:
            c = ws.cell(row=cash_row, column=ci, value=ob.cash_eur_m)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            c = ws.cell(row=cash_row, column=ci,
                        value=f"=${col}${rows['total_le']}-${col}${rows['noncash_assets']}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    out["n_periods"] = str(n)
    out["first_proj_col"] = layout.year_col(h)
    out["last_col"] = layout.year_col(n - 1)
    out["first_col"] = layout.year_col(0)
    return out
