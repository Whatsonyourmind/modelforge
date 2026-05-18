"""Covenants sheet.

For each covenant, per-period:
    - Actual value (formula: leverage = total debt / EBITDA, ICR = EBITDA / |interest|, ...)
    - Threshold (reads from Assumptions, year-indexed)
    - Headroom %
    - Breach flag (0/1)

Breach flag = 1 if covenant tripped. Feeds QC check on "zero breaches in base case".
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.formulas import ebitda_multiple, interest_coverage, headroom
from modelforge.builder.i18n import L
from modelforge.spec.unitranche import UnitrancheSpec


def build(
    ws: Worksheet,
    spec: UnitrancheSpec,
    operating_refs: dict[str, str],
    debt_refs: dict[str, str],
    operating_sheet_name: str,
    debt_sheet_name: str,
) -> dict[str, str]:
    horizon = spec.horizon
    h = horizon.historical_years
    p = horizon.projection_years
    n_years = h + p

    layout.set_column_widths(ws, label_width=42, it_width=34, year_width=12, unit_width=6)

    layout.write_title_block(
        ws, title_en="Covenants", title_it="Covenant",
        subtitle="Period-by-period headroom & breach monitor",
    )
    layout.write_scenario_banner(ws, row=3)

    # Year headers
    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        is_hist = i < h
        c = ws.cell(row=yr_row, column=col_idx,
                    value=f"{'A' if is_hist else 'E'} {yr}")
        styles.style_header(c)

    ebitda_row = int(operating_refs["ebitda_row"])
    interest_row_debt = int(debt_refs["total_interest_row"])
    closing_debt_row = int(debt_refs["total_closing_row"])

    r = 7
    breach_rows: list[int] = []

    for cov in spec.covenants:
        layout.write_section_header(ws, r, cov.name.en, cov.name.secondary)
        r += 1

        # Actual
        actual_row = r
        if cov.kind == "leverage":
            layout.write_row_label(ws, r, L("leverage_ratio").en, L("leverage_ratio").secondary)
            ws.cell(row=r, column=3, value="x").font = styles.font_label_it
            for i in range(n_years):
                col = layout.year_col(i)
                col_idx = ord(col) - ord("A") + 1
                debt_ref = f"'{debt_sheet_name}'!{col}{closing_debt_row}"
                ebitda_ref = f"'{operating_sheet_name}'!{col}{ebitda_row}"
                c = ws.cell(row=r, column=col_idx, value=ebitda_multiple(debt_ref, ebitda_ref))
                styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
        elif cov.kind == "icr":
            layout.write_row_label(ws, r, L("icr").en, L("icr").secondary)
            ws.cell(row=r, column=3, value="x").font = styles.font_label_it
            for i in range(n_years):
                col = layout.year_col(i)
                col_idx = ord(col) - ord("A") + 1
                interest_ref = f"'{debt_sheet_name}'!{col}{interest_row_debt}"
                ebitda_ref = f"'{operating_sheet_name}'!{col}{ebitda_row}"
                c = ws.cell(row=r, column=col_idx,
                            value=interest_coverage(ebitda_ref, interest_ref))
                styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
        else:
            layout.write_row_label(ws, r, cov.name.en, cov.name.secondary)
            ws.cell(row=r, column=3, value="x").font = styles.font_label_it
        r += 1

        # Threshold
        threshold_row = r
        layout.write_row_label(ws, r,
                               f"{cov.name.en} — threshold",
                               f"{cov.name.secondary} — soglia", indent=True)
        ws.cell(row=r, column=3, value="x").font = styles.font_label_it
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i < h:
                c = ws.cell(row=r, column=col_idx, value="")
                styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
            else:
                proj_idx = i - h
                if proj_idx < len(cov.threshold_by_year):
                    a = cov.threshold_by_year[proj_idx]
                    c = ws.cell(row=r, column=col_idx, value=f"={a.name}")
                    styles.style_xref(c, number_format=styles.FMT_MULTIPLE)
        r += 1

        # Headroom
        headroom_row = r
        layout.write_row_label(ws, r,
                               f"{cov.name.en} — headroom",
                               f"{cov.name.secondary} — headroom", indent=True)
        ws.cell(row=r, column=3, value="%").font = styles.font_label_it
        direction = "max" if cov.kind == "leverage" else "min"
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i < h:
                ws.cell(row=r, column=col_idx, value="")
            else:
                actual_ref = f"${col}${actual_row}"
                threshold_ref = f"${col}${threshold_row}"
                c = ws.cell(row=r, column=col_idx,
                            value=headroom(actual_ref, threshold_ref, direction=direction))
                styles.style_formula(c, number_format=styles.FMT_PCT)
        r += 1

        # Breach flag
        #
        # v0.6: covenants only apply when there is meaningful debt
        # activity. Gate on |interest| > 0 so neither the drawdown year
        # (BOP debt = 0 → zero interest → ICR = 0 → falsely below
        # threshold) nor post-maturity years (debt fully repaid → zero
        # interest) trigger artefactual breach flags.
        breach_row = r
        layout.write_row_label(ws, r,
                               f"{cov.name.en} — breach",
                               f"{cov.name.secondary} — violazione", indent=True)
        ws.cell(row=r, column=3, value="").font = styles.font_label_it
        total_interest_row = debt_refs.get("total_interest_row") if debt_refs else None
        for i in range(n_years):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i < h:
                ws.cell(row=r, column=col_idx, value="")
            else:
                actual_ref = f"${col}${actual_row}"
                threshold_ref = f"${col}${threshold_row}"
                if direction == "max":
                    core = f"{actual_ref}>{threshold_ref}"
                else:
                    core = f"{actual_ref}<{threshold_ref}"
                if total_interest_row:
                    # |interest| > 0 proxy: interest is negative by sign
                    # convention, so check it's strictly < 0 (allowing
                    # 1c tolerance for floating-point).
                    int_ref = f"'{debt_sheet_name}'!{col}{total_interest_row}"
                    formula = f"=IF(AND({core},{int_ref}<-0.01),1,0)"
                else:
                    formula = f"=IF({core},1,0)"
                c = ws.cell(row=r, column=col_idx, value=formula)
                styles.style_formula(c, number_format=styles.FMT_INTEGER)
                c.alignment = styles.align_center
        breach_rows.append(breach_row)
        r += 2

    # Aggregate breach counter
    layout.write_section_header(ws, r, "Aggregate breach counter", "Contatore violazioni")
    r += 1
    total_breach_row = r
    layout.write_row_label(ws, r, "Total breaches (projection window)",
                           "Violazioni totali (proiezione)")
    ws.cell(row=r, column=3, value="").font = styles.font_label_it
    # Sum across projection columns and breach rows
    first_proj_col = layout.year_col(h)
    last_proj_col = layout.year_col(n_years - 1)
    parts = [
        f"SUM(${first_proj_col}${br}:${last_proj_col}${br})"
        for br in breach_rows
    ]
    formula = "=" + "+".join(parts) if parts else "=0"
    # Total in col C (a scalar)
    c = ws.cell(row=r, column=3, value=formula)
    styles.style_formula(c, number_format=styles.FMT_INTEGER)
    c.font = styles.font_subheader

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {
        "total_breach_row": str(total_breach_row),
        "total_breach_cell": f"'{ws.title}'!$C${total_breach_row}",
    }
