"""Operating Model sheet.

Revenue → EBITDA → EBIT → EBT → Net Income, fully formulated.
Historical columns are hardcoded from the spec (with source comments).
Projection columns are driven by named-range reads from Assumptions.

Sign convention: costs negative.
Column layout: A=EN label, B=IT label, C=unit, D..D+N = years.

Output: returns a dict of named refs for rows that other sheets consume:
    {"revenue": "'OperatingModel'!$D$10:$J$10",
     "ebitda":  "'OperatingModel'!$D$14:$J$14", ...}
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.formulas import (
    growth, margin, negated, sum_list, ratio, add, diff,
)
from modelforge.builder.i18n import L
from modelforge.graph.schema import LinkageGraph, GraphNode, NodeKind
from modelforge.spec.unitranche import UnitrancheSpec


def build(
    ws: Worksheet,
    spec: UnitrancheSpec,
    graph: LinkageGraph,
    driver_refs: dict[str, str],
) -> dict[str, str]:
    """Emit Operating Model. Returns {line_name: range_ref_string}."""
    horizon = spec.horizon
    h = horizon.historical_years
    p = horizon.projection_years
    n_years = h + p

    layout.set_column_widths(ws, label_width=40, it_width=32, year_width=12, unit_width=6)

    # Title
    layout.write_title_block(
        ws,
        title_en="Operating Model",
        title_it="Modello operativo",
        subtitle=f"{spec.meta.currency} {spec.meta.unit_scale} · sign: {spec.meta.sign_convention}",
    )
    layout.write_scenario_banner(ws, row=3)

    # Year header row (row 5)
    year_header_row = 5
    base_fy_year = spec.target.last_fy_end.year
    ws.cell(row=year_header_row, column=3, value="Unit").font = styles.font_subheader
    for i in range(n_years):
        col = layout.year_col(i)
        yr = base_fy_year - (h - 1) + i  # historical_years back from last FY inclusive
        is_historical = i < h
        cell = ws.cell(row=year_header_row, column=ord(col) - ord("A") + 1,
                       value=f"{'A' if is_historical else 'E'} {yr}")
        styles.style_header(cell)

    # Highlight historical columns for first 6 rows
    for i in range(h):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        for row in range(year_header_row + 1, year_header_row + 40):
            c = ws.cell(row=row, column=col_idx)
            if c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = styles.fill_historical

    # ─── Line layout ───────────────────────────────────────────────────────
    # Row plan:
    #   r_growth   = revenue growth (%) — formula or xref to assumption
    #   r_revenue  = revenue line
    #   r_cogs_placeholder = blank; we model EBITDA margin directly
    #   r_ebitda_margin = EBITDA margin (%)
    #   r_ebitda   = EBITDA
    #   r_da       = D&A (negative)
    #   r_ebit     = EBIT
    #   r_interest = interest placeholder (cross-sheet link set up later by debt.py)
    #   r_ebt      = EBT
    #   r_tax      = taxes (negative)
    #   r_ni       = net income
    #   r_capex_m  = maintenance capex (negative)
    #   r_capex_g  = growth capex (negative)
    #   r_capex_t  = total capex
    #   r_nwc      = Δ NWC (negative if absorbs cash)
    #   r_fcf      = FCF to debt = EBITDA - tax on EBIT - capex - ΔNWC

    rows: dict[str, int] = {}

    r = 7  # content starts row 7

    # Subheader: Revenue build
    layout.write_section_header(ws, r, L("op_revenue_build").en, L("op_revenue_build").secondary)
    r += 1

    # Revenue growth row
    rows["revenue_growth"] = r
    layout.write_row_label(ws, r, L("revenue_growth").en, L("revenue_growth").secondary)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(h):
        # Historical growth: formula vs prior revenue (row below once revenue written).
        # We'll fill these after revenue row exists.
        pass
    for i in range(p):
        # Read from Assumptions: revenue_growth_by_year[i].name
        a = spec.operating.revenue_growth_by_year[i]
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        cell = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(cell, number_format=styles.FMT_PCT)
    r += 1

    # Revenue row
    rows["revenue"] = r
    layout.write_row_label(ws, r, L("revenue").en, L("revenue").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=spec.historical_revenue_eur_m[i])
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        # Comment — anchor to target revenue source
        from openpyxl.comments import Comment
        c.comment = Comment(
            f"Historical actual.\nSource: {spec.target.revenue_source_id}",
            "ModelForge",
        )
        c.comment.width = 220
    for i in range(p):
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        prior_col = layout.year_col(h + i - 1)
        prior_ref = f"${prior_col}${r}"
        growth_ref = f"${layout.year_col(h + i)}${rows['revenue_growth']}"
        cell = ws.cell(row=r, column=col_idx, value=growth(prior_ref, growth_ref))
        styles.style_formula(cell, number_format=styles.FMT_EUR_M)
    # Fill historical growth formulas now that revenue exists
    for i in range(1, h):
        col = layout.year_col(i)
        prior_col = layout.year_col(i - 1)
        col_idx = ord(col) - ord("A") + 1
        cur_ref = f"${col}${rows['revenue']}"
        prior_ref = f"${prior_col}${rows['revenue']}"
        growth_cell = ws.cell(row=rows["revenue_growth"], column=col_idx,
                              value=f"=IFERROR({cur_ref}/{prior_ref}-1,0)")
        styles.style_formula(growth_cell, number_format=styles.FMT_PCT)
    r += 2

    # Subheader: Margin & opex
    layout.write_section_header(ws, r, L("op_margin_opex").en, L("op_margin_opex").secondary)
    r += 1

    # EBITDA margin row
    rows["ebitda_margin"] = r
    layout.write_row_label(ws, r, L("ebitda_margin").en, L("ebitda_margin").secondary)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        col = layout.year_col(i)
        rev_ref = f"${col}${rows['revenue']}"
        # Historical EBITDA margin = historical EBITDA / historical revenue
        hist_ebitda = spec.historical_ebitda_eur_m[i]
        hist_rev = spec.historical_revenue_eur_m[i]
        val = hist_ebitda / hist_rev if hist_rev else 0
        c = ws.cell(row=r, column=col_idx, value=val)
        styles.style_input(c, number_format=styles.FMT_PCT)
    for i in range(p):
        a = spec.operating.ebitda_margin_by_year[i]
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(c, number_format=styles.FMT_PCT)
    r += 1

    # EBITDA row
    rows["ebitda"] = r
    layout.write_row_label(ws, r, L("ebitda").en, L("ebitda").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(h):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=spec.historical_ebitda_eur_m[i])
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        from openpyxl.comments import Comment
        c.comment = Comment(
            f"Historical actual.\nSource: {spec.target.ebitda_source_id}",
            "ModelForge",
        )
    for i in range(p):
        col = layout.year_col(h + i)
        col_idx = ord(col) - ord("A") + 1
        rev_ref = f"${col}${rows['revenue']}"
        mar_ref = f"${col}${rows['ebitda_margin']}"
        c = ws.cell(row=r, column=col_idx, value=margin(rev_ref, mar_ref))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 2

    # D&A row  (cost → negative)
    rows["da"] = r
    layout.write_row_label(ws, r, L("da").en, L("da").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        rev_ref = f"${col}${rows['revenue']}"
        c = ws.cell(row=r, column=col_idx, value=f"=-{rev_ref}*da_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # EBIT row
    rows["ebit"] = r
    layout.write_row_label(ws, r, L("ebit").en, L("ebit").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        ebitda_ref = f"${col}${rows['ebitda']}"
        da_ref = f"${col}${rows['da']}"
        c = ws.cell(row=r, column=col_idx, value=add(ebitda_ref, da_ref))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 2

    # Interest row (placeholder — filled by DebtSchedule via cross-sheet ref)
    rows["interest"] = r
    layout.write_row_label(ws, r, L("interest_expense").en, L("interest_expense").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        # Reference DebtSchedule Interest row (row number will be known after debt sheet).
        # We place a placeholder; debt.py will patch the actual sheet name + row.
        c = ws.cell(row=r, column=col_idx, value=0)
        styles.style_xref(c, number_format=styles.FMT_EUR_M)
    r += 1

    # EBT
    rows["ebt"] = r
    layout.write_row_label(ws, r, L("ebt").en, L("ebt").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(
            row=r,
            column=col_idx,
            value=add(f"${col}${rows['ebit']}", f"${col}${rows['interest']}"),
        )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Tax
    rows["tax"] = r
    layout.write_row_label(ws, r, L("tax").en, L("tax").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        ebt_ref = f"${col}${rows['ebt']}"
        # Only tax positive EBT
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-MAX({ebt_ref},0)*effective_tax_rate")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Net income
    rows["net_income"] = r
    layout.write_row_label(ws, r, L("net_income").en, L("net_income").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(
            row=r,
            column=col_idx,
            value=add(f"${col}${rows['ebt']}", f"${col}${rows['tax']}"),
        )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # ── Cash flow block ───────────────────────────────────────────────────
    layout.write_section_header(ws, r, L("op_fcf_to_debt").en, L("op_fcf_to_debt").secondary)
    r += 1

    # Maintenance capex
    rows["capex_m"] = r
    layout.write_row_label(ws, r, L("capex_maint").en, L("capex_maint").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        rev_ref = f"${col}${rows['revenue']}"
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-{rev_ref}*maintenance_capex_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Growth capex
    rows["capex_g"] = r
    layout.write_row_label(ws, r, L("capex_growth").en, L("capex_growth").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        rev_ref = f"${col}${rows['revenue']}"
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-{rev_ref}*growth_capex_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Total capex
    rows["capex_total"] = r
    layout.write_row_label(ws, r, L("capex_total").en, L("capex_total").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(
            row=r, column=col_idx,
            value=add(f"${col}${rows['capex_m']}", f"${col}${rows['capex_g']}"),
        )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Δ NWC
    rows["nwc"] = r
    layout.write_row_label(ws, r, L("nwc_change").en, L("nwc_change").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        if i == 0:
            c = ws.cell(row=r, column=col_idx, value=0)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior_col = layout.year_col(i - 1)
            rev_cur = f"${col}${rows['revenue']}"
            rev_prior = f"${prior_col}${rows['revenue']}"
            c = ws.cell(
                row=r, column=col_idx,
                value=f"=-({rev_cur}-{rev_prior})*nwc_pct_revenue_delta",
            )
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # FCF to debt
    #
    # v0.6: includes cash interest. Previously FCF summed only
    # EBITDA + tax + capex + ΔNWC, which overstated the base to which
    # the cash sweep was applied (sweep_pct of FCF). Levered FCF
    # available for principal paydown must deduct interest paid:
    #   FCF = EBITDA - tax - capex - ΔNWC - |interest|
    # Since all cost components are already negative, we just add them.
    rows["fcf"] = r
    layout.write_row_label(ws, r, L("fcf_to_debt").en, L("fcf_to_debt").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n_years):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        parts = [
            f"${col}${rows['ebitda']}",
            f"${col}${rows['tax']}",
            f"${col}${rows['capex_total']}",
            f"${col}${rows['nwc']}",
            f"${col}${rows['interest']}",   # v0.6: post-interest FCF
        ]
        c = ws.cell(row=r, column=col_idx, value=sum_list(parts))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    # Freeze panes & print
    ws.freeze_panes = "D7"
    ws.print_title_rows = f"5:5"
    ws.print_title_cols = "A:C"

    # Graph — add cells for every formula cell emitted
    for row_name, row_num in rows.items():
        for i in range(n_years):
            col = layout.year_col(i)
            cell_id = LinkageGraph.cell_id(ws.title, f"{col}{row_num}")
            graph.add_node(GraphNode(id=cell_id, kind=NodeKind.CELL, label=row_name))

    # Return range refs for other sheets
    out: dict[str, str] = {}
    first_col = layout.year_col(0)
    last_col = layout.year_col(n_years - 1)
    first_proj_col = layout.year_col(h)
    last_proj_col = layout.year_col(n_years - 1)
    for name, row_num in rows.items():
        out[name] = f"'{ws.title}'!${first_col}${row_num}:${last_col}${row_num}"
        out[f"{name}_proj"] = f"'{ws.title}'!${first_proj_col}${row_num}:${last_proj_col}${row_num}"
        out[f"{name}_row"] = str(row_num)

    return out
