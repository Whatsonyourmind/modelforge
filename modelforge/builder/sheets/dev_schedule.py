"""Ground-up development RE schedule sheet (annual phased layout).

LAYOUT CHOICE — ANNUAL PHASED (not monthly grid). The development-RE
engine's economics are defined in MONTHS (phased capex, S-curve lease-up,
construction-interest capitalisation, forward-NOI exit, pro-rata loan-to-cost).
A full ~48-column monthly grid with month-level IRR/XIRR risks the certify
recalc engine and build time, so this sheet renders an ANNUAL phased schedule
(construction years → lease-up year → stabilised years → exit) that captures
the SAME economics:

  * Capex is phased by MONTH per the engine convention, then aggregated into
    annual buckets; each annual bucket is emitted as a LIVE Excel formula over
    the capex named ranges (the per-year phasing FRACTIONS are computed in
    Python — deterministic, no clock/RNG — and appear as literals inside the
    formula, exactly the project-finance idiom ``=-total_capex*phasing_y1``).
  * Lease-up occupancy per year = the average of the monthly S-curve over that
    year's operating months, floored at operator_floor_occ (PBSA). Fractions
    computed in Python, emitted as live formulas against the revenue inputs.
  * Construction-phase senior interest CAPITALISES into the loan balance via a
    live opening→interest→capitalised→closing roll-forward; post-delivery
    interest is paid from positive NOI (capped) with the unpaid part
    capitalising; the full balance is repaid at exit.
  * Exit value = forward stabilised NOI / exit_cap_rate, less selling costs,
    booked in the exit year column.

Every OUTPUT cell on this sheet is a LIVE formula (no precomputed numeric
results) so the workbook recalculates end-to-end. The Returns and Waterfall
sheets read this sheet's unlevered CF, equity CF, debt balance and exit rows.

The sheet feeds:
    dev_returns.py     — unlevered/levered IRR, MOIC, NPV, peak debt, exit value
    (waterfall is emitted inline by the template orchestrator using EquityWaterfall)
"""

from __future__ import annotations

import math

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def _monthly_capex_phasing(spec) -> list[float]:
    """Per-month phasing FRACTIONS of total_dev_cost (engine convention).

    Returns a list indexed by month 0..hold_total_months-1 of the fraction of
    TOTAL development cost spent that month. Deterministic (no clock/RNG).

    Convention (per the dev-RE engine):
      acquisition          → month 0
      soft + 30% conting.  → spread over (permit+3) months from t0
      hard + 50% conting. + other
                           → spread over construction_months from month=permit
      ffe  + 20% conting.  → spread over 3 months from max(delivery-3, 0)
    """
    tl = spec.timeline
    cx = spec.capex
    months = tl.hold_total_months
    acq = cx.acquisition_eur_m.base
    hard = cx.hard_costs_eur_m.base
    soft = cx.soft_costs_eur_m.base
    ffe = cx.ffe_eur_m.base
    other = cx.other_dev_charges_eur_m.base
    contingency = (hard + soft + ffe) * cx.contingency_pct.base
    total = acq + hard + soft + ffe + other + contingency
    if total <= 0:
        return [0.0] * months

    spend = [0.0] * months

    def _spread(amount: float, start: int, span: int) -> None:
        span = max(int(span), 1)
        per = amount / span
        for m in range(start, start + span):
            if 0 <= m < months:
                spend[m] += per

    # acquisition at t0
    if 0 < months:
        spend[0] += acq
    # soft + 30% contingency over (permit+3) months from t0
    _spread(soft + 0.30 * contingency, 0, tl.permit_months + 3)
    # hard + 50% contingency + other over construction_months from month=permit
    _spread(hard + 0.50 * contingency + other, tl.permit_months,
            tl.construction_months)
    # ffe + 20% contingency over 3 months from max(delivery-3, 0)
    _spread(ffe + 0.20 * contingency, max(tl.delivery_month - 3, 0), 3)

    return [s / total for s in spend]


def _annualise(monthly: list[float], n_years: int) -> list[float]:
    """Sum a per-month series into annual buckets (year i = months 12i..12i+11)."""
    out = [0.0] * n_years
    for m, v in enumerate(monthly):
        y = m // 12
        if 0 <= y < n_years:
            out[y] += v
    return out


def _annual_occupancy(spec, n_years: int) -> list[float]:
    """Per-year stabilised-occupancy fraction (of the 95% gross potential).

    Occupancy ramps from delivery via the engine's S-curve:
        occ(i) = 0.95 / (1 + exp(-8*(i/leaseup - 0.5)))  for i in 0..leaseup
    averaged over each year's operating months from delivery, then divided by
    0.95 to express it as a FRACTION OF the 95%-occupancy stabilised gross
    (so year-fraction × stabilised_gross is the year's realised gross). For
    PBSA the realised occupancy is floored at operator_floor_occ during
    lease-up; expressed here as a fraction of 0.95.
    """
    tl = spec.timeline
    leaseup_m = tl.leaseup_months
    delivery = tl.delivery_month
    months = tl.hold_total_months

    floor_frac = 0.0
    if spec.revenue.kind == "pbsa" and spec.revenue.operator_floor_occ is not None:
        floor_frac = min(spec.revenue.operator_floor_occ.base / 0.95, 1.0)

    # monthly occupancy fraction-of-stabilised from delivery
    per_month_frac = [0.0] * months
    for m in range(months):
        rel = m - delivery
        if rel < 0:
            per_month_frac[m] = 0.0
            continue
        i = rel  # months since delivery
        # S-curve value (of 0.95) → fraction of stabilised
        x = i / leaseup_m if leaseup_m > 0 else 1.0
        scurve = 0.95 / (1.0 + math.exp(-8.0 * (x - 0.5)))
        frac = scurve / 0.95
        if i <= leaseup_m:
            frac = max(frac, floor_frac)
        else:
            frac = 1.0  # stabilised after lease-up completes
        per_month_frac[m] = min(frac, 1.0)

    # average over each year's months that are post-delivery
    out = [0.0] * n_years
    for y in range(n_years):
        vals = []
        for mm in range(12):
            m = y * 12 + mm
            if m >= months:
                break
            if m >= delivery:
                vals.append(per_month_frac[m])
        out[y] = (sum(vals) / len(vals)) if vals else 0.0
    return out


def build(ws: Worksheet, spec) -> dict[str, str]:
    tl = spec.timeline
    # Annual horizon: year columns t=0..t=H where H = exit year.
    exit_year = math.ceil(tl.hold_total_months / 12)
    delivery_year = tl.delivery_month // 12  # year index occupancy begins
    n = exit_year + 1  # t=0..t=H inclusive

    monthly_phasing = _monthly_capex_phasing(spec)
    annual_phasing = _annualise(monthly_phasing, n)
    annual_occ = _annual_occupancy(spec, n)
    # Grant disbursement: 50% at hard-start month (= permit), 50% at delivery.
    grant_year_hard = tl.permit_months // 12
    grant_year_delivery = tl.delivery_month // 12

    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=10, unit_width=6)
    # Ensure enough year columns exist (default helper only widths 10).
    for i in range(n):
        ws.column_dimensions[layout.year_col(i)].width = 10
    layout.write_title_block(
        ws, title_en="Development Schedule (annual phased)",
        title_it="Piano di sviluppo (annuale, per fasi)",
        subtitle=(f"{spec.meta.currency} {spec.meta.unit_scale} · "
                  f"permit {tl.permit_months}m + construction {tl.construction_months}m "
                  f"→ delivery m{tl.delivery_month}; exit m{tl.hold_total_months}"),
    )
    layout.write_scenario_banner(ws, row=3)

    # Year headers t=0..t=H, with a phase tag
    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if i < delivery_year:
            tag = "C"  # construction
        elif i == delivery_year:
            tag = "D"  # delivery / lease-up
        elif i == exit_year:
            tag = "X"  # exit
        else:
            tag = "O"  # operating / stabilised
        cc = ws.cell(row=yr_row, column=col_idx, value=f"t={i} {tag}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    cx = spec.capex
    acq = cx.acquisition_eur_m.name
    hard = cx.hard_costs_eur_m.name
    soft = cx.soft_costs_eur_m.name
    ffe = cx.ffe_eur_m.name
    other = cx.other_dev_charges_eur_m.name
    cont_pct = cx.contingency_pct.name

    # ── SECTION 1: Development cost ────────────────────────────────────────
    layout.write_section_header(ws, r, "Development cost build", "Costruzione costi di sviluppo")
    r += 1

    rows["contingency"] = r
    layout.write_row_label(ws, r, "Contingency (= (hard+soft+ffe) × %)",
                           "Imprevisti", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=({hard}+{soft}+{ffe})*{cont_pct}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_dev_cost"] = r
    layout.write_row_label(ws, r, "Total development cost (TDC)",
                           "Costo totale di sviluppo")
    cc = ws.cell(row=r, column=4,
                 value=f"={acq}+{hard}+{soft}+{ffe}+{other}+$D${rows['contingency']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Phased dev spend per year (outflow, negative) — phasing fractions baked
    # as deterministic literals (project-finance idiom), multiplied by the live
    # TDC named-range total so scenario worst/best of costs still scale.
    rows["dev_spend"] = r
    layout.write_row_label(ws, r, "Development spend (phased, outflow)",
                           "Spesa di sviluppo (fasata)")
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    tdc_ref = f"$D${rows['total_dev_cost']}"
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        frac = annual_phasing[i]
        if frac > 1e-12:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-{tdc_ref}*{frac:.8f}")
        else:
            cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Per-year development spend = -TDC × (annual phasing fraction). The "
        "phasing fractions are aggregated from the month-level engine "
        "convention (acquisition at t0; soft+30% contingency over permit+3 "
        "months; hard+50% contingency+other over construction; ffe+20% "
        "contingency over the final 3 construction months). TDC is the live "
        "named-range total above, so scenario cost shifts scale the schedule.",
        "ModelForge",
    )
    r += 1

    # Grant inflow (50% at hard-start year, 50% at delivery year)
    grant_assum = spec.capital.public_grant_amount
    rows["grant_in"] = r
    layout.write_row_label(ws, r, f"{spec.capital.grant_name} (inflow)",
                           "Contributo pubblico (entrata)", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if grant_assum is not None:
            gn = grant_assum.name
            share = 0.0
            if i == grant_year_hard:
                share += 0.5
            if i == grant_year_delivery:
                share += 0.5
            if share > 1e-12:
                cc = ws.cell(row=r, column=col_idx, value=f"={gn}*{share:.4f}")
            else:
                cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Net development need = max(0, -dev_spend - grant_in) (positive = to fund)
    rows["net_need"] = r
    layout.write_row_label(ws, r, "Net development need (to fund)",
                           "Fabbisogno netto di sviluppo", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=(f"=MAX(0,-${col}${rows['dev_spend']}"
                            f"-${col}${rows['grant_in']})"))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Grant APPLIED against development spend per period = MIN(grant_in, -dev_spend).
    # A grant tranche disbursed in a period with no remaining spend (e.g. the
    # 50% delivery-month tranche when all spend completes before delivery) is a
    # surplus that flows to equity as a distribution — it does NOT fund the TDC.
    # Only the applied portion belongs in the sources=uses identity; the full
    # grant inflow is still captured in the unlevered/equity cashflows below.
    rows["grant_applied"] = r
    layout.write_row_label(ws, r, "Grant applied against spend (= MIN(grant, spend))",
                           "Contributo applicato alla spesa", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=(f"=MIN(${col}${rows['grant_in']},"
                            f"-${col}${rows['dev_spend']})"))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # ── SECTION 2: Stabilised NOI build ────────────────────────────────────
    layout.write_section_header(ws, r, "Stabilised NOI & lease-up", "NOI stabilizzato & locazione")
    r += 1

    rev = spec.revenue
    # Stabilised gross potential at 95% occ (one cell, t0 column D)
    rows["stab_gross"] = r
    layout.write_row_label(ws, r, "Stabilised gross income (95% occ)",
                           "Reddito lordo stabilizzato (95%)")
    if rev.kind == "pbsa":
        # beds × rent_per_bed_year × 0.95 (effective occ at stabilisation)
        stab_formula = f"={rev.beds.name}*{rev.rent_per_bed_year.name}*0.95/1000000"
        opex_formula = f"={rev.opex_per_unit_year.name}*{rev.beds.name}/1000000"
    else:
        # lettable_sqm × rent_sqm_year × (1 − vacancy)
        stab_formula = (f"={rev.lettable_sqm.name}*{rev.rent_sqm_year.name}"
                        f"*(1-{rev.vacancy_pct.name})/1000000")
        opex_formula = f"={rev.opex_per_unit_year.name}*{rev.lettable_sqm.name}/1000000"
    cc = ws.cell(row=r, column=4, value=stab_formula)
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.comment = Comment(
        "Stabilised gross income at the engine's 95% effective occupancy. "
        "PBSA: beds × rent_per_bed_year × 0.95. Generic: lettable_sqm × "
        "rent_sqm_year × (1 − vacancy). Divided by 1e6 to render in EUR m.",
        "ModelForge",
    )
    r += 1

    rows["stab_opex"] = r
    layout.write_row_label(ws, r, "Stabilised landlord opex", "Opex proprietario stabilizzato",
                           indent=True)
    cc = ws.cell(row=r, column=4, value=opex_formula)
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["stab_noi"] = r
    layout.write_row_label(ws, r, "Stabilised annual NOI", "NOI annuo stabilizzato")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['stab_gross']}-$D${rows['stab_opex']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Per-year realised NOI = stabilised NOI × annual occupancy fraction.
    # (Occupancy fraction baked as deterministic literal of the S-curve.)
    rows["noi"] = r
    layout.write_row_label(ws, r, "Realised NOI (lease-up S-curve)",
                           "NOI realizzato (curva S)")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        frac = annual_occ[i]
        if i == exit_year:
            # No operating NOI booked in the exit column (exit proceeds only);
            # the engine sells at start of exit year on a forward NOI.
            cc = ws.cell(row=r, column=col_idx, value=0)
        elif frac > 1e-12:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=$D${rows['stab_noi']}*{frac:.6f}")
        else:
            cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Realised NOI per year = stabilised NOI × annual occupancy fraction. "
        "The occupancy fraction is the average of the monthly lease-up S-curve "
        "occ(i)=0.95/(1+exp(-8*(i/leaseup-0.5))) over each year's post-delivery "
        "months, expressed as a fraction of the 95% stabilised level and "
        "floored at the operator occupancy floor (PBSA).",
        "ModelForge",
    )
    r += 2

    # ── SECTION 3: Exit ────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "Exit (forward-NOI cap)", "Uscita (cap su NOI forward)")
    r += 1

    # Forward NOI at exit = stabilised NOI × (1+rev_growth)^(years delivery→exit)
    years_to_exit = max(exit_year - delivery_year, 0)
    rows["fwd_noi"] = r
    layout.write_row_label(ws, r, "Forward NOI at exit", "NOI forward a uscita")
    cc = ws.cell(row=r, column=4,
                 value=(f"=$D${rows['stab_noi']}"
                        f"*(1+{rev.rev_growth_pct.name})^{years_to_exit}"))
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["exit_value"] = r
    layout.write_row_label(ws, r, "Gross exit value (fwd NOI / cap)",
                           "Valore uscita lordo")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['fwd_noi']}/{spec.exit.exit_cap_rate.name}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["selling_costs"] = r
    layout.write_row_label(ws, r, "Selling costs", "Costi di vendita", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=-$D${rows['exit_value']}*{spec.exit.selling_costs_pct.name}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_exit"] = r
    layout.write_row_label(ws, r, "Net exit proceeds", "Proventi netti di uscita")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['exit_value']}+$D${rows['selling_costs']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    cc.border = styles.BORDER_TOP_THIN
    r += 2

    # ── SECTION 4: Unlevered project cash flow ─────────────────────────────
    layout.write_section_header(ws, r, "Unlevered project cash flow", "CF progetto (unlevered)")
    r += 1
    rows["unlevered_cf"] = r
    layout.write_row_label(ws, r, "Unlevered CF (dev spend + grant + NOI + exit)",
                           "CF unlevered")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        base = (f"=${col}${rows['dev_spend']}+${col}${rows['grant_in']}"
                f"+${col}${rows['noi']}")
        if i == exit_year:
            base += f"+$D${rows['net_exit']}"
        cc = ws.cell(row=r, column=col_idx, value=base)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # ── SECTION 5: Senior debt roll-forward (pro-rata LTC, IDC capitalised) ──
    layout.write_section_header(ws, r, "Senior debt (pro-rata LTC, IDC capitalised)",
                                "Debito senior (LTC pro-rata, IDC capitalizzato)")
    r += 1

    eq_pct = spec.capital.equity_pct.name
    rate = spec.capital.senior_rate_all_in.name

    # Debt drawdown each period = (1 − equity_pct) × net need
    rows["debt_draw"] = r
    layout.write_row_label(ws, r, "Senior drawdown (= (1−equity%) × net need)",
                           "Tiraggio senior", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i >= delivery_year:
            # No new development draws after delivery in this annual model.
            if i <= delivery_year:
                cc = ws.cell(row=r, column=col_idx,
                             value=f"=(1-{eq_pct})*${col}${rows['net_need']}")
            else:
                cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=(1-{eq_pct})*${col}${rows['net_need']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Opening balance.
    #
    # The senior-debt block emits five rows in this fixed order:
    #   debt_open (r) → debt_interest (r+1) → interest_cap (r+2)
    #   → debt_repay (r+3) → debt_close (r+4)
    # The opening balance for year i>=1 MUST roll forward from the prior year's
    # CLOSING balance (debt_close, r+4), NOT the principal-repayment row (r+3).
    # We capture the closing-balance row symbolically (rows["debt_close"]) so
    # the reference can never silently drift if rows are inserted/reordered.
    rows["debt_open"] = r
    debt_close_row = r + 4  # debt_close is the 5th row in this fixed block
    rows["debt_close"] = debt_close_row
    layout.write_row_label(ws, r, "Opening debt balance", "Debito iniziale", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            prior = layout.year_col(i - 1)
            # Opening(i) = prior-year CLOSING balance (debt_close row).
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${prior}${rows['debt_close']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Interest on opening balance
    rows["debt_interest"] = r
    layout.write_row_label(ws, r, "Interest on opening balance", "Interessi", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['debt_open']}*{rate}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Interest capitalised: during construction (before delivery) ALL interest
    # capitalises; post-delivery, the portion unpaid from NOI capitalises.
    rows["interest_cap"] = r
    layout.write_row_label(ws, r, "Interest capitalised (IDC / unpaid)",
                           "Interessi capitalizzati", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        int_ref = f"${col}${rows['debt_interest']}"
        noi_ref = f"${col}${rows['noi']}"
        if i < delivery_year:
            # Pure construction year: all interest capitalises
            cc = ws.cell(row=r, column=col_idx, value=f"={int_ref}")
        else:
            # Post-delivery: capitalise only interest not covered by positive NOI
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=MAX(0,{int_ref}-MAX({noi_ref},0))")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Repayment at exit (full closing-balance bullet at exit year)
    rows["debt_repay"] = r
    layout.write_row_label(ws, r, "Principal repayment (at exit)",
                           "Rimborso capitale (a uscita)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == exit_year:
            # Repay opening + draw + capitalised interest of the exit year
            cc = ws.cell(row=r, column=col_idx,
                         value=(f"=-(${col}${rows['debt_open']}"
                                f"+${col}${rows['debt_draw']}"
                                f"+${col}${rows['interest_cap']})"))
        else:
            cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Closing balance = opening + draw + capitalised interest + repayment.
    # Row position must match the symbolic debt_close_row reserved above so the
    # opening-balance roll-forward references this exact row.
    assert r == debt_close_row, (
        f"debt_close row drift: expected {debt_close_row}, got {r}"
    )
    rows["debt_close"] = r
    layout.write_row_label(ws, r, "Closing debt balance", "Debito finale")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=(f"=${col}${rows['debt_open']}+${col}${rows['debt_draw']}"
                            f"+${col}${rows['interest_cap']}+${col}${rows['debt_repay']}"))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 1

    # Cash interest paid (post-delivery, from NOI; capped at positive NOI)
    rows["interest_paid"] = r
    layout.write_row_label(ws, r, "Cash interest paid (from NOI)",
                           "Interessi pagati in cassa", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        int_ref = f"${col}${rows['debt_interest']}"
        noi_ref = f"${col}${rows['noi']}"
        if i < delivery_year:
            cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-MIN({int_ref},MAX({noi_ref},0))")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # ── SECTION 6: Equity cash flow ────────────────────────────────────────
    layout.write_section_header(ws, r, "Equity cash flow", "CF equity")
    r += 1

    # Arrangement fee on the debt limit (= total senior drawn), paid by equity
    # at t0. Approximated as fee% × total drawdowns (sum of the draw row).
    first_col = layout.year_col(0); last_col = layout.year_col(n - 1)
    rows["arrangement_fee"] = r
    layout.write_row_label(ws, r, "Arrangement fee (on debt limit, t0)",
                           "Commissione organizzazione (t0)", indent=True)
    arr = spec.capital.arrangement_fee_pct.name
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=(f"=-{arr}*SUM(${first_col}${rows['debt_draw']}"
                                f":${last_col}${rows['debt_draw']})"))
        else:
            cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Equity cash flow:
    #   = equity-funded dev need (negative)
    #     + NOI cash retained after cash interest
    #     + net exit proceeds − debt repayment (at exit)
    #     − arrangement fee (t0)
    # Equity funds (equity_pct × net need); = -(equity_pct × net need)
    rows["equity_cf"] = r
    layout.write_row_label(ws, r, "Equity cash flow", "CF equity")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        parts = [
            f"-({eq_pct})*${col}${rows['net_need']}",  # equity contribution to dev
            f"${col}${rows['noi']}",                    # operating NOI
            f"${col}${rows['interest_paid']}",          # cash interest (negative)
            f"${col}${rows['arrangement_fee']}",        # fee (negative, t0)
        ]
        if i == exit_year:
            parts.append(f"$D${rows['net_exit']}")
            parts.append(f"${col}${rows['debt_repay']}")  # repayment (negative)
        cc = ws.cell(row=r, column=col_idx, value="=" + "+".join(parts))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # ── SECTION 7: Per-sheet QC checks ─────────────────────────────────────
    layout.write_section_header(ws, r, "Schedule QC", "Controlli piano")
    r += 1

    # Sources = Uses: total equity + total debt drawn + grant = TDC
    rows["qc_sources_uses"] = r
    layout.write_row_label(ws, r, "QC: sources = uses (equity+debt+grant = TDC)",
                           "QC: fonti = impieghi", indent=True)
    # Use grant APPLIED (capped at spend per period), not gross grant inflow:
    # a grant tranche landing in a no-spend period is surplus to equity and does
    # not fund the TDC. equity_dev + debt_total + grant_applied = Σ net need +
    # grant_applied = Σ (-dev_spend) = TDC.
    grant_total = (f"SUM(${first_col}${rows['grant_applied']}"
                   f":${last_col}${rows['grant_applied']})")
    # equity portion = eq_pct × Σ net need. eq_pct is a SINGLE-cell named range,
    # so SUMPRODUCT(scalar, array) is dimension-mismatched and returns #VALUE!
    # in the recalc engine — use the equivalent scalar × SUM(array) instead.
    equity_dev = (f"{eq_pct}*SUM(${first_col}${rows['net_need']}"
                  f":${last_col}${rows['net_need']})")
    debt_total = (f"SUM(${first_col}${rows['debt_draw']}:${last_col}${rows['debt_draw']})")
    # equity_dev + debt_total + grant_applied = Σ net need + grant_applied = TDC.
    # Check |equity_dev + debt_total + grant_applied − TDC| < 0.01.
    cc = ws.cell(
        row=r, column=4,
        value=(f"=IF(ABS({equity_dev}+{debt_total}+{grant_total}"
               f"-$D${rows['total_dev_cost']})<0.01,1,0)"),
    )
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.alignment = styles.align_center
    r += 1

    # Equity CF at t0 negative
    rows["qc_t0_negative"] = r
    layout.write_row_label(ws, r, "QC: equity CF at t=0 is negative",
                           "QC: CF equity a t=0 negativo", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IF(${first_col}${rows['equity_cf']}<0,1,0)")
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.alignment = styles.align_center
    r += 1

    # Equity CF at exit positive
    rows["qc_exit_positive"] = r
    layout.write_row_label(ws, r, "QC: equity CF at exit is positive",
                           "QC: CF equity a uscita positivo", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IF(${last_col}${rows['equity_cf']}>0,1,0)")
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.alignment = styles.align_center
    r += 1

    # Debt repaid to 0 at exit
    rows["qc_debt_zero"] = r
    layout.write_row_label(ws, r, "QC: closing debt = 0 at exit",
                           "QC: debito = 0 a uscita", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IF(ABS(${last_col}${rows['debt_close']})<0.01,1,0)")
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.alignment = styles.align_center
    r += 1

    # ── DEBT-CONSERVATION INVARIANT (catches a broken roll-forward) ──────────
    # An external reviewer cannot game this: the senior facility must be a
    # closed system. Everything that was advanced to the borrower — cumulative
    # cash drawdowns PLUS cumulative interest capitalised into the balance —
    # must equal the cumulative principal repaid (in absolute terms). If the
    # opening-balance roll-forward is broken (e.g. it points at the wrong row so
    # interest never compounds and the balance is never carried to exit), this
    # identity breaks by construction. This is the check that auto-catches the
    # historical P0 (opening referenced the repayment row, not the closing row).
    cum_draw = f"SUM(${first_col}${rows['debt_draw']}:${last_col}${rows['debt_draw']})"
    cum_idc = f"SUM(${first_col}${rows['interest_cap']}:${last_col}${rows['interest_cap']})"
    cum_repay = f"SUM(${first_col}${rows['debt_repay']}:${last_col}${rows['debt_repay']})"
    rows["qc_debt_conservation"] = r
    layout.write_row_label(
        ws, r, "QC: senior debt conserved (Σdraws+ΣIDC = Σprincipal repaid)",
        "QC: debito senior conservato", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"=IF(ABS(({cum_draw}+{cum_idc})+{cum_repay})<0.01,1,0)"),
    )
    cc.comment = Comment(
        "Closed-system conservation of the senior facility: cumulative cash "
        "drawdowns + cumulative interest capitalised (IDC) must exactly equal "
        "the cumulative principal repaid (repayments carry a negative sign, so "
        "the sum nets to zero within tolerance). A broken opening-balance "
        "roll-forward — interest not compounding, or the closing balance not "
        "carried to exit — violates this identity and fails the check.",
        "ModelForge",
    )
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.alignment = styles.align_center
    r += 1

    # IDC actually capitalised during construction (the sheet is literally
    # titled "IDC capitalised"; this proves the claim is not hollow).
    rows["qc_idc_positive"] = r
    layout.write_row_label(
        ws, r, "QC: construction interest capitalised (IDC > 0)",
        "QC: interessi di costruzione capitalizzati", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=f"=IF({cum_idc}>0.001,1,0)",
    )
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.alignment = styles.align_center
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    out["n_periods"] = str(n)
    out["exit_year"] = str(exit_year)
    out["first_col"] = first_col
    out["last_col"] = last_col
    return out
