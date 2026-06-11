"""Generic QC sheet — reusable in-workbook check emitter.

Each template can pass a list of (name_en, name_it, formula) checks.
Builds a simple single-column pass/fail sheet with ALL_PASS aggregator.
"""

from __future__ import annotations

from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout

# Absolute QC tolerance, expressed in EUR. €10,000 == 0.01 at unit_scale
# "millions" (the historical hardcoded default). The same absolute amount
# scales to different *workbook-unit* magnitudes depending on meta.unit_scale.
_QC_ABS_TOLERANCE_EUR = 10_000.0

# Workbook-units per EUR (i.e. divide an absolute-EUR amount by this to get
# the value as it appears in the rendered workbook at the given scale).
_UNIT_SCALE_FACTOR = {"actual": 1.0, "thousands": 1_000.0, "millions": 1_000_000.0}


def qc_tolerance(spec, abs_eur: float = _QC_ABS_TOLERANCE_EUR) -> float:
    """QC balance/funding tolerance expressed in *workbook units*.

    The workbook renders values in ``spec.meta.unit_scale`` ("millions" by
    default). A fixed absolute tolerance of ``abs_eur`` EUR therefore maps to
    ``abs_eur / factor`` workbook units, so a materially-unbalanced model is
    still caught at unit_scale "thousands"/"actual" (where the old hardcoded
    0.01 was ~1000x / 1e6x too loose). At "millions" this returns 0.01,
    preserving the historical behaviour byte-for-byte.
    """
    unit_scale = getattr(getattr(spec, "meta", None), "unit_scale", "millions")
    factor = _UNIT_SCALE_FACTOR.get(unit_scale, 1_000_000.0)
    return abs_eur / factor


def fmt_tol(spec, abs_eur: float = _QC_ABS_TOLERANCE_EUR) -> str:
    """``qc_tolerance`` rendered as a compact formula literal (no float noise).

    Produces e.g. "0.01" (millions), "10" (thousands), "10000" (actual).
    """
    tol = qc_tolerance(spec, abs_eur)
    if tol == int(tol):
        return str(int(tol))
    # millions case: keep the historical literal exactly.
    return repr(tol)


def build(ws: Worksheet, checks: list[tuple[str, str, str]]) -> dict[str, str]:
    layout.set_column_widths(ws, label_width=60, it_width=44, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Quality Checks", title_it="Controlli qualità",
        subtitle="All checks must equal 1 before this model leaves the building.",
    )

    ws.cell(row=4, column=1, value="ALL CHECKS PASS").font = styles.font_header
    ws.cell(row=4, column=1).fill = styles.fill_header
    ws.cell(row=4, column=2, value="Tutti i controlli OK").font = styles.font_label_it
    all_pass_cell = ws.cell(row=4, column=3, value=0)
    styles.style_check(all_pass_cell, is_ok=False)

    ws.cell(row=6, column=1, value="Check").font = styles.font_subheader
    ws.cell(row=6, column=2, value="Check (IT)").font = styles.font_subheader
    ws.cell(row=6, column=3, value="Pass").font = styles.font_subheader

    r = 7
    check_refs: list[str] = []
    for en, it, formula in checks:
        layout.write_row_label(ws, r, en, it)
        c = ws.cell(row=r, column=3, value=formula)
        styles.style_formula(c, number_format=styles.FMT_INTEGER)
        c.alignment = styles.align_center
        check_refs.append(f"$C${r}")
        r += 1

    if check_refs:
        all_pass_cell.value = (
            "=IF(SUM(" + ",".join(check_refs) + f")={len(check_refs)},1,0)"
        )

    # Conditional colours
    ws.conditional_formatting.add(
        "C4",
        CellIsRule(operator="equal", formula=["1"],
                   fill=PatternFill("solid", fgColor=styles.COLOR_CHECK_OK),
                   font=Font(name=styles.FONT_BASE, size=styles.FONT_SIZE_HEADER, bold=True, color="006100")),
    )
    ws.conditional_formatting.add(
        "C4",
        CellIsRule(operator="equal", formula=["0"],
                   fill=PatternFill("solid", fgColor=styles.COLOR_CHECK_BAD),
                   font=Font(name=styles.FONT_BASE, size=styles.FONT_SIZE_HEADER, bold=True, color="9C0006")),
    )
    ws.conditional_formatting.add(
        f"C7:C{r}",
        CellIsRule(operator="equal", formula=["1"],
                   fill=PatternFill("solid", fgColor=styles.COLOR_CHECK_OK)),
    )
    ws.conditional_formatting.add(
        f"C7:C{r}",
        CellIsRule(operator="equal", formula=["0"],
                   fill=PatternFill("solid", fgColor=styles.COLOR_CHECK_BAD)),
    )

    ws.freeze_panes = "D7"
    ws.print_title_rows = "4:6"
    ws.print_title_cols = "A:C"

    return {"all_pass_cell": f"'{ws.title}'!$C$4"}
