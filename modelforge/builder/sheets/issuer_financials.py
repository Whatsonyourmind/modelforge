"""Issuer Financials sheet (Minibond Template 2).

Simpler than Unitranche OperatingModel: Revenue → EBITDA → EBIT → NI → FCF,
with FCF/interest coverage (DSCR) prominently computed for bondholders.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.formulas import growth, margin, sum_list, add
from modelforge.builder.i18n import L
from modelforge.graph.schema import LinkageGraph, GraphNode, NodeKind


def build(ws: Worksheet, spec, graph: LinkageGraph, driver_refs: dict[str, str]) -> dict[str, str]:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    n = h + p

    layout.set_column_widths(ws, label_width=40, it_width=32, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Issuer Financials", title_it="Financials dell'emittente",
        subtitle=f"{spec.meta.currency} {spec.meta.unit_scale} · sign: {spec.meta.sign_convention}",
    )
    layout.write_scenario_banner(ws, row=3)

    # Year header row
    yr_row = 5
    base_fy_year = spec.target.last_fy_end.year
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        yr = base_fy_year - (h - 1) + i
        is_hist = i < h
        c = ws.cell(row=yr_row, column=col_idx,
                    value=f"{'A' if is_hist else 'E'} {yr}")
        styles.style_header(c)

    rows: dict[str, int] = {}
    r = 7

    # Revenue growth
    layout.write_section_header(ws, r, "Revenue build", "Costruzione ricavi")
    r += 1
    rows["revenue_growth"] = r
    layout.write_row_label(ws, r, L("revenue_growth").en, L("revenue_growth").secondary)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(p):
        a = spec.operating.revenue_growth_by_year[i]
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(c, number_format=styles.FMT_PCT)
    r += 1

    # Revenue
    rows["revenue"] = r
    layout.write_row_label(ws, r, L("revenue").en, L("revenue").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=spec.historical_revenue_eur_m[i])
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c.comment = Comment(f"Historical.\nSource: {spec.target.revenue_source_id}", "ModelForge")
    for i in range(p):
        col = layout.year_col(h + i)
        col_idx = ord(col) - ord("A") + 1
        prior_col = layout.year_col(h + i - 1)
        c = ws.cell(row=r, column=col_idx,
                    value=growth(f"${prior_col}${r}", f"${col}${rows['revenue_growth']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    # Historical growth formulas
    for i in range(1, h):
        col = layout.year_col(i); prior_col = layout.year_col(i - 1)
        col_idx = ord(col) - ord("A") + 1
        gc = ws.cell(row=rows["revenue_growth"], column=col_idx,
                     value=f"=IFERROR(${col}${rows['revenue']}/${prior_col}${rows['revenue']}-1,0)")
        styles.style_formula(gc, number_format=styles.FMT_PCT)
    r += 2

    # EBITDA margin + EBITDA
    layout.write_section_header(ws, r, "Profitability", "Redditività")
    r += 1
    rows["ebitda_margin"] = r
    layout.write_row_label(ws, r, L("ebitda_margin").en, L("ebitda_margin").secondary)
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        hrev = spec.historical_revenue_eur_m[i]
        val = spec.historical_ebitda_eur_m[i] / hrev if hrev else 0
        c = ws.cell(row=r, column=col_idx, value=val)
        styles.style_input(c, number_format=styles.FMT_PCT)
    for i in range(p):
        a = spec.operating.ebitda_margin_by_year[i]
        col_idx = ord(layout.year_col(h + i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(c, number_format=styles.FMT_PCT)
    r += 1

    rows["ebitda"] = r
    layout.write_row_label(ws, r, L("ebitda").en, L("ebitda").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(h):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=spec.historical_ebitda_eur_m[i])
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c.comment = Comment(f"Historical.\nSource: {spec.target.ebitda_source_id}", "ModelForge")
    for i in range(p):
        col = layout.year_col(h + i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=margin(f"${col}${rows['revenue']}", f"${col}${rows['ebitda_margin']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 2

    # D&A, EBIT
    rows["da"] = r
    layout.write_row_label(ws, r, L("da").en, L("da").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"=-${col}${rows['revenue']}*da_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ebit"] = r
    layout.write_row_label(ws, r, L("ebit").en, L("ebit").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=add(f"${col}${rows['ebitda']}", f"${col}${rows['da']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 2

    # Interest — cross-sheet patched by bond_structure.py
    rows["interest"] = r
    layout.write_row_label(ws, r, L("interest_expense").en, L("interest_expense").secondary, indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    r += 1

    # EBT, tax, NI
    rows["ebt"] = r
    layout.write_row_label(ws, r, L("ebt").en, L("ebt").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=add(f"${col}${rows['ebit']}", f"${col}${rows['interest']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["tax"] = r
    layout.write_row_label(ws, r, L("tax").en, L("tax").secondary, indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-MAX(${col}${rows['ebt']},0)*effective_tax_rate")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_income"] = r
    layout.write_row_label(ws, r, L("net_income").en, L("net_income").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=add(f"${col}${rows['ebt']}", f"${col}${rows['tax']}"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 2

    # Cash flow
    layout.write_section_header(ws, r, "Cash flow (FCF for debt service)", "FCF a servizio del debito")
    r += 1
    rows["capex"] = r
    layout.write_row_label(ws, r, L("capex_total").en, L("capex_total").secondary, indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx, value=f"=-${col}${rows['revenue']}*capex_pct_revenue")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    rows["nwc"] = r
    layout.write_row_label(ws, r, L("nwc_change").en, L("nwc_change").secondary, indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            c = ws.cell(row=r, column=col_idx, value=0)
            styles.style_input(c, number_format=styles.FMT_EUR_M)
        else:
            prior_col = layout.year_col(i - 1)
            c = ws.cell(row=r, column=col_idx,
                        value=f"=-(${col}${rows['revenue']}-${prior_col}${rows['revenue']})*nwc_pct_revenue_delta")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["fcf"] = r
    layout.write_row_label(ws, r, L("fcf_to_debt").en, L("fcf_to_debt").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        parts = [f"${col}${rows['ebitda']}", f"${col}${rows['tax']}",
                 f"${col}${rows['capex']}", f"${col}${rows['nwc']}"]
        c = ws.cell(row=r, column=col_idx, value=sum_list(parts))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    # Register cells in graph
    for rn, rv in rows.items():
        for i in range(n):
            col = layout.year_col(i)
            cid = LinkageGraph.cell_id(ws.title, f"{col}{rv}")
            graph.add_node(GraphNode(id=cid, kind=NodeKind.CELL, label=rn))

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
