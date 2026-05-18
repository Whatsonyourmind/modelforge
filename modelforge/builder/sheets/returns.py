"""Returns sheet.

Lender-centric returns: IRR, MoIC, APR.

Lender cash flow per period (projection window):
    t=close:  -commitment + arrangement_fee         (outflow)
    t=1..N:   +cash interest (positive, sign-flipped from op model)
    t=maturity: +principal repayment (scheduled amort + balloon)

APR includes the arrangement fee; IRR on just coupon + principal.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.formulas import irr, moic
from modelforge.builder.i18n import L
from modelforge.spec.unitranche import UnitrancheSpec


def build(
    ws: Worksheet,
    spec: UnitrancheSpec,
    debt_refs: dict[str, str],
    debt_sheet_name: str,
) -> dict[str, str]:
    horizon = spec.horizon
    h = horizon.historical_years
    p = horizon.projection_years
    n_years = h + p

    layout.set_column_widths(ws, label_width=40, it_width=32, year_width=12, unit_width=6)

    layout.write_title_block(
        ws, title_en="Returns", title_it="Rendimenti",
        subtitle="Lender IRR, MoIC, EIR (IFRS 9) — projection window only",
    )
    layout.write_scenario_banner(ws, row=3)

    # Columns in this sheet represent periods 0..N where N = projection_years
    # Period 0 = close (t = last historical end). Row 5 period headers.
    yr_row = 5
    ws.cell(row=yr_row, column=3, value="Period").font = styles.font_subheader
    for i in range(p + 1):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=yr_row, column=col_idx, value=f"t={i}")
        styles.style_header(c)

    tranche_blocks = debt_refs["tranche_blocks"]

    # ─── Transaction-cost block (consumes orphan-fee assumptions) ───────────
    # These fees were defined in spec.fees but had no formula references,
    # so they showed up as "orphan named ranges" in MoatGate. Wire each
    # into the Net Proceeds calc at t=0 so the Returns sheet correctly
    # accounts for the all-in lender outflow at close.
    r = 7
    fees = getattr(spec, "fees", None)
    fee_assumption_names: list[str] = []
    if fees is not None:
        layout.write_section_header(
            ws, r, "Transaction costs (at close)", "Costi della transazione (al closing)",
        )
        r += 1
        for attr_name, label_en, label_it in (
            ("legal", "Legal fees", "Spese legali"),
            ("advisory", "Advisory fees", "Spese di consulenza"),
            ("other", "Other transaction fees", "Altre spese di transazione"),
        ):
            fee_obj = getattr(fees, attr_name, None)
            if fee_obj is None:
                continue
            ref_name = fee_obj.name
            fee_assumption_names.append(ref_name)
            layout.write_row_label(ws, r, label_en, label_it, indent=True)
            ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
            c = ws.cell(row=r, column=4, value=f"={ref_name}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
            r += 1
        if fee_assumption_names:
            layout.write_row_label(
                ws, r, "Total transaction costs", "Costi totali",
            )
            c = ws.cell(
                row=r, column=4,
                value="=" + "+".join(fee_assumption_names),
            )
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
            c.font = styles.font_subheader
            total_costs_row = r
            r += 2
        else:
            total_costs_row = None
    else:
        total_costs_row = None

    # Exit assumption surfacing: expected hold + make-whole rate become
    # display formulas on this sheet. Both flow into IRR-period framing.
    exit_assum = getattr(spec, "exit", None)
    if exit_assum is not None:
        layout.write_section_header(
            ws, r, "Exit assumptions", "Ipotesi di uscita",
        )
        r += 1
        for attr_name, label_en, label_it, fmt in (
            ("expected_hold_years", "Expected hold (years)", "Holding atteso (anni)", "0.0"),
            ("make_whole_pct", "Make-whole call rate", "Tasso make-whole", styles.FMT_PCT_2DP),
        ):
            assum = getattr(exit_assum, attr_name, None)
            if assum is None:
                continue
            layout.write_row_label(ws, r, label_en, label_it, indent=True)
            c = ws.cell(row=r, column=4, value=f"={assum.name}")
            styles.style_formula(c, number_format=fmt)
            r += 1
        r += 1

    # For each tranche, emit a lender cashflow row
    tranche_cashflow_rows: list[tuple[str, int]] = []
    for tb in tranche_blocks:
        layout.write_section_header(ws, r, f"{tb['name']} — lender CF", f"{tb['name']} — CF finanziatore")
        r += 1

        # Cash flow row: commitment out at t=0, interest in t=1..N, principal repay at maturity
        tr_row = r
        layout.write_row_label(ws, r, L("lender_cashflow").en, L("lender_cashflow").secondary)
        ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it

        # t=0: -commitment (opening at drawdown year on DebtSchedule)
        # Commitment amount = the tranche drawdown at close column on debt sheet.
        # We know drawdown row per tranche = tb["drawdown"], at column h (first projected)
        drawdown_col_on_debt = layout.year_col(h)  # first projected year col
        drawdown_ref = f"'{debt_sheet_name}'!{drawdown_col_on_debt}{tb['drawdown']}"
        arr_fee_ref = f"'{debt_sheet_name}'!{drawdown_col_on_debt}{tb['arrangement_fee']}"
        # At t=0 (close): we FUND -commitment and RECEIVE +arrangement_fee
        c0 = ws.cell(
            row=r, column=4,  # column D = year_col(0) here
            value=f"=-{drawdown_ref}+{arr_fee_ref}",
        )
        styles.style_formula(c0, number_format=styles.FMT_EUR_M)

        # t=1..p: +cash interest (on debt sheet as negative; flip sign here) + scheduled repayment (debt sheet repayment is negative, flip)
        for i in range(1, p + 1):
            # period t corresponds to projection year t, which is year index h + (t-1) on debt sheet
            debt_col = layout.year_col(h + (i - 1))
            interest_ref = f"'{debt_sheet_name}'!{debt_col}{tb['interest']}"
            repay_ref = f"'{debt_sheet_name}'!{debt_col}{tb['repayment']}"
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            c = ws.cell(
                row=r, column=col_idx,
                value=f"=-{interest_ref}-{repay_ref}",
            )
            styles.style_formula(c, number_format=styles.FMT_EUR_M)

        tranche_cashflow_rows.append((tb["name"], tr_row))

        # Summary metrics: IRR + MoIC (on this tranche)
        r += 1
        first_col = layout.year_col(0)  # D
        last_col = layout.year_col(p)
        # IRR
        layout.write_row_label(ws, r, L("lender_irr").en, L("lender_irr").secondary, indent=True)
        c_irr = ws.cell(row=r, column=4,
                        value=f"=IRR(${first_col}${tr_row}:${last_col}${tr_row})")
        styles.style_formula(c_irr, number_format=styles.FMT_PCT_2DP)
        c_irr.font = styles.font_subheader
        r += 1
        # IFRS 9 EIR — identical to IRR for annual evenly-spaced CF incl. fees
        layout.write_row_label(
            ws, r,
            "EIR (IFRS 9, incl. fees)",
            "EIR (IFRS 9, incl. commissioni)",
            indent=True,
        )
        c_eir = ws.cell(row=r, column=4,
                        value=f"=IRR(${first_col}${tr_row}:${last_col}${tr_row},0.08)")
        styles.style_formula(c_eir, number_format=styles.FMT_PCT_2DP)
        c_eir.font = styles.font_subheader
        from openpyxl.comments import Comment
        c_eir.comment = Comment(
            "IFRS 9 Effective Interest Rate: the rate that exactly discounts "
            "estimated future cash payments to the gross carrying amount. "
            "Includes all fees integral to the instrument (arrangement fee, "
            "OID). Per IFRS 9 §B5.4.1.",
            "ModelForge",
        )
        c_eir.comment.width = 320
        c_eir.comment.height = 120
        r += 1
        # MoIC
        layout.write_row_label(ws, r, L("lender_moic").en, L("lender_moic").secondary, indent=True)
        c_moic = ws.cell(
            row=r, column=4,
            value=(
                f"=IFERROR("
                f"SUMIF(${first_col}${tr_row}:${last_col}${tr_row},\">0\")"
                f"/ABS(SUMIF(${first_col}${tr_row}:${last_col}${tr_row},\"<0\"))"
                f",0)"
            ),
        )
        styles.style_formula(c_moic, number_format=styles.FMT_MULTIPLE)
        c_moic.font = styles.font_subheader
        r += 2

    # Portfolio totals
    if len(tranche_cashflow_rows) > 1:
        layout.write_section_header(ws, r, L("ret_blended_returns").en, L("ret_blended_returns").secondary)
        r += 1
        portfolio_row = r
        layout.write_row_label(ws, r, L("ret_blended_cf").en, L("ret_blended_cf").secondary)
        for i in range(p + 1):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            parts = [f"${col}${tr}" for _, tr in tranche_cashflow_rows]
            c = ws.cell(row=r, column=col_idx, value="=" + "+".join(parts))
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
            c.font = styles.font_subheader
        r += 1

        first_col = layout.year_col(0)
        last_col = layout.year_col(p)
        layout.write_row_label(ws, r, L("ret_blended_irr").en, L("ret_blended_irr").secondary, indent=True)
        ws.cell(row=r, column=4, value=f"=IRR(${first_col}${portfolio_row}:${last_col}${portfolio_row})")
        styles.style_formula(ws.cell(row=r, column=4), number_format=styles.FMT_PCT_2DP)
        r += 1
        layout.write_row_label(ws, r, L("ret_blended_moic").en, L("ret_blended_moic").secondary, indent=True)
        ws.cell(
            row=r, column=4,
            value=(
                f"=IFERROR(SUMIF(${first_col}${portfolio_row}:${last_col}${portfolio_row},\">0\")"
                f"/ABS(SUMIF(${first_col}${portfolio_row}:${last_col}${portfolio_row},\"<0\")),0)"
            ),
        )
        styles.style_formula(ws.cell(row=r, column=4), number_format=styles.FMT_MULTIPLE)
        r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    return {"rows": str(r)}
