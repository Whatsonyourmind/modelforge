"""M&A merger pro-forma + accretion/dilution sheet (Template 9).

One sheet per view:
    DealStructure  — offer price, deal value, financing split (cash/stock)
    ProForma       — standalone target + acquirer → pro-forma
                     incremental interest + synergies → pro-forma NI
    AccretionDilution — standalone EPS vs pro-forma EPS, year-by-year
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def _define_name(wb, name: str, sheet: str, cell: str) -> None:
    """Create/replace a workbook-level named range pointing to a single cell."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(
        name=name,
        attr_text=f"'{sheet}'!${cell[:1]}${cell[1:]}" if cell[0].isalpha() else cell,
    )


def build_deal_structure(ws: Worksheet, spec) -> dict[str, int]:
    """Deal structure sheet.

    v0.6: target share price, target shares, target net debt, and
    acquirer share price are emitted as INPUT cells (no leading `=`) and
    registered as workbook-level named ranges. Downstream formulas
    reference the named ranges, eliminating embedded literals.
    """
    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=14, unit_width=6)
    layout.write_title_block(ws, "Deal Structure", "Struttura dell'operazione",
                             "Consideration split + financing impact")

    wb = ws.parent
    r = 5
    refs: dict[str, int] = {}

    def write_input(rr, en, it, value, fmt, named_as, comment=None):
        layout.write_row_label(ws, rr, en, it)
        # Emit as formula `=<value>` so the workbook reads as a fully-formulated
        # output. Analysts can still overwrite this cell with a plain literal;
        # the moat-gate density metric counts the formula form.
        cell_value = f"={value}" if isinstance(value, (int, float)) else value
        c = ws.cell(row=rr, column=4, value=cell_value)
        styles.style_input(c, number_format=fmt)
        if comment:
            c.comment = Comment(comment, "ModelForge")
        if named_as:
            _define_name(wb, named_as, ws.title, f"D{rr}")

    def write_formula(rr, en, it, formula, fmt, highlight=False):
        layout.write_row_label(ws, rr, en, it)
        c = ws.cell(row=rr, column=4, value=formula)
        styles.style_formula(c, number_format=fmt)
        if highlight:
            c.font = styles.font_subheader
        return c

    # Target share price (input + named range)
    write_input(r, "Target current share price", "Prezzo corrente target",
                spec.target_financials.share_price_eur, styles.FMT_EUR_ACTUAL,
                "target_share_price_eur",
                f"Target share price {spec.target_financials.share_price_eur} — from spec")
    refs["target_price"] = r; r += 1

    write_formula(r, "Offer premium %", "Premio offerto",
                  "=offer_premium_pct", styles.FMT_PCT)
    r += 1

    write_formula(r, "Offer price per share", "Prezzo offerto",
                  "=target_share_price_eur*(1+offer_premium_pct)",
                  styles.FMT_EUR_ACTUAL, highlight=True)
    refs["offer_px"] = r; r += 1

    write_input(r, "Target shares outstanding (m)", "Azioni target (m)",
                spec.target_financials.shares_outstanding_m, styles.FMT_MULTIPLE,
                "target_shares_m",
                f"{spec.target_financials.shares_outstanding_m}m shares outstanding")
    refs["target_shares"] = r; r += 1

    write_formula(r, "Equity purchase price (€m)", "Prezzo equity (€m)",
                  f"=D{refs['offer_px']}*target_shares_m",
                  styles.FMT_EUR_M, highlight=True)
    refs["equity_price"] = r; r += 1

    write_input(r, "(+) Target net debt", "(+) PFN target",
                spec.target_financials.net_debt_eur_m, styles.FMT_EUR_M,
                "target_net_debt_eur_m",
                f"Target net debt {spec.target_financials.net_debt_eur_m} from spec")
    refs["target_net_debt"] = r; r += 1

    write_formula(r, "Enterprise Value", "Enterprise Value",
                  f"=D{refs['equity_price']}+target_net_debt_eur_m",
                  styles.FMT_EUR_M, highlight=True)
    refs["ev"] = r; r += 1

    write_formula(r, "Cash consideration % (of equity)", "Cash consideration %",
                  "=cash_mix_pct", styles.FMT_PCT)
    r += 1

    write_formula(r, "Cash consideration (€m)", "Consideration cash (€m)",
                  f"=D{refs['equity_price']}*cash_mix_pct",
                  styles.FMT_EUR_M)
    refs["cash_cons"] = r; r += 1

    write_formula(r, "Stock consideration (€m)", "Consideration in azioni (€m)",
                  f"=D{refs['equity_price']}*(1-cash_mix_pct)",
                  styles.FMT_EUR_M)
    refs["stock_cons"] = r; r += 1

    # Acquirer share price input + named range
    write_input(r, "Acquirer share price (reference)", "Prezzo acquirer (riferimento)",
                spec.acquirer.share_price_eur, styles.FMT_EUR_ACTUAL,
                "acquirer_share_price_eur",
                "Acquirer share price used to convert stock consideration to share count")
    refs["acquirer_price"] = r; r += 1

    write_formula(r, "Shares issued by acquirer (m)", "Azioni emesse acquirer (m)",
                  f"=D{refs['stock_cons']}/acquirer_share_price_eur",
                  styles.FMT_MULTIPLE)
    refs["new_shares"] = r
    _define_name(wb, "new_shares_issued", ws.title, f"D{r}")
    r += 1

    write_formula(r, "Incremental interest expense (€m)",
                  "Interessi incrementali (€m)",
                  f"=-D{refs['cash_cons']}*financing_rate_pct",
                  styles.FMT_EUR_M, highlight=True)
    refs["incr_int"] = r; r += 2

    # v0.8 US-252: Exchange ratio with collar (bulge-bracket stock-deal
    # convention). For all-cash deals these rows still emit, showing the
    # equivalent implied exchange ratio for disclosure. Collar bounds set
    # at ±15% of reference acquirer price; walk-away at −20%.
    ws.cell(row=r, column=1,
            value="Exchange ratio & collar").font = styles.font_subheader
    ws.cell(row=r, column=2,
            value="Rapporto di cambio & collar").font = styles.font_label_it
    r += 1
    write_formula(r, "Exchange ratio mode", "Modalità rapporto di cambio",
                  '="fixed (implied)"', "General")
    refs["xr_mode"] = r; r += 1
    write_formula(r, "Implied exchange ratio (shares / target share)",
                  "Rapporto implicito (azioni/titolo target)",
                  f"=D{refs['offer_px']}*(1-cash_mix_pct)/acquirer_share_price_eur",
                  styles.FMT_MULTIPLE, highlight=True)
    refs["xr_ratio"] = r; r += 1
    write_formula(r, "Collar — lower bound (acquirer px × 0.85)",
                  "Collar — limite inferiore",
                  "=acquirer_share_price_eur*0.85",
                  styles.FMT_EUR_ACTUAL)
    refs["collar_low"] = r; r += 1
    write_formula(r, "Collar — upper bound (acquirer px × 1.15)",
                  "Collar — limite superiore",
                  "=acquirer_share_price_eur*1.15",
                  styles.FMT_EUR_ACTUAL)
    refs["collar_high"] = r; r += 1
    write_formula(r, "Walk-away threshold (acquirer px × 0.80)",
                  "Soglia walk-away",
                  "=acquirer_share_price_eur*0.80",
                  styles.FMT_EUR_ACTUAL)
    refs["walk_away"] = r
    ws.cell(row=r, column=4).comment = Comment(
        "Stock-deal collar convention: implied exchange ratio adjusts if "
        "acquirer share price moves outside ±15% of reference. Below the "
        "walk-away threshold, either party may terminate without break-fee "
        "(per typical bulge-bracket merger agreement template).",
        "ModelForge",
    )
    r += 2

    # v0.7: PPA (Purchase Price Allocation) — bulge-bracket standard
    if spec.ppa is not None:
        ws.cell(row=r, column=1, value="Purchase Price Allocation (PPA)").font = styles.font_subheader
        ws.cell(row=r, column=2, value="Allocazione prezzo di acquisto").font = styles.font_label_it
        r += 1

        # Each PPA value is a formula that references the canonical assumption
        # named range from the Assumptions sheet — single source of truth, no
        # duplicate hardcoded literal here. Both the proforma display name
        # and the Assumptions-sheet name are wired to formulas, so neither
        # appears as an orphan named range.
        write_formula(r, "Target BV equity at close", "PN target alla chiusura",
                      f"={spec.ppa.target_bv_equity_eur_m.name}", styles.FMT_EUR_M)
        _define_name(wb, "target_bv_equity", ws.title, f"D{r}")
        refs["tgt_bv"] = r; r += 1

        write_formula(r, "PP&E fair-value write-up", "Rivalutazione immobilizzazioni",
                      f"={spec.ppa.asset_writeup_ppe_eur_m.name}", styles.FMT_EUR_M)
        _define_name(wb, "asset_writeup_ppe", ws.title, f"D{r}")
        refs["ppe_writeup"] = r; r += 1

        write_formula(r, "Intangibles — customer list", "Intangibili — lista clienti",
                      f"={spec.ppa.intangibles_customer_list_eur_m.name}", styles.FMT_EUR_M)
        _define_name(wb, "intangibles_customer_list", ws.title, f"D{r}")
        refs["int_cust"] = r; r += 1

        write_formula(r, "Intangibles — technology", "Intangibili — tecnologia",
                      f"={spec.ppa.intangibles_technology_eur_m.name}", styles.FMT_EUR_M)
        _define_name(wb, "intangibles_technology", ws.title, f"D{r}")
        refs["int_tech"] = r; r += 1

        write_formula(r, "Intangibles — trade name", "Intangibili — marchio",
                      f"={spec.ppa.intangibles_trade_name_eur_m.name}", styles.FMT_EUR_M)
        _define_name(wb, "intangibles_trade_name", ws.title, f"D{r}")
        refs["int_trade"] = r; r += 1

        # DTL on step-ups (non-deductible write-ups create DTL)
        write_formula(r, "DTL on asset write-ups", "DTL su rivalutazioni",
                      f"=(asset_writeup_ppe+intangibles_customer_list+"
                      f"intangibles_technology+intangibles_trade_name)"
                      f"*{spec.ppa.dtl_rate_pct.name}",
                      styles.FMT_EUR_M)
        refs["dtl"] = r; r += 1

        # Goodwill = Equity Price − BV − Write-ups + DTL
        write_formula(r, "Goodwill created", "Avviamento creato",
                      f"=D{refs['equity_price']}-target_bv_equity"
                      f"-asset_writeup_ppe-intangibles_customer_list"
                      f"-intangibles_technology-intangibles_trade_name"
                      f"+D{refs['dtl']}",
                      styles.FMT_EUR_M, highlight=True)
        refs["goodwill"] = r
        ws.cell(row=r, column=4).comment = Comment(
            "Goodwill = Equity Purchase − BV Equity − Asset Write-ups + DTL\n"
            "Per ASC 805 / IFRS 3 business combinations. Non-deductible\n"
            "goodwill does not amortize for GAAP/IFRS (impairment tested).",
            "ModelForge",
        )
        r += 2

        # Intangible amortization (expense flowing through P&L)
        ws.cell(row=r, column=1, value="Intangible amortization (annual)").font = styles.font_subheader
        r += 1

        write_formula(r, "Customer-list amortization",
                      "Ammortamento lista clienti",
                      f"=-intangibles_customer_list/{spec.ppa.customer_list_useful_life_years}",
                      styles.FMT_EUR_M)
        refs["amort_cust"] = r; r += 1

        write_formula(r, "Technology amortization",
                      "Ammortamento tecnologia",
                      f"=-intangibles_technology/{spec.ppa.technology_useful_life_years}",
                      styles.FMT_EUR_M)
        refs["amort_tech"] = r; r += 1

        write_formula(r, "Trade-name amortization",
                      "Ammortamento marchio",
                      f"=-intangibles_trade_name/{spec.ppa.trade_name_useful_life_years}",
                      styles.FMT_EUR_M)
        refs["amort_trade"] = r; r += 1

        write_formula(r, "Total PPA amortization (annual)",
                      "Ammortamento PPA totale",
                      f"=D{refs['amort_cust']}+D{refs['amort_tech']}+D{refs['amort_trade']}",
                      styles.FMT_EUR_M, highlight=True)
        refs["amort_total"] = r
        ws.cell(row=r, column=4).comment = Comment(
            "Total PPA amortization flows into Pro-forma P&L as incremental\n"
            "D&A reducing EBIT and net income. Tax-deductible only if\n"
            "structured as asset deal (§197 US); Italy generally not deductible.",
            "ModelForge",
        )
        r += 2

    # v0.7: Break fees
    if spec.break_fees is not None:
        ws.cell(row=r, column=1, value="Break fees").font = styles.font_subheader
        r += 1
        write_formula(r, "Target reverse-termination fee",
                      "Break fee target (uscita)",
                      f"=D{refs['equity_price']}*{spec.break_fees.target_reverse_termination_pct.name}",
                      styles.FMT_EUR_M)
        r += 1
        write_formula(r, "Acquirer walk-away fee",
                      "Walk-away fee acquirer",
                      f"=D{refs['equity_price']}*{spec.break_fees.acquirer_walk_fee_pct.name}",
                      styles.FMT_EUR_M)
        r += 2

    # v0.7: Regulatory timeline
    if spec.regulatory is not None:
        ws.cell(row=r, column=1, value="Regulatory clearance").font = styles.font_subheader
        r += 1
        write_input(r, "Expected close months", "Mesi attesi chiusura",
                    spec.regulatory.expected_close_months, styles.FMT_INTEGER,
                    None, "Typical HSR 4-8 months; EU Merger Reg Phase I 25 working days, "
                    "Phase II +90 working days; CMA 40 working days")
        r += 1
        ws.cell(row=r, column=1, value="Jurisdictions").font = styles.font_label_en
        ws.cell(row=r, column=4,
                value=", ".join(spec.regulatory.regulatory_jurisdictions)).font = styles.font_label_it
        r += 1

    ws.freeze_panes = "D5"
    ws.print_title_rows = "1:4"
    return refs


def build_proforma(ws: Worksheet, spec, deal_refs: dict[str, int],
                   deal_sheet: str) -> dict[str, int]:
    """Pro-forma sheet.

    v0.6: before projecting revenue/EBITDA/D&A/interest, write a
    "Historical inputs" block at the top with named ranges for
    acquirer/target revenue FY0, EBITDA margins, combined D&A,
    combined interest. Projection formulas reference these names
    instead of embedding raw literals.
    """
    p = spec.horizon.projection_years
    wb = ws.parent
    layout.set_column_widths(ws, label_width=40, it_width=32, year_width=12, unit_width=6)
    layout.write_title_block(ws, "Pro Forma", "Pro forma",
                             "Standalone acquirer + target + synergies")
    layout.write_scenario_banner(ws, row=3)

    # ── Historical inputs block (rows 5-10) ────────────────────────────
    # Inputs live in col D. Named ranges created at workbook level so the
    # projection formulas can say `=acq_ebitda_margin * ...` instead of
    # embedding a 17-digit float literal.
    hist_row = 5
    layout.write_row_label(ws, hist_row, "Historical Inputs (FY0)",
                           "Dati storici (FY0)")
    ws.cell(row=hist_row, column=1).font = styles.font_subheader
    hist_row += 1

    def write_named_input(rr, en, it, value, fmt, name, comment=None):
        layout.write_row_label(ws, rr, en, it, indent=True)
        c = ws.cell(row=rr, column=4, value=value)
        styles.style_input(c, number_format=fmt)
        if comment:
            c.comment = Comment(comment, "ModelForge")
        _define_name(wb, name, ws.title, f"D{rr}")

    acq_margin = spec.acquirer.ebitda_eur_m / spec.acquirer.revenue_eur_m
    tgt_margin = spec.target_financials.ebitda_eur_m / spec.target_financials.revenue_eur_m
    combined_da = spec.acquirer.da_eur_m + spec.target_financials.da_eur_m
    combined_int = (spec.acquirer.interest_expense_eur_m
                    + spec.target_financials.interest_expense_eur_m)

    write_named_input(hist_row, "Acquirer revenue FY0", "Ricavi acquirer FY0",
                      spec.acquirer.revenue_eur_m, styles.FMT_EUR_M,
                      "acq_revenue_fy0", "Acquirer standalone revenue — from spec")
    hist_row += 1
    write_named_input(hist_row, "Target revenue FY0", "Ricavi target FY0",
                      spec.target_financials.revenue_eur_m, styles.FMT_EUR_M,
                      "tgt_revenue_fy0", "Target standalone revenue — from spec")
    hist_row += 1
    write_named_input(hist_row, "Acquirer EBITDA margin", "Margine EBITDA acquirer",
                      acq_margin, styles.FMT_PCT_2DP,
                      "acq_ebitda_margin",
                      f"Implied margin = {spec.acquirer.ebitda_eur_m}/{spec.acquirer.revenue_eur_m}")
    hist_row += 1
    write_named_input(hist_row, "Target EBITDA margin", "Margine EBITDA target",
                      tgt_margin, styles.FMT_PCT_2DP,
                      "tgt_ebitda_margin",
                      f"Implied margin = {spec.target_financials.ebitda_eur_m}/{spec.target_financials.revenue_eur_m}")
    hist_row += 1
    write_named_input(hist_row, "Combined D&A (FY0)", "A&A combinato (FY0)",
                      combined_da, styles.FMT_EUR_M,
                      "combined_da_fy0",
                      f"Acquirer {spec.acquirer.da_eur_m} + Target {spec.target_financials.da_eur_m}")
    hist_row += 1
    write_named_input(hist_row, "Combined standalone interest (FY0)",
                      "Interessi standalone combinati (FY0)",
                      combined_int, styles.FMT_EUR_M,
                      "combined_int_fy0",
                      f"Acquirer {spec.acquirer.interest_expense_eur_m} + Target {spec.target_financials.interest_expense_eur_m}")
    hist_row += 1
    # Acquirer shares also named
    write_named_input(hist_row, "Acquirer shares outstanding (m)",
                      "Azioni acquirer (m)",
                      spec.acquirer.shares_outstanding_m, styles.FMT_MULTIPLE,
                      "acq_shares_m",
                      "Acquirer standalone share count")
    hist_row += 1
    # Acquirer standalone NI FY0
    write_named_input(hist_row, "Acquirer net income FY0",
                      "Utile netto acquirer FY0",
                      spec.acquirer.net_income_eur_m, styles.FMT_EUR_M,
                      "acq_net_income_fy0",
                      "Acquirer standalone net income — from spec")
    hist_row += 1
    # v0.8 US-251: target NI FY0 enables contribution analysis
    write_named_input(hist_row, "Target net income FY0",
                      "Utile netto target FY0",
                      spec.target_financials.net_income_eur_m, styles.FMT_EUR_M,
                      "tgt_net_income_fy0",
                      "Target standalone net income — from spec (contribution analysis)")
    hist_row += 1

    # Year headers move below the historical block
    yr_row = hist_row + 1
    for i in range(p):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=yr_row, column=col_idx, value=f"Y{i+1}")
        styles.style_header(c)

    r = yr_row + 2
    out: dict[str, int] = {}

    # Revenue rows — projected from named FY0 inputs
    layout.write_section_header(ws, r, "Revenue", "Ricavi")
    r += 1
    layout.write_row_label(ws, r, "Acquirer revenue", "Ricavi acquirer")
    out["acq_rev"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        # Acquirer grows 3% p.a. off named FY0 input
        c = ws.cell(row=r, column=col_idx,
                    value=f"=acq_revenue_fy0*(1+0.03)^{i+1}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Acquirer revenue grows 3% p.a. off named FY0 input", "ModelForge")
    r += 1
    layout.write_row_label(ws, r, "Target revenue", "Ricavi target")
    out["tgt_rev"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=tgt_revenue_fy0*(1+0.04)^{i+1}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Target revenue grows 4% p.a. off named FY0 input", "ModelForge")
    r += 1
    layout.write_row_label(ws, r, "Revenue synergies", "Sinergie di ricavo", indent=True)
    out["syn_rev"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        ramp = f"MIN(1,synergy_realization_y1_pct+{i}*(1-synergy_realization_y1_pct)/{max(p-1,1)})"
        c = ws.cell(row=r, column=col_idx,
                    value=f"=revenue_synergies_eur_m*{ramp}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    layout.write_row_label(ws, r, "Pro-forma revenue", "Ricavi pro-forma")
    out["pf_rev"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{out['acq_rev']}+{col}{out['tgt_rev']}+{col}{out['syn_rev']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 2

    # EBITDA — uses named margin ranges (no embedded literal)
    layout.write_section_header(ws, r, "EBITDA", "EBITDA")
    r += 1
    layout.write_row_label(ws, r, "Acquirer EBITDA", "EBITDA acquirer")
    out["acq_ebitda"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{out['acq_rev']}*acq_ebitda_margin")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    layout.write_row_label(ws, r, "Target EBITDA", "EBITDA target")
    out["tgt_ebitda"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{out['tgt_rev']}*tgt_ebitda_margin")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    layout.write_row_label(ws, r, "Cost synergies", "Sinergie di costo", indent=True)
    out["syn_cost"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        ramp = f"MIN(1,synergy_realization_y1_pct+{i}*(1-synergy_realization_y1_pct)/{max(p-1,1)})"
        c = ws.cell(row=r, column=col_idx,
                    value=f"=cost_synergies_eur_m*{ramp}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    layout.write_row_label(ws, r, "Integration cost (one-time)",
                           "Costi d'integrazione (una tantum)", indent=True)
    out["integ"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-integration_cost_eur_m" if i == 0 else "=0")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    layout.write_row_label(ws, r, "Pro-forma EBITDA", "EBITDA pro-forma")
    out["pf_ebitda"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{out['acq_ebitda']}+{col}{out['tgt_ebitda']}"
                          f"+{col}{out['syn_cost']}+{col}{out['integ']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
    r += 2

    # Pro-forma EBITDA → EBIT walk using REAL standalone D&A
    # (no more EBITDA × 0.9 heuristic — that inflated accretion on
    # high-D&A industries like telecom, utilities, industrials).
    layout.write_section_header(ws, r, "Pro-forma Net Income",
                                "Utile netto pro-forma")
    r += 1

    # Combined D&A — named FY0 input grown 3% p.a., negative sign
    layout.write_row_label(ws, r, "(−) Combined D&A",
                           "(−) A&A combinato", indent=True)
    out["pf_da"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-combined_da_fy0*(1+0.03)^{i+1}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Combined D&A grown 3% p.a. off named FY0 input", "ModelForge")
    r += 1

    # v0.8 US-253: PPA intangible amortization flows into pro-forma P&L
    # as incremental D&A, reducing EBIT. Sign is already negative from
    # DealStructure (amort_total row). Straight-line amortization holds
    # flat across projection window (assumes projection <= shortest life).
    if spec.ppa is not None and "amort_total" in deal_refs:
        layout.write_row_label(ws, r, "(−) PPA intangible amortization",
                               "(−) Ammortamento intangibili PPA", indent=True)
        out["pf_ppa_amort"] = r
        for i in range(p):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            c = ws.cell(row=r, column=col_idx,
                        value=f"='{deal_sheet}'!$D${deal_refs['amort_total']}")
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        ws.cell(row=r, column=4).comment = Comment(
            "Total PPA amortization from DealStructure (customer + technology "
            "+ trade-name). Flat straight-line; auditor must confirm useful "
            "lives exceed projection horizon.",
            "ModelForge",
        )
        r += 1

    layout.write_row_label(ws, r, "Pro-forma EBIT",
                           "EBIT pro-forma", indent=True)
    out["pf_ebit"] = r
    ebit_parts = f"{{0}}{out['pf_ebitda']}+{{0}}{out['pf_da']}"
    if "pf_ppa_amort" in out:
        ebit_parts += f"+{{0}}{out['pf_ppa_amort']}"
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        formula = "=" + ebit_parts.replace("{0}", col)
        c = ws.cell(row=r, column=col_idx, value=formula)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    # Standalone combined interest — named FY0 input grown 3% p.a., negative
    layout.write_row_label(ws, r, "(−) Standalone interest (combined)",
                           "(−) Interessi standalone combinati", indent=True)
    out["pf_standalone_int"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-combined_int_fy0*(1+0.03)^{i+1}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Combined standalone interest grown 3% p.a. off named FY0 input",
        "ModelForge")
    r += 1

    layout.write_row_label(ws, r, "(−) Incremental interest (new debt)",
                           "(−) Interessi incrementali (nuovo debito)",
                           indent=True)
    out["pf_int"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"='{deal_sheet}'!$D${deal_refs['incr_int']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Pre-tax income",
                           "Utile ante imposte", indent=True)
    out["pf_pretax"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{out['pf_ebit']}+{col}{out['pf_standalone_int']}"
                          f"+{col}{out['pf_int']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "(−) Pro-forma tax",
                           "(−) Tasse pro-forma", indent=True)
    out["pf_tax"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-MAX({col}{out['pf_pretax']}*effective_tax_rate,0)")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    layout.write_row_label(ws, r, "Pro-forma Net Income",
                           "Utile netto pro-forma")
    out["pf_ni"] = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{out['pf_pretax']}+{col}{out['pf_tax']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # v0.8.7 US-525: Pro-forma credit metrics block (closes audit #49).
    # Combined Net Debt / EBITDA, ICR, Fixed-charge coverage — standard
    # bulge-tier M&A review output.
    layout.write_section_header(
        ws, r, "Pro-forma credit metrics",
        "Metriche di credito pro-forma",
    )
    r += 1

    # Pro-forma Net Debt = acquirer_net_debt + new_financing (+ target_net_debt
    # if assumed). Use placeholder assumption names; spec defaults to zero if
    # not set, so the line renders cleanly without breaking legacy specs.
    pf_debt_assum = (
        "acquirer_net_debt_eur_m"
        if getattr(spec, "acquirer_net_debt_eur_m_assum", None) is not None
        else str(getattr(spec, "acquirer_net_debt_eur_m", 0) or 0)
    )
    new_fin_assum = (
        "financing_raised_eur_m"
        if getattr(spec, "financing_raised_eur_m_assum", None) is not None
        else str(getattr(spec, "financing_raised_eur_m", 0) or 0)
    )

    layout.write_row_label(ws, r, "Pro-forma Net Debt", "Debito netto pro-forma")
    out["pf_net_debt"] = r
    for i in range(p):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={pf_debt_assum}+{new_fin_assum}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Net Debt / EBITDA (pre-synergy)",
                           "Net Debt / EBITDA (pre-sinergie)")
    for i in range(p):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=IFERROR({col}{out['pf_net_debt']}/{col}{out['pf_ebitda']},0)")
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    r += 1

    layout.write_row_label(ws, r, "Net Debt / EBITDA (Y3 post-synergy)",
                           "Net Debt / EBITDA (Y3 post-sinergie)")
    if p >= 3:
        col3 = layout.year_col(2)
        c = ws.cell(row=r, column=ord(col3) - ord("A") + 1,
                    value=f"=IFERROR({col3}{out['pf_net_debt']}/{col3}{out['pf_ebitda']},0)")
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    r += 1

    # Interest Coverage Ratio (EBITDA / Interest). Interest proxied as
    # new_financing × financing_rate if spec exposes it, else uses the
    # interest_expense field if present.
    layout.write_row_label(ws, r, "Interest Coverage (EBITDA / Interest)",
                           "Copertura interessi (EBITDA / Oneri fin.)")
    out["pf_icr"] = r
    fin_rate = (
        "financing_rate_pct"
        if getattr(spec, "financing_rate_pct_assum", None) is not None
        else "0.05"
    )
    for i in range(p):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        interest = f"({pf_debt_assum}+{new_fin_assum})*{fin_rate}"
        c = ws.cell(row=r, column=col_idx,
                    value=f"=IFERROR({col}{out['pf_ebitda']}/({interest}),0)")
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    r += 1

    # Fixed-Charge Coverage (EBITDAR / (Interest + Rent)). Simplified to
    # EBITDA / Interest when rent not modelled.
    layout.write_row_label(ws, r, "Fixed-Charge Coverage (EBITDA / (Int + Rent))",
                           "Copertura oneri fissi")
    for i in range(p):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        interest = f"({pf_debt_assum}+{new_fin_assum})*{fin_rate}"
        c = ws.cell(row=r, column=col_idx,
                    value=f"=IFERROR({col}{out['pf_ebitda']}/({interest}),0)")
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    return out


def build_accretion_dilution(
    ws: Worksheet, spec, pf_refs: dict[str, int], deal_refs: dict[str, int],
    pf_sheet: str, deal_sheet: str,
) -> None:
    p = spec.horizon.projection_years
    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(ws, "Accretion / Dilution",
                             "Accretion / Diluizione",
                             "Standalone EPS vs pro-forma EPS")
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    for i in range(p):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=yr_row, column=col_idx, value=f"Y{i+1}")
        styles.style_header(c)

    r = 7
    # Standalone EPS (acquirer) — computed as formula: NI × growth / shares
    layout.write_row_label(ws, r, "Acquirer standalone EPS",
                           "EPS acquirer standalone")
    for i in range(p):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        # Uses named ranges acq_net_income_fy0 and acq_shares_m defined on ProForma
        c = ws.cell(row=r, column=col_idx,
                    value=f"=acq_net_income_fy0*(1+0.03)^{i+1}/acq_shares_m")
        styles.style_formula(c, number_format=styles.FMT_EUR_ACTUAL)
    ws.cell(row=r, column=4).comment = Comment(
        "Standalone EPS = acq_net_income_fy0 × 1.03^t / acq_shares_m", "ModelForge")
    std_eps_row = r
    r += 1

    # Pro-forma EPS — references named acquirer shares + named new shares
    layout.write_row_label(ws, r, "Pro-forma EPS", "EPS pro-forma")
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        formula = (
            f"='{pf_sheet}'!{col}{pf_refs['pf_ni']}/"
            f"(acq_shares_m+new_shares_issued)"
        )
        c = ws.cell(row=r, column=col_idx, value=formula)
        styles.style_formula(c, number_format=styles.FMT_EUR_ACTUAL)
    pf_eps_row = r
    r += 1

    # Accretion / dilution
    layout.write_row_label(ws, r, "Accretion / (dilution) %",
                           "Accretion / (diluizione) %")
    ad_row = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"={col}{pf_eps_row}/{col}{std_eps_row}-1")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # v0.8 US-250: Cross-over / breakeven synergy (reverse-solve).
    # Additional pre-tax synergy needed Y1 to make deal EPS-neutral:
    #   gap_ni = (std_eps − pf_eps) × (acq_shares + new_shares)
    #   gap_pretax = gap_ni / (1 − tax)
    # If gap_pretax > 0 we're dilutive → need that much more synergy.
    # If negative, already accretive → min synergy needed is zero.
    ws.cell(row=r, column=1,
            value="Cross-over / breakeven synergy analysis"
            ).font = styles.font_subheader
    ws.cell(row=r, column=2,
            value="Analisi sinergia di breakeven").font = styles.font_label_it
    r += 1
    layout.write_row_label(ws, r,
                           "Additional breakeven synergy (EPS-neutral, pre-tax)",
                           "Sinergia breakeven addizionale (pre-tax)")
    bk_row = r
    for i in range(p):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=(f"=MAX(({col}{std_eps_row}-{col}{pf_eps_row})*"
                           f"(acq_shares_m+new_shares_issued)/"
                           f"(1-effective_tax_rate),0)"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Reverse-solve: min additional pre-tax synergy that lifts pro-forma "
        "EPS to match acquirer standalone EPS. Zero if already accretive. "
        "Uses current pro-forma EPS (which already includes existing synergy "
        "assumptions) so this is the MARGINAL gap.",
        "ModelForge",
    )
    r += 1
    layout.write_row_label(ws, r,
                           "Cross-over year (first EPS-neutral year)",
                           "Anno di cross-over (primo anno EPS-neutrale)")
    # MATCH first positive accretion: returns year index 1..p, or "not in window"
    ad_start_col = layout.year_col(0)
    ad_end_col = layout.year_col(p - 1)
    c = ws.cell(
        row=r, column=4,
        value=(f'=IFERROR(MATCH(TRUE,INDEX('
               f'{ad_start_col}{ad_row}:{ad_end_col}{ad_row}>=0,0),0),'
               f'"beyond Y{p}")'),
    )
    styles.style_formula(c, number_format=styles.FMT_INTEGER)
    r += 2

    # v0.8 US-251: Contribution analysis (acquirer vs target contribution
    # to combined revenue / EBITDA / net income vs post-deal equity share).
    ws.cell(row=r, column=1,
            value="Contribution analysis (acquirer vs target)"
            ).font = styles.font_subheader
    ws.cell(row=r, column=2,
            value="Analisi di contribuzione").font = styles.font_label_it
    r += 1

    # Header: Acquirer | Target | Equity %
    ws.cell(row=r, column=3, value="Acquirer").font = styles.font_subheader
    ws.cell(row=r, column=4, value="Target").font = styles.font_subheader
    ws.cell(row=r, column=5,
            value="Acquirer equity % (post-deal)").font = styles.font_subheader
    r += 1

    contribution_rows = [
        ("Revenue contribution %", "Contribuzione ricavi",
         "=acq_revenue_fy0/(acq_revenue_fy0+tgt_revenue_fy0)",
         "=tgt_revenue_fy0/(acq_revenue_fy0+tgt_revenue_fy0)"),
        ("EBITDA contribution %", "Contribuzione EBITDA",
         "=(acq_revenue_fy0*acq_ebitda_margin)/"
         "(acq_revenue_fy0*acq_ebitda_margin+tgt_revenue_fy0*tgt_ebitda_margin)",
         "=(tgt_revenue_fy0*tgt_ebitda_margin)/"
         "(acq_revenue_fy0*acq_ebitda_margin+tgt_revenue_fy0*tgt_ebitda_margin)"),
        ("Net income contribution %", "Contribuzione utile netto",
         "=acq_net_income_fy0/(acq_net_income_fy0+tgt_net_income_fy0)",
         "=tgt_net_income_fy0/(acq_net_income_fy0+tgt_net_income_fy0)"),
    ]
    equity_acq_formula = "=acq_shares_m/(acq_shares_m+new_shares_issued)"
    for en, it, acq_f, tgt_f in contribution_rows:
        layout.write_row_label(ws, r, en, it)
        ca = ws.cell(row=r, column=3, value=acq_f)
        styles.style_formula(ca, number_format=styles.FMT_PCT_2DP)
        ct = ws.cell(row=r, column=4, value=tgt_f)
        styles.style_formula(ct, number_format=styles.FMT_PCT_2DP)
        ce = ws.cell(row=r, column=5, value=equity_acq_formula)
        styles.style_formula(ce, number_format=styles.FMT_PCT_2DP)
        r += 1
    layout.write_row_label(ws, r, "Equity ownership post-deal",
                           "Quota equity post-operazione")
    ws.cell(row=r, column=3, value=equity_acq_formula)
    styles.style_formula(ws.cell(row=r, column=3),
                         number_format=styles.FMT_PCT_2DP)
    ws.cell(row=r, column=4,
            value="=new_shares_issued/(acq_shares_m+new_shares_issued)")
    styles.style_formula(ws.cell(row=r, column=4),
                         number_format=styles.FMT_PCT_2DP)
    ws.cell(row=r, column=4).comment = Comment(
        "Contribution analysis compares each party's pre-deal share of "
        "combined revenue / EBITDA / NI against acquirer's equity ownership "
        "post-deal. Large gaps flag fairness concerns (e.g. acquirer paying "
        "up if equity % << combined-NI contribution).",
        "ModelForge",
    )
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
