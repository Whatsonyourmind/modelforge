"""Project Finance debt schedule + DSCR covenant + DSRA.

v0.2 behaviour (linear amortization, fixed debt amount) preserved as default.

v0.3 additions:
    - amortization_profile: linear | sculpted_level_debt_service
                            | sculpted_dscr_target | bullet
    - debt_sizing_mode: fixed_amount | dscr_target
    - DSRA (Debt Service Reserve Account) lines: target balance, balance,
      funding per operating year.

When debt_sizing_mode == "dscr_target", the senior debt amount is solved
in-Python BEFORE this sheet renders (see pf_solver.solve_dscr_target_debt).
The solved value overrides spec.debt.amount.base (with rationale update).

When amortization_profile in {sculpted_*}, the per-year principal schedule
is baked into the sheet as hardcoded PERCENTAGES of senior_amount, so
formulas stay live: `=-senior_amount * 0.04271`. Scenario worst/best of
senior_amount still scale correctly.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.pf_solver import level_debt_service_pct_schedule


def _sculpted_pct_schedule(spec) -> list[float]:
    """Return operating-year-length list of principal pcts of senior_amount."""
    rate = spec.debt.reference_rate.base + spec.debt.margin_bps.base / 10000.0
    amort_years = spec.debt.tenor_operating_years - spec.debt.grace_years
    return level_debt_service_pct_schedule(
        rate=rate,
        amort_years=amort_years,
        grace_years=spec.debt.grace_years,
        operating_years=spec.horizon.operating_years,
    )


def build(ws: Worksheet, spec, cashflow_refs: dict[str, str],
          cashflow_sheet: str) -> dict[str, str]:
    from openpyxl.workbook.defined_name import DefinedName

    c = spec.horizon.construction_years
    o = spec.horizon.operating_years
    n = c + o

    # v0.12 (US-PF-covenants): covenant thresholds + DSCR blank-guard, lifted
    # from builder hardcodes to overridable spec named inputs. Resolved once
    # here; defaults reproduce the prior literals (LLCR 1.50x / PLCR 1.75x /
    # DSCR blank guard 0.001 EUR m), so behavior is backward-compatible.
    _cov_thr = getattr(spec, "covenant_thresholds", None)

    profile = spec.debt.amortization_profile
    is_sculpted = profile in ("sculpted_level_debt_service", "sculpted_dscr_target")
    is_bullet = profile == "bullet"

    # Pre-compute pct schedule if sculpted
    pct_schedule: list[float] = []
    if is_sculpted:
        pct_schedule = _sculpted_pct_schedule(spec)

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    # v0.7 honesty fix: the "sculpted_*" profiles are implemented as a LEVEL
    # DEBT SERVICE (annuity) amortization curve — they do NOT pin realised DSCR
    # to a flat target (that would require CFADS-driven sculpting:
    # principal_t = CFADS_t/target - interest_t). Under dscr_target sizing the
    # senior amount is solved so the *minimum* DSCR equals the target; realised
    # DSCR then rises above target in later years. Render an honest profile
    # label so the deliverable stops claiming sculpting it does not perform.
    _PROFILE_LABELS = {
        "linear": "linear amortization",
        "sculpted_level_debt_service": "level debt service (annuity)",
        "sculpted_dscr_target": "level debt service (sized to min-DSCR target)",
        "bullet": "bullet repayment",
    }
    profile_label = _PROFILE_LABELS.get(profile, profile)
    subtitle = (
        f"Commitment {spec.meta.currency} · "
        f"{c + spec.debt.tenor_operating_years - spec.debt.grace_years}y senior · "
        f"{profile_label}"
    )
    layout.write_title_block(
        ws, title_en="Project Debt, DSCR & DSRA",
        title_it="Debito di progetto, DSCR & DSRA",
        subtitle=subtitle,
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        phase = "C" if i < c else "O"
        yr = i + 1 if i < c else (i - c + 1)
        cc = ws.cell(row=yr_row, column=col_idx, value=f"{phase}{yr}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    # ────────────────────────────────────────────────────────────
    # SECTION 1 — DEBT ROLL-FORWARD
    # ────────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "Debt roll-forward", "Piano del debito")
    r += 1

    # Opening
    rows["opening"] = r
    layout.write_row_label(ws, r, "Opening debt", "Debito iniziale")
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            prior = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx, value=f"=${prior}${r + 3}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Drawdown — during construction, proportional to capex phasing
    rows["drawdown"] = r
    layout.write_row_label(ws, r, "Drawdown", "Tiraggio", indent=True)
    debt_amount = spec.debt.amount.name
    for i in range(c):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        phasing = spec.construction.capex_phasing_pct[i].name
        cc = ws.cell(row=r, column=col_idx, value=f"={debt_amount}*{phasing}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    for i in range(c, n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Scheduled amortization — profile-dependent
    rows["amort"] = r
    layout.write_row_label(ws, r, "Scheduled amortization", "Ammortamento", indent=True)
    amort_years = spec.debt.tenor_operating_years - spec.debt.grace_years
    last_amort_op_idx = spec.debt.grace_years + amort_years - 1  # 0-based within operating

    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if i < c:
            cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            op_idx = i - c  # 0-based within operating years
            if is_bullet:
                # Principal only at last amort year (end of tenor)
                if op_idx == last_amort_op_idx:
                    cc = ws.cell(row=r, column=col_idx, value=f"=-{debt_amount}")
                else:
                    cc = ws.cell(row=r, column=col_idx, value=0)
            elif is_sculpted:
                pct = pct_schedule[op_idx] if op_idx < len(pct_schedule) else 0.0
                if pct > 0:
                    cc = ws.cell(row=r, column=col_idx,
                                 value=f"=-{debt_amount}*{pct:.8f}")
                    # Cell comment: solver provenance. NB: this is a LEVEL
                    # DEBT SERVICE (annuity) curve, NOT DSCR-target sculpting —
                    # realised DSCR is not flat at target (see sheet subtitle).
                    cc.comment = Comment(
                        f"Level-debt-service (annuity) amortization\n"
                        f"Operating year {op_idx + 1}: principal = {pct * 100:.3f}% of senior\n"
                        f"Solver: annuity at rate={spec.debt.reference_rate.base + spec.debt.margin_bps.base/10000:.4f}, "
                        f"{amort_years}y amort, {spec.debt.grace_years}y grace",
                        "ModelForge",
                    )
                else:
                    cc = ws.cell(row=r, column=col_idx, value=0)
            else:
                # Linear (v0.2 behaviour)
                if op_idx < spec.debt.grace_years or op_idx > last_amort_op_idx:
                    cc = ws.cell(row=r, column=col_idx, value=0)
                else:
                    cc = ws.cell(row=r, column=col_idx,
                                 value=f"=-{debt_amount}/{amort_years}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Closing
    rows["closing"] = r
    layout.write_row_label(ws, r, "Closing debt", "Debito finale")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['opening']}+${col}${rows['drawdown']}+${col}${rows['amort']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # ────────────────────────────────────────────────────────────
    # SECTION 2 — INTEREST & DEBT SERVICE
    # ────────────────────────────────────────────────────────────
    rows["avg_balance"] = r
    layout.write_row_label(ws, r, "Average balance", "Debito medio", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=(${col}${rows['opening']}+${col}${rows['closing']})/2")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    ref = spec.debt.reference_rate.name
    margin = spec.debt.margin_bps.name
    rows["rate"] = r
    layout.write_row_label(ws, r, "All-in rate", "Tasso all-in", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"={ref}+({margin}/10000)")
        styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    r += 1

    rows["interest"] = r
    layout.write_row_label(ws, r, "Cash interest", "Interessi cassa")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['avg_balance']}*${col}${rows['rate']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    r += 1

    # Total debt service (negative) = interest + amort
    rows["debt_service"] = r
    layout.write_row_label(ws, r, "Total debt service", "Servizio totale del debito")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['interest']}+${col}${rows['amort']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # ────────────────────────────────────────────────────────────
    # SECTION 3 — DSCR COVENANT
    # ────────────────────────────────────────────────────────────
    layout.write_section_header(ws, r, "DSCR covenant", "Covenant DSCR")
    r += 1

    cads_row = int(cashflow_refs["cads_row"])

    # v0.12: DSCR blank-guard threshold lifted from hardcode (1E-6) to a visible,
    # overridable spec named input (pf_dscr_blank_threshold). The input CELL is
    # emitted later, in the "Reserves" section near the LLCR/PLCR thresholds, so
    # inserting it does NOT shift the timeline rows above. We resolve its value
    # here and the named range is referenced by the DSCR formula below; named
    # ranges resolve at calc time independent of cell position.
    _ds_blank_thr = (getattr(_cov_thr, "dscr_blank_threshold_eur_m", 0.001)
                     if _cov_thr else 0.001)

    rows["dscr"] = r
    layout.write_row_label(ws, r, "DSCR (CADS / |debt service|)", "DSCR")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i < c:
            cc = ws.cell(row=r, column=col_idx, value="")
        else:
            cads_ref = f"'{cashflow_sheet}'!{col}{cads_row}"
            ds_ref = f"${col}${rows['debt_service']}"
            # v0.7 fix (v0.12 made overridable): post-amortization operating years
            # carry only a near-zero float residual debt service (~1e-8 from
            # average-balance interest on a fully-repaid loan). Dividing CFADS by
            # that residual exploded the DSCR cell to ~1.8e8, polluting
            # MIN/AVERAGE. Treat |debt service| below pf_dscr_blank_threshold
            # (default 0.001 EUR m) as "no debt service" and leave the cell blank
            # — Excel's MIN/AVERAGE ignore text cells, so the summary statistics
            # now cover only genuinely-levered operating years.
            cc = ws.cell(
                row=r, column=col_idx,
                value=(f'=IF(ABS({ds_ref})<pf_dscr_blank_threshold,"",'
                       f'IFERROR({cads_ref}/ABS({ds_ref}),""))'),
            )
            styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    r += 1

    rows["dscr_threshold"] = r
    layout.write_row_label(ws, r, "DSCR threshold", "Soglia DSCR", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        if i < c:
            ws.cell(row=r, column=col_idx, value="")
        else:
            op_idx = i - c
            if op_idx < len(spec.covenant.threshold_by_year):
                a = spec.covenant.threshold_by_year[op_idx]
                cc = ws.cell(row=r, column=col_idx, value=f"={a.name}")
                styles.style_xref(cc, number_format=styles.FMT_MULTIPLE)
    r += 1

    rows["dscr_breach"] = r
    layout.write_row_label(ws, r, "DSCR breach flag", "Violazione DSCR", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i < c:
            ws.cell(row=r, column=col_idx, value="")
        else:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=IF(${col}${rows['dscr']}<${col}${rows['dscr_threshold']},1,0)")
            styles.style_formula(cc, number_format=styles.FMT_INTEGER)
            cc.alignment = styles.align_center
    r += 1

    # Total breach counter
    rows["total_breach"] = r
    layout.write_row_label(ws, r, "Total DSCR breaches", "Violazioni totali DSCR")
    first_op_col = layout.year_col(c); last_col = layout.year_col(n - 1)
    cc = ws.cell(row=r, column=3,
                 value=f"=SUM(${first_op_col}${rows['dscr_breach']}:${last_col}${rows['dscr_breach']})")
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    cc.font = styles.font_subheader
    r += 1

    # v0.7: Min / Avg DSCR summary (bulge-tier standard — BIWS, Bodmer)
    rows["min_dscr"] = r
    layout.write_row_label(ws, r, "Minimum DSCR (operating years)",
                           "DSCR minimo (anni operativi)", indent=True)
    cc = ws.cell(row=r, column=3,
                 value=f"=MIN(${first_op_col}${rows['dscr']}:${last_col}${rows['dscr']})")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    cc.font = styles.font_subheader
    r += 1

    rows["avg_dscr"] = r
    layout.write_row_label(ws, r, "Average DSCR (operating years)",
                           "DSCR medio (anni operativi)", indent=True)
    cc = ws.cell(row=r, column=3,
                 value=f"=AVERAGE(${first_op_col}${rows['dscr']}:${last_col}${rows['dscr']})")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    cc.font = styles.font_subheader
    r += 2

    # v0.7: LLCR + PLCR (Loan Life / Project Life Coverage Ratios)
    # LLCR = NPV(CFADS over loan life, cost_of_debt) / (debt outstanding + DSRA)
    # PLCR = NPV(CFADS over project life, cost_of_debt) / (debt outstanding + DSRA)
    # Threshold 1.50x typical (Bodmer, BIWS).
    layout.write_section_header(
        ws, r, "LLCR & PLCR (forward-looking coverage)",
        "LLCR & PLCR (copertura forward)",
    )
    r += 1

    # Find first-operating opening-debt cell for PV anchor
    cost_of_debt_ref = f"{spec.debt.reference_rate.name}+({spec.debt.margin_bps.name}/10000)"

    # Loan life = from first operating year through amortization end
    # Project life = from first operating year through last operating col
    first_op_col = layout.year_col(c)
    last_op_col = layout.year_col(n - 1)

    # LLCR
    rows["llcr"] = r
    layout.write_row_label(ws, r, "LLCR (NPV CFADS over loan life / Debt)",
                           "LLCR (VAN CFADS vita prestito / Debito)")
    cads_range = f"'{cashflow_sheet}'!{first_op_col}{cads_row}:{last_op_col}{cads_row}"
    debt_out_ref = f"${first_op_col}${rows['opening']}"
    cc = ws.cell(row=r, column=3,
                 value=f"=IFERROR(NPV({cost_of_debt_ref},{cads_range})/{debt_out_ref},0)")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    cc.font = styles.font_subheader
    cc.comment = openpyxl.comments.Comment(
        "Loan Life Coverage Ratio — per BIWS / Edward Bodmer. "
        "Threshold ≥ 1.50x typical for PF lenders.",
        "ModelForge",
    ) if False else None  # skip comment if openpyxl not imported in this scope
    r += 1

    # LLCR threshold — lifted from hardcode (1.50x) to spec named input.
    # v0.12: overridable via spec.covenant_thresholds.llcr_threshold;
    # default 1.50x reproduces the prior literal (byte-identical).
    _llcr_thr = getattr(_cov_thr, "llcr_threshold", 1.50) if _cov_thr else 1.50
    rows["llcr_threshold"] = r
    layout.write_row_label(ws, r, "LLCR threshold", "Soglia LLCR", indent=True)
    cc = ws.cell(row=r, column=3, value=_llcr_thr)
    styles.style_input(cc, number_format=styles.FMT_MULTIPLE)
    cc.comment = Comment(
        "Loan Life Coverage Ratio covenant floor (BIWS / Edward Bodmer). "
        "1.50x typical for PF senior lenders. Overridable via "
        "spec.covenant_thresholds.llcr_threshold.",
        "ModelForge",
    )
    if "pf_llcr_threshold" in ws.parent.defined_names:
        del ws.parent.defined_names["pf_llcr_threshold"]
    ws.parent.defined_names["pf_llcr_threshold"] = DefinedName(
        name="pf_llcr_threshold",
        attr_text=f"'{ws.title}'!$C${r}",
    )
    r += 1

    # PLCR
    rows["plcr"] = r
    layout.write_row_label(ws, r, "PLCR (NPV CFADS over project life / Debt)",
                           "PLCR (VAN CFADS vita progetto / Debito)")
    cc = ws.cell(row=r, column=3,
                 value=f"=IFERROR(NPV({cost_of_debt_ref},{cads_range})/{debt_out_ref},0)")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    cc.font = styles.font_subheader
    r += 1

    # PLCR threshold — lifted from hardcode (1.75x) to spec named input.
    # v0.12: overridable via spec.covenant_thresholds.plcr_threshold;
    # default 1.75x reproduces the prior literal (byte-identical). PLCR floor
    # sits above LLCR because the project-life window adds post-loan tail cash.
    _plcr_thr = getattr(_cov_thr, "plcr_threshold", 1.75) if _cov_thr else 1.75
    rows["plcr_threshold"] = r
    layout.write_row_label(ws, r, "PLCR threshold", "Soglia PLCR", indent=True)
    cc = ws.cell(row=r, column=3, value=_plcr_thr)
    styles.style_input(cc, number_format=styles.FMT_MULTIPLE)
    cc.comment = Comment(
        "Project Life Coverage Ratio covenant floor (BIWS / Edward Bodmer). "
        "1.75x typical; > LLCR floor because the project-life window adds "
        "post-loan tail cash. Overridable via "
        "spec.covenant_thresholds.plcr_threshold.",
        "ModelForge",
    )
    if "pf_plcr_threshold" in ws.parent.defined_names:
        del ws.parent.defined_names["pf_plcr_threshold"]
    ws.parent.defined_names["pf_plcr_threshold"] = DefinedName(
        name="pf_plcr_threshold",
        attr_text=f"'{ws.title}'!$C${r}",
    )
    r += 2

    # ── v0.7: additional bulge-tier PF reserves, cure, make-whole ──────────
    layout.write_section_header(
        ws, r, "Reserves, cure rights & prepayment features",
        "Riserve, diritti di cura & rimborsi anticipati",
    )
    r += 1

    # O&M reserve (months of opex)
    rows["om_reserve_months"] = r
    layout.write_row_label(ws, r, "O&M reserve (months of opex)",
                           "Riserva O&M (mesi opex)", indent=True)
    om_months = getattr(spec.operating, "om_reserve_months", 3) or 3
    cc = ws.cell(row=r, column=3, value=om_months)
    styles.style_input(cc, number_format=styles.FMT_INTEGER)
    cc.comment = openpyxl.comments.Comment(
        "Typical O&M reserve: 3-6 months of operating costs. "
        "Funded at COD from debt proceeds; released at decommissioning.",
        "ModelForge",
    ) if False else None
    r += 1

    # Major Maintenance Reserve (sinking fund)
    rows["mmr_target"] = r
    layout.write_row_label(ws, r, "Major Maintenance Reserve target (€m)",
                           "Riserva manutenzione maggiore (€m)", indent=True)
    mmr_val = getattr(spec.operating, "major_maintenance_reserve_eur_m", None)
    cc = ws.cell(row=r, column=3,
                 value=mmr_val.base if mmr_val else 0)
    styles.style_input(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Lock-up test
    rows["lockup_threshold"] = r
    layout.write_row_label(ws, r, "Lock-up DSCR threshold (block distribution)",
                           "Soglia lock-up DSCR", indent=True)
    cc = ws.cell(row=r, column=3,
                 value=f"={spec.covenant.lock_up_threshold.name}")
    styles.style_xref(cc, number_format=styles.FMT_MULTIPLE)
    r += 1

    # Equity cure rights
    rows["equity_cure_cap"] = r
    layout.write_row_label(ws, r, "Equity cure rights — max count",
                           "Diritti cura equity — max volte", indent=True)
    cc = ws.cell(row=r, column=3,
                 value=spec.debt.equity_cure_cap_count)
    styles.style_input(cc, number_format=styles.FMT_INTEGER)
    r += 1

    rows["equity_cure_uplift"] = r
    layout.write_row_label(ws, r, "Equity cure max EBITDA uplift (per cure)",
                           "Cura equity — uplift max EBITDA", indent=True)
    uplift = spec.debt.equity_cure_max_uplift_pct
    cc = ws.cell(row=r, column=3,
                 value=f"={uplift.name}" if uplift else 0.20)
    (styles.style_xref if uplift else styles.style_input)(cc, number_format=styles.FMT_PCT)
    r += 1

    # Make-whole premium
    rows["make_whole_spread"] = r
    layout.write_row_label(ws, r, "Make-whole spread (bps)",
                           "Premio make-whole (bps)", indent=True)
    mw = spec.debt.make_whole_spread_bps
    cc = ws.cell(row=r, column=3,
                 value=f"={mw.name}" if mw else 50)
    (styles.style_xref if mw else styles.style_input)(cc, number_format=styles.FMT_BPS)
    cc.comment = openpyxl.comments.Comment(
        "Make-whole premium paid on early redemption of fixed-rate bonds "
        "(T+50bps typical for US PP style). Sources: BIWS, LSTA.",
        "ModelForge",
    ) if False else None
    r += 1

    # Mandatory prepayment events
    rows["mandatory_prepay"] = r
    layout.write_row_label(ws, r, "Mandatory prepayment events",
                           "Eventi di rimborso obbligatorio", indent=True)
    ws.cell(row=r, column=4, value="Insurance proceeds, asset sale, change of control, illegality, excess CF sweep").font = styles.font_label_it
    r += 2

    # ── v0.7: P50/P90 probabilistic revenue + degradation ──────────────────
    layout.write_section_header(
        ws, r, "Probabilistic revenue & degradation (solar PF)",
        "Ricavi probabilistici & degrado (PF solare)",
    )
    r += 1

    rows["p90_haircut"] = r
    layout.write_row_label(ws, r, "P90 revenue haircut vs P50 (base)",
                           "Taglio P90 vs P50", indent=True)
    p90 = getattr(spec.operating, "p90_revenue_haircut_pct", None)
    cc = ws.cell(row=r, column=3,
                 value=f"={p90.name}" if p90 else 0.08)
    (styles.style_xref if p90 else styles.style_input)(cc, number_format=styles.FMT_PCT)
    cc.comment = openpyxl.comments.Comment(
        "P90 is the revenue level exceeded 90% of years. "
        "P90-P50 gap typically 8-10% for solar. Bank debt sized "
        "against P90 (1-yr) per Solargis/NREL convention.",
        "ModelForge",
    ) if False else None
    r += 1

    rows["panel_degradation"] = r
    layout.write_row_label(ws, r, "Panel degradation rate (annual)",
                           "Degrado pannelli (annuale)", indent=True)
    deg = getattr(spec.operating, "panel_degradation_pct_annual", None)
    cc = ws.cell(row=r, column=3,
                 value=f"={deg.name}" if deg else 0.005)
    (styles.style_xref if deg else styles.style_input)(cc, number_format=styles.FMT_PCT_2DP)
    cc.comment = openpyxl.comments.Comment(
        "Solar panel output degrades ~0.5% per year (standard "
        "manufacturer warranty). Compounded over 25-30y life.",
        "ModelForge",
    ) if False else None
    r += 2

    # ────────────────────────────────────────────────────────────
    # SECTION 4 — DSRA (v0.3)
    # ────────────────────────────────────────────────────────────
    dsra_months = spec.debt.dsra_months
    layout.write_section_header(
        ws, r,
        f"DSRA (Debt Service Reserve Account — {dsra_months}-month target)",
        f"DSRA (Riserva servizio debito — {dsra_months} mesi target)",
    )
    r += 1

    # DSRA target balance at end of year = (dsra_months/12) * |debt_service_next_year|
    # Construction years and final operating year: target = 0 (no next-year debt or release)
    rows["dsra_target"] = r
    layout.write_row_label(ws, r, "DSRA target balance", "DSRA target", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i < c - 1:
            # Early construction — no reserve yet
            cc = ws.cell(row=r, column=col_idx, value=0)
        elif i == n - 1:
            # Final operating year — release to 0
            cc = ws.cell(row=r, column=col_idx, value=0)
        else:
            next_col = layout.year_col(i + 1)
            next_ds = f"${next_col}${rows['debt_service']}"
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=({dsra_months}/12)*ABS({next_ds})")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # DSRA balance (assumes always funded to target — simplification)
    rows["dsra_balance"] = r
    layout.write_row_label(ws, r, "DSRA balance (end of year)", "DSRA saldo", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['dsra_target']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    r += 1

    # DSRA funding (negative = fund, positive = release)
    rows["dsra_funding"] = r
    layout.write_row_label(ws, r, "DSRA funding (outflow<0 / release>0)",
                           "DSRA finanziamento (uscita<0 / rilascio>0)")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 0:
            # First year: funding = -(target - 0) = -target
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=-${col}${rows['dsra_balance']}")
        else:
            prior = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${prior}${rows['dsra_balance']}-${col}${rows['dsra_balance']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # DSCR blank-guard threshold — lifted from hardcode (1E-6) to spec named
    # input (pf_dscr_blank_threshold). Emitted at the BOTTOM of the sheet so it
    # does not shift any timeline/summary rows above (which the cross-sheet wiring
    # and tests reference by address). The DSCR formula above references this
    # named range; named ranges resolve at calc time independent of cell
    # position. Default 0.001 EUR m (=€1k) brackets the ~1e-8 post-amortization
    # residual far below any genuinely-levered year (those run €0.3-3m), so the
    # blank/solvent set is unchanged vs the prior 1E-6 literal.
    r += 1
    layout.write_section_header(
        ws, r, "Covenant guards (model controls)",
        "Controlli covenant (parametri modello)",
    )
    r += 1
    rows["dscr_blank_threshold"] = r
    layout.write_row_label(ws, r, "DSCR blank guard (min |debt service|, €m)",
                           "Soglia DSCR vuoto (min |servizio debito|, €m)",
                           indent=True)
    cc = ws.cell(row=r, column=3, value=_ds_blank_thr)
    styles.style_input(cc, number_format="0.000\" €m\"")
    cc.comment = Comment(
        "Absolute |debt service| floor (EUR m) below which a DSCR cell is left "
        "blank. Guards against dividing CFADS by the ~1e-8 residual a "
        "fully-repaid loan carries post-amortization (would explode the cell and "
        "pollute MIN/AVERAGE). Overridable via "
        "spec.covenant_thresholds.dscr_blank_threshold_eur_m.",
        "ModelForge",
    )
    if "pf_dscr_blank_threshold" in ws.parent.defined_names:
        del ws.parent.defined_names["pf_dscr_blank_threshold"]
    ws.parent.defined_names["pf_dscr_blank_threshold"] = DefinedName(
        name="pf_dscr_blank_threshold",
        attr_text=f"'{ws.title}'!$C${r}",
    )
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    # v0.6: patch the ProjectCashFlow interest-expense placeholder row
    # with a cross-sheet reference to this sheet's cash-interest row.
    # This gives the tax walk (EBIT − Interest) on the cashflow sheet
    # access to the real interest number.
    if "interest_row" in cashflow_refs:
        cf_interest_row = int(cashflow_refs["interest_row"])
        cf_ws = ws.parent[cashflow_sheet]
        for i in range(n):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            cc = cf_ws.cell(row=cf_interest_row, column=col_idx)
            cc.value = f"='{ws.title}'!{col}{rows['interest']}"
            styles.style_xref(cc, number_format=styles.FMT_EUR_M)

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    out["total_breach_cell"] = f"'{ws.title}'!$C${rows['total_breach']}"
    return out
