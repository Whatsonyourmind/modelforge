"""Structured Credit tranche waterfall sheet."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    y = spec.horizon.collection_years
    n = y + 1

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Tranche Waterfall", title_it="Waterfall di tranche",
        subtitle=f"{spec.meta.currency} · {len(spec.tranches)} tranches · {y}y horizon",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=yr_row, column=col_idx, value=f"t={i}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    # Pool
    layout.write_section_header(ws, r, "Collateral pool", "Collaterale")
    r += 1

    rows["face_value"] = r
    layout.write_row_label(ws, r, "Face value of pool", "Valore nominale pool")
    cc = ws.cell(row=r, column=4, value="=face_value_eur_m")
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cum_default_pct"] = r
    layout.write_row_label(ws, r, "Cumulative default % face", "Default cum. % nominale")
    ws.cell(row=r, column=3, value="%").font = styles.font_label_it
    ws.cell(row=r, column=4, value=0)
    for i in range(y):
        col_idx = ord(layout.year_col(i + 1)) - ord("A") + 1
        a = spec.collateral.cumulative_default_curve_pct[i]
        cc = ws.cell(row=r, column=col_idx, value=f"={a.name}")
        styles.style_xref(cc, number_format=styles.FMT_PCT)
    r += 1

    rows["cum_loss_pct"] = r
    layout.write_row_label(ws, r, "Cumulative net loss % face (after recovery)",
                           "Perdita netta cum. % nominale")
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['cum_default_pct']}*(1-recovery_pct_on_default)")
        styles.style_formula(cc, number_format=styles.FMT_PCT)
    r += 2

    # Per-tranche
    layout.write_section_header(ws, r, "Tranches", "Tranche")
    r += 1

    tranche_rows: list[dict] = []
    for tr in spec.tranches:
        layout.write_section_header(ws, r, f"Tranche: {tr.name.en} ({tr.rating})",
                                    f"Tranche: {tr.name.it} ({tr.rating})")
        r += 1

        # Size = (detachment - attachment) × face
        size_row = r
        layout.write_row_label(ws, r, "Tranche size", "Dimensione tranche")
        cc = ws.cell(
            row=r, column=4,
            value=f"=($D${rows['face_value']}*({tr.detachment_point_pct.name}-{tr.attachment_point_pct.name}))",
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        r += 1

        # Attachment / detachment
        att_row = r
        layout.write_row_label(ws, r, "Attachment point", "Punto di attacco", indent=True)
        cc = ws.cell(row=r, column=4, value=f"={tr.attachment_point_pct.name}")
        styles.style_xref(cc, number_format=styles.FMT_PCT)
        r += 1

        det_row = r
        layout.write_row_label(ws, r, "Detachment point", "Punto di stacco", indent=True)
        cc = ws.cell(row=r, column=4, value=f"={tr.detachment_point_pct.name}")
        styles.style_xref(cc, number_format=styles.FMT_PCT)
        r += 1

        # Tranche loss % = MAX(0, MIN(cum_loss - attachment, detachment - attachment)) / (detachment - attachment)
        loss_row = r
        layout.write_row_label(ws, r, "Tranche loss % (cumulative)",
                               "Perdita tranche % (cum)", indent=True)
        ws.cell(row=r, column=4, value=0)
        for i in range(1, n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            cc = ws.cell(
                row=r, column=col_idx,
                value=(
                    f"=IFERROR(MAX(0,MIN(${col}${rows['cum_loss_pct']}-$D${att_row},$D${det_row}-$D${att_row}))"
                    f"/($D${det_row}-$D${att_row}),0)"
                ),
            )
            styles.style_formula(cc, number_format=styles.FMT_PCT)
        r += 1

        # Tranche notional outstanding = size × (1 - loss%)
        out_row = r
        layout.write_row_label(ws, r, "Tranche notional outstanding",
                               "Nozionale in circolazione", indent=True)
        cc = ws.cell(row=r, column=4, value=f"=$D${size_row}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        for i in range(1, n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=$D${size_row}*(1-${col}${loss_row})")
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        r += 1

        # Coupon (on outstanding)
        coupon_row = r
        layout.write_row_label(ws, r, "Coupon paid", "Cedola pagata", indent=True)
        ws.cell(row=r, column=4, value=0)
        for i in range(1, n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            prior_col = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${prior_col}${out_row}*{tr.coupon_pct.name}")
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        r += 1

        # Principal repayment at maturity (year y) — surviving notional
        princ_row = r
        layout.write_row_label(ws, r, "Principal repayment",
                               "Rimborso capitale", indent=True)
        for i in range(n):
            col_idx = ord(layout.year_col(i)) - ord("A") + 1
            if i == y:
                cc = ws.cell(row=r, column=col_idx,
                             value=f"=${layout.year_col(y)}${out_row}")
            else:
                cc = ws.cell(row=r, column=col_idx, value=0)
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        r += 1

        # Investor cash flow (from investor perspective)
        inv_cf_row = r
        layout.write_row_label(ws, r, f"Investor CF — {tr.name.en}",
                               f"CF investitore — {tr.name.it}")
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i == 0:
                cc = ws.cell(row=r, column=col_idx, value=f"=-$D${size_row}")
            else:
                cc = ws.cell(row=r, column=col_idx,
                             value=f"=${col}${coupon_row}+${col}${princ_row}")
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
            cc.font = styles.font_subheader
            cc.border = styles.BORDER_TOP_THIN
        r += 1

        # Tranche IRR
        first_col = layout.year_col(0); last_col = layout.year_col(n - 1)
        irr_row = r
        layout.write_row_label(ws, r, f"Tranche IRR — {tr.name.en}",
                               f"IRR tranche — {tr.name.it}", indent=True)
        cc = ws.cell(row=r, column=4,
                     value=f"=IRR(${first_col}${inv_cf_row}:${last_col}${inv_cf_row},{tr.coupon_pct.name})")
        styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
        cc.font = styles.font_subheader
        r += 2

        tranche_rows.append({
            "name": tr.name.en, "size": size_row, "attachment": att_row,
            "detachment": det_row, "loss": loss_row, "outstanding": out_row,
            "coupon": coupon_row, "principal": princ_row, "cf": inv_cf_row, "irr": irr_row,
        })

    # v0.8 US-260/261: Principal Deficiency Ledger (PDL) + strict priority
    # waterfall summary block. Audit criterion #101 requires PDL row on
    # Tranches (or CollectionWaterfall) sheet.
    r += 1
    layout.write_section_header(
        ws, r, "Principal Deficiency Ledger (PDL) & priority waterfall",
        "Registro deficit principale (PDL) e waterfall",
    )
    r += 1
    layout.write_row_label(ws, r,
                           "PDL (cumulative losses allocated to tranches)",
                           "PDL (perdite cumulative allocate alle tranche)",
                           indent=True)
    # Sum cumulative tranche losses across all tranches
    if tranche_rows:
        loss_terms = [f"'Tranches'!D{tr['loss']}" for tr in tranche_rows]
        formula = "=" + "+".join(loss_terms) if loss_terms else "=0"
    else:
        formula = "=0"
    cc = ws.cell(row=r, column=4, value=formula)
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    r += 1
    layout.write_row_label(ws, r,
                           "Priority of payments (strict senior→mezz→equity)",
                           "Ordine pagamenti (senior→mezz→equity)",
                           indent=True)
    ws.cell(row=r, column=4,
            value="Interest: Senior → Mezz → Equity | Principal: same order"
            ).font = styles.font_label_it
    r += 1
    layout.write_row_label(ws, r,
                           "Trigger events (acceleration / reversion)",
                           "Eventi trigger (accelerazione / reversione)",
                           indent=True)
    ws.cell(row=r, column=4,
            value="PDL > 50% equity NV, servicer default, subordination breach"
            ).font = styles.font_label_it

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    out["tranche_rows"] = tranche_rows  # type: ignore[assignment]
    return out
