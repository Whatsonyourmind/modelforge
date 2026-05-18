"""Cover sheet.

Bulge-tier front page. Contains:
    - Project code, target, sector, deliverable
    - Analyst, valuation date, version, status
    - Scenario toggle (named range `scenario_index`, 1/2/3)
    - Active scenario label (formula CHOOSE(scenario_index, ...))
    - Currency, unit scale, sign convention
    - Revision log
    - Sheet index with hyperlinks

The scenario toggle is the single point of control for the whole model.
"""

from __future__ import annotations

from openpyxl.workbook.defined_name import DefinedName

from modelforge.builder import styles, layout
from modelforge.builder.i18n import L
from modelforge.graph.schema import LinkageGraph, GraphNode, NodeKind


def build(ws, spec, graph: LinkageGraph) -> None:
    layout.set_column_widths(ws, label_width=34, it_width=30, year_width=22)

    layout.write_title_block(
        ws,
        title_en=f"{spec.target.name} — {spec.meta.deliverable.en}",
        title_it=spec.meta.deliverable.secondary,
        subtitle=f"{spec.meta.confidentiality} · {spec.meta.status.upper()} · {spec.meta.version}",
    )

    # ── Metadata block (rows 4..)
    rows = [
        ("project_code", spec.meta.project_code, None),
        ("analyst", spec.meta.analyst, None),
        ("valuation_date", spec.meta.valuation_date.isoformat(), styles.FMT_DATE),
        ("version", spec.meta.version, None),
        ("status", spec.meta.status, None),
        ("currency", spec.meta.currency, None),
        ("unit_scale", spec.meta.unit_scale, None),
        ("sign_convention", spec.meta.sign_convention, None),
        ("target", spec.target.name, None),
        ("sector", spec.target.sector.en, None),
        ("country", spec.target.country, None),
    ]
    r = 4
    for key, value, nfmt in rows:
        lbl = L(key)
        layout.write_row_label(ws, r, lbl.en, lbl.secondary)
        c = ws.cell(row=r, column=3, value=value)
        styles.style_input(c, number_format=nfmt or "General")
        r += 1

    # ── Scenario toggle block
    r += 1
    layout.write_section_header(ws, r, "Scenario control", "Controllo scenari")
    r += 1

    lbl_active = L("scenario_active")
    layout.write_row_label(ws, r, lbl_active.en, lbl_active.secondary)
    # Toggle cell: integer 1/2/3
    toggle_cell = ws.cell(row=r, column=3, value=2)  # default = BASE
    styles.style_input(toggle_cell, number_format=styles.FMT_INTEGER)
    toggle_cell.comment = None
    toggle_addr = f"'{ws.title}'!$C${r}"

    # Named range: scenario_index
    _register_name(ws.parent, "scenario_index", toggle_addr)

    # Graph: register the toggle as a driver.
    drv_id = LinkageGraph.driver_id("scenario_index")
    graph.add_node(
        GraphNode(
            id=drv_id,
            kind=NodeKind.DRIVER,
            label="scenario_index",
            payload={"unit": "index", "scope": "global"},
        )
    )
    cell_id = LinkageGraph.cell_id(ws.title, f"C{r}")
    graph.add_node(GraphNode(id=cell_id, kind=NodeKind.CELL, label="scenario_index"))
    graph.lives_on(drv_id, cell_id)

    r += 1
    # Resolved scenario label
    from modelforge.builder.formulas import scenario_pick
    lbl_pair = L("scenario")
    layout.write_row_label(ws, r, f"{lbl_pair.en} (resolved)", f"{lbl_pair.secondary} (attivo)")
    worst_txt = f'"{L("scenario_worst").en}"'
    base_txt = f'"{L("scenario_base").en}"'
    best_txt = f'"{L("scenario_best").en}"'
    resolved_cell = ws.cell(row=r, column=3, value=f"=CHOOSE(scenario_index,{worst_txt},{base_txt},{best_txt})")
    styles.style_formula(resolved_cell, number_format="General")
    resolved_cell.alignment = styles.align_center
    resolved_cell.font = styles.font_subheader
    r += 2

    # ── Revision log
    layout.write_section_header(ws, r, L("revision_log").en, L("revision_log").secondary)
    r += 1
    header_cols = ["Version", "Date", "Analyst", "Note"]
    for i, h in enumerate(header_cols):
        c = ws.cell(row=r, column=1 + i, value=h)
        styles.style_header(c)
    r += 1
    for entry in spec.meta.revision_log:
        ws.cell(row=r, column=1, value=entry.version).font = styles.font_label_en
        d = ws.cell(row=r, column=2, value=entry.date.isoformat())
        styles.style_input(d, number_format=styles.FMT_DATE)
        ws.cell(row=r, column=3, value=entry.analyst).font = styles.font_label_en
        ws.cell(row=r, column=4, value=entry.note).font = styles.font_label_en
        r += 1

    r += 2

    # ── Sheet index (hyperlinks to each tab)
    layout.write_section_header(ws, r, "Sheet index", "Indice fogli")
    r += 1
    sheet_order = ["Cover", "Sources", "Assumptions", "OperatingModel",
                   "DebtSchedule", "Covenants", "Returns", "QC"]
    for name in sheet_order:
        if name == "Cover":
            continue
        c = ws.cell(row=r, column=1, value=name)
        c.hyperlink = f"#'{name}'!A1"
        c.font = styles.Font(
            name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
            color="0563C1", underline="single",
        )
        r += 1

    # Print area
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.fitToWidth = 1
    ws.print_title_rows = "1:2"


def _register_name(wb, name: str, addr: str) -> None:
    """Register/replace a workbook-level defined name."""
    if name in wb.defined_names:
        del wb.defined_names[name]
    dn = DefinedName(name=name, attr_text=addr)
    wb.defined_names[name] = dn
