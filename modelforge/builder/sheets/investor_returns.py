"""Investor Returns sheet (Minibond Template 2).

Gross-to-net view from the bondholder's perspective:
    Gross cash flow (coupon + principal at maturity/amort)
    – withholding tax
    – transaction costs at entry
    = Net cash flow
    → Gross YTM, Net YTM (after-tax), MoIC
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.i18n import L

# Bold key-metric font that ALSO carries an explicit colour, so a bold summary
# cell still satisfies the certify styling gate (which requires both an explicit
# font colour and a number_format). styles.font_subheader is bold but colourless.
_FONT_METRIC_BOLD = Font(
    name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY, bold=True,
    color=styles.COLOR_FORMULA,
)


def build(ws: Worksheet, spec, bond_refs: dict[str, str], bond_sheet_name: str) -> None:
    h = spec.horizon.historical_years
    p = spec.horizon.projection_years

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Investor Returns", title_it="Rendimento investitore",
        subtitle="Gross YTM · Net YTM (after Italian withholding) · IFRS 9 EIR",
    )
    layout.write_scenario_banner(ws, row=3)

    # Periods 0..p (annual)
    yr_row = 5
    ws.cell(row=yr_row, column=3, value="Period").font = styles.font_subheader
    for i in range(p + 1):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=yr_row, column=col_idx, value=f"t={i}")
        styles.style_header(c)

    r = 7
    rows: dict[str, int] = {}
    interest_row = int(bond_refs["interest_row"])
    amort_row = int(bond_refs["amort_row"])

    layout.write_section_header(ws, r, "Gross bondholder cash flow", "CF lordo al bondholder")
    r += 1
    rows["gross_cf"] = r
    layout.write_row_label(ws, r, "Gross cash flow", "CF lordo")

    # t=0: -notional + transaction cost (negative, cost to investor)
    notional = spec.bond.notional.name
    tx_cost = spec.bond.arrangement_fee_pct.name  # investor-side we use transaction_cost_bps
    tx_cost_inv = spec.investor_adjustments.transaction_cost_bps.name
    c0 = ws.cell(row=r, column=4,
                 value=f"=-{notional}-({notional}*{tx_cost_inv}/10000)")
    styles.style_formula(c0, number_format=styles.FMT_EUR_M)

    for i in range(1, p + 1):
        debt_col = layout.year_col(h + (i - 1))
        interest_ref = f"'{bond_sheet_name}'!{debt_col}{interest_row}"
        amort_ref = f"'{bond_sheet_name}'!{debt_col}{amort_row}"
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-{interest_ref}-{amort_ref}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 2

    # Withholding tax (on interest only, principal is return of capital)
    layout.write_section_header(ws, r, "Net bondholder cash flow (after WHT)",
                                "CF netto al bondholder (al netto WHT)")
    r += 1
    rows["wht"] = r
    layout.write_row_label(ws, r, "Withholding tax on coupon", "Ritenuta d'acconto su cedola", indent=True)
    c0 = ws.cell(row=r, column=4, value=0)
    styles.style_formula(c0, number_format=styles.FMT_EUR_M)
    for i in range(1, p + 1):
        debt_col = layout.year_col(h + (i - 1))
        interest_ref = f"'{bond_sheet_name}'!{debt_col}{interest_row}"
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        # withholding = positive_coupon * wht_rate, which reduces investor CF
        c = ws.cell(row=r, column=col_idx,
                    value=f"=-(-{interest_ref})*withholding_tax_pct")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_cf"] = r
    layout.write_row_label(ws, r, "Net cash flow", "CF netto")
    for i in range(p + 1):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        c = ws.cell(row=r, column=col_idx,
                    value=f"=${col}${rows['gross_cf']}+${col}${rows['wht']}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = styles.font_subheader
        c.border = styles.BORDER_TOP_THIN
    r += 2

    # Summary metrics
    layout.write_section_header(ws, r, "Return metrics", "Metriche di rendimento")
    r += 1
    first_col = layout.year_col(0); last_col = layout.year_col(p)

    layout.write_row_label(ws, r, "Gross YTM", "YTM lordo", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=IRR(${first_col}${rows['gross_cf']}:${last_col}${rows['gross_cf']},0.05)")
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    c.font = styles.font_subheader
    r += 1

    layout.write_row_label(ws, r, "Net YTM (after WHT)", "YTM netto (post WHT)", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=IRR(${first_col}${rows['net_cf']}:${last_col}${rows['net_cf']},0.04)")
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    c.font = styles.font_subheader
    r += 1

    layout.write_row_label(ws, r, "EIR (IFRS 9, incl. fees)",
                           "EIR (IFRS 9, incl. commissioni)", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=IRR(${first_col}${rows['gross_cf']}:${last_col}${rows['gross_cf']},0.05)")
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    c.font = styles.font_subheader
    c.comment = Comment(
        "IFRS 9 §B5.4.1 — EIR includes all fees that are an integral part "
        "of the instrument. Principal and coupon cash flows on the gross CF "
        "line already include the upfront transaction cost.",
        "ModelForge",
    )
    r += 1

    layout.write_row_label(ws, r, "MoIC (gross)", "MoIC (lordo)", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=IFERROR(SUMIF(${first_col}${rows['gross_cf']}:${last_col}${rows['gross_cf']},\">0\")"
                      f"/ABS(SUMIF(${first_col}${rows['gross_cf']}:${last_col}${rows['gross_cf']},\"<0\")),0)")
    styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    c.font = styles.font_subheader
    r += 2

    # ── Fixed-income analytics (Macaulay/modified duration + issuer cost) ──
    #
    # SCOPE: rendered only when the bond schedule exposes a coupon rate +
    # net-proceeds row (i.e. the minibond render path). `bond_refs` from the
    # unitranche `returns` sheet does NOT carry these keys, so the block is a
    # no-op there and that template is unaffected.
    coupon_rate_ref = None
    if getattr(spec, "bond", None) is not None and getattr(spec.bond, "coupon", None) is not None:
        if spec.bond.coupon.kind == "fixed":
            coupon_rate_ref = spec.bond.coupon.fixed_rate.name
    has_issuer_proceeds = "net_proceeds_row" in bond_refs and "amort_row" in bond_refs

    if coupon_rate_ref is not None and has_issuer_proceeds:
        layout.write_section_header(ws, r, "Fixed-income analytics", "Analitiche obbligazionarie")
        r += 1

        # Discount-rate convention: duration is reported at the PAR yield, which
        # for a bond priced at par equals the coupon (the textbook anchor). The
        # gross CF row already holds the par inflows (coupon + principal) at
        # t=1..p; t=0 is the −notional outlay and is excluded from the weights.
        # Helper rows make the PV-weighting auditable cell-by-cell.

        # Period index t (0..p). Pure integers — blue inputs (a reviewer may
        # see they are the period axis, not a formula).
        period_row = r
        layout.write_row_label(ws, r, "Period (t)", "Periodo (t)", indent=True)
        for i in range(p + 1):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            cc = ws.cell(row=r, column=col_idx, value=i)
            styles.style_input(cc, number_format=styles.FMT_INTEGER)
        r += 1

        # PV of each inflow at the coupon (par) yield: CF_t / (1+coupon)^t.
        # t=0 is forced to 0 (the −notional outlay is not an inflow weight).
        pvcf_row = r
        layout.write_row_label(ws, r, "PV of inflow @ coupon yield",
                               "VA del flusso @ cedola", indent=True)
        for i in range(p + 1):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i == 0:
                cc = ws.cell(row=r, column=col_idx, value=0)
            else:
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=(f"=${col}${rows['gross_cf']}"
                           f"/(1+{coupon_rate_ref})^${col}${period_row}"),
                )
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        r += 1

        # Price = PV of all inflows (== face at par). Denominator for duration.
        price_row = r
        layout.write_row_label(ws, r, "Price (Σ PV of inflows)",
                               "Prezzo (Σ VA flussi)", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=SUM(${first_col}${pvcf_row}:${last_col}${pvcf_row})")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # Macaulay duration = Σ(t · PV_t) / Σ PV_t  — PV-weighted average life.
        mac_row = r
        layout.write_row_label(ws, r, "Macaulay duration", "Duration di Macaulay", indent=True)
        c = ws.cell(
            row=r, column=4,
            value=(f"=SUMPRODUCT(${first_col}${period_row}:${last_col}${period_row},"
                   f"${first_col}${pvcf_row}:${last_col}${pvcf_row})/$D${price_row}"),
        )
        styles.style_formula(c, number_format=styles.FMT_YEARS)
        c.font = _FONT_METRIC_BOLD
        c.comment = Comment(
            "Macaulay duration — PV-weighted average time to the coupon+"
            "principal inflows, discounted at the par (coupon) yield. "
            "Σ(t·PV_t)/Σ(PV_t).",
            "ModelForge",
        )
        r += 1

        # Modified duration = Macaulay / (1 + coupon yield). dP/P ≈ −ModDur·dy.
        layout.write_row_label(ws, r, "Modified duration", "Duration modificata", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${mac_row}/(1+{coupon_rate_ref})")
        styles.style_formula(c, number_format=styles.FMT_YEARS)
        c.font = _FONT_METRIC_BOLD
        c.comment = Comment(
            "Modified duration = Macaulay / (1 + ytm). First-order price "
            "sensitivity: a +1% yield move moves price by ≈ −ModDur%.",
            "ModelForge",
        )
        r += 2

        # ── Issuer all-in cost of debt ──────────────────────────────────
        # The issuer receives NET proceeds (face − arrangement − legal −
        # listing − rating) but services the FULL coupon + principal, so its
        # all-in cost is the IRR of [net proceeds; −(coupon+principal) …].
        # With positive upfront fees this lands ABOVE the coupon.
        layout.write_section_header(ws, r, "Issuer all-in cost of debt",
                                    "Costo all-in del debito (emittente)")
        r += 1

        # Issuer cashflow: t=0 net proceeds IN (xref BondStructure); each later
        # period the coupon + scheduled amort OUT. BondStructure stores both
        # interest and amort as NEGATIVE, so their sum is already the issuer's
        # signed outflow.
        issuer_cf_row = r
        np_row = int(bond_refs["net_proceeds_row"])
        int_row = int(bond_refs["interest_row"])
        amort_row_b = int(bond_refs["amort_row"])
        layout.write_row_label(ws, r, "Issuer cash flow", "Flusso di cassa emittente", indent=True)
        for i in range(p + 1):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i == 0:
                cc = ws.cell(row=r, column=col_idx,
                             value=f"='{bond_sheet_name}'!$D${np_row}")
            else:
                debt_col = layout.year_col(h + (i - 1))
                # tenor years only carry coupon/principal; beyond maturity the
                # BondStructure rows are 0, so summing the full t=1..p span is safe.
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=(f"='{bond_sheet_name}'!{debt_col}{int_row}"
                           f"+'{bond_sheet_name}'!{debt_col}{amort_row_b}"),
                )
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        r += 1

        layout.write_row_label(ws, r, "Issuer all-in cost (IRR)",
                               "Costo all-in emittente (IRR)", indent=True)
        c = ws.cell(
            row=r, column=4,
            value=(f"=IRR(${first_col}${issuer_cf_row}:${last_col}${issuer_cf_row},"
                   f"{coupon_rate_ref})"),
        )
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        c.font = _FONT_METRIC_BOLD
        c.comment = Comment(
            "Issuer all-in cost of funds — IRR of the issuer cashflow: net "
            "proceeds received at close (face less arrangement/legal/listing/"
            "rating fees) against the coupon + principal it services. With "
            "upfront fees this exceeds the headline coupon.",
            "ModelForge",
        )
        r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"
