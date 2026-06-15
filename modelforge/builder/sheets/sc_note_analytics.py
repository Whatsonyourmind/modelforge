"""Note analytics sheet — Template 19.

Per-note investor metrics derived from the liability waterfall: weighted-average
life (WAL), principal / interest received over the deal life, any writedown,
investor IRR, the credit-enhancement available at close and an indicative
rating proxy mapped from that enhancement (downgraded to "Impaired" if the note
takes a principal loss).

The principal-received time series (one row per note) is materialised here so
WAL is a transparent SUMPRODUCT(period, principal) / SUM(principal); the
remaining metrics SUM the corresponding waterfall rows directly.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, wf_refs: dict[str, str], wf_sheet: str) -> dict[str, str]:
    periods = spec.horizon.periods
    n = periods + 1
    cur = spec.meta.currency
    notes = spec.notes
    senior = int(wf_refs["senior_idx"])
    eq = int(wf_refs["eq_idx"])
    debt = list(range(len(notes) - 1))
    last_col = layout.year_col(periods)

    def wf(col, key):
        return f"'{wf_sheet}'!${col}${wf_refs[key]}"

    layout.set_column_widths(ws, label_width=30, it_width=20, year_width=11, unit_width=5)
    for ci in range(4, 4 + 10):  # widen metric columns
        ws.column_dimensions[chr(ord("A") + ci - 1)].width = 12
    layout.write_title_block(
        ws, title_en="Note Analytics (WAL / IRR / rating)",
        title_it="Analisi note (WAL / IRR / rating)",
        subtitle=f"{cur} · per-note investor metrics · {len(debt)} debt + residual",
    )
    layout.write_scenario_banner(ws, row=3)
    yr_row = 5
    for i in range(n):
        ci = ord(layout.year_col(i)) - ord("A") + 1
        styles.style_header(ws.cell(row=yr_row, column=ci, value=f"t={i}"))

    def _ci(i):
        return ord(layout.year_col(i)) - ord("A") + 1

    rows: dict[str, int] = {}
    r = 7

    # ── Period index + principal-received time series (for WAL) ────────────
    layout.write_section_header(ws, r, "Principal received by note (per period)",
                                "Capitale ricevuto per nota")
    r += 1
    period_row = r
    layout.write_row_label(ws, r, "Period index (t)", "Indice periodo (t)", indent=True)
    for i in range(n):
        c = ws.cell(row=r, column=_ci(i), value=i)
        styles.style_input(c, number_format=styles.FMT_INTEGER)
    r += 1

    prin_rows: list[int] = []
    for j in range(len(notes)):
        layout.write_row_label(ws, r, f"Principal — {notes[j].name.en}",
                               f"Capitale — {notes[j].name.secondary}", indent=True)
        ws.cell(row=r, column=3, value=cur).font = styles.font_label_it
        for i in range(n):
            col = layout.year_col(i)
            if i == 0:
                c = ws.cell(row=r, column=_ci(i), value=0)
            elif j == eq:
                c = ws.cell(row=r, column=_ci(i), value=f"={wf(col, 'prin_to_residual_row')}")
            else:
                # debt note principal = collections paydown + turbo + maturity reserve cure
                c = ws.cell(row=r, column=_ci(i),
                            value=(f"={wf(col, f'prin_paid_{j}_row')}"
                                   f"+{wf(col, f'turbo_{j}_row')}+{wf(col, f'reserve_cure_{j}_row')}"))
            styles.style_xref(c, number_format=styles.FMT_EUR_M)
        prin_rows.append(r)
        r += 1
    r += 1

    # ── Per-note analytics table (one row per note, metrics across columns) ─
    layout.write_section_header(ws, r, "Per-note analytics", "Analisi per nota")
    r += 1
    hdr_row = r
    metric_cols = {
        "rating": 4, "initial": 5, "coupon": 6, "ce": 7, "wal": 8,
        "prin_life": 9, "int_life": 10, "loss": 11, "irr": 12, "proxy": 13,
    }
    titles = {
        "rating": "Rating", "initial": "Initial", "coupon": "Coupon",
        "ce": "CE @close", "wal": "WAL (y)", "prin_life": "Prin recv",
        "int_life": "Int recv", "loss": "Loss", "irr": "IRR", "proxy": "Rating proxy",
    }
    ws.cell(row=hdr_row, column=1, value="Note").font = styles.font_subheader
    for k, ci in metric_cols.items():
        c = ws.cell(row=hdr_row, column=ci, value=titles[k])
        c.font = styles.font_subheader
        c.fill = styles.fill_subheader
        c.alignment = styles.align_center
    r += 1

    def adv(j):
        return notes[j].advance_pct.name

    for j in range(len(notes)):
        rr = r
        layout.write_row_label(ws, rr, notes[j].name.en, notes[j].name.secondary)
        prow = prin_rows[j]
        # Rating (input from spec)
        ws.cell(row=rr, column=metric_cols["rating"], value=notes[j].rating).font = styles.font_static
        # Initial balance = advance × initial pool (LoanTape initial_pool via Waterfall xref base)
        c = ws.cell(row=rr, column=metric_cols["initial"],
                    value=f"={wf('D', f'eop_{j}_row')}")  # t=0 EOP == initial balance
        styles.style_xref(c, number_format=styles.FMT_EUR_M)
        # Coupon
        c = ws.cell(row=rr, column=metric_cols["coupon"], value=f"={notes[j].coupon_pct.name}")
        styles.style_xref(c, number_format=styles.FMT_PCT_2DP)
        # Credit enhancement at close = Σ advance of notes junior to j (% of pool)
        juniors = [adv(k) for k in range(j + 1, len(notes))]
        ce_formula = "=" + ("+".join(juniors) if juniors else "0")
        c = ws.cell(row=rr, column=metric_cols["ce"], value=ce_formula)
        styles.style_formula(c, number_format=styles.FMT_PCT)
        # WAL = SUMPRODUCT(period, principal) / SUM(principal)
        pr = f"$D${period_row}:${last_col}${period_row}"
        kr = f"$D${prow}:${last_col}${prow}"
        c = ws.cell(row=rr, column=metric_cols["wal"],
                    value=f"=IFERROR(SUMPRODUCT({pr},{kr})/SUM({kr}),0)")
        styles.style_formula(c, number_format=styles.FMT_YEARS)
        # Principal received (life)
        c = ws.cell(row=rr, column=metric_cols["prin_life"], value=f"=SUM({kr})")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        # Interest received (life)
        if j == eq:
            # Residual interest received = excess spread NET of SPV tax (spv_tax is
            # negative). The reserve return is RETURN OF CAPITAL, not interest, so
            # it is excluded here (it appears in the total-cash / IRR via inv_cf).
            int_terms = (f"SUM('{wf_sheet}'!$D${wf_refs['residual_int_to_equity_row']}:"
                         f"${last_col}${wf_refs['residual_int_to_equity_row']})"
                         f"+SUM('{wf_sheet}'!$D${wf_refs['spv_tax_row']}:"
                         f"${last_col}${wf_refs['spv_tax_row']})")
        else:
            int_terms = (f"SUM('{wf_sheet}'!$D${wf_refs[f'int_paid_{j}_row']}:"
                         f"${last_col}${wf_refs[f'int_paid_{j}_row']})")
        c = ws.cell(row=rr, column=metric_cols["int_life"], value=f"={int_terms}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        # Loss / writedown = MAX(initial − principal received, 0)
        c = ws.cell(row=rr, column=metric_cols["loss"],
                    value=(f"=MAX(${chr(ord('A')+metric_cols['initial']-1)}{rr}"
                           f"-${chr(ord('A')+metric_cols['prin_life']-1)}{rr},0)"))
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        # IRR over the note's investor-CF row
        icf = f"'{wf_sheet}'!$D${wf_refs[f'inv_cf_{j}_row']}:${last_col}${wf_refs[f'inv_cf_{j}_row']}"
        guess = notes[j].coupon_pct.base if notes[j].coupon_pct.base > 0 else 0.12
        c = ws.cell(row=rr, column=metric_cols["irr"], value=f"=IFERROR(IRR({icf},{guess}),-1)")
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        # Rating proxy from CE (downgraded to Impaired if loss > 0)
        ce_cell = f"${chr(ord('A')+metric_cols['ce']-1)}{rr}"
        loss_cell = f"${chr(ord('A')+metric_cols['loss']-1)}{rr}"
        if j == eq:
            # The first-loss residual is designed to absorb principal losses —
            # a writedown is it doing its job, not an impairment. Always label
            # it "First-loss" (no Impaired override).
            proxy_formula = '="First-loss"'
        else:
            # Subordination-implied bucket (NOT a rating opinion). Calibrated
            # conservatively for granular non-prime collateral so a thin
            # mezzanine is not flattered: AAA needs ~20% CE, AA ~14%, A ~9%.
            proxy = (f'IF({ce_cell}>=0.20,"AAA-ind",IF({ce_cell}>=0.14,"AA-ind",'
                     f'IF({ce_cell}>=0.09,"A-ind",IF({ce_cell}>=0.05,"BBB-ind",'
                     f'IF({ce_cell}>=0.03,"BB-ind","B-ind")))))')
            proxy_formula = f'=IF({loss_cell}>0.01,"Impaired",{proxy})'
        c = ws.cell(row=rr, column=metric_cols["proxy"], value=proxy_formula)
        c.font = styles.font_static
        rows[f"note_{j}"] = rr
        r += 1
    r += 1

    # ── Deal headline (single scalars in col D for sensitivity / readers) ──
    layout.write_section_header(ws, r, "Deal headline", "Sintesi operazione")
    r += 1
    irr_col = chr(ord("A") + metric_cols["irr"] - 1)
    wal_col = chr(ord("A") + metric_cols["wal"] - 1)
    rows["residual_irr"] = r
    layout.write_row_label(ws, r, "Residual IRR", "IRR residuale")
    c = ws.cell(row=r, column=4, value=f"=${irr_col}${rows[f'note_{eq}']}")
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    c.font = styles.font_subheader
    r += 1
    rows["senior_wal"] = r
    layout.write_row_label(ws, r, "Senior WAL (y)", "WAL senior", indent=True)
    c = ws.cell(row=r, column=4, value=f"=${wal_col}${rows[f'note_{senior}']}")
    styles.style_formula(c, number_format=styles.FMT_YEARS)
    r += 1

    rows["period"] = period_row
    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"
    return {f"{k}_row": str(v) for k, v in rows.items()}
