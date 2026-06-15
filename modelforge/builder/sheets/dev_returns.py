"""Development RE returns + European promote waterfall sheet.

Reads the DevSchedule sheet's unlevered CF, equity CF, debt balance and exit
rows and renders:
    - Unlevered project IRR / NPV / MOIC
    - Levered equity IRR / NPV / MOIC
    - Equity invested (Σ equity contributions), peak senior debt, exit value
    - A European whole-fund LP/GP promote waterfall (pref → catch-up → 80/20
      residual) on the total equity contribution / distribution.

Every output cell is a LIVE Excel formula (IRR / NPV / SUMIF / MAX / MIN) so
the workbook recalculates end-to-end. IRR/NPV are taken over the contiguous
annual period row on DevSchedule — the proven-certifiable idiom shared with
pf_returns / re_financing.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, sched_refs: dict[str, str], sched_sheet: str) -> dict[str, str]:
    first_col = sched_refs["first_col"]
    last_col = sched_refs["last_col"]
    n = int(sched_refs["n_periods"])
    n_years_hold = n - 1

    unlev_row = int(sched_refs["unlevered_cf_row"])
    equity_row = int(sched_refs["equity_cf_row"])
    debt_close_row = int(sched_refs["debt_close_row"])
    net_exit_row = int(sched_refs["net_exit_row"])

    layout.set_column_widths(ws, label_width=48, it_width=36, year_width=12, unit_width=6)
    layout.write_title_block(
        ws, title_en="Development Returns & Promote",
        title_it="Rendimenti di sviluppo & promote",
        subtitle="Unlevered & levered IRR / MOIC / NPV + European whole-fund waterfall",
    )
    layout.write_scenario_banner(ws, row=3)

    def _sched_range(row: int) -> str:
        return f"'{sched_sheet}'!{first_col}{row}:'{sched_sheet}'!{last_col}{row}"

    # Contiguous range on the schedule sheet (single sheet qualifier form for IRR)
    def _rng(row: int) -> str:
        return f"'{sched_sheet}'!${first_col}${row}:${last_col}${row}"

    rows: dict[str, int] = {}
    r = 7

    # ── Unlevered project returns ──────────────────────────────────────────
    layout.write_section_header(ws, r, "Unlevered project returns", "Rendimenti progetto (unlevered)")
    r += 1

    rows["unlev_irr"] = r
    layout.write_row_label(ws, r, "Unlevered IRR (annual)", "IRR unlevered (annuale)", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=IRR({_rng(unlev_row)},0.10)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.font = styles.font_subheader
    r += 1

    rows["unlev_npv"] = r
    layout.write_row_label(ws, r, "Unlevered NPV @ discount rate", "VAN unlevered", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"='{sched_sheet}'!${first_col}${unlev_row}"
               f"+NPV({spec.discount_rate.name},"
               f"'{sched_sheet}'!{layout.year_col(1)}{unlev_row}:"
               f"{last_col}{unlev_row})"),
    )
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["unlev_moic"] = r
    layout.write_row_label(ws, r, "Unlevered MOIC", "MOIC unlevered", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"=IFERROR(SUMIF({_rng(unlev_row)},\">0\")"
               f"/ABS(SUMIF({_rng(unlev_row)},\"<0\")),0)"),
    )
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    r += 2

    # ── Levered equity returns ─────────────────────────────────────────────
    layout.write_section_header(ws, r, "Levered equity returns", "Rendimenti equity (levered)")
    r += 1

    rows["equity_irr"] = r
    layout.write_row_label(ws, r, "Levered equity IRR (annual)", "IRR equity (annuale)", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=IRR({_rng(equity_row)},0.12)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.font = styles.font_subheader
    r += 1

    rows["equity_npv"] = r
    layout.write_row_label(ws, r, "Levered equity NPV @ discount rate", "VAN equity", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"='{sched_sheet}'!${first_col}${equity_row}"
               f"+NPV({spec.discount_rate.name},"
               f"'{sched_sheet}'!{layout.year_col(1)}{equity_row}:"
               f"{last_col}{equity_row})"),
    )
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["equity_moic"] = r
    layout.write_row_label(ws, r, "Levered equity MOIC", "MOIC equity", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"=IFERROR(SUMIF({_rng(equity_row)},\">0\")"
               f"/ABS(SUMIF({_rng(equity_row)},\"<0\")),0)"),
    )
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    cc.font = styles.font_subheader
    r += 2

    # ── Headline capital metrics ───────────────────────────────────────────
    layout.write_section_header(ws, r, "Capital summary", "Riepilogo capitale")
    r += 1

    rows["equity_invested"] = r
    layout.write_row_label(ws, r, "Total equity invested", "Equity totale investito", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=f"=ABS(SUMIF({_rng(equity_row)},\"<0\"))",
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["peak_debt"] = r
    layout.write_row_label(ws, r, "Peak senior debt", "Debito senior di picco", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=MAX({_rng(debt_close_row)})")
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["exit_value"] = r
    layout.write_row_label(ws, r, "Net exit proceeds", "Proventi netti di uscita", indent=True)
    cc = ws.cell(row=r, column=4, value=f"='{sched_sheet}'!$D${net_exit_row}")
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # ── European whole-fund promote waterfall ──────────────────────────────
    layout.write_section_header(
        ws, r, "European whole-fund waterfall (pref + catch-up + promote)",
        "Waterfall whole-fund europeo (pref + catch-up + promote)",
    )
    r += 1
    ws.cell(
        row=r, column=1,
        value=("Whole-fund European waterfall: LP pref compounded on total "
               "equity, then GP catch-up to the promote threshold, then an "
               "80/20 LP/GP split on the residual. Pref / promote are live "
               "spec-driven named inputs."),
    ).font = styles.font_label_it
    r += 1

    lp_pct = spec.waterfall.lp_capital_commitment_pct.name

    rows["total_contrib"] = r
    layout.write_row_label(ws, r, "Total equity contribution", "Capitale equity totale", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=$D${rows['equity_invested']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_distrib"] = r
    layout.write_row_label(ws, r, "Total equity distributions (gross)",
                           "Distribuzioni equity (lorde)", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=SUMIF({_rng(equity_row)},\">0\")")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["total_profit"] = r
    layout.write_row_label(ws, r, "Total profit", "Profitto totale", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['total_distrib']}-$D${rows['total_contrib']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    r += 2

    # Pref + promote knobs (live spec inputs)
    rows["pref_return"] = r
    layout.write_row_label(ws, r, "LP preferred return (compounded)",
                           "Rendimento preferenziale LP", indent=True)
    cc = ws.cell(row=r, column=4, value="=lp_preferred_return_pct")
    styles.style_input(cc, number_format=styles.FMT_PCT_2DP)
    cc.comment = Comment(
        "LP preferred return, compounded annually on contributed capital before "
        "any GP promote. Spec-driven named input "
        "(spec.waterfall.lp_preferred_return_pct, resolved from the pref tier).",
        "ModelForge",
    )
    r += 1

    rows["promote_lp_share"] = r
    layout.write_row_label(ws, r, "LP share of residual (= 1 − GP promote)",
                           "Quota LP residuo", indent=True)
    cc = ws.cell(row=r, column=4, value="=1-gp_promote_pct")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.comment = Comment(
        "LP share of the residual band = 1 − gp_promote_pct. GP promote is the "
        "spec-driven named input (spec.waterfall.gp_promote_pct). 80/20 promote "
        "→ GP 20%, LP 80%.",
        "ModelForge",
    )
    r += 2

    # Tier 1 — LP pref threshold (capital × (1+pref)^hold) — whole-fund
    ws.cell(row=r, column=1, value="Tier 1: Return of capital + pref").font = styles.font_subheader
    r += 1
    rows["pref_threshold"] = r
    layout.write_row_label(ws, r, "LP pref threshold (capital × (1+pref)^hold)",
                           "Soglia pref LP", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"=$D${rows['total_contrib']}*{lp_pct}"
               f"*(1+$D${rows['pref_return']})^{n_years_hold}"),
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["lp_tier1"] = r
    layout.write_row_label(ws, r, "LP receives in Tier 1", "LP riceve in Tier 1", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=f"=MIN($D${rows['total_distrib']},$D${rows['pref_threshold']})",
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Tier 2 — GP catch-up (to promote threshold)
    ws.cell(row=r, column=1, value="Tier 2: GP catch-up").font = styles.font_subheader
    r += 1
    rows["gp_catchup"] = r
    layout.write_row_label(ws, r, "GP catch-up amount", "Catch-up GP", indent=True)
    # catch-up = max(0, profit above pref) × (1−LP_share)/LP_share
    cc = ws.cell(
        row=r, column=4,
        value=(f"=MAX(0,($D${rows['total_profit']}"
               f"-($D${rows['pref_threshold']}-$D${rows['total_contrib']}*{lp_pct}))"
               f")*(1-$D${rows['promote_lp_share']})/$D${rows['promote_lp_share']}"),
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Tier 3 — 80/20 residual split
    ws.cell(row=r, column=1, value="Tier 3: 80/20 residual split").font = styles.font_subheader
    r += 1
    rows["lp_tier3"] = r
    layout.write_row_label(ws, r, "LP share of residual", "Quota LP residuo", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"=MAX(0,$D${rows['total_distrib']}-$D${rows['lp_tier1']}"
               f"-$D${rows['gp_catchup']})*$D${rows['promote_lp_share']}"),
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1
    rows["gp_tier3"] = r
    layout.write_row_label(ws, r, "GP share of residual (promote)",
                           "Quota GP residuo (promote)", indent=True)
    cc = ws.cell(
        row=r, column=4,
        value=(f"=MAX(0,$D${rows['total_distrib']}-$D${rows['lp_tier1']}"
               f"-$D${rows['gp_catchup']})*(1-$D${rows['promote_lp_share']})"),
    )
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # Totals by partner
    ws.cell(row=r, column=1, value="Totals by partner (post-waterfall)").font = styles.font_subheader
    r += 1
    rows["lp_total"] = r
    layout.write_row_label(ws, r, "LP total post-waterfall", "LP totale post-waterfall", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=$D${rows['lp_tier1']}+$D${rows['lp_tier3']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    r += 1
    rows["gp_total"] = r
    layout.write_row_label(ws, r, "GP total post-waterfall", "GP totale post-waterfall", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=$D${rows['gp_catchup']}+$D${rows['gp_tier3']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
