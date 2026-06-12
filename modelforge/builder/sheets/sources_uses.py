"""Sources & Uses + LBO-specific blocks — Sponsor LBO template (v0.8).

Sheet layout:
    1. Sources & Uses — balanced equation (sources = uses)
    2. Purchase price build (offer × FD + net debt + fees)
    3. PPA (goodwill + intangibles + DTL)
    4. Transaction fees split (M&A advisory + financing)
    5. Sponsor capital structure (equity + rollover + MIP)
    6. NWC closing adjustment
    7. Earnout / CVR
    8. Dividend recap
    9. Exit scenarios (×3: strategic / IPO / secondary)
    10. Hurdle analysis (reverse-solve max PP at IRR)
    11. GP promote waterfall (pref + catchup + carry)
    12. Returns summary (IRR / MoIC / CoC per scenario)

References: Macabacus LBO Long; BIWS; Rosenbaum-Pearl Investment Banking.
"""

from __future__ import annotations

from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.defined_name import DefinedName

from modelforge.builder import styles, layout


def _def(wb, name: str, sheet: str, cell: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=f"'{sheet}'!{cell}")


def build(ws: Worksheet, spec, debt_refs: dict | None = None,
          debt_sheet_name: str = "DebtSchedule") -> dict[str, int]:
    """Emit full Sources & Uses + LBO-specific blocks sheet.

    ``debt_refs`` (from the DebtSchedule builder) supplies ``total_closing_row``
    so exit proceeds and the dividend recap can net the debt actually
    outstanding at the relevant year. Passed by the sponsor_lbo template; when
    absent the debt-netting refs degrade to 0 (legacy enterprise-value-only
    behaviour) rather than crashing.
    """
    layout.set_column_widths(ws, label_width=46, it_width=34, year_width=14, unit_width=6)
    layout.write_title_block(
        ws, "Sources & Uses + LBO Structuring",
        "Sources & Uses + Struttura LBO",
        "Balanced S&U, PPA, sponsor capital, exits, hurdle, promote",
    )

    wb = ws.parent
    refs: dict[str, int] = {}
    r = 5

    # ── LBO-2 FIX (v0.11): cross-sheet references to the OperatingModel's
    #    PROJECTED EBITDA. Exit proceeds, the hurdle reverse-solve, and the
    #    dividend recap previously multiplied their exit/recap multiples by
    #    `historical_ebitda_lfy` (entry LTM EBITDA), discarding all modeled
    #    growth (~46 EUR m understated at exit). We instead point at the
    #    projected EBITDA cell for the relevant year on OperatingModel.
    #
    #    OperatingModel layout (operating.build): year columns start at D
    #    (layout.year_col(0)); EBITDA is row 13. Projection year `y` (1-indexed)
    #    lives at column index h + y - 1. We clamp into the projection window.
    OPMODEL_EBITDA_ROW = 13  # operating.build: EBITDA line row (fixed layout)
    _h = spec.horizon.historical_years
    _p = spec.horizon.projection_years

    def _proj_ebitda_ref(proj_year_1indexed: int) -> str:
        """Cross-sheet ref to projected EBITDA for a 1-indexed projection year,
        clamped to the projection window [1.._p]."""
        y = max(1, min(int(proj_year_1indexed), _p))
        col = layout.year_col(_h + y - 1)
        return f"OperatingModel!{col}{OPMODEL_EBITDA_ROW}"

    _exit_year_val = getattr(spec, "exit_year", 5)
    _recap_year_val = getattr(spec, "div_recap_year", 3)
    proj_exit_ebitda_ref = _proj_ebitda_ref(_exit_year_val)
    proj_recap_ebitda_ref = _proj_ebitda_ref(_recap_year_val)

    # LBO-EXIT FIX (audit 2026-06): exit equity to the sponsor = exit EV − net
    # debt OUTSTANDING at exit (the prior code booked the full enterprise value
    # as the sponsor's inflow, never repaying the ~€35m of tranche debt, which
    # overstated strategic IRR by ~6.5pp and MoIC by ~0.8×). Likewise the
    # dividend recap pays (new debt raised to target leverage) − (debt already
    # outstanding at the recap date), NOT minus the sponsor equity cheque. Both
    # net the DebtSchedule 'Total debt outstanding' at the matching year column.
    _total_closing_row = None
    if debt_refs and debt_refs.get("total_closing_row") is not None:
        _total_closing_row = int(debt_refs["total_closing_row"])

    def _proj_debt_ref(proj_year_1indexed: int) -> str:
        """Cross-sheet ref to total debt outstanding at a 1-indexed projection
        year (same column convention as _proj_ebitda_ref). Returns '0' when the
        debt row is unknown, degrading to enterprise-value-only behaviour."""
        if _total_closing_row is None:
            return "0"
        y = max(1, min(int(proj_year_1indexed), _p))
        col = layout.year_col(_h + y - 1)
        return f"'{debt_sheet_name}'!{col}{_total_closing_row}"

    proj_exit_debt_ref = _proj_debt_ref(_exit_year_val)
    proj_recap_debt_ref = _proj_debt_ref(_recap_year_val)

    # ── SECTION 1: Sources & Uses (US-201) ────────────────────────────────
    layout.write_section_header(ws, r, "Sources & Uses",
                                "Fonti e impieghi (S&U)")
    r += 1

    # Aggregate tranche amounts for Sources
    senior_amount = 0.0
    mezz_amount = 0.0
    for t in spec.debt.tranches:
        if t.seniority == "senior":
            senior_amount += t.amount.base
        elif t.seniority in ("mezz", "sub"):
            mezz_amount += t.amount.base

    rcf_amount = 0.0
    if spec.debt.rcf.enabled and spec.debt.rcf.amount is not None:
        rcf_amount = spec.debt.rcf.amount.base

    sponsor_equity = (getattr(spec, "sponsor_equity_eur_m", None)
                      .base if getattr(spec, "sponsor_equity_eur_m", None)
                      else 0.0)
    mgmt_rollover = getattr(spec, "mgmt_rollover_eur_m", 0.0) or 0.0
    ma_fees = getattr(spec, "ma_advisory_fees_eur_m", 0.0) or 0.0
    fin_fees = getattr(spec, "financing_fees_eur_m", 0.0) or 0.0

    # SOURCES
    ws.cell(row=r, column=1, value="SOURCES").font = styles.font_subheader
    ws.cell(row=r, column=2, value="FONTI").font = styles.font_label_it
    r += 1
    refs["sources_senior"] = r
    layout.write_row_label(ws, r, "Senior debt (Term Loan B / unitranche)",
                           "Debito senior (TLB / unitranche)", indent=True)
    c = ws.cell(row=r, column=4, value=senior_amount)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["sources_mezz"] = r
    layout.write_row_label(ws, r, "Mezzanine / subordinated debt",
                           "Debito mezzanino", indent=True)
    c = ws.cell(row=r, column=4, value=mezz_amount)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["sources_rcf"] = r
    layout.write_row_label(ws, r, "Revolver facility (drawn at close)",
                           "Revolver (tirato alla chiusura)", indent=True)
    c = ws.cell(row=r, column=4, value=0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["sources_sponsor_eq"] = r
    layout.write_row_label(ws, r, "Sponsor equity (new money)",
                           "Equity dello sponsor (denaro nuovo)", indent=True)
    c = ws.cell(row=r, column=4, value=sponsor_equity)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["sources_rollover"] = r
    layout.write_row_label(ws, r, "Management rollover equity",
                           "Rollover equity del management", indent=True)
    c = ws.cell(row=r, column=4, value=mgmt_rollover)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["sources_total"] = r
    layout.write_row_label(ws, r, "Total sources", "Totale fonti")
    c = ws.cell(row=r, column=4,
                value=(f"=D{refs['sources_senior']}+D{refs['sources_mezz']}"
                       f"+D{refs['sources_rcf']}+D{refs['sources_sponsor_eq']}"
                       f"+D{refs['sources_rollover']}"))
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.border = styles.BORDER_TOP_THIN
    r += 2

    # USES
    ws.cell(row=r, column=1, value="USES").font = styles.font_subheader
    ws.cell(row=r, column=2, value="IMPIEGHI").font = styles.font_label_it
    r += 1
    refs["uses_equity_pp"] = r
    layout.write_row_label(ws, r, "Purchase equity (target shares × offer px)",
                           "Equity da acquisire", indent=True)
    # LBO-3 FIX (v0.11): the Uses "Purchase equity" line previously computed
    # offer_px × FD shares ONLY (71.5), omitting the option buyout that the
    # canonical Purchase-price-build equity cell (Section 2, pp_equity = 73.5)
    # includes. The two equity figures must agree. We write a placeholder value
    # here and, once Section 2's `pp_equity` cell exists, patch this cell to a
    # LIVE reference to it (see the "LBO-3 patch" below). The placeholder value
    # already includes the option buyout so the workbook is correct even if the
    # patch order ever changes.
    offer_px_assum = getattr(spec, "offer_premium_pct", None)
    target_fd = getattr(spec, "target_fd_shares_m", 0.0) or 0.0
    target_px = getattr(spec, "target_share_price_eur", 0.0) or 0.0
    option_buyout = getattr(spec, "option_buyout_eur_m", 0.0) or 0.0
    if target_fd and target_px and offer_px_assum is not None:
        equity_pp = target_fd * target_px * (1 + offer_px_assum.base) + option_buyout
    else:
        equity_pp = senior_amount + mezz_amount + sponsor_equity + mgmt_rollover \
                    - ma_fees - fin_fees  # fallback: plug
    c = ws.cell(row=r, column=4, value=equity_pp)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["uses_refi_debt"] = r
    layout.write_row_label(ws, r, "Refinance target net debt",
                           "Rifinanziamento debito target", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "target_net_debt_close_eur_m", 0.0) or 0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["uses_ma_fees"] = r
    layout.write_row_label(ws, r, "M&A advisory fees (expensed)",
                           "Spese M&A (spesate)", indent=True)
    c = ws.cell(row=r, column=4, value=ma_fees)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["uses_fin_fees"] = r
    layout.write_row_label(ws, r, "Financing fees (capitalized + amortized)",
                           "Commissioni di finanziamento (capitalizzate)",
                           indent=True)
    c = ws.cell(row=r, column=4, value=fin_fees)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["uses_oid"] = r
    layout.write_row_label(ws, r, "OID discount (original issue)",
                           "Sconto OID", indent=True)
    oid_agg = sum((t.oid_pct.base * t.amount.base for t in spec.debt.tranches
                   if hasattr(t, "oid_pct") and t.oid_pct is not None),
                  start=0.0)
    c = ws.cell(row=r, column=4, value=oid_agg)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["uses_min_cash"] = r
    layout.write_row_label(ws, r, "Minimum cash to balance sheet",
                           "Cassa minima al bilancio", indent=True)
    c = ws.cell(row=r, column=4, value=5.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["uses_total"] = r
    layout.write_row_label(ws, r, "Total uses", "Totale impieghi")
    c = ws.cell(row=r, column=4,
                value=(f"=D{refs['uses_equity_pp']}+D{refs['uses_refi_debt']}"
                       f"+D{refs['uses_ma_fees']}+D{refs['uses_fin_fees']}"
                       f"+D{refs['uses_oid']}+D{refs['uses_min_cash']}"))
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.border = styles.BORDER_TOP_THIN
    r += 1

    # ── LBO-1 FIX (v0.11): sponsor NEW-MONEY equity is the BALANCING PLUG, not a
    #    fixed spec input. Standard LBO S&U: the sponsor writes whatever equity
    #    cheque makes Sources == Uses, given the debt drawn and management
    #    rollover. So:
    #        sponsor_equity = total_uses − (senior + mezz + rcf) − rollover
    #    Written as a live formula on the Sources sponsor-equity cell, this makes
    #    Total sources == Total uses and the S&U check ≈ 0 BY CONSTRUCTION. The
    #    prior code planted the spec's sponsor_equity_eur_m here, leaving a −7.3
    #    imbalance whenever it disagreed with the financing need.
    debt_source_refs = (
        f"D{refs['sources_senior']}+D{refs['sources_mezz']}"
        f"+D{refs['sources_rcf']}"
    )
    plug_cell = ws.cell(
        row=refs["sources_sponsor_eq"], column=4,
        value=(f"=D{refs['uses_total']}-({debt_source_refs})"
               f"-D{refs['sources_rollover']}"),
    )
    styles.style_formula(plug_cell, number_format=styles.FMT_EUR_M)
    plug_cell.comment = Comment(
        "Sponsor new-money equity is the S&U balancing plug: "
        "Total Uses − debt drawn (senior+mezz+RCF) − management rollover. "
        "This guarantees Sources = Uses by construction (standard LBO "
        "structuring).",
        "ModelForge",
    )

    # ── LBO-3 PATCH (v0.11): point the Uses "Purchase equity" cell at the
    #    canonical Purchase-price-build equity cell (Section 2 `pp_equity`,
    #    which includes the option buyout) so the two never diverge. Section 2
    #    is emitted below; `refs['pp_equity']` will exist by then, so we defer
    #    the rewrite to just after Section 2 (see `_patch_uses_equity_pp`).
    def _patch_uses_equity_pp() -> None:
        cc = ws.cell(row=refs["uses_equity_pp"], column=4,
                     value=f"=D{refs['pp_equity']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)

    refs["sau_check"] = r
    layout.write_row_label(ws, r, "S&U check (sources − uses)",
                           "S&U check (fonti − impieghi)")
    c = ws.cell(row=r, column=4,
                value=f"=D{refs['sources_total']}-D{refs['uses_total']}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.comment = Comment(
        "S&U must balance. Non-zero means sponsor equity or debt "
        "quantum is misstated. Default TL/mezz sizing can be inverted "
        "via reverse-solve hurdle block below.",
        "ModelForge",
    )
    r += 2

    # ── SECTION 2: Purchase price build (US-202) ──────────────────────────
    layout.write_section_header(ws, r, "Purchase price build",
                                "Costruzione prezzo di acquisto")
    r += 1
    refs["pp_offer_px"] = r
    layout.write_row_label(ws, r, "Offer price per share",
                           "Prezzo di offerta per azione", indent=True)
    c = ws.cell(row=r, column=4, value=target_px * (1 + (offer_px_assum.base
                                       if offer_px_assum else 0.30)))
    styles.style_input(c, number_format=styles.FMT_EUR_ACTUAL)
    r += 1
    refs["pp_fd_shares"] = r
    layout.write_row_label(ws, r, "Fully-diluted shares outstanding (m)",
                           "Azioni fully-diluted (m)", indent=True)
    c = ws.cell(row=r, column=4, value=target_fd)
    # A share count is a count, not a multiple — use the accounting number
    # format, not the 'x' multiple format.
    styles.style_input(c, number_format=styles.FMT_EUR_ACTUAL)
    r += 1
    refs["pp_option_buyout"] = r
    layout.write_row_label(ws, r, "Option buyout (treasury / in-the-money)",
                           "Acquisto opzioni (metodo tesoriera / ITM)",
                           indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "option_buyout_eur_m", 0.0) or 0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["pp_equity"] = r
    layout.write_row_label(ws, r, "Equity purchase price",
                           "Prezzo di acquisto equity")
    c = ws.cell(row=r, column=4,
                value=(f"=D{refs['pp_offer_px']}*D{refs['pp_fd_shares']}"
                       f"+D{refs['pp_option_buyout']}"))
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    r += 1
    # LBO-3: now that the canonical equity-purchase-price cell exists, point the
    # Uses "Purchase equity" line at it so the two figures always agree.
    _patch_uses_equity_pp()
    refs["pp_net_debt"] = r
    layout.write_row_label(ws, r, "(+) Target net debt assumed",
                           "(+) PFN target assunto", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "target_net_debt_close_eur_m", 0.0) or 0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["pp_transaction_fees"] = r
    layout.write_row_label(ws, r, "(+) Transaction fees (aggregate)",
                           "(+) Spese totali", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=D{refs['uses_ma_fees']}+D{refs['uses_fin_fees']}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["pp_enterprise"] = r
    layout.write_row_label(ws, r, "purchase_price — Enterprise Value",
                           "Prezzo acquisto — Enterprise Value")
    c = ws.cell(row=r, column=4,
                value=(f"=D{refs['pp_equity']}+D{refs['pp_net_debt']}"
                       f"+D{refs['pp_transaction_fees']}"))
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.border = styles.BORDER_TOP_THIN
    _def(wb, "purchase_price_eur_m", ws.title, f"$D${r}")
    r += 2

    # ── SECTION 3: PPA — goodwill + intangibles + DTL (US-203) ────────────
    layout.write_section_header(ws, r, "Purchase Price Allocation (PPA)",
                                "Allocazione prezzo di acquisto (PPA)")
    r += 1
    refs["ppa_bv"] = r
    layout.write_row_label(ws, r, "Target book value of equity at close",
                           "PN target alla chiusura", indent=True)
    bv = getattr(spec, "target_bv_equity_eur_m", None)
    c = ws.cell(row=r, column=4, value=bv.base if bv else 50.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["ppa_ppe"] = r
    layout.write_row_label(ws, r, "PP&E fair-value write-up",
                           "Rivalutazione immobilizzazioni", indent=True)
    ppe = getattr(spec, "ppe_writeup_eur_m", None)
    c = ws.cell(row=r, column=4, value=ppe.base if ppe else 0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["ppa_cust"] = r
    layout.write_row_label(ws, r, "Intangibles — customer list",
                           "Intangibili — lista clienti", indent=True)
    ic = getattr(spec, "intangibles_customer_list_eur_m", None)
    c = ws.cell(row=r, column=4, value=ic.base if ic else 30.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["ppa_tech"] = r
    layout.write_row_label(ws, r, "Intangibles — technology",
                           "Intangibili — tecnologia", indent=True)
    it = getattr(spec, "intangibles_technology_eur_m", None)
    c = ws.cell(row=r, column=4, value=it.base if it else 15.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["ppa_trade"] = r
    layout.write_row_label(ws, r, "Intangibles — trade name",
                           "Intangibili — marchio", indent=True)
    tn = getattr(spec, "intangibles_trade_name_eur_m", None)
    c = ws.cell(row=r, column=4, value=tn.base if tn else 10.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["ppa_dtl_rate"] = r
    layout.write_row_label(ws, r, "DTL rate on asset step-ups",
                           "Aliquota DTL su rivalutazioni", indent=True)
    dtl_r = getattr(spec, "dtl_rate_pct", None)
    c = ws.cell(row=r, column=4, value=dtl_r.base if dtl_r else 0.24)
    styles.style_input(c, number_format=styles.FMT_PCT)
    r += 1
    refs["ppa_dtl"] = r
    layout.write_row_label(ws, r, "DTL on asset write-ups",
                           "DTL su rivalutazioni", indent=True)
    c = ws.cell(row=r, column=4,
                value=(f"=(D{refs['ppa_ppe']}+D{refs['ppa_cust']}"
                       f"+D{refs['ppa_tech']}+D{refs['ppa_trade']})"
                       f"*D{refs['ppa_dtl_rate']}"))
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["ppa_goodwill"] = r
    layout.write_row_label(ws, r, "Goodwill created at close",
                           "Avviamento creato")
    c = ws.cell(
        row=r, column=4,
        value=(f"=D{refs['pp_equity']}-D{refs['ppa_bv']}"
               f"-D{refs['ppa_ppe']}-D{refs['ppa_cust']}"
               f"-D{refs['ppa_tech']}-D{refs['ppa_trade']}"
               f"+D{refs['ppa_dtl']}"),
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    c.comment = Comment(
        "Goodwill = Purchase Equity − BV Equity − Write-ups + DTL "
        "(ASC 805 / IFRS 3). Not deductible under GAAP/IFRS; annual "
        "impairment testing rather than amortization.",
        "ModelForge",
    )
    r += 1

    # Intangible amortization (straight-line per useful life)
    cust_life = getattr(spec, "customer_list_useful_life_years", 10)
    tech_life = getattr(spec, "technology_useful_life_years", 7)
    trade_life = getattr(spec, "trade_name_useful_life_years", 15)
    refs["amort_cust"] = r
    layout.write_row_label(ws, r,
                           f"Customer-list amortization (SL, {cust_life}y)",
                           f"Ammortamento lista clienti (SL, {cust_life} anni)",
                           indent=True)
    c = ws.cell(row=r, column=4, value=f"=-D{refs['ppa_cust']}/{cust_life}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["amort_tech"] = r
    layout.write_row_label(ws, r,
                           f"Technology amortization (SL, {tech_life}y)",
                           f"Ammortamento tecnologia (SL, {tech_life} anni)",
                           indent=True)
    c = ws.cell(row=r, column=4, value=f"=-D{refs['ppa_tech']}/{tech_life}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["amort_trade"] = r
    layout.write_row_label(ws, r,
                           f"Trade-name amortization (SL, {trade_life}y)",
                           f"Ammortamento marchio (SL, {trade_life} anni)",
                           indent=True)
    c = ws.cell(row=r, column=4, value=f"=-D{refs['ppa_trade']}/{trade_life}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["amort_total"] = r
    layout.write_row_label(ws, r, "Total PPA amortization (annual)",
                           "Ammortamento PPA totale (annuale)")
    c = ws.cell(row=r, column=4,
                value=(f"=D{refs['amort_cust']}+D{refs['amort_tech']}"
                       f"+D{refs['amort_trade']}"))
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    r += 2

    # ── SECTION 4: Transaction fees split (#35) ───────────────────────────
    layout.write_section_header(ws, r, "Transaction fees — split",
                                "Spese di transazione — dettaglio")
    r += 1
    refs["txn_ma"] = r
    layout.write_row_label(ws, r, "M&A advisory fees (expensed at close)",
                           "Spese M&A (spesate alla chiusura)", indent=True)
    c = ws.cell(row=r, column=4, value=ma_fees)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["txn_fin"] = r
    layout.write_row_label(ws, r,
                           "Financing fees (capitalized, amortized over tenor)",
                           "Commissioni di finanziamento (capitalizzate)",
                           indent=True)
    c = ws.cell(row=r, column=4, value=fin_fees)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 2

    # ── SECTION 5: Sponsor capital structure (US-207) ─────────────────────
    layout.write_section_header(ws, r, "Sponsor capital structure",
                                "Struttura capitale dello sponsor")
    r += 1
    refs["cap_sponsor_eq"] = r
    layout.write_row_label(ws, r, "Sponsor equity (new money)",
                           "Equity dello sponsor", indent=True)
    c = ws.cell(row=r, column=4, value=sponsor_equity)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["cap_rollover"] = r
    layout.write_row_label(ws, r, "Management rollover",
                           "Rollover management", indent=True)
    c = ws.cell(row=r, column=4, value=mgmt_rollover)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    mip_pct = getattr(spec, "mip_pool_pct", 0.10)
    mip_years = getattr(spec, "mip_vesting_years", 4)
    refs["cap_mip"] = r
    layout.write_row_label(ws, r,
                           f"MIP pool (post-close equity, {mip_years}y vest)",
                           f"Pool MIP ({mip_years} anni vesting)", indent=True)
    c = ws.cell(row=r, column=4, value=mip_pct)
    styles.style_input(c, number_format=styles.FMT_PCT)
    r += 1
    refs["cap_total_eq"] = r
    layout.write_row_label(ws, r, "Total equity at close (sponsor + rollover)",
                           "Equity totale alla chiusura")
    c = ws.cell(row=r, column=4,
                value=f"=D{refs['cap_sponsor_eq']}+D{refs['cap_rollover']}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = styles.font_subheader
    r += 2

    # ── SECTION 6: NWC closing adjustment (US-213) ────────────────────────
    layout.write_section_header(ws, r, "NWC closing adjustment",
                                "Adeguamento CCN alla chiusura")
    r += 1
    refs["nwc_peg"] = r
    layout.write_row_label(ws, r, "NWC target peg (at close)",
                           "Target CCN alla chiusura", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "nwc_target_peg_eur_m", 0.0) or 0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["nwc_true_up"] = r
    layout.write_row_label(ws, r, "NWC true-up at close (actual − peg)",
                           "Aggiustamento CCN (effettivo − target)",
                           indent=True)
    c = ws.cell(row=r, column=4, value=0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 2

    # ── SECTION 7: Earnout / CVR (US-209) ─────────────────────────────────
    layout.write_section_header(ws, r, "Earnout / CVR",
                                "Earnout / CVR")
    r += 1
    refs["earnout_fv"] = r
    layout.write_row_label(ws, r, "Earnout fair value at close",
                           "Fair value earnout alla chiusura", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "earnout_fair_value_eur_m", 0.0) or 0.0)
    styles.style_input(c, number_format=styles.FMT_EUR_M)
    r += 1
    refs["earnout_year"] = r
    layout.write_row_label(ws, r, "Earnout payment year",
                           "Anno pagamento earnout", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "earnout_year", 2))
    styles.style_input(c, number_format=styles.FMT_INTEGER)
    r += 1

    # v0.8.9 US-589: Earnout accretion through P&L. Default 5% discount
    # rate on fair value; accretes each period until payment.
    refs["earnout_discount"] = r
    layout.write_row_label(
        ws, r, "Earnout discount rate (accretion)",
        "Tasso sconto earnout", indent=True,
    )
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "earnout_discount_pct", 0.05))
    styles.style_input(c, number_format=styles.FMT_PCT_2DP)
    _def(wb, "earnout_discount_pct", ws.title, f"$D${r}")
    r += 1

    # Total accretion over life = FV × ((1+disc)^year − 1)
    refs["earnout_accretion"] = r
    layout.write_row_label(
        ws, r, "Earnout accretion through P&L (total)",
        "Accrual earnout a P&L (totale)", indent=True,
    )
    c = ws.cell(
        row=r, column=4,
        value=(f"=D{refs['earnout_fv']}*"
               f"((1+earnout_discount_pct)^D{refs['earnout_year']}-1)"),
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Earnout / CVR accretion through P&L per IFRS 3.58: contingent "
        "consideration at FV unwinds at the discount rate each period. "
        "Total accretion over life = FV × ((1+disc)^year − 1).",
        "ModelForge",
    )
    r += 2

    # ── SECTION 8: Dividend recap (US-208) ────────────────────────────────
    layout.write_section_header(ws, r, "Dividend recap",
                                "Dividend recap")
    r += 1
    refs["recap_enabled"] = r
    layout.write_row_label(ws, r, "Dividend recap enabled",
                           "Dividend recap abilitato", indent=True)
    c = ws.cell(row=r, column=4,
                value=1 if getattr(spec, "div_recap_enabled", False) else 0)
    styles.style_input(c, number_format=styles.FMT_INTEGER)
    r += 1
    refs["recap_year"] = r
    layout.write_row_label(ws, r, "Recap year",
                           "Anno recap", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "div_recap_year", 3))
    styles.style_input(c, number_format=styles.FMT_INTEGER)
    r += 1
    refs["recap_leverage"] = r
    layout.write_row_label(ws, r, "Recap target leverage (× EBITDA)",
                           "Leva target recap", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "div_recap_target_leverage", 4.0))
    styles.style_input(c, number_format=styles.FMT_MULTIPLE)
    _def(wb, "recap_target_leverage", ws.title, f"$D${r}")
    r += 1

    # v0.8.9 US-584: Recap sponsor share — fraction of recap proceeds
    # distributed to sponsor (vs held as cash on BS). Default 1.0.
    refs["recap_sponsor_share"] = r
    layout.write_row_label(ws, r, "Recap sponsor share of proceeds",
                           "Quota sponsor su proventi recap", indent=True)
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "div_recap_sponsor_share_pct", 1.0))
    styles.style_input(c, number_format=styles.FMT_PCT)
    _def(wb, "recap_sponsor_share_pct", ws.title, f"$D${r}")
    r += 2

    # ── SECTION 9: Exit scenarios ×3 (US-210, #37) ────────────────────────
    layout.write_section_header(ws, r, "Exit scenarios",
                                "Scenari di uscita")
    r += 1
    strat_x = getattr(spec, "exit_strategic_multiple", 10.0)
    ipo_x = getattr(spec, "exit_ipo_multiple", 13.0)
    sec_x = getattr(spec, "exit_secondary_multiple", 8.0)
    exit_year = getattr(spec, "exit_year", 5)

    refs["exit_year"] = r
    layout.write_row_label(ws, r, "Exit year", "Anno di uscita", indent=True)
    c = ws.cell(row=r, column=4, value=exit_year)
    styles.style_input(c, number_format=styles.FMT_INTEGER)
    _def(wb, "exit_year_input", ws.title, f"$D${r}")
    r += 1

    refs["exit_strategic"] = r
    layout.write_row_label(ws, r, "Exit — strategic sale (EV/EBITDA ×)",
                           "Uscita — vendita strategica", indent=True)
    c = ws.cell(row=r, column=4, value=strat_x)
    styles.style_input(c, number_format=styles.FMT_MULTIPLE)
    r += 1
    refs["exit_ipo"] = r
    layout.write_row_label(ws, r, "Exit — IPO (P/E ×)",
                           "Uscita — IPO", indent=True)
    c = ws.cell(row=r, column=4, value=ipo_x)
    styles.style_input(c, number_format=styles.FMT_MULTIPLE)
    r += 1
    refs["exit_secondary"] = r
    layout.write_row_label(ws, r, "Exit — secondary LBO (leverage cap)",
                           "Uscita — secondary LBO", indent=True)
    c = ws.cell(row=r, column=4, value=sec_x)
    styles.style_input(c, number_format=styles.FMT_MULTIPLE)
    r += 2

    # ── SECTION 10: Hurdle analysis (US-211, #40) ─────────────────────────
    layout.write_section_header(ws, r, "Hurdle analysis — reverse-solve max PP",
                                "Hurdle — prezzo max al ritorno obiettivo")
    r += 1
    hurdle_pcts = getattr(spec, "hurdle_irr_pcts", [0.20, 0.25, 0.30])
    for i, pct in enumerate(hurdle_pcts):
        refs[f"hurdle_{i}"] = r
        layout.write_row_label(ws, r,
                               f"Hurdle IRR {pct:.0%} — implied max purchase price",
                               f"Hurdle IRR {pct:.0%} — PP max implicito",
                               indent=True)
        # Approximate reverse-solve: at exit, assume 3x terminal MoIC → PP = Exit/(1+IRR)^year / 3
        # Simple placeholder formula; full reverse-solve requires iterative Excel.
        # Uses exit EQUITY (EV − net debt at exit), consistent with the returns.
        c = ws.cell(
            row=r, column=4,
            value=(f"=(D{refs['exit_strategic']}*{proj_exit_ebitda_ref}"
                   f"-{proj_exit_debt_ref})/"
                   f"(1+{pct})^D{refs['exit_year']}/3"),
        )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1
    ws.cell(row=r, column=1, value="Reverse-solve — max PP for 20/25/30% IRR").font = styles.font_label_it
    r += 2

    # ── SECTION 11: GP promote waterfall (US-212, #39) ────────────────────
    layout.write_section_header(ws, r, "Sponsor GP promote",
                                "Waterfall GP promote")
    r += 1
    gp_pref = getattr(spec, "gp_pref_pct", 0.08)
    gp_catch = getattr(spec, "gp_catchup_pct", 1.0)
    gp_carry = getattr(spec, "gp_carry_pct", 0.20)
    gp_type = getattr(spec, "gp_waterfall_type", "european")
    refs["promote_pref"] = r
    layout.write_row_label(ws, r, f"LP preferred return (hurdle)",
                           "Rendimento preferred LP", indent=True)
    c = ws.cell(row=r, column=4, value=gp_pref)
    styles.style_input(c, number_format=styles.FMT_PCT)
    r += 1
    refs["promote_catchup"] = r
    layout.write_row_label(ws, r, "GP catch-up rate",
                           "Catch-up GP", indent=True)
    c = ws.cell(row=r, column=4, value=gp_catch)
    styles.style_input(c, number_format=styles.FMT_PCT)
    r += 1
    refs["promote_carry"] = r
    layout.write_row_label(ws, r, "GP carry / promote",
                           "Carry / promote GP", indent=True)
    c = ws.cell(row=r, column=4, value=gp_carry)
    styles.style_input(c, number_format=styles.FMT_PCT)
    r += 1
    refs["promote_type"] = r
    layout.write_row_label(ws, r, "Waterfall type",
                           "Tipo waterfall", indent=True)
    c = ws.cell(row=r, column=4, value=gp_type)
    r += 2

    # ── SECTION 12a: Sponsor cash-flow series (US-570) ────────────────────
    # Real per-year series so exit IRR can use IRR() — equal-annual
    # periods (no volatile TODAY() needed). Columns D..M = Y0..Y9.
    layout.write_section_header(ws, r, "Sponsor cash-flow series (Y0..Y9)",
                                "Serie di cassa sponsor")
    r += 1
    max_hold = 10
    # Year headers
    refs["cf_year_hdr"] = r
    ws.cell(row=r, column=2, value="Year").font = styles.font_label_it
    for y in range(max_hold):
        col_idx = 4 + y  # D..M
        c = ws.cell(row=r, column=col_idx, value=y)
        styles.style_header(c)
    r += 1

    # Sponsor CF series — row per scenario
    exit_labels = [
        ("Strategic sale", "Vendita strategica", "exit_strategic"),
        ("IPO", "IPO", "exit_ipo"),
        ("Secondary LBO", "Secondary LBO", "exit_secondary"),
    ]
    cf_rows: dict[str, int] = {}
    for en_label, it_label, mult_key in exit_labels:
        ws.cell(row=r, column=1, value=en_label).font = styles.font_subheader
        ws.cell(row=r, column=2, value=it_label).font = styles.font_label_it
        cf_rows[mult_key] = r
        for y in range(max_hold):
            col_idx = 4 + y
            col_letter = chr(ord("A") + col_idx - 1)
            if y == 0:
                # LBO-EQUITY FIX (v0.11): the sponsor t0 equity outflow MUST be
                # the S&U balancing plug (true new money invested = Total Uses −
                # debt drawn − rollover), NOT the fixed capital-structure input
                # cell `cap_sponsor_eq`. Prior code outflowed `cap_sponsor_eq`
                # (35.0 spec input) while the S&U plug requires 44.3, so the
                # reported equity IRR/MoIC were computed on a DIFFERENT equity
                # number than the deal actually needs. Pointing at the plug cell
                # (`sources_sponsor_eq`, row 10) unifies the two: the IRR series
                # now consumes the same new-money cheque the S&U balances to.
                formula = f"=-D{refs['sources_sponsor_eq']}"
            else:
                # Dividend during hold: via recap in recap_year (if enabled),
                # else 0. Exit proceeds if y == exit_year.
                # LBO-2: exit proceeds use the PROJECTED exit-year EBITDA
                # (captures modeled growth), not entry LTM EBITDA. Exit FIX:
                # net the debt outstanding at exit so the sponsor receives
                # EQUITY (EV − net debt), not the whole enterprise value.
                exit_proceeds = (
                    f"IF({y}=D{refs['exit_year']},"
                    f"D{refs[mult_key]}*{proj_exit_ebitda_ref}"
                    f"-{proj_exit_debt_ref},0)"
                )
                # LBO-2: recap leverage applies to the PROJECTED recap-year
                # EBITDA (the deleveraged/grown EBITDA at the recap date).
                # Recap FIX: net the debt already OUTSTANDING at the recap date
                # (what the new debt refinances), not the sponsor equity cheque.
                recap_div = (
                    f"IF(AND(D{refs['recap_enabled']}=1,{y}=D{refs['recap_year']}),"
                    f"(recap_target_leverage*{proj_recap_ebitda_ref}"
                    f"-{proj_recap_debt_ref})*recap_sponsor_share_pct,0)"
                )
                # Earnout payment in earnout_year (neutral to sponsor — adds to
                # exit proceeds when relevant; placeholder 0 for cleanliness)
                formula = f"={exit_proceeds}+{recap_div}"
            c = ws.cell(row=r, column=col_idx, value=formula)
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1
    r += 1

    # ── SECTION 12: Returns summary (#38) ─────────────────────────────────
    layout.write_section_header(ws, r, "Returns summary",
                                "Riepilogo ritorni (IRR / MoIC / CoC)")
    r += 1

    for en_label, it_label, mult_key in exit_labels:
        ws.cell(row=r, column=1, value=en_label).font = styles.font_subheader
        ws.cell(row=r, column=2, value=it_label).font = styles.font_label_it
        r += 1
        # IRR via Excel IRR() on the per-year sponsor cash-flow series
        # (US-570). Equal-annual periods → no date array required. This
        # replaces the prior geometric approximation and handles variable
        # exit years + interim distributions (recap, dividends) correctly.
        cf_row_num = cf_rows[mult_key]
        cf_range = f"D{cf_row_num}:M{cf_row_num}"
        refs[f"irr_{mult_key}"] = r
        layout.write_row_label(ws, r, "IRR (on cash-flow series)",
                               "IRR (sulla serie di cassa)", indent=True)
        c = ws.cell(
            row=r, column=4,
            value=f"=IFERROR(IRR({cf_range}),0)",
        )
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        r += 1
        refs[f"moic_{mult_key}"] = r
        layout.write_row_label(ws, r, "MoIC (cash-on-cash)",
                               "MoIC (multiplo su equity)", indent=True)
        # MoIC = sum of positive CFs / -sum of negative CFs
        c = ws.cell(
            row=r, column=4,
            value=(f"=IFERROR(SUMIF({cf_range},\">0\")/"
                   f"-SUMIF({cf_range},\"<0\"),0)"),
        )
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
        r += 1
        refs[f"coc_{mult_key}"] = r
        layout.write_row_label(ws, r, "Cash-on-cash (coc)",
                               "Cash-on-cash", indent=True)
        c = ws.cell(row=r, column=4, value=f"=D{refs[f'moic_{mult_key}']}")
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
        r += 2

    # ── SECTION 13: PIK compound accrual + Dividend recap flow (US-570) ───
    layout.write_section_header(
        ws, r, "PIK accrual + Dividend recap (live compute)",
        "Accrual PIK + Dividend recap (computo dinamico)",
    )
    r += 1

    # v0.8.9 US-580: PIK coupon rate as workbook-level named range.
    # Input cell first, then the accrued-principal formula below refers
    # to the named range rather than a literal.
    refs["pik_coupon_input"] = r
    layout.write_row_label(
        ws, r, "PIK coupon rate (compound)",
        "Cedola PIK (compound)", indent=True,
    )
    c = ws.cell(row=r, column=4,
                value=getattr(spec, "pik_coupon_pct", 0.075))
    styles.style_input(c, number_format=styles.FMT_PCT)
    _def(wb, "pik_coupon_pct", ws.title, f"$D${r}")
    r += 1

    # PIK compound accrual: principal × (1 + pik_coupon_pct)^exit_year.
    # Both factors are now named-range driven.
    refs["pik_accrued_exit"] = r
    layout.write_row_label(
        ws, r, "PIK tranche accrued principal at exit",
        "Principale PIK accruato all'uscita", indent=True,
    )
    pik_initial_ref = (
        f"D{refs.get('cap_pik_debt', refs['cap_sponsor_eq'])}"
    )
    c = ws.cell(
        row=r, column=4,
        value=f"={pik_initial_ref}*(1+pik_coupon_pct)^D{refs['exit_year']}",
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "PIK compound accrual: principal × (1+pik_coupon_pct)^exit_year. "
        "Coupon named range pik_coupon_pct is editable. Assumes PIK "
        "coupon rolls into principal each period.",
        "ModelForge",
    )
    r += 1

    # v0.8.9 US-584: Dividend recap with live sponsor-share fraction.
    # Distribution = (target_lev × EBITDA − sponsor_eq) × sponsor_share_pct.
    refs["recap_distribution"] = r
    layout.write_row_label(
        ws, r, "Dividend recap — distribution to sponsor",
        "Dividend recap — distribuzione", indent=True,
    )
    c = ws.cell(
        row=r, column=4,
        value=(f"=IF(D{refs['recap_enabled']}=1,"
               f"(recap_target_leverage*{proj_recap_ebitda_ref}"
               f"-{proj_recap_debt_ref})*recap_sponsor_share_pct,0)"),
    )
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    ws.cell(row=r, column=4).comment = Comment(
        "Dividend recap distribution = IF(recap enabled, "
        "target_leverage × projected recap-year EBITDA − debt outstanding at "
        "recap, 0). New debt is issued up to target leverage on the GROWN "
        "(projected) EBITDA at the recap date; proceeds net of the debt being "
        "refinanced fund the sponsor distribution. Simplified: assumes no "
        "refinance fees (add via financing_fee_pct).",
        "ModelForge",
    )
    r += 2

    ws.freeze_panes = "D5"
    ws.print_title_rows = "1:4"
    return refs


def build_historical_ebitda_lfy(wb, spec) -> None:
    """Register `historical_ebitda_lfy` named range so SourcesUses formulas
    can reference last-FY EBITDA without cross-sheet lookup.

    Scalar workbook-level defined name pointing to the last entry of
    spec.historical_ebitda_eur_m.
    """
    from openpyxl.workbook.defined_name import DefinedName
    if not getattr(spec, "historical_ebitda_eur_m", None):
        return
    last = spec.historical_ebitda_eur_m[-1]
    # Store on hidden cell via Assumptions sheet or Cover? Simplest: put on
    # SourcesUses A1 corner... but that conflicts. Use Cover helper instead.
    # We'll emit a cell on SourcesUses at Z1 as scratch.
    if "SourcesUses" in wb.sheetnames:
        ws = wb["SourcesUses"]
        ws.cell(row=1, column=26, value=last)
        if "historical_ebitda_lfy" in wb.defined_names:
            del wb.defined_names["historical_ebitda_lfy"]
        wb.defined_names["historical_ebitda_lfy"] = DefinedName(
            name="historical_ebitda_lfy", attr_text="'SourcesUses'!$Z$1",
        )
