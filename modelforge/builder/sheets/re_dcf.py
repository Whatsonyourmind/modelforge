"""Real Estate NOI build + DCF + equity waterfall."""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    h = spec.horizon.hold_years
    n = h + 1  # t=0..t=h

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="RE DCF & NOI", title_it="DCF immobiliare e NOI",
        subtitle=f"{spec.meta.currency} {spec.meta.unit_scale} · {h}y hold",
    )
    layout.write_scenario_banner(ws, row=3)

    # Year headers t=0..t=h
    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=yr_row, column=col_idx, value=f"t={i}")
        styles.style_header(cc)

    rows: dict[str, int] = {}
    r = 7

    # Acquisition at t=0
    layout.write_section_header(ws, r, "Acquisition", "Acquisizione")
    r += 1
    rows["acq_price"] = r
    layout.write_row_label(ws, r, "Acquisition price", "Prezzo di acquisto")
    cc = ws.cell(row=r, column=4, value="=-acquisition_price_eur_m")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # NOI build
    layout.write_section_header(ws, r, "NOI build", "Costruzione NOI")
    r += 1

    rows["gross_rent"] = r
    layout.write_row_label(ws, r, "Gross potential rent (€/sqm × area)",
                           "Canone lordo potenziale")
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == 1:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=lettable_area_sqm*rent_eur_sqm_year1/1000000")
        else:
            prior = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${prior}${r}*(1+rent_indexation_pct)")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["vacancy"] = r
    layout.write_row_label(ws, r, "Vacancy loss", "Perdita per sfitto", indent=True)
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['gross_rent']}*vacancy_pct")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["effective_rent"] = r
    layout.write_row_label(ws, r, "Effective rent", "Canone effettivo")
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['gross_rent']}+${col}${rows['vacancy']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["property_opex"] = r
    layout.write_row_label(ws, r, "Property opex", "Opex immobiliare", indent=True)
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['gross_rent']}*opex_pct_gross_rent")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["noi"] = r
    layout.write_row_label(ws, r, "NOI (Net Operating Income)", "NOI")
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['effective_rent']}+${col}${rows['property_opex']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 1

    rows["capex"] = r
    layout.write_row_label(ws, r, "Maintenance capex", "Capex manutenzione", indent=True)
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['gross_rent']}*capex_pct_gross_rent")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["cfads"] = r
    layout.write_row_label(ws, r, "CF after capex (pre-debt, pre-tax)",
                           "CF post capex")
    ws.cell(row=r, column=4, value=0)
    for i in range(1, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['noi']}+${col}${rows['capex']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    r += 2

    # Exit
    layout.write_section_header(ws, r, "Exit", "Uscita")
    r += 1
    last_col = layout.year_col(h)
    rows["exit_noi"] = r
    layout.write_row_label(ws, r, "Exit-year NOI", "NOI anno di uscita")
    cc = ws.cell(row=r, column=4,
                 value=f"=${last_col}${rows['noi']}")
    styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["exit_value"] = r
    layout.write_row_label(ws, r, "Gross sale proceeds (NOI / cap)", "Corrispettivo lordo")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['exit_noi']}/exit_cap_rate")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    c_ann_col_idx = ord(last_col) - ord("A") + 1
    exit_val_target = ws.cell(row=r, column=c_ann_col_idx, value=f"=$D${r}")  # duplicate on exit col for CF line
    styles.style_formula(exit_val_target, number_format=styles.FMT_EUR_M)
    r += 1

    rows["transaction_costs"] = r
    layout.write_row_label(ws, r, "Transaction costs", "Costi di transazione", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=-$D${rows['exit_value']}*transaction_costs_pct")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["net_exit"] = r
    layout.write_row_label(ws, r, "Net sale proceeds", "Corrispettivo netto")
    cc = ws.cell(row=r, column=4,
                 value=f"=$D${rows['exit_value']}+$D${rows['transaction_costs']}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    cc.font = styles.font_subheader
    cc.border = styles.BORDER_TOP_THIN
    r += 2

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    return out
