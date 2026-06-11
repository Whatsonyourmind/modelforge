"""Project Finance cashflow sheet.

Construction phase: capex per year, IDC capitalized, no revenue.
Operating phase: revenue → opex → EBITDA → tax → CADS → debt service → dividends.

Layout: columns D..D+N where N = construction_years + operating_years.

v0.3: CFADS (DSCR numerator) unchanged; after pf_debt has emitted, the
template orchestrator calls `append_distributable_cash(ws, cashflow_refs,
debt_refs)` which adds:
    - Senior debt service (from DebtDSCR)
    - DSRA funding (from DebtDSCR)
    - Distributable cash to equity

DSRA funding is DOWNSTREAM of the DSCR check per market convention.
"""

from __future__ import annotations

import openpyxl.comments
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout
from modelforge.builder.i18n import L


def build(ws: Worksheet, spec, driver_refs: dict[str, str]) -> dict[str, str]:
    c = spec.horizon.construction_years
    o = spec.horizon.operating_years
    n = c + o

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Project Cash Flow", title_it="Flusso di cassa del progetto",
        subtitle=f"{c}y construction · {o}y operating · {spec.meta.currency} {spec.meta.unit_scale}",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    ws.cell(row=yr_row, column=3, value="Phase").font = styles.font_subheader
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        phase = "C" if i < c else "O"
        yr = i + 1 if i < c else (i - c + 1)
        c_cell = ws.cell(row=yr_row, column=col_idx, value=f"{phase}{yr}")
        styles.style_header(c_cell)

    rows: dict[str, int] = {}
    r = 7

    # CONSTRUCTION
    layout.write_section_header(ws, r, "Construction phase", "Fase di costruzione")
    r += 1

    rows["capex"] = r
    layout.write_row_label(ws, r, "Capex (outflow)", "Capex (uscita)", indent=True)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    capex_total = spec.construction.total_capex_eur_m.name
    for i in range(c):
        phasing = spec.construction.capex_phasing_pct[i].name
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"=-{capex_total}*{phasing}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    for i in range(c, n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 2

    # OPERATING
    layout.write_section_header(ws, r, "Operating phase", "Fase operativa")
    r += 1

    rows["revenue"] = r
    layout.write_row_label(ws, r, L("revenue").en, L("revenue").secondary)
    ws.cell(row=r, column=3, value=spec.meta.currency).font = styles.font_label_it
    yr1 = spec.operating.availability_payment_eur_m_yr1.name
    idx = spec.operating.revenue_indexation_pct.name
    # v0.8 US-240: panel degradation compounds into revenue when present.
    # Revenue_t = Revenue_{t-1} × (1 + indexation) × (1 − degradation).
    deg_assum = getattr(spec.operating, "panel_degradation_pct_annual", None)
    deg_factor = f"*(1-{deg_assum.name})" if deg_assum is not None else ""
    for i in range(c):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    for i in range(c, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == c:
            cc = ws.cell(row=r, column=col_idx, value=f"={yr1}")
        else:
            prior_col = layout.year_col(i - 1)
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${prior_col}${r}*(1+{idx}){deg_factor}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    if deg_assum is not None:
        ws.cell(row=r, column=4).comment = openpyxl.comments.Comment(
            f"Revenue compounds at (1+indexation)×(1−{deg_assum.name}) per "
            "year (panel degradation, solar PV convention 0.5% p.a. "
            "per manufacturer warranty).",
            "ModelForge",
        )
    r += 1

    # v0.8 US-241: P90 alternative revenue row (downside scenario used
    # by the debt sizer when debt_sizing_mode='dscr_target_p90').
    # P90 = P50 × (1 − p90_haircut). Held as a parallel row so both
    # P50 and P90 CFADS are available to the audit trail.
    p90_assum = getattr(spec.operating, "p90_revenue_haircut_pct", None)
    if p90_assum is not None:
        rows["revenue_p90"] = r
        layout.write_row_label(ws, r, "Revenue — P90 (downside, haircut)",
                               "Ricavi — P90 (scenario downside)", indent=True)
        for i in range(c):
            col_idx = ord(layout.year_col(i)) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
        for i in range(c, n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['revenue']}*(1-{p90_assum.name})")
            styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        ws.cell(row=r, column=4).comment = openpyxl.comments.Comment(
            "P90 revenue = P50 × (1 − p90_haircut). Used by debt sizer "
            "under dscr_target_p90 mode per Solargis/NREL bankability "
            "convention. Both P50 and P90 CFADS kept live for audit trail.",
            "ModelForge",
        )
        r += 1

    rows["opex"] = r
    layout.write_row_label(ws, r, "Opex (cost)", "Opex (costo)", indent=True)
    for i in range(c):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    for i in range(c, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-${col}${rows['revenue']}*opex_pct_revenue")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ebitda"] = r
    layout.write_row_label(ws, r, L("ebitda").en, L("ebitda").secondary)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['revenue']}+${col}${rows['opex']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    r += 1

    # v0.6: Full tax walk with D&A and interest shields.
    # D&A straight-line over the operating period, starting at COD.
    rows["da"] = r
    layout.write_row_label(ws, r, "D&A (straight-line)", "A&A (lineare)", indent=True)
    capex_total = spec.construction.total_capex_eur_m.name
    for i in range(c):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    for i in range(c, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"=-{capex_total}/{o}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    rows["ebit"] = r
    layout.write_row_label(ws, r, "EBIT", "EBIT", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['ebitda']}+${col}${rows['da']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Interest expense — placeholder; pf_debt will patch to cross-sheet ref.
    rows["interest"] = r
    layout.write_row_label(ws, r, "Interest expense (ref)",
                           "Oneri finanziari (rif.)", indent=True)
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=0)
        styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Taxable income = EBIT + Interest (Interest is negative)
    rows["taxable"] = r
    layout.write_row_label(ws, r, "Taxable income", "Imponibile", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=${col}${rows['ebit']}+${col}${rows['interest']}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Tax — only positive taxable income is taxed (no NOL carry-forward yet)
    rows["tax"] = r
    layout.write_row_label(ws, r, "Taxes (on EBIT − Interest)",
                           "Imposte (su EBIT − Oneri)", indent=True)
    for i in range(c):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    for i in range(c, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"=-MAX(${col}${rows['taxable']},0)*effective_tax_rate")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # v0.8.9 US-588: ΔWC = (rev_t − rev_{t-1}) × nwc_pct_revenue.
    # When spec.operating.nwc_pct_revenue is absent, default to 0.
    nwc_pct_assum = getattr(spec.operating, "nwc_pct_revenue", None)
    nwc_pct_ref = (
        nwc_pct_assum.name if nwc_pct_assum is not None else "0"
    )
    rows["delta_wc"] = r
    layout.write_row_label(
        ws, r, "(−) Δ Working capital", "(−) Δ capitale circolante",
    )
    for i in range(c):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    for i in range(c, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i == c:
            # First operating year: ΔWC = rev × nwc_pct (seed working cap)
            cc = ws.cell(
                row=r, column=col_idx,
                value=f"=-${col}${rows['revenue']}*{nwc_pct_ref}",
            )
        else:
            prev_col = layout.year_col(i - 1)
            cc = ws.cell(
                row=r, column=col_idx,
                value=(f"=-(${col}${rows['revenue']}"
                       f"-${prev_col}${rows['revenue']})*{nwc_pct_ref}"),
            )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # v0.8.9 US-588: Maintenance capex = -revenue × maintenance_reserve_pct_revenue.
    # Leverages the existing OperatingPhase.maintenance_reserve_pct_revenue
    # assumption already on the Assumptions sheet as a named range.
    maint_pct_assum = getattr(
        spec.operating, "maintenance_reserve_pct_revenue", None,
    )
    maint_pct_ref = (
        maint_pct_assum.name if maint_pct_assum is not None else "0"
    )
    rows["maint_capex"] = r
    layout.write_row_label(
        ws, r, "(−) Maintenance capex", "(−) Capex di manutenzione",
    )
    for i in range(c):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        ws.cell(row=r, column=col_idx, value=0)
    for i in range(c, n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(
            row=r, column=col_idx,
            value=f"=-${col}${rows['revenue']}*{maint_pct_ref}",
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # CFADS = EBITDA − cash taxes − ΔWC − maintenance capex
    # Full bulge-tier identity per Breaking Into Wall Street / Edward Bodmer.
    # (Growth capex is funded from equity/debt, not from operating CF, so it
    # does not belong in CFADS — kept on the Capex schedule.)
    rows["cads"] = r  # Cash Available for Debt Service
    layout.write_row_label(
        ws, r,
        "CFADS (= EBITDA − cash taxes − ΔWC − maintenance capex)",
        "CFADS (= EBITDA − imposte − Δ WC − capex manutenzione)",
    )
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(
            row=r, column=col_idx,
            value=(
                f"=${col}${rows['ebitda']}+${col}${rows['tax']}"
                f"+${col}${rows['delta_wc']}+${col}${rows['maint_capex']}"
            ),
        )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"

    out = {f"{k}_row": str(v) for k, v in rows.items()}
    # Track where the distributable-cash section should begin (append later)
    out["_next_row"] = str(r + 1)
    return out


def append_distributable_cash(
    ws: Worksheet,
    spec,
    cashflow_refs: dict[str, str],
    debt_refs: dict[str, str],
    debt_sheet: str,
) -> dict[str, str]:
    """Append debt-service + DSRA + distributable-cash rows.

    Called by the template orchestrator AFTER pf_debt.build has emitted, so
    debt_refs are known. DSCR numerator stays CFADS (above); DSRA only
    affects distributable cash per market convention.
    """
    c = spec.horizon.construction_years
    o = spec.horizon.operating_years
    n = c + o

    r = int(cashflow_refs["_next_row"])

    layout.write_section_header(
        ws, r, "Waterfall — equity distributions", "Waterfall — distribuzioni equity",
    )
    r += 1

    debt_service_row = int(debt_refs["debt_service_row"])
    dsra_funding_row = int(debt_refs["dsra_funding_row"])
    cads_row = int(cashflow_refs["cads_row"])

    rows: dict[str, int] = {}

    # CADS passthrough for clarity
    rows["cads_ref"] = r
    layout.write_row_label(ws, r, "CADS (from above)", "CADS (dal sopra)", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx, value=f"=${col}${cads_row}")
        styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # Debt service (cross-sheet pull, negative)
    rows["ds_pull"] = r
    layout.write_row_label(ws, r, "Senior debt service", "Servizio debito senior", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"='{debt_sheet}'!{col}{debt_service_row}")
        styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # DSRA funding (cross-sheet pull, negative when funding, positive on release)
    rows["dsra_pull"] = r
    layout.write_row_label(ws, r, "DSRA funding / release", "DSRA finanziamento", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=f"='{debt_sheet}'!{col}{dsra_funding_row}")
        styles.style_xref(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # ── v0.8.8 US-560: O&M reserve funding live compute ──────────────────
    # Target balance = om_reserve_months × (annual_opex / 12).
    # Year 1 of operations: fund full target as a one-time debit.
    # Subsequent years: top up for opex growth (positive when opex grows).
    # Terminal year: full release back to equity.
    om_months = int(getattr(spec.operating, "om_reserve_months", 0) or 0)
    opex_row = int(cashflow_refs.get("opex_row", 0))
    rows["om_reserve_balance"] = r
    layout.write_row_label(ws, r, "O&M reserve balance (target)",
                           "Riserva O&M (obiettivo)", indent=True)
    if om_months and opex_row:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c:
                ws.cell(row=r, column=col_idx, value=0)
            else:
                # Target = |opex| × months/12 (opex is negative on the sheet)
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=f"=-${col}${opex_row}*{om_months}/12",
                )
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    else:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
    r += 1

    rows["om_reserve_cf"] = r
    layout.write_row_label(
        ws, r, "(−) O&M reserve funding / (+) release",
        "(−) Finanziamento riserva O&M / (+) rilascio", indent=True,
    )
    if om_months and opex_row:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c:
                ws.cell(row=r, column=col_idx, value=0)
            elif i == c:
                # Full funding at COD
                cc = ws.cell(row=r, column=col_idx,
                             value=f"=-${col}${rows['om_reserve_balance']}")
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
            elif i == n - 1:
                # Full release at terminal year (post-decommissioning)
                prev_col = layout.year_col(i - 1)
                cc = ws.cell(row=r, column=col_idx,
                             value=f"=${prev_col}${rows['om_reserve_balance']}")
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
            else:
                # Top up / release difference vs prior year target
                prev_col = layout.year_col(i - 1)
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=(f"=-(${col}${rows['om_reserve_balance']}"
                           f"-${prev_col}${rows['om_reserve_balance']})"),
                )
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    else:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
    r += 1

    # ── v0.8.8 US-561 + v0.8.9 US-581: MMR sinking fund (fully live).
    # Target = major_maintenance_reserve_eur_m (named range).
    # Build years + event-cycle length live via mmr_build_years named
    # range. Balance formula: target × MOD(ops_year, mmr_build_years) /
    # mmr_build_years. Event year (MOD = 0) resets balance to 0.
    # v0.12 (US-PF-covenants): the build-years value is lifted from the hardcode
    # `5` to spec.mmr_build_years (default 5 → byte-identical when omitted).
    mmr_target_assum = getattr(spec.operating, "major_maintenance_reserve_eur_m", None)
    mmr_target_name = mmr_target_assum.name if mmr_target_assum else None
    mmr_build_years = getattr(spec, "mmr_build_years", 5) or 5

    if mmr_target_name:
        # Register mmr_build_years as workbook-level named range first
        from openpyxl.workbook.defined_name import DefinedName
        wb = ws.parent
        layout.write_row_label(
            ws, r, "MMR build-years (sinking schedule)",
            "Anni di accantonamento MMR", indent=True,
        )
        # NB: keep this cell in its own variable (NOT `c`, which holds
        # construction-years and is reused below for the balance roll-forward).
        mmr_cell = ws.cell(row=r, column=4, value=mmr_build_years)
        styles.style_input(mmr_cell, number_format=styles.FMT_INTEGER)
        mmr_cell.comment = openpyxl.comments.Comment(
            "Major-Maintenance-Reserve sinking-fund build period: operating "
            "years over which the MMR ramps to its target before each "
            "maintenance event (MOD == 0 resets the balance). Overridable via "
            "spec.mmr_build_years.",
            "ModelForge",
        )
        if "mmr_build_years" in wb.defined_names:
            del wb.defined_names["mmr_build_years"]
        wb.defined_names["mmr_build_years"] = DefinedName(
            name="mmr_build_years",
            attr_text=f"'{ws.title}'!$D${r}",
        )
        r += 1

    rows["mmr_balance"] = r
    layout.write_row_label(ws, r, "MMR balance (sinking fund)",
                           "Riserva manutenzione straordinaria", indent=True)
    if mmr_target_name:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c:
                ws.cell(row=r, column=col_idx, value=0)
            else:
                ops_year = i - c + 1  # 1-indexed operating year
                # Formula: target × MOD(ops_year, build_years) / build_years
                # Event year (MOD = 0) resets balance to 0 automatically.
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=(f"={mmr_target_name}*"
                           f"MOD({ops_year},mmr_build_years)"
                           f"/mmr_build_years"),
                )
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    else:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
    r += 1

    rows["mmr_cf"] = r
    layout.write_row_label(ws, r, "(−) MMR funding / (+) draw",
                           "(−) Finanziamento MMR / (+) utilizzo", indent=True)
    if mmr_target_name:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c or i == c:
                ws.cell(row=r, column=col_idx, value=0)
            else:
                prev_col = layout.year_col(i - 1)
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=(f"=-(${col}${rows['mmr_balance']}"
                           f"-${prev_col}${rows['mmr_balance']})"),
                )
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    else:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
    r += 1

    # Distributable cash (gross, pre lock-up) = CADS + debt service + DSRA
    # + O&M reserve CF + MMR CF (all inflows/outflows contribute directly).
    rows["distributable_gross"] = r
    layout.write_row_label(ws, r, "Distributable cash — pre lock-up",
                           "Cassa distribuibile — pre lock-up", indent=True)
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        cc = ws.cell(row=r, column=col_idx,
                     value=(f"=${col}${rows['cads_ref']}"
                            f"+${col}${rows['ds_pull']}"
                            f"+${col}${rows['dsra_pull']}"
                            f"+${col}${rows['om_reserve_cf']}"
                            f"+${col}${rows['mmr_cf']}"))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    # v0.8 US-244: Lock-up test pass flag — DSCR ≥ lock_up_threshold.
    # Reads DSCR from DebtDSCR sheet, threshold from named range.
    dscr_row = int(debt_refs.get("dscr_row", 0))
    rows["lockup_pass"] = r
    layout.write_row_label(ws, r, "Lock-up test pass (DSCR ≥ threshold)",
                           "Lock-up test — superato", indent=True)
    lockup_name = spec.covenant.lock_up_threshold.name
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i < c or dscr_row == 0:
            ws.cell(row=r, column=col_idx, value=0)
        else:
            cc = ws.cell(
                row=r, column=col_idx,
                value=(f"=IF('{debt_sheet}'!{col}{dscr_row}>="
                       f"{lockup_name},1,0)"),
            )
            styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    r += 1

    # ── v0.8.8 US-562: Equity cure (iterative calc). Turns on workbook
    # iterative calc so the cure can cascade through DSCR. When DSCR <
    # covenant_min AND cures_used < cap, sponsor injects equity equal to
    # the covenant shortfall × (1 + haircut). Simple haircut: 0 (1:1).
    # Cure count is tracked cumulatively across operating years.
    rows["cure_cumulative"] = r
    layout.write_row_label(ws, r, "Equity cures used (cumulative)",
                           "Cure equity cumulate", indent=True)
    cure_cap = int(getattr(spec.debt, "equity_cure_cap_count", 0) or 0)
    cure_enabled = cure_cap > 0 and dscr_row != 0
    if cure_enabled:
        ws.parent.calculation.iterate = True
        ws.parent.calculation.iterateCount = 100
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c:
                ws.cell(row=r, column=col_idx, value=0)
            else:
                if i == c:
                    cc = ws.cell(row=r, column=col_idx,
                                 value=f"=IF(${col}${rows['lockup_pass']}=0,1,0)")
                else:
                    prev_col = layout.year_col(i - 1)
                    cc = ws.cell(
                        row=r, column=col_idx,
                        value=(f"=${prev_col}${r}"
                               f"+IF(AND(${col}${rows['lockup_pass']}=0,"
                               f"${prev_col}${r}<{cure_cap}),1,0)"),
                    )
                styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    else:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
    r += 1

    rows["cure_cf"] = r
    layout.write_row_label(ws, r, "(+) Equity cure injection",
                           "(+) Iniezione equity cure", indent=True)
    if cure_enabled:
        # Cure shortfall = covenant_min × debt_service − CADS  (positive
        # when short). Covenant proxy: lock-up threshold (close enough
        # for sponsor cure sizing).
        lockup_name = spec.covenant.lock_up_threshold.name
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c:
                ws.cell(row=r, column=col_idx, value=0)
            else:
                # Cure fires only in breach years that are still within cap.
                trigger = (
                    f"AND(${col}${rows['lockup_pass']}=0,"
                    f"${col}${rows['cure_cumulative']}<={cure_cap})"
                )
                shortfall = (
                    f"(-${col}${rows['ds_pull']}*{lockup_name})"
                    f"-${col}${rows['cads_ref']}"
                )
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=f"=IF({trigger},MAX({shortfall},0),0)",
                )
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    else:
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            ws.cell(row=r, column=col_idx, value=0)
    r += 1

    # ── v0.8.8 US-563 + v0.8.9 US-586: Make-whole on early redemption,
    # fully live. Premium = principal_outstanding × make_whole_spread_bps
    # × remaining_tenor_years / 10000, applied in the early-redemption year.
    # Triggered by spec.debt.early_redemption_flag = 1. Principal proxy
    # = initial_debt × early_redemption_principal_pct.
    mw_spread_assum = getattr(spec.debt, "make_whole_spread_bps", None)
    mw_name = mw_spread_assum.name if mw_spread_assum else None
    rows["make_whole_cf"] = r
    layout.write_row_label(
        ws, r, "(−) Make-whole premium on early redemption",
        "(−) Premio make-whole su rimborso anticipato", indent=True,
    )
    er_flag = int(getattr(spec.debt, "early_redemption_flag", 0) or 0)
    er_year = int(getattr(spec.debt, "early_redemption_year", 0) or 0)
    er_pct = float(getattr(spec.debt, "early_redemption_principal_pct", 0.8) or 0.0)
    total_op_years = o  # operating-phase tenor proxy
    # Find debt-principal reference (sum of debt-tranche amounts)
    initial_debt_ref = "+".join(
        f"{t.amount.name}" for t in spec.debt.tranches
    ) if getattr(spec.debt, "tranches", None) else "0"
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i < c:
            ws.cell(row=r, column=col_idx, value=0)
        else:
            ops_year = i - c + 1
            if mw_name and er_flag and ops_year == er_year:
                remaining = max(total_op_years - ops_year, 0)
                cc = ws.cell(
                    row=r, column=col_idx,
                    value=(f"=-({initial_debt_ref})*{er_pct}"
                           f"*{mw_name}/10000*{remaining}"),
                )
                styles.style_formula(cc, number_format=styles.FMT_EUR_M)
            else:
                ws.cell(row=r, column=col_idx, value=0)
    ws.cell(row=r, column=4).comment = openpyxl.comments.Comment(
        "Make-whole premium: -principal_outstanding × make_whole_spread_bps "
        "/ 10000 × remaining_tenor_years. Triggers only when "
        "early_redemption_flag=1 and ops_year=early_redemption_year.",
        "ModelForge",
    )
    r += 1

    # ── v0.8.8 US-565 + v0.8.9 US-587: Mandatory prepayment 5-event live.
    # Each event row: flag × expected_cash in year = trigger_year.
    # Default trigger_year = first operating year unless spec overrides.
    mp_events = [
        ("mp_insurance_flag", "mp_insurance_eur_m",
         "Insurance proceeds",
         "Proventi assicurativi"),
        ("mp_asset_sale_flag", "mp_asset_sale_eur_m",
         "Asset sale",
         "Vendita cespiti"),
        ("mp_coc_flag", "mp_coc_eur_m",
         "Change of control",
         "Cambio controllo"),
        ("mp_illegality_flag", "mp_illegality_eur_m",
         "Illegality event",
         "Illegalità"),
        ("mp_cf_sweep_flag", "mp_cf_sweep_eur_m",
         "Excess CF sweep",
         "Eccesso CF sweep"),
    ]
    mp_sub_rows: list[int] = []
    for flag_attr, cash_attr, en, it in mp_events:
        flag = int(getattr(spec.debt, flag_attr, 0) or 0)
        cash = float(getattr(spec.debt, cash_attr, 0.0) or 0.0)
        mp_sub_rows.append(r)
        layout.write_row_label(
            ws, r, f"  ◦ {en} (flag × amount)",
            f"  ◦ {it}", indent=True,
        )
        for i in range(n):
            col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
            if i < c:
                ws.cell(row=r, column=col_idx, value=0)
            else:
                ops_year = i - c + 1
                # Event fires in first operating year by default.
                if ops_year == 1 and flag:
                    cc = ws.cell(row=r, column=col_idx, value=-cash)
                    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
                else:
                    ws.cell(row=r, column=col_idx, value=0)
        r += 1

    # Mandatory-prepay aggregate row (sum of 5 event rows)
    rows["mandatory_prepay_cf"] = r
    layout.write_row_label(
        ws, r, "(−) Mandatory prepayment — total",
        "(−) Rimborso obbligatorio — totale", indent=True,
    )
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        parts = [f"${col}${sr}" for sr in mp_sub_rows]
        cc = ws.cell(row=r, column=col_idx, value=f"=" + "+".join(parts))
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
    ws.cell(row=r, column=4).comment = openpyxl.comments.Comment(
        "Sum of five mandatory-prepayment events per Reg. EU 575/2013 "
        "Art. 178 and standard PF loan covenants. Each event toggleable "
        "via spec.debt.mp_*_flag + mp_*_eur_m input.",
        "ModelForge",
    )
    r += 1

    # Distributable cash (net, post lock-up) = IF lock-up pass, gross; else 0
    # Adjusted for cure injection (+), make-whole (−), mandatory prepay (−).
    rows["distributable"] = r
    layout.write_row_label(ws, r, "Distributable cash to equity (post lock-up)",
                           "Cassa distribuibile agli equity holder")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        if i < c or dscr_row == 0:
            cc = ws.cell(row=r, column=col_idx,
                         value=f"=${col}${rows['distributable_gross']}")
        else:
            cc = ws.cell(
                row=r, column=col_idx,
                value=(
                    f"=IF(${col}${rows['lockup_pass']}=1,"
                    f"${col}${rows['distributable_gross']},0)"
                    f"+${col}${rows['cure_cf']}"
                    f"+${col}${rows['make_whole_cf']}"
                    f"+${col}${rows['mandatory_prepay_cf']}"
                ),
            )
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    ws.cell(row=r, column=4).comment = openpyxl.comments.Comment(
        "Distributable cash is blocked (=0) when DSCR < lock_up_threshold "
        "per bulge-bracket covenant standard. Gross pre-lock-up figure "
        "retained above for auditor inspection. v0.8.8 adds equity cure "
        "injection (+), make-whole premium (−), and mandatory prepayment "
        "event toggles (−).",
        "ModelForge",
    )
    r += 2

    # ── v0.8.8 US-564: Real-vs-nominal inflation QC check ─────────────
    # Sanity: tariff escalator ≈ opex inflation ≈ nominal-WACC inflation
    # baseline. Flag red if any diverge by >50bps.
    rows["inflation_qc"] = r
    layout.write_row_label(
        ws, r, "QC: inflation consistency (tariff vs opex)",
        "QC: coerenza inflazione (tariffa vs costi)", indent=True,
    )
    tariff_idx = spec.operating.revenue_indexation_pct.name
    opex_idx = spec.operating.opex_indexation_pct.name
    # Pass (=1) if |tariff − opex| <= 0.005 (50bps); fail otherwise
    c_idx = 4
    cc = ws.cell(
        row=r, column=c_idx,
        value=f"=IF(ABS({tariff_idx}-{opex_idx})<=0.005,1,0)",
    )
    styles.style_formula(cc, number_format=styles.FMT_INTEGER)
    ws.cell(row=r, column=c_idx).comment = openpyxl.comments.Comment(
        "Real-vs-nominal consistency: tariff escalator and opex "
        "inflation assumption should be within 50bps of each other "
        "(both in same real or both in same nominal basis). Diverging "
        "values imply unit-mismatch between revenue and cost sides.",
        "ModelForge",
    )
    r += 1

    return {f"{k}_row": str(v) for k, v in rows.items()}
