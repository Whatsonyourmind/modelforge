"""Project Finance equity returns sheet.

Equity CF = CADS + debt drawdown − capex − debt service (constr + op phases).
IRR + NPV for the project sponsor.
"""

from __future__ import annotations

from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles, layout


def build(ws: Worksheet, spec, cashflow_refs: dict[str, str], debt_refs: dict[str, str],
          cashflow_sheet: str, debt_sheet: str) -> None:
    c = spec.horizon.construction_years
    o = spec.horizon.operating_years
    n = c + o

    layout.set_column_widths(ws, label_width=44, it_width=34, year_width=11, unit_width=6)
    layout.write_title_block(
        ws, title_en="Sponsor Equity Returns", title_it="Rendimento dello sponsor",
        subtitle="Project equity CF + IRR + NPV",
    )
    layout.write_scenario_banner(ws, row=3)

    yr_row = 5
    for i in range(n):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        phase = "C" if i < c else "O"
        yr = i + 1 if i < c else (i - c + 1)
        cc = ws.cell(row=yr_row, column=col_idx, value=f"{phase}{yr}")
        styles.style_header(cc)

    capex_row = int(cashflow_refs["capex_row"])
    cads_row = int(cashflow_refs["cads_row"])
    draw_row = int(debt_refs["drawdown_row"])
    ds_row = int(debt_refs["debt_service_row"])

    r = 7
    layout.write_section_header(ws, r, "Sponsor equity cash flow", "CF equity sponsor")
    r += 1

    cf_row = r
    layout.write_row_label(ws, r, "Equity cash flow", "CF equity")
    for i in range(n):
        col = layout.year_col(i); col_idx = ord(col) - ord("A") + 1
        capex_ref = f"'{cashflow_sheet}'!{col}{capex_row}"
        cads_ref = f"'{cashflow_sheet}'!{col}{cads_row}"
        draw_ref = f"'{debt_sheet}'!{col}{draw_row}"
        ds_ref = f"'{debt_sheet}'!{col}{ds_row}"
        # Equity CF = capex (negative) + draw (offsets capex) + CADS + debt service (negative)
        cc = ws.cell(row=r, column=col_idx,
                     value=f"={capex_ref}+{draw_ref}+{cads_ref}+{ds_ref}")
        styles.style_formula(cc, number_format=styles.FMT_EUR_M)
        cc.font = styles.font_subheader
        cc.border = styles.BORDER_TOP_THIN
    r += 2

    # Return metrics
    layout.write_section_header(ws, r, "Return metrics", "Metriche")
    r += 1
    first_col = layout.year_col(0); last_col = layout.year_col(n - 1)

    layout.write_row_label(ws, r, "Equity IRR", "IRR equity", indent=True)
    cc = ws.cell(row=r, column=4, value=f"=IRR(${first_col}${cf_row}:${last_col}${cf_row},0.08)")
    styles.style_formula(cc, number_format=styles.FMT_PCT_2DP)
    cc.font = styles.font_subheader
    r += 1

    layout.write_row_label(ws, r, "Target IRR (sponsor)", "IRR target (sponsor)", indent=True)
    cc = ws.cell(row=r, column=4, value="=target_irr")
    styles.style_xref(cc, number_format=styles.FMT_PCT_2DP)
    r += 1

    layout.write_row_label(ws, r, "Equity NPV @ target IRR", "VAN equity @ IRR target", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=NPV(target_irr,${layout.year_col(1)}${cf_row}:${last_col}${cf_row})+${first_col}${cf_row}")
    styles.style_formula(cc, number_format=styles.FMT_EUR_M)
    r += 1

    layout.write_row_label(ws, r, "Equity MoIC", "MoIC equity", indent=True)
    cc = ws.cell(row=r, column=4,
                 value=f"=IFERROR(SUMIF(${first_col}${cf_row}:${last_col}${cf_row},\">0\")"
                       f"/ABS(SUMIF(${first_col}${cf_row}:${last_col}${cf_row},\"<0\")),0)")
    styles.style_formula(cc, number_format=styles.FMT_MULTIPLE)
    r += 1

    ws.freeze_panes = "D7"
    ws.print_title_rows = "5:5"
    ws.print_title_cols = "A:C"
