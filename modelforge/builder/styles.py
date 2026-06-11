"""Bulge-tier Excel formatting.

Color-coded fonts, number formats, borders. These are the visual rules a
senior MD checks at a glance. Defined once, applied everywhere.

Colour convention (Goldman/Morgan Stanley standard):
    Blue   — hardcoded input
    Black  — formula
    Green  — cross-sheet reference
    Red    — external link, warning, or error

Number formats use accounting conventions with negatives in parentheses.
Sign convention (COSTS NEGATIVE) is enforced at the spec/formula layer,
not here — this module is purely visual.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side


# ─────────────────────────────────────────────────────────────────────────────
# Colours
# ─────────────────────────────────────────────────────────────────────────────

COLOR_INPUT = "0000FF"       # Blue — hardcoded value
COLOR_FORMULA = "000000"     # Black — formula
COLOR_XREF = "006100"        # Green — cross-sheet ref
COLOR_WARNING = "C00000"     # Red — warning / external
COLOR_HEADER_FILL = "1F3864" # Dark navy header fill
COLOR_HEADER_TEXT = "FFFFFF" # White header text
COLOR_CHECK_OK = "C6EFCE"    # Pale green
COLOR_CHECK_BAD = "FFC7CE"   # Pale red
COLOR_SUBHEADER = "D9E1F2"   # Pale navy sub-header
COLOR_HISTORICAL = "FFF2CC"  # Pale yellow — historical (actuals)
COLOR_PROJECTED = "FFFFFF"   # White — projected
COLOR_STATIC_FILL = "F2F2F2" # Light grey — pre-computed static (non-formula) output
COLOR_STATIC_TEXT = "595959" # Dark grey — static-value font


# ─────────────────────────────────────────────────────────────────────────────
# Number formats (bulge-tier)
# ─────────────────────────────────────────────────────────────────────────────

# v0.7: bulge-tier accounting formats with PARENTHESIS negatives
# (Goldman/MS/JPM standard: negatives are shown as (1,234) not -1,234,
# rendered in red for emphasis). Zeros shown as "-" per accounting
# convention. Previous patterns used leading-minus with red — correct
# for code but not bracket-style per TTS/Macabacus.
FMT_EUR_M = '_(* #,##0.0_);[Red]_(* (#,##0.0);_(* "-"??_);_(@_)'      # 1 dp
FMT_EUR_K = '_(* #,##0_);[Red]_(* (#,##0);_(* "-"??_);_(@_)'          # 0 dp
FMT_EUR_ACTUAL = '_(* #,##0_);[Red]_(* (#,##0);_(* "-"??_);_(@_)'
FMT_PCT = "0.0%;[Red](0.0%);0.0%"
FMT_PCT_2DP = "0.00%;[Red](0.00%);0.00%"
FMT_MULTIPLE = "0.00\"x\";[Red](0.00\"x\");\"-\""
FMT_YEARS = "0.0\" y\""
FMT_INTEGER = "#,##0;[Red](#,##0);-"
FMT_BPS = "#,##0\" bps\";[Red](#,##0\" bps\")"
FMT_DATE = "dd-mmm-yyyy"
FMT_YEAR = "yyyy"


# ─────────────────────────────────────────────────────────────────────────────
# Borders
# ─────────────────────────────────────────────────────────────────────────────

_thin = Side(style="thin", color="808080")
_thick = Side(style="medium", color="000000")
_dotted = Side(style="dotted", color="A0A0A0")

BORDER_NONE = Border()
BORDER_TOP_THIN = Border(top=_thin)
BORDER_BOTTOM_THIN = Border(bottom=_thin)
BORDER_TOP_THICK = Border(top=_thick)
BORDER_BOTTOM_THICK = Border(bottom=_thick)
BORDER_BOX = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_HEADER = Border(top=_thick, bottom=_thick)


# ─────────────────────────────────────────────────────────────────────────────
# Fonts
# ─────────────────────────────────────────────────────────────────────────────

FONT_BASE = "Calibri"
FONT_SIZE_BODY = 10
FONT_SIZE_HEADER = 11
FONT_SIZE_TITLE = 14

font_input = Font(name=FONT_BASE, size=FONT_SIZE_BODY, color=COLOR_INPUT)
font_formula = Font(name=FONT_BASE, size=FONT_SIZE_BODY, color=COLOR_FORMULA)
font_xref = Font(name=FONT_BASE, size=FONT_SIZE_BODY, color=COLOR_XREF)
font_warning = Font(name=FONT_BASE, size=FONT_SIZE_BODY, color=COLOR_WARNING)
font_header = Font(
    name=FONT_BASE,
    size=FONT_SIZE_HEADER,
    color=COLOR_HEADER_TEXT,
    bold=True,
)
font_subheader = Font(name=FONT_BASE, size=FONT_SIZE_BODY, bold=True)
font_title = Font(name=FONT_BASE, size=FONT_SIZE_TITLE, bold=True)
font_label_en = Font(name=FONT_BASE, size=FONT_SIZE_BODY, bold=False)
font_label_it = Font(name=FONT_BASE, size=FONT_SIZE_BODY - 1, italic=True, color="666666")
# Static (pre-computed, non-formula) numeric output. Italic + dark grey so a
# reviewer can tell at a glance that the cell is NOT a live formula — it was
# computed in Python and hardcoded. Distinct from blue inputs (which a user
# may legitimately override) and black formulas (which recalc live).
font_static = Font(name=FONT_BASE, size=FONT_SIZE_BODY, italic=True,
                   color=COLOR_STATIC_TEXT)


# ─────────────────────────────────────────────────────────────────────────────
# Fills
# ─────────────────────────────────────────────────────────────────────────────

fill_header = PatternFill("solid", fgColor=COLOR_HEADER_FILL)
fill_subheader = PatternFill("solid", fgColor=COLOR_SUBHEADER)
fill_check_ok = PatternFill("solid", fgColor=COLOR_CHECK_OK)
fill_check_bad = PatternFill("solid", fgColor=COLOR_CHECK_BAD)
fill_historical = PatternFill("solid", fgColor=COLOR_HISTORICAL)
fill_static = PatternFill("solid", fgColor=COLOR_STATIC_FILL)
fill_none = PatternFill()


# ─────────────────────────────────────────────────────────────────────────────
# Alignments
# ─────────────────────────────────────────────────────────────────────────────

align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_center = Alignment(horizontal="center", vertical="center")
align_label = Alignment(horizontal="left", vertical="center", indent=0, wrap_text=False)
align_label_indent = Alignment(horizontal="left", vertical="center", indent=1)


# ─────────────────────────────────────────────────────────────────────────────
# Cell styling helpers
# ─────────────────────────────────────────────────────────────────────────────


def style_input(cell, number_format: str = FMT_EUR_M) -> None:
    cell.font = font_input
    cell.number_format = number_format
    cell.alignment = align_right


def style_formula(cell, number_format: str = FMT_EUR_M) -> None:
    cell.font = font_formula
    cell.number_format = number_format
    cell.alignment = align_right


def style_xref(cell, number_format: str = FMT_EUR_M) -> None:
    cell.font = font_xref
    cell.number_format = number_format
    cell.alignment = align_right


def style_static_value(cell, number_format: str = FMT_EUR_M) -> None:
    """Style a pre-computed, non-formula numeric output.

    For values that were computed in Python and written as a literal (not a
    live ``=`` formula). Uses a distinct light-grey fill + italic dark-grey
    font so a reviewer can immediately distinguish a hardcoded precomputed
    output from a blue user input or a black live formula. Carries an explicit
    number_format so the cell is never an "unstyled numeric" (the certify
    gate's styling-gap check).
    """
    cell.font = font_static
    cell.fill = fill_static
    cell.number_format = number_format
    cell.alignment = align_right


def style_warning(cell, number_format: str = FMT_EUR_M) -> None:
    cell.font = font_warning
    cell.number_format = number_format
    cell.alignment = align_right


def style_header(cell) -> None:
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = BORDER_HEADER


def style_subheader(cell) -> None:
    cell.font = font_subheader
    cell.fill = fill_subheader
    cell.alignment = align_left
    cell.border = BORDER_BOTTOM_THIN


def style_title(cell) -> None:
    cell.font = font_title
    cell.alignment = align_left


def style_label_en(cell) -> None:
    cell.font = font_label_en
    cell.alignment = align_label


def style_label_it(cell) -> None:
    cell.font = font_label_it
    cell.alignment = align_label


def style_check(cell, is_ok: bool = True, number_format: str = FMT_INTEGER) -> None:
    cell.font = Font(name=FONT_BASE, size=FONT_SIZE_BODY, bold=True)
    cell.fill = fill_check_ok if is_ok else fill_check_bad
    cell.number_format = number_format
    cell.alignment = align_center
    cell.border = BORDER_BOX


# ─────────────────────────────────────────────────────────────────────────────
# v0.8.7 US-505 — Macabacus AutoColor parity (post-build pass)
# ─────────────────────────────────────────────────────────────────────────────
#
# Macabacus AutoColor convention:
#   Blue   — hardcoded input
#   Black  — formula (same-sheet)
#   Green  — formula referencing another sheet ( =Sheet!Cell )
#   Red    — external link ( =[Book.xlsx]Sheet!Cell ) / warning
#
# We already tag blue and black at write-time via style_input / style_formula.
# The green xref tagging is hard to retrofit at every write site across the
# codebase, so v0.8.7 does it in a post-build pass that walks the workbook
# once and promotes cross-sheet formula cells to green, and external-link
# cells to red.


def _formula_is_xref(formula: str) -> bool:
    """True if formula references another sheet.

    Looks for a sheet-qualified ref like ``Sheet!A1`` or ``'Named Sheet'!A1``.
    Named-range references without a ``!`` count as same-sheet for colouring.
    External workbook refs (``[Book.xlsx]Sheet!A1``) are handled separately.
    """
    if not formula or not isinstance(formula, str):
        return False
    if not formula.startswith("="):
        return False
    # External link marker
    if "[" in formula and "]" in formula and "!" in formula:
        return False  # external — tagged separately
    return "!" in formula


def _formula_is_external_link(formula: str) -> bool:
    """True if formula references an external workbook."""
    if not formula or not isinstance(formula, str):
        return False
    if not formula.startswith("="):
        return False
    return "[" in formula and "]" in formula and "!" in formula


def auto_color_xrefs(wb) -> int:
    """Walk every cell in every sheet; green cross-sheet refs, red externals.

    Called as the last step of the build pipeline. Idempotent — re-running
    on a coloured workbook produces no change.

    Returns the count of cells promoted to green or red (useful for tests
    and telemetry).
    """
    promoted = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if not isinstance(v, str) or not v.startswith("="):
                    continue
                # Skip if already red (warning) — don't downgrade
                current_color = getattr(c.font, "color", None)
                rgb = None
                if current_color is not None:
                    rgb = getattr(current_color, "rgb", None)
                if isinstance(rgb, str) and rgb.upper().endswith("C00000"):
                    continue

                if _formula_is_external_link(v):
                    c.font = font_warning
                    promoted += 1
                elif _formula_is_xref(v):
                    # Promote if currently black or unset (default). Skip
                    # cells with explicit blue (input) or other colors, so
                    # special styling (warning orange, subheader) is kept.
                    is_black = (
                        isinstance(rgb, str)
                        and rgb.upper().endswith("000000")
                    ) or rgb is None
                    if is_black:
                        is_bold = bool(getattr(c.font, "bold", False))
                        size = c.font.size or FONT_SIZE_BODY
                        italic = bool(getattr(c.font, "italic", False))
                        c.font = Font(
                            name=FONT_BASE,
                            size=size,
                            color=COLOR_XREF,
                            bold=is_bold,
                            italic=italic,
                        )
                        promoted += 1
    return promoted
