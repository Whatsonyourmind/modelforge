"""Shared base workbook builder.

Every template goes through this orchestrator. It emits the 4 shared
sheets (Cover, Sources, Assumptions, QC) and calls back to a
template-specific function for the core sheets in between.

    template                build_core_sheets(wb, spec, graph, driver_refs,
                                              sheet_index=2)
        ↓                     — emits template-specific sheets after
    base orchestrator            Assumptions, returns dict of cross-sheet
                                 refs to feed QC with.
"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, Protocol

from openpyxl import Workbook

from modelforge.builder.sheets import assumptions, cover, sources
from modelforge.graph.schema import LinkageGraph
from modelforge.graph.store import GraphStore


class CoreSheetBuilder(Protocol):
    """Callback signature for template-specific sheet emission."""

    def __call__(
        self,
        wb: Workbook,
        spec,  # BaseModelSpec
        graph: LinkageGraph,
        driver_refs: dict[str, str],
        source_rows: dict[str, int],
    ) -> dict[str, str]:
        """Emit template-specific sheets. Return refs for QC."""
        ...


def build_base_workbook(
    spec,  # BaseModelSpec
    out_path: Path | str,
    core_builder: CoreSheetBuilder,
    graph_db_path: Path | str | None = None,
) -> tuple[Path, Path, LinkageGraph]:
    """Orchestrate the shared workbook skeleton.

    Steps:
        1. Create Workbook.
        2. Emit Cover (defines scenario_index named range).
        3. Emit Sources (returns {S-id: row}).
        4. Emit Assumptions (returns {driver: active_ref}).
        5. Call template's core_builder to emit middle sheets.
        6. Save workbook + graph.

    The core builder is responsible for emitting its own QC sheet since
    each template's checks differ. A shared minimum-QC skeleton is
    accessible via modelforge.qc.runner.run_qc on the saved file.

    Returns (xlsx_path, graph_db_path, graph).
    """
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    if graph_db_path is None:
        graph_db_path = out_path.with_suffix(".graph.db")
    graph_db_path = Path(graph_db_path)

    graph = LinkageGraph(model_id=spec.meta.project_code)

    wb = Workbook()
    wb.remove(wb.active)

    # Cover — must be first; sets scenario_index named range
    cover_ws = wb.create_sheet("Cover")
    cover.build(cover_ws, spec, graph)

    # Sources
    sources_ws = wb.create_sheet("Sources")
    source_rows = sources.build(sources_ws, spec, graph)

    # Assumptions — creates every driver's named range
    assum_ws = wb.create_sheet("Assumptions")
    driver_refs = assumptions.build(assum_ws, spec, graph, source_rows)

    # Template-specific core sheets
    core_builder(wb, spec, graph, driver_refs, source_rows)

    # Save
    wb.save(out_path)

    # Persist graph
    store = GraphStore(graph_db_path)
    store.save(graph)

    return out_path, graph_db_path, graph
