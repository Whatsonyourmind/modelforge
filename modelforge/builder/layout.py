"""Shared layout helpers for sheet builders.

A `SheetCursor` tracks the current row on a sheet as builders emit content.
Column layout is centralised: columns A–B are labels (EN + IT), C onwards
are years or scenario values. Every sheet follows the same skeleton:

    Row 1: Title block
    Row 2: Subtitle / unit / currency / scenario badge
    Row 3: blank
    Row 4+: Check row (QC pings, coloured)
    Row 5: blank
    Row 6+: content
"""

from __future__ import annotations

from dataclasses import dataclass

from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter

from modelforge.builder import styles


# Column layout constants
COL_EN = "A"
COL_IT = "B"
COL_UNIT = "C"
COL_FIRST_YEAR = "D"  # Historical year 1 starts here


def year_col(index_from_zero: int) -> str:
    """Column letter for the Nth year column (0-indexed)."""
    return get_column_letter(ord(COL_FIRST_YEAR) - ord("A") + 1 + index_from_zero)


@dataclass
class SheetCursor:
    ws: Worksheet
    row: int = 1

    def skip(self, n: int = 1) -> None:
        self.row += n

    def here(self) -> int:
        return self.row

    def advance(self) -> int:
        r = self.row
        self.row += 1
        return r


def set_column_widths(ws: Worksheet, label_width: int = 42, it_width: int = 36, year_width: int = 12, unit_width: int = 8) -> None:
    ws.column_dimensions[COL_EN].width = label_width
    ws.column_dimensions[COL_IT].width = it_width
    ws.column_dimensions[COL_UNIT].width = unit_width
    for i in range(10):  # up to 10 year columns by default
        col = year_col(i)
        ws.column_dimensions[col].width = year_width


def write_title_block(ws: Worksheet, title_en: str, title_it: str, subtitle: str = "") -> None:
    c = ws.cell(row=1, column=1, value=title_en)
    styles.style_title(c)
    ws.cell(row=1, column=2, value=title_it).font = styles.font_label_it
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = styles.font_label_it


def write_row_label(ws: Worksheet, row: int, en: str, it: str = "", indent: bool = False) -> None:
    c_en = ws.cell(row=row, column=1, value=en)
    styles.style_label_en(c_en)
    if indent:
        c_en.alignment = styles.align_label_indent
    c_it = ws.cell(row=row, column=2, value=it)
    styles.style_label_it(c_it)


def write_section_header(ws: Worksheet, row: int, en: str, it: str = "", span_cols: int = 12) -> None:
    c = ws.cell(row=row, column=1, value=en)
    styles.style_subheader(c)
    c_it = ws.cell(row=row, column=2, value=it)
    c_it.font = styles.font_subheader
    c_it.fill = styles.fill_subheader
    # Extend the subheader fill across the row
    for i in range(2, span_cols):
        cc = ws.cell(row=row, column=i + 1)
        if cc.value is None:
            cc.fill = styles.fill_subheader


def freeze_panes_at_first_year(ws: Worksheet) -> None:
    ws.freeze_panes = f"{COL_FIRST_YEAR}7"


def write_scenario_banner(ws: Worksheet, row: int = 3) -> None:
    """Bulge-tier: every sheet shows active scenario prominently.

    Emits a 1-row banner:
        A: "ACTIVE SCENARIO →"
        C: formula CHOOSE(scenario_index, "WORST", "BASE", "BEST")

    Conditional colour applied so the banner shifts: amber/green/blue
    depending on scenario_index.
    """
    from openpyxl.formatting.rule import CellIsRule
    from openpyxl.styles import PatternFill, Font

    c_lbl = ws.cell(row=row, column=1, value="ACTIVE SCENARIO →")
    c_lbl.font = Font(name=FONT_BASE_LOCAL, size=10, bold=True, color="555555")
    c_lbl.alignment = styles.align_right

    c_lbl_it = ws.cell(row=row, column=2, value="SCENARIO ATTIVO →")
    c_lbl_it.font = Font(name=FONT_BASE_LOCAL, size=9, italic=True, color="888888")
    c_lbl_it.alignment = styles.align_right

    c_val = ws.cell(
        row=row, column=3,
        value='=CHOOSE(scenario_index,"WORST","BASE","BEST")',
    )
    c_val.font = Font(name=FONT_BASE_LOCAL, size=11, bold=True, color="FFFFFF")
    c_val.alignment = styles.align_center
    c_val.fill = PatternFill("solid", fgColor="1F3864")

    # Conditional colouring — override fill based on scenario_index
    addr = c_val.coordinate
    ws.conditional_formatting.add(
        addr,
        CellIsRule(
            operator="equal", formula=['"WORST"'],
            fill=PatternFill("solid", fgColor="9C0006"),
            font=Font(name=FONT_BASE_LOCAL, size=11, bold=True, color="FFFFFF"),
        ),
    )
    ws.conditional_formatting.add(
        addr,
        CellIsRule(
            operator="equal", formula=['"BASE"'],
            fill=PatternFill("solid", fgColor="1F3864"),
            font=Font(name=FONT_BASE_LOCAL, size=11, bold=True, color="FFFFFF"),
        ),
    )
    ws.conditional_formatting.add(
        addr,
        CellIsRule(
            operator="equal", formula=['"BEST"'],
            fill=PatternFill("solid", fgColor="006100"),
            font=Font(name=FONT_BASE_LOCAL, size=11, bold=True, color="FFFFFF"),
        ),
    )


FONT_BASE_LOCAL = "Calibri"
