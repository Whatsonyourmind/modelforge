"""XLSX reader — one chunk per sheet, flattened as markdown table.

Uses openpyxl (already a ModelForge dep). Output chunks carry
meta={"sheet_name": ..., "nrows": ..., "ncols": ...}.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from modelforge.ingest.readers.base import DocChunk, DocIndex


def read_xlsx(path: Path) -> DocIndex:
    """Return DocIndex with one chunk per sheet.

    Each chunk's text is a markdown table of the sheet's used range
    (rows x cols). Formulas aren't evaluated — we read values. Large
    sheets are capped at 200 rows × 30 cols for context-budget sanity.
    """
    path = Path(path)
    chunks: list[DocChunk] = []
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        for sidx, sheet_name in enumerate(wb.sheetnames, start=1):
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True, max_row=200, max_col=30))
            if not rows:
                continue
            md = _rows_to_markdown(rows)
            chunks.append(DocChunk(
                doc_filename=path.name,
                page=sidx,   # treat sheet index as "page"
                text=f"## Sheet: {sheet_name}\n\n{md}",
                kind="table",
                meta={"sheet_name": sheet_name,
                      "nrows": len(rows),
                      "ncols": max((len(r) for r in rows), default=0)},
            ))
    finally:
        wb.close()

    return DocIndex(
        doc_filename=path.name,
        path=path,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        total_pages=len(chunks),
        chunks=chunks,
    )


def _rows_to_markdown(rows: list[tuple]) -> str:
    """Convert list-of-tuples (openpyxl output) to markdown table.

    Strips fully-empty trailing columns; treats first non-blank row as header.
    """
    # Find first non-empty row as header
    header_idx = 0
    for i, r in enumerate(rows):
        if any(cell is not None and str(cell).strip() for cell in r):
            header_idx = i
            break
    useful = rows[header_idx:]
    if not useful:
        return ""
    # Determine column count
    ncols = max(len(r) for r in useful)
    header = [str(c) if c is not None else "" for c in useful[0]]
    header = header + [""] * (ncols - len(header))
    lines = ["| " + " | ".join(header) + " |",
             "|" + "|".join(["---"] * ncols) + "|"]
    for r in useful[1:]:
        cells = [str(c) if c is not None else "" for c in r]
        cells = cells + [""] * (ncols - len(cells))
        lines.append("| " + " | ".join(cells) + " |")
    return "\n".join(lines)
