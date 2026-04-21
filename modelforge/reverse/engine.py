"""Reverse-engineer engine.

Four passes over a legacy workbook:

    Pass 1 — Sheet classification
        Maps each sheet name + content to a canonical category:
        cover / sources / assumptions / operating / debt / returns /
        valuation / waterfall / tranches / covenants / qc / other.

    Pass 2 — Input extraction
        For each assumption-like sheet, extract the labeled inputs
        (row label in col A/B, numeric value in col C–E or later).
        Heuristic: input = cell with a numeric value adjacent to a
        non-empty string label.

    Pass 3 — Formula clustering
        Redacts cell refs + literals to shape-patterns and tallies
        most-frequent shapes per sheet (same technique as the
        dossier's formula inventory).

    Pass 4 — Template detection
        Matches sheet-kind frequencies against each ModelForge
        template's signature to suggest the closest fit.

The spec skeleton emitter produces a YAML with meta + target
placeholders + whatever assumptions the extractor could recover.
A human fills the gaps using the REVERSE_REPORT.md.
"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ─── Data classes ───────────────────────────────────────────────────────────


@dataclass
class ExtractedInput:
    sheet: str
    cell: str
    label: str
    value: float


@dataclass
class SheetProfile:
    name: str
    kind: str                # "cover" | "assumptions" | "operating" | ...
    rows: int
    cols: int
    formula_cells: int
    hardcoded_cells: int
    top_formula_shapes: list[tuple[str, int]] = field(default_factory=list)


@dataclass
class ReverseReport:
    path: Path
    sheets: list[SheetProfile]
    inputs: list[ExtractedInput]
    named_ranges: dict[str, str]
    detected_template: str          # best-guess model_type
    template_confidence: float      # 0..1
    template_scores: dict[str, float] = field(default_factory=dict)
    notes: list[str] = field(default_factory=list)

    @property
    def n_inputs(self) -> int:
        return len(self.inputs)


# ─── Sheet classification ──────────────────────────────────────────────────


_SHEET_KIND_PATTERNS: list[tuple[str, str]] = [
    (r"^cover$|^title$|^front$",                  "cover"),
    (r"^sources?$|^bibliography$|^citations?$",   "sources"),
    (r"^assumption|^inputs?$|^drivers?$",         "assumptions"),
    (r"^operat|income|p.{0,2}l|profit|3.statement|model$|^fcf",
                                                  "operating"),
    (r"debt|loan|facility|bond|tranche|capital.structure|financing",
                                                  "debt"),
    (r"covenant",                                 "covenants"),
    (r"returns?|irr|moic|accret",                 "returns"),
    (r"valuation|dcf|wacc|football|range",        "valuation"),
    (r"waterfall|distribution|collection",        "waterfall"),
    (r"qc|check|test|audit",                      "qc"),
    (r"sensitivity|tornado|scenario",             "sensitivity"),
    (r"monte.?carlo|simulation",                  "monte_carlo"),
    (r"repro|reproduc|meta",                      "metadata"),
]


def classify_sheet(name: str) -> str:
    """Classify a sheet by name against canonical categories."""
    low = name.lower().strip()
    for pattern, kind in _SHEET_KIND_PATTERNS:
        if re.search(pattern, low):
            return kind
    return "other"


def _profile_sheet(ws: Worksheet) -> SheetProfile:
    kind = classify_sheet(ws.title)
    formula_cells = 0
    hardcoded_cells = 0
    patterns: Counter = Counter()
    for row in ws.iter_rows():
        for c in row:
            v = c.value
            if isinstance(v, str) and v.startswith("="):
                formula_cells += 1
                p = re.sub(r"[A-Z]+\d+", "CELL", v)
                p = re.sub(r"\d+(\.\d+)?", "N", p)
                patterns[p[:120]] += 1
            elif isinstance(v, (int, float)):
                hardcoded_cells += 1
    return SheetProfile(
        name=ws.title, kind=kind,
        rows=ws.max_row, cols=ws.max_column,
        formula_cells=formula_cells, hardcoded_cells=hardcoded_cells,
        top_formula_shapes=patterns.most_common(5),
    )


# ─── Input extraction ───────────────────────────────────────────────────────


def _extract_inputs(ws: Worksheet, profile: SheetProfile) -> list[ExtractedInput]:
    """Only mine sheets likely to hold assumptions for inputs."""
    if profile.kind not in ("assumptions", "operating", "debt", "valuation"):
        return []
    out: list[ExtractedInput] = []
    for row in ws.iter_rows(max_col=min(ws.max_column, 20)):
        # Find the label: first non-empty string cell in cols A..B
        label: Optional[str] = None
        for c in row[:2]:
            if isinstance(c.value, str) and c.value.strip():
                label = c.value.strip()
                break
        if not label:
            continue
        # Find numeric input values in subsequent columns
        for c in row[2:]:
            if isinstance(c.value, (int, float)) and not isinstance(c.value, bool):
                out.append(ExtractedInput(
                    sheet=ws.title, cell=c.coordinate,
                    label=label, value=float(c.value),
                ))
                # Only take the first numeric per row
                break
    return out


# ─── Template type detection ───────────────────────────────────────────────


# Signature of "what a ModelForge template looks like" in kind counts
_TEMPLATE_SIGNATURES: dict[str, dict[str, float]] = {
    "unitranche": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "debt": 1, "covenants": 1, "returns": 1, "qc": 1,
    },
    "credit_memo": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "debt": 1, "covenants": 1, "returns": 1, "qc": 1,
    },
    "minibond": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "debt": 1, "covenants": 1, "returns": 1, "qc": 1,
    },
    "project_finance": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "debt": 1, "returns": 1, "qc": 1,
    },
    "real_estate": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "debt": 1, "waterfall": 1, "qc": 1,
    },
    "npl": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "waterfall": 1, "qc": 1,
    },
    "structured_credit": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "debt": 1, "qc": 1,       # "Tranches" classified as debt
    },
    "three_statement": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "qc": 1,
    },
    "dcf": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "operating": 1, "valuation": 1, "qc": 1,
    },
    "merger": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "debt": 1, "operating": 1, "returns": 1, "qc": 1,
    },
    "fairness": {
        "cover": 1, "sources": 1, "assumptions": 1,
        "valuation": 1, "qc": 1,
    },
}


# Sheet kinds that strongly distinguish one template from another.
# Bonus score when the observed workbook has these *and* the template
# signature mentions them; penalty when absent.
_DISTINCTIVE_KINDS_BY_TEMPLATE: dict[str, set[str]] = {
    "dcf":        {"valuation"},
    "fairness":   {"valuation"},
    "npl":        {"waterfall"},
    "real_estate": {"waterfall"},
    "merger":     {"returns"},  # accretion/dilution proxy
}


def _sheet_name_boost(sheets: list[SheetProfile]) -> dict[str, float]:
    """Boost certain templates based on sheet name patterns unique to them."""
    names = [s.name.lower() for s in sheets]
    boosts: dict[str, float] = {}
    # Each boost adds up to +0.10 to a template's final score.
    if any("football" in n or "football_field" in n for n in names):
        boosts["fairness"] = boosts.get("fairness", 0) + 0.25
    if any("wacc" in n for n in names):
        boosts["dcf"] = boosts.get("dcf", 0) + 0.25
    if any("fcfforecast" in n or "fcf_forecast" in n or "fcf forecast" in n
           for n in names):
        boosts["dcf"] = boosts.get("dcf", 0) + 0.10
    # Real estate: Financing + DCF (NOI/waterfall-style) pattern distinct
    # from senior-debt-schedule templates
    has_financing = any(n == "financing" or "financing" in n for n in names)
    has_dcf_sheet = any(n == "dcf" for n in names)
    if has_financing and has_dcf_sheet:
        boosts["real_estate"] = boosts.get("real_estate", 0) + 0.45
    elif has_financing and not any("tranche" in n or "debtschedule" in n.replace(" ", "")
                                    or "bondstructure" in n.replace(" ", "") for n in names):
        boosts["real_estate"] = boosts.get("real_estate", 0) + 0.15
    # Penalize structured_credit when no tranche-like sheet is present
    if not any("tranche" in n for n in names):
        boosts["structured_credit"] = boosts.get("structured_credit", 0) - 0.20
    if any("accretion" in n or "acc_dil" in n for n in names):
        boosts["merger"] = boosts.get("merger", 0) + 0.10
    # v0.8.7 US-550: stronger merger signals to outweigh the v0.7
    # enrichment sheets (ComparableBetas, ComplianceCheck, PPA) that
    # made three_statement score high on enriched merger workbooks.
    has_dealstruct = any("dealstructure" in n.replace(" ", "") for n in names)
    has_proforma = any(n == "proforma" or "pro_forma" in n for n in names)
    has_accdil = any("accretiondilution" in n.replace(" ", "")
                      or "accretion" in n for n in names)
    if has_dealstruct and has_proforma and has_accdil:
        # Full merger triad — unambiguous signal
        boosts["merger"] = boosts.get("merger", 0) + 0.30
        # And suppress three_statement since the operating model here is
        # a pro-forma combination, not a standalone 3-stmt.
        boosts["three_statement"] = boosts.get("three_statement", 0) - 0.20
    elif has_dealstruct or has_proforma:
        boosts["merger"] = boosts.get("merger", 0) + 0.10
    if any("investorreturn" in n.replace(" ", "") for n in names):
        boosts["minibond"] = boosts.get("minibond", 0) + 0.10
    if any("bondstructure" in n.replace(" ", "") for n in names):
        boosts["minibond"] = boosts.get("minibond", 0) + 0.10
    if any("projectcashflow" in n.replace(" ", "") or "dscr" in n for n in names):
        boosts["project_finance"] = boosts.get("project_finance", 0) + 0.10
    if any("sponsorequity" in n.replace(" ", "")
           or "equityreturns" in n.replace(" ", "") for n in names):
        boosts["project_finance"] = boosts.get("project_finance", 0) + 0.05
    if any("tranche" in n for n in names):
        boosts["structured_credit"] = boosts.get("structured_credit", 0) + 0.10
    if any("collectionwaterfall" in n.replace(" ", "") for n in names):
        boosts["npl"] = boosts.get("npl", 0) + 0.10
    if any("covenant" in n for n in names):
        # Covenants are on unitranche/credit_memo/minibond — subtle lift
        boosts["unitranche"] = boosts.get("unitranche", 0) + 0.03
        boosts["credit_memo"] = boosts.get("credit_memo", 0) + 0.03
    if any("creditopinion" in n.replace(" ", "") for n in names):
        boosts["credit_memo"] = boosts.get("credit_memo", 0) + 0.10
    return boosts


def detect_template_type(
    sheets: list[SheetProfile],
) -> tuple[str, float, dict[str, float]]:
    """Score each template signature against observed sheet kinds, then
    apply sheet-name-specific boosts to break ties (e.g. a "WACCBuild"
    sheet strongly suggests DCF).

    Returns (best_template_type, confidence, full_scores).
    """
    observed = Counter(p.kind for p in sheets)
    scores: dict[str, float] = {}
    for name, sig in _TEMPLATE_SIGNATURES.items():
        score = 0.0
        for kind, weight in sig.items():
            if observed.get(kind, 0) > 0:
                score += weight
        max_score = sum(sig.values())
        base = score / max_score if max_score > 0 else 0.0
        # Distinctive-kind requirement: if a template REQUIRES valuation
        # (DCF / fairness) and observed has none, cap at 60%.
        required = _DISTINCTIVE_KINDS_BY_TEMPLATE.get(name, set())
        if required and not any(observed.get(k, 0) > 0 for k in required):
            base = min(base, 0.60)
        scores[name] = base

    # Apply sheet-name boosts (allow scores > 1.0 so distinctive
    # features break ties even when multiple templates reach base=1.0)
    for k, boost in _sheet_name_boost(sheets).items():
        scores[k] = scores.get(k, 0.0) + boost

    best = max(scores, key=scores.get)
    # Confidence display capped at 1.0 for the primary return, but
    # internal scores preserved for the full table.
    confidence = min(scores[best], 1.0)
    return best, confidence, scores


# ─── Public API ────────────────────────────────────────────────────────────


def analyze_workbook(path: Path | str) -> ReverseReport:
    path = Path(path)
    wb = load_workbook(path, data_only=False, keep_links=True)

    sheets: list[SheetProfile] = []
    inputs: list[ExtractedInput] = []
    notes: list[str] = []

    for name in wb.sheetnames:
        ws = wb[name]
        profile = _profile_sheet(ws)
        sheets.append(profile)
        inputs.extend(_extract_inputs(ws, profile))

    named_ranges = {n: dn.attr_text for n, dn in wb.defined_names.items()}
    template, conf, scores = detect_template_type(sheets)

    if len(inputs) < 5:
        notes.append("Very few numeric inputs detected — the workbook may use "
                     "external data refs or all-formula cells.")
    if not named_ranges:
        notes.append("No workbook-level named ranges — ModelForge conventions "
                     "suggest defining them for every driver.")
    if conf < 0.5:
        notes.append(f"Low confidence template match ({conf:.0%}); manual "
                     "review recommended.")

    return ReverseReport(
        path=path, sheets=sheets, inputs=inputs,
        named_ranges=named_ranges,
        detected_template=template, template_confidence=conf,
        template_scores=scores, notes=notes,
    )


def render_markdown(rep: ReverseReport) -> str:
    out = [f"# Reverse-engineering report — {rep.path.name}", ""]
    out.append(f"**Detected template type:** `{rep.detected_template}` "
               f"(confidence {rep.template_confidence:.0%})")
    out.append("")

    out.append("## Template-match scores")
    out.append("")
    out.append("| Template | Score |")
    out.append("|---|---|")
    for k, v in sorted(rep.template_scores.items(), key=lambda x: x[1], reverse=True):
        out.append(f"| {k} | {v:.0%} |")
    out.append("")

    out.append("## Sheet analysis")
    out.append("")
    out.append("| Sheet | Kind | Rows × Cols | Formulas | Inputs |")
    out.append("|---|---|---|---|---|")
    for p in rep.sheets:
        out.append(f"| {p.name} | `{p.kind}` | {p.rows} × {p.cols} | "
                   f"{p.formula_cells} | {p.hardcoded_cells} |")
    out.append("")

    out.append(f"## Named ranges ({len(rep.named_ranges)})")
    out.append("")
    if rep.named_ranges:
        out.append("| Name | Attr |")
        out.append("|---|---|")
        for n, a in sorted(rep.named_ranges.items())[:40]:
            out.append(f"| `{n}` | `{a}` |")
        if len(rep.named_ranges) > 40:
            out.append(f"| ... | *(+{len(rep.named_ranges) - 40} more)* |")
    else:
        out.append("*(none — ModelForge convention: define workbook-level names for every driver)*")
    out.append("")

    out.append(f"## Extracted inputs ({rep.n_inputs})")
    out.append("")
    if rep.inputs:
        out.append("| Sheet | Cell | Label | Value |")
        out.append("|---|---|---|---|")
        for inp in rep.inputs[:50]:
            out.append(f"| {inp.sheet} | {inp.cell} | {inp.label} | {inp.value} |")
        if rep.n_inputs > 50:
            out.append(f"| ... | *(+{rep.n_inputs - 50} more — see spec skeleton)* |")
    else:
        out.append("*(no inputs extracted)*")
    out.append("")

    if rep.notes:
        out.append("## Notes / caveats")
        out.append("")
        for n in rep.notes:
            out.append(f"- {n}")
        out.append("")

    out.append("## Next steps")
    out.append("")
    out.append(f"1. Review the spec skeleton output (write via `--spec-out`).")
    out.append(f"2. Fill in sources, meta, target from the original workbook's cover.")
    out.append(f"3. Run `modelforge build <spec.yaml>` to emit a ModelForge-native workbook.")
    out.append(f"4. Compare with original using `modelforge diff original.xlsx new.xlsx`.")
    return "\n".join(out)


def render_spec_skeleton(rep: ReverseReport) -> str:
    """Emit a partial YAML spec — the user must complete meta + target
    + rationales + source IDs. Assumptions are prefilled from inputs.
    """
    import yaml as _yaml
    assums: list[dict] = []
    seen: set[str] = set()
    for i, inp in enumerate(rep.inputs[:30], start=1):
        # Generate a snake_case name from the label
        name = re.sub(r"[^a-z0-9_]+", "_", inp.label.lower()).strip("_") or f"driver_{i}"
        name = name[:40]
        # De-duplicate
        if name in seen:
            name = f"{name}_{i}"
        seen.add(name)
        assums.append({
            "id": f"A-{i:03d}",
            "name": name,
            "label": {"en": inp.label[:80], "it": inp.label[:80]},
            "unit": "eur_m" if abs(inp.value) > 1 else "pct",
            "base": round(float(inp.value), 6),
            "rationale": f"(auto-extracted from {inp.sheet}!{inp.cell}; "
                         "fill in rationale from original analyst notes)",
            "confidence": "M",
        })

    skeleton = {
        "model_type": rep.detected_template,
        "meta": {
            "project_code": f"REVERSED-{rep.path.stem.upper()[:20]}",
            "deliverable": {"en": "Reverse-engineered spec (TO COMPLETE)",
                            "it": "Spec ricostruita (DA COMPLETARE)"},
            "analyst": "(analyst name)",
            "version": "v0.1-reversed",
            "status": "draft",
            "valuation_date": "2026-04-17",
            "currency": "EUR",
            "unit_scale": "millions",
            "sign_convention": "costs_negative",
            "revision_log": [
                {"version": "v0.1-reversed",
                 "date": "2026-04-17",
                 "analyst": "(analyst)",
                 "note": f"Reverse-engineered from {rep.path.name}"},
            ],
        },
        "target": {
            "name": "(target name)",
            "sector": {"en": "(sector)", "it": "(settore)"},
            "country": "IT",
            "currency": "EUR",
            "revenue_last_fy_eur_m": 0.0,
            "revenue_source_id": "S-001",
            "ebitda_last_fy_eur_m": 0.0,
            "ebitda_source_id": "S-001",
            "last_fy_end": "2025-12-31",
        },
        "sources": [
            {"id": "S-001", "doc": f"{rep.path.name}",
             "publisher": "(unknown)", "date": "2026-04-17",
             "verified": False,
             "note": "Original source of reverse-engineered inputs"},
        ],
        "_reverse_engineering_notes": {
            "source_workbook": rep.path.name,
            "detected_template": rep.detected_template,
            "template_confidence": rep.template_confidence,
            "inputs_extracted": rep.n_inputs,
            "hint": "Inputs below are auto-extracted; rename/regroup as needed "
                    "per template sub-model (operating / debt / covenants / etc). "
                    "Delete _reverse_engineering_notes before `modelforge build`.",
        },
        "_extracted_assumptions": assums,
    }
    return _yaml.safe_dump(skeleton, sort_keys=False, allow_unicode=True,
                           default_flow_style=False)
