"""Conservation gate — recompute the in-workbook QC checks and require ALL_PASS.

``certify`` proves a workbook has **zero formula errors** (no ``#REF!``/``#DIV0!``
…). It says nothing about a conservation check that legitimately *recomputes to
0* — a balance sheet that does not balance, a debt roll-forward that telescopes,
``NI != EBT + tax`` — because such a cell is a perfectly valid formula returning
``0``, not an Excel error. So a model can FAIL its own QC sheet and still certify
CERTIFIED today.

This auditor closes that gap: it recomputes every formula with the third-party
``formulas`` engine (via :func:`recompute_cell_values`), reads the **recomputed**
``ALL CHECKS PASS`` aggregator plus every per-check result on the QC sheet, and
requires the aggregator to evaluate to ``1`` with no per-check evaluating to
``0``. It is deterministic, free/local, and zero-LLM — the property is
machine-checkable, not LM-judged. Promoting recomputed economic conservation to
a PASS/FAIL gate is the contribution; the underlying invariants are the model's
own QC checks.

Scope: strength is bounded by each template's QC coverage (it gates exactly the
checks the model declares). Wired opt-in (``certify --conservation``) so it never
changes the default ``certify`` verdict for existing users. Missing recalc engine
or no QC sheet → INDETERMINATE, which the gate treats as a failure.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from modelforge.qc.workbook_audit import _coerce_error, recompute_cell_values

# i18n / template variants of the aggregator label, all in column A of the QC sheet.
_ALL_PASS_LABELS = {
    "ALL CHECKS PASS",
    "TUTTI I CONTROLLI OK",
    "ALLE PRÜFUNGEN BESTANDEN",
    "ALLE PRUFUNGEN BESTANDEN",
}


@dataclass
class ConservationFinding:
    """A per-check (or aggregator) that did not recompute to a pass."""

    ref: str
    label: str
    status: str  # "fail" (recomputed 0) or "error" (check formula errored)
    recomputed: object


@dataclass
class ConservationAuditReport:
    workbook: str
    recalc_ran: bool = False
    qc_sheet: str | None = None
    all_pass_ref: str | None = None
    all_pass_value: object = None
    n_checks: int = 0
    findings: list[ConservationFinding] = field(default_factory=list)
    status: str = "INDETERMINATE"  # PASS / FAIL / INDETERMINATE
    notes: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        """True only on a clean PASS. INDETERMINATE is a gate failure."""
        return self.status == "PASS"

    @property
    def n_findings(self) -> int:
        return len(self.findings)

    @property
    def verdict(self) -> str:
        return self.status

    def summary(self) -> dict:
        return {
            "status": self.status,
            "qc_sheet": self.qc_sheet,
            "all_pass_ref": self.all_pass_ref,
            "all_pass_value": self.all_pass_value,
            "n_checks": self.n_checks,
            "n_findings": self.n_findings,
            "recalc_ran": self.recalc_ran,
        }

    def print(self) -> None:  # pragma: no cover - console formatting
        from rich.console import Console
        from rich.table import Table

        console = Console()
        if self.status == "PASS":
            badge = "[bold green]CONSERVATION PASS[/bold green]"
        elif self.status == "FAIL":
            badge = "[bold red]CONSERVATION FAIL[/bold red]"
        else:
            badge = "[bold yellow]CONSERVATION INDETERMINATE[/bold yellow]"
        console.print(
            f"{badge}  {Path(self.workbook).name}  "
            f"checks={self.n_checks}  failing={self.n_findings}  "
            f"(recalc {'ran' if self.recalc_ran else 'skipped'})"
        )
        if self.findings:
            tbl = Table(title="Conservation findings (certify-blind)")
            tbl.add_column("Cell", style="bold")
            tbl.add_column("Check")
            tbl.add_column("Recomputed")
            for f in self.findings[:20]:
                tbl.add_row(f.ref, f.label, f"{f.recomputed} ({f.status})")
            console.print(tbl)
            if self.n_findings > 20:
                console.print(f"[dim]… and {self.n_findings - 20} more.[/dim]")
        for n in self.notes:
            console.print(f"[dim]note: {n}[/dim]")


def _classify(value: object) -> tuple[str, float | None]:
    """Map a recomputed check value to (status, numeric).

    status ∈ {"pass", "fail", "error", "unknown"}. A check passes at ~1, fails at
    ~0; an Excel error → "error"; anything else → "unknown".
    """
    if _coerce_error(value) is not None:
        return "error", None
    try:
        f = float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return "unknown", None
    if abs(f - 1.0) < 0.5:
        return "pass", f
    if abs(f - 0.0) < 0.5:
        return "fail", f
    return "unknown", f


def _find_qc_sheet(wb):
    """Locate the QC sheet + the ALL_PASS aggregator cell by its column-A label.

    Returns (sheet_name, anchor_row, anchor_col) or (None, None, None).
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip().upper() in _ALL_PASS_LABELS:
                    return ws.title, c.row, c.column
    return None, None, None


def audit_conservation(xlsx_path: Path | str) -> ConservationAuditReport:
    """Recompute the QC sheet and require ALL_PASS==1 with no per-check==0.

    ``report.passed`` is True only on a clean PASS; FAIL (a check recomputed 0 or
    errored) and INDETERMINATE (no recalc engine / no QC sheet / non-numeric
    aggregator) both return False so the gate fails closed.
    """
    path = Path(xlsx_path)
    report = ConservationAuditReport(workbook=str(path))

    values, recalc_ran, notes = recompute_cell_values(path)
    report.recalc_ran = recalc_ran
    report.notes.extend(notes)

    if not recalc_ran:
        report.notes.append("Recalculation engine unavailable — conservation INDETERMINATE.")
        report.status = "INDETERMINATE"
        return report

    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:  # pragma: no cover
        report.notes.append(f"Could not open workbook: {e}")
        report.status = "INDETERMINATE"
        return report

    qc_sheet, anchor_row, anchor_col = _find_qc_sheet(wb)
    if qc_sheet is None:
        report.notes.append("No QC sheet (no 'ALL CHECKS PASS' aggregator found) — INDETERMINATE.")
        report.status = "INDETERMINATE"
        return report

    report.qc_sheet = qc_sheet
    ws = wb[qc_sheet]
    result_col = anchor_col + 2  # label in col A → 1/0 result in col C

    # (1) The ALL_PASS aggregator.
    all_pass_ref = f"{qc_sheet}!{ws.cell(row=anchor_row, column=result_col).coordinate}"
    report.all_pass_ref = all_pass_ref
    all_pass_val = values.get(all_pass_ref)
    report.all_pass_value = all_pass_val
    ap_status, _ = _classify(all_pass_val)

    # (2) Every per-check: a labelled row whose result cell holds a formula.
    for r in range(anchor_row + 1, ws.max_row + 1):
        label_cell = ws.cell(row=r, column=anchor_col)
        result_cell = ws.cell(row=r, column=result_col)
        label = label_cell.value
        formula = result_cell.value
        if not (isinstance(label, str) and label.strip()):
            continue
        if not (isinstance(formula, str) and formula.startswith("=")):
            continue
        report.n_checks += 1
        ref = f"{qc_sheet}!{result_cell.coordinate}"
        c_status, c_val = _classify(values.get(ref))
        if c_status in ("fail", "error"):
            report.findings.append(
                ConservationFinding(ref=ref, label=label.strip(), status=c_status, recomputed=values.get(ref))
            )

    # (3) Verdict.
    if ap_status in ("error", "unknown") or all_pass_val is None:
        report.notes.append(
            f"ALL_PASS aggregator {all_pass_ref} did not recompute to a 0/1 value "
            f"(got {all_pass_val!r}) — INDETERMINATE."
        )
        report.status = "INDETERMINATE"
    elif ap_status == "pass" and not report.findings:
        report.status = "PASS"
    else:
        report.status = "FAIL"
        if ap_status == "fail":
            report.findings.insert(
                0,
                ConservationFinding(
                    ref=all_pass_ref, label="ALL CHECKS PASS (aggregator)",
                    status="fail", recomputed=all_pass_val,
                ),
            )

    return report
