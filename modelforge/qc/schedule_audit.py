"""Schedule auditor — the certify-blind "hardcode wedged in a formula series" gate.

``certify`` (see :mod:`modelforge.qc.workbook_audit`) proves a workbook has
**zero formula errors** — every cell that *is* a formula evaluates cleanly. It
says nothing about a cell that *should* be a formula but is a bare number.

The classic, certify-invisible defect: an analyst types a hardcoded value over
one period of an otherwise-formula projection row — e.g. year 4 of a 10-year
debt schedule becomes ``1234`` instead of ``=opening+draw+repay``. The number
is valid, the workbook still "has no #REF!", and certification passes — but the
roll-forward is silently broken from that period on.

This module flags exactly that: a **non-innocuous hardcoded number that sits
between formula cells inside a contiguous period series** ("interior hardcode").
It is deliberately high-precision:

* legitimate first-/last-period inputs live at the *edge* of the series (no
  formula on one side) and are NOT flagged;
* innocuous literals (0, 1, 12, 100, …) are skipped;
* input / reference / audit sheets — where hardcoded numbers are expected — are
  excluded via :func:`modelforge.moat.classifier.classify_sheet`.

It is **advisory** (a REVIEW signal, never a hard FAIL on its own); a caller may
opt in to gating with ``audit-schedule --strict`` or a future ``certify
--schedule`` flag. The module is read-only and never mutates the workbook.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from modelforge.moat.classifier import classify_sheet
from modelforge.moat.gate import _INNOCUOUS

# Sheets where hardcoded numbers are legitimate (inputs, data tables, meta).
_EXCLUDED_SHEET_CLASSES = {"input", "reference", "audit"}

# A real period series needs at least this many contiguous numeric/formula
# cells — below it, "between two formulas" is not meaningfully a series.
_MIN_BAND = 3

# Roll-forward check: row-label patterns + how far below an opening row its
# paired closing row may sit, and a cell-reference parser.
_OPENING_RE = re.compile(r"\b(opening|beginning|brought[ _-]?forward)\b|\bbo[pf]\b|\bb/?f\b", re.I)
_CLOSING_RE = re.compile(r"\b(closing|ending|carried[ _-]?forward)\b|\beo[pf]\b|\bc/?f\b", re.I)
_REF_RE = re.compile(r"\$?([A-Z]{1,3})\$?(\d+)")
_BLOCK_WINDOW = 12  # max rows between an opening row and its paired closing row


@dataclass
class ScheduleFinding:
    sheet: str
    cell: str
    row: int
    detail: str
    kind: str = "interior_hardcode"  # or "rollforward"
    value: float | None = None

    @property
    def ref(self) -> str:
        return f"{self.sheet}!{self.cell}"


@dataclass
class ScheduleAuditReport:
    workbook: str
    findings: list[ScheduleFinding] = field(default_factory=list)
    sheets_scanned: int = 0
    sheets_skipped: list[str] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    @property
    def n_findings(self) -> int:
        return len(self.findings)

    @property
    def passed(self) -> bool:
        """Clean == no interior-hardcode findings."""
        return self.n_findings == 0

    @property
    def verdict(self) -> str:
        # Advisory by design: CLEAN or REVIEW (never a hard FAIL on its own).
        return "CLEAN" if self.passed else "REVIEW"

    def summary(self) -> dict:
        return {
            "workbook": self.workbook,
            "verdict": self.verdict,
            "passed": self.passed,
            "findings": self.n_findings,
            "sheets_scanned": self.sheets_scanned,
        }

    def print(self) -> None:
        """Render a human-readable report (rich if available, else plain)."""
        try:
            from rich.console import Console
            from rich.table import Table

            console = Console()
            badge = "[bold green]CLEAN[/bold green]" if self.passed \
                else "[bold yellow]REVIEW[/bold yellow]"
            console.print(
                f"{badge}  {Path(self.workbook).name}  "
                f"({self.n_findings} finding(s), {self.sheets_scanned} sheet(s) scanned)"
            )
            if self.findings:
                tbl = Table(title="Schedule findings (certify-blind)")
                tbl.add_column("cell")
                tbl.add_column("kind")
                tbl.add_column("detail")
                for f in self.findings:
                    tbl.add_row(f.ref, f.kind, f.detail)
                console.print(tbl)
            for n in self.notes:
                console.print(f"  note: {n}")
        except ImportError:  # pragma: no cover - rich is a project dep
            print(f"{self.verdict}  {Path(self.workbook).name}  "
                  f"({self.n_findings} finding(s))")
            for f in self.findings:
                print(f"  {f.ref} [{f.kind}] — {f.detail}")


# ─────────────────────────────────────────────────────────────────────────────
# Internals
# ─────────────────────────────────────────────────────────────────────────────


def _is_formula(v) -> bool:
    return isinstance(v, str) and v.startswith("=")


def _is_number(v) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


def _is_innocuous(n: float) -> bool:
    """Mirror the moat gate's innocuous-literal set (0/1/12/100/…)."""
    if n in _INNOCUOUS:
        return True
    if isinstance(n, float) and n.is_integer() and int(n) in _INNOCUOUS:
        return True
    return False


def _scan_sheet(ws, report: ScheduleAuditReport) -> None:
    """Flag interior hardcodes within each row's contiguous period series."""
    for row in ws.iter_rows():
        # Collect numeric/formula cells in column order. Text/blank cells are
        # band breakers (they are simply not collected, creating a column gap).
        cells: list[tuple[int, str, str, float | str]] = []
        for c in row:
            v = c.value
            if _is_formula(v):
                cells.append((c.column, c.coordinate, "formula", v))
            elif _is_number(v):
                cells.append((c.column, c.coordinate, "number", float(v)))
        if len(cells) < _MIN_BAND:
            continue

        # Split into contiguous bands (consecutive column indices). A gap of a
        # blank/text column (column delta > 1) starts a new band.
        bands: list[list[tuple[int, str, str, float | str]]] = []
        band = [cells[0]]
        for prev, cur in zip(cells, cells[1:]):
            if cur[0] == prev[0] + 1:
                band.append(cur)
            else:
                bands.append(band)
                band = [cur]
        bands.append(band)

        for band in bands:
            if len(band) < _MIN_BAND:
                continue
            if not any(k == "formula" for (_, _, k, _) in band):
                continue
            for i, (col, coord, kind, val) in enumerate(band):
                if kind != "number" or _is_innocuous(val):
                    continue
                has_formula_left = any(band[j][2] == "formula" for j in range(i))
                has_formula_right = any(
                    band[j][2] == "formula" for j in range(i + 1, len(band))
                )
                if has_formula_left and has_formula_right:
                    report.findings.append(
                        ScheduleFinding(
                            sheet=ws.title,
                            cell=coord,
                            row=int(coord_row(coord)),
                            value=float(val),
                            detail=(
                                f"hardcoded {val!r} sits between formula cells in a "
                                f"{len(band)}-cell period series — likely a number "
                                f"typed over a formula"
                            ),
                        )
                    )


def coord_row(coordinate: str) -> int:
    """Row number from a cell coordinate like 'E12' -> 12."""
    digits = "".join(ch for ch in coordinate if ch.isdigit())
    return int(digits) if digits else 0


def _row_label(ws, r: int) -> str:
    """First non-empty text label in columns A/B of row r."""
    for col in (1, 2):
        v = ws.cell(row=r, column=col).value
        if isinstance(v, str) and v.strip():
            return v
    return ""


def _period_cols(ws, r: int) -> list[int]:
    """Ordered column indices in row r holding a formula or a number."""
    cols: list[int] = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=r, column=c).value
        if _is_formula(v) or _is_number(v):
            cols.append(c)
    return cols


def _refs_to_row(formula: str, target_row: int) -> set[int]:
    """Column indices of cell references to ``target_row`` in ``formula``.

    Ignores cross-sheet references (``Sheet!$D$11``) — those are not a
    same-sheet roll-forward — so a coincidental row-number match on another
    sheet cannot cause a false positive.
    """
    cols: set[int] = set()
    for m in _REF_RE.finditer(formula):
        if formula[max(0, m.start() - 1):m.start()] == "!":
            continue  # cross-sheet reference
        if int(m.group(2)) == target_row:
            try:
                cols.add(column_index_from_string(m.group(1)))
            except ValueError:
                continue
    return cols


def _scan_rollforward(ws, report: ScheduleAuditReport) -> None:
    """Flag opening rows whose period-t cell references the closing row at the
    WRONG prior period (roll-forward period drift) — the bug class certify is
    blind to because every cell involved is still a valid formula.

    High-precision: a finding is raised ONLY when an opening cell positively
    references its paired closing row at a column that is NOT the immediately
    prior period. Opening cells that are hardcodes (handled by the interior
    check), or that reference the closing row at the correct prior column, or
    that don't reference the closing row by coordinate at all (named ranges,
    cross-sheet), are never flagged.
    """
    max_row = ws.max_row
    openings: list[int] = []
    closings: list[int] = []
    for r in range(1, max_row + 1):
        lab = _row_label(ws, r)
        if not lab:
            continue
        if _OPENING_RE.search(lab):
            openings.append(r)
        elif _CLOSING_RE.search(lab):
            closings.append(r)

    for r_o in openings:
        # Pair with the nearest closing row below, within the block window.
        r_c = next((rc for rc in closings if r_o < rc <= r_o + _BLOCK_WINDOW), None)
        if r_c is None:
            continue
        pcols = _period_cols(ws, r_c)
        if len(pcols) < 2:
            continue
        for i in range(1, len(pcols)):
            col_t, col_prev = pcols[i], pcols[i - 1]
            ocell = ws.cell(row=r_o, column=col_t)
            f = ocell.value
            if not _is_formula(f):
                continue  # hardcode -> interior-hardcode check's domain
            refs = _refs_to_row(f, r_c)
            if not refs or col_prev in refs:
                continue  # no coordinate ref to closing row, or correct prior ref
            wrong = ", ".join(get_column_letter(c) for c in sorted(refs))
            report.findings.append(
                ScheduleFinding(
                    sheet=ws.title,
                    cell=ocell.coordinate,
                    row=r_o,
                    kind="rollforward",
                    detail=(
                        f"opening row references closing row {r_c} at the wrong "
                        f"period (expected prior col {get_column_letter(col_prev)}, "
                        f"found {wrong}) — roll-forward period drift"
                    ),
                )
            )


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────


def audit_schedule(xlsx_path: Path | str) -> ScheduleAuditReport:
    """Audit a built workbook for two certify-blind schedule defects.

    1. **Interior hardcode** — a non-innocuous bare number wedged between
       formula cells inside a contiguous period series (a value typed over a
       formula).
    2. **Roll-forward period drift** — an opening row whose period-t cell
       references its paired closing row at the WRONG prior period column.

    Returns a :class:`ScheduleAuditReport`. ``report.passed`` is True iff
    neither defect is found on any non-input/-reference/-audit sheet. Both
    checks are high-precision (zero findings across all certified examples).
    """
    path = Path(xlsx_path)
    report = ScheduleAuditReport(workbook=str(path))

    wb = load_workbook(path, data_only=False)
    for name in wb.sheetnames:
        if classify_sheet(name) in _EXCLUDED_SHEET_CLASSES:
            report.sheets_skipped.append(name)
            continue
        report.sheets_scanned += 1
        _scan_sheet(wb[name], report)
        _scan_rollforward(wb[name], report)

    # Deterministic ordering for stable reports/tests.
    report.findings.sort(key=lambda f: (f.sheet, f.row, f.cell, f.kind))
    return report
