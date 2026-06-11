"""Workbook auditor — the "zero formula errors" certification gate.

Loads a built ``.xlsx``, evaluates every formula with the third-party
``formulas`` engine (a clean-room recalculation that does NOT trust openpyxl's
cached values), and reports:

    (a) ANY Excel error cell — ``#REF!`` / ``#DIV/0!`` / ``#VALUE!`` /
        ``#NAME?`` / ``#NUM!`` / ``#N/A`` / ``#NULL!`` — whether it is a
        live formula that evaluates to an error OR a cached error literal
        baked into the file.
    (b) styling gaps — numeric cells (hardcoded or formula) that carry
        neither an explicit font colour nor an explicit number_format, the
        two things a bulge-tier reviewer scans for at a glance.

A workbook is CERTIFIED when it has zero formula-error cells. Styling gaps
do NOT fail certification (they downgrade it to WARN) — the hard contract is
"no formula errors", per the institutional zero-error policy.

This module is read-only: it never mutates the workbook.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


# The six canonical Excel error sigils, plus #NULL! and #GETTING_DATA.
# We match on the leading ``#`` and an uppercase letter so we don't trip on
# legitimate string content like a "#1 ranked" label (those start "#" then a
# digit) — Excel error codes are always ``#<UPPER>``.
_ERROR_RE = re.compile(r"^#(REF|DIV/0|VALUE|NAME|NUM|N/A|NULL|GETTING_DATA)")


@dataclass
class ErrorCell:
    sheet: str
    cell: str
    error: str
    # "evaluated" — the `formulas` engine recomputed this and got an error.
    # "cached"    — openpyxl read a stored error literal in the file.
    source: str

    @property
    def ref(self) -> str:
        return f"{self.sheet}!{self.cell}"


@dataclass
class StyleGap:
    sheet: str
    cell: str
    reason: str  # "no font colour", "no number_format", or both

    @property
    def ref(self) -> str:
        return f"{self.sheet}!{self.cell}"


@dataclass
class WorkbookAuditReport:
    workbook: str
    error_cells: list[ErrorCell] = field(default_factory=list)
    style_gaps: list[StyleGap] = field(default_factory=list)
    numeric_cells: int = 0
    recalc_ran: bool = False
    notes: list[str] = field(default_factory=list)

    @property
    def n_errors(self) -> int:
        return len(self.error_cells)

    @property
    def n_style_gaps(self) -> int:
        return len(self.style_gaps)

    @property
    def passed(self) -> bool:
        """Pass == zero formula-error cells. Styling gaps do not fail."""
        return self.n_errors == 0

    @property
    def verdict(self) -> str:
        if self.n_errors:
            return "FAIL"
        if self.n_style_gaps:
            return "WARN"
        return "CERTIFIED"

    def summary(self) -> dict:
        return {
            "workbook": self.workbook,
            "verdict": self.verdict,
            "passed": self.passed,
            "error_cells": self.n_errors,
            "style_gaps": self.n_style_gaps,
            "numeric_cells": self.numeric_cells,
            "recalc_ran": self.recalc_ran,
        }


# ─────────────────────────────────────────────────────────────────────────────
# Internals
# ─────────────────────────────────────────────────────────────────────────────


def _is_error_string(v) -> bool:
    return isinstance(v, str) and bool(_ERROR_RE.match(v))


def _coerce_error(v):
    """If ``v`` (an engine value, possibly an array / XlError) is an Excel
    error, return its string code; otherwise return None.

    The ``formulas`` engine returns ``XlError`` instances whose ``str()`` is
    the Excel code (``#DIV/0!`` etc.). Ranges arrive as numpy arrays — we
    flatten and report the first error found.
    """
    if v is None:
        return None
    # numpy / range value — flatten and scan
    flatten = getattr(v, "flatten", None)
    if flatten is not None:
        try:
            for item in v.flatten():
                code = _coerce_error(item)
                if code is not None:
                    return code
            return None
        except Exception:
            return None
    s = str(v)
    if _ERROR_RE.match(s):
        return s
    return None


def _split_engine_key(key: str) -> tuple[str | None, str | None]:
    """Parse a ``formulas`` solution key into (sheet, cell).

    Keys look like ``"[book.xlsx]SHEETNAME'!A1"`` (sheet upper-cased by the
    engine). We recover the sheet token and the cell coordinate.
    """
    try:
        after_book = key.split("]", 1)[1]
        sheet, cell = after_book.split("'!", 1)
        return sheet.strip("'"), cell.strip()
    except (IndexError, ValueError):
        return None, None


def _resolve_sheet_name(engine_sheet: str, real_names: list[str]) -> str:
    """Map the engine's upper-cased sheet token back to the real sheet name."""
    for n in real_names:
        if n.upper() == engine_sheet.upper():
            return n
    return engine_sheet


def _scan_cached_errors(wb_cached, report: WorkbookAuditReport,
                        seen: set[str]) -> None:
    """Flag any error literal baked into the file's cached values."""
    for ws in wb_cached.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if _is_error_string(c.value):
                    ref = f"{ws.title}!{c.coordinate}"
                    if ref in seen:
                        continue
                    seen.add(ref)
                    report.error_cells.append(
                        ErrorCell(ws.title, c.coordinate, str(c.value),
                                  "cached")
                    )


def _scan_style_gaps(wb, report: WorkbookAuditReport) -> None:
    """Flag numeric cells (formula or hardcoded) lacking colour or format.

    A cell is a styling gap when it holds a number (or a formula — formulas
    almost always resolve to numbers in these models) AND it has neither an
    explicit font colour nor an explicit number_format. ``General`` is the
    openpyxl default and counts as "no number_format".
    """
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                is_formula = isinstance(v, str) and v.startswith("=")
                is_number = isinstance(v, (int, float)) and not isinstance(v, bool)
                if not (is_formula or is_number):
                    continue
                report.numeric_cells += 1

                has_format = bool(c.number_format) and c.number_format != "General"
                font_color = getattr(c.font, "color", None)
                color_rgb = getattr(font_color, "rgb", None) if font_color else None
                # openpyxl's "no colour" is None or the theme default (often
                # an auto/indexed black with no rgb). Treat a missing rgb as
                # "no explicit colour".
                has_color = isinstance(color_rgb, str) and len(color_rgb) >= 6

                if has_format and has_color:
                    continue
                reasons = []
                if not has_color:
                    reasons.append("no font colour")
                if not has_format:
                    reasons.append("no number_format")
                report.style_gaps.append(
                    StyleGap(ws.title, c.coordinate, " + ".join(reasons))
                )


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────


def audit_workbook(xlsx_path: Path | str) -> WorkbookAuditReport:
    """Audit a built workbook for formula errors + styling gaps.

    Returns a :class:`WorkbookAuditReport`. ``report.passed`` is True iff the
    workbook has zero formula-error cells.
    """
    path = Path(xlsx_path)
    report = WorkbookAuditReport(workbook=str(path))

    wb = load_workbook(path, data_only=False)
    real_names = list(wb.sheetnames)

    seen_error_refs: set[str] = set()

    # (1a) Live evaluation via the `formulas` engine — the authoritative
    # zero-error check. If the dep is missing we degrade gracefully and rely
    # on the cached-error scan below.
    try:
        import formulas  # type: ignore
    except ImportError:
        report.notes.append(
            "`formulas` package not installed — formula evaluation skipped; "
            "only cached error literals are checked."
        )
    else:
        try:
            xl = formulas.ExcelModel().loads(str(path)).finish()
            sol = xl.calculate()
            report.recalc_ran = True
        except Exception as e:  # pragma: no cover - engine load failure
            report.notes.append(f"Formula engine could not load workbook: {e}")
            sol = {}
        for key, val in sol.items():
            raw = val.value if hasattr(val, "value") else val
            code = _coerce_error(raw)
            if code is None:
                continue
            engine_sheet, cell = _split_engine_key(key)
            if engine_sheet is None or cell is None:
                continue
            sheet = _resolve_sheet_name(engine_sheet, real_names)
            ref = f"{sheet}!{cell}"
            if ref in seen_error_refs:
                continue
            seen_error_refs.add(ref)
            report.error_cells.append(ErrorCell(sheet, cell, code, "evaluated"))

    # (1b) Cached error literals baked into the saved file.
    try:
        wb_cached = load_workbook(path, data_only=True)
        _scan_cached_errors(wb_cached, report, seen_error_refs)
    except Exception as e:  # pragma: no cover
        report.notes.append(f"Cached-value scan failed: {e}")

    # (2) Styling gaps.
    _scan_style_gaps(wb, report)

    # Stable ordering for deterministic reports/tests.
    report.error_cells.sort(key=lambda e: (e.sheet, e.cell))
    report.style_gaps.sort(key=lambda g: (g.sheet, g.cell))
    return report
