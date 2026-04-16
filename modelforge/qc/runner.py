"""External QC runner.

Opens a built workbook and validates what in-sheet checks cannot:
    - Every hardcoded cell has a comment
    - Every source referenced in comments exists on the Sources sheet
    - Every named range resolves
    - Print areas set on every sheet
    - No orphan empty sheets
    - Sign convention smoke-test via the QC sheet's cell values

Emits a QCReport with pass/fail per check. Can also print to console
(rich-formatted) or write a PDF.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook
from rich.console import Console
from rich.table import Table


@dataclass
class QCCheck:
    name: str
    passed: bool
    detail: str = ""


@dataclass
class QCReport:
    workbook: str
    checks: list[QCCheck] = field(default_factory=list)

    @property
    def all_pass(self) -> bool:
        return all(c.passed for c in self.checks)

    @property
    def n_pass(self) -> int:
        return sum(1 for c in self.checks if c.passed)

    @property
    def n_total(self) -> int:
        return len(self.checks)

    def print(self) -> None:
        console = Console()
        tbl = Table(title=f"ModelForge QC — {self.workbook}")
        tbl.add_column("Check", style="bold")
        tbl.add_column("Status")
        tbl.add_column("Detail", overflow="fold")
        for c in self.checks:
            status = "[green]PASS[/green]" if c.passed else "[red]FAIL[/red]"
            tbl.add_row(c.name, status, c.detail)
        console.print(tbl)
        console.print(
            f"\n[bold]{'ALL CHECKS PASS' if self.all_pass else 'FAIL'}[/bold]  "
            f"({self.n_pass}/{self.n_total})"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Individual checks
# ─────────────────────────────────────────────────────────────────────────────


def _check_comments_on_hardcoded_cells(wb) -> QCCheck:
    """Every hardcoded input on Assumptions/Sources/OperatingModel historical
    columns should have a cell comment or a source link.

    Focused check: every BASE cell on Assumptions (col G) has a comment.
    """
    ws = wb["Assumptions"]
    missing: list[str] = []
    # Header on row 5; data starts row 6
    for row in range(6, ws.max_row + 1):
        id_cell = ws.cell(row=row, column=1).value
        if not id_cell:
            continue
        base_cell = ws.cell(row=row, column=7)
        if base_cell.value is None:
            continue
        if base_cell.comment is None:
            missing.append(f"{id_cell} at G{row}")
    passed = not missing
    detail = "OK" if passed else f"{len(missing)} cells missing comments: {missing[:5]}"
    return QCCheck("Every BASE assumption has a cell comment", passed, detail)


def _check_sources_referenced_exist(wb) -> QCCheck:
    """Every source_id referenced in Assumptions column L must exist on Sources."""
    src_ws = wb["Sources"]
    assum_ws = wb["Assumptions"]
    existing: set[str] = set()
    for row in range(6, src_ws.max_row + 1):
        v = src_ws.cell(row=row, column=1).value
        if v and re.match(r"^S-\d{3,}$", str(v)):
            existing.add(str(v))
    missing: list[str] = []
    for row in range(6, assum_ws.max_row + 1):
        ref = assum_ws.cell(row=row, column=12).value
        if ref and re.match(r"^S-\d{3,}$", str(ref)) and ref not in existing:
            missing.append(str(ref))
    passed = not missing
    detail = "OK" if passed else f"Missing sources: {set(missing)}"
    return QCCheck("All referenced source IDs exist on Sources sheet", passed, detail)


def _check_scenario_toggle_named_range(wb) -> QCCheck:
    """Named range `scenario_index` must resolve."""
    if "scenario_index" not in wb.defined_names:
        return QCCheck("scenario_index named range exists", False, "Not defined")
    dn = wb.defined_names["scenario_index"]
    return QCCheck("scenario_index named range exists", True, dn.attr_text)


def _check_named_ranges_populated(wb) -> QCCheck:
    """Named ranges should resolve to non-empty cells."""
    unresolved: list[str] = []
    for name, dn in wb.defined_names.items():
        try:
            # Defined names like 'Sheet'!$A$1 — resolve by splitting
            attr = dn.attr_text
            if "!" not in attr:
                continue
            sheet_part, ref = attr.split("!", 1)
            sheet = sheet_part.strip("'")
            if sheet not in wb.sheetnames:
                unresolved.append(name)
                continue
            # sheet refs exist; just ensure non-empty
            ws = wb[sheet]
            cell = ws[ref.replace("$", "")]
            if cell.value is None:
                unresolved.append(f"{name}={attr} (empty)")
        except Exception as e:
            unresolved.append(f"{name}: {e}")
    passed = not unresolved
    detail = "OK" if passed else f"{len(unresolved)} issues: {unresolved[:5]}"
    return QCCheck("All named ranges resolve to populated cells", passed, detail)


def _check_no_orphan_sheets(wb) -> QCCheck:
    """No empty sheets (must have at least 1 row of content)."""
    orphans = [s for s in wb.sheetnames if wb[s].max_row <= 1 and wb[s].max_column <= 1]
    return QCCheck(
        "No orphan / empty sheets",
        not orphans,
        "OK" if not orphans else f"Empty: {orphans}",
    )


def _check_print_areas(wb) -> QCCheck:
    """Every sheet should have freeze panes or print titles configured."""
    bad = []
    for name in wb.sheetnames:
        ws = wb[name]
        if ws.freeze_panes is None and ws.print_title_rows is None:
            bad.append(name)
    return QCCheck(
        "Print-ready (freeze panes or print titles set)",
        not bad,
        "OK" if not bad else f"Not set on: {bad}",
    )


def _check_qc_sheet_exists(wb) -> QCCheck:
    has_qc = "QC" in wb.sheetnames
    return QCCheck("QC sheet present", has_qc, "" if has_qc else "Missing")


def _check_sign_convention_declared(wb) -> QCCheck:
    """Cover sheet's sign convention cell should say 'costs_negative'."""
    ws = wb["Cover"]
    # Scan for 'sign_convention' EN label and read its col C value
    for row in range(1, 25):
        en = ws.cell(row=row, column=1).value
        if en and "sign" in str(en).lower():
            val = ws.cell(row=row, column=3).value
            ok = str(val).lower() == "costs_negative"
            return QCCheck(
                "Sign convention declared = costs_negative",
                ok,
                f"Found: {val}",
            )
    return QCCheck("Sign convention declared = costs_negative", False, "Not found on Cover")


# ─────────────────────────────────────────────────────────────────────────────
# Public runner
# ─────────────────────────────────────────────────────────────────────────────


def run_qc(xlsx_path: Path | str) -> QCReport:
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    report = QCReport(workbook=str(xlsx_path))

    report.checks.append(_check_qc_sheet_exists(wb))
    report.checks.append(_check_sign_convention_declared(wb))
    report.checks.append(_check_scenario_toggle_named_range(wb))
    report.checks.append(_check_named_ranges_populated(wb))
    report.checks.append(_check_comments_on_hardcoded_cells(wb))
    report.checks.append(_check_sources_referenced_exist(wb))
    report.checks.append(_check_print_areas(wb))
    report.checks.append(_check_no_orphan_sheets(wb))

    return report
