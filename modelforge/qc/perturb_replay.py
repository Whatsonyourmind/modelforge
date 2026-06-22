"""Perturbation-replay conservation gate — invariants must HOLD under input shocks.

The static conservation gate (:mod:`modelforge.qc.conservation_audit`) recomputes
the QC sheet at the *baseline* inputs only. A latent conservation defect can hide
there: a "Total" hardcoded to its baseline value passes ``Total == Σ parts`` at
the built numbers, then silently breaks the identity for any other inputs. The
static gate is blind to it (the cell is a valid number, the check passes at
baseline).

This gate closes that gap deterministically. It compiles the workbook once with
the ``formulas`` engine, then for each resolved INPUT driver (a numeric leaf
constant whose perturbation re-propagates into the QC sheet) it shocks the driver
by a small relative delta, recomputes the whole workbook, and re-reads the QC
checks. A conservation **identity** holds for ALL valid inputs in a correctly
wired model, so any per-check that PASSES at baseline but FLIPS to fail/error
under the shock is a genuine input-dependent conservation defect — fail-closed.

Honest scope (the INVARIANTS-HOLD half of the perturb-replay design):
  * It gates exactly the checks the model's own QC sheet declares.
  * Small shocks (default ±5%) keep the model in a valid region, so robust
    validity thresholds do not trip (verified zero-flip on certified examples);
    it is NOT a stress test of threshold checks.
  * Drivers with no recomputable cone to the QC sheet are reported, not failed;
    no QC-reaching driver ⇒ INCONCLUSIVE (cannot exercise the invariants).
  * Missing recalc engine / no QC sheet / a baseline that is not already clean
    ⇒ INDETERMINATE, which the gate treats as a failure.

Zero-LLM, free/local, deterministic. Complementary to the static schedule
(``audit_schedule``) and conservation (``audit_conservation``) gates.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook

from modelforge.qc.conservation_audit import _classify, _find_qc_sheet
from modelforge.qc.workbook_audit import _scalarize

# Bounds (opt-in gate; keep runtime predictable).
_REL_DELTA = 0.05
_MAX_PROBE = 16          # candidate drivers to probe (each = TWO recomputes, ±δ)
_MIN_MOVE = 1e-9         # downstream-movement threshold (relative)


@dataclass
class PerturbFinding:
    """A QC check that passed at baseline but broke under an input shock."""

    driver: str          # the perturbed input cell (Sheet!Cell)
    check_ref: str       # the QC check cell that flipped
    check_label: str
    status: str          # "fail" (recomputed 0) or "error"
    recomputed: object


@dataclass
class PerturbReplayReport:
    workbook: str
    recalc_ran: bool = False
    qc_sheet: str | None = None
    drivers_probed: int = 0
    drivers_active: int = 0       # perturbation moved ≥1 non-QC model cell
    rel_delta: float = _REL_DELTA
    findings: list[PerturbFinding] = field(default_factory=list)
    status: str = "INDETERMINATE"  # PASS / FAIL / INCONCLUSIVE / INDETERMINATE
    notes: list[str] = field(default_factory=list)

    @property
    def passed(self) -> bool:
        return self.status == "PASS"

    @property
    def n_findings(self) -> int:
        return len(self.findings)

    @property
    def verdict(self) -> str:
        return self.status

    def summary(self) -> dict:
        return {
            "status": self.status,
            "qc_sheet": self.qc_sheet,
            "drivers_probed": self.drivers_probed,
            "drivers_active": self.drivers_active,
            "rel_delta": self.rel_delta,
            "n_findings": self.n_findings,
            "recalc_ran": self.recalc_ran,
        }

    def print(self) -> None:  # pragma: no cover - console formatting
        from rich.console import Console
        from rich.table import Table

        console = Console()
        badge = {
            "PASS": "[bold green]PERTURB PASS[/bold green]",
            "FAIL": "[bold red]PERTURB FAIL[/bold red]",
            "INCONCLUSIVE": "[bold yellow]PERTURB INCONCLUSIVE[/bold yellow]",
            "INDETERMINATE": "[bold yellow]PERTURB INDETERMINATE[/bold yellow]",
        }.get(self.status, self.status)
        console.print(
            f"{badge}  {Path(self.workbook).name}  "
            f"active drivers={self.drivers_active}/{self.drivers_probed}  "
            f"shock=±{self.rel_delta:.0%}  breaks={self.n_findings}"
        )
        if self.findings:
            tbl = Table(title="Input-dependent conservation breaks (certify-blind)")
            tbl.add_column("Driver", style="bold")
            tbl.add_column("Broken check")
            tbl.add_column("Recomputed")
            for f in self.findings[:20]:
                tbl.add_row(f.driver, f.check_label, f"{f.recomputed} ({f.status})")
            console.print(tbl)
        for n in self.notes:
            console.print(f"[dim]note: {n}[/dim]")


# ── helpers ───────────────────────────────────────────────────────────────────

def _engine_key(stem: str, sheet: str, coord: str) -> str:
    return f"'[{stem}]{sheet.upper()}'!{coord}"


def _val(sol, key):
    v = sol.get(key)
    if v is None:
        return None
    raw = v.value if hasattr(v, "value") else v
    return _scalarize(raw)


def _fnum(x):
    try:
        return float(x)
    except (TypeError, ValueError):
        return None


def _resolve_drivers(wb, stem: str) -> list[tuple[str, float, str]]:
    """Candidate input drivers: numeric leaf CONSTANTS (not formulas).

    Returns (engine_key, value, label) ordered by likely QC impact: registered
    defined-name inputs first, then constants on assumption/input sheets. Only
    constants are perturbable — overriding a formula cell is futile (the engine
    recomputes it from its formula).
    """
    seen: set[str] = set()
    named: list[tuple[str, float, str]] = []
    others: list[tuple[str, float, str]] = []

    # 1. Defined-name targets that point at a single numeric constant.
    try:
        names = dict(wb.defined_names)
    except Exception:  # pragma: no cover
        names = {}
    for nm, dn in names.items():
        m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)", getattr(dn, "attr_text", "") or "")
        if not m:
            continue
        sheet, coord = m.group(1), f"{m.group(2)}{m.group(3)}"
        try:
            cell = wb[sheet][coord]
        except Exception:
            continue
        v = cell.value
        if isinstance(v, (int, float)) and not isinstance(v, bool) and abs(v) > _MIN_MOVE:
            key = _engine_key(stem, sheet, coord)
            if key not in seen:
                seen.add(key)
                named.append((key, float(v), f"{sheet}!{coord} ({nm})"))

    # 2. Numeric constants on assumption/input sheets.
    from modelforge.moat.classifier import classify_sheet
    for ws in wb.worksheets:
        cls = classify_sheet(ws.title)
        if not (cls == "input" or ws.title.lower() in ("assumptions", "inputs", "drivers")):
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, (int, float)) and not isinstance(v, bool) and abs(v) > _MIN_MOVE:
                    key = _engine_key(stem, ws.title, c.coordinate)
                    if key not in seen:
                        seen.add(key)
                        others.append((key, float(v), f"{ws.title}!{c.coordinate}"))

    return named + others


def _qc_check_keys(wb, stem: str):
    """Return (qc_sheet, all_pass_key, [(check_key, label), ...]) or (None, ...)."""
    qc_sheet, arow, acol = _find_qc_sheet(wb)
    if qc_sheet is None:
        return None, None, []
    ws = wb[qc_sheet]
    rcol = acol + 2
    ap_key = _engine_key(stem, qc_sheet, ws.cell(row=arow, column=rcol).coordinate)
    checks: list[tuple[str, str]] = []
    for r in range(arow + 1, ws.max_row + 1):
        lab = ws.cell(row=r, column=acol).value
        fml = ws.cell(row=r, column=rcol).value
        if isinstance(lab, str) and lab.strip() and isinstance(fml, str) and fml.startswith("="):
            key = _engine_key(stem, qc_sheet, ws.cell(row=r, column=rcol).coordinate)
            checks.append((key, lab.strip()))
    return qc_sheet, ap_key, checks


# ── public entry point ─────────────────────────────────────────────────────────

def audit_perturb_replay(
    xlsx_path: Path | str,
    rel_delta: float = _REL_DELTA,
    max_probe: int = _MAX_PROBE,
) -> PerturbReplayReport:
    """Shock each input driver and require the QC conservation checks to HOLD."""
    path = Path(xlsx_path)
    report = PerturbReplayReport(workbook=str(path), rel_delta=rel_delta)

    try:
        import formulas  # type: ignore
    except ImportError:
        report.notes.append("`formulas` package not installed — perturb-replay INDETERMINATE.")
        return report

    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:  # pragma: no cover
        report.notes.append(f"Could not open workbook: {e}")
        return report

    stem = path.name
    qc_sheet, ap_key, checks = _qc_check_keys(wb, stem)
    if qc_sheet is None:
        report.notes.append("No QC sheet (no 'ALL CHECKS PASS' aggregator) — INDETERMINATE.")
        return report
    report.qc_sheet = qc_sheet

    try:
        xl = formulas.ExcelModel().loads(str(path)).finish()
        base = xl.calculate()
        report.recalc_ran = True
    except Exception as e:  # pragma: no cover
        report.notes.append(f"Formula engine could not compile/calculate: {e}")
        return report

    # Baseline must already be clean (else this is conservation_audit's job).
    base_ap = _classify(_val(base, ap_key))[0]
    if base_ap != "pass":
        report.notes.append(
            f"Baseline ALL_PASS does not recompute to a clean pass ({base_ap}) — "
            f"run `certify --conservation` first. INDETERMINATE."
        )
        return report

    # Re-evaluate every per-check AND the aggregator under each shock.
    eval_checks = list(checks) + [(ap_key, "ALL CHECKS PASS (aggregator)")]
    base_status = {k: _classify(_val(base, k))[0] for k, _ in eval_checks}

    # Baseline numeric snapshot of every cell, for activity detection. A driver
    # is "active" when its shock moves ≥1 NON-QC model cell (the QC result cells
    # are binary and stay 1 for a holding invariant, so they cannot signal flow).
    qc_up = qc_sheet.upper()

    def _is_qc(k: str) -> bool:
        try:
            return k.split("]", 1)[1].split("'!", 1)[0].upper() == qc_up
        except Exception:  # pragma: no cover
            return False

    base_all = {k: _fnum(_val(base, k)) for k in base.keys()}

    drivers = _resolve_drivers(wb, stem)
    drivers = [d for d in drivers if base.get(d[0]) is not None][:max_probe]

    def _moved_any_nonqc(pert) -> bool:
        for k, bv in base_all.items():
            if bv is None or _is_qc(k):
                continue
            pv = _fnum(_val(pert, k))
            if pv is None:
                continue
            if abs(pv - bv) > _MIN_MOVE * max(1.0, abs(bv)):
                return True
        return False

    for key, val0, label in drivers:
        report.drivers_probed += 1
        up = val0 * (1.0 + rel_delta) or (val0 + rel_delta)
        dn = val0 * (1.0 - rel_delta) or (val0 - rel_delta)
        try:
            pert_up = xl.calculate(inputs={key: up})
            pert_dn = xl.calculate(inputs={key: dn})
        except Exception:
            continue  # engine declined this input — inconclusive for this driver

        if not (_moved_any_nonqc(pert_up) or _moved_any_nonqc(pert_dn)):
            continue  # inactive driver (e.g. an inactive scenario column)
        report.drivers_active += 1

        # A conservation IDENTITY (A == B+C) breaks under BOTH ±δ shocks; a
        # one-sided validity threshold ("debt fully amortized by term") breaks in
        # only one direction. Flag only the two-sided breaks — those are genuine
        # input-dependent conservation defects, not legitimate threshold limits.
        for k, clabel in eval_checks:
            if base_status.get(k) != "pass":
                continue
            up_status = _classify(_val(pert_up, k))[0]
            dn_status = _classify(_val(pert_dn, k))[0]
            if up_status in ("fail", "error") and dn_status in ("fail", "error"):
                report.findings.append(
                    PerturbFinding(
                        driver=label,
                        check_ref=k.split("]", 1)[-1],
                        check_label=clabel,
                        status=up_status,
                        recomputed=_val(pert_up, k),
                    )
                )

    # Verdict.
    if report.findings:
        report.status = "FAIL"
    elif report.drivers_active == 0:
        report.status = "INCONCLUSIVE"
        report.notes.append(
            "No resolved driver materially moved the model — could not exercise "
            "the conservation invariants under perturbation."
        )
    else:
        report.status = "PASS"
    return report
