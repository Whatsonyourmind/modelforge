"""FastAPI app factory."""

from __future__ import annotations

import hashlib
import os
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from openpyxl import load_workbook
from pydantic import BaseModel, Field


# ── In-memory + disk-backed workbook store ──────────────────────────────────


@dataclass
class StoredWorkbook:
    workbook_id: str     # content SHA-256 first 16 chars
    original_name: str
    path: Path
    sheet_names: list[str]
    primary_output_ref: Optional[str]
    reproducibility: dict = field(default_factory=dict)


class WorkbookStore:
    """File-system-backed registry keyed by content hash."""

    def __init__(self, session_dir: Optional[Path] = None) -> None:
        self.session_dir = session_dir or Path(tempfile.mkdtemp(
            prefix="modelforge_web_"))
        self.session_dir.mkdir(parents=True, exist_ok=True)
        self.workbooks: dict[str, StoredWorkbook] = {}
        self._rescan()

    def _rescan(self) -> None:
        """Re-populate from disk (survives server restart)."""
        for f in self.session_dir.glob("*.xlsx"):
            # workbook_id encoded in the stem
            wid = f.stem.split("_", 1)[0]
            if wid in self.workbooks:
                continue
            self._register_existing(f, wid)

    def _register_existing(self, path: Path, wid: str) -> StoredWorkbook:
        wb = load_workbook(path, data_only=False, keep_links=True,
                           read_only=False)
        primary = None
        if "primary_output" in wb.defined_names:
            primary = wb.defined_names["primary_output"].attr_text
        repro = {}
        if "Reproducibility" in wb.sheetnames:
            ws = wb["Reproducibility"]
            for r in range(5, ws.max_row + 1):
                k = ws.cell(row=r, column=1).value
                v = ws.cell(row=r, column=2).value
                if k and v is not None:
                    repro[str(k)] = str(v)
        stored = StoredWorkbook(
            workbook_id=wid,
            original_name=path.name.split("_", 1)[1] if "_" in path.name else path.name,
            path=path,
            sheet_names=list(wb.sheetnames),
            primary_output_ref=primary,
            reproducibility=repro,
        )
        self.workbooks[wid] = stored
        return stored

    def add(self, content: bytes, original_name: str) -> StoredWorkbook:
        wid = hashlib.sha256(content).hexdigest()[:16]
        if wid in self.workbooks:
            return self.workbooks[wid]
        safe_name = "".join(c if c.isalnum() or c in "-._" else "_"
                             for c in original_name)
        path = self.session_dir / f"{wid}_{safe_name}"
        path.write_bytes(content)
        return self._register_existing(path, wid)

    def get(self, wid: str) -> StoredWorkbook:
        if wid not in self.workbooks:
            raise KeyError(wid)
        return self.workbooks[wid]

    def list_all(self) -> list[StoredWorkbook]:
        return list(self.workbooks.values())


# ── Request models ──────────────────────────────────────────────────────────


class RiskRequest(BaseModel):
    equity_value_eur_m: float = Field(gt=0)
    equity_volatility: float = Field(gt=0)
    debt_face_value_eur_m: float = Field(gt=0)
    risk_free_rate: float = 0.039
    horizon_years: float = 1.0
    lgd: float = 0.45
    eir: float = 0.05
    maturity_years: int = 5
    days_past_due: int = 0
    counterparty: Optional[str] = None


# ── App factory ─────────────────────────────────────────────────────────────


def create_app(session_dir: Optional[Path] = None) -> FastAPI:
    app = FastAPI(
        title="ModelForge Web",
        description="Bulge-tier Excel model factory — HTTP surface.",
        version="0.5.8",
    )
    store = WorkbookStore(session_dir=session_dir)

    # D5: SaaS API surface — tenant management, auth, billing, tenant-scoped
    # workbooks. Mounted under /api/v1/*. Existing demo routes (/, /upload,
    # /workbook/{id}/...) stay intact — production deploys can disable them
    # by setting MODELFORGE_DISABLE_DEMO=1 (handled below).
    from modelforge.web.saas_routes import build_saas_router
    app.include_router(build_saas_router())

    @app.get("/healthz", include_in_schema=False)
    def root_healthz() -> dict:
        return {"status": "ok", "service": "modelforge-web"}

    if os.environ.get("MODELFORGE_DISABLE_DEMO") == "1":
        # Production: don't expose the unauthenticated demo upload UI.
        # SaaS callers use /api/v1/* exclusively.
        @app.get("/", response_class=HTMLResponse)
        def _root_prod() -> str:
            return (
                "<h1>ModelForge SaaS</h1><p>API at <code>/api/v1/*</code>. "
                "See <code>/docs</code> for OpenAPI.</p>"
            )
        return app

    # ── Index
    @app.get("/", response_class=HTMLResponse)
    def index() -> str:
        rows = "\n".join(
            f"<tr><td><code>{w.workbook_id}</code></td>"
            f"<td>{w.original_name}</td>"
            f"<td>{len(w.sheet_names)}</td>"
            f"<td><code>{w.primary_output_ref or '—'}</code></td>"
            f"<td><a href='/workbook/{w.workbook_id}/view'>view</a> · "
            f"<a href='/workbook/{w.workbook_id}/dossier'>dossier</a> · "
            f"<a href='/workbook/{w.workbook_id}/drift'>drift</a></td></tr>"
            for w in sorted(store.list_all(), key=lambda x: x.original_name)
        )
        return _INDEX_HTML.replace("__ROWS__", rows or
                                    "<tr><td colspan='5'><i>(no uploads yet)</i></td></tr>")

    # ── Upload
    @app.post("/upload")
    async def upload(file: UploadFile = File(...)) -> dict:
        content = await file.read()
        if not content:
            raise HTTPException(400, "empty file")
        if not (file.filename or "").lower().endswith(".xlsx"):
            raise HTTPException(400, "only .xlsx accepted")
        try:
            stored = store.add(content, file.filename or "upload.xlsx")
        except Exception as e:
            raise HTTPException(400, f"failed to parse workbook: {e}")
        return {
            "workbook_id": stored.workbook_id,
            "original_name": stored.original_name,
            "sheet_names": stored.sheet_names,
            "primary_output_ref": stored.primary_output_ref,
            "reproducibility": stored.reproducibility,
        }

    # ── JSON metadata
    @app.get("/workbook/{wid}")
    def get_workbook(wid: str) -> dict:
        try:
            stored = store.get(wid)
        except KeyError:
            raise HTTPException(404, "workbook not found")
        return {
            "workbook_id": stored.workbook_id,
            "original_name": stored.original_name,
            "sheet_names": stored.sheet_names,
            "primary_output_ref": stored.primary_output_ref,
            "reproducibility": stored.reproducibility,
        }

    # ── HTML view
    @app.get("/workbook/{wid}/view", response_class=HTMLResponse)
    def view_workbook(wid: str) -> str:
        try:
            stored = store.get(wid)
        except KeyError:
            raise HTTPException(404, "workbook not found")
        sheets_html = "\n".join(f"<li>{s}</li>" for s in stored.sheet_names)
        repro_rows = "\n".join(
            f"<tr><td>{k}</td><td><code>{v}</code></td></tr>"
            for k, v in stored.reproducibility.items()
        )
        return _VIEW_HTML.replace("__ID__", stored.workbook_id)\
                        .replace("__NAME__", stored.original_name)\
                        .replace("__PRIMARY__", stored.primary_output_ref or "—")\
                        .replace("__SHEETS__", sheets_html)\
                        .replace("__REPRO__", repro_rows or
                                 "<tr><td colspan='2'>(no repro block)</td></tr>")

    # ── Dossier
    @app.get("/workbook/{wid}/dossier")
    def get_dossier(wid: str) -> FileResponse:
        from modelforge.dossier import generate_dossier
        try:
            stored = store.get(wid)
        except KeyError:
            raise HTTPException(404, "workbook not found")
        pdf = stored.path.with_suffix(".dossier.pdf")
        if not pdf.exists():
            generate_dossier(stored.path, pdf)
        return FileResponse(pdf, media_type="application/pdf",
                            filename=f"{stored.original_name}.dossier.pdf")

    # ── Drift
    @app.get("/workbook/{wid}/drift")
    def check_drift_endpoint(wid: str) -> dict:
        from modelforge.drift import check_drift
        try:
            stored = store.get(wid)
        except KeyError:
            raise HTTPException(404, "workbook not found")
        rep = check_drift(stored.path)
        return {
            "workbook_id": wid,
            "checked": rep.checked_drivers,
            "flagged": rep.n_flagged,
            "clean": rep.clean,
            "items": [
                {
                    "driver": i.driver_name, "assumed": i.assumed_value,
                    "current": i.current_value, "delta_bps": i.delta_bps,
                    "delta_rel": i.delta_rel, "flagged": i.flagged,
                    "source": i.source, "kind": i.kind,
                }
                for i in rep.items
            ],
            "missing": rep.missing_drivers,
        }

    # ── Diff
    @app.get("/diff", response_class=HTMLResponse)
    def diff_endpoint(a: str = Query(...), b: str = Query(...)) -> str:
        from modelforge.diff import compute_diff, render_html
        try:
            wa, wb_ = store.get(a), store.get(b)
        except KeyError as e:
            raise HTTPException(404, f"workbook not found: {e}")
        res = compute_diff(wa.path, wb_.path)
        return render_html(res)

    # ── Risk
    @app.post("/risk")
    def risk_endpoint(req: RiskRequest) -> dict:
        from modelforge.risk import (
            ECLInputs, MertonInputs, calibrate_pd_kmv, compute_ecl,
            solve_merton,
        )
        merton = solve_merton(MertonInputs(
            equity_value=req.equity_value_eur_m,
            equity_volatility=req.equity_volatility,
            debt_face_value=req.debt_face_value_eur_m,
            risk_free_rate=req.risk_free_rate,
            horizon_years=req.horizon_years,
        ))
        kmv_pd = calibrate_pd_kmv(merton.distance_to_default)
        pd_ = max(merton.probability_of_default, kmv_pd)
        ecl_inp = ECLInputs(
            exposure_at_default_eur_m=req.debt_face_value_eur_m,
            loss_given_default=req.lgd,
            effective_interest_rate=req.eir,
            maturity_years=req.maturity_years,
            pd_curve_annual=[pd_] * req.maturity_years,
            current_pd_12m=pd_, origination_pd_12m=pd_,
            days_past_due=req.days_past_due,
        )
        ecl = compute_ecl(ecl_inp, req.counterparty or "counterparty")
        return {
            "counterparty": req.counterparty,
            "merton": {
                "asset_value": merton.asset_value,
                "asset_volatility": merton.asset_volatility,
                "distance_to_default": merton.distance_to_default,
                "pd": merton.probability_of_default,
                "converged": merton.converged,
            },
            "kmv_pd": kmv_pd,
            "ecl": {
                "stage": ecl.stage.value,
                "ecl_12_month_eur_m": ecl.ecl_12_month_eur_m,
                "ecl_lifetime_eur_m": ecl.ecl_lifetime_eur_m,
                "ecl_eur_m": ecl.ecl_eur_m,
                "implied_rate": ecl.implied_rate_pct,
                "notes": ecl.notes,
            },
        }

    # ── Health
    @app.get("/health")
    def health() -> dict:
        return {"status": "ok", "workbooks": len(store.workbooks)}

    # Expose for tests / CLI
    app.state.store = store
    return app


# ── HTML templates (inline) ─────────────────────────────────────────────────

_INDEX_HTML = """<!doctype html>
<html lang="en"><head><meta charset='utf-8'>
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>ModelForge Web — Bulge-Tier Financial Model Factory</title>
<meta name="description" content="Upload, view, diff, export bulge-tier ModelForge workbooks via HTTP.">
<script src="https://cdn.tailwindcss.com"></script>
<style>
  body { font-family: 'Inter', system-ui, -apple-system, sans-serif; }
  .gradient-bg { background: linear-gradient(135deg, #102C57 0%, #1A3F7A 50%, #2C5BA0 100%); }
  pre, code, .mono { font-family: 'JetBrains Mono', 'Consolas', monospace; }
</style></head>
<body class="bg-slate-50 text-slate-900">

<header class="gradient-bg text-white">
  <div class="max-w-6xl mx-auto px-6 py-5 flex items-center justify-between">
    <div>
      <div class="text-xl font-bold tracking-tight">ModelForge Web</div>
      <p class="text-blue-200 text-xs mt-0.5">v0.9.3 · MCP-native · 14 templates · 431 tests</p>
    </div>
    <nav class="space-x-5 text-sm">
      <a href="/health" class="hover:text-blue-200">Health</a>
      <a href="https://github.com/Whatsonyourmind/modelforge" class="hover:text-blue-200">GitHub</a>
      <a href="https://github.com/Whatsonyourmind/modelforge/blob/master/SCORECARD_v3.md" class="hover:text-blue-200">Scorecard</a>
      <a href="https://github.com/Whatsonyourmind/modelforge/blob/master/GTM_STRATEGY.md" class="hover:text-blue-200">GTM</a>
    </nav>
  </div>
</header>

<main class="max-w-6xl mx-auto px-6 py-8">
  <section class="mb-8">
    <h1 class="text-3xl font-bold mb-2">Upload a workbook</h1>
    <p class="text-slate-600 mb-4 leading-relaxed">
      Upload a <code class="mono bg-slate-200 px-1.5 py-0.5 rounded">.xlsx</code> generated by
      <code class="mono bg-slate-200 px-1.5 py-0.5 rounded">modelforge build</code> to view metadata,
      export an audit dossier, check drift, or diff against another version.
      Source traceability is preserved through every operation.
    </p>
    <form action='/upload' method='post' enctype='multipart/form-data' class="bg-white border border-slate-200 rounded-lg p-6 flex flex-wrap items-center gap-4">
      <input type='file' name='file' accept='.xlsx' required
             class="block text-sm text-slate-600 file:mr-3 file:py-2 file:px-4 file:rounded file:border-0 file:bg-blue-600 file:text-white hover:file:bg-blue-700 file:cursor-pointer cursor-pointer">
      <button type='submit' class="bg-slate-900 text-white px-5 py-2 rounded font-semibold hover:bg-slate-800 transition">Upload workbook</button>
    </form>
  </section>

  <section>
    <h2 class="text-2xl font-bold mb-4">Workbooks in this session</h2>
    <div class="bg-white border border-slate-200 rounded-lg overflow-hidden">
      <table class="w-full text-sm">
        <thead>
          <tr class="bg-slate-100 text-left text-slate-700 border-b border-slate-200">
            <th class="px-4 py-3 font-semibold">ID</th>
            <th class="px-4 py-3 font-semibold">Name</th>
            <th class="px-4 py-3 font-semibold">Sheets</th>
            <th class="px-4 py-3 font-semibold">Primary output</th>
            <th class="px-4 py-3 font-semibold">Actions</th>
          </tr>
        </thead>
        <tbody>
__ROWS__
        </tbody>
      </table>
    </div>
  </section>

  <section class="mt-10 grid md:grid-cols-2 gap-6">
    <div class="bg-white border border-slate-200 rounded-lg p-5">
      <h3 class="font-bold text-lg mb-2">API endpoints</h3>
      <ul class="text-sm text-slate-600 space-y-1 mono">
        <li><span class="text-blue-700">POST</span> /upload</li>
        <li><span class="text-blue-700">GET</span>  /workbook/&lt;id&gt;</li>
        <li><span class="text-blue-700">GET</span>  /workbook/&lt;id&gt;/view</li>
        <li><span class="text-blue-700">GET</span>  /workbook/&lt;id&gt;/dossier</li>
        <li><span class="text-blue-700">GET</span>  /workbook/&lt;id&gt;/drift</li>
        <li><span class="text-blue-700">GET</span>  /diff?a=&lt;id&gt;&amp;b=&lt;id&gt;</li>
        <li><span class="text-blue-700">POST</span> /risk</li>
        <li><span class="text-blue-700">GET</span>  /health</li>
      </ul>
    </div>
    <div class="bg-white border border-slate-200 rounded-lg p-5">
      <h3 class="font-bold text-lg mb-2">Use it inside Claude Code / Cursor / ChatGPT</h3>
      <p class="text-sm text-slate-600 leading-relaxed mb-2">ModelForge is also an MCP server. Wire it into your MCP client:</p>
      <pre class="text-xs bg-slate-100 p-3 rounded overflow-x-auto"><code>{
  "mcpServers": {
    "modelforge": { "command": "modelforge-mcp" }
  }
}</code></pre>
    </div>
  </section>
</main>

<footer class="mt-12 bg-slate-900 text-slate-400 py-6 text-sm">
  <div class="max-w-6xl mx-auto px-6 flex flex-wrap justify-between gap-4">
    <div>© 2026 Luka Stanisljevic · ModelForge v0.9.3 · Milan</div>
    <div class="space-x-4">
      <a href="https://github.com/Whatsonyourmind/modelforge" class="hover:text-white">GitHub</a>
      <a href="mailto:redacted@example.com" class="hover:text-white">Contact</a>
    </div>
  </div>
</footer>
</body></html>
"""


_VIEW_HTML = """<!doctype html>
<html lang="en"><head><meta charset='utf-8'>
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>ModelForge — __NAME__</title>
<script src="https://cdn.tailwindcss.com"></script>
<style>
  body { font-family: 'Inter', system-ui, -apple-system, sans-serif; }
  .gradient-bg { background: linear-gradient(135deg, #102C57 0%, #1A3F7A 50%, #2C5BA0 100%); }
  .mono { font-family: 'JetBrains Mono', 'Consolas', monospace; }
</style></head>
<body class="bg-slate-50 text-slate-900">

<header class="gradient-bg text-white">
  <div class="max-w-6xl mx-auto px-6 py-5 flex items-center justify-between">
    <a href="/" class="text-xl font-bold tracking-tight hover:opacity-90">ModelForge Web</a>
    <nav class="space-x-4 text-sm">
      <a href="/" class="hover:text-blue-200">← All workbooks</a>
    </nav>
  </div>
</header>

<main class="max-w-5xl mx-auto px-6 py-8">
  <div class="mb-8">
    <h1 class="text-3xl font-bold mb-2">__NAME__</h1>
    <p class="text-sm text-slate-500 mono">
      ID: <span class="bg-slate-200 px-2 py-0.5 rounded">__ID__</span>
      · Primary output: <span class="text-green-700 bg-slate-200 px-2 py-0.5 rounded">__PRIMARY__</span>
    </p>
  </div>

  <div class="flex flex-wrap gap-3 mb-8">
    <a href='/workbook/__ID__/dossier' class="bg-slate-900 text-white px-4 py-2 rounded font-semibold hover:bg-slate-800 transition text-sm">Download dossier PDF</a>
    <a href='/workbook/__ID__/drift' class="bg-blue-600 text-white px-4 py-2 rounded font-semibold hover:bg-blue-700 transition text-sm">Check drift (JSON)</a>
    <a href='/workbook/__ID__' class="bg-slate-200 text-slate-900 px-4 py-2 rounded font-semibold hover:bg-slate-300 transition text-sm">Raw JSON metadata</a>
  </div>

  <section class="bg-white border border-slate-200 rounded-lg p-6 mb-6">
    <h2 class="text-xl font-bold mb-3">Sheets</h2>
    <ul class="grid grid-cols-2 md:grid-cols-3 gap-2 text-sm">__SHEETS__</ul>
  </section>

  <section class="bg-white border border-slate-200 rounded-lg overflow-hidden">
    <h2 class="text-xl font-bold p-6 pb-3">Reproducibility</h2>
    <table class="w-full text-sm">
      <thead>
        <tr class="bg-slate-100 text-left text-slate-700 border-b border-slate-200">
          <th class="px-6 py-2 font-semibold">Field</th>
          <th class="px-6 py-2 font-semibold">Value</th>
        </tr>
      </thead>
      <tbody>__REPRO__</tbody>
    </table>
  </section>
</main>

<footer class="mt-12 bg-slate-900 text-slate-400 py-4 text-sm">
  <div class="max-w-6xl mx-auto px-6">
    © 2026 ModelForge · <a href="https://github.com/Whatsonyourmind/modelforge" class="hover:text-white">GitHub</a>
  </div>
</footer>
</body></html>
"""


# Module-level app instance for ASGI servers (uvicorn modelforge.web.app:app)
app = create_app()
