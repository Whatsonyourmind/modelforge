"""Build the chat system prompt from a workbook + linkage graph.

The prompt captures enough structural detail — every assumption, every
source, every formula shape, and the full named-range map — that
Claude can answer questions like "why is Y3 revenue growth 7%?" or
"which cells depend on assumption A-012?" with cited references.

Designed to be CACHED — the prompt is deterministic for a given
workbook, so calls across a chat session hit the Anthropic prompt
cache.
"""

from __future__ import annotations

import re
import sqlite3
from collections import Counter
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook


def workbook_summary(xlsx_path: Path | str, graph_db: Optional[Path | str] = None) -> str:
    """Produce a text summary of a workbook for use as a system prompt.

    Format (human-readable, stable across rebuilds):

        # ModelForge workbook summary
        File: <path>
        Sheets: [...]
        Primary output: 'Sheet'!$X$Y
        Reproducibility: SHA=... version=... timestamp=...

        ## Sources (S-###)
        ...

        ## Assumptions (A-###)
        ...

        ## Formula shapes by sheet
        ...

        ## Linkage graph stats
        ...
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)

    out: list[str] = []
    out.append("# ModelForge workbook summary")
    out.append(f"File: {xlsx_path.name}")
    out.append(f"Sheets: {wb.sheetnames}")
    if "primary_output" in wb.defined_names:
        out.append(f"primary_output: {wb.defined_names['primary_output'].attr_text}")
    if "Reproducibility" in wb.sheetnames:
        r_ws = wb["Reproducibility"]
        meta: list[str] = []
        for r in range(5, r_ws.max_row + 1):
            label = r_ws.cell(row=r, column=1).value
            value = r_ws.cell(row=r, column=2).value
            if label and value is not None:
                meta.append(f"{label}={value}")
        if meta:
            out.append("Reproducibility: " + "; ".join(meta))

    out.append("")

    # Sources
    if "Sources" in wb.sheetnames:
        out.append("## Sources (S-###)")
        sws = wb["Sources"]
        for r in range(6, sws.max_row + 1):
            vid = sws.cell(row=r, column=1).value
            if not vid:
                continue
            doc = sws.cell(row=r, column=2).value or ""
            page = sws.cell(row=r, column=3).value or ""
            publisher = sws.cell(row=r, column=4).value or ""
            date = sws.cell(row=r, column=5).value or ""
            url = sws.cell(row=r, column=6).value or ""
            verified = sws.cell(row=r, column=7).value or ""
            out.append(
                f"{vid}: {doc} (p.{page}) — {publisher}, {date}"
                f"{'  [verified]' if str(verified).strip() else ''}"
                f"{'  ' + str(url) if url else ''}"
            )
        out.append("")

    # Assumptions
    if "Assumptions" in wb.sheetnames:
        out.append("## Assumptions (A-###)")
        out.append("Format: ID | name | unit | worst/base/best | conf | source | rationale")
        aws = wb["Assumptions"]
        for r in range(6, aws.max_row + 1):
            vid = aws.cell(row=r, column=1).value
            if not vid:
                continue
            name = aws.cell(row=r, column=2).value
            unit = aws.cell(row=r, column=5).value
            worst = aws.cell(row=r, column=6).value
            base = aws.cell(row=r, column=7).value
            best = aws.cell(row=r, column=8).value
            conf = aws.cell(row=r, column=11).value
            src = aws.cell(row=r, column=12).value
            rationale = aws.cell(row=r, column=10).value
            out.append(
                f"{vid} | {name} | {unit} | {worst}/{base}/{best} "
                f"| {conf or '-'} | {src or '-'} | {rationale or ''}"
            )
        out.append("")

    # Formula shapes
    out.append("## Formula shapes by sheet (top 10 per sheet)")
    for sheet in wb.sheetnames:
        if sheet in ("Cover", "Sources", "Reproducibility"):
            continue
        ws = wb[sheet]
        patterns: Counter = Counter()
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    p = re.sub(r"[A-Z]+\d+", "CELL", v)
                    p = re.sub(r"\d+(\.\d+)?", "N", p)
                    patterns[p[:120]] += 1
        if not patterns:
            continue
        out.append(f"\n### {sheet}")
        for p, n in patterns.most_common(10):
            out.append(f"  ×{n}: {p}")
    out.append("")

    # Graph stats
    gd = Path(graph_db) if graph_db else xlsx_path.with_suffix(".graph.db")
    if gd.exists():
        out.append("## Linkage graph summary")
        try:
            with sqlite3.connect(gd) as conn:
                node_rows = conn.execute(
                    "SELECT kind, COUNT(*) FROM nodes GROUP BY kind"
                ).fetchall()
                edge_rows = conn.execute(
                    "SELECT kind, COUNT(*) FROM edges GROUP BY kind"
                ).fetchall()
            out.append(f"Nodes: {dict(node_rows)}")
            out.append(f"Edges: {dict(edge_rows)}")
        except sqlite3.OperationalError as e:
            out.append(f"(graph read error: {e})")
        out.append("")

    return "\n".join(out)


def build_system_prompt(xlsx_path: Path | str,
                        graph_db: Optional[Path | str] = None) -> str:
    """System prompt for Claude-powered chat on a workbook."""
    summary = workbook_summary(xlsx_path, graph_db)
    return (
        "You are ModelForge's lineage Q&A assistant. The user has built "
        "the Excel financial model summarised below and is asking "
        "questions about its cells, assumptions, and source citations. "
        "Answer with the following rules:\n"
        "\n"
        "  1. Always cite Assumption IDs (A-###) or Source IDs (S-###) "
        "when a number you mention has them. Use the format "
        "(A-012, confidence M) or (S-004 p.14).\n"
        "  2. When asked 'why is X?', walk the lineage: which driver "
        "sets it, which assumption or source provides the driver, and "
        "what rationale the analyst gave.\n"
        "  3. When asked 'which cells depend on A-###?', list the "
        "sheets and approximate locations; suggest running "
        "`modelforge lineage <graph.db> <CELL:id>` for the exact walk.\n"
        "  4. If the user asks about a number that is not in the "
        "summary (e.g. a computed value), explain what formula produces "
        "it by reading the formula shapes section.\n"
        "  5. Be concise. A senior MD reviews these answers — no "
        "filler, no meta-commentary on your process.\n"
        "  6. If something genuinely can't be answered from the "
        "summary, say so — don't fabricate.\n"
        "\n"
        "Here is the workbook summary (ground truth for this session):\n"
        "\n" + summary
    )


__all__ = ["build_system_prompt", "workbook_summary"]
