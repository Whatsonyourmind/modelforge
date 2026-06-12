"""Portfolio Review template — Template 16 (v0.10 PREVIEW).

Aggregator template: N portfolio companies × monitoring metrics in one
workbook. Different shape from single-deal templates — does not use
build_base_workbook framework.

v0.10: minimal renderer (Cover + Portfolio matrix + Aggregate summary + QC).
v0.11: sparklines, heatmaps, exception flagging.
v0.12: fund-level performance — Fund Returns sheet (MOIC, DPI, RVPI, TVPI,
gross & net IRR) when spec.fund_cashflows is supplied.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font

from modelforge.builder import layout, styles
from modelforge.builder.i18n import L

# Bold key-metric font that ALSO carries an explicit colour, so a bold summary
# cell still satisfies the certify styling gate (which requires both an explicit
# font colour and a number_format). styles.font_subheader is bold but colourless.
_FONT_METRIC_BOLD = Font(
    name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY, bold=True,
    color=styles.COLOR_FORMULA,
)


def build(spec, out_path: Path | str, graph_db_path=None):
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _build_cover_sheet(wb, spec)
    _build_portfolio_matrix_sheet(wb, spec)
    # Fund-level performance is rendered only when the fund cashflow stream is
    # supplied; otherwise the workbook stays a pure covenant/leverage monitor
    # (back-compatible with v0.10 manual builds that pass no fund_cashflows).
    if getattr(spec, "fund_cashflows", None):
        _build_fund_returns_sheet(wb, spec)
    _build_aggregate_summary_sheet(wb, spec)
    _build_qc_sheet(wb, spec)

    wb.save(out_path)

    # Portfolio review template doesn't use the linkage graph (no formulas
    # cross-reference per-portco rows). Return a null graph path so the
    # registry signature stays consistent.
    return out_path, out_path.with_suffix(".graph.db")


def _build_cover_sheet(wb, spec) -> None:
    ws = wb.create_sheet("Cover")
    layout.set_column_widths(ws, label_width=28, it_width=28, year_width=18)
    layout.write_title_block(
        ws,
        title_en=f"{spec.fund_name} — Portfolio Review {spec.review_quarter}",
        title_it=f"{spec.fund_name} — Riepilogo portafoglio {spec.review_quarter}",
        subtitle=f"{spec.meta.confidentiality} · {spec.meta.status.upper()} · {spec.meta.version}",
    )

    rows = [
        ("Fund", "Fondo", spec.fund_name),
        ("Vintage", "Annata", str(spec.fund_vintage) if spec.fund_vintage else ""),
        ("AUM (in unit_scale)", "AUM", str(spec.fund_aum) if spec.fund_aum else ""),
        ("Review quarter", "Trimestre", spec.review_quarter),
        ("Portfolio companies", "Società in portafoglio", str(len(spec.portfolio))),
        ("Analyst", "Analista", spec.meta.analyst),
        ("Valuation date", "Data valutazione", spec.meta.valuation_date.isoformat()),
        ("Version", "Versione", spec.meta.version),
    ]
    r = 4
    for en, it, val in rows:
        layout.write_row_label(ws, r, en, it)
        ws.cell(row=r, column=3, value=val)
        r += 1

    # ── Honest feature-scope note (SHIPPED vs roadmap) ────────────────────
    # The template is a covenant/leverage MONITOR plus (v0.12) a fund-level
    # performance section when a fund cashflow stream is supplied. Stated here
    # so the deliverable neither over- nor under-claims.
    has_fund = bool(getattr(spec, "fund_cashflows", None))
    r += 1
    scope = ws.cell(
        row=r, column=1,
        value="Feature scope — covenant / leverage monitor + fund returns",
    )
    styles.style_subheader(scope)
    r += 1
    fund_line_en = (
        "SHIPPED (v0.12): fund-level performance on the Fund Returns sheet — "
        "MOIC, DPI, RVPI, TVPI (= DPI + RVPI), gross & net IRR from the "
        "capital-call / distribution / NAV stream."
        if has_fund else
        "Fund-level performance (Fund Returns sheet) renders when a fund "
        "cashflow stream is supplied; none provided in this build."
    )
    fund_line_it = (
        "Performance a livello fondo: MOIC, DPI, RVPI, TVPI, IRR lordo e netto."
        if has_fund else
        "Performance a livello fondo disponibile se forniti i flussi del fondo."
    )
    scope_rows = [
        ("SHIPPED (v0.10): per-company covenant cushion, leverage, "
         "EBITDA actual-vs-plan, cash-trap flag, internal-rating roll-up.",
         "In ambito: cushion covenant, leva, EBITDA vs piano, rating."),
        (fund_line_en, fund_line_it),
        ("ROADMAP: trend sparklines, covenant cushion heatmap, automated "
         "exception flagging.",
         "Roadmap: sparkline di trend, heatmap covenant, flagging eccezioni."),
    ]
    for en, it in scope_rows:
        c = ws.cell(row=r, column=1, value=en)
        c.font = styles.font_label_en
        c.alignment = styles.align_left
        ws.cell(row=r, column=2, value=it).font = styles.font_label_it
        r += 1


def _build_portfolio_matrix_sheet(wb, spec) -> None:
    """N rows (one per portco) × M columns (metrics)."""
    ws = wb.create_sheet("Portfolio")
    layout.set_column_widths(ws, label_width=22, it_width=22, year_width=14)

    layout.write_title_block(
        ws,
        title_en="Portfolio Matrix",
        title_it="Matrice del portafoglio",
        subtitle="One row per portfolio company. Sort/filter as needed.",
    )

    # Header row at row 4
    headers = [
        "ID", "Name", "Sector", "Country",
        "Entry Lev", "Curr Lev",
        "Plan EBITDA (Q)", "Actual EBITDA (Q)", "Δ % vs Plan",
        "Cov Cushion %", "Cash Trap", "Rating",
        "Next Cov Test", "Narrative",
    ]
    r = 4
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=r, column=c, value=h)
        styles.style_label_en(cell)

    # Data rows
    for i, portco in enumerate(spec.portfolio):
        r = 5 + i
        cells_data = [
            portco.portco_id, portco.name, portco.sector, portco.country,
            portco.entry_leverage, portco.current_leverage,
            portco.plan_ebitda_q, portco.actual_ebitda_q,
            # Δ% vs Plan as formula
            None,
            portco.covenant_cushion_pct,
            "YES" if portco.cash_trap_active else "no",
            portco.rating_internal or "",
            portco.next_covenant_test_date.isoformat() if portco.next_covenant_test_date else "",
            portco.narrative,
        ]
        for c, v in enumerate(cells_data, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            # Δ% vs Plan: formula at column 9
            if c == 9 and portco.plan_ebitda_q and portco.actual_ebitda_q:
                # Excel cells use 1-indexed column letters
                from openpyxl.utils import get_column_letter
                plan_col = get_column_letter(7)
                actual_col = get_column_letter(8)
                cell.value = f"={actual_col}{r}/{plan_col}{r}-1"
                cell.number_format = "0.00%"


def _build_fund_returns_sheet(wb, spec) -> None:
    """Fund-level performance — MOIC / DPI / RVPI / TVPI / gross & net IRR.

    Renders the fund's own (LP-facing) cashflow stream and the standard PE
    fund KPIs from it. Multiples and IRR are emitted as LIVE Excel formulas
    over the cashflow table so a reviewer can audit them cell-by-cell; the
    same values are also reconciled in Python (finance_core + numpy_financial
    cross-check lives in hardtest_portfolio_review.py).

    Sign convention: capital calls and distributions are entered as positive
    magnitudes; the net LP cashflow (the IRR vector) is
    ``distribution - capital_call`` each period, with the TERMINAL period also
    crediting the residual NAV as a realisation inflow.
    """
    from openpyxl.utils import get_column_letter

    ws = wb.create_sheet("FundReturns")
    layout.set_column_widths(ws, label_width=40, it_width=30, year_width=14)
    layout.write_title_block(
        ws,
        title_en="Fund Returns",
        title_it="Rendimenti del fondo",
        subtitle="MOIC · DPI · RVPI · TVPI (= DPI + RVPI) · gross & net IRR "
                 f"({spec.meta.currency}, {spec.meta.unit_scale})",
    )

    cfs = sorted(spec.fund_cashflows, key=lambda c: c.period)
    n = len(cfs)
    last_i = n - 1

    # ── Cashflow table: metrics in rows, periods across columns D.. ──────────
    hdr_row = 4
    layout.write_section_header(ws, hdr_row, "Fund cashflow stream",
                                "Flussi di cassa del fondo")
    # Period header axis (integers — styled as inputs: they are the period axis)
    period_row = hdr_row + 1
    layout.write_row_label(ws, period_row, "Period (t)", "Periodo (t)")
    for i, cf in enumerate(cfs):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=period_row, column=col_idx, value=cf.period)
        styles.style_input(c, number_format=styles.FMT_INTEGER)

    call_row = period_row + 1
    layout.write_row_label(ws, call_row, "Capital call (paid-in)",
                           "Richiamo di capitale", indent=True)
    dist_row = call_row + 1
    layout.write_row_label(ws, dist_row, "Distribution to LPs",
                           "Distribuzione agli LP", indent=True)
    nav_row = dist_row + 1
    layout.write_row_label(ws, nav_row, "Residual NAV (end of period)",
                           "NAV residuo (fine periodo)", indent=True)
    for i, cf in enumerate(cfs):
        col_idx = ord(layout.year_col(i)) - ord("A") + 1
        c = ws.cell(row=call_row, column=col_idx, value=cf.capital_call)
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c = ws.cell(row=dist_row, column=col_idx, value=cf.distribution)
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        c = ws.cell(row=nav_row, column=col_idx, value=cf.nav)
        styles.style_input(c, number_format=styles.FMT_EUR_M)

    # Net LP cashflow (IRR vector): distribution - call, + terminal residual NAV
    vec_row = nav_row + 1
    layout.write_row_label(ws, vec_row, "Net LP cashflow (IRR vector)",
                           "Flusso netto LP (vettore IRR)")
    for i in range(n):
        col = layout.year_col(i)
        col_idx = ord(col) - ord("A") + 1
        if i == last_i:
            # terminal period realises the residual NAV
            f = f"={col}{dist_row}-{col}{call_row}+{col}{nav_row}"
        else:
            f = f"={col}{dist_row}-{col}{call_row}"
        c = ws.cell(row=vec_row, column=col_idx, value=f)
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = _FONT_METRIC_BOLD
        c.border = styles.BORDER_TOP_THIN
    ws.cell(row=vec_row, column=ord(layout.year_col(last_i)) - ord("A") + 1).comment = Comment(
        "Net LP cashflow each period = distribution - capital call. The final "
        "period also credits residual NAV as a realisation inflow, so the IRR "
        "vector treats end-of-life NAV as a terminal distribution.",
        "ModelForge",
    )

    first_col = layout.year_col(0)
    last_col = layout.year_col(last_i)

    # ── KPI block ────────────────────────────────────────────────────────────
    r = vec_row + 2
    layout.write_section_header(ws, r, "Fund performance — multiples & IRR",
                                "Performance del fondo — multipli e IRR")
    r += 1

    # Paid-in / distributed / terminal NAV / total value (live SUM formulas)
    paidin_row = r
    layout.write_row_label(ws, r, "Total paid-in (Σ capital calls)",
                           "Capitale richiamato (Σ)", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=SUM({first_col}{call_row}:{last_col}{call_row})")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    distrib_row = r
    layout.write_row_label(ws, r, "Total distributed (Σ distributions)",
                           "Distribuito (Σ)", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=SUM({first_col}{dist_row}:{last_col}{dist_row})")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    termnav_row = r
    layout.write_row_label(ws, r, "Terminal residual NAV",
                           "NAV residuo finale", indent=True)
    c = ws.cell(row=r, column=4, value=f"={last_col}{nav_row}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    r += 1

    totval_row = r
    layout.write_row_label(ws, r, "Total value (distributions + terminal NAV)",
                           "Valore totale (distribuzioni + NAV)", indent=True)
    c = ws.cell(row=r, column=4,
                value=f"=$D${distrib_row}+$D${termnav_row}")
    styles.style_formula(c, number_format=styles.FMT_EUR_M)
    c.font = _FONT_METRIC_BOLD
    r += 2

    # MOIC = total value / paid-in
    moic_row = r
    layout.write_row_label(ws, r, "MOIC (total value / paid-in)",
                           "MOIC (valore totale / richiamato)", indent=True)
    c = ws.cell(row=r, column=4, value=f"=$D${totval_row}/$D${paidin_row}")
    styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    c.font = _FONT_METRIC_BOLD
    c.comment = Comment(
        "MOIC = (cumulative distributions + terminal residual NAV) / total "
        "paid-in capital. Equals fund-level TVPI for a single LP class.",
        "ModelForge",
    )
    r += 1

    # DPI = distributions / paid-in
    dpi_row = r
    layout.write_row_label(ws, r, "DPI (distributions / paid-in)",
                           "DPI (distribuzioni / richiamato)", indent=True)
    c = ws.cell(row=r, column=4, value=f"=$D${distrib_row}/$D${paidin_row}")
    styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    r += 1

    # RVPI = residual NAV / paid-in
    rvpi_row = r
    layout.write_row_label(ws, r, "RVPI (residual NAV / paid-in)",
                           "RVPI (NAV residuo / richiamato)", indent=True)
    c = ws.cell(row=r, column=4, value=f"=$D${termnav_row}/$D${paidin_row}")
    styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    r += 1

    # TVPI = DPI + RVPI (model-independent identity, rendered as the sum)
    tvpi_row = r
    layout.write_row_label(ws, r, "TVPI (= DPI + RVPI)",
                           "TVPI (= DPI + RVPI)", indent=True)
    c = ws.cell(row=r, column=4, value=f"=$D${dpi_row}+$D${rvpi_row}")
    styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
    c.font = _FONT_METRIC_BOLD
    c.comment = Comment(
        "TVPI = DPI + RVPI by definition (total value to paid-in = realised "
        "plus unrealised value over paid-in). Rendered as the sum so the "
        "identity is auditable on the face of the workbook.",
        "ModelForge",
    )
    r += 1

    # Gross IRR (live IRR over the net LP cashflow vector)
    irr_row = r
    layout.write_row_label(ws, r, "Gross IRR (fund cashflow IRR)",
                           "IRR lordo (IRR flussi del fondo)", indent=True)
    c = ws.cell(
        row=r, column=4,
        value=f"=IRR({first_col}{vec_row}:{last_col}{vec_row},0.1)",
    )
    styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
    c.font = _FONT_METRIC_BOLD
    c.comment = Comment(
        "Gross fund IRR — the periodic IRR of the net LP cashflow vector "
        "(distributions less calls, with terminal NAV as a realisation). "
        "Cross-checked against numpy_financial.irr in the hardtest.",
        "ModelForge",
    )
    r += 1

    # ── Net (gross-to-net) bridge: management fee + carried interest ──────────
    # European whole-fund waterfall: carry applies to profit ABOVE an 8%-compounded
    # preferred return on paid-in, after management fees. All terms come from the
    # spec; the bridge is rendered with live formulas + a Python cross-check.
    if spec.mgmt_fee_rate is not None and spec.carry_rate is not None and spec.pref_rate is not None:
        committed = spec.fund_committed if spec.fund_committed is not None else sum(
            cf.capital_call for cf in cfs
        )
        years = cfs[last_i].period - cfs[0].period
        if years <= 0:
            years = n - 1

        r += 1
        layout.write_section_header(ws, r, "Gross-to-net bridge (2/20 over pref)",
                                    "Bridge lordo-netto (2/20 su hurdle)")
        r += 1

        # committed capital (input)
        committed_row = r
        layout.write_row_label(ws, r, "Committed capital", "Capitale impegnato", indent=True)
        c = ws.cell(row=r, column=4, value=float(committed))
        styles.style_input(c, number_format=styles.FMT_EUR_M)
        r += 1

        # management fee rate / carry rate / pref rate / fee years (inputs)
        feerate_row = r
        layout.write_row_label(ws, r, "Management fee rate (p.a.)", "Commissione gestione (p.a.)", indent=True)
        c = ws.cell(row=r, column=4, value=float(spec.mgmt_fee_rate))
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        r += 1
        feeyears_row = r
        layout.write_row_label(ws, r, "Fee years", "Anni di commissione", indent=True)
        c = ws.cell(row=r, column=4, value=int(years))
        styles.style_input(c, number_format=styles.FMT_INTEGER)
        r += 1
        carryrate_row = r
        layout.write_row_label(ws, r, "Carry rate", "Carried interest", indent=True)
        c = ws.cell(row=r, column=4, value=float(spec.carry_rate))
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        r += 1
        prefrate_row = r
        layout.write_row_label(ws, r, "Preferred return (hurdle)", "Hurdle (rendimento preferito)", indent=True)
        c = ws.cell(row=r, column=4, value=float(spec.pref_rate))
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        r += 1

        # total management fees = rate * committed * fee_years
        mgmtfee_row = r
        layout.write_row_label(ws, r, "Total management fees", "Commissioni totali", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${feerate_row}*$D${committed_row}*$D${feeyears_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # preferred amount = paid-in * ((1+pref)^years - 1). The hurdle compounds
        # on capital actually drawn, not committed (committed overstates the
        # hurdle and understates carry). Management fees stay on committed.
        pref_row = r
        layout.write_row_label(ws, r, "Preferred (hurdle) amount", "Importo hurdle", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${paidin_row}*((1+$D${prefrate_row})^$D${feeyears_row}-1)")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # gross profit = total value - committed
        grossprofit_row = r
        layout.write_row_label(ws, r, "Gross profit (total value − committed)",
                               "Profitto lordo", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${totval_row}-$D${committed_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # carry base = max(profit - pref - fees, 0)
        carrybase_row = r
        layout.write_row_label(ws, r, "Carry base = max(profit − pref − fees, 0)",
                               "Base carry", indent=True)
        c = ws.cell(
            row=r, column=4,
            value=f"=MAX($D${grossprofit_row}-$D${pref_row}-$D${mgmtfee_row},0)",
        )
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # carried interest = carry_rate * carry_base
        carry_row = r
        layout.write_row_label(ws, r, "Carried interest", "Carried interest", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${carryrate_row}*$D${carrybase_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        # net value = total value - fees - carry  ; net MOIC = net value / paid-in
        netvalue_row = r
        layout.write_row_label(ws, r, "Net value (gross − fees − carry)",
                               "Valore netto", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${totval_row}-$D${mgmtfee_row}-$D${carry_row}")
        styles.style_formula(c, number_format=styles.FMT_EUR_M)
        c.font = _FONT_METRIC_BOLD
        r += 1

        netmoic_row = r
        layout.write_row_label(ws, r, "Net MOIC (net value / paid-in)",
                               "MOIC netto", indent=True)
        c = ws.cell(row=r, column=4,
                    value=f"=$D${netvalue_row}/$D${paidin_row}")
        styles.style_formula(c, number_format=styles.FMT_MULTIPLE)
        c.font = _FONT_METRIC_BOLD
        c.comment = Comment(
            "Net MOIC = (total value − management fees − carried interest) / "
            "paid-in. Always ≤ gross MOIC. Carry is charged only on profit "
            "above the compounded preferred return (European whole-fund "
            "waterfall).",
            "ModelForge",
        )
        r += 1

        # Net IRR — the deliverable promises "gross & net IRR" but previously
        # rendered only gross. Build an auditable post-fee LP cashflow row, then
        # take its IRR. Fee/carry timing matches the gross-to-net bridge above:
        # annual management fee (committed × rate) on periods 1..N, total carried
        # interest charged at the terminal period.
        netvec_row = r
        layout.write_row_label(ws, r, "Net LP cashflow (after fees & carry)",
                               "CF netto LP (dopo fee e carry)", indent=True)
        for i in range(n):
            col = layout.year_col(i)
            col_idx = ord(col) - ord("A") + 1
            if i == 0:
                f = f"={col}{vec_row}"
            elif i == last_i:
                f = (f"={col}{vec_row}-($D${feerate_row}*$D${committed_row})"
                     f"-$D${carry_row}")
            else:
                f = f"={col}{vec_row}-($D${feerate_row}*$D${committed_row})"
            c = ws.cell(row=r, column=col_idx, value=f)
            styles.style_formula(c, number_format=styles.FMT_EUR_M)
        r += 1

        netirr_row = r
        layout.write_row_label(ws, r, "Net IRR (after fees & carry)",
                               "IRR netto (dopo fee e carry)", indent=True)
        c = ws.cell(
            row=r, column=4,
            value=f"=IRR({first_col}{netvec_row}:{last_col}{netvec_row},0.1)",
        )
        styles.style_formula(c, number_format=styles.FMT_PCT_2DP)
        c.font = _FONT_METRIC_BOLD
        c.comment = Comment(
            "Net IRR = periodic IRR of the LP cashflow net of annual management "
            "fees (committed × fee rate, periods 1..N) and carried interest "
            "(charged at the terminal period). Always ≤ gross IRR.",
            "ModelForge",
        )
        r += 1

    ws.freeze_panes = "D5"


def _build_aggregate_summary_sheet(wb, spec) -> None:
    """Roll-up summary across the portfolio."""
    ws = wb.create_sheet("Summary")
    layout.set_column_widths(ws, label_width=34, it_width=30, year_width=18)
    layout.write_title_block(
        ws,
        title_en="Aggregate Summary",
        title_it="Riepilogo aggregato",
        subtitle="Roll-up metrics across the portfolio.",
    )

    n = len(spec.portfolio)
    lev_values = [p.current_leverage for p in spec.portfolio if p.current_leverage is not None]
    cushion_values = [p.covenant_cushion_pct for p in spec.portfolio if p.covenant_cushion_pct is not None]
    cash_traps = sum(1 for p in spec.portfolio if p.cash_trap_active)
    rating_dist: dict[str, int] = {}
    for p in spec.portfolio:
        if p.rating_internal:
            rating_dist[p.rating_internal] = rating_dist.get(p.rating_internal, 0) + 1

    rows = [
        ("Total portfolio companies", "Totale società", n),
        ("Avg current leverage (×)", "Leva media corrente (×)",
         round(sum(lev_values) / len(lev_values), 2) if lev_values else "n/a"),
        ("Max current leverage (×)", "Leva max corrente (×)",
         round(max(lev_values), 2) if lev_values else "n/a"),
        ("Min covenant cushion %", "Min cushion covenant %",
         round(min(cushion_values), 1) if cushion_values else "n/a"),
        ("Cash-trap-active portcos", "Portco con cash-trap attivo", cash_traps),
        ("Rating 1 (outperform)", "Rating 1", rating_dist.get("1", 0)),
        ("Rating 2", "Rating 2", rating_dist.get("2", 0)),
        ("Rating 3 (on-plan)", "Rating 3 (on-plan)", rating_dist.get("3", 0)),
        ("Rating 4 (watch)", "Rating 4 (watch)", rating_dist.get("4", 0)),
        ("Rating 5 (workout)", "Rating 5 (workout)", rating_dist.get("5", 0)),
    ]
    r = 4
    for en, it, val in rows:
        layout.write_row_label(ws, r, en, it)
        c = ws.cell(row=r, column=3, value=val)
        if isinstance(val, float):
            c.number_format = "#,##0.00"
        r += 1


def _build_qc_sheet(wb, spec) -> None:
    """Minimal QC — every portco has portco_id + name; covenant cushion in plausible range."""
    ws = wb.create_sheet("QC")
    layout.write_title_block(
        ws,
        title_en="QC checks",
        title_it="Controlli QC",
        subtitle="Portfolio review v0.10 PREVIEW — minimal check set.",
    )
    layout.set_column_widths(ws, label_width=50, it_width=40, year_width=12)

    checks = [
        (
            f"All {len(spec.portfolio)} portcos have portco_id + name",
            f"Tutte le {len(spec.portfolio)} portco hanno ID + nome",
            all(p.portco_id and p.name for p in spec.portfolio),
        ),
        (
            "No duplicate portco_ids",
            "Nessun ID duplicato",
            len({p.portco_id for p in spec.portfolio}) == len(spec.portfolio),
        ),
        (
            "Covenant cushion %, when present, is in [-100, 500] range",
            "Cushion in range plausibile",
            all(
                (p.covenant_cushion_pct is None or -100.0 <= p.covenant_cushion_pct <= 500.0)
                for p in spec.portfolio
            ),
        ),
        (
            "Current leverage, when present, is positive",
            "Leva corrente positiva",
            all(
                (p.current_leverage is None or p.current_leverage > 0)
                for p in spec.portfolio
            ),
        ),
    ]
    r = 4
    for en, it, ok in checks:
        layout.write_row_label(ws, r, en, it)
        c = ws.cell(row=r, column=3, value="PASS" if ok else "FAIL")
        styles.style_label_en(c)
        r += 1
