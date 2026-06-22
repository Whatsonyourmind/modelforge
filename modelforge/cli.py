"""ModelForge CLI.

    modelforge build <spec.yaml>           → emit .xlsx + graph.db
    modelforge qc    <model.xlsx>          → run QC gate, print report
    modelforge sources <model.xlsx>        → list all sources used
    modelforge lineage <graph.db> <cell>   → walk back from cell to doc page
    modelforge stats <graph.db>            → graph node/edge counts
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Callable

import click
import yaml
from rich.console import Console
from rich.table import Table

from modelforge.graph.store import GraphStore
from modelforge.qc import run_qc
from modelforge.spec.unitranche import UnitrancheSpec
from modelforge.templates import REGISTRY, build_model

# Force UTF-8 output so help/text containing non-ASCII characters (e.g. the
# "χ²" in a backtest command docstring) doesn't raise UnicodeEncodeError on
# Windows cp1252 consoles.
for _stream in (sys.stdout, sys.stderr):
    _reconfigure = getattr(_stream, "reconfigure", None)
    if _reconfigure is not None:
        try:
            _reconfigure(encoding="utf-8", errors="replace")
        except (ValueError, OSError):
            pass


console = Console()


@click.group()
def main() -> None:
    """ModelForge — bulge-tier Excel model factory."""


# Map every model_type to a zero-arg loader that lazily imports + returns its
# Pydantic spec class. Lazy so a missing optional dep in one template can't
# break CLI startup. Keys MUST stay in lockstep with templates.REGISTRY — the
# _load_spec_class guard below asserts that invariant so the two can never
# silently drift again (the root cause of the original "Unknown model_type"
# bug for the 4 newest templates).
def _spec_loader_map() -> dict[str, "Callable[[], type]"]:
    def _unitranche():
        return UnitrancheSpec

    def _minibond():
        from modelforge.spec.minibond import MinibondSpec
        return MinibondSpec

    def _credit_memo():
        from modelforge.spec.credit_memo import CreditMemoSpec
        return CreditMemoSpec

    def _project_finance():
        from modelforge.spec.project_finance import ProjectFinanceSpec
        return ProjectFinanceSpec

    def _real_estate():
        from modelforge.spec.real_estate import RealEstateSpec
        return RealEstateSpec

    def _npl():
        from modelforge.spec.npl import NPLSpec
        return NPLSpec

    def _structured_credit():
        from modelforge.spec.structured_credit import StructuredCreditSpec
        return StructuredCreditSpec

    def _three_statement():
        from modelforge.spec.three_statement import ThreeStatementSpec
        return ThreeStatementSpec

    def _dcf():
        from modelforge.spec.dcf import DCFSpec
        return DCFSpec

    def _merger():
        from modelforge.spec.merger import MergerSpec
        return MergerSpec

    def _fairness():
        from modelforge.spec.fairness import FairnessSpec
        return FairnessSpec

    def _sponsor_lbo():
        from modelforge.spec.sponsor_lbo import SponsorLBOSpec
        return SponsorLBOSpec

    def _ipo():
        from modelforge.spec.ipo import IPOSpec
        return IPOSpec

    def _restructuring():
        from modelforge.spec.restructuring import RestructuringSpec
        return RestructuringSpec

    def _hgb_carveout():
        from modelforge.spec.hgb_carveout import HGBCarveoutSpec
        return HGBCarveoutSpec

    def _portfolio_review():
        from modelforge.spec.portfolio_review import PortfolioReviewSpec
        return PortfolioReviewSpec

    def _development_re():
        from modelforge.spec.development_re import DevelopmentRESpec
        return DevelopmentRESpec

    def _bank_fig():
        from modelforge.spec.bank_fig import BankFigSpec
        return BankFigSpec

    def _loan_tape_securitization():
        from modelforge.spec.loan_tape_securitization import LoanTapeSecuritizationSpec
        return LoanTapeSecuritizationSpec

    return {
        "unitranche": _unitranche,
        "minibond": _minibond,
        "credit_memo": _credit_memo,
        "project_finance": _project_finance,
        "real_estate": _real_estate,
        "npl": _npl,
        "structured_credit": _structured_credit,
        "three_statement": _three_statement,
        "dcf": _dcf,
        "merger": _merger,
        "fairness": _fairness,
        "sponsor_lbo": _sponsor_lbo,
        "ipo": _ipo,
        "restructuring": _restructuring,
        "hgb_carveout": _hgb_carveout,
        "portfolio_review": _portfolio_review,
        "development_re": _development_re,
        "bank_fig": _bank_fig,
        "loan_tape_securitization": _loan_tape_securitization,
    }


# Backwards-compat alias retained for any external caller importing this name.
SPEC_CLASSES = {
    "unitranche": UnitrancheSpec,
}


def _load_spec_class(model_type: str):
    """Resolve a model_type to its Pydantic spec class (registry-driven).

    The loader map is kept in lockstep with ``templates.REGISTRY`` so the CLI
    can build every shipped template — it can no longer drift behind the
    registry the way the original hard-coded if/elif chain did. The "Known:"
    list in the error is rebuilt dynamically from the registry.
    """
    loaders = _spec_loader_map()
    loader = loaders.get(model_type)
    if loader is not None:
        return loader()
    known = ", ".join(sorted(REGISTRY))
    raise ValueError(f"Unknown model_type {model_type!r}. Known: {known}")


def _warn_unknown_top_keys(raw: dict, SpecClass, source_name: str) -> None:
    """Warn (not fail) about unknown TOP-LEVEL spec keys.

    Pydantic silently drops an unknown key, so a fat-fingered field name
    (e.g. ``risk_free_ratee``) is dropped and the model builds with the default
    — a silent wrong-number risk. Surface it as a visible warning so it is never
    silent. (Stays a warning, not a hard error, so a slightly-ahead-of-schema
    spec still builds.)
    """
    # `screening` is an OPTIONAL, build-irrelevant convenience block read only by
    # the deal-screener (`modelforge screen`); it is intentionally not a spec
    # field, so tolerate it here rather than flag it as a typo on every spec that
    # opts into screening.
    tolerated = {"screening"}
    extras = sorted(set(raw) - set(SpecClass.model_fields) - tolerated)
    if extras:
        console.print(
            f"[yellow]warning:[/yellow] {source_name}: unknown field(s) ignored "
            f"(typo?): {', '.join(extras)}"
        )


def _check_dangling_sources(spec, source_name: str) -> None:
    """Fail on a referenced source_id that is not defined in `sources`.

    A `*_source_id` that points at a non-existent Source plants a fabricated
    "Source: S-xxx" citation on the workbook — an integrity defect in a finance
    deliverable. Walk the validated spec, collect every Source.id (defined) and
    every `*_source_id`/`source_id` value (referenced), and exit non-zero if any
    referenced id is undefined.
    """
    from pydantic import BaseModel as _BM
    from modelforge.spec.base import Source

    defined: set[str] = set()
    referenced: set[str] = set()
    seen: set[int] = set()

    def walk(o):
        if id(o) in seen:
            return
        if isinstance(o, _BM):
            seen.add(id(o))
            if isinstance(o, Source) and isinstance(o.id, str) and o.id.strip():
                defined.add(o.id)
            for fname, fval in o.__dict__.items():
                if ((fname == "source_id" or fname.endswith("_source_id"))
                        and isinstance(fval, str) and fval.strip()):
                    referenced.add(fval)
                walk(fval)
        elif isinstance(o, (list, tuple)):
            for it in o:
                walk(it)
        elif isinstance(o, dict):
            for it in o.values():
                walk(it)

    walk(spec)
    dangling = sorted(referenced - defined)
    if dangling:
        console.print(
            f"[red]{source_name}: source id(s) referenced but not defined in "
            f"`sources`: {', '.join(dangling)}. Add them to the sources list "
            f"(or fix the reference) — a dangling source_id plants a citation "
            f"for a source that isn't in the workbook.[/red]"
        )
        sys.exit(1)


def _load_and_validate_spec(spec_path: Path):
    """Parse + validate a YAML spec with FRIENDLY errors. Shared by build /
    certify / monte-carlo / audit so none of them ever dumps a raw Python
    traceback on a malformed or incomplete spec. Exits non-zero with a plain-
    language message on: YAML parse errors, a non-mapping document, an unknown
    model_type, or validation failures. Returns (spec, model_type, spec_bytes).
    """
    from pydantic import ValidationError
    from modelforge.spec.errors import format_validation_error

    spec_bytes = spec_path.read_bytes()
    try:
        raw = yaml.safe_load(spec_bytes)
    except yaml.YAMLError as e:
        console.print(f"[red]YAML parse error in {spec_path.name}:[/red] {e}")
        sys.exit(1)
    if not isinstance(raw, dict):
        console.print(f"[red]{spec_path.name} is not a YAML mapping "
                      f"(expected a spec document).[/red]")
        sys.exit(1)

    model_type = raw.get("model_type", "unitranche")
    try:
        SpecClass = _load_spec_class(model_type)
    except ValueError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(2)

    try:
        spec = SpecClass.model_validate(raw)
    except ValidationError as e:
        console.print(f"[red]{format_validation_error(e)}[/red]")
        sys.exit(1)

    _warn_unknown_top_keys(raw, SpecClass, spec_path.name)
    _check_dangling_sources(spec, spec_path.name)
    return spec, model_type, spec_bytes


def _finish_after_injection(xlsx, spec, spec_bytes, spec_path) -> None:
    """Re-run the deterministic finishing chain after Trust/Moat injection.

    ``build_model()`` already styles, timestamp-pins and manifest-hashes the
    core workbook, but the CLI injects the RedFlags + MOAT sheets *afterwards*.
    Without re-finishing, the delivered file carries unstyled injected cells
    (uncertifiable), wall-clock zip mtimes (non-deterministic bytes), and a
    manifest ``workbook_sha256`` that no longer matches the shipped bytes
    (``verify_manifest`` fails). Re-running the three steps here — as the last
    workbook-mutating action — makes the post-injection bytes the ones that get
    styled, pinned and hashed. Each step is independently guarded so a
    post-processing hiccup never blocks an already-built workbook. The styler is
    clock/RNG-free, so a same-spec rebuild stays byte-identical.
    """
    try:
        from openpyxl import load_workbook
        from modelforge.builder.styles import auto_color_xrefs, auto_style_gaps
        wb = load_workbook(xlsx, keep_links=True)
        auto_color_xrefs(wb)
        auto_style_gaps(wb)
        wb.save(xlsx)
    except Exception:
        pass
    try:
        from modelforge.analytics.reproducibility import finalize_determinism
        finalize_determinism(xlsx, spec, spec_source_bytes=spec_bytes)
    except Exception:
        pass
    try:
        from modelforge.analytics.manifest import write_manifest
        write_manifest(
            xlsx, spec, spec_source_bytes=spec_bytes, spec_source_path=spec_path,
        )
    except Exception:
        pass


def _inject_trust_moat_and_finish(xlsx, spec, spec_bytes, spec_path, *,
                                  trust_strict: bool = False,
                                  quiet: bool = False) -> None:
    """Inject the Trust-Layer RedFlags + Moat sheets, then re-finish the file.

    Shared by ``build`` and ``certify <spec>`` so both audit/ship the *same*
    delivered artifact. ``quiet=True`` suppresses console output (certify path).
    """
    # Trust Layer — plausibility RedFlags sheet.
    try:
        from modelforge.trust import (
            DEFAULT_RULES, TrustEngine, inject_red_flag_sheet,
        )
        report = TrustEngine(rules=DEFAULT_RULES).evaluate(xlsx, spec)
        inject_red_flag_sheet(xlsx, report)
    except Exception as e:
        if not quiet:
            console.print(f"[yellow]Trust Layer skipped:[/yellow] {e}")
    else:
        if not quiet:
            s = report.summary()
            if report.has_failures():
                console.print(
                    f"[red]Trust:[/red] FAIL={s['fail']} WARN={s['warn']} "
                    f"INFO={s['info']}  (see RedFlags sheet)"
                )
            elif report.has_warnings():
                console.print(
                    f"[yellow]Trust:[/yellow] WARN={s['warn']} INFO={s['info']}  "
                    "(see RedFlags sheet)"
                )
            else:
                console.print(
                    f"[green]Trust:[/green] all {s['rules_run']} rules pass"
                )
        if trust_strict and report.has_failures():
            # Keep even the aborted artifact internally consistent.
            _finish_after_injection(xlsx, spec, spec_bytes, spec_path)
            sys.exit(3)

    # MoatGate — "fully-formulated live Excel outputs" moat.
    try:
        from modelforge.moat import MoatGate, inject_moat_sheet
        moat_report = MoatGate().evaluate(xlsx)
        inject_moat_sheet(xlsx, moat_report)
    except Exception as e:
        if not quiet:
            console.print(f"[yellow]Moat gate skipped:[/yellow] {e}")
    else:
        if not quiet:
            summary = moat_report.summary()
            if moat_report.passes():
                console.print(
                    f"[green]Moat:[/green] {summary['gates_passed']}/"
                    f"{summary['gates_total']} gates  "
                    f"core-output density {summary['core_output_density']*100:.1f}%"
                )
            else:
                failed = [g.name for g in moat_report.gate_results if not g.passed]
                console.print(
                    f"[red]Moat:[/red] FAIL on {','.join(failed)}  "
                    f"(core-output density "
                    f"{summary['core_output_density']*100:.1f}%; see MOAT sheet)"
                )

    _finish_after_injection(xlsx, spec, spec_bytes, spec_path)


@main.command("build")
@click.argument("spec_path", type=click.Path(exists=True, path_type=Path))
@click.option("--out", "out_path", type=click.Path(path_type=Path), default=None,
              help="Output .xlsx path. Defaults to output/<spec_stem>.xlsx.")
@click.option("--no-trust", is_flag=True, default=False,
              help="Skip the Trust Layer plausibility pass (faster, less safe).")
@click.option("--trust-strict", is_flag=True, default=False,
              help="Exit non-zero if any FAIL-severity violation is found.")
def build_cmd(spec_path: Path, out_path: Path | None,
              no_trust: bool, trust_strict: bool) -> None:
    """Build an Excel workbook from a YAML spec.

    By default the Trust Layer runs after the build, injecting a `RedFlags`
    sheet with any plausibility violations. Use --no-trust to skip it
    or --trust-strict to fail the build on FAIL-severity issues.
    """
    spec, model_type, spec_bytes = _load_and_validate_spec(spec_path)

    if out_path is None:
        out_path = Path("output") / f"{spec_path.stem}.xlsx"

    xlsx, graph = build_model(
        spec, out_path,
        spec_source_bytes=spec_bytes,
        spec_source_path=spec_path,
    )
    console.print(f"[green]Built:[/green] {xlsx}  [dim]({model_type})[/dim]")
    console.print(f"[green]Graph:[/green] {graph}")

    if no_trust:
        return

    # Inject RedFlags + MOAT, then re-run the styler/determinism/manifest chain
    # so the *delivered* workbook is certifiable, byte-deterministic, and its
    # manifest hash matches the shipped bytes.
    _inject_trust_moat_and_finish(
        xlsx, spec, spec_bytes, spec_path,
        trust_strict=trust_strict, quiet=False,
    )


def _template_description(name: str) -> str:
    """Best-effort one-line description for a model_type.

    Prefers a curated blurb; otherwise falls back to the first line of the
    template module's docstring so newly-added templates self-document
    instead of showing a blank cell.
    """
    curated = {
        "unitranche": "Italian mid-market unitranche LBO (senior direct lending)",
        "minibond": "Italian minibond pricing + investor returns",
        "credit_memo": "Credit memo with covenant headroom + LGD + recovery analysis",
        "project_finance": "Infrastructure/RE project finance, DSCR-driven, construction + operating",
        "real_estate": "RE DCF with NOI build, exit cap, equity waterfall (pref + promote)",
        "npl": "NPL portfolio recovery waterfall (collection curves, IRR)",
        "structured_credit": "Securitization tranche waterfall (senior/mezz/junior)",
        "three_statement": "Classic 3-statement corporate model (P&L + BS + CFS)",
        "dcf": "DCF / WACC enterprise valuation with terminal value + sensitivity",
        "merger": "M&A merger model with EPS accretion/dilution + synergies",
        "fairness": "Fairness opinion football field (trading/transaction comps + DCF/LBO)",
        "sponsor_lbo": "Sponsor LBO — sources & uses, debt schedule, returns waterfall",
        "ipo": "IPO valuation + offering structure (primary/secondary, greenshoe)",
        "restructuring": "Restructuring / distressed recovery waterfall by class",
        "hgb_carveout": "HGB carve-out — standalone EBITDA bridge + carve-out EV",
        "portfolio_review": "PE fund portfolio review — TVPI/DPI/RVPI + gross & net IRR",
        "development_re": "Ground-up RE development — phased capex, lease-up S-curve, forward-NOI exit, LTC debt, promote",
        "bank_fig": "Bank / FIG — NII, RWA, CET1 & leverage ratios, MDA-gated dividends & buybacks",
        "loan_tape_securitization": "Loan-tape securitization (CLO/RMBS) — stratified tape, pool cashflow (CPR/CDR/recovery), sequential-pay turbo waterfall with OC/IC + reserve, note WAL/IRR/rating",
    }
    if name in curated:
        return curated[name]
    try:
        import importlib
        mod = importlib.import_module(f"modelforge.templates.{name}")
        doc = (mod.__doc__ or "").strip()
        if doc:
            return doc.splitlines()[0].strip()
    except Exception:
        pass
    return ""


@main.command("list-templates")
def list_templates_cmd() -> None:
    """List all available model templates."""
    from modelforge.templates import PREVIEW_TEMPLATES, REGISTRY
    tbl = Table(title="ModelForge templates")
    tbl.add_column("model_type", style="bold")
    tbl.add_column("Description")
    tbl.add_column("Status")
    # Registry-driven so every shipped template shows — the loader map and
    # the template registry are kept in lockstep, so all 16 appear here.
    for name in sorted(REGISTRY):
        if name in PREVIEW_TEMPLATES:
            status = "[yellow]preview[/yellow]"
        else:
            status = "[green]OK[/green]"
        tbl.add_row(name, _template_description(name), status)
    console.print(tbl)


@main.command("validate")
@click.argument("spec_path", type=click.Path(exists=True, path_type=Path))
@click.option("--max-errors", default=5, show_default=True, type=int,
              help="Maximum number of errors to show.")
def validate_cmd(spec_path: Path, max_errors: int) -> None:
    """Validate a YAML spec without building — fast pre-flight check.

    Parses the spec, resolves its model_type, and runs full Pydantic
    validation. On success prints a one-line OK and exits 0. On failure
    prints up to --max-errors friendly, plain-language fixes (e.g.
    "Missing required field: operating.ebitda_margin_by_year") and exits 1.
    """
    from pydantic import ValidationError
    from modelforge.spec.errors import format_validation_error

    try:
        raw = yaml.safe_load(spec_path.read_bytes())
    except yaml.YAMLError as e:
        console.print(f"[red]YAML parse error in {spec_path.name}:[/red] {e}")
        sys.exit(1)
    if not isinstance(raw, dict):
        console.print(f"[red]{spec_path.name} is not a YAML mapping "
                      f"(expected a spec document).[/red]")
        sys.exit(1)

    model_type = raw.get("model_type", "unitranche")
    try:
        SpecClass = _load_spec_class(model_type)
    except ValueError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(1)

    try:
        spec = SpecClass.model_validate(raw)
    except ValidationError as e:
        console.print(f"[red]{format_validation_error(e, limit=max_errors)}[/red]")
        sys.exit(1)

    _warn_unknown_top_keys(raw, SpecClass, spec_path.name)
    _check_dangling_sources(spec, spec_path.name)
    console.print(f"[green]OK[/green] {spec_path.name} is a valid "
                  f"[bold]{model_type}[/bold] spec.")
    sys.exit(0)


@main.command("schema")
@click.argument("model_type")
@click.option("--indent", default=2, show_default=True, type=int,
              help="JSON indent width.")
def schema_cmd(model_type: str, indent: int) -> None:
    """Print the JSON Schema for a model_type (for IDE autocomplete).

    Pipe to a file and reference it from your editor's YAML schema mapping
    to get inline completion + validation while authoring a spec:

        modelforge schema dcf > dcf.schema.json
    """
    import json
    try:
        SpecClass = _load_spec_class(model_type)
    except ValueError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(1)
    schema = SpecClass.model_json_schema()
    # Prepend the JSON-Schema dialect + an $id so editors/IDEs treat the output
    # as a real schema for the autocomplete use case this command advertises.
    schema = {
        "$schema": "https://json-schema.org/draft/2020-12/schema",
        "$id": f"modelforge://{model_type}.schema.json",
        **schema,
    }
    # Plain stdout (not rich) so the output is valid, pipeable JSON.
    print(json.dumps(schema, indent=indent, ensure_ascii=False))


@main.command("scaffold")
@click.argument("model_type")
@click.option("-o", "--output", "out_path", type=click.Path(path_type=Path),
              default=None, help="Write to a file instead of stdout.")
def scaffold_cmd(model_type: str, out_path: Path | None) -> None:
    """Emit a starter YAML spec skeleton for a model_type.

    Prints a ready-to-edit spec seeded from a shipped example, with a banner
    reminding you to replace the illustrative placeholder values and
    source IDs with your deal's real figures. Then:

        modelforge scaffold dcf -o my_dcf.yaml
        modelforge validate my_dcf.yaml
    """
    from modelforge.spec.scaffold import scaffold_yaml, _SEED_EXAMPLE
    try:
        SpecClass = _load_spec_class(model_type)
    except ValueError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(1)
    try:
        text = scaffold_yaml(model_type, SpecClass)
    except KeyError:
        console.print(f"[red]No scaffold available for {model_type!r}.[/red]")
        sys.exit(1)

    # ipo/restructuring ship no full example yet — their stub is a minimal
    # required-field skeleton that needs values before it validates/builds.
    is_stub = model_type not in _SEED_EXAMPLE

    if out_path is not None:
        out_path.write_text(text, encoding="utf-8")
        console.print(f"[green]Scaffold written:[/green] {out_path}")
        if is_stub:
            console.print(
                f"[yellow]note:[/yellow] {model_type} ships no full example yet — "
                f"this is a minimal skeleton; fill the required fields before "
                f"`validate`/`build`."
            )
        console.print(f"[dim]Next: modelforge validate {out_path}[/dim]")
    else:
        # Plain stdout so it pipes cleanly to a file.
        print(text)


@main.command("qc")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
def qc_cmd(xlsx_path: Path) -> None:
    """Run QC gate on a built workbook."""
    report = run_qc(xlsx_path)
    report.print()
    sys.exit(0 if report.all_pass else 1)


@main.command("audit-schedule")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("--strict", is_flag=True, default=False,
              help="Exit non-zero when any interior hardcode is found "
                   "(default is advisory: report but exit 0).")
def audit_schedule_cmd(xlsx_path: Path, strict: bool) -> None:
    """Flag hardcoded numbers wedged inside a formula series (certify-blind).

    `certify` proves a workbook has zero formula errors but says nothing about a
    cell that *should* be a formula but was overwritten with a bare number. This
    catches exactly that: a non-innocuous hardcode sitting between formula cells
    in a contiguous period series (e.g. a value typed over year 4 of a 10-year
    debt schedule). Input/reference/audit sheets and edge-period inputs are
    excluded. Advisory by default; pass --strict to gate on it.
    """
    from modelforge.qc import audit_schedule

    report = audit_schedule(xlsx_path)
    report.print()
    sys.exit(0 if (report.passed or not strict) else 1)


@main.command("audit-conservation")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("--strict", is_flag=True, default=False,
              help="Exit non-zero unless conservation PASSES (default is "
                   "advisory: report but exit 0).")
def audit_conservation_cmd(xlsx_path: Path, strict: bool) -> None:
    """Recompute the QC sheet and require ALL_PASS==1 (certify-blind).

    `certify` proves zero formula errors but a conservation check that
    legitimately recomputes to 0 (a balance sheet that doesn't balance, a debt
    roll-forward that telescopes) is a valid formula returning 0 — not an Excel
    error — so it certifies CERTIFIED today. This recomputes every formula with
    the `formulas` engine, reads the recomputed ALL CHECKS PASS aggregator plus
    every per-check, and requires the aggregator to be 1 with no per-check at 0.
    Advisory by default; pass --strict to gate on it. No recalc engine / no QC
    sheet → INDETERMINATE (a failure under --strict).
    """
    from modelforge.qc import audit_conservation

    report = audit_conservation(xlsx_path)
    report.print()
    sys.exit(0 if (report.passed or not strict) else 1)


@main.command("certify")
@click.argument("target", type=click.Path(exists=True, path_type=Path))
@click.option("--out", "out_path", type=click.Path(path_type=Path), default=None,
              help="Output .xlsx path when TARGET is a spec. "
                   "Defaults to output/<spec_stem>.xlsx.")
@click.option("--max-gaps", default=8, show_default=True, type=int,
              help="Max styling-gap cells to list.")
@click.option("--no-trust", is_flag=True, default=False,
              help="When TARGET is a spec, audit the core build only "
                   "(skip the Trust/Moat injection the `build` command ships).")
@click.option("--strict", is_flag=True, default=False,
              help="Exit non-zero on WARN too (treat styling gaps as failures).")
@click.option("--schedule", "schedule_check", is_flag=True, default=False,
              help="Also run the schedule auditor (interior hardcodes + "
                   "roll-forward period drift) and exit non-zero on any finding.")
@click.option("--conservation", "conservation_check", is_flag=True, default=False,
              help="Also recompute the QC sheet and require ALL_PASS==1 (no "
                   "per-check at 0); exit non-zero on a conservation FAIL or "
                   "INDETERMINATE result.")
def certify_cmd(target: Path, out_path: Path | None, max_gaps: int,
                no_trust: bool, strict: bool, schedule_check: bool,
                conservation_check: bool) -> None:
    """Certify a workbook has zero formula errors (the "no #REF!" gate).

    TARGET may be a built ``.xlsx`` or a ``.yaml``/``.yml`` spec. When a spec
    is given it is built exactly as the ``build`` command ships it — including
    the Trust/Moat sheets — so the badge is earned on the delivered artifact
    (pass ``--no-trust`` to audit only the core build). The auditor recomputes
    every formula with the third-party ``formulas`` engine and reports any Excel
    error cell (#REF!/#DIV0!/#VALUE!/#NAME?/#NUM!/#N/A) plus numeric cells
    lacking a font colour or number_format.

    Prints CERTIFIED (zero errors, no styling gaps), WARN (zero errors but
    some styling gaps), or FAIL (≥1 formula error). Exits 0 unless a formula
    error is found (exit 1); with ``--strict``, WARN also exits 1. Pass
    ``--schedule`` to additionally run the schedule auditor (interior hardcodes
    + roll-forward period drift — defects that are still valid formulas, so the
    formula-error audit cannot see them) and exit 1 on any finding.
    """
    from modelforge.qc import audit_workbook

    if target.suffix.lower() in (".yaml", ".yml"):
        spec, model_type, spec_bytes = _load_and_validate_spec(target)
        if out_path is None:
            out_path = Path("output") / f"{target.stem}.xlsx"
        xlsx, _ = build_model(
            spec, out_path,
            spec_source_bytes=spec_bytes,
            spec_source_path=target,
        )
        if not no_trust:
            # Build the same artifact `build` delivers, so the badge matches.
            _inject_trust_moat_and_finish(
                xlsx, spec, spec_bytes, target, quiet=True,
            )
        console.print(f"[green]Built:[/green] {xlsx}  [dim]({model_type})[/dim]")
        audit_target = xlsx
    else:
        audit_target = target

    report = audit_workbook(audit_target)
    s = report.summary()

    verdict = report.verdict
    if verdict == "CERTIFIED":
        badge = "[bold green]CERTIFIED[/bold green]"
    elif verdict == "WARN":
        badge = "[bold yellow]WARN[/bold yellow]"
    else:
        badge = "[bold red]FAIL[/bold red]"

    console.print(
        f"{badge}  {Path(audit_target).name}  "
        f"errors={s['error_cells']}  style-gaps={s['style_gaps']}  "
        f"numeric-cells={s['numeric_cells']}  "
        f"(recalc {'ran' if s['recalc_ran'] else 'skipped'})"
    )

    if report.error_cells:
        tbl = Table(title="Formula-error cells")
        tbl.add_column("Cell", style="bold")
        tbl.add_column("Error")
        tbl.add_column("Found via")
        for e in report.error_cells[:20]:
            tbl.add_row(e.ref, e.error, e.source)
        console.print(tbl)
        if report.n_errors > 20:
            console.print(f"[dim]… and {report.n_errors - 20} more.[/dim]")

    if report.style_gaps and max_gaps > 0:
        shown = report.style_gaps[:max_gaps]
        gap_refs = ", ".join(f"{g.ref} ({g.reason})" for g in shown)
        more = (f" … and {report.n_style_gaps - max_gaps} more"
                if report.n_style_gaps > max_gaps else "")
        console.print(f"[yellow]Styling gaps:[/yellow] {gap_refs}{more}")

    for n in report.notes:
        console.print(f"[dim]note: {n}[/dim]")

    # Opt-in schedule gate: interior hardcodes + roll-forward period drift —
    # defects that are valid formulas, so the formula-error audit above is blind
    # to them. Zero findings on every certified example, so this never fails a
    # ModelForge-built workbook; it catches a number typed over a formula or a
    # roll-forward repointed at the wrong period in a hand-edited model.
    schedule_ok = True
    if schedule_check:
        from modelforge.qc import audit_schedule

        srep = audit_schedule(audit_target)
        sbadge = "[bold green]SCHEDULE CLEAN[/bold green]" if srep.passed \
            else "[bold yellow]SCHEDULE REVIEW[/bold yellow]"
        console.print(
            f"{sbadge}  {srep.n_findings} finding(s), "
            f"{srep.sheets_scanned} sheet(s) scanned"
        )
        if srep.findings:
            tbl = Table(title="Schedule findings (certify-blind)")
            tbl.add_column("Cell", style="bold")
            tbl.add_column("Kind")
            tbl.add_column("Detail")
            for f in srep.findings[:20]:
                tbl.add_row(f.ref, f.kind, f.detail)
            console.print(tbl)
            if srep.n_findings > 20:
                console.print(f"[dim]… and {srep.n_findings - 20} more.[/dim]")
        schedule_ok = srep.passed

    # Opt-in conservation gate: recompute the QC sheet and require ALL_PASS==1.
    # A conservation check that recomputes to 0 (balance sheet that doesn't
    # balance, debt roll-forward that telescopes) is a valid formula returning 0
    # — not an Excel error — so the formula-error audit above is blind to it.
    conservation_ok = True
    if conservation_check:
        from modelforge.qc import audit_conservation

        crep = audit_conservation(audit_target)
        if crep.status == "PASS":
            cbadge = "[bold green]CONSERVATION PASS[/bold green]"
        elif crep.status == "FAIL":
            cbadge = "[bold red]CONSERVATION FAIL[/bold red]"
        else:
            cbadge = "[bold yellow]CONSERVATION INDETERMINATE[/bold yellow]"
        console.print(
            f"{cbadge}  {crep.n_checks} check(s), {crep.n_findings} failing"
        )
        if crep.findings:
            tbl = Table(title="Conservation findings (certify-blind)")
            tbl.add_column("Cell", style="bold")
            tbl.add_column("Check")
            tbl.add_column("Recomputed")
            for f in crep.findings[:20]:
                tbl.add_row(f.ref, f.label, f"{f.recomputed} ({f.status})")
            console.print(tbl)
            if crep.n_findings > 20:
                console.print(f"[dim]… and {crep.n_findings - 20} more.[/dim]")
        for n in crep.notes:
            console.print(f"[dim]note: {n}[/dim]")
        conservation_ok = crep.passed

    # Default: fail only on formula errors. --strict also fails on WARN
    # (styling gaps); --schedule also fails on any schedule finding;
    # --conservation also fails on a conservation FAIL/INDETERMINATE.
    ok = (report.passed and (verdict == "CERTIFIED" or not strict)
          and schedule_ok and conservation_ok)
    sys.exit(0 if ok else 1)


@main.command("sources")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
def sources_cmd(xlsx_path: Path) -> None:
    """List sources used in a workbook."""
    from openpyxl import load_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["Sources"]
    tbl = Table(title=f"Sources in {xlsx_path.name}")
    tbl.add_column("ID")
    tbl.add_column("Document")
    tbl.add_column("Page")
    tbl.add_column("Publisher")
    tbl.add_column("Verified")
    tbl.add_column("URL", overflow="fold")
    for row in range(6, ws.max_row + 1):
        vid = ws.cell(row=row, column=1).value
        if not vid:
            continue
        tbl.add_row(
            str(vid),
            str(ws.cell(row=row, column=2).value or ""),
            str(ws.cell(row=row, column=3).value or ""),
            str(ws.cell(row=row, column=4).value or ""),
            "✔" if ws.cell(row=row, column=7).value == "✔" else "",
            str(ws.cell(row=row, column=6).value or ""),
        )
    console.print(tbl)


@main.command("lineage")
@click.argument("graph_db", type=click.Path(exists=True, path_type=Path))
@click.argument("cell_id")
def lineage_cmd(graph_db: Path, cell_id: str) -> None:
    """Walk the linkage graph back from a cell.

    Example: modelforge lineage model.graph.db "CELL:OperatingModel!D10"
    """
    store = GraphStore(graph_db)
    # Infer model_id from any stored row
    import sqlite3
    with sqlite3.connect(graph_db) as conn:
        row = conn.execute("SELECT DISTINCT model_id FROM nodes LIMIT 1").fetchone()
    if not row:
        console.print("[red]Graph DB empty.[/red]")
        sys.exit(2)
    model_id = row[0]
    hops = store.lineage(model_id, cell_id)
    tbl = Table(title=f"Lineage of {cell_id}")
    tbl.add_column("ID")
    tbl.add_column("Kind")
    tbl.add_column("Label")
    for h in hops:
        tbl.add_row(h["id"], h["kind"], h["label"])
    console.print(tbl)


@main.command("ingest")
@click.argument("dataroom_dir", type=click.Path(exists=True, file_okay=False, path_type=Path))
@click.option("-t", "--template", default="project_finance",
              help="Target template (default: project_finance).")
@click.option("-o", "--output", "output_yaml", type=click.Path(path_type=Path), default=None,
              help="Output YAML path. Defaults to output/<dir_name>.yaml.")
@click.option("--model", default="claude-opus-4-7",
              help="Anthropic model (default: claude-opus-4-7).")
@click.option("--backend", "backend_name", default="cli",
              type=click.Choice(["cli", "api"]),
              help="LLM backend: 'cli' (Claude Code, no API key) or 'api' (Anthropic SDK).")
@click.option("--max-docs", type=int, default=50,
              help="Cap docs scanned (default: 50).")
@click.option("--strict", is_flag=True,
              help="Fail on any Pydantic validation error (default: best-effort).")
@click.option("--dry-run", is_flag=True,
              help="Classify + report only; skip extraction.")
@click.option("--no-cache", is_flag=True,
              help="Disable Anthropic prompt caching (api backend only).")
@click.option("-v", "--verbose", is_flag=True,
              help="Print per-step progress.")
def ingest_cmd(
    dataroom_dir: Path, template: str, output_yaml: Path | None,
    model: str, backend_name: str, max_docs: int, strict: bool,
    dry_run: bool, no_cache: bool, verbose: bool,
) -> None:
    """Ingest a data room (PDFs/XLSX/CSV) and produce a ModelForge YAML spec.

    Default backend is 'cli' which uses your Claude Code subscription
    (no API key needed). Use '--backend api' with ANTHROPIC_API_KEY for
    the Anthropic SDK path with prompt caching.
    """
    from modelforge.ingest.pipeline import ingest
    if output_yaml is None:
        output_yaml = Path("output") / f"{dataroom_dir.name}.yaml"

    def log(msg: str) -> None:
        if verbose:
            console.print(f"[dim]{msg}[/dim]")

    console.print(f"[bold]Ingesting[/bold] {dataroom_dir} -> {output_yaml}")
    console.print(f"Template: {template}  Backend: {backend_name}  Model: {model}")
    try:
        result = ingest(
            dataroom_dir=dataroom_dir,
            template=template,
            output_yaml=output_yaml,
            max_docs=max_docs,
            model=model,
            use_cache=not no_cache,
            strict=strict,
            dry_run=dry_run,
            backend_name=backend_name,
            log=log,
        )
    except RuntimeError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(2)
    except ValueError as e:
        console.print(f"[red]{e}[/red]")
        sys.exit(2)

    status = "[green]PASS[/green]" if result.spec_valid else "[yellow]needs review[/yellow]"
    console.print(f"\nSpec validation: {status}")
    console.print(f"Cache hit rate:  {result.cache_hit_rate*100:.1f}%")
    console.print(f"Tokens:          {result.total_input_tokens:,} in / {result.total_output_tokens:,} out")
    console.print(f"Elapsed:         {result.elapsed_seconds:.1f}s")
    console.print(f"\n[green]YAML[/green]:   {result.yaml_path}")
    console.print(f"[green]Report[/green]: {result.report_path}")
    if not dry_run and result.spec_valid:
        console.print(f"\nNext: [cyan]modelforge build {result.yaml_path}[/cyan]")
    sys.exit(0 if result.spec_valid or dry_run else 1)


@main.command("verify")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("--spec", "spec_path", type=click.Path(exists=True, path_type=Path),
              default=None, help="YAML spec to recompute hash from.")
def verify_cmd(xlsx_path: Path, spec_path: Path | None) -> None:
    """Read Reproducibility metadata and (optionally) verify spec hash.

    Without --spec, prints the stored metadata (hash, version, build
    timestamp). With --spec, recomputes the SHA-256 from the given YAML
    bytes and compares to the stored hash; exits 0 on match, 1 on
    mismatch.
    """
    from modelforge.analytics.reproducibility import (
        read_reproducibility,
        verify_spec_hash,
    )

    meta = read_reproducibility(xlsx_path)
    if not meta:
        console.print(f"[red]No Reproducibility metadata in {xlsx_path.name}. "
                      f"Rebuild with modelforge build.[/red]")
        sys.exit(2)

    tbl = Table(title=f"Reproducibility — {xlsx_path.name}")
    tbl.add_column("Field", style="bold")
    tbl.add_column("Value", overflow="fold")
    for k, v in meta.items():
        tbl.add_row(k, v)
    console.print(tbl)

    if spec_path is None:
        sys.exit(0)

    spec_bytes = spec_path.read_bytes()
    match, stored, recomputed = verify_spec_hash(xlsx_path, spec_bytes)
    if match:
        console.print(f"\n[green]PASS[/green] spec hash matches: {stored}")
        sys.exit(0)
    console.print(
        f"\n[red]FAIL[/red] spec hash mismatch.\n"
        f"  stored:     {stored}\n"
        f"  recomputed: {recomputed}"
    )
    sys.exit(1)


@main.command("monte-carlo")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("--spec", "spec_path", type=click.Path(exists=True, path_type=Path),
              required=True,
              help="YAML spec the workbook was built from (drives factor "
                   "selection + the shadow engine for the exact resim).")
@click.option("--runs", "n_runs", type=int, default=None,
              help="Number of Monte-Carlo draws. Default: MCConfig default (1000).")
@click.option("--dist", "distribution",
              type=click.Choice(["triangular", "normal", "lognormal"]),
              default=None, help="Per-factor shock distribution. Default: triangular.")
@click.option("--seed", "seed", type=int, default=None,
              help="RNG seed. Default: MCConfig deterministic seed.")
def monte_carlo_cmd(xlsx_path: Path, spec_path: Path,
                    n_runs: int | None, distribution: str | None,
                    seed: int | None) -> None:
    """Refresh the MonteCarlo sheet on a built workbook.

    Re-runs the SAMPLED Monte-Carlo simulation (a sampled distribution cannot
    recompute inside vanilla Excel, so its stats + histogram are an honest
    build-time snapshot) and re-writes the MonteCarlo sheet. The sheet also
    carries a genuinely-LIVE parametric (normal-approx) band — P5/P50/P95/mean
    as Excel formulas off ``primary_output`` + the editable ``mc_vol`` input —
    which recomputes in-cell on a scenario flip or vol edit; this command
    re-seeds ``mc_vol`` to the freshly-sampled std so the two blocks agree at
    the as-built scenario.

    This is the refresh path the snapshot banner points analysts to.
    """
    from modelforge.analytics.monte_carlo import (
        MCConfig, append_monte_carlo_sheet,
    )

    spec, model_type, spec_bytes = _load_and_validate_spec(spec_path)

    base = MCConfig()
    config = MCConfig(
        n_runs=n_runs if n_runs is not None else base.n_runs,
        distribution=distribution if distribution is not None else base.distribution,
        seed=seed if seed is not None else base.seed,
    )

    result = append_monte_carlo_sheet(xlsx_path, spec, config=config)
    if result is None:
        console.print(
            f"[yellow]Monte-Carlo skipped:[/yellow] no locatable primary "
            f"output / applicable factors in {xlsx_path.name}."
        )
        sys.exit(1)
    console.print(
        f"[green]Monte-Carlo refreshed:[/green] {result}  "
        f"[dim]({config.n_runs:,}-run {config.distribution}, seed={config.seed}; "
        f"sampled block = snapshot, parametric band = LIVE off primary_output/"
        f"mc_vol)[/dim]"
    )


@main.command("roi")
@click.option("--deals", "deals_per_year", type=int, default=20,
              show_default=True, help="Deals closed per year.")
@click.option("--hours-legacy", "hours_legacy", type=float, default=40.0,
              show_default=True, help="Analyst hours per deal (hand-built).")
@click.option("--hours-mf", "hours_mf", type=float, default=6.0,
              show_default=True, help="Analyst hours per deal with ModelForge.")
@click.option("--rate", "rate", type=float, default=180.0, show_default=True,
              help="Loaded analyst cost per hour (€).")
@click.option("--legacy-error-rate", type=float, default=0.15,
              show_default=True, help="Share of deals needing rework today.")
@click.option("--mf-error-rate", type=float, default=0.03, show_default=True,
              help="Share of deals needing rework with ModelForge.")
@click.option("--audit-hours-legacy", type=float, default=20.0,
              show_default=True, help="Audit hours per deal today.")
@click.option("--audit-hours-mf", type=float, default=4.0, show_default=True,
              help="Audit hours per deal with ModelForge.")
@click.option("--seats", type=int, default=3, show_default=True)
@click.option("--seat-price", "seat_price", type=float, default=499.0,
              show_default=True, help="Monthly price per seat (€).")
@click.option("--customer", default="(customer)",
              help="Customer name for the markdown header.")
@click.option("-o", "--output", "md_path", type=click.Path(path_type=Path),
              default=None, help="Markdown one-pager export path.")
def roi_cmd(deals_per_year: int, hours_legacy: float, hours_mf: float,
            rate: float, legacy_error_rate: float, mf_error_rate: float,
            audit_hours_legacy: float, audit_hours_mf: float,
            seats: int, seat_price: float,
            customer: str, md_path: Path | None) -> None:
    """Compute the business case for a prospective ModelForge buyer.

    Takes fund-specific assumptions (deal volume, analyst cost, current
    error + audit workload), outputs annual savings, ROI, and payback.
    Deterministic — every number traces to Python in
    `modelforge.roi.calculator`.
    """
    from modelforge.roi import ROIInputs, compute_roi, render_markdown
    inp = ROIInputs(
        deals_per_year=deals_per_year,
        hours_per_deal_legacy=hours_legacy,
        hours_per_deal_modelforge=hours_mf,
        loaded_analyst_cost_eur_per_hour=rate,
        legacy_error_rate=legacy_error_rate,
        modelforge_error_rate=mf_error_rate,
        audit_hours_legacy=audit_hours_legacy,
        audit_hours_modelforge=audit_hours_mf,
        seats=seats,
        monthly_price_per_seat_eur=seat_price,
    )
    res = compute_roi(inp)

    tbl = Table(title=f"ModelForge ROI — {customer}")
    tbl.add_column("Metric", style="bold")
    tbl.add_column("Value", justify="right")
    tbl.add_row("Hours saved per deal", f"{res.hours_saved_per_deal:.1f}")
    tbl.add_row("Annual time savings",
                f"EUR {res.annual_time_savings_eur:,.0f}")
    tbl.add_row("Rework reduction savings",
                f"EUR {res.rework_savings_eur:,.0f}")
    tbl.add_row("Audit time savings",
                f"EUR {res.audit_savings_eur:,.0f}")
    tbl.add_row("[bold]Gross annual savings[/bold]",
                f"[bold]EUR {res.total_gross_savings_eur:,.0f}[/bold]")
    tbl.add_row("Subscription cost",
                f"EUR {res.subscription_cost_eur:,.0f}")
    tbl.add_row("[bold]Net annual savings[/bold]",
                f"[bold]EUR {res.net_savings_eur:,.0f}[/bold]")
    tbl.add_row("1-year ROI", f"{res.roi_1y_pct:.1%}")
    tbl.add_row("Payback period",
                f"{res.payback_months:.1f} months")
    console.print(tbl)

    for n in res.notes:
        console.print(f"[yellow]note:[/yellow] {n}")

    if md_path is not None:
        md_path.write_text(render_markdown(res, customer=customer),
                            encoding="utf-8")
        console.print(f"[green]Markdown:[/green] {md_path}")


@main.command("scan")
@click.argument("folder", type=click.Path(exists=True, file_okay=False,
                                           path_type=Path))
@click.option("--persist", is_flag=True,
              help="Save the new state as the baseline (default: report-only).")
@click.option("-o", "--output", "md_path", type=click.Path(path_type=Path),
              default=None, help="Optional markdown report path.")
def scan_cmd(folder: Path, persist: bool, md_path: Path | None) -> None:
    """Diff a data-room folder against a persisted baseline.

    First run with --persist stores a baseline at
    `.modelforge/baseline.json` inside the folder. Subsequent runs
    report what's been added, modified, or removed since. Foundation
    for a v1.0 dataroom-watcher agent that auto-triggers ingest +
    diff + alerting on change.
    """
    from modelforge.watch import scan_folder
    change, _ = scan_folder(folder, persist=persist)

    tbl = Table(title=f"Data-room scan — {folder} "
                       f"({len(change.added)} added, "
                       f"{len(change.modified)} modified, "
                       f"{len(change.removed)} removed)")
    tbl.add_column("Change", style="bold")
    tbl.add_column("Path")
    tbl.add_column("Size (B)", justify="right")
    for fp in change.added:
        tbl.add_row("[green]+ added[/green]", fp.path, f"{fp.size:,}")
    for old, new in change.modified:
        tbl.add_row("[yellow]~ modified[/yellow]", new.path,
                    f"{old.size:,} -> {new.size:,}")
    for fp in change.removed:
        tbl.add_row("[red]- removed[/red]", fp.path, f"{fp.size:,}")
    if change.clean:
        tbl.add_row("[dim]no changes[/dim]", "", "")
    console.print(tbl)

    if md_path is not None:
        md = [f"# Data-room scan — {folder}", "",
              f"**{change.n_changes} change(s)**", ""]
        if change.added:
            md.append("## Added")
            for fp in change.added:
                md.append(f"- `{fp.path}` ({fp.size:,} bytes)")
            md.append("")
        if change.modified:
            md.append("## Modified")
            for old, new in change.modified:
                md.append(f"- `{new.path}` ({old.size:,} -> {new.size:,} bytes)")
            md.append("")
        if change.removed:
            md.append("## Removed")
            for fp in change.removed:
                md.append(f"- `{fp.path}`")
            md.append("")
        md_path.write_text("\n".join(md), encoding="utf-8")
        console.print(f"[green]Markdown:[/green] {md_path}")

    sys.exit(0 if change.clean else 1)


@main.command("watch")
@click.argument("folder", type=click.Path(exists=True, file_okay=False,
                                           path_type=Path))
@click.option("--interval", default=60, show_default=True, type=int,
              help="Polling interval in seconds.")
@click.option("--no-persist", "no_persist", is_flag=True,
              help="Do not save the baseline between scans.")
def watch_cmd(folder: Path, interval: int, no_persist: bool) -> None:
    """Continuously poll a data-room folder for changes.

    Prints a running summary every `interval` seconds. Ctrl-C to stop.
    For a one-shot report, use `modelforge scan` instead.
    """
    import time
    from datetime import datetime as _dt
    from modelforge.watch import scan_folder

    console.print(f"[bold]Watching[/bold] {folder} (interval={interval}s) — "
                  "Ctrl-C to stop.")
    persist = not no_persist
    try:
        while True:
            change, _ = scan_folder(folder, persist=persist)
            stamp = _dt.now().strftime("%H:%M:%S")
            if change.clean:
                console.print(f"[dim]{stamp}[/dim]  no changes")
            else:
                console.print(
                    f"[dim]{stamp}[/dim]  [green]+{len(change.added)}[/green] "
                    f"[yellow]~{len(change.modified)}[/yellow] "
                    f"[red]-{len(change.removed)}[/red]"
                )
                for fp in change.added:
                    console.print(f"  [green]+ added[/green] {fp.path}")
                for old, new in change.modified:
                    console.print(f"  [yellow]~ modified[/yellow] {new.path}")
                for fp in change.removed:
                    console.print(f"  [red]- removed[/red] {fp.path}")
            time.sleep(interval)
    except KeyboardInterrupt:
        console.print("\n[dim]stopped.[/dim]")


@main.command("serve")
@click.option("--host", default="127.0.0.1", show_default=True,
              help="Bind host. Use 0.0.0.0 to expose on the network.")
@click.option("--port", default=8000, show_default=True, type=int)
@click.option("--session-dir", "session_dir", type=click.Path(path_type=Path),
              default=None,
              help="Persist uploaded workbooks in this directory.")
def serve_cmd(host: str, port: int, session_dir: Path | None) -> None:
    """Launch the ModelForge web thin layer (FastAPI + uvicorn).

    Exposes HTTP endpoints: upload xlsx, view metadata, download
    dossier PDF, drift check vs live feeds, diff between two uploaded
    workbooks, one-shot risk analysis. In-memory + disk-backed
    workbook registry keyed by content hash.
    """
    try:
        import uvicorn
    except ImportError:
        console.print("[red]uvicorn not installed. Run "
                      "`pip install 'modelforge[web]'`.[/red]")
        sys.exit(2)
    from modelforge.web import create_app
    app = create_app(session_dir=session_dir)
    console.print(f"[bold]ModelForge web[/bold] — http://{host}:{port}/")
    console.print("[dim]Ctrl-C to stop.[/dim]")
    uvicorn.run(app, host=host, port=port, log_level="info")


@main.command("drift")
@click.argument("path", type=click.Path(exists=True, path_type=Path))
@click.option("--portfolio", is_flag=True,
              help="PATH is a directory — scan every .xlsx inside.")
@click.option("--threshold-bps", default=50.0, show_default=True, type=float,
              help="Absolute bps delta to flag for rate drivers.")
@click.option("--threshold-rel", default=0.10, show_default=True, type=float,
              help="Relative delta to flag for non-rate drivers.")
@click.option("-o", "--output", "md_path", type=click.Path(path_type=Path),
              default=None, help="Optional markdown report path.")
def drift_cmd(path: Path, portfolio: bool, threshold_bps: float,
              threshold_rel: float, md_path: Path | None) -> None:
    """Flag workbook drivers that have drifted from current market values.

    Reads ECB + Damodaran feeds and compares to each workbook's
    Assumption BASE cells. Flags any driver where |Δbps| ≥
    threshold-bps (rate drivers) or |Δ%| ≥ threshold-rel (value
    drivers). Pass --portfolio to sweep every .xlsx in a folder.

    Exits 0 clean, 1 on any flag.
    """
    from modelforge.drift import (
        check_drift, check_portfolio, render_markdown, render_portfolio_markdown,
    )
    if portfolio:
        if not path.is_dir():
            console.print(f"[red]{path} is not a directory.[/red]")
            sys.exit(2)
        p_rep = check_portfolio(path, threshold_bps=threshold_bps,
                                 threshold_rel=threshold_rel)
        tbl = Table(title=f"Portfolio drift — {path} "
                           f"({p_rep.n_workbooks} workbook(s); "
                           f"{p_rep.n_flagged_workbooks} with flags; "
                           f"{p_rep.total_flags} total flags)")
        tbl.add_column("Workbook", style="bold")
        tbl.add_column("Checked", justify="right")
        tbl.add_column("Flagged", justify="right")
        tbl.add_column("Top driver drift")
        for r in p_rep.per_workbook:
            top = max(r.flagged, key=lambda i: abs(i.delta_bps), default=None)
            top_str = (f"{top.driver_name} {top.delta_bps:+,.0f}bps"
                       if top else "—")
            tbl.add_row(r.xlsx_path.name, str(r.checked_drivers),
                        str(r.n_flagged), top_str)
        console.print(tbl)
        if md_path is not None:
            md_path.write_text(render_portfolio_markdown(p_rep),
                                encoding="utf-8")
            console.print(f"[green]Markdown:[/green] {md_path}")
        sys.exit(0 if p_rep.clean else 1)

    # Single-workbook mode
    rep = check_drift(path, threshold_bps=threshold_bps,
                      threshold_rel=threshold_rel)

    tbl = Table(title=f"Drift report — {path.name} "
                       f"({rep.checked_drivers} drivers checked, "
                       f"{rep.n_flagged} flagged)")
    tbl.add_column("Driver", style="bold")
    tbl.add_column("Assumed", justify="right")
    tbl.add_column("Current", justify="right")
    tbl.add_column("Δ bps", justify="right")
    tbl.add_column("Δ %", justify="right")
    tbl.add_column("Source")
    tbl.add_column("Flag")
    for it in rep.items:
        assumed_f = (f"{it.assumed_value:.4%}" if it.kind == "rate"
                     else f"{it.assumed_value:,.3f}")
        current_f = (f"{it.current_value:.4%}" if it.kind == "rate"
                     else f"{it.current_value:,.3f}")
        flag = "[red]⚠ FLAG[/red]" if it.flagged else "[green]OK[/green]"
        tbl.add_row(it.driver_name, assumed_f, current_f,
                    f"{it.delta_bps:+,.1f}", f"{it.delta_rel:+.2%}",
                    it.source, flag)
    console.print(tbl)

    if rep.missing_drivers:
        console.print(f"[dim]Missing from workbook ({len(rep.missing_drivers)}): "
                      f"{', '.join(rep.missing_drivers[:8])}"
                      f"{'...' if len(rep.missing_drivers) > 8 else ''}[/dim]")

    if md_path is not None:
        md_path.write_text(render_markdown(rep), encoding="utf-8")
        console.print(f"[green]Markdown:[/green] {md_path}")

    sys.exit(0 if rep.clean else 1)


@main.command("edgar")
@click.argument("ticker")
@click.option("--years", default=5, show_default=True, type=int,
              help="Number of historical fiscal years to pull.")
@click.option("--offline", is_flag=True,
              help="Use bundled sample (AAPL) instead of live EDGAR.")
@click.option("--json-out", "json_out", type=click.Path(path_type=Path),
              default=None, help="Optional JSON export path.")
def edgar_cmd(ticker: str, years: int, offline: bool,
              json_out: Path | None) -> None:
    """Pull XBRL-tagged US financials from SEC EDGAR (no API key).

    Example: `modelforge edgar AAPL --years 5` pulls Apple's last 5
    fiscal years of revenue / operating income / net income, plus
    most-recent total assets and long-term debt. Public fair-use API;
    sends a User-Agent per SEC policy.
    """
    from modelforge.ingest.edgar import fetch_company_financials

    fin = fetch_company_financials(ticker, years=years, prefer_bundled=offline)
    if fin is None:
        console.print(
            f"[red]Could not fetch EDGAR facts for {ticker.upper()!r}. "
            f"Either the ticker is not US-listed or EDGAR is unreachable. "
            f"Try --offline for the bundled AAPL sample.[/red]")
        sys.exit(2)

    tbl = Table(title=f"EDGAR company facts — {fin.ticker} "
                       f"({fin.entity_name}, CIK {fin.cik})")
    tbl.add_column("Fiscal Year", style="bold")
    tbl.add_column("Revenue ($m)")
    tbl.add_column("Operating income ($m)")
    tbl.add_column("Net income ($m)")
    for i, fy in enumerate(fin.fiscal_years):
        tbl.add_row(
            str(fy),
            f"{fin.revenue_usd_m[i]:,.0f}",
            f"{fin.operating_income_usd_m[i]:,.0f}",
            f"{fin.net_income_usd_m[i]:,.0f}",
        )
    console.print(tbl)

    summary = Table(title="Snapshot (most recent FY)")
    summary.add_column("Metric", style="bold")
    summary.add_column("Value")
    if fin.total_assets_usd_m is not None:
        summary.add_row("Total assets ($m)", f"{fin.total_assets_usd_m:,.0f}")
    if fin.long_term_debt_usd_m is not None:
        summary.add_row("Long-term debt ($m)",
                        f"{fin.long_term_debt_usd_m:,.0f}")
    summary.add_row("Source", fin.source_url)
    summary.add_row("Fetched from", fin.fetched_from)
    console.print(summary)

    for n in fin.notes:
        console.print(f"[yellow]note:[/yellow] {n}")

    if json_out is not None:
        import json as _json
        from dataclasses import asdict
        json_out.write_text(_json.dumps(asdict(fin), indent=2,
                                         default=str), encoding="utf-8")
        console.print(f"[green]JSON:[/green] {json_out}")


@main.command("risk")
@click.option("--equity", "equity_value", type=float, required=True,
              help="Market equity value (€m) — market cap.")
@click.option("--equity-vol", "equity_volatility", type=float, required=True,
              help="Annualized equity volatility (decimal, e.g. 0.30).")
@click.option("--debt", "debt_face_value", type=float, required=True,
              help="Debt face value at maturity (€m).")
@click.option("--risk-free", "risk_free_rate", type=float, default=0.039,
              show_default=True,
              help="Continuously compounded risk-free rate (decimal).")
@click.option("--horizon", "horizon_years", type=float, default=1.0,
              show_default=True, help="Default-horizon years.")
@click.option("--lgd", type=float, default=0.45, show_default=True,
              help="Loss given default (decimal).")
@click.option("--eir", type=float, default=0.05, show_default=True,
              help="Effective interest rate for ECL discounting.")
@click.option("--maturity", type=int, default=5, show_default=True,
              help="Loan maturity (years) for lifetime ECL.")
@click.option("--dpd", "days_past_due", type=int, default=0, show_default=True,
              help="Days past due — drives stage inference.")
@click.option("--exposure", "ead", type=float, default=None,
              help="EAD (€m). Defaults to debt face value.")
@click.option("--origination-pd", "origination_pd", type=float, default=None,
              help="12-mo PD at origination (for SICR detection).")
@click.option("--counterparty", default=None, help="Display name for the report.")
@click.option("--json-out", "json_out", type=click.Path(path_type=Path),
              default=None, help="Optional JSON export path.")
def risk_cmd(equity_value: float, equity_volatility: float,
             debt_face_value: float, risk_free_rate: float,
             horizon_years: float, lgd: float, eir: float,
             maturity: int, days_past_due: int,
             ead: float | None, origination_pd: float | None,
             counterparty: str | None, json_out: Path | None) -> None:
    """Run Merton + KMV + IFRS 9 ECL in one shot on a counterparty.

    Example: `modelforge risk --equity 500 --equity-vol 0.32 --debt 400
    --lgd 0.45 --maturity 7 --counterparty "CDMO SpA"`.
    """
    from modelforge.risk import (
        ECLInputs, MertonInputs, calibrate_pd_kmv, compute_ecl, solve_merton,
    )

    merton = solve_merton(MertonInputs(
        equity_value=equity_value, equity_volatility=equity_volatility,
        debt_face_value=debt_face_value, risk_free_rate=risk_free_rate,
        horizon_years=horizon_years,
    ))
    kmv_pd = calibrate_pd_kmv(merton.distance_to_default)

    # Use max(Merton-theoretical, KMV-empirical) for the ECL curve —
    # rating-agency-defensible conservative choice.
    pd_for_ecl = max(merton.probability_of_default, kmv_pd)
    ecl_ead = ead if ead is not None else debt_face_value
    ecl_inp = ECLInputs(
        exposure_at_default_eur_m=ecl_ead,
        loss_given_default=lgd,
        effective_interest_rate=eir,
        maturity_years=maturity,
        pd_curve_annual=[pd_for_ecl] * maturity,
        current_pd_12m=pd_for_ecl,
        origination_pd_12m=origination_pd or pd_for_ecl,
        days_past_due=days_past_due,
    )
    ecl = compute_ecl(ecl_inp, counterparty or "counterparty")

    cp = counterparty or "counterparty"
    tbl = Table(title=f"Risk analysis — {cp}")
    tbl.add_column("Metric", style="bold")
    tbl.add_column("Value")
    tbl.add_row("Equity value (€m)", f"{equity_value:,.2f}")
    tbl.add_row("Equity volatility", f"{equity_volatility:.2%}")
    tbl.add_row("Debt face (€m)", f"{debt_face_value:,.2f}")
    tbl.add_row("Risk-free rate", f"{risk_free_rate:.2%}")
    tbl.add_row("", "")
    tbl.add_row("Merton asset value (€m)", f"{merton.asset_value:,.2f}")
    tbl.add_row("Merton asset σ_V", f"{merton.asset_volatility:.2%}")
    tbl.add_row("Distance to default", f"{merton.distance_to_default:+.3f}")
    tbl.add_row("Merton PD (theoretical)",
                f"{merton.probability_of_default:.4%}")
    tbl.add_row("KMV PD (empirical)", f"{kmv_pd:.4%}")
    tbl.add_row("PD used for ECL (max)", f"{pd_for_ecl:.4%}")
    tbl.add_row("", "")
    tbl.add_row("IFRS 9 stage",
                f"{ecl.stage.value.upper()} — "
                f"{'; '.join(ecl.notes) if ecl.notes else 'no notes'}")
    tbl.add_row("12-month ECL (€m)", f"{ecl.ecl_12_month_eur_m:.4f}")
    tbl.add_row("Lifetime ECL (€m)", f"{ecl.ecl_lifetime_eur_m:.4f}")
    tbl.add_row("Reported ECL (€m)", f"{ecl.ecl_eur_m:.4f}")
    tbl.add_row("Implied provision rate", f"{ecl.implied_rate_pct:.2%}")
    console.print(tbl)

    if json_out is not None:
        import json
        payload = {
            "counterparty": cp,
            "inputs": {
                "equity_value": equity_value,
                "equity_volatility": equity_volatility,
                "debt_face_value": debt_face_value,
                "risk_free_rate": risk_free_rate,
                "horizon_years": horizon_years,
                "lgd": lgd, "eir": eir, "maturity": maturity,
                "days_past_due": days_past_due,
            },
            "merton": {
                "asset_value": merton.asset_value,
                "asset_volatility": merton.asset_volatility,
                "distance_to_default": merton.distance_to_default,
                "pd": merton.probability_of_default,
                "converged": merton.converged,
                "iterations": merton.iterations,
            },
            "kmv_pd": kmv_pd,
            "ecl": {
                "stage": ecl.stage.value,
                "ecl_12_month_eur_m": ecl.ecl_12_month_eur_m,
                "ecl_lifetime_eur_m": ecl.ecl_lifetime_eur_m,
                "ecl_eur_m": ecl.ecl_eur_m,
                "implied_rate": ecl.implied_rate_pct,
                "notes": ecl.notes,
            },
        }
        json_out.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        console.print(f"[green]JSON:[/green] {json_out}")


@main.command("reverse")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("-o", "--report", "report_path", type=click.Path(path_type=Path),
              default=None, help="Output REVERSE_REPORT.md path.")
@click.option("--spec-out", "spec_path", type=click.Path(path_type=Path),
              default=None, help="Output partial spec YAML path.")
def reverse_cmd(xlsx_path: Path, report_path: Path | None,
                spec_path: Path | None) -> None:
    """Reverse-engineer a legacy Excel model into a ModelForge spec skeleton.

    Classifies sheets, extracts inputs, clusters formulas, suggests the
    closest ModelForge template type. Produces a markdown report and
    optional partial YAML spec.
    """
    from modelforge.reverse import (
        analyze_workbook, render_markdown, render_spec_skeleton,
    )
    rep = analyze_workbook(xlsx_path)

    md = render_markdown(rep)
    if report_path is None:
        report_path = xlsx_path.with_name(xlsx_path.stem + ".REVERSE_REPORT.md")
    report_path.write_text(md, encoding="utf-8")
    console.print(f"[green]Report:[/green] {report_path}")
    console.print(f"[bold]Detected template:[/bold] `{rep.detected_template}` "
                  f"(confidence {rep.template_confidence:.0%})")
    console.print(f"[bold]Inputs extracted:[/bold] {rep.n_inputs}")

    if spec_path is not None:
        spec_yaml = render_spec_skeleton(rep)
        spec_path.write_text(spec_yaml, encoding="utf-8")
        console.print(f"[green]Spec skeleton:[/green] {spec_path}")
    else:
        console.print("[dim]Pass --spec-out <path.yaml> to also write a partial "
                      "YAML spec skeleton.[/dim]")


@main.group("feeds")
def feeds_group() -> None:
    """Market data feeds — EURIBOR / ECB rates, Damodaran ERPs."""


@feeds_group.command("list")
def feeds_list_cmd() -> None:
    """Print current bundled/cached feed values."""
    from modelforge.feeds import ECBFeed, DamodaranFeed
    ecb = ECBFeed.load()
    tbl = Table(title=f"ECB rates (as of {ecb.as_of})")
    tbl.add_column("Metric", style="bold")
    tbl.add_column("Value")
    for k, v in ecb.as_rows():
        tbl.add_row(k, f"{v:.4%}")
    console.print(tbl)

    dam = DamodaranFeed.load()
    tbl2 = Table(title=f"Damodaran country risk (as of {dam.snapshot.data.get('as_of', '?')})")
    tbl2.add_column("ISO", style="bold")
    tbl2.add_column("CRP")
    tbl2.add_column("Total ERP")
    tbl2.add_column("Rating")
    for iso, crp, erp, rating in dam.as_rows():
        tbl2.add_row(iso, f"{crp:.4%}", f"{erp:.4%}", rating)
    console.print(tbl2)


@feeds_group.command("refresh")
@click.option("--timeout", default=10.0, type=float,
              help="Network timeout per request (seconds).")
def feeds_refresh_cmd(timeout: float) -> None:
    """Pull the latest ECB observations from the SDW live API."""
    from modelforge.feeds import ECBFeed, DamodaranFeed
    console.print("[bold]Refreshing ECB feed from SDW...[/bold]")
    ecb = ECBFeed.load().refresh(timeout=timeout)
    console.print(f"[green]ECB refreshed[/green]  as_of={ecb.as_of}")
    # Damodaran is annual — refresh() is a no-op that keeps bundled data
    DamodaranFeed.load()
    console.print("[dim]Damodaran snapshot bundled (annual cadence; manual update per January release).[/dim]")


@main.command("diff")
@click.argument("v1_xlsx", type=click.Path(exists=True, path_type=Path))
@click.argument("v2_xlsx", type=click.Path(exists=True, path_type=Path))
@click.option("--format", "fmt", type=click.Choice(["md", "html", "both"]),
              default="md", show_default=True)
@click.option("-o", "--output", "out_path", type=click.Path(path_type=Path),
              default=None, help="Output file (md, html, or stem for 'both').")
def diff_cmd(v1_xlsx: Path, v2_xlsx: Path, fmt: str, out_path: Path | None) -> None:
    """Git-style structured diff between two built workbooks.

    Diff dimensions: assumptions (A-id keyed), sources (S-id),
    formulas, structural (sheets/named ranges), reproducibility
    metadata. Clean diff exits 0; any change exits 0 still (not a
    failure condition — diff is informational).
    """
    from modelforge.diff import compute_diff, render_markdown, render_html
    res = compute_diff(v1_xlsx, v2_xlsx)
    md = render_markdown(res)
    hh = render_html(res)

    if out_path is None:
        # Print markdown to stdout
        console.print(md)
        return

    if fmt == "md" or fmt == "both":
        md_path = out_path if fmt == "md" else out_path.with_suffix(".md")
        md_path.write_text(md, encoding="utf-8")
        console.print(f"[green]Markdown:[/green] {md_path}")
    if fmt == "html" or fmt == "both":
        html_path = out_path if fmt == "html" else out_path.with_suffix(".html")
        html_path.write_text(hh, encoding="utf-8")
        console.print(f"[green]HTML:[/green] {html_path}")


@main.command("backtest")
@click.argument("predicted_csv", type=click.Path(exists=True, path_type=Path))
@click.argument("realized_csv", type=click.Path(exists=True, path_type=Path))
@click.option("--n-groups", default=10, show_default=True, type=int,
              help="Number of PD deciles for the H-L grouping.")
def backtest_cmd(predicted_csv: Path, realized_csv: Path,
                 n_groups: int) -> None:
    """Hosmer-Lemeshow χ² backtest on a bundled PD curve.

    Expected CSV format (no header):
        predicted_csv : one PD per line (decimal, e.g. 0.023)
        realized_csv  : one default flag per line (0 or 1)

    Both files must have identical row counts. Exit 0 if calibration
    passes (p > 0.05), 1 otherwise.
    """
    from modelforge.risk import hosmer_lemeshow
    preds = [float(l.strip()) for l in predicted_csv.read_text().splitlines()
             if l.strip()]
    reals = [int(l.strip()) for l in realized_csv.read_text().splitlines()
             if l.strip()]
    if len(preds) != len(reals):
        console.print(f"[red]Length mismatch: {len(preds)} vs {len(reals)}[/red]")
        sys.exit(2)
    chi2, p = hosmer_lemeshow(preds, reals, n_groups=n_groups)

    tbl = Table(title=f"Hosmer-Lemeshow backtest ({len(preds)} exposures, "
                       f"{n_groups} groups)")
    tbl.add_column("Statistic", style="bold")
    tbl.add_column("Value")
    tbl.add_row("chi2", f"{chi2:.3f}")
    tbl.add_row("p-value", f"{p:.4f}")
    tbl.add_row("Predicted total defaults", f"{sum(preds):.2f}")
    tbl.add_row("Realised defaults", f"{sum(reals)}")
    tbl.add_row("Calibration",
                "[green]PASS[/green] (p > 0.05)" if p > 0.05
                else "[red]FAIL[/red] (p <= 0.05)")
    console.print(tbl)
    sys.exit(0 if p > 0.05 else 1)


@main.command("chat")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("--backend", type=click.Choice(["api", "dry"]), default="api",
              help="`api` uses Anthropic SDK (requires ANTHROPIC_API_KEY); "
                   "`dry` prints the prompt without calling.")
@click.option("--model", default="claude-opus-4-7", show_default=True)
@click.option("--export", "export_path", type=click.Path(path_type=Path),
              default=None, help="Write conversation to markdown on exit.")
def chat_cmd(xlsx_path: Path, backend: str, model: str,
             export_path: Path | None) -> None:
    """Lineage Q&A REPL against a built workbook.

    Answers cite Assumption IDs (A-###) and Source IDs (S-###) and can
    walk the linkage graph to explain any cell's provenance. Type
    'exit' or 'quit' to leave; 'history' to print all turns so far.
    """
    from modelforge.chat import ChatSession
    console.print(f"[bold]ModelForge chat[/bold]  {xlsx_path.name}  "
                  f"[dim]({backend} backend · {model})[/dim]")
    console.print("[dim]Type your question, or 'exit' / 'quit' to leave. "
                  "'history' prints the conversation so far.[/dim]\n")
    session = ChatSession(xlsx_path=xlsx_path, model=model,
                          backend=backend)  # type: ignore[arg-type]
    try:
        while True:
            try:
                q = input("You › ").strip()
            except (EOFError, KeyboardInterrupt):
                console.print()
                break
            if not q:
                continue
            if q.lower() in ("exit", "quit"):
                break
            if q.lower() == "history":
                console.print(session.to_markdown())
                continue
            try:
                reply = session.ask(q)
            except RuntimeError as e:
                console.print(f"[red]{e}[/red]")
                continue
            console.print(f"[green]ModelForge ›[/green] {reply}\n")
    finally:
        if export_path is not None:
            export_path.write_text(session.to_markdown(), encoding="utf-8")
            console.print(f"[green]Conversation exported to[/green] {export_path}")


@main.command("dossier")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("-o", "--output", "output_pdf", type=click.Path(path_type=Path),
              default=None, help="Output PDF path. Defaults to <xlsx>.dossier.pdf.")
def dossier_cmd(xlsx_path: Path, output_pdf: Path | None) -> None:
    """Generate a regulator-grade audit dossier PDF for a built workbook.

    Includes: cover + metadata, executive summary, assumptions register,
    source registry, formula inventory, lineage graph summary, QC
    sign-off with signature block, bilingual glossary.
    """
    from modelforge.dossier import generate_dossier

    pdf_path = generate_dossier(xlsx_path, output_pdf)
    console.print(f"[green]Dossier:[/green] {pdf_path}")


@main.command("stats")
@click.argument("graph_db", type=click.Path(exists=True, path_type=Path))
def stats_cmd(graph_db: Path) -> None:
    """Print graph statistics."""
    store = GraphStore(graph_db)
    import sqlite3
    with sqlite3.connect(graph_db) as conn:
        models = [r[0] for r in conn.execute("SELECT DISTINCT model_id FROM nodes").fetchall()]
    for m in models:
        stats = store.stats(m)
        tbl = Table(title=f"Graph stats — {m}")
        tbl.add_column("Bucket")
        tbl.add_column("Count")
        for k, v in sorted(stats.items()):
            tbl.add_row(k, str(v))
        console.print(tbl)


@main.command("moat")
@click.argument("xlsx_path", type=click.Path(exists=True, path_type=Path))
@click.option("--strict", is_flag=True, default=False,
              help="Exit non-zero if any moat gate fails.")
@click.option("--no-inject", is_flag=True, default=False,
              help="Don't inject the MOAT sheet into the workbook.")
def moat_cmd(xlsx_path: Path, strict: bool, no_inject: bool) -> None:
    """Verify "fully-formulated live Excel outputs" moat gates.

    Runs four gates and prints the verdict:
      • formula_density   — output sheets ≥ 90% formulas
      • reference_graph   — output formulas have no magic-number literals
      • no_orphan_inputs  — every named range is referenced by a formula
      • recalculation     — third-party engine recomputes identical values
    """
    from modelforge.moat import MoatGate, inject_moat_sheet
    report = MoatGate().evaluate(xlsx_path)
    if not no_inject:
        inject_moat_sheet(xlsx_path, report)

    tbl = Table(title=f"MOAT — {xlsx_path.name}")
    tbl.add_column("Gate")
    tbl.add_column("Verdict")
    tbl.add_column("Metric", justify="right")
    tbl.add_column("Detail")
    for g in report.gate_results:
        verdict = "[green]PASS[/green]" if g.passed else "[red]FAIL[/red]"
        metric = (f"{g.metric:.4f}" if isinstance(g.metric, float) else
                  (str(g.metric) if g.metric is not None else ""))
        tbl.add_row(g.name, verdict, metric, g.detail[:80])
    console.print(tbl)
    console.print(
        f"Core-output formula density: [bold]{report.core_output_density()*100:.1f}%[/bold]"
    )
    if strict and not report.passes():
        sys.exit(4)


@main.command("audit-all")
@click.option("--examples-dir", "examples_dir", type=click.Path(path_type=Path),
              default=Path("examples"),
              help="Directory containing YAML specs to build.")
@click.option("--out-dir", "out_dir", type=click.Path(path_type=Path),
              default=Path("output/audit"),
              help="Directory to write the built workbooks.")
@click.option("--report", "report_path", type=click.Path(path_type=Path),
              default=Path("AUDIT_REPORT.md"),
              help="Markdown report path (default: AUDIT_REPORT.md).")
def audit_all_cmd(examples_dir: Path, out_dir: Path, report_path: Path) -> None:
    """Build every example, run Trust + Moat gates, write a markdown report.

    The audit report is the evidence pack for design partners — proves
    that every demo template builds, recalculates, and meets the moat
    + plausibility gates.
    """
    from modelforge.moat import MoatGate
    from modelforge.trust import DEFAULT_RULES, TrustEngine
    out_dir.mkdir(parents=True, exist_ok=True)

    specs = sorted(examples_dir.glob("*.yaml"))
    if not specs:
        console.print(f"[yellow]No specs found in {examples_dir}[/yellow]")
        sys.exit(2)

    console.print(f"Auditing {len(specs)} specs from {examples_dir}…")
    rows = []
    moat_engine = MoatGate()
    trust_engine = TrustEngine(rules=DEFAULT_RULES)

    for spec_path in specs:
        try:
            spec_bytes = spec_path.read_bytes()
            raw = yaml.safe_load(spec_bytes)
            mt = raw.get("model_type", "unitranche")
            SpecClass = _load_spec_class(mt)
            spec = SpecClass.model_validate(raw)
            xlsx_out = out_dir / f"{spec_path.stem}.xlsx"
            xlsx, _ = build_model(spec, xlsx_out,
                                  spec_source_bytes=spec_bytes,
                                  spec_source_path=spec_path)
            tr = trust_engine.evaluate(xlsx, spec)
            mr = moat_engine.evaluate(xlsx)
            rows.append({
                "spec": spec_path.name,
                "template": mt,
                "trust_fail": len(tr.by_severity("fail")),
                "trust_warn": len(tr.by_severity("warn")),
                "moat_passes": mr.passes(),
                "core_output_density": mr.core_output_density(),
                "moat_failed_gates": [g.name for g in mr.gate_results if not g.passed],
                "error": None,
            })
        except Exception as e:
            rows.append({
                "spec": spec_path.name, "template": "?",
                "trust_fail": 0, "trust_warn": 0,
                "moat_passes": False,
                "core_output_density": 0.0,
                "moat_failed_gates": ["build_failed"],
                "error": str(e)[:120],
            })

    # Write report
    lines = ["# ModelForge Audit Report",
             f"\n_Generated by `modelforge audit-all` against `{examples_dir}` — "
             f"{len(rows)} specs._\n",
             "## Summary\n"]
    n_pass_trust = sum(1 for r in rows if r["trust_fail"] == 0 and r["error"] is None)
    n_pass_moat = sum(1 for r in rows if r["moat_passes"] and r["error"] is None)
    n_built = sum(1 for r in rows if r["error"] is None)
    lines.append(
        f"- Built: **{n_built}/{len(rows)}** templates compile end-to-end\n"
        f"- Trust Layer FAIL-clean: **{n_pass_trust}/{len(rows)}**\n"
        f"- Moat gates PASS: **{n_pass_moat}/{len(rows)}**\n"
        f"- Avg core-output formula density: "
        f"**{(sum(r['core_output_density'] for r in rows)/len(rows))*100:.1f}%**\n"
    )
    lines.append("\n## Per-spec\n")
    lines.append("| Spec | Template | Built | Trust FAIL | Trust WARN | "
                 "Moat | Core density | Failed gates / error |")
    lines.append("|---|---|:-:|:-:|:-:|:-:|:-:|---|")
    for r in rows:
        built = "✓" if r["error"] is None else "✗"
        moat = "✓" if r["moat_passes"] else "✗"
        density = (f"{r['core_output_density']*100:.1f}%"
                   if r["error"] is None else "—")
        gates = (r["error"] if r["error"]
                 else (", ".join(r["moat_failed_gates"]) or "—"))
        lines.append(
            f"| {r['spec']} | {r['template']} | {built} | "
            f"{r['trust_fail']} | {r['trust_warn']} | {moat} | "
            f"{density} | {gates} |"
        )
    report_path.write_text("\n".join(lines), encoding="utf-8")
    console.print(f"[green]Wrote[/green] {report_path}")
    console.print(f"  Built: {n_built}/{len(rows)}  Trust-clean: {n_pass_trust}  "
                  f"Moat-PASS: {n_pass_moat}")


@main.command("deck")
@click.argument("target", type=click.Path(exists=True, path_type=Path))
@click.option("--type", "deck_type", type=click.Choice(["ic_memo", "teaser"]),
              default="ic_memo", show_default=True,
              help="Deck composer to use.")
@click.option("--theme", default=None,
              help="Deck theme (e.g. finance-pro, executive-dark). "
                   "Defaults to the composer's theme.")
@click.option("--out", "out_path", type=click.Path(path_type=Path), default=None,
              help="Output .pptx path. Defaults to "
                   "<workbook_stem>_<type>.pptx next to the workbook.")
def deck_cmd(target: Path, deck_type: str, theme: str | None,
             out_path: Path | None) -> None:
    """Render a certified board deck (.pptx) from a spec or built workbook.

    TARGET may be a ``.yaml``/``.yml`` spec or a built ``.xlsx``. A spec is
    built exactly as the ``build`` command ships it (Trust/Moat sheets +
    deterministic finishing + manifest) before the deck chain runs. A
    workbook requires its ``<stem>.manifest.json`` sidecar next to it.

    Fail-closed: the deck is only rendered when the manifest verifies AND
    the certification audit returns CERTIFIED. Every deck ends with a
    mandatory "Certification & Red Flags" slide and embeds
    spec_sha256 + workbook_sha256 in the .pptx core properties; the file is
    byte-deterministic for a given workbook (same workbook in →
    byte-identical deck out).
    """
    try:
        from modelforge.deck.pipeline import DeckAdapterError, build_deck_from_workbook
    except ImportError as e:
        console.print(
            "[red]deck rendering requires the deck extras: "
            "pip install 'modelforge-finance\\[deck]'[/red]"
        )
        console.print(f"[dim](missing dependency: {getattr(e, 'name', None) or e})[/dim]")
        sys.exit(1)

    suffix = target.suffix.lower()
    if suffix in (".yaml", ".yml"):
        spec, model_type, spec_bytes = _load_and_validate_spec(target)
        xlsx_out = Path("output") / f"{target.stem}.xlsx"
        xlsx, _ = build_model(
            spec, xlsx_out,
            spec_source_bytes=spec_bytes,
            spec_source_path=target,
        )
        # Ship the same artifact `build` delivers (RedFlags + MOAT +
        # styler/determinism/manifest) so the deck certifies the real file.
        _inject_trust_moat_and_finish(xlsx, spec, spec_bytes, target, quiet=True)
        console.print(f"[green]Built:[/green] {xlsx}  [dim]({model_type})[/dim]")
        workbook = Path(xlsx)
    elif suffix == ".xlsx":
        workbook = target
    else:
        console.print(
            f"[red]{target.name}: expected a .yaml/.yml spec or a built "
            f".xlsx workbook.[/red]"
        )
        sys.exit(1)

    try:
        result = build_deck_from_workbook(
            workbook, deck_type=deck_type, theme=theme, out_path=out_path,
        )
    except DeckAdapterError as e:
        from rich.markup import escape as _rich_escape

        # escape: adapter errors may contain literal brackets (e.g. the
        # "pip install 'modelforge-finance[deck]'" hint) that rich would
        # otherwise swallow as markup tags.
        console.print(f"[red]{_rich_escape(str(e))}[/red]")
        sys.exit(1)
    except ImportError as e:
        console.print(
            "[red]deck rendering requires the deck extras: "
            "pip install 'modelforge-finance\\[deck]'[/red]"
        )
        console.print(f"[dim](missing dependency: {getattr(e, 'name', None) or e})[/dim]")
        sys.exit(1)

    console.print(
        f"[green]Deck:[/green] {result.pptx_path}  "
        f"[dim]({result.deck_type} · {result.slide_count} slides · "
        f"theme {result.theme})[/dim]"
    )
    console.print(
        f"[green]Certified:[/green] {result.audit_verdict}  "
        f"[dim]workbook {result.workbook_sha256[:16]}…  "
        f"spec {result.spec_sha256[:16]}…  (embedded in core properties)[/dim]"
    )
    if result.red_flag_count:
        console.print(
            f"[yellow]Red flags:[/yellow] {result.red_flag_count} Trust-Layer "
            f"entr{'y' if result.red_flag_count == 1 else 'ies'} carried onto "
            f"the certification slide."
        )


def _parse_kv_pairs(pairs: tuple[str, ...], *, what: str) -> dict[str, str]:
    """Parse repeatable ``key=value`` CLI options into a dict.

    Friendly-exits (code 2) on a token missing ``=`` or with an empty key, so a
    fat-fingered ``--filter sectorindustrials`` is a clear message rather than a
    silent no-op or a traceback.
    """
    out: dict[str, str] = {}
    for raw in pairs:
        if "=" not in raw:
            console.print(
                f"[red]--{what} expects key=value, got {raw!r} (no '='). "
                f"Example: --{what} "
                f"{'sector=industrials' if what == 'filter' else 'irr_base=1.0'}.[/red]"
            )
            sys.exit(2)
        key, value = raw.split("=", 1)
        key = key.strip()
        if not key:
            console.print(f"[red]--{what} has an empty key in {raw!r}.[/red]")
            sys.exit(2)
        out[key] = value.strip()
    return out


def _coerce_filter_value(value: str):
    """Coerce a filter value string to bool/number/list when it clearly is one.

    Keeps screening filters ergonomic on the command line: ``deal_size_eur_m_min``
    compares numerically, ``geography_in`` accepts a comma list, and an
    obviously-numeric equality (``vintage=2026``) matches an int spec value. Falls
    back to the raw string so ``sector=industrials`` stays an exact text match.
    """
    v = value.strip()
    if "," in v:
        return [item.strip() for item in v.split(",") if item.strip()]
    low = v.lower()
    if low in ("true", "false"):
        return low == "true"
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        pass
    return v


@main.command("screen")
@click.argument("spec_dir", type=click.Path(exists=False, path_type=Path))
@click.option("--filter", "filters", multiple=True, metavar="KEY=VAL",
              help="Repeatable filter predicate. Suffix the key with _min (>=), "
                   "_max (<=), or _in (comma-list membership); a bare key is an "
                   "exact match. E.g. --filter ebitda_margin_min=0.20 "
                   "--filter sector=industrials.")
@click.option("--rank", "ranks", multiple=True, metavar="METRIC=WEIGHT",
              help="Repeatable ranking weight (float). Positive = higher-is-"
                   "better, negative = lower-is-better. E.g. --rank irr_base=1.0 "
                   "--rank leverage_x=-0.3.")
@click.option("--top", "top_n", type=int, default=25, show_default=True,
              help="Maximum number of ranked deals to show.")
@click.option("--glob", "glob_pattern", default="**/*.yaml", show_default=True,
              help="Glob used to discover spec files under SPEC_DIR.")
def screen_cmd(spec_dir: Path, filters: tuple[str, ...],
               ranks: tuple[str, ...], top_n: int,
               glob_pattern: str) -> None:
    """Filter + rank a directory of deal spec YAMLs — without building each one.

    Triage many candidates at the spec layer (screening 1,000 deals never builds
    1,000 workbooks). A spec is screenable when it carries an optional top-level
    ``screening:`` block, e.g.::

        screening:
          sector: "US SaaS"
          geography: "US"
          deal_size_eur_m: 75
          vintage: 2026
          irr_base: 0.24
          leverage_x: 4.1

    Filter keys support ``_min`` (>=), ``_max`` (<=), ``_in`` (comma-list
    membership) or a bare key (equality). Rank weights are floats; a negative
    weight means lower-is-better. Prints a ranked table and exits 0.
    """
    from modelforge.screening import screen

    filter_raw = _parse_kv_pairs(filters, what="filter")
    parsed_filters = {k: _coerce_filter_value(v) for k, v in filter_raw.items()}

    rank_raw = _parse_kv_pairs(ranks, what="rank")
    parsed_ranks: dict[str, float] = {}
    for metric, weight in rank_raw.items():
        try:
            parsed_ranks[metric] = float(weight)
        except ValueError:
            console.print(
                f"[red]--rank weight for {metric!r} must be a number, "
                f"got {weight!r}.[/red]"
            )
            sys.exit(2)

    if not spec_dir.exists():
        console.print(
            f"[red]{spec_dir} does not exist.[/red] Point `screen` at a folder "
            f"of deal spec YAMLs."
        )
        sys.exit(2)
    if not spec_dir.is_dir():
        console.print(
            f"[red]{spec_dir} is not a directory.[/red] `screen` takes a folder "
            f"of spec YAMLs (use `certify`/`build` for a single spec)."
        )
        sys.exit(2)

    results = screen(
        spec_dir,
        filters=parsed_filters,
        rank_by=parsed_ranks,
        top_n=top_n,
        glob_pattern=glob_pattern,
    )

    if not results:
        console.print(
            f"[yellow]No screenable deals matched in {spec_dir}.[/yellow]\n"
            f"[dim]A spec is screenable only when it carries an optional "
            f"top-level `screening:` block (keys like sector / geography / "
            f"deal_size_eur_m / vintage / irr_base). Either no spec under "
            f"`{glob_pattern}` has one, or your filters excluded them all.[/dim]"
        )
        sys.exit(0)

    # Build a compact ranked table. Beyond rank/deal_id/score, surface the most
    # useful summary fields when present (in a stable, finance-readable order),
    # so the table reads like a screening grid rather than a JSON dump.
    preferred = ["sector", "geography", "deal_size_eur_m",
                 "vintage", "irr_base", "leverage_x"]
    present = [f for f in preferred if any(f in r.summary for r in results)]

    tbl = Table(title=f"Deal screen — {spec_dir} "
                       f"({len(results)} of "
                       f"{len(results) if len(results) < top_n else f'top {top_n}'} "
                       f"shown)")
    tbl.add_column("#", justify="right", style="bold")
    tbl.add_column("Deal", style="bold")
    tbl.add_column("Score", justify="right")
    _labels = {
        "deal_size_eur_m": "Size (m)",
        "irr_base": "IRR base",
        "leverage_x": "Leverage",
    }
    for f in present:
        justify = "right" if f not in ("sector", "geography") else "left"
        tbl.add_column(_labels.get(f, f.replace("_", " ").title()),
                       justify=justify)

    def _fmt(field: str, val) -> str:
        if val is None:
            return "—"
        if field == "irr_base":
            try:
                return f"{float(val):.1%}"
            except (TypeError, ValueError):
                return str(val)
        if field in ("deal_size_eur_m", "leverage_x"):
            try:
                return f"{float(val):,.1f}"
            except (TypeError, ValueError):
                return str(val)
        return str(val)

    for i, r in enumerate(results, start=1):
        row = [str(i), r.deal_id, f"{r.score:.3f}"]
        row.extend(_fmt(f, r.summary.get(f)) for f in present)
        tbl.add_row(*row)
    console.print(tbl)

    if parsed_ranks:
        weights = ", ".join(f"{m}={w:+g}" for m, w in parsed_ranks.items())
        console.print(f"[dim]Ranked by: {weights} "
                      f"(min-max normalized; negative = lower-is-better).[/dim]")
    else:
        console.print("[dim]No --rank given — all matches scored 0.0 "
                      "(filter-only screen).[/dim]")
    sys.exit(0)


if __name__ == "__main__":
    main()
