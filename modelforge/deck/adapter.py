"""Certified workbook → deck Facts adapter (the deck pipeline's trust gate).

Turns a **built, certified** ModelForge workbook into the typed Facts models
the deck composers consume (``DealFacts`` for the IC memo, ``TeaserFacts``
for the teaser), with full lineage: every numeric fact carries the
``Sheet!Cell`` reference it was read from.

FAIL-CLOSED contract (in order):

1. The workbook's manifest sidecar (``<stem>.manifest.json``) must exist and
   ``verify_manifest`` must pass — a missing or stale manifest means the
   bytes on disk are not the bytes that were built, so no deck is produced.
2. The QC workbook audit (``modelforge.qc.workbook_audit``) must return
   verdict ``CERTIFIED`` (zero formula-error cells AND zero styling gaps).
   ``WARN``/``FAIL`` workbooks are refused.
3. Only templates with a verified field mapping are adapted. Anything else
   raises a friendly "not deck-mappable yet" error naming the supported set.

Numeric facts that live in formula cells are recomputed with the same
third-party ``formulas`` engine the certification audit uses — the adapter
never trusts a cached value it cannot recompute.

THREAT MODEL (v1): the manifest sidecar is *self-attested* SHA-256 —
``verify_manifest`` protects against drift and accidental edits (a workbook
whose bytes changed after build no longer matches its recorded hashes), NOT
against an adversary who can rewrite the sidecar alongside the workbook.
There is no signature or HMAC in v1; treat the manifest as an integrity
checksum, not an authenticity proof.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook

__all__ = [
    "CellFact",
    "DeckAdapterError",
    "SUPPORTED_DECK_TEMPLATES",
    "WorkbookFacts",
    "adapt_workbook",
    "detect_template",
]


SUPPORTED_DECK_TEMPLATES: tuple[str, ...] = (
    "sponsor_lbo", "real_estate", "development_re",
)


class DeckAdapterError(Exception):
    """Friendly, fail-closed error raised anywhere along the deck chain."""


# ─────────────────────────────────────────────────────────────────────────────
# Data shapes
# ─────────────────────────────────────────────────────────────────────────────


@dataclass
class CellFact:
    """A numeric fact plus the workbook cell(s) it was read from."""

    value: Any
    ref: str  # "Sheet!D10" (or a composite like "SourcesUses!D75 × OperatingModel!K13")

    def __float__(self) -> float:  # convenience
        return float(self.value)


@dataclass
class WorkbookFacts:
    """Everything the deck pipeline needs, extracted from ONE certified workbook."""

    workbook: Path
    template: str
    manifest: Any  # analytics.manifest.BuildManifest
    audit_verdict: str
    audit_summary: dict[str, Any]
    red_flags: list[dict[str, str]] = field(default_factory=list)
    facts: dict[str, CellFact] = field(default_factory=dict)

    @property
    def source_refs(self) -> dict[str, str]:
        """fact name → sheet!cell lineage map (for stamping/footnotes)."""
        return {k: v.ref for k, v in self.facts.items()}

    # Composed lazily so the adapter has no import-time dependency on compose.
    def deal_facts(self):
        """Build the ``DealFacts`` model (IC-memo composer input)."""
        return _to_deal_facts(self)

    def teaser_facts(self):
        """Build the ``TeaserFacts`` model (teaser composer input)."""
        return _to_teaser_facts(self)


# ─────────────────────────────────────────────────────────────────────────────
# Formula-engine evaluation (cached per workbook content hash)
# ─────────────────────────────────────────────────────────────────────────────

_ENGINE_CACHE: dict[str, dict[tuple[str, str], Any]] = {}
_AUDIT_CACHE: dict[str, Any] = {}
_CACHE_MAX = 4


def _sha256_file(path: Path) -> str:
    import hashlib

    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def _scalarize(raw: Any) -> Any:
    """Collapse a `formulas` engine value (possibly a numpy array) to a scalar."""
    if raw is None:
        return None
    flatten = getattr(raw, "flatten", None)
    if flatten is not None:
        try:
            flat = raw.flatten()
            return _scalarize(flat[0]) if len(flat) else None
        except Exception:
            return None
    return raw


def _evaluate_workbook(xlsx_path: Path, wb_sha: str) -> dict[tuple[str, str], Any]:
    """Recompute every formula with the `formulas` engine.

    Returns {(SHEET_UPPER, CELL): value}. Cached on workbook content hash so
    repeated deck builds from the same bytes do not pay the engine twice.
    """
    cached = _ENGINE_CACHE.get(wb_sha)
    if cached is not None:
        return cached
    try:
        import formulas  # type: ignore
    except ImportError as e:  # pragma: no cover - dep ships with modelforge
        raise DeckAdapterError(
            "The deck adapter needs the `formulas` package to recompute "
            "workbook values (pip install formulas)."
        ) from e

    from modelforge.qc.workbook_audit import _split_engine_key

    try:
        xl = formulas.ExcelModel().loads(str(xlsx_path)).finish()
        sol = xl.calculate()
    except Exception as e:
        raise DeckAdapterError(
            f"Formula engine could not recompute {xlsx_path.name}: {e}"
        ) from e

    out: dict[tuple[str, str], Any] = {}
    for key, val in sol.items():
        sheet, cell = _split_engine_key(key)
        if sheet is None or cell is None:
            continue
        raw = val.value if hasattr(val, "value") else val
        out[(sheet.upper(), cell.upper())] = _scalarize(raw)

    if len(_ENGINE_CACHE) >= _CACHE_MAX:
        _ENGINE_CACHE.pop(next(iter(_ENGINE_CACHE)))
    _ENGINE_CACHE[wb_sha] = out
    return out


# ─────────────────────────────────────────────────────────────────────────────
# Workbook reading helpers
# ─────────────────────────────────────────────────────────────────────────────


def _find_label_rows(ws, label: str, col: int = 1) -> list[int]:
    """All row numbers whose column-`col` text equals `label` (stripped)."""
    target = label.strip()
    rows: list[int] = []
    for row in ws.iter_rows(min_col=col, max_col=col):
        c = row[0]
        if isinstance(c.value, str) and c.value.strip() == target:
            rows.append(c.row)
    return rows


def _find_label_row(ws, label: str, col: int = 1) -> Optional[int]:
    rows = _find_label_rows(ws, label, col=col)
    return rows[0] if rows else None


class _Reader:
    """Bound reader over (workbook, engine solution) producing CellFacts."""

    def __init__(self, wb, solution: dict[tuple[str, str], Any]):
        self.wb = wb
        self.solution = solution

    def cell_fact(self, sheet: str, coord: str) -> CellFact:
        """Value of sheet!coord — raw input cells read directly, formula
        cells via the recompute engine (never the cached value)."""
        ws = self.wb[sheet]
        coord = coord.replace("$", "").upper()
        raw = ws[coord].value
        ref = f"{sheet}!{coord}"
        if isinstance(raw, str) and raw.startswith("="):
            val = self.solution.get((sheet.upper(), coord))
            if val is None:
                raise DeckAdapterError(
                    f"Could not recompute formula cell {ref} — the deck "
                    f"adapter refuses to guess. Re-run `modelforge certify` "
                    f"on the workbook."
                )
            return CellFact(value=val, ref=ref)
        return CellFact(value=raw, ref=ref)

    def labeled_fact(self, sheet: str, label: str, value_col: str = "D",
                     occurrence: int = 0) -> CellFact:
        ws = self.wb[sheet]
        rows = _find_label_rows(ws, label)
        if not rows or occurrence >= len(rows):
            raise DeckAdapterError(
                f"Could not locate the labeled row {label!r} on sheet "
                f"{sheet!r} — the workbook layout does not match the "
                f"{'sponsor_lbo'} template this adapter was verified against."
            )
        return self.cell_fact(sheet, f"{value_col}{rows[occurrence]}")

    def named_fact(self, name: str) -> CellFact:
        dn = self.wb.defined_names.get(name) if hasattr(self.wb.defined_names, "get") else None
        if dn is None and name in self.wb.defined_names:
            dn = self.wb.defined_names[name]
        if dn is None:
            raise DeckAdapterError(
                f"Named range {name!r} not found in the workbook."
            )
        m = re.match(r"'?([^'!]+)'?!\$?([A-Z]+)\$?(\d+)", dn.attr_text or "")
        if not m:
            raise DeckAdapterError(
                f"Named range {name!r} has an unsupported reference "
                f"({dn.attr_text!r})."
            )
        sheet, col, row = m.group(1), m.group(2), m.group(3)
        return self.cell_fact(sheet, f"{col}{row}")


# ─────────────────────────────────────────────────────────────────────────────
# Template detection
# ─────────────────────────────────────────────────────────────────────────────


def detect_template(wb) -> str:
    """Best-effort model_type detection for a built workbook.

    Primary source: the Trust Layer ``RedFlags`` sheet carries a literal
    ``Template: <model_type>`` line. Fallback: discriminating sheet names.
    """
    if "RedFlags" in wb.sheetnames:
        ws = wb["RedFlags"]
        for row in ws.iter_rows(min_row=1, max_row=6, min_col=1, max_col=1):
            v = row[0].value
            if isinstance(v, str) and v.strip().lower().startswith("template:"):
                return v.split(":", 1)[1].strip()
    sheets = set(wb.sheetnames)
    if "SourcesUses" in sheets:
        return "sponsor_lbo"
    # development_re is the ground-up development template: a DevSchedule sheet
    # (phased capex / S-curve lease-up / IDC roll-forward) paired with a Returns
    # sheet carrying the unlevered+levered IRR/MOIC and the European promote
    # waterfall. The discriminating sheet is "DevSchedule" (unique to this
    # template; the stabilised real_estate template emits DCF+Financing instead).
    if "DevSchedule" in sheets:
        return "development_re"
    # real_estate is the DCF + Financing pair emitted by templates/real_estate.py
    # (NOI build + exit on the DCF sheet, senior mortgage + equity waterfall on
    # the Financing sheet). The legacy "REModel"/"EquityWaterfall" names were
    # never emitted by the builder; the live discriminator is {DCF, Financing}.
    if {"DCF", "Financing"} <= sheets or "REModel" in sheets:
        return "real_estate"
    if "CreditOpinion" in sheets:
        return "unitranche"
    if "DCFValuation" in sheets:
        return "dcf"
    return "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────


def adapt_workbook(
    xlsx_path: Path | str,
    manifest_path: Optional[Path | str] = None,
) -> WorkbookFacts:
    """Certified workbook (+ manifest sidecar) → :class:`WorkbookFacts`.

    Raises :class:`DeckAdapterError` (fail-closed) when the manifest is
    missing/stale, when the certification audit is not CERTIFIED, or when
    the template has no verified deck mapping.
    """
    xlsx = Path(xlsx_path)
    if not xlsx.exists():
        raise DeckAdapterError(f"Workbook not found: {xlsx}")

    # ── 1. Manifest sidecar must exist and verify ─────────────────────────
    from modelforge.analytics.manifest import read_manifest, verify_manifest

    mpath = Path(manifest_path) if manifest_path else xlsx.with_suffix(".manifest.json")
    if not mpath.exists():
        raise DeckAdapterError(
            f"Manifest sidecar not found: {mpath.name} (expected next to "
            f"{xlsx.name}). A certified deck can only be generated from a "
            f"workbook built by `modelforge build`, which writes the "
            f"manifest. Rebuild the workbook, or pass the manifest path "
            f"explicitly."
        )
    verdict = verify_manifest(xlsx, manifest_path=mpath)
    if not verdict.ok:
        issues = "; ".join(verdict.issues) or "hash mismatch"
        raise DeckAdapterError(
            f"Manifest verification FAILED for {xlsx.name}: {issues}. The "
            f"workbook bytes on disk are not the bytes that were built — "
            f"refusing to generate a deck from an unverified workbook."
        )
    manifest = read_manifest(mpath)

    wb_sha = verdict.workbook_actual or _sha256_file(xlsx)

    # ── 2. Certification audit must be CERTIFIED ──────────────────────────
    from modelforge.qc.workbook_audit import audit_workbook

    report = _AUDIT_CACHE.get(wb_sha)
    if report is None:
        report = audit_workbook(xlsx)
        if len(_AUDIT_CACHE) >= _CACHE_MAX:
            _AUDIT_CACHE.pop(next(iter(_AUDIT_CACHE)))
        _AUDIT_CACHE[wb_sha] = report
    if report.verdict != "CERTIFIED":
        first = ", ".join(e.ref for e in report.error_cells[:5])
        gaps = ", ".join(g.ref for g in report.style_gaps[:5])
        detail = (f"formula errors at {first}" if report.error_cells
                  else f"styling gaps at {gaps}")
        raise DeckAdapterError(
            f"Workbook {xlsx.name} audit verdict is {report.verdict}, not "
            f"CERTIFIED ({detail}). The deck pipeline is fail-closed: fix "
            f"the workbook (`modelforge certify {xlsx.name}`) before "
            f"exporting a deck."
        )

    # ── 3. Template must have a verified deck mapping ─────────────────────
    wb = load_workbook(xlsx, data_only=False, keep_links=True)
    template = detect_template(wb)
    if template not in SUPPORTED_DECK_TEMPLATES:
        raise DeckAdapterError(
            f"template {template!r} not deck-mappable yet; supported: "
            f"{list(SUPPORTED_DECK_TEMPLATES)}. Build the deck from a "
            f"supported template, or use `modelforge export` paths for "
            f"other templates."
        )

    # ── 4. Recompute + extract facts ──────────────────────────────────────
    solution = _evaluate_workbook(xlsx, wb_sha)
    reader = _Reader(wb, solution)

    extractors = {
        "sponsor_lbo": _extract_sponsor_lbo,
        "real_estate": _extract_real_estate,
        "development_re": _extract_development_re,
    }
    facts = extractors[template](reader)
    red_flags = _read_red_flags(wb)

    return WorkbookFacts(
        workbook=xlsx,
        template=template,
        manifest=manifest,
        audit_verdict=report.verdict,
        audit_summary=report.summary(),
        red_flags=red_flags,
        facts=facts,
    )


# ─────────────────────────────────────────────────────────────────────────────
# RedFlags sheet reader
# ─────────────────────────────────────────────────────────────────────────────


def _read_red_flags(wb) -> list[dict[str, str]]:
    """Parse Trust-Layer entries from the RedFlags sheet (header at row 7)."""
    if "RedFlags" not in wb.sheetnames:
        return []
    ws = wb["RedFlags"]
    flags: list[dict[str, str]] = []
    for r in range(8, ws.max_row + 1):
        severity = ws.cell(row=r, column=2).value
        rule = ws.cell(row=r, column=3).value
        if severity is None and rule is None:
            continue
        if str(severity).strip().upper() == "ALL CLEAR":
            return []
        flags.append({
            "severity": str(severity or "").strip(),
            "rule": str(rule or "").strip(),
            "cell": str(ws.cell(row=r, column=4).value or "").strip(),
            "message": str(ws.cell(row=r, column=7).value or "").strip(),
        })
    return flags


# ─────────────────────────────────────────────────────────────────────────────
# sponsor_lbo extraction
# ─────────────────────────────────────────────────────────────────────────────

_SU = "SourcesUses"

# Exact col-A labels emitted by builder/sheets/sources_uses.py (v0.8+).
_LBL_SENIOR = "Senior debt (Term Loan B / unitranche)"
_LBL_MEZZ = "Mezzanine / subordinated debt"
_LBL_RCF = "Revolver facility (drawn at close)"
_LBL_SPONSOR_EQ = "Sponsor equity (new money)"
_LBL_ROLLOVER = "Management rollover equity"
_LBL_TOTAL_SOURCES = "Total sources"
_LBL_TOTAL_USES = "Total uses"
_LBL_MA_FEES = "M&A advisory fees (expensed)"
_LBL_FIN_FEES = "Financing fees (capitalized + amortized)"
_LBL_REFI = "Refinance target net debt"
_LBL_MIN_CASH = "Minimum cash to balance sheet"
_LBL_EQ_PP = "Equity purchase price"
_LBL_EXIT_STRATEGIC = "Exit — strategic sale (EV/EBITDA ×)"
_LBL_IRR = "IRR (on cash-flow series)"
_LBL_MOIC = "MoIC (cash-on-cash)"
_LBL_BREACHES = "Total breaches (projection window)"

_HURDLE_RE = re.compile(
    r"=\(D\d+\*(?P<ebitda>'?[A-Za-z0-9_ ]+'?![A-Z]+\d+)"
    r"-(?P<debt>'?[A-Za-z0-9_ ]+'?![A-Z]+\d+|0)\)"
)


def _ref_fact(reader: _Reader, ref: str) -> CellFact:
    """CellFact from a cross-sheet ref string like ``'DebtSchedule'!K60``."""
    m = re.match(r"'?([^'!]+)'?!([A-Z]+\d+)", ref)
    if not m:
        raise DeckAdapterError(f"Unparseable cell reference {ref!r}.")
    return reader.cell_fact(m.group(1), m.group(2))


def _cover_value(reader: _Reader, label: str) -> CellFact:
    return reader.labeled_fact("Cover", label, value_col="C")


def _extract_sponsor_lbo(reader: _Reader) -> dict[str, CellFact]:
    wb = reader.wb
    for required in (_SU, "Cover", "OperatingModel", "Covenants"):
        if required not in wb.sheetnames:
            raise DeckAdapterError(
                f"Sheet {required!r} missing — this workbook does not match "
                f"the sponsor_lbo layout the deck adapter was verified "
                f"against."
            )

    f: dict[str, CellFact] = {}

    # Cover identity block (labels in col A, values in col C)
    f["deal_name"] = _cover_value(reader, "Target")
    f["sector"] = _cover_value(reader, "Sector")
    f["country"] = _cover_value(reader, "Country")
    f["currency"] = _cover_value(reader, "Currency")
    f["project_code"] = _cover_value(reader, "Project code")
    f["analyst"] = _cover_value(reader, "Analyst")
    f["valuation_date"] = _cover_value(reader, "Valuation date")

    # Sources & Uses
    f["senior_debt"] = reader.labeled_fact(_SU, _LBL_SENIOR)
    f["mezz_debt"] = reader.labeled_fact(_SU, _LBL_MEZZ)
    f["rcf_drawn"] = reader.labeled_fact(_SU, _LBL_RCF)
    f["sponsor_equity"] = reader.labeled_fact(_SU, _LBL_SPONSOR_EQ)  # S&U plug
    f["mgmt_rollover"] = reader.labeled_fact(_SU, _LBL_ROLLOVER)
    f["total_sources"] = reader.labeled_fact(_SU, _LBL_TOTAL_SOURCES)
    f["total_uses"] = reader.labeled_fact(_SU, _LBL_TOTAL_USES)
    f["ma_fees"] = reader.labeled_fact(_SU, _LBL_MA_FEES)
    f["financing_fees"] = reader.labeled_fact(_SU, _LBL_FIN_FEES)
    f["refi_net_debt"] = reader.labeled_fact(_SU, _LBL_REFI)
    f["min_cash"] = reader.labeled_fact(_SU, _LBL_MIN_CASH)
    f["equity_purchase_price"] = reader.labeled_fact(_SU, _LBL_EQ_PP)

    # Entry EV — prefer the registered named range
    f["entry_ev"] = reader.named_fact("purchase_price_eur_m")
    f["exit_year"] = reader.named_fact("exit_year_input")
    f["exit_multiple_strategic"] = reader.labeled_fact(_SU, _LBL_EXIT_STRATEGIC)

    # Returns block: IRR/MoIC rows appear in scenario order
    # strategic / IPO / secondary — take occurrence 0 (strategic).
    f["irr_strategic"] = reader.labeled_fact(_SU, _LBL_IRR, occurrence=0)
    f["moic_strategic"] = reader.labeled_fact(_SU, _LBL_MOIC, occurrence=0)
    f["irr_ipo"] = reader.labeled_fact(_SU, _LBL_IRR, occurrence=1)
    f["irr_secondary"] = reader.labeled_fact(_SU, _LBL_IRR, occurrence=2)

    # Exit-year EBITDA + debt outstanding at exit: the hurdle reverse-solve
    # formula embeds the canonical cross-sheet refs — parse them rather than
    # hard-coding column math.
    su = wb[_SU]
    hurdle_row = None
    for row in su.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if isinstance(v, str) and v.startswith("Hurdle IRR"):
            hurdle_row = row[0].row
            break
    if hurdle_row is None:
        raise DeckAdapterError(
            "Hurdle-analysis block not found on SourcesUses — cannot locate "
            "the exit-year EBITDA/debt references."
        )
    hurdle_formula = str(su.cell(row=hurdle_row, column=4).value or "")
    m = _HURDLE_RE.match(hurdle_formula)
    if not m:
        raise DeckAdapterError(
            f"Could not parse the hurdle formula {hurdle_formula!r} for "
            f"exit EBITDA/debt references."
        )
    exit_ebitda = _ref_fact(reader, m.group("ebitda"))
    f["exit_ebitda"] = exit_ebitda
    if m.group("debt") == "0":
        f["exit_net_debt"] = CellFact(value=0.0, ref=f"{_SU}!D{hurdle_row} (no debt ref)")
    else:
        f["exit_net_debt"] = _ref_fact(reader, m.group("debt"))

    exit_mult = float(f["exit_multiple_strategic"].value)
    exit_ev_val = exit_mult * float(exit_ebitda.value)
    f["exit_ev"] = CellFact(
        value=exit_ev_val,
        ref=f"{f['exit_multiple_strategic'].ref} × {exit_ebitda.ref}",
    )
    f["exit_equity"] = CellFact(
        value=exit_ev_val - float(f["exit_net_debt"].value),
        ref=f"{f['exit_ev'].ref} − {f['exit_net_debt'].ref}",
    )

    # Last-actual revenue / EBITDA from OperatingModel (header row 5 carries
    # "A YYYY"/"E YYYY"; last "A" column = last historical FY).
    om = wb["OperatingModel"]
    last_actual_col = None
    for c in om[5]:
        if isinstance(c.value, str) and c.value.strip().startswith("A "):
            last_actual_col = c.column_letter
    if last_actual_col is not None:
        rev_row = _find_label_row(om, "Revenue")
        ebitda_row = _find_label_row(om, "EBITDA")
        if rev_row:
            f["revenue_last_fy"] = reader.cell_fact(
                "OperatingModel", f"{last_actual_col}{rev_row}")
        if ebitda_row:
            f["ebitda_last_fy"] = reader.cell_fact(
                "OperatingModel", f"{last_actual_col}{ebitda_row}")

    # Covenant summary: every "<name> — threshold" row; first projection-year
    # threshold + same-column actual from the row above.
    cov = wb["Covenants"]
    first_proj_col = None
    for c in cov[5]:
        if isinstance(c.value, str) and c.value.strip().startswith("E "):
            first_proj_col = c.column_letter
            break
    covenant_idx = 0
    if first_proj_col is not None:
        for row in cov.iter_rows(min_col=1, max_col=1):
            v = row[0].value
            if not (isinstance(v, str) and v.strip().endswith("— threshold")):
                continue
            name = v.strip()[: -len("— threshold")].strip(" —")
            r = row[0].row
            try:
                thr = reader.cell_fact("Covenants", f"{first_proj_col}{r}")
                act = reader.cell_fact("Covenants", f"{first_proj_col}{r - 1}")
            except DeckAdapterError:
                continue
            if thr.value is None or act.value is None:
                continue
            f[f"covenant_{covenant_idx}_name"] = CellFact(name, f"Covenants!A{r}")
            f[f"covenant_{covenant_idx}_threshold"] = thr
            f[f"covenant_{covenant_idx}_actual"] = act
            covenant_idx += 1
    breach_row = _find_label_row(cov, _LBL_BREACHES)
    if breach_row:
        f["covenant_breaches"] = reader.cell_fact("Covenants", f"C{breach_row}")

    return f


# ─────────────────────────────────────────────────────────────────────────────
# Facts → composer models
# ─────────────────────────────────────────────────────────────────────────────


def _fnum(facts: dict[str, CellFact], key: str, default: float = 0.0) -> float:
    cf = facts.get(key)
    if cf is None or cf.value is None:
        return default
    try:
        return float(cf.value)
    except (TypeError, ValueError):
        return default


def _fstr(facts: dict[str, CellFact], key: str, default: str = "") -> str:
    cf = facts.get(key)
    return str(cf.value) if cf is not None and cf.value is not None else default


def _fdate(facts: dict[str, CellFact], key: str) -> date:
    cf = facts.get(key)
    raw = cf.value if cf is not None else None
    if isinstance(raw, date):
        return raw
    if isinstance(raw, str):
        try:
            return date.fromisoformat(raw[:10])
        except ValueError:
            pass
    return date(2026, 1, 1)


def _covenant_lines(facts: dict[str, CellFact]) -> list[str]:
    lines: list[str] = []
    i = 0
    while f"covenant_{i}_name" in facts:
        name = _fstr(facts, f"covenant_{i}_name")
        thr = _fnum(facts, f"covenant_{i}_threshold")
        act = _fnum(facts, f"covenant_{i}_actual")
        lines.append(
            f"Covenant {name}: threshold {thr:.2f}x vs first-year "
            f"{act:.2f}x ({facts[f'covenant_{i}_actual'].ref})"
        )
        i += 1
    return lines


def _to_deal_facts(wf: WorkbookFacts):
    """Certified-workbook facts → ``DealFacts`` (IC-memo composer input)."""
    if wf.template == "development_re":
        return _dev_to_deal_facts(wf)
    if wf.template == "real_estate":
        return _re_to_deal_facts(wf)
    return _lbo_to_deal_facts(wf)


def _lbo_to_deal_facts(wf: WorkbookFacts):
    """sponsor_lbo facts → ``DealFacts`` (IC-memo composer input)."""
    from modelforge.deck.compose.ic_memo import DealFacts

    f = wf.facts
    currency = _fstr(f, "currency", "EUR")
    equity = _fnum(f, "sponsor_equity")
    senior = _fnum(f, "senior_debt")
    mezz = _fnum(f, "mezz_debt")
    rcf = _fnum(f, "rcf_drawn")
    rollover = _fnum(f, "mgmt_rollover")
    debt_total = senior + mezz + rcf
    entry_ev = _fnum(f, "entry_ev")
    exit_ev = _fnum(f, "exit_ev")
    exit_net_debt = _fnum(f, "exit_net_debt")
    exit_equity = _fnum(f, "exit_equity")
    irr = _fnum(f, "irr_strategic")
    moic = _fnum(f, "moic_strategic")
    ebitda = _fnum(f, "ebitda_last_fy")

    thesis = [
        f"Entry EV {entry_ev:,.1f}m ({currency}) — certified purchase-price "
        f"build ({f['entry_ev'].ref})",
        f"Strategic-exit IRR {irr:.1%} / MoIC {moic:.2f}x on the live "
        f"sponsor cash-flow series ({f['irr_strategic'].ref})",
    ]
    if ebitda:
        thesis.append(
            f"Debt at close {debt_total:,.1f}m = "
            f"{debt_total / ebitda:.1f}x LTM EBITDA {ebitda:,.1f}m "
            f"({f['senior_debt'].ref})"
        )
    thesis.extend(_covenant_lines(f)[:2])

    stack_pairs = [
        ("Senior", senior), ("Mezzanine", mezz), ("Revolver (drawn)", rcf),
        ("Rollover equity", rollover), ("Sponsor equity", equity),
    ]
    stack_pairs = [(n, v) for n, v in stack_pairs if abs(v) > 1e-9]

    if wf.red_flags:
        risks = [
            f"[{rf['severity']}] {rf['rule']}: {rf['message']}"
            for rf in wf.red_flags[:5]
        ]
    else:
        risks = ["No Trust-Layer red flags raised at build time (ALL CLEAR)."]
    breaches = int(_fnum(f, "covenant_breaches"))
    if breaches:
        risks.append(
            f"{breaches} covenant breach period(s) flagged in the projection "
            f"window ({f['covenant_breaches'].ref})."
        )
    mitigants = [
        f"Workbook certified: zero formula errors (verdict "
        f"{wf.audit_verdict}).",
        f"Bytes verified against build manifest "
        f"(sha256 {wf.manifest.workbook_sha256[:12]}…).",
    ]

    return DealFacts(
        deal_name=_fstr(f, "deal_name", "Certified Deal"),
        sector=_fstr(f, "sector", "Private Markets"),
        vertical="pe",
        country=_fstr(f, "country", "IT"),
        deal_date=_fdate(f, "valuation_date"),
        author=_fstr(f, "analyst", "Deal Team"),
        company=_fstr(f, "project_code", "ModelForge"),
        confidentiality="confidential",
        total_size_eur_m=_fnum(f, "total_uses"),
        equity_required_eur_m=equity,
        debt_eur_m=debt_total,
        hold_period_years=max(1, int(_fnum(f, "exit_year", 5))),
        investment_thesis_bullets=thesis,
        target_irr_pct=irr,
        target_moic=moic,
        waterfall_categories=[
            "Sponsor Equity (Y0)", "Exit EV (strategic)",
            "Net Debt at Exit", "Net Equity to Sponsor",
        ],
        waterfall_values=[-equity, exit_ev, -exit_net_debt, exit_equity],
        capital_stack_tranches=[n for n, _ in stack_pairs],
        capital_stack_values=[v for _, v in stack_pairs],
        risks=risks,
        mitigants=mitigants,
        recommendation="approve",
        recommendation_rationale=(
            f"All figures extracted from the certified workbook "
            f"{wf.workbook.name} (audit verdict {wf.audit_verdict}, "
            f"0 formula errors); returns from the live IRR/MoIC block on "
            f"SourcesUses. Workbook currency: {currency} millions."
        ),
        ask_eur_m=equity,
        sector_metrics={
            "entry_ev": entry_ev,
            "exit_ev": exit_ev,
            "exit_year": int(_fnum(f, "exit_year", 5)),
            "currency": currency,
            "covenants": _covenant_lines(f),
        },
    )


def _to_teaser_facts(wf: WorkbookFacts):
    """Certified-workbook facts → ``TeaserFacts`` (teaser composer input)."""
    if wf.template == "development_re":
        return _dev_to_teaser_facts(wf)
    if wf.template == "real_estate":
        return _re_to_teaser_facts(wf)
    return _lbo_to_teaser_facts(wf)


def _lbo_to_teaser_facts(wf: WorkbookFacts):
    """sponsor_lbo facts → ``TeaserFacts`` (teaser composer input)."""
    from modelforge.deck.compose.teaser import TeaserFacts

    f = wf.facts
    revenue = _fnum(f, "revenue_last_fy") or None
    ebitda = _fnum(f, "ebitda_last_fy") or None
    margin = (ebitda / revenue) if (revenue and ebitda) else None
    irr = _fnum(f, "irr_strategic")
    moic = _fnum(f, "moic_strategic")
    currency = _fstr(f, "currency", "EUR")

    return TeaserFacts(
        project_codename=_fstr(f, "project_code", "Project Forge"),
        company_real_name=_fstr(f, "deal_name") or None,
        anonymized=True,
        sector=_fstr(f, "sector", "Private Markets"),
        vertical="pe",
        country=_fstr(f, "country", "IT"),
        deal_date=_fdate(f, "valuation_date"),
        author=_fstr(f, "analyst", "Advisor"),
        company=_fstr(f, "project_code", "ModelForge"),
        confidentiality="confidential",
        one_line_thesis=(
            f"Certified LBO model: {irr:.0%} strategic-exit IRR / "
            f"{moic:.1f}x MoIC ({currency} millions)."
        ),
        revenue_eur_m=revenue,
        ebitda_eur_m=ebitda,
        ebitda_margin_pct=margin,
        ask_eur_m=_fnum(f, "sponsor_equity") or None,
        enterprise_value_eur_m=_fnum(f, "entry_ev") or None,
        investment_highlights=[
            f"Strategic-exit IRR {irr:.1%} / MoIC {moic:.2f}x "
            f"({f['irr_strategic'].ref})",
            f"Entry EV {_fnum(f, 'entry_ev'):,.1f}m "
            f"({f['entry_ev'].ref})",
            "Every figure recomputed and certified — zero formula errors "
            "(ModelForge Trust Layer).",
        ],
        contact_name=_fstr(f, "analyst", "Deal Team"),
        contact_role="Deal Team",
    )


# ─────────────────────────────────────────────────────────────────────────────
# real_estate extraction
# ─────────────────────────────────────────────────────────────────────────────

# Exact col-A labels emitted by builder/sheets/re_dcf.py and re_financing.py.
_RE_LBL_ACQ = "Acquisition price"
_RE_LBL_NOI = "NOI (Net Operating Income)"
_RE_LBL_EXIT_NOI = "Exit-year NOI"
_RE_LBL_EXIT_VALUE = "Gross sale proceeds (NOI / cap)"
_RE_LBL_TXN_COSTS = "Transaction costs"
_RE_LBL_NET_EXIT = "Net sale proceeds"
_RE_LBL_LOAN = "Loan amount (at close)"
_RE_LBL_EQUITY_CF = "Equity cash flow"
_RE_LBL_EQUITY_IRR = "Equity IRR"
_RE_LBL_EQUITY_MOIC = "Equity MoIC"
_RE_LBL_LP_PREF = "LP preferred return (compounded)"
_RE_LBL_TOTAL_EQ_CONTRIB = "Total equity contribution (t=0)"
_RE_LBL_TOTAL_EQ_DISTRIB = "Total equity distributions (gross)"
_RE_LBL_LP_TOTAL = "LP total post-waterfall"
_RE_LBL_GP_TOTAL = "GP total post-waterfall"


def _re_hold_years(reader: _Reader) -> Optional[CellFact]:
    """Hold years from the DCF year-header row (largest ``t=N`` header).

    The RE workbook carries no hold-years named range; the t-headers on the
    DCF year row (row 5) are the authoritative source. Returns the integer
    hold horizon with the header cell it was read from, or None.
    """
    ws = reader.wb["DCF"]
    best: Optional[int] = None
    best_ref: Optional[str] = None
    for row in ws.iter_rows(min_row=5, max_row=5):
        for c in row:
            v = c.value
            if isinstance(v, str):
                m = re.fullmatch(r"\s*t\s*=\s*(\d+)\s*", v)
                if m:
                    n = int(m.group(1))
                    if best is None or n > best:
                        best = n
                        best_ref = f"DCF!{c.coordinate}"
    if best is None:
        return None
    return CellFact(value=best, ref=best_ref or "DCF!t-headers")


def _extract_real_estate(reader: _Reader) -> dict[str, CellFact]:
    """Extract DealFacts-compatible CellFacts from a certified real_estate
    workbook (DCF + Financing sheets, shared Cover identity block).

    Every numeric fact keeps its ``Sheet!Cell`` source ref. Headline fields
    the RE workbook does NOT contain (LTM revenue/EBITDA, exit-multiple comps,
    sensitivity grid) are simply not populated so the composer drops those
    slides — the no-hollow-slides invariant is preserved.
    """
    wb = reader.wb
    for required in ("DCF", "Financing", "Cover"):
        if required not in wb.sheetnames:
            raise DeckAdapterError(
                f"Sheet {required!r} missing — this workbook does not match "
                f"the real_estate layout the deck adapter was verified "
                f"against."
            )

    f: dict[str, CellFact] = {}

    # ── Cover identity block (shared; labels col A, values col C) ──────────
    f["deal_name"] = _cover_value(reader, "Target")
    f["sector"] = _cover_value(reader, "Sector")
    f["country"] = _cover_value(reader, "Country")
    f["currency"] = _cover_value(reader, "Currency")
    f["project_code"] = _cover_value(reader, "Project code")
    f["analyst"] = _cover_value(reader, "Analyst")
    f["valuation_date"] = _cover_value(reader, "Valuation date")

    # ── Going-in / acquisition ────────────────────────────────────────────
    # Acquisition price is the going-in gross asset value (positive). Prefer
    # the registered named-range input; the DCF row carries it negated.
    f["acquisition_price"] = reader.named_fact("acquisition_price_eur_m")
    f["exit_cap_rate"] = reader.named_fact("exit_cap_rate")
    f["ltv_pct"] = reader.named_fact("ltv_pct")
    f["lp_capital_pct"] = reader.named_fact("lp_capital_commitment_pct")

    # ── DCF: NOI + exit value ─────────────────────────────────────────────
    f["exit_noi"] = reader.labeled_fact("DCF", _RE_LBL_EXIT_NOI)
    f["exit_value_gross"] = reader.labeled_fact("DCF", _RE_LBL_EXIT_VALUE)
    f["exit_txn_costs"] = reader.labeled_fact("DCF", _RE_LBL_TXN_COSTS)
    f["net_sale_proceeds"] = reader.labeled_fact("DCF", _RE_LBL_NET_EXIT)

    # Stabilised going-in NOI = year-1 NOI (column to the right of t=0/col D
    # on the annual NOI row). Used for the going-in yield-on-cost metric.
    noi_rows = _find_label_rows(wb["DCF"], _RE_LBL_NOI)
    if noi_rows:
        # year-1 is the first projection column (column E, t=1).
        try:
            f["going_in_noi"] = reader.cell_fact("DCF", f"E{noi_rows[0]}")
        except DeckAdapterError:
            pass

    # ── Financing: debt + equity returns + waterfall ──────────────────────
    f["loan_amount"] = reader.labeled_fact("Financing", _RE_LBL_LOAN)
    f["equity_cf_t0"] = reader.labeled_fact("Financing", _RE_LBL_EQUITY_CF)
    f["levered_equity_irr"] = reader.labeled_fact("Financing", _RE_LBL_EQUITY_IRR)
    f["equity_moic"] = reader.labeled_fact("Financing", _RE_LBL_EQUITY_MOIC)
    f["total_equity_contrib"] = reader.labeled_fact(
        "Financing", _RE_LBL_TOTAL_EQ_CONTRIB)
    f["lp_pref"] = reader.labeled_fact("Financing", _RE_LBL_LP_PREF)
    f["lp_total_post"] = reader.labeled_fact("Financing", _RE_LBL_LP_TOTAL)
    f["gp_total_post"] = reader.labeled_fact("Financing", _RE_LBL_GP_TOTAL)

    # Total gross equity distributions (D32) is a standalone `SUMIF(...,">0")`
    # that the `formulas` engine declines to recompute as an isolated cell
    # (it resolves fine when nested in downstream cells). It only enriches the
    # sector-metrics block, so it is OPTIONAL — never let a recompute quirk on
    # a non-headline cell fail the whole deck. If it cannot be recomputed it is
    # simply omitted (the composer never depends on it).
    try:
        f["total_equity_distrib"] = reader.labeled_fact(
            "Financing", _RE_LBL_TOTAL_EQ_DISTRIB)
    except DeckAdapterError:
        pass

    # GP promote (carried interest) = 1 − LP share of residual.
    try:
        gp_promote = reader.named_fact("gp_promote_pct")
        f["gp_promote"] = gp_promote
    except DeckAdapterError:
        pass

    hold = _re_hold_years(reader)
    if hold is not None:
        f["hold_years"] = hold

    return f


def _re_to_deal_facts(wf: WorkbookFacts):
    """real_estate facts → ``DealFacts`` (IC-memo composer input).

    Honest mapping: RE populates the levered equity IRR / MoIC, the going-in
    cap & yield-on-cost, the entry→exit equity cash-flow waterfall, the
    Senior-debt / LP-equity / GP-equity capital stack, and the LP-pref /
    GP-promote economics. It does NOT carry LTM revenue/EBITDA, exit-multiple
    comparable transactions, or an IRR sensitivity grid — those DealFacts
    fields are deliberately left empty so the composer omits the comps and
    sensitivity slides rather than ship them hollow.
    """
    from modelforge.deck.compose.ic_memo import DealFacts

    f = wf.facts
    currency = _fstr(f, "currency", "EUR")
    acq = _fnum(f, "acquisition_price")
    debt = _fnum(f, "loan_amount")
    equity = _fnum(f, "total_equity_contrib")
    if equity <= 0:  # fall back to abs(t=0 equity CF)
        equity = abs(_fnum(f, "equity_cf_t0"))
    irr = _fnum(f, "levered_equity_irr")
    moic = _fnum(f, "equity_moic")
    exit_noi = _fnum(f, "exit_noi")
    exit_value = _fnum(f, "exit_value_gross")
    net_exit = _fnum(f, "net_sale_proceeds")
    cap_rate = _fnum(f, "exit_cap_rate")
    going_in_noi = _fnum(f, "going_in_noi")
    yield_on_cost = (going_in_noi / acq) if (going_in_noi and acq) else None
    lp_pref = _fnum(f, "lp_pref")
    gp_promote = _fnum(f, "gp_promote")
    hold_years = max(1, int(_fnum(f, "hold_years", 7)))

    thesis: list[str] = [
        f"Going-in asset value {acq:,.1f}m ({currency}) at "
        f"{(debt / acq) if acq else 0:.0%} LTV "
        f"({f['acquisition_price'].ref})",
        f"Levered equity IRR {irr:.1%} / MoIC {moic:.2f}x on the certified "
        f"equity cash-flow series ({f['levered_equity_irr'].ref})",
        f"Exit on a {cap_rate:.2%} cap rate → gross sale proceeds "
        f"{exit_value:,.1f}m on exit-year NOI {exit_noi:,.2f}m "
        f"({f['exit_value_gross'].ref})",
    ]
    if yield_on_cost is not None:
        thesis.append(
            f"Going-in yield-on-cost {yield_on_cost:.2%} "
            f"(year-1 NOI {going_in_noi:,.2f}m / cost {acq:,.1f}m; "
            f"{f['going_in_noi'].ref})"
        )

    # Entry → exit equity waterfall (EUR M): equity in (−), net sale proceeds
    # (+), debt repaid at exit (−), net equity to investors (=).
    debt_repaid = -debt
    net_equity = net_exit + debt_repaid
    waterfall_categories = [
        "Equity Invested (t=0)", "Net Sale Proceeds",
        "Senior Debt Repaid", "Net Equity to Investors",
    ]
    waterfall_values = [-equity, net_exit, debt_repaid, net_equity]

    # Capital stack: Senior mortgage + LP equity + GP equity.
    lp_pct = _fnum(f, "lp_capital_pct")
    lp_equity = equity * lp_pct
    gp_equity = equity * (1.0 - lp_pct)
    stack_pairs = [
        ("Senior Mortgage", debt),
        ("LP Equity", lp_equity),
        ("GP Equity", gp_equity),
    ]
    stack_pairs = [(n, v) for n, v in stack_pairs if abs(v) > 1e-9]

    if wf.red_flags:
        risks = [
            f"[{rf['severity']}] {rf['rule']}: {rf['message']}"
            for rf in wf.red_flags[:5]
        ]
    else:
        risks = ["No Trust-Layer red flags raised at build time (ALL CLEAR)."]
    # Honest-label: the workbook waterfall is a simplified illustrative LP/GP
    # structure (not a full tier-by-tier waterfall). Surface that on the IC
    # memo so the committee sees the limitation explicitly.
    risks.append(
        "LP/GP waterfall is a simplified illustrative allocation (LP pref + "
        "GP catch-up + promote), not a full tier-by-tier model."
    )
    mitigants = [
        f"Workbook certified: zero formula errors (verdict "
        f"{wf.audit_verdict}).",
        f"Bytes verified against build manifest "
        f"(sha256 {wf.manifest.workbook_sha256[:12]}…).",
        f"LP preferred return {lp_pref:.2%}, GP promote {gp_promote:.0%} "
        f"on the residual ({f['lp_pref'].ref}).",
    ]

    sector_metrics = {
        "exit_cap_rate": cap_rate,
        "exit_noi": exit_noi,
        "exit_value_gross": exit_value,
        "net_sale_proceeds": net_exit,
        "lp_preferred_return": lp_pref,
        "gp_promote": gp_promote,
        "currency": currency,
    }
    if yield_on_cost is not None:
        sector_metrics["going_in_yield_on_cost"] = yield_on_cost

    return DealFacts(
        deal_name=_fstr(f, "deal_name", "Certified Asset"),
        sector=_fstr(f, "sector", "Real Estate"),
        vertical="re",
        country=_fstr(f, "country", "IT"),
        deal_date=_fdate(f, "valuation_date"),
        author=_fstr(f, "analyst", "Deal Team"),
        company=_fstr(f, "project_code", "ModelForge"),
        confidentiality="confidential",
        total_size_eur_m=acq,
        equity_required_eur_m=equity,
        debt_eur_m=debt,
        hold_period_years=hold_years,
        investment_thesis_title="Investment Thesis",
        investment_thesis_bullets=thesis,
        target_irr_pct=irr,
        target_moic=moic,
        levered_irr_pct=irr,
        yield_on_cost_pct=yield_on_cost,
        waterfall_categories=waterfall_categories,
        waterfall_values=waterfall_values,
        capital_stack_tranches=[n for n, _ in stack_pairs],
        capital_stack_values=[v for _, v in stack_pairs],
        risks=risks,
        mitigants=mitigants,
        recommendation="approve",
        recommendation_rationale=(
            f"All figures extracted from the certified workbook "
            f"{wf.workbook.name} (audit verdict {wf.audit_verdict}, "
            f"0 formula errors); levered equity returns from the IRR/MoIC "
            f"block on Financing. Workbook currency: {currency} millions."
        ),
        ask_eur_m=equity,
        sector_metrics=sector_metrics,
    )


def _re_to_teaser_facts(wf: WorkbookFacts):
    """real_estate facts → ``TeaserFacts`` (teaser composer input).

    RE has no LTM revenue/EBITDA, so ``revenue_eur_m`` / ``ebitda_eur_m`` are
    left None (the teaser snapshot KPI cards drop accordingly). The asset
    economics (going-in value, levered IRR/MoIC, exit cap, NOI) carry through
    the EV field and the highlight bullets instead.
    """
    from modelforge.deck.compose.teaser import TeaserFacts

    f = wf.facts
    currency = _fstr(f, "currency", "EUR")
    acq = _fnum(f, "acquisition_price")
    equity = _fnum(f, "total_equity_contrib")
    if equity <= 0:
        equity = abs(_fnum(f, "equity_cf_t0"))
    irr = _fnum(f, "levered_equity_irr")
    moic = _fnum(f, "equity_moic")
    cap_rate = _fnum(f, "exit_cap_rate")
    exit_noi = _fnum(f, "exit_noi")
    going_in_noi = _fnum(f, "going_in_noi")
    yield_on_cost = (going_in_noi / acq) if (going_in_noi and acq) else None

    highlights = [
        f"Levered equity IRR {irr:.1%} / MoIC {moic:.2f}x "
        f"({f['levered_equity_irr'].ref})",
        f"Going-in asset value {acq:,.1f}m ({currency}) "
        f"({f['acquisition_price'].ref})",
        f"Exit cap rate {cap_rate:.2%} on exit-year NOI {exit_noi:,.2f}m "
        f"({f['exit_value_gross'].ref})",
    ]
    if yield_on_cost is not None:
        highlights.append(
            f"Going-in yield-on-cost {yield_on_cost:.2%} "
            f"({f['going_in_noi'].ref})"
        )
    highlights.append(
        "Every figure recomputed and certified — zero formula errors "
        "(ModelForge Trust Layer)."
    )

    snapshot = [
        f"Sector: {_fstr(f, 'sector', 'Real Estate')}",
        f"Geography: {_fstr(f, 'country', 'IT')}",
        f"Going-in value: {currency} {acq:,.1f}m",
        f"Exit cap rate: {cap_rate:.2%}",
    ]

    return TeaserFacts(
        project_codename=_fstr(f, "project_code", "Project Forge"),
        company_real_name=_fstr(f, "deal_name") or None,
        anonymized=True,
        sector=_fstr(f, "sector", "Real Estate"),
        vertical="re",
        country=_fstr(f, "country", "IT"),
        deal_date=_fdate(f, "valuation_date"),
        author=_fstr(f, "analyst", "Advisor"),
        company=_fstr(f, "project_code", "ModelForge"),
        confidentiality="confidential",
        one_line_thesis=(
            f"Certified RE model: {irr:.0%} levered equity IRR / "
            f"{moic:.1f}x MoIC, {cap_rate:.2%} exit cap ({currency} millions)."
        ),
        revenue_eur_m=None,
        ebitda_eur_m=None,
        ebitda_margin_pct=None,
        ask_eur_m=equity or None,
        enterprise_value_eur_m=acq or None,
        company_snapshot_bullets=snapshot,
        investment_highlights=highlights,
        contact_name=_fstr(f, "analyst", "Deal Team"),
        contact_role="Deal Team",
    )


# ─────────────────────────────────────────────────────────────────────────────
# development_re extraction (ground-up development underwriting)
# ─────────────────────────────────────────────────────────────────────────────

# Exact col-A labels emitted by builder/sheets/dev_schedule.py + dev_returns.py.
_DEV_LBL_TDC = "Total development cost (TDC)"
_DEV_LBL_EXIT_VALUE = "Gross exit value (fwd NOI / cap)"
_DEV_LBL_FWD_NOI = "Forward NOI at exit"
_DEV_LBL_STAB_NOI = "Stabilised annual NOI"
_DEV_LBL_UNLEV_IRR = "Unlevered IRR (annual)"
_DEV_LBL_UNLEV_MOIC = "Unlevered MOIC"
_DEV_LBL_LEV_IRR = "Levered equity IRR (annual)"
_DEV_LBL_LEV_MOIC = "Levered equity MOIC"
_DEV_LBL_EQUITY_INVESTED = "Total equity invested"
_DEV_LBL_PEAK_DEBT = "Peak senior debt"
_DEV_LBL_NET_EXIT = "Net exit proceeds"
_DEV_LBL_PREF_RETURN = "LP preferred return (compounded)"
_DEV_LBL_PREF_THRESHOLD = "LP pref threshold (capital × (1+pref)^hold)"
_DEV_LBL_LP_TOTAL = "LP total post-waterfall"
_DEV_LBL_GP_TOTAL = "GP total post-waterfall"


def _dev_equity_aggregate(reader: _Reader, sign: str) -> CellFact:
    """Aggregate the DevSchedule equity cash-flow row by sign.

    sign == "<0":  ABS(Σ negative equity-CF periods) → total equity invested.
    sign == ">0":  Σ positive equity-CF periods       → gross distributions.

    Reads each year cell of the equity-CF DATA row (the LAST "Equity cash flow"
    row on DevSchedule — the section header carries the same label above it) and
    sums in Python, so it never depends on a standalone-``SUMIF`` output cell the
    `formulas` engine declines to recompute in isolation.
    """
    ws = reader.wb["DevSchedule"]
    rows = _find_label_rows(ws, "Equity cash flow")
    if not rows:
        raise DeckAdapterError(
            "Equity cash-flow row not found on DevSchedule — workbook does not "
            "match the development_re layout."
        )
    data_row = rows[-1]  # section header is the earlier occurrence
    # Year columns are D.. (col 4) onward; sum across the populated year cells.
    total = 0.0
    last_col = "D"
    for col_idx in range(4, ws.max_column + 1):
        coord = f"{ws.cell(row=data_row, column=col_idx).column_letter}{data_row}"
        try:
            cf = reader.cell_fact("DevSchedule", coord)
        except DeckAdapterError:
            continue
        if cf.value is None:
            continue
        try:
            v = float(cf.value)
        except (TypeError, ValueError):
            continue
        if (sign == "<0" and v < 0) or (sign == ">0" and v > 0):
            total += v
            last_col = ws.cell(row=data_row, column=col_idx).column_letter
    value = abs(total) if sign == "<0" else total
    label = "Σ<0" if sign == "<0" else "Σ>0"
    return CellFact(
        value=value,
        ref=f"DevSchedule!D{data_row}:{last_col}{data_row} ({label})",
    )


def _extract_development_re(reader: _Reader) -> dict[str, CellFact]:
    """Extract DealFacts-compatible CellFacts from a certified development_re
    workbook (DevSchedule + Returns sheets, shared Cover identity block).

    The development template is an ANNUAL-PHASED ground-up underwriting: phased
    capex → S-curve lease-up → forward-NOI cap-rate exit, financed pro-rata
    loan-to-cost with construction interest capitalised, and distributed through
    a European whole-fund LP/GP promote. Every numeric fact keeps its
    ``Sheet!Cell`` source ref.

    Headline fields the development workbook does NOT carry (LTM revenue/EBITDA,
    exit-multiple comps, sensitivity grid) are simply not populated, so the
    composer drops those slides — the no-hollow-slides invariant is preserved.
    """
    wb = reader.wb
    for required in ("DevSchedule", "Returns", "Cover"):
        if required not in wb.sheetnames:
            raise DeckAdapterError(
                f"Sheet {required!r} missing — this workbook does not match "
                f"the development_re layout the deck adapter was verified "
                f"against."
            )

    f: dict[str, CellFact] = {}

    # ── Cover identity block (shared; labels col A, values col C) ──────────
    f["deal_name"] = _cover_value(reader, "Target")
    f["sector"] = _cover_value(reader, "Sector")
    f["country"] = _cover_value(reader, "Country")
    f["currency"] = _cover_value(reader, "Currency")
    f["project_code"] = _cover_value(reader, "Project code")
    f["analyst"] = _cover_value(reader, "Analyst")
    f["valuation_date"] = _cover_value(reader, "Valuation date")

    # ── DevSchedule: total dev cost + exit value ──────────────────────────
    f["total_dev_cost"] = reader.labeled_fact("DevSchedule", _DEV_LBL_TDC)
    f["exit_value_gross"] = reader.labeled_fact("DevSchedule", _DEV_LBL_EXIT_VALUE)
    f["forward_noi"] = reader.labeled_fact("DevSchedule", _DEV_LBL_FWD_NOI)
    f["stabilised_noi"] = reader.labeled_fact("DevSchedule", _DEV_LBL_STAB_NOI)

    # ── Returns: unlevered + levered IRR/MOIC, equity invested, peak debt ──
    f["unlevered_irr"] = reader.labeled_fact("Returns", _DEV_LBL_UNLEV_IRR)
    f["unlevered_moic"] = reader.labeled_fact("Returns", _DEV_LBL_UNLEV_MOIC)
    f["levered_equity_irr"] = reader.labeled_fact("Returns", _DEV_LBL_LEV_IRR)
    f["equity_moic"] = reader.labeled_fact("Returns", _DEV_LBL_LEV_MOIC)
    f["peak_debt"] = reader.labeled_fact("Returns", _DEV_LBL_PEAK_DEBT)
    # "Net exit proceeds" appears on both DevSchedule and Returns; read the
    # Returns occurrence (it cross-refs the schedule's net-exit cell).
    f["net_exit_proceeds"] = reader.labeled_fact("Returns", _DEV_LBL_NET_EXIT)

    # Total equity invested. The Returns "Total equity invested" cell is a
    # standalone ``ABS(SUMIF(...,"<0"))`` that the `formulas` engine declines to
    # recompute as an ISOLATED cell (it resolves fine when nested in downstream
    # cells, but the adapter never trusts an unrecomputable cached value). The
    # equity cash-flow ROW on the schedule, by contrast, recomputes cell-by-cell
    # cleanly — sum its negative periods for the contribution and its positive
    # periods for the gross distribution, with a composite source ref. This is
    # the same engine-quirk workaround the real_estate extractor uses.
    f["total_equity_contrib"] = _dev_equity_aggregate(reader, "<0")
    try:
        f["total_equity_distrib"] = _dev_equity_aggregate(reader, ">0")
    except DeckAdapterError:
        pass

    # ── Returns: European promote waterfall economics ─────────────────────
    f["lp_pref"] = reader.labeled_fact("Returns", _DEV_LBL_PREF_RETURN)
    f["lp_pref_threshold"] = reader.labeled_fact("Returns", _DEV_LBL_PREF_THRESHOLD)
    f["lp_total_post"] = reader.labeled_fact("Returns", _DEV_LBL_LP_TOTAL)
    f["gp_total_post"] = reader.labeled_fact("Returns", _DEV_LBL_GP_TOTAL)

    # ── Named-range inputs (exit cap, LP%, GP promote) ────────────────────
    f["exit_cap_rate"] = reader.named_fact("dev_exit_cap_rate")
    try:
        f["lp_capital_pct"] = reader.named_fact("lp_capital_commitment_pct")
    except DeckAdapterError:
        pass
    try:
        f["gp_promote"] = reader.named_fact("gp_promote_pct")
    except DeckAdapterError:
        pass

    # ── Hold (years) from the DevSchedule t-headers (largest t=N tag) ──────
    hold = _dev_hold_years(reader)
    if hold is not None:
        f["hold_years"] = hold

    return f


def _dev_hold_years(reader: _Reader) -> Optional[CellFact]:
    """Hold years = the largest ``t=N`` tag on the DevSchedule header row (5).

    The DevSchedule headers read ``t=0 C``, ``t=1 C`` … ``t=H X``; the exit
    year is the largest N. Returns the integer hold horizon with the header cell
    it was read from, or None.
    """
    ws = reader.wb["DevSchedule"]
    best: Optional[int] = None
    best_ref: Optional[str] = None
    for row in ws.iter_rows(min_row=5, max_row=5):
        for c in row:
            v = c.value
            if isinstance(v, str):
                m = re.match(r"\s*t\s*=\s*(\d+)", v)
                if m:
                    nval = int(m.group(1))
                    if best is None or nval > best:
                        best = nval
                        best_ref = f"DevSchedule!{c.coordinate}"
    if best is None:
        return None
    return CellFact(value=best, ref=best_ref or "DevSchedule!t-headers")


def _dev_to_deal_facts(wf: WorkbookFacts):
    """development_re facts → ``DealFacts`` (IC-memo composer input).

    Honest mapping: development_re populates the levered equity IRR / MoIC, the
    UNLEVERED project IRR (a development carries a real unlevered return), the
    total development cost, the forward-NOI cap-rate exit value, the equity
    invested → distribution waterfall, the Senior-debt / LP-equity / GP-equity
    capital stack, and the LP-pref / GP-promote economics. It does NOT carry LTM
    revenue/EBITDA, exit-multiple comparable transactions, or an IRR sensitivity
    grid — those DealFacts fields are deliberately left empty so the composer
    omits the comps and sensitivity slides rather than ship them hollow.
    """
    from modelforge.deck.compose.ic_memo import DealFacts

    f = wf.facts
    currency = _fstr(f, "currency", "EUR")
    tdc = _fnum(f, "total_dev_cost")
    equity = _fnum(f, "total_equity_contrib")
    peak_debt = _fnum(f, "peak_debt")
    irr = _fnum(f, "levered_equity_irr")
    unlev_irr = _fnum(f, "unlevered_irr")
    moic = _fnum(f, "equity_moic")
    exit_value = _fnum(f, "exit_value_gross")
    net_exit = _fnum(f, "net_exit_proceeds")
    fwd_noi = _fnum(f, "forward_noi")
    cap_rate = _fnum(f, "exit_cap_rate")
    lp_pref = _fnum(f, "lp_pref")
    gp_promote = _fnum(f, "gp_promote")
    gp_total = _fnum(f, "gp_total_post")
    hold_years = max(1, int(_fnum(f, "hold_years", 5)))
    ltc = (peak_debt / tdc) if tdc else 0.0

    thesis: list[str] = [
        f"Total development cost {tdc:,.1f}m ({currency}) at {ltc:.0%} "
        f"loan-to-cost ({f['total_dev_cost'].ref})",
        f"Levered equity IRR {irr:.1%} / MoIC {moic:.2f}x on the certified "
        f"equity cash-flow series ({f['levered_equity_irr'].ref})",
        f"Unlevered project IRR {unlev_irr:.1%} on the project cash flow "
        f"({f['unlevered_irr'].ref})",
        f"Stabilised exit on a {cap_rate:.2%} cap rate → gross exit value "
        f"{exit_value:,.1f}m on forward NOI {fwd_noi:,.2f}m "
        f"({f['exit_value_gross'].ref})",
    ]

    # Entry → exit equity waterfall (EUR M): equity in (−), net exit proceeds
    # (+), senior debt repaid at exit (−), net equity to investors (=).
    debt_repaid = -peak_debt
    net_equity = net_exit + debt_repaid
    waterfall_categories = [
        "Equity Invested", "Net Exit Proceeds",
        "Senior Debt Repaid", "Net Equity to Investors",
    ]
    waterfall_values = [-equity, net_exit, debt_repaid, net_equity]

    # Capital stack: Senior development debt (peak) + LP equity + GP equity.
    lp_pct = _fnum(f, "lp_capital_pct")
    lp_equity = equity * lp_pct
    gp_equity = equity * (1.0 - lp_pct)
    stack_pairs = [
        ("Senior Development Debt", peak_debt),
        ("LP Equity", lp_equity),
        ("GP Equity", gp_equity),
    ]
    stack_pairs = [(n, v) for n, v in stack_pairs if abs(v) > 1e-9]

    if wf.red_flags:
        risks = [
            f"[{rf['severity']}] {rf['rule']}: {rf['message']}"
            for rf in wf.red_flags[:5]
        ]
    else:
        risks = ["No Trust-Layer red flags raised at build time (ALL CLEAR)."]
    risks.append(
        "Ground-up development risk: programme, construction-cost and lease-up "
        "timing drive the levered return; the lease-up S-curve and "
        "construction-interest capitalisation are modelled annually."
    )
    risks.append(
        "LP/GP waterfall is a European whole-fund allocation (LP pref + GP "
        "catch-up + promote), an illustrative structure, not a full "
        "tier-by-tier model."
    )
    mitigants = [
        f"Workbook certified: zero formula errors (verdict "
        f"{wf.audit_verdict}).",
        f"Bytes verified against build manifest "
        f"(sha256 {wf.manifest.workbook_sha256[:12]}…).",
        f"LP preferred return {lp_pref:.2%}, GP promote {gp_promote:.0%} "
        f"on the residual ({f['lp_pref'].ref}).",
    ]

    sector_metrics = {
        "total_development_cost": tdc,
        "peak_senior_debt": peak_debt,
        "loan_to_cost": ltc,
        "exit_cap_rate": cap_rate,
        "forward_noi": fwd_noi,
        "exit_value_gross": exit_value,
        "net_exit_proceeds": net_exit,
        "unlevered_irr": unlev_irr,
        "lp_preferred_return": lp_pref,
        "gp_promote": gp_promote,
        "gp_total_post_waterfall": gp_total,
        "currency": currency,
    }

    return DealFacts(
        deal_name=_fstr(f, "deal_name", "Certified Development"),
        sector=_fstr(f, "sector", "Real Estate Development"),
        vertical="re",
        country=_fstr(f, "country", "IT"),
        deal_date=_fdate(f, "valuation_date"),
        author=_fstr(f, "analyst", "Deal Team"),
        company=_fstr(f, "project_code", "ModelForge"),
        confidentiality="confidential",
        total_size_eur_m=tdc,
        equity_required_eur_m=equity,
        debt_eur_m=peak_debt,
        hold_period_years=hold_years,
        investment_thesis_title="Investment Thesis",
        investment_thesis_bullets=thesis,
        target_irr_pct=irr,
        target_moic=moic,
        levered_irr_pct=irr,
        unlevered_irr_pct=unlev_irr,
        waterfall_categories=waterfall_categories,
        waterfall_values=waterfall_values,
        capital_stack_tranches=[n for n, _ in stack_pairs],
        capital_stack_values=[v for _, v in stack_pairs],
        risks=risks,
        mitigants=mitigants,
        recommendation="approve",
        recommendation_rationale=(
            f"All figures extracted from the certified workbook "
            f"{wf.workbook.name} (audit verdict {wf.audit_verdict}, "
            f"0 formula errors); unlevered + levered returns from the IRR/MoIC "
            f"block on Returns. Workbook currency: {currency} millions."
        ),
        ask_eur_m=equity,
        sector_metrics=sector_metrics,
    )


def _dev_to_teaser_facts(wf: WorkbookFacts):
    """development_re facts → ``TeaserFacts`` (teaser composer input).

    A ground-up development has no LTM revenue/EBITDA, so ``revenue_eur_m`` /
    ``ebitda_eur_m`` are left None (the teaser snapshot KPI cards drop). The
    development economics (total dev cost, levered + unlevered IRR/MoIC, exit
    cap, forward NOI) carry through the EV field and the highlight bullets.
    """
    from modelforge.deck.compose.teaser import TeaserFacts

    f = wf.facts
    currency = _fstr(f, "currency", "EUR")
    tdc = _fnum(f, "total_dev_cost")
    equity = _fnum(f, "total_equity_contrib")
    irr = _fnum(f, "levered_equity_irr")
    unlev_irr = _fnum(f, "unlevered_irr")
    moic = _fnum(f, "equity_moic")
    cap_rate = _fnum(f, "exit_cap_rate")
    fwd_noi = _fnum(f, "forward_noi")
    exit_value = _fnum(f, "exit_value_gross")
    peak_debt = _fnum(f, "peak_debt")
    ltc = (peak_debt / tdc) if tdc else 0.0

    highlights = [
        f"Levered equity IRR {irr:.1%} / MoIC {moic:.2f}x "
        f"({f['levered_equity_irr'].ref})",
        f"Unlevered project IRR {unlev_irr:.1%} ({f['unlevered_irr'].ref})",
        f"Total development cost {tdc:,.1f}m ({currency}) at {ltc:.0%} "
        f"loan-to-cost ({f['total_dev_cost'].ref})",
        f"Stabilised exit cap rate {cap_rate:.2%} on forward NOI "
        f"{fwd_noi:,.2f}m ({f['exit_value_gross'].ref})",
        "Every figure recomputed and certified — zero formula errors "
        "(ModelForge Trust Layer).",
    ]

    snapshot = [
        f"Sector: {_fstr(f, 'sector', 'Real Estate Development')}",
        f"Geography: {_fstr(f, 'country', 'IT')}",
        f"Total development cost: {currency} {tdc:,.1f}m",
        f"Gross exit value: {currency} {exit_value:,.1f}m",
        f"Exit cap rate: {cap_rate:.2%}",
    ]

    return TeaserFacts(
        project_codename=_fstr(f, "project_code", "Project Forge"),
        company_real_name=_fstr(f, "deal_name") or None,
        anonymized=True,
        sector=_fstr(f, "sector", "Real Estate Development"),
        vertical="re",
        country=_fstr(f, "country", "IT"),
        deal_date=_fdate(f, "valuation_date"),
        author=_fstr(f, "analyst", "Advisor"),
        company=_fstr(f, "project_code", "ModelForge"),
        confidentiality="confidential",
        one_line_thesis=(
            f"Certified ground-up development: {irr:.0%} levered equity IRR / "
            f"{moic:.1f}x MoIC, {cap_rate:.2%} exit cap ({currency} millions)."
        ),
        revenue_eur_m=None,
        ebitda_eur_m=None,
        ebitda_margin_pct=None,
        ask_eur_m=equity or None,
        enterprise_value_eur_m=tdc or None,
        company_snapshot_bullets=snapshot,
        investment_highlights=highlights,
        contact_name=_fstr(f, "analyst", "Deal Team"),
        contact_role="Deal Team",
    )
