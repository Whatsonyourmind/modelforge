"""Regulator-grade PDF dossier generator.

Reads a built ModelForge workbook and produces a PDF containing:

    1. Cover page — deal name, version, spec SHA-256, build timestamp
    2. Executive Summary — target, deliverable, scenario, primary output
    3. Assumptions Register — every Assumption with unit, scenario
       values, rationale, confidence, source link
    4. Source Registry — every S-### doc with publisher, page, URL,
       verified status
    5. Formula Inventory — unique formula shapes per sheet (sampled)
    6. Lineage Graph Summary — node / edge counts per kind
    7. QC Sign-off — the 8-check external QC gate result + auditor
       signature field
    8. Glossary — bilingual EN/IT terminology

The PDF embeds the spec SHA-256 on the cover so the dossier itself is
version-locked. Rebuild the workbook, rebuild the dossier.
"""

from __future__ import annotations

import html
import sqlite3
import re
from collections import Counter
from pathlib import Path
from typing import Optional


def _esc(s) -> str:
    """Escape text for reportlab Paragraph (which uses XML-like markup)."""
    if s is None:
        return ""
    return html.escape(str(s), quote=False)

from openpyxl import load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm, mm
from reportlab.platypus import (
    HRFlowable,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from modelforge.analytics.reproducibility import read_reproducibility
from modelforge.qc import run_qc


# ─── Styles ───────────────────────────────────────────────────────────────────


def _styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    s = {
        "title": ParagraphStyle("title", parent=base["Title"],
                                fontName="Helvetica-Bold", fontSize=22,
                                textColor=colors.HexColor("#1F3864"),
                                spaceAfter=12),
        "h1": ParagraphStyle("h1", parent=base["Heading1"],
                             fontName="Helvetica-Bold", fontSize=16,
                             textColor=colors.HexColor("#1F3864"),
                             spaceAfter=8, spaceBefore=14),
        "h2": ParagraphStyle("h2", parent=base["Heading2"],
                             fontName="Helvetica-Bold", fontSize=12,
                             textColor=colors.HexColor("#2F5496"),
                             spaceAfter=4, spaceBefore=8),
        "body": ParagraphStyle("body", parent=base["BodyText"],
                               fontName="Helvetica", fontSize=9,
                               textColor=colors.black, leading=12),
        "mono": ParagraphStyle("mono", parent=base["BodyText"],
                               fontName="Courier", fontSize=8,
                               textColor=colors.HexColor("#333333"),
                               leading=10),
        "cite": ParagraphStyle("cite", parent=base["BodyText"],
                               fontName="Helvetica-Oblique", fontSize=8,
                               textColor=colors.HexColor("#555555")),
        "cover_meta": ParagraphStyle("cover_meta", parent=base["BodyText"],
                                     fontName="Helvetica", fontSize=10,
                                     leading=14),
    }
    return s


_TABLE_STYLE = TableStyle([
    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F3864")),
    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
    ("FONTSIZE", (0, 0), (-1, -1), 8),
    ("ALIGN", (0, 0), (-1, -1), "LEFT"),
    ("VALIGN", (0, 0), (-1, -1), "TOP"),
    ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#BBBBBB")),
    ("ROWBACKGROUNDS", (0, 1), (-1, -1),
     [colors.white, colors.HexColor("#F4F6FA")]),
])


# ─── Section builders ─────────────────────────────────────────────────────────


def _cover(wb, xlsx_path: Path, meta: dict, styles) -> list:
    out: list = []
    title = wb["Cover"].cell(row=1, column=1).value or xlsx_path.stem
    out.append(Paragraph(_esc(title), styles["title"]))
    out.append(HRFlowable(width="100%", thickness=1.5,
                          color=colors.HexColor("#1F3864")))
    out.append(Spacer(1, 6 * mm))

    # Pull metadata from the Cover sheet's key/value rows (col A label, col C value)
    cov = wb["Cover"]
    kv: list[tuple[str, str]] = []
    for r in range(4, 30):
        label = cov.cell(row=r, column=1).value
        value = cov.cell(row=r, column=3).value
        if label and value is not None and not str(label).startswith("Scenario control"):
            kv.append((str(label), str(value)))
    rows = [[Paragraph("<b>Field</b>", styles["body"]),
             Paragraph("<b>Value</b>", styles["body"])]]
    for k, v in kv[:15]:
        rows.append([Paragraph(_esc(k), styles["body"]),
                     Paragraph(_esc(v), styles["body"])])
    tbl = Table(rows, colWidths=[5 * cm, 12 * cm])
    tbl.setStyle(_TABLE_STYLE)
    out.append(tbl)
    out.append(Spacer(1, 6 * mm))

    # Reproducibility block
    out.append(Paragraph("Reproducibility", styles["h2"]))
    repro_rows = [[Paragraph("<b>Field</b>", styles["body"]),
                   Paragraph("<b>Value</b>", styles["body"])]]
    for k, v in meta.items():
        repro_rows.append([
            Paragraph(_esc(k), styles["body"]),
            Paragraph(_esc(v), styles["mono"] if k == "Spec SHA-256" else styles["body"]),
        ])
    if len(repro_rows) > 1:
        t = Table(repro_rows, colWidths=[5 * cm, 12 * cm])
        t.setStyle(_TABLE_STYLE)
        out.append(t)

    out.append(Spacer(1, 10 * mm))
    out.append(Paragraph(
        "This dossier embeds the spec SHA-256 and ModelForge build "
        "metadata. Any change to the spec produces a different hash. "
        "To verify a workbook matches this dossier, run "
        "<i>modelforge verify &lt;xlsx&gt; --spec &lt;yaml&gt;</i>.",
        styles["cite"],
    ))
    return out


def _executive_summary(wb, styles) -> list:
    out: list = [Paragraph("Executive Summary", styles["h1"])]
    # Best-effort: pull deliverable + target + sector + primary output
    cov = wb["Cover"]
    summary_bits: list[str] = []
    for r in range(4, 20):
        label = cov.cell(row=r, column=1).value
        value = cov.cell(row=r, column=3).value
        if not label:
            continue
        key = str(label).lower()
        if key in ("target", "sector", "country", "currency", "analyst",
                   "valuation_date", "version", "status"):
            summary_bits.append(f"<b>{_esc(label)}:</b> {_esc(value)}")
    out.append(Paragraph(" &nbsp;·&nbsp; ".join(summary_bits), styles["body"]))
    out.append(Spacer(1, 4 * mm))

    # Primary output (if registered)
    if "primary_output" in wb.defined_names:
        attr = wb.defined_names["primary_output"].attr_text
        out.append(Paragraph(
            f"<b>Primary output named range:</b> "
            f"<font face='Courier'>{_esc(attr)}</font>",
            styles["body"],
        ))

    out.append(Spacer(1, 4 * mm))
    out.append(Paragraph(
        "Every number in the attached workbook is a live formula whose "
        "inputs are named ranges on the Assumptions sheet. Every "
        "assumption cites either a source document (S-###) or an "
        "analyst rationale (A-###). This dossier reproduces the full "
        "trail for auditor and rating-agency review.",
        styles["body"],
    ))
    return out


def _assumptions_register(wb, styles) -> list:
    if "Assumptions" not in wb.sheetnames:
        return []
    ws = wb["Assumptions"]
    out: list = [Paragraph("Assumptions Register", styles["h1"])]
    out.append(Paragraph(
        "Every driver on the Assumptions sheet, its base value, "
        "rationale, confidence (H/M/L), and cited source. The "
        "workbook's Active column (col I) is a CHOOSE(scenario_index, "
        "worst, base, best) formula — every downstream cell reads it.",
        styles["body"],
    ))
    out.append(Spacer(1, 3 * mm))

    rows = [[Paragraph(f"<b>{h}</b>", styles["body"]) for h in
             ("ID", "Driver", "Unit", "Worst", "Base", "Best",
              "Conf.", "Source", "Rationale")]]
    for r in range(6, ws.max_row + 1):
        vid = ws.cell(row=r, column=1).value
        if not vid:
            continue
        rows.append([
            Paragraph(_esc(vid), styles["mono"]),
            Paragraph(_esc(ws.cell(row=r, column=2).value), styles["mono"]),
            Paragraph(_esc(ws.cell(row=r, column=5).value), styles["body"]),
            Paragraph(_esc(_fmt_num(ws.cell(row=r, column=6).value)), styles["body"]),
            Paragraph(_esc(_fmt_num(ws.cell(row=r, column=7).value)), styles["body"]),
            Paragraph(_esc(_fmt_num(ws.cell(row=r, column=8).value)), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=11).value), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=12).value), styles["mono"]),
            Paragraph(_esc(ws.cell(row=r, column=10).value), styles["cite"]),
        ])
    tbl = Table(rows, colWidths=[1.3 * cm, 3.5 * cm, 1.0 * cm, 1.6 * cm, 1.6 * cm,
                                 1.6 * cm, 0.9 * cm, 1.3 * cm, 4.0 * cm],
                repeatRows=1)
    tbl.setStyle(_TABLE_STYLE)
    out.append(tbl)
    return out


def _fmt_num(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (int, float)):
        if abs(v) < 1 and v != 0:
            return f"{v:.4f}"
        return f"{v:,.2f}"
    return str(v)


def _source_registry(wb, styles) -> list:
    if "Sources" not in wb.sheetnames:
        return []
    ws = wb["Sources"]
    out: list = [Paragraph("Source Registry", styles["h1"])]
    out.append(Paragraph(
        "Every source document referenced by Assumptions. The verified "
        "flag indicates whether the analyst cross-checked the value to "
        "the specific page cited.",
        styles["body"],
    ))
    out.append(Spacer(1, 3 * mm))
    rows = [[Paragraph(f"<b>{h}</b>", styles["body"]) for h in
             ("ID", "Document", "Page", "Publisher", "Date",
              "Verified", "URL / Note")]]
    for r in range(6, ws.max_row + 1):
        vid = ws.cell(row=r, column=1).value
        if not vid:
            continue
        rows.append([
            Paragraph(_esc(vid), styles["mono"]),
            Paragraph(_esc(ws.cell(row=r, column=2).value), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=3).value), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=4).value), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=5).value), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=7).value), styles["body"]),
            Paragraph(_esc(ws.cell(row=r, column=6).value), styles["cite"]),
        ])
    tbl = Table(rows, colWidths=[1.3 * cm, 4.2 * cm, 0.9 * cm, 2.8 * cm,
                                 1.8 * cm, 1.2 * cm, 4.4 * cm], repeatRows=1)
    tbl.setStyle(_TABLE_STYLE)
    out.append(tbl)
    return out


def _formula_inventory(wb, styles) -> list:
    out: list = [Paragraph("Formula Inventory", styles["h1"])]
    out.append(Paragraph(
        "Unique formula shapes per sheet (named ranges redacted to "
        "driver placeholders). Sampled up to 20 patterns per sheet.",
        styles["body"],
    ))
    out.append(Spacer(1, 3 * mm))
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        patterns: Counter = Counter()
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    # Redact numbers and sheet-qualified cell refs to cluster formulas
                    p = re.sub(r"[A-Z]+\d+", "CELL", v)
                    p = re.sub(r"\d+(\.\d+)?", "N", p)
                    patterns[p[:120]] += 1
        if not patterns:
            continue
        out.append(Paragraph(_esc(sheet), styles["h2"]))
        rows = [[Paragraph("<b>Pattern</b>", styles["body"]),
                 Paragraph("<b>Count</b>", styles["body"])]]
        for p, n in patterns.most_common(20):
            rows.append([Paragraph(_esc(p), styles["mono"]),
                         Paragraph(str(n), styles["body"])])
        tbl = Table(rows, colWidths=[14.5 * cm, 2.5 * cm], repeatRows=1)
        tbl.setStyle(_TABLE_STYLE)
        out.append(tbl)
        out.append(Spacer(1, 3 * mm))
    return out


def _lineage_graph_summary(graph_db: Optional[Path], styles) -> list:
    out: list = [Paragraph("Lineage Graph Summary", styles["h1"])]
    out.append(Paragraph(
        "Node / edge counts by kind from the SQLite linkage graph. "
        "Each CELL links back to a DRIVER or SOURCE. Walk any cell "
        "with <i>modelforge lineage &lt;graph.db&gt; CELL:SHEET!REF</i>.",
        styles["body"],
    ))
    out.append(Spacer(1, 3 * mm))
    if graph_db is None or not graph_db.exists():
        out.append(Paragraph(
            "<i>(graph DB not found — lineage data unavailable)</i>",
            styles["cite"],
        ))
        return out
    with sqlite3.connect(graph_db) as conn:
        node_rows = conn.execute(
            "SELECT kind, COUNT(*) FROM nodes GROUP BY kind ORDER BY 2 DESC"
        ).fetchall()
        edge_rows = conn.execute(
            "SELECT kind, COUNT(*) FROM edges GROUP BY kind ORDER BY 2 DESC"
        ).fetchall()
    rows = [[Paragraph("<b>Kind</b>", styles["body"]),
             Paragraph("<b>Nodes</b>", styles["body"])]]
    for k, n in node_rows:
        rows.append([Paragraph(_esc(k), styles["body"]),
                     Paragraph(str(n), styles["body"])])
    rows.append([Paragraph("<i>Edges by kind:</i>", styles["body"]),
                 Paragraph("", styles["body"])])
    for k, n in edge_rows:
        rows.append([Paragraph(_esc(k), styles["body"]),
                     Paragraph(str(n), styles["body"])])
    tbl = Table(rows, colWidths=[6 * cm, 11 * cm])
    tbl.setStyle(_TABLE_STYLE)
    out.append(tbl)
    return out


def _qc_signoff(xlsx_path: Path, styles) -> list:
    out: list = [Paragraph("QC Sign-off", styles["h1"])]
    report = run_qc(xlsx_path)
    rows = [[Paragraph("<b>Check</b>", styles["body"]),
             Paragraph("<b>Status</b>", styles["body"]),
             Paragraph("<b>Detail</b>", styles["body"])]]
    for c in report.checks:
        status = "PASS" if c.passed else "FAIL"
        color_tag = ("green" if c.passed else "red")
        rows.append([
            Paragraph(_esc(c.name), styles["body"]),
            Paragraph(f'<font color="{color_tag}"><b>{status}</b></font>',
                      styles["body"]),
            Paragraph(_esc(c.detail or ""), styles["body"]),
        ])
    tbl = Table(rows, colWidths=[8 * cm, 2 * cm, 7 * cm], repeatRows=1)
    tbl.setStyle(_TABLE_STYLE)
    out.append(tbl)
    out.append(Spacer(1, 4 * mm))
    out.append(Paragraph(
        f"<b>Result:</b> {report.n_pass}/{report.n_total} checks pass. "
        f"{'ALL CHECKS PASS' if report.all_pass else 'SOME FAILED'}.",
        styles["body"],
    ))
    out.append(Spacer(1, 10 * mm))

    # Sign-off block
    out.append(Paragraph("Sign-off", styles["h2"]))
    sign_rows = [
        ["Reviewer name", "________________________________"],
        ["Role / Firm", "________________________________"],
        ["Date", "________________________________"],
        ["Signature", "________________________________"],
    ]
    srows = []
    for label, val in sign_rows:
        srows.append([Paragraph(label, styles["body"]),
                      Paragraph(val, styles["mono"])])
    t = Table(srows, colWidths=[5 * cm, 11 * cm])
    t.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    out.append(t)
    return out


def _glossary(styles) -> list:
    out: list = [Paragraph("Glossary (EN / IT)", styles["h1"])]
    items = [
        ("Assumption (A-###)", "Analyst-driven input with rationale and confidence."),
        ("Source (S-###)", "Citable document (filename + page + publisher) referenced by Assumptions."),
        ("Active scenario", "CHOOSE(scenario_index, Worst, Base, Best). Toggle on Cover sheet."),
        ("Sign convention", "Costs negative (not parenthesis-positive). Every row format follows."),
        ("QC gate", "Eight external checks that a built workbook must pass before delivery."),
        ("Primary output", "Workbook-level named range pointing at the deal's key metric (Blended Lender IRR, Sponsor Equity IRR, etc.)."),
        ("Sensitivity tornado", "Factor-by-factor ±shock chart on primary_output."),
        ("Monte Carlo", "1000-run simulation with per-factor elasticities on primary_output."),
        ("Linkage graph", "SQLite DAG linking every cell ↔ driver ↔ source. Queryable with modelforge lineage."),
        ("FAST Standard", "Flexible / Appropriate / Structured / Transparent modelling conventions. ModelForge is FAST-compliant by design."),
        ("IFRS 9 §B5.4.1", "Effective Interest Rate — amortized-cost discounting rule cited on Returns/InvestorReturns."),
        ("Damodaran Italy ERP", "Country equity risk premium (6.7% base / 4.23% mature) used in WACC build on DCF."),
    ]
    rows = [[Paragraph("<b>Term</b>", styles["body"]),
             Paragraph("<b>Definition</b>", styles["body"])]]
    for k, v in items:
        rows.append([Paragraph(_esc(k), styles["mono"]),
                     Paragraph(_esc(v), styles["body"])])
    tbl = Table(rows, colWidths=[5 * cm, 12 * cm], repeatRows=1)
    tbl.setStyle(_TABLE_STYLE)
    out.append(tbl)
    return out


# ─── Public API ───────────────────────────────────────────────────────────────


def generate_dossier(
    xlsx_path: Path | str,
    output_pdf: Optional[Path | str] = None,
    graph_db: Optional[Path | str] = None,
) -> Path:
    """Generate an audit dossier PDF for a built workbook.

    Returns the PDF path. Raises FileNotFoundError if the workbook
    doesn't exist.
    """
    xlsx_path = Path(xlsx_path)
    if not xlsx_path.exists():
        raise FileNotFoundError(xlsx_path)

    if output_pdf is None:
        output_pdf = xlsx_path.with_suffix(".dossier.pdf")
    output_pdf = Path(output_pdf)

    if graph_db is None:
        gd = xlsx_path.with_suffix(".graph.db")
        graph_db = gd if gd.exists() else None
    else:
        graph_db = Path(graph_db)

    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    meta = read_reproducibility(xlsx_path)
    styles = _styles()

    story: list = []
    story += _cover(wb, xlsx_path, meta, styles)
    story.append(PageBreak())
    story += _executive_summary(wb, styles)
    story.append(Spacer(1, 4 * mm))
    story += _assumptions_register(wb, styles)
    story.append(PageBreak())
    story += _source_registry(wb, styles)
    story.append(PageBreak())
    story += _formula_inventory(wb, styles)
    story.append(PageBreak())
    story += _lineage_graph_summary(graph_db, styles)
    story.append(Spacer(1, 6 * mm))
    story += _qc_signoff(xlsx_path, styles)
    story.append(PageBreak())
    story += _glossary(styles)

    doc = SimpleDocTemplate(
        str(output_pdf), pagesize=A4,
        leftMargin=2 * cm, rightMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm,
        title=f"ModelForge dossier — {xlsx_path.stem}",
        author="ModelForge",
    )
    doc.build(story)
    return output_pdf


__all__ = ["generate_dossier"]
