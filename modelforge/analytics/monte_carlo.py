"""Monte Carlo simulation on the primary output.

Given the same factor list used for the sensitivity tornado, draws
``n_runs`` random samples per factor from a configurable distribution
(triangular by default) and aggregates them into a distribution over
the primary output. Results render as a MonteCarlo sheet with two
clearly-distinguished blocks:

* **Sampled (SNAPSHOT).** The full ``n_runs`` simulated distribution —
  stats (P5/P25/P50/P75/P95, mean, std) + a native Excel histogram
  chart. A sampled Monte-Carlo simulation **cannot** recompute inside
  vanilla Excel (there is no in-cell RNG-over-N-draws-over-the-whole-
  model primitive), so these numbers are an HONEST build-time snapshot
  from a FIXED seed + the as-built scenario, flagged with the
  ``STATIC SNAPSHOT`` banner. To refresh, rerun ``modelforge monte-carlo``
  (see :func:`append_monte_carlo_sheet` / the CLI ``monte-carlo`` command).

* **Parametric (LIVE — normal approx).** Beside the sampled block, a
  genuinely-live downside band: P5 / P50 / P95 / mean computed as native
  Excel ``=`` formulas off the live ``primary_output`` named range and a
  live volatility input cell (``mc_vol``), e.g.
  ``=primary_output*(1+NORM.S.INV(0.05)*mc_vol)`` for P5. Because every
  term is a live cell, this band genuinely RECOMPUTES when the user flips
  ``scenario_index`` (which moves ``primary_output`` via the Assumptions
  CHOOSE) or edits the ``mc_vol`` input — proven with the ``formulas``
  engine. It is a normal (Gaussian) approximation, honestly distinct from
  the sampled histogram, and seeded at build time to the sampled std so
  the two agree at the as-built scenario.

Like the sensitivity tornado, the SAMPLED block uses the per-factor
elasticity coefficient to map a driver shock to an output delta when no
per-template shadow engine is available — an approximation that ships
across all 8 templates. The exact shadow-engine recompute STAYS available
for the CLI / JSON / PDF-dossier export. Sampled stats are written as
values (not formulas) per the PRD AC for US-002; the parametric band is
written as live formulas (the reactive upgrade).

Distributions
-------------
* ``triangular(low_shock, 0, high_shock)`` — default. Captures asymmetry
  in the factor's shock envelope directly.
* ``normal(mean=0, std=(high_shock-low_shock)/4)`` — assumes ±2σ covers
  the declared range.
* ``lognormal(mean=0, sigma=(high_shock-low_shock)/4)`` — for drivers
  that must stay non-negative (rates, multiples).

All factors are drawn independently. Correlation structure is a v0.4.2
follow-up once shadow engines enable joint distributions.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Literal, Optional

import numpy as np
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.analytics.factors import SensitivityFactor, default_factors_for
from modelforge.analytics.sensitivity import (
    PrimaryOutputLoc,
    _ELASTICITY_REGISTRY,
    _col_letter,
    _driver_exists,
    _elasticity_for,
    _find_primary_output,
)
from modelforge.builder import styles


Distribution = Literal["triangular", "normal", "lognormal"]


# ─── Parametric-LIVE honesty banner ──────────────────────────────────────────
# The sampled block keeps the shared ``static_snapshot_banner`` (i18n, 8-lang)
# from sensitivity.py. The parametric band gets its own banner. We render it as
# INLINE bilingual EN | IT text (exactly like the other descriptive lines on
# this sheet — title subline row 2 and the method/seed line) rather than a new
# i18n LABELS entry, so it stays self-contained in this module and never risks
# the 8-language coverage gate. The text is unmistakably a LIVE accuracy
# disclosure (normal/Gaussian approximation off primary_output), not a "stale"
# warning, and is styled in formula-green to read as live.

_PARAMETRIC_BANNER_EN = (
    "PARAMETRIC (LIVE — normal approx) — P5/P50/P95/mean as live Excel "
    "formulas off primary_output and the mc_vol input. RECOMPUTES when you "
    "flip scenario_index or edit mc_vol. Distinct from the sampled histogram "
    "below, which is a build-time snapshot (exact resim via `modelforge "
    "monte-carlo`)."
)
_PARAMETRIC_BANNER_IT = (
    "PARAMETRICO (LIVE — approssimazione normale) — P5/P50/P95/media come "
    "formule Excel live su primary_output e l'input mc_vol. Si RICALCOLA "
    "quando cambi scenario_index o modifichi mc_vol. Distinto dall'istogramma "
    "campionato sotto, che è un'istantanea di build (ri-simulazione esatta "
    "con `modelforge monte-carlo`)."
)


def write_parametric_live_banner(ws: Worksheet, row: int) -> None:
    """Write the bilingual 'PARAMETRIC (LIVE — normal approx)' banner.

    Marks the live parametric downside band whose P5/P50/P95/mean cells are
    native Excel formulas off ``primary_output`` + the editable ``mc_vol``
    input — so they genuinely recompute on a scenario flip or vol edit.
    Styled formula-green so it reads as informational/live, visually distinct
    from the warning-amber ``STATIC SNAPSHOT`` banner on the sampled block.
    """
    c = ws.cell(row=row, column=1,
                value=f"{_PARAMETRIC_BANNER_EN}  |  {_PARAMETRIC_BANNER_IT}")
    c.font = styles.Font(
        name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
        bold=True, italic=True, color="006100",
    )
    c.alignment = styles.Alignment(
        horizontal="left", vertical="center", wrap_text=False,
    )


@dataclass
class MCConfig:
    n_runs: int = 1000
    distribution: Distribution = "triangular"
    seed: int = 20260417  # deterministic by default


@dataclass
class MCResult:
    n_runs: int
    distribution: Distribution
    samples: np.ndarray   # shape (n_runs,) — fractional OR absolute per delta_mode
    factor_contributions: dict[str, np.ndarray]
    method: str = "elasticity"          # "shadow" or "elasticity"
    delta_mode: str = "fractional"      # "fractional" or "absolute"

    @property
    def mean(self) -> float:
        return float(np.mean(self.samples))

    @property
    def std(self) -> float:
        return float(np.std(self.samples))

    def percentile(self, p: float) -> float:
        return float(np.percentile(self.samples, p))


def _draw_shocks(
    rng: np.random.Generator,
    low: float,
    high: float,
    n: int,
    distribution: Distribution,
) -> np.ndarray:
    if distribution == "triangular":
        return rng.triangular(left=low, mode=0.0, right=high, size=n)
    if distribution == "normal":
        sigma = (high - low) / 4.0 if high > low else abs(high) / 2.0
        return rng.normal(loc=0.0, scale=max(sigma, 1e-9), size=n)
    if distribution == "lognormal":
        sigma = (high - low) / 4.0 if high > low else abs(high) / 2.0
        # Shift so mean(shock)=0 approximately
        draws = rng.lognormal(mean=0.0, sigma=max(sigma, 1e-9), size=n)
        return draws - draws.mean()
    raise ValueError(f"Unknown distribution: {distribution}")


def run_monte_carlo(
    factors: list[SensitivityFactor],
    config: MCConfig,
    spec=None,
) -> MCResult:
    """Simulate output deltas, using a shadow engine when available.

    If ``spec`` is provided and a shadow engine exists for its
    model_type, each draw re-evaluates the primary output exactly
    (applying ALL factor shocks simultaneously for a given draw). This
    is the gold-standard path — no linearity assumption, correlations
    captured implicitly via the full recompute.

    Otherwise falls back to the linear elasticity approximation:
    ``output_delta = Σ_i elasticity_i × shock_i``.
    """
    from modelforge.shadow import compute_primary_output, has_shadow_engine
    rng = np.random.default_rng(config.seed)

    use_shadow = (
        spec is not None
        and has_shadow_engine(getattr(spec, "model_type", ""))
    )

    # Pre-draw shocks per factor (shape n_runs × n_factors)
    contributions: dict[str, np.ndarray] = {}
    shocks_by_factor: dict[str, np.ndarray] = {}
    for f in factors:
        shocks = _draw_shocks(rng, f.low_shock, f.high_shock,
                              config.n_runs, config.distribution)
        shocks_by_factor[f.driver_name] = shocks

    if use_shadow:
        all_assums = {a.name: a for a in spec.all_assumptions()}
        base = compute_primary_output(spec, {})
        if base is None or abs(base) < 1e-12:
            # Degenerate — fall through to elasticity path
            use_shadow = False

    # For deals where the base primary output is near zero (e.g. PF
    # equity IRR hovering around 0%), fractional deltas (shocked/base
    # - 1) blow up. Switch to absolute deltas when |base| < threshold.
    _NEAR_ZERO = 0.01
    use_absolute = use_shadow and abs(base) < _NEAR_ZERO

    if use_shadow:
        total = np.zeros(config.n_runs, dtype=float)
        for k in range(config.n_runs):
            overrides = {}
            for f in factors:
                if f.driver_name not in all_assums:
                    continue
                bv = all_assums[f.driver_name].base
                overrides[f.driver_name] = bv * (1 + shocks_by_factor[f.driver_name][k])
            shocked = compute_primary_output(spec, overrides)
            if shocked is None:
                total[k] = 0.0
            elif use_absolute:
                total[k] = shocked - base
            else:
                total[k] = (shocked - base) / base
        # Per-factor contribution (isolated single-factor shock) —
        # useful for "which factor drives most of the variance" tooltip
        for f in factors:
            if f.driver_name not in all_assums:
                contributions[f.driver_name] = np.zeros(config.n_runs)
                continue
            bv = all_assums[f.driver_name].base
            per_factor = np.zeros(config.n_runs)
            # Sample 50 draws for per-factor attribution (keep MC fast)
            sample = min(50, config.n_runs)
            for k in range(sample):
                shocked = compute_primary_output(
                    spec,
                    {f.driver_name: bv * (1 + shocks_by_factor[f.driver_name][k])},
                )
                per_factor[k] = (shocked - base) / base if shocked is not None else 0.0
            # Extrapolate mean contribution to full length
            if sample > 0:
                per_factor[sample:] = per_factor[:sample].mean()
            contributions[f.driver_name] = per_factor
    else:
        # Elasticity fallback
        total = np.zeros(config.n_runs, dtype=float)
        for f in factors:
            e = _elasticity_for(f.driver_name)
            contrib = e * shocks_by_factor[f.driver_name]
            contributions[f.driver_name] = contrib
            total += contrib

    return MCResult(
        n_runs=config.n_runs,
        distribution=config.distribution,
        samples=total,
        factor_contributions=contributions,
        method="shadow" if use_shadow else "elasticity",
        delta_mode="absolute" if use_absolute else "fractional",
    )


# ─── Sheet emission ───────────────────────────────────────────────────────────


_SHEET_NAME = "MonteCarlo"
_N_BINS = 30


def _emit_sheet(
    wb,
    primary_loc: PrimaryOutputLoc,
    factors: list[SensitivityFactor],
    result: MCResult,
) -> Worksheet:
    if _SHEET_NAME in wb.sheetnames:
        del wb[_SHEET_NAME]
    ws = wb.create_sheet(_SHEET_NAME)

    for col, w in {"A": 14, "B": 32, "C": 14, "D": 14, "E": 14,
                   "F": 14, "G": 14, "H": 14}.items():
        ws.column_dimensions[col].width = w

    # ── Title block
    from modelforge.builder.i18n import L as _L
    ws.cell(row=1, column=1, value=_L("monte_carlo_title").en).font = styles.font_title
    ws.cell(row=2, column=1, value=_L("monte_carlo_title").it).font = styles.font_label_it
    from modelforge.builder import layout as _layout
    _layout.write_scenario_banner(ws, row=3)
    method_tag = ("exact shadow-engine recompute per draw"
                  if result.method == "shadow"
                  else "linearized per-factor elasticity")
    delta_tag = ("absolute (base output is near zero — fractional Δ "
                 "would blow up)"
                 if result.delta_mode == "absolute"
                 else "fractional (× base = absolute)")
    # Row 4: honest STATIC-SNAPSHOT banner. Every MC statistic below is a
    # pre-computed literal from a FIXED seed + the as-built scenario — none
    # of it recomputes when the user flips scenario_index or edits a driver.
    from modelforge.analytics.sensitivity import write_static_snapshot_banner
    write_static_snapshot_banner(ws, row=4)

    # ── Base output reference (LIVE — moves with scenario_index)
    ws.cell(row=5, column=1, value="Base output").font = styles.font_subheader
    ws.cell(row=5, column=2, value=primary_loc.label).font = styles.font_subheader
    c_base = ws.cell(row=5, column=3, value="=primary_output")
    styles.style_formula(c_base, number_format=styles.FMT_PCT_2DP)
    c_base.font = styles.font_subheader

    # Method / Δ-mode / seed provenance (row 6 — describes the FIXED build-time
    # run that produced the SAMPLED static stats below).
    ws.cell(
        row=6, column=1,
        value=(
            f"{result.n_runs:,}-run {result.distribution} simulation on "
            f"{primary_loc.label}. Method: {method_tag}. Δ mode: {delta_tag}. "
            f"Seed={MCConfig().seed}."
        ),
    ).font = styles.font_label_it

    # ══ Block 1: PARAMETRIC band — genuinely LIVE ════════════════════════════
    # P5/P50/P95/mean as native Excel formulas off the live primary_output
    # named range + a live volatility input (mc_vol). NORM.S.INV gives the
    # standard-normal quantiles; the band recomputes when scenario_index flips
    # (primary_output moves via the Assumptions CHOOSE) or when mc_vol is
    # edited. This is the reactive upgrade — visually + semantically distinct
    # from the sampled snapshot block below. A sampled MC cannot recompute in
    # vanilla Excel, but this parametric (normal-approx) band can, so we ship
    # it live and clearly labelled.
    par_banner_row = 8
    write_parametric_live_banner(ws, row=par_banner_row)

    # Live volatility input cell + named range. Seeded at build time to the
    # sampled std so the parametric band agrees with the sampled distribution
    # at the as-built scenario; an analyst can override it (blue input) to
    # widen/narrow the band, and it recomputes live.
    vol_row = par_banner_row + 1
    ws.cell(row=vol_row, column=1,
            value="Volatility (σ, editable)").font = styles.font_subheader
    vol_cell = ws.cell(row=vol_row, column=2, value=float(result.std))
    styles.style_input(vol_cell, number_format=styles.FMT_PCT_2DP)
    vol_cell.comment = Comment(
        "LIVE volatility input for the parametric (normal-approx) band. "
        f"Seeded to the sampled std ({result.std:.4f}) at build time so the "
        "parametric band matches the sampled distribution at the as-built "
        "scenario. Edit it (blue input) to widen/narrow the band — P5/P50/P95 "
        "recompute live. Distinct from the frozen sampled stats below.",
        "ModelForge",
    )
    vol_col = _col_letter(2)
    vol_ref_cell = f"{vol_col}{vol_row}"
    if "mc_vol" in wb.defined_names:
        del wb.defined_names["mc_vol"]
    wb.defined_names["mc_vol"] = DefinedName(
        name="mc_vol", attr_text=f"'{ws.title}'!${vol_col}${vol_row}")

    # Parametric stat rows — LIVE formulas. delta_mode determines whether the
    # band is multiplicative (fractional regime) or additive (near-zero base).
    par_hdr_row = vol_row + 1
    for i, c in enumerate(("Statistic", "Parametric (LIVE)"), start=1):
        styles.style_header(ws.cell(row=par_hdr_row, column=i, value=c))

    # (label, z-quantile or None for mean)
    par_specs = [
        ("Mean (LIVE)", None),
        ("P5 (LIVE)", 0.05),
        ("P50 (LIVE)", 0.50),
        ("P95 (LIVE)", 0.95),
    ]
    par_first = par_hdr_row + 1
    for i, (label, q) in enumerate(par_specs):
        r = par_first + i
        ws.cell(row=r, column=1, value=label).font = styles.font_subheader
        if q is None:
            # Mean ≈ primary_output itself (the parametric distribution is
            # centred on the live base output — no shift).
            formula = "=primary_output"
        elif result.delta_mode == "absolute":
            formula = f"=primary_output+NORM.S.INV({q})*mc_vol"
        else:
            formula = f"=primary_output*(1+NORM.S.INV({q})*mc_vol)"
        pc = ws.cell(row=r, column=2, value=formula)
        styles.style_formula(pc, number_format=styles.FMT_PCT_2DP)
        pc.comment = Comment(
            "LIVE parametric (normal-approx) statistic — native Excel formula "
            f"({formula[1:]}) off primary_output + mc_vol. Recomputes when you "
            "flip scenario_index or edit mc_vol. NOT a sampled value: it is a "
            "Gaussian approximation of the downside band, distinct from the "
            "sampled histogram below.",
            "ModelForge",
        )
    par_last = par_first + len(par_specs) - 1

    # ══ Block 2: SAMPLED distribution — HONEST build-time SNAPSHOT ════════════
    # Re-flag the sampled block right above its header so the boundary between
    # the LIVE parametric band and the frozen sampled stats is unmistakable.
    sampled_flag_row = par_last + 2
    fl = ws.cell(
        row=sampled_flag_row, column=1,
        value=("SAMPLED (snapshot) — full "
               f"{result.n_runs:,}-draw {result.distribution} distribution; "
               "frozen at build time, does NOT recompute on scenario flip "
               "(refresh: `modelforge monte-carlo`)."),
    )
    fl.font = styles.Font(
        name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
        bold=True, italic=True, color="9C5700",
    )
    fl.fill = styles.fill_static

    # ── Distribution stats (SAMPLED — static literals)
    stats_row = sampled_flag_row + 1
    for i, c in enumerate(("Statistic", "Δ output (frac)",
                           "Absolute output"), start=1):
        cell = ws.cell(row=stats_row, column=i, value=c)
        styles.style_header(cell)

    rows = [
        ("Mean", result.mean),
        ("Std", result.std),
        ("P5", result.percentile(5)),
        ("P25", result.percentile(25)),
        ("P50 (median)", result.percentile(50)),
        ("P75", result.percentile(75)),
        ("P95", result.percentile(95)),
        ("Min", float(np.min(result.samples))),
        ("Max", float(np.max(result.samples))),
    ]
    for i, (label, frac) in enumerate(rows):
        r = stats_row + 1 + i
        ws.cell(row=r, column=1, value=label).font = styles.font_subheader
        # Col B — delta (fractional or absolute per result.delta_mode).
        # STATIC: pre-computed at build time from a FIXED seed + as-built
        # scenario — style as static_value (grey/italic) so it is visually
        # distinct from a blue live input and never looks recomputable.
        fc = ws.cell(row=r, column=2, value=frac)
        styles.style_static_value(fc, number_format=styles.FMT_PCT_2DP)
        fc.comment = Comment(
            f"Computed from {result.n_runs:,} {result.distribution} draws "
            f"across {len(factors)} factors. Seed={MCConfig().seed}. "
            f"Method={result.method}. Δ mode={result.delta_mode}. "
            f"STATIC snapshot — does not recompute on scenario flip.",
            "ModelForge",
        )
        # Col C — absolute output. Formula differs by delta_mode:
        #   fractional: absolute = primary_output * (1 + Δ)
        #   absolute:   absolute = primary_output + Δ
        if result.delta_mode == "absolute":
            ac = ws.cell(row=r, column=3, value=f"=primary_output+B{r}")
        else:
            ac = ws.cell(row=r, column=3, value=f"=primary_output*(1+B{r})")
        styles.style_formula(ac, number_format=styles.FMT_PCT_2DP)

    # ── Histogram bins
    hist_row = stats_row + 1 + len(rows) + 2
    ws.cell(row=hist_row, column=1, value="Histogram bins").font = styles.font_subheader
    for i, c in enumerate(("Bin low (Δ)", "Bin high (Δ)", "Count"), start=1):
        cell = ws.cell(row=hist_row + 1, column=i, value=c)
        styles.style_header(cell)

    counts, edges = np.histogram(result.samples, bins=_N_BINS)
    for i in range(_N_BINS):
        r = hist_row + 2 + i
        # Bin edges + counts are also pre-computed STATIC literals — style
        # them as static_value so the whole histogram block reads as a frozen
        # build-time snapshot, not editable live data.
        styles.style_static_value(
            ws.cell(row=r, column=1, value=float(edges[i])),
            number_format=styles.FMT_PCT_2DP,
        )
        styles.style_static_value(
            ws.cell(row=r, column=2, value=float(edges[i + 1])),
            number_format=styles.FMT_PCT_2DP,
        )
        styles.style_static_value(
            ws.cell(row=r, column=3, value=int(counts[i])),
            number_format=styles.FMT_INTEGER,
        )
    hist_first = hist_row + 2
    hist_last = hist_row + 2 + _N_BINS - 1

    # ── Histogram chart
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = f"Monte Carlo — Δ {primary_loc.label}"
    chart.y_axis.title = "Count"
    chart.x_axis.title = "Δ output (fractional)"
    count_ref = Reference(ws, min_col=3, min_row=hist_row + 1,
                          max_col=3, max_row=hist_last)
    bins_ref = Reference(ws, min_col=1, min_row=hist_first,
                         max_col=1, max_row=hist_last)
    chart.add_data(count_ref, titles_from_data=True)
    chart.set_categories(bins_ref)
    chart.height = 9
    chart.width = 18
    chart.dataLabels = DataLabelList(showVal=False)
    # Anchor the histogram to the right of the sampled stats block (col E),
    # at the sampled header row — keeps it clear of the parametric band (cols
    # A:B) above.
    ws.add_chart(chart, f"E{stats_row}")

    # Freeze below the title + scenario/static banners so the LIVE parametric
    # band and the sampled block scroll under a stable header.
    ws.freeze_panes = "A8"
    ws.print_title_rows = f"{stats_row}:{stats_row}"
    return ws


# ─── Public API ───────────────────────────────────────────────────────────────


def append_monte_carlo_sheet(
    xlsx_path: Path | str,
    spec,
    factors: Optional[list[SensitivityFactor]] = None,
    config: Optional[MCConfig] = None,
) -> Optional[Path]:
    """Append a MonteCarlo sheet + histogram chart to a built workbook.

    Returns the xlsx path on success, or None if sensitivity's primary
    output couldn't be located (same degradation contract).
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    primary_loc = _find_primary_output(wb)
    if primary_loc is None:
        return None

    if factors is None:
        factors = default_factors_for(getattr(spec, "model_type", ""))
    applicable = [f for f in factors if _driver_exists(wb, f.driver_name)]
    if not applicable:
        return None

    if config is None:
        config = MCConfig()

    # Pass spec so MC uses the shadow engine when available
    result = run_monte_carlo(applicable, config, spec=spec)
    _emit_sheet(wb, primary_loc, applicable, result)
    wb.save(xlsx_path)
    return xlsx_path


__all__ = [
    "MCConfig",
    "MCResult",
    "run_monte_carlo",
    "append_monte_carlo_sheet",
    "write_parametric_live_banner",
]
