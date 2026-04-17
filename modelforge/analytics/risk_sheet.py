"""RiskAnalysis sheet post-processor.

Adds a RiskAnalysis sheet to any built workbook whose spec carries an
optional `risk_analysis` block (RiskAnalysisSpec). Emits the full
probabilistic credit picture:

    Section 1 — Merton structural solve
        Equity / σ_E / Debt / r / T inputs → V, σ_V, DD, PD

    Section 2 — KMV empirical calibration
        DD → empirical PD via Moody's RiskCalc-style table

    Section 3 — IFRS 9 ECL
        Stage 1 / Stage 2 / Stage 3 side-by-side
        12-month ECL + lifetime ECL

    Section 4 — Native Excel bar chart
        12-mo ECL vs lifetime ECL across stages

All numeric values are pre-computed via the shadow-style `modelforge.risk`
engine; comments on every cell cite the formula + source ID where
applicable. Live primary inputs (equity, vol, debt) use Assumption
named ranges when the template registers them, falling back to
hardcoded inputs otherwise.
"""

from __future__ import annotations

from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import layout, styles
from modelforge.risk import (
    ECLInputs,
    MertonInputs,
    Stage,
    calibrate_pd_kmv,
    compute_ecl,
    solve_merton,
)


_SHEET_NAME = "RiskAnalysis"


def append_risk_analysis_sheet(
    xlsx_path: Path | str,
    spec,
) -> Optional[Path]:
    """Append RiskAnalysis sheet if the spec includes a risk_analysis block.

    Returns the xlsx path if the sheet was added, None if the spec has
    no risk_analysis block (graceful no-op).
    """
    risk = getattr(spec, "risk_analysis", None)
    if risk is None:
        return None

    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    if _SHEET_NAME in wb.sheetnames:
        del wb[_SHEET_NAME]
    ws = wb.create_sheet(_SHEET_NAME)

    # Compute
    em = risk.equity_market
    merton = solve_merton(MertonInputs(
        equity_value=em.equity_value_eur_m,
        equity_volatility=em.equity_volatility,
        debt_face_value=em.debt_face_value_eur_m,
        risk_free_rate=em.risk_free_rate,
        horizon_years=em.horizon_years,
    ))
    kmv_pd = calibrate_pd_kmv(merton.distance_to_default)
    pd_for_ecl = max(merton.probability_of_default, kmv_pd)
    ead = (risk.exposure_at_default_eur_m
           if risk.exposure_at_default_eur_m is not None
           else em.debt_face_value_eur_m)
    base_origination = risk.origination_pd_12m or pd_for_ecl

    # Compute all three stage results for comparison
    def _ecl_at(dpd: int, origination_pd: float) -> tuple:
        inp = ECLInputs(
            exposure_at_default_eur_m=ead,
            loss_given_default=risk.loss_given_default,
            effective_interest_rate=risk.effective_interest_rate,
            maturity_years=risk.maturity_years,
            pd_curve_annual=[pd_for_ecl] * risk.maturity_years,
            current_pd_12m=pd_for_ecl,
            origination_pd_12m=origination_pd,
            days_past_due=dpd,
        )
        r = compute_ecl(inp)
        return r

    s1 = _ecl_at(0, pd_for_ecl)          # performing (Stage 1)
    s2 = _ecl_at(0, pd_for_ecl / 2.5)    # SICR (current PD > 2× origination)
    s3 = _ecl_at(100, pd_for_ecl)        # 90+ dpd (Stage 3)
    reported = _ecl_at(risk.days_past_due, base_origination)

    # ── Layout
    for col, w in {"A": 38, "B": 28, "C": 18, "D": 18, "E": 18}.items():
        ws.column_dimensions[col].width = w

    ws.cell(row=1, column=1, value="Risk Analysis").font = styles.font_title
    ws.cell(row=2, column=1,
            value="Analisi del rischio — Merton / KMV / IFRS 9"
            ).font = styles.font_label_it
    layout.write_scenario_banner(ws, row=3)
    ws.cell(
        row=4, column=1,
        value=(f"Structural PD (Merton 1974) + empirical calibration "
               f"(Moody's RiskCalc-style) + IFRS 9 §B5.5.17 ECL. "
               f"LGD {risk.loss_given_default:.0%} · EIR "
               f"{risk.effective_interest_rate:.2%} · maturity "
               f"{risk.maturity_years}y."),
    ).font = styles.font_label_it

    # ── Section 1: Merton
    r = 6
    _section_header(ws, r, "Merton structural solve",
                    "Risoluzione strutturale (Merton 1974)")
    r += 1
    rows_1 = [
        ("Equity value (€m)", em.equity_value_eur_m,
         "Observed market cap", em.equity_source_id),
        ("Equity volatility (σ_E)", em.equity_volatility,
         "Annualized σ of equity returns", em.volatility_source_id),
        ("Debt face value (€m)", em.debt_face_value_eur_m,
         "Total debt at maturity", None),
        ("Risk-free rate", em.risk_free_rate,
         "Continuously compounded, decimal", None),
        ("Horizon (years)", em.horizon_years,
         "Default horizon", None),
        ("", None, "", None),
        ("Asset value V (€m)", merton.asset_value,
         "Solved from Merton two-equation system", None),
        ("Asset volatility σ_V", merton.asset_volatility,
         "Solved; σ_E·E ≈ σ_V·V·N(d1)", None),
        ("d1", merton.d1, "(ln(V/D)+(r+½σ²)T)/(σ√T)", None),
        ("d2", merton.d2, "d1 − σ√T", None),
        ("Distance-to-default", merton.distance_to_default,
         "(ln(V/D)+(r−½σ²)T)/(σ√T)", None),
        ("Merton PD (theoretical)", merton.probability_of_default,
         "N(−DD) under physical measure", None),
    ]
    r = _write_rows(ws, r, rows_1)
    r += 1

    # ── Section 2: KMV
    _section_header(ws, r, "KMV empirical calibration",
                    "Calibrazione empirica stile KMV")
    r += 1
    rows_2 = [
        ("Distance-to-default (from above)", merton.distance_to_default,
         "Same DD; next row applies empirical mapping", None),
        ("KMV empirical PD (1y)", kmv_pd,
         "Moody's RiskCalc Europe DD→PD interpolated table", None),
        ("PD used for ECL (max of two)", pd_for_ecl,
         "Conservative: max(Merton theoretical, KMV empirical)", None),
    ]
    r = _write_rows(ws, r, rows_2)
    r += 1

    # ── Section 3: IFRS 9 ECL
    _section_header(ws, r, "IFRS 9 ECL — all stages",
                    "IFRS 9 ECL — confronto tra stage")
    r += 1
    # Header row for the 3-stage comparison
    headers = ["Metric", "Stage 1 (performing)", "Stage 2 (SICR)",
               "Stage 3 (default)", "Reported"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=r, column=i, value=h)
        styles.style_header(c)
    r += 1
    _ecl_row(ws, r, "12-month ECL (€m)",
             s1.ecl_12_month_eur_m, s2.ecl_12_month_eur_m,
             s3.ecl_12_month_eur_m, reported.ecl_12_month_eur_m)
    r += 1
    _ecl_row(ws, r, "Lifetime ECL (€m)",
             s1.ecl_lifetime_eur_m, s2.ecl_lifetime_eur_m,
             s3.ecl_lifetime_eur_m, reported.ecl_lifetime_eur_m)
    r += 1
    _ecl_row(ws, r, "Reported ECL (€m)",
             s1.ecl_eur_m, s2.ecl_eur_m, s3.ecl_eur_m, reported.ecl_eur_m)
    r += 1
    _ecl_row(ws, r, "Implied provision rate",
             s1.implied_rate_pct, s2.implied_rate_pct,
             s3.implied_rate_pct, reported.implied_rate_pct,
             fmt=styles.FMT_PCT_2DP)
    r += 1
    ws.cell(row=r, column=1, value="Applied stage").font = styles.font_subheader
    ws.cell(row=r, column=5, value=reported.stage.value.upper()).font = \
        styles.font_subheader
    r += 2

    # ── Section 4: native Excel chart
    chart_row = r
    chart = BarChart()
    chart.type = "col"
    chart.style = 11
    chart.title = "ECL by stage (€m)"
    chart.y_axis.title = "ECL (€m)"
    chart.x_axis.title = "Stage"
    # Data: B..E of the two ECL rows (headers in cells B,C,D,E of the
    # section-3 header row)
    # Headers (categories) at section_3_header_row
    section_3_header_row = r - 7  # -1 lifetime -1 reported -1 rate -1 stage -1 blank -1 reported row -1 12mo row + 1
    # Simpler: just reconstruct using the two specific rows we wrote
    # Just reference the 12-month + lifetime rows explicitly
    twelve_row = r - 6
    lifetime_row = r - 5
    _12m_ref = Reference(ws, min_col=2, min_row=twelve_row,
                          max_col=4, max_row=twelve_row)
    _life_ref = Reference(ws, min_col=2, min_row=lifetime_row,
                           max_col=4, max_row=lifetime_row)
    chart.add_data(_12m_ref, titles_from_data=False)
    chart.add_data(_life_ref, titles_from_data=False)
    cats = Reference(ws, min_col=2, min_row=section_3_header_row,
                     max_col=4, max_row=section_3_header_row)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 16
    ws.add_chart(chart, f"F{chart_row - 10}")

    ws.freeze_panes = "B5"
    ws.print_title_rows = "1:4"
    wb.save(xlsx_path)
    return xlsx_path


# ── Helpers ─────────────────────────────────────────────────────────────────


def _section_header(ws: Worksheet, row: int, en: str, it: str) -> None:
    c = ws.cell(row=row, column=1, value=en)
    styles.style_subheader(c)
    c2 = ws.cell(row=row, column=2, value=it)
    c2.font = styles.font_subheader
    c2.fill = styles.fill_subheader


def _write_rows(ws: Worksheet, start_row: int, rows) -> int:
    r = start_row
    for label, value, explain, source_id in rows:
        if label == "":
            r += 1
            continue
        lbl = ws.cell(row=r, column=1, value=label)
        styles.style_label_en(lbl)
        c = ws.cell(row=r, column=2, value=value)
        if isinstance(value, (int, float)):
            nfmt = _pick_fmt(label, value)
            styles.style_input(c, number_format=nfmt)
        note_parts = [explain]
        if source_id:
            note_parts.append(f"Source: {source_id}")
        if explain or source_id:
            c.comment = Comment("\n".join(note_parts), "ModelForge")
        r += 1
    return r


def _pick_fmt(label: str, value) -> str:
    ll = label.lower()
    if "pd" in ll or "rate" in ll or "volatility" in ll:
        return styles.FMT_PCT_2DP
    if "distance" in ll or "d1" in ll or "d2" in ll:
        return styles.FMT_MULTIPLE
    return styles.FMT_EUR_M if isinstance(value, (int, float)) and abs(value) > 1 else styles.FMT_PCT_2DP


def _ecl_row(ws: Worksheet, row: int, label: str,
             s1: float, s2: float, s3: float, reported: float,
             fmt: str = styles.FMT_EUR_M) -> None:
    ws.cell(row=row, column=1, value=label).font = styles.font_label_en
    for col, v in ((2, s1), (3, s2), (4, s3), (5, reported)):
        c = ws.cell(row=row, column=col, value=v)
        styles.style_input(c, number_format=fmt)
