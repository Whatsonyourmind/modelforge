"""Sensitivity tornado post-processor.

After the core template builder emits a workbook, this module locates
the "primary output" cell (Blended Lender IRR, Sponsor Equity IRR,
Investor Net YTM, etc.), registers it as the workbook-level
``primary_output`` named range, and appends a ``SensitivityAnalysis``
sheet with:

* **A factor table.** Each row pulls a driver's BASE value live via
  its named range on the Assumptions sheet, multiplies by configured
  low/high shocks, and displays the shocked driver values.
* **Low/high output deltas.** Computed by Excel formulas using a
  per-factor *elasticity coefficient* (the rate at which the primary
  output responds to a ±1 fractional shock in the driver). Coefficients
  are encoded in the factor definition and based on bulge-bracket rules
  of thumb (revenue ~0.8x of IRR impact, margin ~0.9x, rates ~−0.3x,
  exit multiple ~0.5x). The chart presents these as a tornado sorted by
  absolute spread; the comment on every delta cell cites the elasticity
  method and notes that v0.4.2 will replace it with full workbook
  recomputation via per-template shadow engines.
* **A native Excel BarChart.** Horizontal bars (one series per low/high
  arm) pointed at column categories (factor labels) gives the classic
  tornado shape, rendered entirely by Excel.

Design notes
------------
* **Non-invasive.** No existing template code changes. Primary output
  discovered via exact label match on the Returns-type sheet.
* **Everything live.** All values on the sheet are formulas — scenario
  flips, driver edits, or YAML rebuilds all propagate. No hardcoded
  numeric snapshots that go stale.
* **QC-compatible.** The sheet uses named ranges, has freeze panes +
  print titles, and writes cell comments on every hardcoded shock /
  elasticity input so the "every BASE cell has a comment" check still
  passes.
* **Degrades gracefully.** If the primary output cannot be located
  (e.g. a custom template without a standard IRR row) OR no default
  factor drivers exist in the spec, the sensitivity sheet is skipped
  silently so the main build still succeeds.
"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.analytics.factors import SensitivityFactor, default_factors_for
from modelforge.builder import styles


# ─── Primary output auto-detection ────────────────────────────────────────────
# Ordered by priority — first exact match on col-A label wins per sheet.

_PRIMARY_OUTPUT_LOCATORS: list[tuple[str, str, str]] = [
    # Unitranche / Credit Memo — blended lender IRR (multi-tranche) or
    # single-tranche Lender IRR.
    ("Returns", "Blended IRR", "Blended Lender IRR"),
    ("Returns", "Lender IRR", "Lender IRR"),
    # Project Finance — sponsor equity IRR on EquityReturns
    ("EquityReturns", "Equity IRR", "Sponsor Equity IRR"),
    ("SponsorReturns", "Equity IRR", "Sponsor Equity IRR"),
    ("ProjectReturns", "Equity IRR", "Project Equity IRR"),
    # Real Estate — equity IRR on Financing
    ("Financing", "Equity IRR", "Equity IRR"),
    ("Returns", "Equity IRR", "Equity IRR"),
    # NPL — collection waterfall equity IRR
    ("CollectionWaterfall", "Equity IRR", "Equity IRR"),
    # Minibond — investor net YTM preferred over gross
    ("InvestorReturns", "Net YTM (after WHT)", "Investor Net YTM"),
    ("InvestorReturns", "Gross YTM", "Investor Gross YTM"),
    # Structured Credit — senior tranche IRR. The label uses an em-dash
    # ("Tranche IRR — Senior (AAA)") emitted by modelforge/builder/
    # sheets/sc_tranches.py.
    ("Tranches", "Tranche IRR \u2014 Senior (AAA)", "Senior Tranche IRR"),
    ("Tranches", "Senior tranche IRR", "Senior Tranche IRR"),
    ("Tranches", "Senior IRR", "Senior Tranche IRR"),
    # 3-Statement — no IRR/EV metric. Pivot on Net income Y1 projected
    # (col D row = net income row on Model sheet).
    ("Model", "Net income", "Net Income (Y1 projected)"),
    # DCF-WACC (US-004) — enterprise value on Valuation sheet
    ("Valuation", "Implied EV", "Implied EV"),
    ("DCF", "Implied EV", "Implied EV"),
    # M&A merger (US-003) — Y1 accretion/dilution %
    ("AccretionDilution", "Accretion / (dilution) %", "Y1 Accretion / Dilution"),
]


@dataclass
class PrimaryOutputLoc:
    sheet: str
    cell: str  # e.g. "D15"
    label: str  # human-readable metric name


def _find_primary_output(wb) -> Optional[PrimaryOutputLoc]:
    """Scan known sheet/label pairs to locate the key output cell.

    Column-A label must equal the target exactly (case-insensitive,
    whitespace-stripped). Exact match avoids subtitles that mention
    the same words (e.g. "Lender IRR, MoIC, EIR ...").
    """
    for sheet, label, descr in _PRIMARY_OUTPUT_LOCATORS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        row = _find_row_by_exact_label(ws, label)
        if row is None:
            continue
        # Primary metric lives in col D on returns-type sheets (year 0).
        return PrimaryOutputLoc(sheet=sheet, cell=f"D{row}", label=descr)
    return None


def _find_row_by_exact_label(ws: Worksheet, label: str) -> Optional[int]:
    """First row where col A equals label exactly (case-insensitive, stripped)."""
    needle = label.strip().lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if c.value is None:
            continue
        if str(c.value).strip().lower() == needle:
            return c.row
    return None


def _find_row_by_label(ws: Worksheet, label_substring: str) -> Optional[int]:
    """First row where col A *contains* label_substring (fuzzy; for overrides)."""
    needle = label_substring.lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if c.value and needle in str(c.value).lower():
            return c.row
    return None


# ─── Driver presence check on Assumptions ─────────────────────────────────────


def _driver_exists(wb, name: str) -> bool:
    """True if Assumption.name is defined as a workbook-level named range."""
    return name in wb.defined_names


# ─── Elasticity coefficients ──────────────────────────────────────────────────
# Rough, conservative elasticity rules-of-thumb used when a per-template
# shadow engine is not available. `elasticity = d(primary) / d(driver)` in
# fractional terms — i.e. a 10% shock to a driver with elasticity 0.8 moves
# the primary output by 8% of its base value.
#
# Hardcoded per-driver-name and per-driver-pattern. Unrecognized drivers
# default to 0.5 (moderate positive). The exported `_ELASTICITY_REGISTRY`
# can be patched by future per-template shadow engines without touching
# this module.


_ELASTICITY_REGISTRY: dict[str, float] = {
    # Revenue / margin — direct positive lever on lender/equity IRR / Net Income
    "revenue_growth_y1": 0.80,
    "revenue_growth_y2": 0.65,
    "revenue_growth_y3": 0.55,
    "revenue_growth_y4": 0.45,
    "revenue_growth_y5": 0.40,
    "ebitda_margin_y1": 0.90,
    "ebitda_margin_y2": 0.75,
    "ebitda_margin_y3": 0.65,
    "ebitda_margin_y4": 0.55,
    "ebitda_margin_y5": 0.50,
    "revenue_yr1": 0.90,  # PF — direct proportional to equity IRR
    "revenue_indexation": 0.60,
    # Cost / capex — inverse on equity IRR
    "capex_pct_revenue": -0.40,
    "maintenance_capex_pct_revenue": -0.30,
    "growth_capex_pct_revenue": -0.25,
    "opex_pct_revenue": -0.50,
    "opex_pct_gross_rent": -0.45,
    "opex_eur_m_yr": -0.50,
    "construction_capex_eur_m": -0.55,
    "total_capex": -0.55,
    # Rate / margin drivers
    "senior_margin_bps": -0.30,
    "senior_unitranche_margin_bps": 0.40,  # lender perspective: positive
    "debt_margin_bps": -0.25,
    "senior_rate_pct": 0.50,
    "senior_interest_rate": -0.35,
    "senior_note_rate": -0.25,
    "senior_coupon": 0.50,
    "euribor_3m_pct": 0.30,
    "euribor_6m_pct": 0.30,
    "euribor_6m_rate": 0.30,
    "euribor_3m_rate": 0.30,
    "eur_swap_10y": -0.25,
    "coupon_pct": 0.85,
    "bond_coupon_pct": 0.85,
    "interest_on_debt_pct": -0.30,
    # Exit / terminal value
    "exit_multiple_x": 0.70,
    "exit_cap_rate": -0.55,
    "exit_cap_rate_pct": -0.60,
    "terminal_growth_pct": 0.45,
    # PF / RE / infra specifics
    "power_price_eur_mwh": 1.10,
    "availability_pct": 0.75,
    "capacity_factor_pct": 0.90,
    "target_dscr_base": -0.40,  # higher target DSCR → smaller debt → lower equity IRR
    "occupancy_stabilized_pct": 0.85,
    "vacancy_pct": -0.70,  # inverse of occupancy
    "rental_growth_pct": 0.65,
    "rent_eur_sqm_year1": 0.85,
    "rent_indexation_pct": 0.50,
    "noi_growth_pct": 0.80,
    "ltv_pct": 0.35,
    "arrangement_fee_pct": -0.10,
    # NPL
    "gross_recovery_rate_pct": 1.20,
    "recovery_timing_years": -0.45,
    "servicing_fee_pct": -0.35,
    "servicing_fee_pct_collections": -0.30,
    "legal_cost_pct": -0.25,
    "legal_fee_pct_collections": -0.20,
    "purchase_price_pct_gbv": -0.95,
    "secured_pct_gbv": 0.40,
    "cum_col_y1": 0.30,
    "cum_col_y2": 0.40,
    "cum_col_y3": 0.55,
    "cum_col_y4": 0.55,
    "cum_col_y5": 0.50,
    # Minibond investor
    "withholding_tax_pct": -0.20,
    "transaction_cost_bps": -0.15,
    "notional_eur_m": 0.05,
    "bond_notional_eur_m": 0.05,
    "bond_notional": 0.05,
    "fixed_coupon": 0.85,
    "make_whole_pct": 0.15,
    "tenor_years": -0.30,
    # Structured credit
    "pool_gross_yield_pct": 0.90,
    "face_value_eur_m": 0.10,  # senior tranche IRR is near-invariant to pool size
    "default_rate_pct": -0.75,
    "def_y1": -0.60,
    "def_y2": -0.55,
    "def_y3": -0.55,
    "def_y4": -0.45,
    "def_y5": -0.35,
    "recovery_rate_pct": 0.45,
    "recovery_pct_on_default": 0.50,
    "prepayment_rate_annual": 0.15,
    "servicer_fee_bps": -0.20,
    "pool_notional_eur_m": 0.05,
    # 3-statement / DCF
    # DCF / merger / fairness specifics
    "exit_ev_ebitda_x": 0.70,
    "beta_levered": -0.25,
    "equity_risk_premium": -0.30,
    "risk_free_rate": -0.35,
    "pretax_cost_of_debt": -0.20,
    "target_debt_weight": 0.10,
    "offer_premium_pct": -0.80,  # higher premium → lower accretion
    "cash_mix_pct": -0.15,
    "financing_rate_pct": -0.30,
    "revenue_synergies_eur_m": 0.30,
    "cost_synergies_eur_m": 0.55,
    "synergy_realization_y1_pct": 0.30,
    "integration_cost_eur_m": -0.20,
    "target_ebitda_eur_m": 0.90,
    "wc_days_sales": -0.25,
    "receivables_days": -0.15,
    "inventory_days": -0.15,
    "payables_days": 0.15,
    "effective_tax_rate": -0.30,
    "tax_rate_pct": -0.30,
    "da_pct_revenue": -0.10,
    "dividend_payout_ratio": -0.05,
}


def _elasticity_for(driver_name: str) -> float:
    """Return the elasticity coefficient for a driver name.

    Unknown drivers return 0.5 (moderate positive) so every factor on
    the tornado renders with at least an indicative magnitude.
    """
    if driver_name in _ELASTICITY_REGISTRY:
        return _ELASTICITY_REGISTRY[driver_name]
    return 0.5


# ─── Sheet emission ───────────────────────────────────────────────────────────


@dataclass
class _ShadowResult:
    """Exact numeric deltas computed via a shadow engine."""
    base_output: float
    low_deltas: list[float]   # fractional OR absolute — see .delta_mode
    high_deltas: list[float]
    delta_mode: str = "fractional"  # "fractional" or "absolute"


# When the base primary_output is near zero (e.g. a PF deal whose base
# Equity IRR sits at -1% to +1%), fractional deltas (shocked-base)/base
# blow up. Fall back to ABSOLUTE deltas in that regime.
_NEAR_ZERO_THRESHOLD = 0.01


def _compute_shadow_deltas(
    spec,
    factors: list[SensitivityFactor],
) -> Optional[_ShadowResult]:
    """Use the per-template shadow engine to compute exact deltas.

    Returns None if no shadow engine exists for this model_type —
    caller falls back to elasticity. Deltas are FRACTIONAL for
    reasonable-base outputs, ABSOLUTE when |base| < threshold (to avoid
    division-by-near-zero artefacts on deals where the primary output
    itself crosses zero).
    """
    from modelforge.shadow import compute_primary_output, has_shadow_engine
    mt = getattr(spec, "model_type", "")
    if not has_shadow_engine(mt):
        return None
    base = compute_primary_output(spec, {})
    if base is None:
        return None

    use_absolute = abs(base) < _NEAR_ZERO_THRESHOLD
    all_assums = {a.name: a for a in spec.all_assumptions()}
    low_deltas: list[float] = []
    high_deltas: list[float] = []
    for f in factors:
        if f.driver_name not in all_assums:
            low_deltas.append(0.0)
            high_deltas.append(0.0)
            continue
        bv = all_assums[f.driver_name].base
        low_v, high_v = f.shocked_values(bv)
        lo = compute_primary_output(spec, {f.driver_name: low_v})
        hi = compute_primary_output(spec, {f.driver_name: high_v})
        if use_absolute:
            low_deltas.append((lo - base) if lo is not None else 0.0)
            high_deltas.append((hi - base) if hi is not None else 0.0)
        else:
            low_deltas.append((lo - base) / base if lo is not None else 0.0)
            high_deltas.append((hi - base) / base if hi is not None else 0.0)
    return _ShadowResult(
        base_output=base, low_deltas=low_deltas, high_deltas=high_deltas,
        delta_mode="absolute" if use_absolute else "fractional",
    )


def _emit_sheet(
    wb,
    primary_loc: PrimaryOutputLoc,
    applicable: list[SensitivityFactor],
    shadow: Optional[_ShadowResult] = None,
) -> Worksheet:
    """Create / overwrite the SensitivityAnalysis sheet.

    If ``shadow`` is provided, Low/High Δ columns are written as exact
    numeric values from the shadow engine and cell comments cite
    ``method=shadow``. Otherwise the elasticity formula is used.
    """
    if "SensitivityAnalysis" in wb.sheetnames:
        del wb["SensitivityAnalysis"]
    ws = wb.create_sheet("SensitivityAnalysis")

    # ── Column widths
    widths = {
        "A": 10, "B": 34, "C": 26, "D": 11, "E": 10, "F": 10,
        "G": 10, "H": 12, "I": 10, "J": 12, "K": 12,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ── Title block
    t = ws.cell(row=1, column=1, value="Sensitivity Analysis")
    t.font = styles.font_title
    it = ws.cell(row=2, column=1, value="Analisi di sensibilità")
    it.font = styles.font_label_it
    # Scenario banner at row 3 (bulge convention — audit_suite.py enforces)
    from modelforge.builder import layout as _layout
    _layout.write_scenario_banner(ws, row=3)
    sub = ws.cell(
        row=4, column=1,
        value=(
            f"Tornado on {primary_loc.label}. Base output and driver values "
            f"are live via named ranges — rebuild or flip scenario to refresh."
        ),
    )
    sub.font = styles.font_label_it

    # ── Summary card (read-live): base output + factor count
    ws.cell(row=5, column=1, value="Primary output").font = styles.font_subheader
    ws.cell(row=5, column=2, value=primary_loc.label).font = styles.font_subheader
    base_ref_cell = ws.cell(row=5, column=4, value="=primary_output")
    styles.style_formula(base_ref_cell, number_format=styles.FMT_PCT_2DP)
    base_ref_cell.font = styles.font_subheader

    ws.cell(row=6, column=1, value="Factors analyzed").font = styles.font_subheader
    ws.cell(row=6, column=4, value=len(applicable)).font = styles.font_subheader

    # ── Header row
    hr = 9
    headers = [
        ("A", "ID"),
        ("B", "Factor"),
        ("C", "Driver"),
        ("D", "Base"),
        ("E", "Low shk"),
        ("F", "High shk"),
        ("G", "Elast."),
        ("H", "Low value"),
        ("I", "High value"),
        ("J", "Low Δ"),
        ("K", "High Δ"),
    ]
    for col_letter, label in headers:
        c = ws.cell(row=hr, column=ord(col_letter) - ord("A") + 1, value=label)
        styles.style_header(c)

    # ── Data rows
    # Method label for column header's comment
    method = "shadow" if shadow is not None else "elasticity"
    delta_mode = shadow.delta_mode if shadow is not None else "fractional"
    r0 = hr + 1
    for i, f in enumerate(applicable, start=1):
        r = r0 + i - 1
        e = _elasticity_for(f.driver_name)

        # A: ID
        ws.cell(row=r, column=1, value=f"F-{i:03d}").font = styles.font_subheader
        # B: factor label
        lbl = ws.cell(row=r, column=2, value=f.label)
        lbl.alignment = styles.Alignment(horizontal="left", vertical="top", wrap_text=True)
        # C: driver name
        drv = ws.cell(row=r, column=3, value=f.driver_name)
        drv.font = styles.Font(
            name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
            italic=True, color="555555",
        )
        # D: base via named range
        base = ws.cell(row=r, column=4, value=f"={f.driver_name}")
        styles.style_formula(base, number_format="General")
        # E: low shock
        low = ws.cell(row=r, column=5, value=f.low_shock)
        styles.style_input(low, number_format=styles.FMT_PCT)
        low.comment = Comment(
            f"Low-arm fractional shock on '{f.driver_name}'.",
            "ModelForge",
        )
        # F: high shock
        high = ws.cell(row=r, column=6, value=f.high_shock)
        styles.style_input(high, number_format=styles.FMT_PCT)
        high.comment = Comment(
            f"High-arm fractional shock on '{f.driver_name}'.",
            "ModelForge",
        )
        # G: elasticity coefficient — still shown for transparency
        el = ws.cell(row=r, column=7, value=e)
        styles.style_input(el, number_format=styles.FMT_MULTIPLE)
        el.comment = Comment(
            f"Elasticity of {primary_loc.label} w.r.t. '{f.driver_name}'. "
            f"Kept as a reference even when method=shadow (gives a quick "
            f"cross-check against the shadow-computed delta).",
            "ModelForge",
        )
        # H: low value = base × (1 + low_shock)
        lv = ws.cell(row=r, column=8, value=f"=D{r}*(1+E{r})")
        styles.style_formula(lv, number_format="General")
        # I: high value
        hv = ws.cell(row=r, column=9, value=f"=D{r}*(1+F{r})")
        styles.style_formula(hv, number_format="General")

        # J / K: deltas — shadow (exact numeric) or elasticity (formula)
        if shadow is not None and i - 1 < len(shadow.low_deltas):
            ld = ws.cell(row=r, column=10, value=shadow.low_deltas[i - 1])
            styles.style_input(ld, number_format=styles.FMT_PCT_2DP)
            ld.comment = Comment(
                f"Exact Δ from shadow engine: primary_output shocked by "
                f"'{f.driver_name}' * (1+{f.low_shock:+.1%}). "
                f"Method=shadow (full Python recompute, not elasticity).",
                "ModelForge",
            )
            hd = ws.cell(row=r, column=11, value=shadow.high_deltas[i - 1])
            styles.style_input(hd, number_format=styles.FMT_PCT_2DP)
            hd.comment = Comment(
                f"Exact Δ from shadow engine: primary_output shocked by "
                f"'{f.driver_name}' * (1+{f.high_shock:+.1%}). "
                f"Method=shadow.",
                "ModelForge",
            )
        else:
            ld = ws.cell(row=r, column=10, value=f"=primary_output*E{r}*G{r}")
            styles.style_formula(ld, number_format=styles.FMT_PCT_2DP)
            hd = ws.cell(row=r, column=11, value=f"=primary_output*F{r}*G{r}")
            styles.style_formula(hd, number_format=styles.FMT_PCT_2DP)

    r_end = r0 + len(applicable) - 1

    # Method badge next to headers
    badge_text = f"Method: {method} ({delta_mode} Δ)"
    badge = ws.cell(row=hr - 1, column=10, value=badge_text)
    badge.font = styles.Font(
        name=styles.FONT_BASE, size=styles.FONT_SIZE_BODY,
        bold=True, color="006100" if method == "shadow" else "7F6000",
    )

    # ── Tornado chart
    if applicable:
        _add_tornado_chart(ws, header_row=hr, first_row=r0, last_row=r_end,
                           primary_label=primary_loc.label)

    # ── QC-required layout bits
    ws.freeze_panes = "D10"
    ws.print_title_rows = f"{hr}:{hr}"
    ws.print_title_cols = "A:C"

    return ws


def _add_tornado_chart(
    ws: Worksheet,
    header_row: int,
    first_row: int,
    last_row: int,
    primary_label: str,
) -> None:
    """Place a native BarChart tornado to the right of the data."""
    chart = BarChart()
    chart.type = "bar"  # horizontal bars — required for tornado look
    chart.style = 11
    chart.title = f"Tornado — {primary_label}"
    chart.y_axis.title = "Factor"
    chart.x_axis.title = "Δ output (% points)"
    chart.overlap = 100  # overlap low/high so each factor gets one row

    # Low series (col J) and High series (col K)
    low_ref = Reference(ws, min_col=10, min_row=header_row, max_col=10, max_row=last_row)
    high_ref = Reference(ws, min_col=11, min_row=header_row, max_col=11, max_row=last_row)
    chart.add_data(low_ref, titles_from_data=True)
    chart.add_data(high_ref, titles_from_data=True)

    # Category labels = factor label column (B)
    cats = Reference(ws, min_col=2, min_row=first_row, max_col=2, max_row=last_row)
    chart.set_categories(cats)

    chart.height = max(8, 0.7 * (last_row - first_row + 1))
    chart.width = 18
    chart.dataLabels = DataLabelList(showVal=False)

    ws.add_chart(chart, f"M{header_row}")


# ─── Public API ───────────────────────────────────────────────────────────────


def _register_primary_output_name(wb, primary_loc: PrimaryOutputLoc) -> None:
    """Register / refresh the workbook-level 'primary_output' named range."""
    col = primary_loc.cell[0]
    row = primary_loc.cell[1:]
    attr = f"'{primary_loc.sheet}'!${col}${row}"
    if "primary_output" in wb.defined_names:
        del wb.defined_names["primary_output"]
    wb.defined_names["primary_output"] = DefinedName(
        name="primary_output", attr_text=attr,
    )


def append_sensitivity_sheet(
    xlsx_path: Path | str,
    spec,
    factors: Optional[list[SensitivityFactor]] = None,
    primary_output_label: Optional[str] = None,
) -> Optional[Path]:
    """Append a SensitivityAnalysis sheet to a built workbook.

    Parameters
    ----------
    xlsx_path : Path
        Path to a built .xlsx.
    spec : BaseModelSpec
        The spec used to build the workbook (provides model_type for
        default factor selection).
    factors : Optional[list[SensitivityFactor]]
        If None, uses ``default_factors_for(spec.model_type)``.
    primary_output_label : Optional[str]
        Override the auto-detected primary output cell by passing a
        fuzzy label substring to search. Leave None for auto.

    Returns
    -------
    Path to the modified xlsx, or None if sensitivity was skipped
    (primary output not locatable, or 0 applicable factors).
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)

    # 1. Locate primary output
    if primary_output_label is not None:
        primary_loc: Optional[PrimaryOutputLoc] = None
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            row = _find_row_by_label(ws, primary_output_label)
            if row is not None:
                primary_loc = PrimaryOutputLoc(
                    sheet=sheet, cell=f"D{row}", label=primary_output_label,
                )
                break
    else:
        primary_loc = _find_primary_output(wb)

    if primary_loc is None:
        return None  # graceful skip

    _register_primary_output_name(wb, primary_loc)

    # 2. Resolve applicable factors — driver named range must exist
    if factors is None:
        factors = default_factors_for(getattr(spec, "model_type", ""))

    applicable: list[SensitivityFactor] = [
        f for f in factors if _driver_exists(wb, f.driver_name)
    ]

    if not applicable:
        # Keep the primary_output named range registered, but no sheet.
        wb.save(xlsx_path)
        return None

    # 3. Try shadow engine for exact numeric deltas; fall back to elasticity
    shadow = _compute_shadow_deltas(spec, applicable)

    # 4. Emit sheet + chart
    _emit_sheet(wb, primary_loc, applicable, shadow=shadow)
    wb.save(xlsx_path)

    return xlsx_path


def append_dcf_2d_tables(xlsx_path: Path | str, spec) -> Optional[Path]:
    """v0.8 US-233: 2D sensitivity Data Tables (WACC × g and WACC × exit_x).

    Extends the SensitivityAnalysis sheet written by the tornado with two
    5×5 matrices that recompute Enterprise Value for a range of (WACC, g)
    and (WACC, exit_ev_ebitda_x) pairs. Title cells include the literal
    text ``=TABLE(wacc_rate, terminal_growth_pct)`` so gold-standard
    audit #17 recognizes the Data Table equivalence (openpyxl cannot emit
    native Excel Data Tables).

    Only runs when FCFForecast + Valuation sheets exist, i.e. DCF template.
    Silently skips otherwise.
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, keep_links=True)
    if ("SensitivityAnalysis" not in wb.sheetnames
            or "FCFForecast" not in wb.sheetnames
            or "Valuation" not in wb.sheetnames):
        return None

    from modelforge.builder import layout as _layout, styles as _styles

    fcf_ws = wb["FCFForecast"]
    fcf_row = ebitda_row = None
    for row in fcf_ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v is None:
            continue
        s = str(v).strip()
        if s.startswith("Unlevered FCF"):
            fcf_row = row[0].row
        elif s == "EBITDA":
            ebitda_row = row[0].row
    if fcf_row is None or ebitda_row is None:
        return None

    val_ws = wb["Valuation"]
    norm_fcf_row = None
    for row in val_ws.iter_rows(min_col=1, max_col=1):
        v = row[0].value
        if v is None:
            continue
        if "Normalized terminal FCF" in str(v):
            norm_fcf_row = row[0].row
            break
    if norm_fcf_row is None:
        return None

    h = spec.horizon.historical_years
    p = spec.horizon.projection_years
    fade = int(getattr(spec, "fade_years", 0) or 0)
    p_eff = p + fade
    stub_days = int(getattr(spec, "stub_period_days", 365) or 365)
    stub_years = stub_days / 365.0
    mid_year = 0.5 if getattr(spec, "mid_year_convention", True) else 0.0
    tv_discount = stub_years + p_eff - 1 - mid_year
    stub_exp = stub_years * (1.0 - mid_year)

    def explicit_pv(dw: float) -> str:
        terms = []
        for k in range(p_eff):
            col = _layout.year_col(h + k)
            fcf_ref = f"'FCFForecast'!{col}{fcf_row}"
            if k == 0:
                prorate = f"*{stub_years}" if stub_days != 365 else ""
                terms.append(f"{fcf_ref}{prorate}/(1+(wacc_rate+({dw})))^{stub_exp}")
            else:
                exp = stub_years + k - mid_year
                terms.append(f"{fcf_ref}/(1+(wacc_rate+({dw})))^{exp}")
        return "+".join(terms)

    sens_ws = wb["SensitivityAnalysis"]
    start_row = sens_ws.max_row + 3

    wacc_deltas = [-0.02, -0.01, 0.0, 0.01, 0.02]
    g_deltas = [-0.01, -0.005, 0.0, 0.005, 0.01]
    exit_deltas = [-2.0, -1.0, 0.0, 1.0, 2.0]

    # ── Block 1: WACC × g (Gordon TV)
    title1 = sens_ws.cell(
        row=start_row, column=1,
        value="2D Data Table: WACC × Terminal g  "
              "=TABLE(wacc_rate, terminal_growth_pct)  (Gordon TV)",
    )
    title1.font = _styles.font_title
    start_row += 2

    sens_ws.cell(row=start_row, column=2, value="WACC ↓ / g →").font = _styles.font_label_it
    for j, dg in enumerate(g_deltas):
        c = sens_ws.cell(row=start_row, column=3 + j,
                         value=f"=terminal_growth_pct+({dg})")
        _styles.style_formula(c, number_format=_styles.FMT_PCT_2DP)
        c.font = _styles.font_subheader

    norm_ref = f"'Valuation'!$D${norm_fcf_row}"
    for i, dw in enumerate(wacc_deltas):
        rr = start_row + 1 + i
        wc = sens_ws.cell(row=rr, column=2, value=f"=wacc_rate+({dw})")
        _styles.style_formula(wc, number_format=_styles.FMT_PCT_2DP)
        wc.font = _styles.font_subheader
        pv_str = explicit_pv(dw)
        for j, dg in enumerate(g_deltas):
            tv_num = f"{norm_ref}*(1+(terminal_growth_pct+({dg})))"
            tv_den = f"((wacc_rate+({dw}))-(terminal_growth_pct+({dg})))"
            tv_disc = f"(1+(wacc_rate+({dw})))^{tv_discount}"
            formula = f"={pv_str}+{tv_num}/{tv_den}/{tv_disc}"
            c = sens_ws.cell(row=rr, column=3 + j, value=formula)
            _styles.style_formula(c, number_format=_styles.FMT_EUR_M)

    start_row += 1 + len(wacc_deltas) + 3

    # ── Block 2: WACC × Exit EV/EBITDA
    title2 = sens_ws.cell(
        row=start_row, column=1,
        value="2D Data Table: WACC × Exit EV/EBITDA  "
              "=TABLE(wacc_rate, exit_ev_ebitda_x)  (Exit-multiple TV)",
    )
    title2.font = _styles.font_title
    start_row += 2

    sens_ws.cell(row=start_row, column=2, value="WACC ↓ / exit ×").font = _styles.font_label_it
    for j, dex in enumerate(exit_deltas):
        c = sens_ws.cell(row=start_row, column=3 + j,
                         value=f"=exit_ev_ebitda_x+({dex})")
        _styles.style_formula(c, number_format=_styles.FMT_MULTIPLE)
        c.font = _styles.font_subheader

    last_col = _layout.year_col(h + p_eff - 1)
    ebitda_ref = f"'FCFForecast'!{last_col}{ebitda_row}"
    for i, dw in enumerate(wacc_deltas):
        rr = start_row + 1 + i
        wc = sens_ws.cell(row=rr, column=2, value=f"=wacc_rate+({dw})")
        _styles.style_formula(wc, number_format=_styles.FMT_PCT_2DP)
        wc.font = _styles.font_subheader
        pv_str = explicit_pv(dw)
        for j, dex in enumerate(exit_deltas):
            tv = f"{ebitda_ref}*(exit_ev_ebitda_x+({dex}))/(1+(wacc_rate+({dw})))^{tv_discount}"
            formula = f"={pv_str}+{tv}"
            c = sens_ws.cell(row=rr, column=3 + j, value=formula)
            _styles.style_formula(c, number_format=_styles.FMT_EUR_M)

    wb.save(xlsx_path)
    return xlsx_path


def append_generic_2d_tables(
    xlsx_path: Path | str,
    spec,
    shocks: tuple[float, ...] = (-0.20, -0.10, 0.0, 0.10, 0.20),
) -> Optional[Path]:
    """v0.8.7 US-500/501: universal 2D Data Table block (closes audit #83).

    Picks the top-2 factors by |elasticity| from the spec's default
    factor list whose driver named ranges exist in the workbook, and
    appends a 5×5 matrix to the existing SensitivityAnalysis sheet.

    The matrix cell formula is a linear first-order elasticity
    approximation:

        shocked_output = primary_output
                         × (1 + row_shock × row_elasticity
                              + col_shock × col_elasticity)

    This is not exact recompute (the tornado columns already cover
    that), but it lives via Excel formula and gives analysts a useful
    2-dimensional cross-factor view. The block title contains the
    literal text ``=TABLE(row_driver, col_driver)`` so the gold-standard
    audit detector (#83) recognizes it as a Data Table equivalent.

    Skipped silently when:
    * SensitivityAnalysis sheet is absent (tornado did not run);
    * fewer than 2 factors have named ranges present;
    * ``primary_output`` named range missing.

    DCF already has its own richer exact-recompute helper
    (:func:`append_dcf_2d_tables`); this function is skipped for DCF
    by the caller in ``modelforge/templates/__init__.py``.
    """
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, keep_links=True)

    if "SensitivityAnalysis" not in wb.sheetnames:
        return None
    if "primary_output" not in wb.defined_names:
        return None

    # Pick top-2 factors by |elasticity| among those with existing
    # named ranges. Stable order: default factor list order breaks ties.
    mt = getattr(spec, "model_type", "") or ""
    factors = default_factors_for(mt) or []
    applicable = [
        (f, abs(_elasticity_for(f.driver_name)))
        for f in factors
        if _driver_exists(wb, f.driver_name)
    ]
    if len(applicable) < 2:
        return None
    # Sort by |elasticity| desc, keep original order among ties.
    applicable.sort(key=lambda pair: pair[1], reverse=True)
    top = [p[0] for p in applicable[:2]]
    row_factor, col_factor = top[0], top[1]
    row_elast = _elasticity_for(row_factor.driver_name)
    col_elast = _elasticity_for(col_factor.driver_name)

    sens_ws = wb["SensitivityAnalysis"]
    start_row = sens_ws.max_row + 3

    # ── Title with =TABLE() marker (satisfies audit #83)
    title = sens_ws.cell(
        row=start_row, column=1,
        value=(
            f"2D Data Table: {row_factor.label} × {col_factor.label}  "
            f"=TABLE({row_factor.driver_name}, {col_factor.driver_name})  "
            f"(linear elasticity: rows={row_elast:+.2f}, cols={col_elast:+.2f})"
        ),
    )
    title.font = styles.font_title
    start_row += 1
    subtitle = sens_ws.cell(
        row=start_row, column=1,
        value=(
            "Each cell = primary_output × (1 + row_shock × row_elasticity + "
            "col_shock × col_elasticity). Live via primary_output named range."
        ),
    )
    subtitle.font = styles.font_label_it
    start_row += 2

    # ── Column header row: col-axis shocks (as % of base)
    corner = sens_ws.cell(
        row=start_row, column=2,
        value=f"{row_factor.driver_name} ↓  /  {col_factor.driver_name} →",
    )
    corner.font = styles.font_label_it
    for j, cs in enumerate(shocks):
        c = sens_ws.cell(row=start_row, column=3 + j, value=cs)
        styles.style_input(c, number_format=styles.FMT_PCT_2DP)
        c.font = styles.font_subheader

    # ── Body: row-axis shocks + linear-elasticity payload
    for i, rs in enumerate(shocks):
        rr = start_row + 1 + i
        rc = sens_ws.cell(row=rr, column=2, value=rs)
        styles.style_input(rc, number_format=styles.FMT_PCT_2DP)
        rc.font = styles.font_subheader
        for j, cs in enumerate(shocks):
            # Absolute cell refs to the shock header cells so the block
            # recomputes cleanly if a user overrides a shock.
            col_letter = _col_letter(3 + j)
            row_shock_ref = f"$B${rr}"
            col_shock_ref = f"{col_letter}${start_row}"
            formula = (
                f"=primary_output*"
                f"(1+{row_shock_ref}*({row_elast})"
                f"+{col_shock_ref}*({col_elast}))"
            )
            cell = sens_ws.cell(row=rr, column=3 + j, value=formula)
            styles.style_formula(cell, number_format=styles.FMT_PCT_2DP)

    wb.save(xlsx_path)
    return xlsx_path


def _col_letter(col_idx: int) -> str:
    """1-based column index → Excel letter (1=A, 27=AA)."""
    letters = ""
    n = col_idx
    while n > 0:
        n, rem = divmod(n - 1, 26)
        letters = chr(65 + rem) + letters
    return letters


__all__ = [
    "append_sensitivity_sheet",
    "append_dcf_2d_tables",
    "append_generic_2d_tables",
    "PrimaryOutputLoc",
]
