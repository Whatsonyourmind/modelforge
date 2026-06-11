"""Reproducibility metadata — spec hash, ModelForge version, Python
version, build timestamp.

Every workbook emitted by ``modelforge build`` carries a
``Reproducibility`` sheet plus matching named ranges
(``mf_spec_sha256``, ``mf_version``, ``mf_python_version``,
``mf_build_timestamp``). Auditors and the v0.4 dossier exporter rely on
these to confirm a workbook was built from a known spec and has not
been tampered with.

The spec hash is computed over the *exact bytes* of the YAML spec file
(or over a canonical JSON serialization of the spec if bytes are not
provided). Two workbooks built from identical spec bytes will carry the
same ``mf_spec_sha256`` — useful for "was this the version we
approved?" committee questions.

Determinism
-----------
Builds are reproducible by default: two builds of the *same spec bytes*
produce a byte-identical workbook. The build timestamp is **not** the
wall clock — it is derived deterministically from the spec hash (so a
given spec always carries the same timestamp), unless overridden:

* ``SOURCE_DATE_EPOCH`` — POSIX-standard reproducible-build env var
  (Unix seconds). When set, it pins the build timestamp for every
  workbook in the run. Useful for "this whole batch was built at T".
* ``MODELFORGE_WALL_CLOCK_TIMESTAMP=1`` — opt back in to the real wall
  clock (non-reproducible) for the rare case where a live build time is
  genuinely wanted.

Besides the visible timestamp cell, ``append_reproducibility_block``
(the final save of the build) also pins the workbook's core-properties
``created``/``modified`` dates and every zip-member mtime to the same
deterministic instant — otherwise openpyxl would stamp ``docProps/
core.xml`` and the zip directory with ``datetime.now()`` and break
byte-identity even though every cell matched.
"""

from __future__ import annotations

import hashlib
import json
import os
import platform
import sys
import zipfile
from datetime import datetime, timezone
from importlib import metadata
from pathlib import Path
from typing import Optional

from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet

from modelforge.builder import styles


SHEET_NAME = "Reproducibility"

# Sentinel epoch used when no spec hash is available to derive from and no
# override is set. 2020-01-01T00:00:00Z — a fixed, recognizable instant.
_SENTINEL_EPOCH = 1577836800
# Env var (POSIX reproducible-builds standard) that pins the build instant.
ENV_SOURCE_DATE_EPOCH = "SOURCE_DATE_EPOCH"
# Opt back in to a real, non-reproducible wall-clock timestamp.
ENV_WALL_CLOCK = "MODELFORGE_WALL_CLOCK_TIMESTAMP"

# Named-range identifiers. Keep the `mf_` prefix so they don't clash
# with business drivers on Assumptions.
NAME_SHA = "mf_spec_sha256"
NAME_VERSION = "mf_version"
NAME_PYTHON = "mf_python_version"
NAME_TIMESTAMP = "mf_build_timestamp"
NAME_SPEC_SOURCE = "mf_spec_source"


def _modelforge_version() -> str:
    try:
        return metadata.version("modelforge")
    except metadata.PackageNotFoundError:
        return "0.0.0+local"


def _canonical_spec_bytes(spec) -> bytes:
    """Fallback hash source when the YAML source bytes are unknown.

    Uses ``spec.model_dump(mode='json')`` with sorted keys and no
    whitespace so a given spec always produces identical bytes.
    """
    payload = spec.model_dump(mode="json")
    return json.dumps(payload, sort_keys=True, separators=(",", ":"),
                      default=str).encode("utf-8")


def compute_spec_hash(
    spec,
    spec_source_bytes: Optional[bytes] = None,
) -> tuple[str, str]:
    """Return (sha256_hex, source_descriptor).

    Prefers the raw YAML bytes for deterministic cross-platform hashing.
    Falls back to a canonical JSON serialization of the Pydantic spec.
    """
    if spec_source_bytes is not None:
        return hashlib.sha256(spec_source_bytes).hexdigest(), "yaml_bytes"
    return hashlib.sha256(_canonical_spec_bytes(spec)).hexdigest(), "canonical_json"


def _resolve_build_datetime(sha: Optional[str]) -> datetime:
    """Return the (timezone-aware UTC) build timestamp for this build.

    Resolution order (first match wins):

    1. ``MODELFORGE_WALL_CLOCK_TIMESTAMP`` truthy → real wall clock
       (explicitly non-reproducible; opt-in only).
    2. ``SOURCE_DATE_EPOCH`` set → that Unix epoch (whole run pinned).
    3. A valid spec ``sha`` → an epoch derived deterministically from the
       hash, so the *same spec bytes always carry the same timestamp*
       while different specs differ. The derived value lands in a sane,
       human-plausible range (post-2020) so the cell reads naturally.
    4. Fallback → the fixed ``_SENTINEL_EPOCH``.

    Same spec bytes → same hash → same timestamp → byte-identical
    workbook. This is the core of the determinism guarantee.
    """
    if os.environ.get(ENV_WALL_CLOCK, "").strip().lower() in ("1", "true", "yes", "on"):
        return datetime.now(timezone.utc).replace(microsecond=0)

    raw_epoch = os.environ.get(ENV_SOURCE_DATE_EPOCH)
    if raw_epoch is not None:
        try:
            return datetime.fromtimestamp(int(raw_epoch.strip()), tz=timezone.utc)
        except (ValueError, OverflowError, OSError):
            pass  # malformed override → fall through to deterministic derivation

    if sha:
        # Map the first 8 hex digits of the hash into a deterministic
        # offset (0 .. ~50 years of seconds) added to the sentinel epoch.
        # Bounded so it never overflows or lands in an absurd year.
        try:
            offset = int(sha[:8], 16) % (50 * 365 * 24 * 3600)  # < 50 years
        except (ValueError, TypeError):
            offset = 0
        return datetime.fromtimestamp(_SENTINEL_EPOCH + offset, tz=timezone.utc)

    return datetime.fromtimestamp(_SENTINEL_EPOCH, tz=timezone.utc)


def _pin_core_modified(core_xml: bytes, dt: datetime) -> bytes:
    """Replace ``dcterms:modified`` in core.xml with the deterministic instant.

    openpyxl overwrites ``properties.modified`` with ``datetime.now()``
    *inside* ``save()`` (openpyxl.writer.excel line ~291), so pinning it
    on the in-memory workbook is clobbered. We patch the serialized XML
    after the fact, matching the W3CDTF instant that ``created`` already
    carries. ``re`` is used on the single element only — the rest of the
    document is left byte-identical.
    """
    import re as _re
    iso_z = dt.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
    return _re.sub(
        rb"(<dcterms:modified[^>]*>)[^<]*(</dcterms:modified>)",
        lambda m: m.group(1) + iso_z.encode("ascii") + m.group(2),
        core_xml,
        count=1,
    )


def _normalize_zip_timestamps(xlsx_path: Path, dt: datetime) -> None:
    """Make the .xlsx archive deterministic.

    Two sources of wall-clock noise survive an openpyxl ``save()``:

    1. Every zip member is stamped with ``datetime.now()`` as its mtime.
    2. ``docProps/core.xml``'s ``dcterms:modified`` is re-set to
       ``datetime.now()`` *during* the save, clobbering any value we
       pinned on the in-memory workbook.

    We repack the archive (preserving member order, bytes, and flags),
    pinning every member's ``date_time`` to ``dt`` and rewriting the
    ``modified`` element of core.xml to the same deterministic instant.
    """
    dt_tuple = (dt.year, dt.month, dt.day, dt.hour, dt.minute, dt.second)
    with zipfile.ZipFile(xlsx_path, "r") as zin:
        infos = zin.infolist()
        members = [(info, zin.read(info.filename)) for info in infos]

    tmp_path = xlsx_path.with_suffix(xlsx_path.suffix + ".repro.tmp")
    with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, data in members:
            if info.filename == "docProps/core.xml":
                data = _pin_core_modified(data, dt)
            new_info = zipfile.ZipInfo(filename=info.filename, date_time=dt_tuple)
            new_info.compress_type = info.compress_type
            new_info.external_attr = info.external_attr
            new_info.internal_attr = info.internal_attr
            new_info.create_system = info.create_system
            zout.writestr(new_info, data)
    os.replace(tmp_path, xlsx_path)


def _register_name(wb, name: str, addr: str) -> None:
    if name in wb.defined_names:
        del wb.defined_names[name]
    wb.defined_names[name] = DefinedName(name=name, attr_text=addr)


def _emit_sheet(
    wb,
    sha: str,
    source_descriptor: str,
    spec_source_path: Optional[Path],
    build_dt: datetime,
) -> Worksheet:
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 70

    from modelforge.builder.i18n import L as _L
    ws.cell(row=1, column=1, value=_L("reproducibility_title").en).font = styles.font_title
    ws.cell(row=2, column=1, value=_L("reproducibility_title").it).font = styles.font_label_it
    from modelforge.builder import layout as _layout
    _layout.write_scenario_banner(ws, row=3)
    ws.cell(
        row=4, column=1,
        value=(
            "Every workbook emitted by `modelforge build` carries this "
            "metadata. The spec SHA-256 is deterministic: two workbooks "
            "built from identical spec bytes will match. Use "
            "`modelforge verify <xlsx> --spec <yaml>` to confirm a workbook "
            "was built from a given spec."
        ),
    ).font = styles.font_label_it

    rows = [
        ("Spec SHA-256", sha, NAME_SHA),
        ("ModelForge version", _modelforge_version(), NAME_VERSION),
        ("Python version", platform.python_version(), NAME_PYTHON),
        ("Build timestamp (UTC)",
         build_dt.isoformat(timespec="seconds"),
         NAME_TIMESTAMP),
        ("Spec source",
         str(spec_source_path) if spec_source_path else f"({source_descriptor})",
         NAME_SPEC_SOURCE),
        ("Host", f"{platform.system()} {platform.release()}", None),
        ("Python impl", f"{platform.python_implementation()} "
                        f"{sys.version_info.major}.{sys.version_info.minor}", None),
    ]

    r0 = 5
    for i, (label, value, name) in enumerate(rows):
        r = r0 + i
        lbl = ws.cell(row=r, column=1, value=label)
        lbl.font = styles.font_subheader
        v = ws.cell(row=r, column=2, value=value)
        styles.style_input(v, number_format="General")
        if name == NAME_SHA:
            v.font = styles.Font(
                name="Consolas", size=styles.FONT_SIZE_BODY,
            )
        if name is not None:
            addr = f"'{SHEET_NAME}'!$B${r}"
            _register_name(wb, name, addr)

    ws.freeze_panes = "B5"
    ws.print_title_rows = "1:3"
    return ws


def append_reproducibility_block(
    xlsx_path: Path | str,
    spec,
    spec_source_bytes: Optional[bytes] = None,
    spec_source_path: Optional[Path | str] = None,
) -> Path:
    """Append Reproducibility sheet + named ranges to a built workbook.

    Parameters
    ----------
    xlsx_path : Path
        Built workbook.
    spec : BaseModelSpec
        The Pydantic spec used.
    spec_source_bytes : Optional[bytes]
        Raw YAML bytes (preferred) for deterministic hashing.
    spec_source_path : Optional[Path]
        For display on the Reproducibility sheet.
    """
    xlsx_path = Path(xlsx_path)
    sha, descriptor = compute_spec_hash(spec, spec_source_bytes)
    build_dt = _resolve_build_datetime(sha)

    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    _emit_sheet(wb, sha, descriptor,
                Path(spec_source_path) if spec_source_path else None,
                build_dt)
    # Pin core-properties dates so docProps/core.xml is deterministic.
    # (openpyxl defaults created/modified to datetime.now().)
    naive_dt = build_dt.astimezone(timezone.utc).replace(tzinfo=None)
    wb.properties.created = naive_dt
    wb.properties.modified = naive_dt
    wb.save(xlsx_path)
    # Normalize zip-member mtimes so the archive directory is deterministic.
    _normalize_zip_timestamps(xlsx_path, build_dt)
    return xlsx_path


def finalize_determinism(
    xlsx_path: Path | str,
    spec,
    spec_source_bytes: Optional[bytes] = None,
) -> Path:
    """Pin core-properties dates + zip mtimes on an already-saved workbook.

    Call this as the **last** workbook-mutating step of a build. Any later
    ``openpyxl`` ``save`` re-stamps ``docProps/core.xml`` (created/modified)
    and the zip directory with ``datetime.now()`` — so the post-processor
    that runs last (e.g. AutoColor) must finish with this call, or the
    byte-identity guarantee is lost even though every cell matches.

    Idempotent: derives the same deterministic instant from the spec hash
    that ``append_reproducibility_block`` used.
    """
    xlsx_path = Path(xlsx_path)
    sha, _ = compute_spec_hash(spec, spec_source_bytes)
    build_dt = _resolve_build_datetime(sha)
    naive_dt = build_dt.astimezone(timezone.utc).replace(tzinfo=None)

    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    wb.properties.created = naive_dt
    wb.properties.modified = naive_dt
    wb.save(xlsx_path)
    _normalize_zip_timestamps(xlsx_path, build_dt)
    return xlsx_path


def read_reproducibility(xlsx_path: Path | str) -> dict:
    """Extract stored metadata from a workbook (for `modelforge verify`)."""
    xlsx_path = Path(xlsx_path)
    wb = load_workbook(xlsx_path, data_only=False, keep_links=True)
    if SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[SHEET_NAME]
    out: dict[str, str] = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=2).value
        if label and value is not None:
            out[str(label)] = str(value)
    return out


def verify_spec_hash(
    xlsx_path: Path | str,
    spec_source_bytes: bytes,
) -> tuple[bool, str, str]:
    """Compare stored SHA-256 to a recomputed one from given YAML bytes.

    Returns (match, stored, recomputed).
    """
    stored_meta = read_reproducibility(xlsx_path)
    stored = stored_meta.get("Spec SHA-256", "")
    recomputed = hashlib.sha256(spec_source_bytes).hexdigest()
    return stored == recomputed, stored, recomputed


__all__ = [
    "SHEET_NAME",
    "ENV_SOURCE_DATE_EPOCH",
    "ENV_WALL_CLOCK",
    "compute_spec_hash",
    "append_reproducibility_block",
    "finalize_determinism",
    "read_reproducibility",
    "verify_spec_hash",
]
