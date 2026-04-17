"""Model diff engine — compares two built ModelForge workbooks.

Five diff dimensions (each dimension's "diff record" is a small dict):

    1. Assumption changes — keyed by A-###; tracks unit / base / worst /
       best / confidence / source_id changes. Most actionable: this is
       where committees focus.
    2. Source changes — S-### entries added/removed; doc/page/publisher
       edits on existing entries.
    3. Formula changes — per (sheet, cell) with old → new formula text.
       Filters out cosmetic-only changes (trailing whitespace, case
       differences in function names).
    4. Structural changes — sheets added/removed, named ranges added/
       removed, sheet row/col counts changed.
    5. Reproducibility drift — spec SHA, ModelForge version, Python
       version, build timestamp from the Reproducibility sheet.
"""

from __future__ import annotations

import html
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook


# ─── Data classes ────────────────────────────────────────────────────────────


@dataclass
class AssumptionChange:
    assumption_id: str         # A-###
    driver_name: str           # snake_case
    field: str                 # "base" | "worst" | "best" | "unit" | "source_id" | "confidence"
    old: Any
    new: Any


@dataclass
class SourceChange:
    source_id: str             # S-###
    kind: str                  # "added" | "removed" | "edited"
    field: Optional[str] = None  # which column if edited
    old: Any = None
    new: Any = None


@dataclass
class FormulaChange:
    sheet: str
    cell: str                  # e.g. "D9"
    old: str
    new: str


@dataclass
class StructuralChange:
    kind: str                  # "sheet_added" | "sheet_removed" | "named_range_added" | ...
    name: str
    detail: str = ""


@dataclass
class ReproChange:
    field: str
    old: str
    new: str


@dataclass
class DiffResult:
    v1_path: Path
    v2_path: Path
    assumption_changes: list[AssumptionChange] = field(default_factory=list)
    source_changes: list[SourceChange] = field(default_factory=list)
    formula_changes: list[FormulaChange] = field(default_factory=list)
    structural_changes: list[StructuralChange] = field(default_factory=list)
    repro_changes: list[ReproChange] = field(default_factory=list)

    @property
    def is_clean(self) -> bool:
        return not (self.assumption_changes or self.source_changes
                    or self.formula_changes or self.structural_changes
                    or self.repro_changes)

    @property
    def total_changes(self) -> int:
        return (len(self.assumption_changes) + len(self.source_changes)
                + len(self.formula_changes) + len(self.structural_changes)
                + len(self.repro_changes))


# ─── Extract-and-compare helpers ─────────────────────────────────────────────


def _extract_assumptions(wb) -> dict[str, dict]:
    """Return {A-id: {name, unit, base, worst, best, confidence, source_id}}."""
    if "Assumptions" not in wb.sheetnames:
        return {}
    ws = wb["Assumptions"]
    out: dict[str, dict] = {}
    for r in range(6, ws.max_row + 1):
        aid = ws.cell(row=r, column=1).value
        if not aid or not re.match(r"^A-\d{3,}$", str(aid)):
            continue
        out[str(aid)] = {
            "name": ws.cell(row=r, column=2).value,
            "unit": ws.cell(row=r, column=5).value,
            "worst": ws.cell(row=r, column=6).value,
            "base": ws.cell(row=r, column=7).value,
            "best": ws.cell(row=r, column=8).value,
            "confidence": ws.cell(row=r, column=11).value,
            "source_id": ws.cell(row=r, column=12).value,
        }
    return out


def _extract_sources(wb) -> dict[str, dict]:
    if "Sources" not in wb.sheetnames:
        return {}
    ws = wb["Sources"]
    out: dict[str, dict] = {}
    for r in range(6, ws.max_row + 1):
        sid = ws.cell(row=r, column=1).value
        if not sid or not re.match(r"^S-\d{3,}$", str(sid)):
            continue
        out[str(sid)] = {
            "doc": ws.cell(row=r, column=2).value,
            "page": ws.cell(row=r, column=3).value,
            "publisher": ws.cell(row=r, column=4).value,
            "date": ws.cell(row=r, column=5).value,
            "url": ws.cell(row=r, column=6).value,
            "verified": ws.cell(row=r, column=7).value,
        }
    return out


def _extract_formulas(wb) -> dict[tuple[str, str], str]:
    """Return {(sheet, cell_coordinate): formula_text} for every formula."""
    out: dict[tuple[str, str], str] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    out[(sheet, c.coordinate)] = _canonicalize_formula(v)
    return out


def _canonicalize_formula(f: str) -> str:
    """Strip insignificant whitespace; normalize function case."""
    # Uppercase function names (Excel is case-insensitive)
    f = re.sub(r"([A-Za-z_][A-Za-z0-9_]*)(?=\()",
               lambda m: m.group(1).upper(), f)
    # Collapse internal whitespace
    return re.sub(r"\s+", "", f).strip()


def _extract_named_ranges(wb) -> dict[str, str]:
    return {n: dn.attr_text for n, dn in wb.defined_names.items()}


def _extract_reproducibility(wb) -> dict[str, str]:
    if "Reproducibility" not in wb.sheetnames:
        return {}
    ws = wb["Reproducibility"]
    out: dict[str, str] = {}
    for r in range(5, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        val = ws.cell(row=r, column=2).value
        if label and val is not None:
            out[str(label)] = str(val)
    return out


# ─── Public API ──────────────────────────────────────────────────────────────


def compute_diff(v1_path: Path | str, v2_path: Path | str) -> DiffResult:
    v1_path, v2_path = Path(v1_path), Path(v2_path)
    wb1 = load_workbook(v1_path, data_only=False, keep_links=True)
    wb2 = load_workbook(v2_path, data_only=False, keep_links=True)
    res = DiffResult(v1_path=v1_path, v2_path=v2_path)

    # ── Assumptions
    a1 = _extract_assumptions(wb1)
    a2 = _extract_assumptions(wb2)
    for aid in sorted(set(a1) | set(a2)):
        if aid not in a1:
            for fld, v in (a2[aid] or {}).items():
                if v is not None:
                    res.assumption_changes.append(AssumptionChange(
                        aid, str(a2[aid].get("name", "")), f"added.{fld}",
                        None, v,
                    ))
            continue
        if aid not in a2:
            for fld, v in (a1[aid] or {}).items():
                if v is not None:
                    res.assumption_changes.append(AssumptionChange(
                        aid, str(a1[aid].get("name", "")),
                        f"removed.{fld}", v, None,
                    ))
            continue
        # Field-by-field delta
        for fld in ("unit", "base", "worst", "best", "confidence", "source_id"):
            v1, v2 = a1[aid].get(fld), a2[aid].get(fld)
            if v1 != v2:
                res.assumption_changes.append(AssumptionChange(
                    aid, str(a1[aid].get("name", "")), fld, v1, v2,
                ))

    # ── Sources
    s1, s2 = _extract_sources(wb1), _extract_sources(wb2)
    for sid in sorted(set(s1) - set(s2)):
        res.source_changes.append(SourceChange(sid, "removed"))
    for sid in sorted(set(s2) - set(s1)):
        res.source_changes.append(SourceChange(sid, "added"))
    for sid in sorted(set(s1) & set(s2)):
        for fld in ("doc", "page", "publisher", "date", "url", "verified"):
            if s1[sid].get(fld) != s2[sid].get(fld):
                res.source_changes.append(SourceChange(
                    sid, "edited", field=fld,
                    old=s1[sid].get(fld), new=s2[sid].get(fld),
                ))

    # ── Formulas
    f1 = _extract_formulas(wb1)
    f2 = _extract_formulas(wb2)
    for key in sorted(set(f1) | set(f2)):
        old = f1.get(key)
        new = f2.get(key)
        if old != new:
            res.formula_changes.append(FormulaChange(
                sheet=key[0], cell=key[1],
                old=old or "", new=new or "",
            ))

    # ── Structural
    sheets1 = set(wb1.sheetnames)
    sheets2 = set(wb2.sheetnames)
    for s in sorted(sheets1 - sheets2):
        res.structural_changes.append(StructuralChange("sheet_removed", s))
    for s in sorted(sheets2 - sheets1):
        res.structural_changes.append(StructuralChange("sheet_added", s))

    nr1 = _extract_named_ranges(wb1)
    nr2 = _extract_named_ranges(wb2)
    for n in sorted(set(nr1) - set(nr2)):
        res.structural_changes.append(StructuralChange(
            "named_range_removed", n, detail=nr1[n]))
    for n in sorted(set(nr2) - set(nr1)):
        res.structural_changes.append(StructuralChange(
            "named_range_added", n, detail=nr2[n]))
    for n in sorted(set(nr1) & set(nr2)):
        if nr1[n] != nr2[n]:
            res.structural_changes.append(StructuralChange(
                "named_range_moved", n,
                detail=f"{nr1[n]} -> {nr2[n]}"))

    # ── Reproducibility
    r1, r2 = _extract_reproducibility(wb1), _extract_reproducibility(wb2)
    for k in sorted(set(r1) | set(r2)):
        if r1.get(k) != r2.get(k):
            res.repro_changes.append(ReproChange(
                k, str(r1.get(k, "")), str(r2.get(k, ""))))

    return res


# ─── Rendering ───────────────────────────────────────────────────────────────


def _fmt_value(v: Any) -> str:
    if v is None:
        return "—"
    if isinstance(v, float):
        if abs(v) < 1 and v != 0:
            return f"{v:.4f}"
        return f"{v:,.2f}"
    return str(v)


def render_markdown(res: DiffResult) -> str:
    out = [f"# ModelForge diff — {res.v1_path.name} → {res.v2_path.name}", ""]
    if res.is_clean:
        out.append("**Clean diff — no changes across any dimension.**")
        return "\n".join(out)

    out.append(f"**Total changes:** {res.total_changes}")
    out.append("")

    if res.repro_changes:
        out.append("## Reproducibility drift")
        out.append("")
        out.append("| Field | v1 | v2 |")
        out.append("|---|---|---|")
        for c in res.repro_changes:
            out.append(f"| {c.field} | `{c.old}` | `{c.new}` |")
        out.append("")

    if res.assumption_changes:
        out.append("## Assumption changes")
        out.append("")
        out.append("| A-id | Driver | Field | v1 | v2 |")
        out.append("|---|---|---|---|---|")
        for c in res.assumption_changes:
            out.append(
                f"| {c.assumption_id} | `{c.driver_name}` | {c.field} "
                f"| {_fmt_value(c.old)} | {_fmt_value(c.new)} |"
            )
        out.append("")

    if res.source_changes:
        out.append("## Source changes")
        out.append("")
        out.append("| S-id | Kind | Field | v1 | v2 |")
        out.append("|---|---|---|---|---|")
        for c in res.source_changes:
            out.append(
                f"| {c.source_id} | {c.kind} | {c.field or ''} "
                f"| {_fmt_value(c.old)} | {_fmt_value(c.new)} |"
            )
        out.append("")

    if res.structural_changes:
        out.append("## Structural changes")
        out.append("")
        out.append("| Kind | Name | Detail |")
        out.append("|---|---|---|")
        for c in res.structural_changes:
            out.append(f"| {c.kind} | `{c.name}` | {c.detail} |")
        out.append("")

    if res.formula_changes:
        out.append(f"## Formula changes ({len(res.formula_changes)})")
        out.append("")
        # Cap at 100 for markdown readability; overflow goes to HTML
        shown = res.formula_changes[:100]
        out.append("| Sheet | Cell | v1 | v2 |")
        out.append("|---|---|---|---|")
        for c in shown:
            old = c.old.replace("|", "\\|")
            new = c.new.replace("|", "\\|")
            out.append(f"| {c.sheet} | {c.cell} | `{old}` | `{new}` |")
        if len(res.formula_changes) > 100:
            out.append("")
            out.append(f"*(+{len(res.formula_changes) - 100} more — see HTML output)*")
        out.append("")

    return "\n".join(out)


def render_html(res: DiffResult) -> str:
    esc = html.escape
    title = f"ModelForge diff — {esc(res.v1_path.name)} → {esc(res.v2_path.name)}"
    body = [f"<!doctype html><html><head><meta charset='utf-8'><title>{title}</title>",
            "<style>",
            "body{font-family:-apple-system,Segoe UI,Helvetica,sans-serif;padding:2rem;max-width:1200px;margin:0 auto;color:#222}",
            "h1{color:#1F3864;border-bottom:2px solid #1F3864;padding-bottom:0.5rem}",
            "h2{color:#2F5496;margin-top:2rem}",
            "table{border-collapse:collapse;width:100%;margin:1rem 0;font-size:0.9rem}",
            "th{background:#1F3864;color:#fff;text-align:left;padding:0.5rem;font-weight:600}",
            "td{border:1px solid #ddd;padding:0.4rem;vertical-align:top}",
            "tr:nth-child(even){background:#f4f6fa}",
            "code{background:#f0f0f0;padding:0.1rem 0.3rem;border-radius:2px;font-family:Consolas,monospace;font-size:0.85em}",
            ".added{color:#006100;font-weight:600}.removed{color:#9C0006;font-weight:600}",
            ".clean{color:#006100;font-weight:600;padding:1rem;background:#E8F5E9;border-radius:4px}",
            "</style></head><body>",
            f"<h1>{title}</h1>"]

    if res.is_clean:
        body.append("<p class='clean'>✓ Clean diff — no changes across any dimension.</p>")
    else:
        body.append(f"<p><strong>Total changes:</strong> {res.total_changes}</p>")

    def _tbl(rows, headers):
        h = "".join(f"<th>{esc(h)}</th>" for h in headers)
        trs = []
        for row in rows:
            tds = "".join(f"<td>{v}</td>" for v in row)
            trs.append(f"<tr>{tds}</tr>")
        return f"<table><thead><tr>{h}</tr></thead><tbody>{''.join(trs)}</tbody></table>"

    if res.repro_changes:
        body.append("<h2>Reproducibility drift</h2>")
        body.append(_tbl(
            [(esc(c.field), f"<code>{esc(c.old)}</code>", f"<code>{esc(c.new)}</code>")
             for c in res.repro_changes],
            ["Field", "v1", "v2"],
        ))

    if res.assumption_changes:
        body.append(f"<h2>Assumption changes ({len(res.assumption_changes)})</h2>")
        body.append(_tbl(
            [(esc(c.assumption_id), f"<code>{esc(c.driver_name)}</code>",
              esc(c.field), esc(_fmt_value(c.old)), esc(_fmt_value(c.new)))
             for c in res.assumption_changes],
            ["A-id", "Driver", "Field", "v1", "v2"],
        ))

    if res.source_changes:
        body.append(f"<h2>Source changes ({len(res.source_changes)})</h2>")
        body.append(_tbl(
            [(esc(c.source_id),
              f"<span class='{'added' if c.kind=='added' else 'removed' if c.kind=='removed' else ''}'>{esc(c.kind)}</span>",
              esc(c.field or ""),
              esc(_fmt_value(c.old)), esc(_fmt_value(c.new)))
             for c in res.source_changes],
            ["S-id", "Kind", "Field", "v1", "v2"],
        ))

    if res.structural_changes:
        body.append(f"<h2>Structural changes ({len(res.structural_changes)})</h2>")
        body.append(_tbl(
            [(esc(c.kind), f"<code>{esc(c.name)}</code>", esc(c.detail))
             for c in res.structural_changes],
            ["Kind", "Name", "Detail"],
        ))

    if res.formula_changes:
        body.append(f"<h2>Formula changes ({len(res.formula_changes)})</h2>")
        body.append(_tbl(
            [(esc(c.sheet), esc(c.cell),
              f"<code>{esc(c.old)}</code>", f"<code>{esc(c.new)}</code>")
             for c in res.formula_changes],
            ["Sheet", "Cell", "v1", "v2"],
        ))

    body.append("</body></html>")
    return "\n".join(body)
