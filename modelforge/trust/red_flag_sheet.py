"""Auto-inject a 'RedFlags' worksheet into a built workbook.

Reviewers see the violation list at the top of the workbook without
having to rerun the engine. Format mirrors the QC sheet for consistency.
"""

from __future__ import annotations

from pathlib import Path
from typing import Union

from modelforge.trust.violations import TrustReport


_SHEET_NAME = "RedFlags"

_SEVERITY_COLOR = {
    "fail": "FFC7CE",   # excel red-fill background
    "warn": "FFEB9C",   # yellow
    "info": "DDEBF7",   # light blue
}
_SEVERITY_LABEL = {"fail": "FAIL", "warn": "WARN", "info": "INFO"}


def inject_red_flag_sheet(xlsx_path: Union[str, Path], report: TrustReport) -> Path:
    """Open ``xlsx_path``, replace any existing RedFlags sheet with the report,
    save in place. Returns the path."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    path = Path(xlsx_path)
    wb = load_workbook(path)
    if _SHEET_NAME in wb.sheetnames:
        del wb[_SHEET_NAME]
    ws = wb.create_sheet(_SHEET_NAME, index=0)  # surface at the front

    # Headers + summary
    summary = report.summary()
    title_font = Font(bold=True, size=14, color="000000")
    label_font = Font(bold=True, size=10)
    body_font = Font(size=10)
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.cell(row=1, column=1, value="ModelForge Trust Layer — Red Flags").font = title_font
    ws.cell(row=2, column=1, value=f"Workbook: {Path(report.workbook).name}").font = body_font
    ws.cell(row=3, column=1, value=f"Template: {report.template}").font = body_font
    ws.cell(row=4, column=1,
            value=(f"Rules run: {summary['rules_run']}  •  "
                   f"FAIL: {summary['fail']}  WARN: {summary['warn']}  INFO: {summary['info']}")
            ).font = label_font

    if report.error_messages:
        ws.cell(row=5, column=1,
                value=f"Engine errors: {len(report.error_messages)}  (rule misconfigurations, not data issues)"
                ).font = body_font

    # Table headers
    hr = 7
    headers = ["#", "Severity", "Rule", "Cell / Anchor", "Actual",
               "Expected", "Message", "Recommendation"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = label_font
        c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center")

    # Body rows
    if not report.violations:
        ws.cell(row=hr + 1, column=2, value="ALL CLEAR").font = Font(bold=True, color="00B050")
        ws.cell(row=hr + 1, column=3, value="No plausibility violations triggered").font = body_font
    else:
        for i, v in enumerate(report.violations, start=1):
            row = hr + i
            fill_color = _SEVERITY_COLOR.get(v.severity, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            cells = [
                (1, i),
                (2, _SEVERITY_LABEL.get(v.severity, v.severity.upper())),
                (3, v.rule_name),
                (4, v.cell or ""),
                (5, _fmt_num(v.actual)),
                (6, _fmt_band(v.expected_low, v.expected_high)),
                (7, v.message),
                (8, v.recommendation or ""),
            ]
            for col, val in cells:
                c = ws.cell(row=row, column=col, value=val)
                c.fill = fill
                c.border = border
                c.font = body_font
                c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            ws.row_dimensions[row].height = 30

    # Column widths
    widths = [4, 10, 32, 22, 14, 18, 60, 60]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{hr + 1}"
    ws.print_title_rows = f"{hr}:{hr}"

    wb.save(path)
    return path


def _fmt_num(v):
    if v is None:
        return ""
    try:
        if abs(v) >= 1000:
            return f"{v:,.2f}"
        if abs(v) < 0.01:
            return f"{v:.4f}"
        return f"{v:.4f}"
    except (TypeError, ValueError):
        return str(v)


def _fmt_band(low, high):
    if low is None and high is None:
        return ""
    if low is None:
        return f"≤ {_fmt_num(high)}"
    if high is None:
        return f"≥ {_fmt_num(low)}"
    return f"[{_fmt_num(low)}, {_fmt_num(high)}]"
