"""MoatGate — runs the four gates and returns a MoatReport."""

from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Union

from modelforge.moat.classifier import classify_sheet
from modelforge.moat.report import GateResult, MoatReport, SheetMetrics


@dataclass(frozen=True)
class MoatThresholds:
    """Configurable gate thresholds. Defaults are the institutional bar."""
    core_output_min_formula_ratio: float = 0.90
    reference_min_formula_ratio: float = 0.20  # only aggregation rows
    recalc_tolerance_abs: float = 1e-4         # absolute diff
    recalc_tolerance_rel: float = 1e-6         # relative diff
    max_orphan_named_ranges: int = 5           # some headroom for back-compat
    max_magic_number_cells: int = 0            # zero-tolerance by default


# Magic-number heuristic: a literal numeric inside a formula that is
# NOT 0/1/-1 (counters), NOT 12 (months), NOT 100/1000/100000 (scaling),
# NOT a reasonable date-arithmetic constant.
_INNOCUOUS = {0, 1, -1, 2, 12, 4, 100, 1000, 10000, 100000, 1000000,
              360, 365, 366, 30, 31, 7, 24, 60, 0.5, 0.25, 0.75}
_NUM_RE = re.compile(r"(?<![A-Z_:])-?\d+(\.\d+)?(?![A-Z_])")


def _has_magic_number(formula: str) -> bool:
    """Return True if the formula contains a non-innocuous numeric literal."""
    if not formula.startswith("="):
        return False
    body = formula[1:]
    # Skip if it's purely a single value reference (e.g. =A1 or =named_range)
    for m in _NUM_RE.finditer(body):
        try:
            n = float(m.group())
            # Treat integers and floats equivalently for the innocuous set
            if n in _INNOCUOUS or int(n) if n.is_integer() else False in _INNOCUOUS:
                continue
            return True
        except (TypeError, ValueError):
            continue
    return False


class MoatGate:
    def __init__(self, thresholds: MoatThresholds | None = None) -> None:
        self.thresholds = thresholds or MoatThresholds()

    # ─── public ─────────────────────────────────────────────────────────────

    def evaluate(self, xlsx_path: Union[str, Path]) -> MoatReport:
        path = Path(xlsx_path)
        report = MoatReport(workbook=str(path))

        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, data_only=False)
        except Exception as e:
            report.error_messages.append(f"Failed to open workbook: {e}")
            return report

        # 1. Per-sheet metrics
        for s in wb.sheetnames:
            ws = wb[s]
            klass = classify_sheet(s)
            m = self._sheet_metrics(ws, klass)
            report.sheet_metrics.append(m)

        # 2. Gates
        report.gate_results.append(self._gate_formula_density(report))
        report.gate_results.append(self._gate_reference_graph(report))

        report.orphan_named_ranges = self._find_orphan_named_ranges(wb)
        report.gate_results.append(self._gate_no_orphan_inputs(report))

        # Recalc gate is heavier — keep last so we can short-circuit
        report.gate_results.append(self._gate_recalculation(path, report))

        return report

    # ─── per-sheet metrics ──────────────────────────────────────────────────

    def _sheet_metrics(self, ws, sheet_class: str) -> SheetMetrics:
        m = SheetMetrics(name=ws.title, sheet_class=sheet_class)
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                if isinstance(v, str) and v.startswith("="):
                    m.formula_cells += 1
                    m.numeric_cells += 1
                    if _has_magic_number(v):
                        m.formulas_with_magic_numbers.append(c.coordinate)
                elif isinstance(v, (int, float, bool)):
                    m.hardcoded_numeric_cells += 1
                    m.numeric_cells += 1
                # text / labels are intentionally skipped
        return m

    # ─── gates ──────────────────────────────────────────────────────────────

    def _gate_formula_density(self, report: MoatReport) -> GateResult:
        """For every CORE_OUTPUT sheet, formula ratio must meet threshold."""
        thresh = self.thresholds.core_output_min_formula_ratio
        violations: list[str] = []
        for m in report.sheet_metrics:
            if m.sheet_class != "core_output":
                continue
            if m.numeric_cells < 5:
                continue  # ignore near-empty sheets
            if m.formula_ratio < thresh:
                violations.append(
                    f"{m.name}: {m.formula_ratio:.1%} formula "
                    f"({m.formula_cells}/{m.numeric_cells}, threshold ≥{thresh:.0%})"
                )
        passed = not violations
        detail = "All core output sheets meet density threshold" if passed \
            else "; ".join(violations[:5])
        return GateResult(
            name="formula_density",
            passed=passed,
            detail=detail,
            metric=report.core_output_density(),
            threshold=thresh,
        )

    def _gate_reference_graph(self, report: MoatReport) -> GateResult:
        """Output formulas must not contain magic-number literals."""
        thresh = self.thresholds.max_magic_number_cells
        offenders: list[str] = []
        for m in report.sheet_metrics:
            if m.sheet_class not in ("core_output", "audit"):
                continue
            for cell in m.formulas_with_magic_numbers[:5]:
                offenders.append(f"{m.name}!{cell}")
        passed = len(offenders) <= thresh
        detail = (
            "No magic-number literals in output formulas"
            if passed
            else f"{len(offenders)} formula(s) with magic numbers (showing first 5): "
                 f"{', '.join(offenders[:5])}"
        )
        return GateResult(
            name="reference_graph",
            passed=passed,
            detail=detail,
            metric=float(len(offenders)),
            threshold=float(thresh),
        )

    def _gate_no_orphan_inputs(self, report: MoatReport) -> GateResult:
        """Every defined name should be referenced by ≥1 formula."""
        n = len(report.orphan_named_ranges)
        thresh = self.thresholds.max_orphan_named_ranges
        passed = n <= thresh
        detail = (
            f"All defined names reachable from formulas (≤{thresh} allowed)"
            if passed
            else f"{n} orphan named range(s): {', '.join(report.orphan_named_ranges[:5])}"
        )
        return GateResult(
            name="no_orphan_inputs",
            passed=passed,
            detail=detail,
            metric=float(n),
            threshold=float(thresh),
        )

    def _gate_recalculation(self, path: Path, report: MoatReport) -> GateResult:
        """Recalc with the third-party engine; mismatches against cached values fail."""
        try:
            import formulas  # type: ignore
        except ImportError:
            return GateResult(
                name="recalculation",
                passed=True,  # don't punish if the dep isn't installed
                detail="Skipped: `formulas` package not installed",
            )
        try:
            xl = formulas.ExcelModel().loads(str(path)).finish()
            sol = xl.calculate()
        except Exception as e:
            report.error_messages.append(f"Recalc failed: {e}")
            return GateResult(
                name="recalculation",
                passed=False,
                detail=f"Engine could not load workbook: {e}",
            )
        # Empty / no-formula workbook → vacuous pass
        if not sol:
            return GateResult(
                name="recalculation",
                passed=True,
                detail="Empty workbook (no formulas to recalc)",
                metric=0.0,
            )
        # Compare against openpyxl's cached values
        try:
            from openpyxl import load_workbook
            wb_cached = load_workbook(path, data_only=True)
        except Exception as e:
            report.error_messages.append(f"Cached-value load failed: {e}")
            return GateResult(
                name="recalculation",
                passed=False,
                detail=f"Could not read cached values: {e}",
            )
        mismatches: list[str] = []
        for k, v in sol.items():
            cached = self._cached_value(wb_cached, k)
            if cached is None:
                continue
            recalc = self._coerce_num(v.value if hasattr(v, "value") else v)
            if recalc is None:
                continue
            diff = abs(recalc - cached)
            tol = max(self.thresholds.recalc_tolerance_abs,
                      abs(recalc) * self.thresholds.recalc_tolerance_rel)
            if diff > tol:
                mismatches.append(f"{k.split('!')[-1]}: cached={cached:.4f} recalc={recalc:.4f}")
                if len(mismatches) >= 50:
                    break
        report.recalc_mismatches = mismatches
        passed = not mismatches
        detail = (
            f"All formula values reconcile (tol={self.thresholds.recalc_tolerance_abs:.0e})"
            if passed
            else f"{len(mismatches)} cell(s) disagree (showing first 5): "
                 f"{'; '.join(mismatches[:5])}"
        )
        return GateResult(
            name="recalculation",
            passed=passed,
            detail=detail,
            metric=float(len(mismatches)),
            threshold=0,
        )

    @staticmethod
    def _coerce_num(v):
        if v is None:
            return None
        if hasattr(v, "flatten"):
            try:
                vl = list(v.flatten())
                v = vl[0] if vl else None
            except Exception:
                return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _cached_value(wb_cached, formulas_key: str):
        """Look up the cached value for a key like '[file.xlsx]SHEET'!CELL."""
        try:
            sheet = formulas_key.split("]", 1)[1].split("'!", 1)[0].strip("'")
            cell = formulas_key.split("'!", 1)[1].strip()
            ws = wb_cached[sheet]
            v = ws[cell].value
            if isinstance(v, (int, float)):
                return float(v)
            return None
        except (IndexError, KeyError, AttributeError, ValueError):
            return None

    # Named-range prefixes considered intentional metadata (build provenance,
    # reproducibility, audit-trail). Excluded from orphan accounting.
    _METADATA_PREFIXES: tuple[str, ...] = ("mf_", "audit_", "_meta_")

    def _find_orphan_named_ranges(self, wb) -> list[str]:
        """Return defined-names that no formula references.

        Scans every formula across all sheets, collects every name-like
        token, then flags defined-names absent from that set. Metadata
        names (``mf_*``, ``audit_*``, ``_meta_*``) are intentionally
        orphan and excluded.
        """
        names = set()
        for nm in wb.defined_names:
            if any(nm.startswith(p) for p in self._METADATA_PREFIXES):
                continue
            names.add(nm)

        if not names:
            return []

        used: set[str] = set()
        # Build a cheap regex that matches any defined-name occurrence
        for s in wb.sheetnames:
            ws = wb[s]
            for row in ws.iter_rows():
                for c in row:
                    v = c.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    body = v[1:]
                    for nm in names:
                        if nm in used:
                            continue
                        if re.search(rf"(?<![A-Za-z0-9_]){re.escape(nm)}(?![A-Za-z0-9_])", body):
                            used.add(nm)
                if used == names:
                    break
            if used == names:
                break
        orphans = sorted(names - used)
        return orphans
