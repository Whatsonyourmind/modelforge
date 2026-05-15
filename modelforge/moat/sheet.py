"""Inject a 'MOAT' sheet into a workbook with the gate verdict."""

from __future__ import annotations

from pathlib import Path
from typing import Union

from modelforge.moat.report import MoatReport


_SHEET_NAME = "MOAT"


def inject_moat_sheet(xlsx_path: Union[str, Path], report: MoatReport) -> Path:
    """Add a MOAT sheet to the workbook so reviewers see the verdict in-place."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    path = Path(xlsx_path)
    wb = load_workbook(path)
    if _SHEET_NAME in wb.sheetnames:
        del wb[_SHEET_NAME]
    # Place after RedFlags / QC if they exist, else at the front
    pos = 0
    for preferred in ("RedFlags", "QC"):
        if preferred in wb.sheetnames:
            pos = wb.sheetnames.index(preferred) + 1
    ws = wb.create_sheet(_SHEET_NAME, index=pos)

    title_font = Font(bold=True, size=14)
    label_font = Font(bold=True, size=10)
    body = Font(size=10)
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    thin = Side(border_style="thin", color="888888")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "ModelForge MOAT — Live-Formula Verification"
    ws["A1"].font = title_font
    ws["A2"] = f"Workbook: {path.name}"
    ws["A2"].font = body
    ws["A3"] = (
        f"Core-output formula density: {report.core_output_density():.1%}"
    )
    ws["A3"].font = label_font
    overall = "PASS" if report.passes() else "FAIL"
    ws["A4"] = f"Overall: {overall}  ({sum(1 for g in report.gate_results if g.passed)}/{len(report.gate_results)} gates)"
    ws["A4"].font = label_font
    ws["A4"].fill = pass_fill if report.passes() else fail_fill

    # Gate table
    hr = 6
    headers = ["Gate", "Verdict", "Metric", "Threshold", "Detail"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=hr, column=i, value=h)
        c.font = label_font
        c.border = border
    for i, g in enumerate(report.gate_results, start=1):
        row = hr + i
        fill = pass_fill if g.passed else fail_fill
        for col, val in enumerate(
            [g.name, "PASS" if g.passed else "FAIL",
             _fmt(g.metric), _fmt(g.threshold), g.detail],
            start=1,
        ):
            c = ws.cell(row=row, column=col, value=val)
            c.fill = fill
            c.border = border
            c.font = body
            c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 30

    # Per-sheet metrics table
    sr = hr + len(report.gate_results) + 3
    ws.cell(row=sr - 1, column=1, value="Per-sheet formula density").font = label_font

    sub_headers = ["Sheet", "Class", "Numeric cells", "Formula cells", "Formula %", "Magic-number cells"]
    for i, h in enumerate(sub_headers, start=1):
        c = ws.cell(row=sr, column=i, value=h)
        c.font = label_font
        c.border = border

    for i, m in enumerate(report.sheet_metrics, start=1):
        row = sr + i
        cells = [
            m.name,
            m.sheet_class,
            m.numeric_cells,
            m.formula_cells,
            f"{m.formula_ratio*100:.1f}%" if m.numeric_cells else "n/a",
            len(m.formulas_with_magic_numbers),
        ]
        for col, val in enumerate(cells, start=1):
            c = ws.cell(row=row, column=col, value=val)
            c.font = body
            c.border = border
            c.alignment = Alignment(wrap_text=True, vertical="top")

    # Orphan + recalc summaries
    if report.orphan_named_ranges:
        orow = sr + len(report.sheet_metrics) + 3
        ws.cell(row=orow, column=1,
                value=f"Orphan named ranges ({len(report.orphan_named_ranges)}):").font = label_font
        ws.cell(row=orow + 1, column=1,
                value=", ".join(report.orphan_named_ranges[:30])
                + ("…" if len(report.orphan_named_ranges) > 30 else "")).font = body
    if report.recalc_mismatches:
        mrow = sr + len(report.sheet_metrics) + 5
        ws.cell(row=mrow, column=1,
                value=f"Recalc mismatches ({len(report.recalc_mismatches)}, first 10):").font = label_font
        for i, msg in enumerate(report.recalc_mismatches[:10], start=1):
            ws.cell(row=mrow + i, column=1, value=msg).font = body

    widths = [22, 14, 14, 14, 12, 60]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = f"A{hr + 1}"
    wb.save(path)
    return path


def _fmt(v) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        if 0 < abs(v) < 1:
            return f"{v:.4f}"
        return f"{v:,.2f}"
    return str(v)
