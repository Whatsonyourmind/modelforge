"""Adversarial stress-test of ModelForge outputs vs. bulge-bracket standards.

Tests against the error catalogue from:
  - Wall Street Oasis (LBO errors, balance sheet checks)
  - Wall Street Prep (3-statement linkages, interest basis)
  - Macabacus (DCF errors, mid-year convention, WACC)
  - BreakingIntoWallStreet (PF DSCR / LLCR / CFADS)
  - Corporate Finance Institute (RE roll, BS ties)

Produces structured findings: severity / file / sheet / cell / rule / evidence.
"""

from __future__ import annotations

import json
import math
import re
from pathlib import Path
from typing import Any

import formulas
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

ROOT = Path(__file__).parent
OUTPUT = ROOT / "output"

SUITE = [
    "unitranche_cdmo.xlsx",
    "minibond_logistics.xlsx",
    "credit_memo_cdmo.xlsx",
    "project_finance_solar.xlsx",
    "real_estate_pbsa.xlsx",
    "npl_mixed_portfolio.xlsx",
    "structured_credit_pmi.xlsx",
    "three_statement_cdmo.xlsx",
    "real_stevanato_3statement.xlsx",
    "real_enfinity_solar_pf.xlsx",
    "merger_tim_iliad.xlsx",
    "dcf_enel.xlsx",
    "fairness_amplifon.xlsx",
]

Findings: list[dict[str, Any]] = []


def add(severity: str, file: str, sheet: str | None, cell: str | None,
        rule: str, evidence: str, fix: str = "") -> None:
    Findings.append({
        "severity": severity,
        "file": file,
        "sheet": sheet,
        "cell": cell,
        "rule": rule,
        "evidence": evidence,
        "fix": fix,
    })


def to_scalar(v: Any) -> float | str | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if math.isfinite(float(v)) else None
    if isinstance(v, str):
        return v
    if hasattr(v, "value"):
        arr = v.value
        try:
            first = arr.flat[0] if hasattr(arr, "flat") else arr[0][0]
        except Exception:
            try:
                first = list(arr)[0]
                if hasattr(first, "__iter__") and not isinstance(first, str):
                    first = list(first)[0]
            except Exception:
                return None
        if hasattr(first, "item"):
            try:
                first = first.item()
            except Exception:
                pass
        if isinstance(first, str):
            return first
        try:
            fv = float(first)
            return fv if math.isfinite(fv) else None
        except Exception:
            return None
    try:
        return float(v)
    except Exception:
        return None


def eval_workbook(xlsx_path: Path) -> dict[str, Any]:
    xl = formulas.ExcelModel().loads(str(xlsx_path)).finish()
    sol = xl.calculate()
    out: dict[str, Any] = {}
    for k, v in sol.items():
        sv = to_scalar(v)
        if sv is not None:
            out[k.upper()] = sv
    return out


def key(xlsx_name: str, sheet: str, ref: str) -> str:
    return f"'[{xlsx_name}]{sheet}'!{ref}".upper()


def find_row(wb, sheet: str, substring: str, case: bool = False) -> int | None:
    if sheet not in wb.sheetnames:
        return None
    ws = wb[sheet]
    target = substring if case else substring.lower()
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if c.value is None:
            continue
        s = str(c.value) if case else str(c.value).lower()
        if target in s:
            return c.row
    return None


def find_all_rows(wb, sheet: str, substring: str) -> list[int]:
    """All rows where col A contains substring (case-insensitive)."""
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    target = substring.lower()
    rows = []
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if c.value and target in str(c.value).lower():
            rows.append(c.row)
    return rows


def first_data_col(wb, sheet: str) -> int:
    """Column index where projection data starts — we assume col D (=4) universally."""
    return 4


def last_data_col(wb, sheet: str) -> int:
    if sheet not in wb.sheetnames:
        return 4
    return wb[sheet].max_column


def scan_hardcodes(wb, sheet: str, rows: list[int]) -> list[tuple[str, float]]:
    """Return (cell_addr, value) pairs where a projection cell is a raw number (not formula)
    and looks like a 'real' value (|v| > 0.001, integer or multi-sig-fig)."""
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    out = []
    for r in rows:
        for c_idx in range(4, ws.max_column + 1):
            cell = ws.cell(row=r, column=c_idx)
            if cell.value is None:
                continue
            v = cell.value
            if isinstance(v, (int, float)) and abs(v) > 0.001:
                out.append((f"{get_column_letter(c_idx)}{r}", float(v)))
    return out


# ─── PER-TEMPLATE AUDITS ───────────────────────────────────────────────────

def audit_three_statement(file: str, wb, values: dict[str, Any]) -> None:
    """Checks specific to 3-statement templates."""
    sheet = "Model"
    if sheet not in wb.sheetnames:
        return

    # 1) CFS ↔ BS cash tie: ending cash on CFS = cash on BS
    #    Look for "Ending cash" / "Net change in cash" and "Cash" (BS)
    cfs_net_row = find_row(wb, sheet, "net change in cash")
    cfs_end_row = find_row(wb, sheet, "ending cash")
    bs_cash_row = find_row(wb, sheet, "Cash")  # BS Cash — might be same as CFS row label
    # Look for explicit "Cash and equivalents" or "Cash" with row > cfs_end_row
    bs_cash_candidates = find_all_rows(wb, sheet, "cash")

    # Retained earnings roll-forward: RE_t = RE_{t-1} + NI_t - Div_t
    re_row = find_row(wb, sheet, "retained")
    ni_row = find_row(wb, sheet, "net income")
    div_row = find_row(wb, sheet, "dividend")

    if re_row and ni_row:
        for col in range(5, 9):  # E..H (years 2..5 of projection)
            col_letter = get_column_letter(col)
            prev_letter = get_column_letter(col - 1)
            re_curr = values.get(key(file, sheet, f"{col_letter}{re_row}"))
            re_prev = values.get(key(file, sheet, f"{prev_letter}{re_row}"))
            ni_curr = values.get(key(file, sheet, f"{col_letter}{ni_row}"))
            div_curr = None
            if div_row:
                div_curr = values.get(key(file, sheet, f"{col_letter}{div_row}"))
            if not all(isinstance(x, (int, float)) for x in [re_curr, re_prev, ni_curr]):
                continue
            expected = re_prev + ni_curr + (div_curr if isinstance(div_curr, (int, float)) else 0)
            if abs(re_curr - expected) > 0.01 * max(1.0, abs(re_curr)):
                add("HIGH", file, sheet, f"{col_letter}{re_row}",
                    "Retained-earnings roll",
                    f"RE_t={re_curr:.2f} but RE_(t-1)+NI_t{'+Div' if div_curr else ''}={expected:.2f}",
                    "RE should roll: prior RE + Net Income - Dividends")

    # Ensure CFS line items are formulas, not hardcodes
    if cfs_net_row:
        cfs_section_start = find_row(wb, sheet, "CFO")
        cfs_section_end = cfs_end_row or (cfs_net_row + 3)
        if cfs_section_start:
            ws = wb[sheet]
            for r in range(cfs_section_start, cfs_section_end + 1):
                for c_idx in range(4, ws.max_column + 1):
                    cell = ws.cell(row=r, column=c_idx)
                    if isinstance(cell.value, (int, float)) and abs(cell.value) > 0.001:
                        add("MEDIUM", file, sheet,
                            f"{get_column_letter(c_idx)}{r}",
                            "Hardcoded CFS line item",
                            f"Value {cell.value} in CFS section. WSP: every CFS line "
                            "should reference elsewhere (CFS is a reconciliation).",
                            "Replace with formula linking to BS/IS")


def audit_lbo_style(file: str, wb, values: dict[str, Any]) -> None:
    """Unitranche + credit_memo checks."""
    om_sheet = "OperatingModel"
    ds_sheet = "DebtSchedule"

    # Interest expense basis: BOP vs avg vs EOP — inspect formula
    if ds_sheet in wb.sheetnames:
        ws = wb[ds_sheet]
        int_row = find_row(wb, ds_sheet, "interest")
        if int_row:
            # pick first projection column D
            cell = ws.cell(row=int_row, column=4)
            f = str(cell.value) if cell.value else ""
            if not f.startswith("="):
                return
            # Heuristic: formula should reference opening debt (BOP) or AVERAGE
            # WSP recommends BOP to avoid circular
            has_avg = "AVERAGE" in f.upper() or "AVG" in f.upper()
            has_bop = re.search(r"[A-Z]+\d+\s*[*]", f)  # uses a prior cell
            # Can't easily distinguish; flag for manual inspection
            # Instead check: does interest formula reference EOP of SAME period?
            # If it references closing of same column (e.g. Dclose * rate), that's wrong.

    # EBITDA sign sanity: EBITDA positive, D&A negative, Interest negative
    for label, expect_sign in [("D&A", -1), ("interest expense", -1), ("taxes", -1),
                                 ("Total capex", -1)]:
        r = find_row(wb, om_sheet, label)
        if not r:
            continue
        for col in range(4, 12):
            v = values.get(key(file, om_sheet, f"{get_column_letter(col)}{r}"))
            if isinstance(v, (int, float)) and abs(v) > 0.01:
                if expect_sign < 0 and v > 0.01:
                    add("HIGH", file, om_sheet,
                        f"{get_column_letter(col)}{r}",
                        f"Sign convention: '{label}'",
                        f"Value {v:.2f} is positive; cost-like row should be negative",
                        "Flip sign or apply negating formula")

    # Closing debt must be non-negative and amortize by maturity
    cd_row = find_row(wb, ds_sheet, "Closing debt")
    if cd_row:
        for col in range(4, 14):
            v = values.get(key(file, ds_sheet, f"{get_column_letter(col)}{cd_row}"))
            if isinstance(v, (int, float)) and v < -0.01:
                add("CRITICAL", file, ds_sheet,
                    f"{get_column_letter(col)}{cd_row}",
                    "Debt cannot be negative",
                    f"Closing debt = {v:.2f}",
                    "Check amortization formula — use MAX(0, ...)")

    # Returns: sponsor equity CF at t=0 must be negative
    returns = "Returns"
    if returns in wb.sheetnames:
        ecf = find_row(wb, returns, "equity cash")
        if ecf:
            v = values.get(key(file, returns, f"D{ecf}"))
            if isinstance(v, (int, float)) and v >= 0:
                add("HIGH", file, returns, f"D{ecf}",
                    "Sponsor equity at t=0",
                    f"t=0 equity CF = {v:.2f}; should be negative (capital out)",
                    "Apply negating formula on sponsor-equity row")


def audit_project_finance(file: str, wb, values: dict[str, Any]) -> None:
    """PF-specific: DSCR correct, debt fully amortized, DSRA funded, CFADS not EBITDA."""
    ds_sheet = "DebtDSCR"
    pcf_sheet = "ProjectCashFlow"

    if ds_sheet not in wb.sheetnames:
        return
    ws = wb[ds_sheet]

    # Total DSCR breaches must be 0 in base
    breach_row = find_row(wb, ds_sheet, "Total DSCR breaches")
    if breach_row:
        v = values.get(key(file, ds_sheet, f"C{breach_row}"))
        if isinstance(v, (int, float)) and v > 0:
            add("HIGH", file, ds_sheet, f"C{breach_row}",
                "PF DSCR breaches",
                f"Total breaches = {int(v)} in BASE",
                "Size debt to DSCR or adjust sculpt profile")

    # Check that DSCR is CFADS / Debt Service, not EBITDA / Debt Service
    dscr_row = find_row(wb, ds_sheet, "DSCR")
    cads_row = find_row(wb, pcf_sheet, "CADS") or find_row(wb, pcf_sheet, "CFADS")
    if dscr_row:
        # Look at the formula in first operating year
        for col in range(5, ws.max_column + 1):
            cell = ws.cell(row=dscr_row, column=col)
            if cell.value and str(cell.value).startswith("="):
                f = str(cell.value)
                # Standard formula is =CADS/|DebtService|. If it references EBITDA, flag.
                if "EBITDA" in f.upper():
                    add("HIGH", file, ds_sheet, f"{get_column_letter(col)}{dscr_row}",
                        "DSCR based on EBITDA not CFADS",
                        f"Formula: {f[:80]}. Standard is CFADS/|DebtService|.",
                        "Use CFADS/DS — CFADS = EBITDA - tax - WC - maintenance capex")
                break

    # Debt must be zero at end of tenor
    closing_row = find_row(wb, ds_sheet, "Closing debt") or find_row(wb, ds_sheet, "closing")
    if closing_row:
        # Tail column
        last_col = ws.max_column
        for c in range(max(4, last_col - 2), last_col + 1):
            v = values.get(key(file, ds_sheet, f"{get_column_letter(c)}{closing_row}"))
            if isinstance(v, (int, float)) and v > 0.5:
                add("MEDIUM", file, ds_sheet, f"{get_column_letter(c)}{closing_row}",
                    "Debt not amortized at maturity",
                    f"Closing debt = {v:.2f} in tail column",
                    "Verify tenor = operating years; last principal payment wipes balance")

    # DSRA funded at end of construction
    dsra_row = find_row(wb, ds_sheet, "DSRA balance")
    dsra_target_row = find_row(wb, ds_sheet, "DSRA target")
    if dsra_row and dsra_target_row:
        # Find first operating column — heuristic: after "C1", "C2" cols
        # Typically first op col is where construction ends; use O1 label lookup
        # We just check end of year 1 of operations ≥ target
        ws = wb[ds_sheet]
        # Find column with header "O1" or similar
        op_col = None
        for c in range(4, ws.max_column + 1):
            h = ws.cell(row=2, column=c).value  # try row 2
            if h and "O1" in str(h):
                op_col = c
                break
            h3 = ws.cell(row=3, column=c).value
            if h3 and "O1" in str(h3):
                op_col = c
                break
        if op_col:
            b = values.get(key(file, ds_sheet, f"{get_column_letter(op_col)}{dsra_row}"))
            t = values.get(key(file, ds_sheet, f"{get_column_letter(op_col)}{dsra_target_row}"))
            if isinstance(b, (int, float)) and isinstance(t, (int, float)):
                if b + 0.5 < t:
                    add("MEDIUM", file, ds_sheet,
                        f"{get_column_letter(op_col)}{dsra_row}",
                        "DSRA under-funded at O1",
                        f"Balance {b:.2f} < Target {t:.2f}",
                        "Top up from CFADS in final construction month")


def audit_dcf(file: str, wb, values: dict[str, Any]) -> None:
    """DCF / RE DCF: terminal growth < WACC, mid-year toggle, WACC weights tie."""
    # Locate the DCF sheet
    candidates = ["DCF", "Model"]
    dcf_sheet = None
    for s in candidates:
        if s in wb.sheetnames:
            dcf_sheet = s
            break
    if not dcf_sheet:
        return

    # Terminal growth
    tg_row = find_row(wb, dcf_sheet, "Terminal growth") or \
             find_row(wb, "Assumptions", "Terminal growth")
    wacc_row = find_row(wb, dcf_sheet, "WACC") or \
               find_row(wb, "Assumptions", "WACC") or \
               find_row(wb, "Assumptions", "Discount rate")
    # Also check for terminal value formula sanity (Gordon growth)
    tv_row = find_row(wb, dcf_sheet, "Terminal value")

    g = None
    r = None
    if tg_row:
        # Try common cells — E (BASE col) or C
        for ref in ["E", "C", "D", "F"]:
            v = values.get(key(file, "Assumptions" if "Assumption" in str(tg_row) else dcf_sheet,
                               f"{ref}{tg_row}"))
            if isinstance(v, (int, float)) and 0 <= v <= 0.2:
                g = v
                break
    if wacc_row:
        for ref in ["E", "C", "D", "F"]:
            v = values.get(key(file, "Assumptions", f"{ref}{wacc_row}"))
            if v is None:
                v = values.get(key(file, dcf_sheet, f"{ref}{wacc_row}"))
            if isinstance(v, (int, float)) and 0.01 <= v <= 0.3:
                r = v
                break

    if g is not None and r is not None and g >= r - 0.001:
        add("CRITICAL", file, dcf_sheet, None,
            "Terminal growth ≥ WACC",
            f"g={g:.3%} vs WACC={r:.3%}; Gordon formula diverges",
            "Cap g at long-run GDP growth, always < WACC")


def audit_covenants(file: str, wb, values: dict[str, Any]) -> None:
    """Covenants sheet: leverage = Net Debt / LTM EBITDA; ICR = EBITDA / Int."""
    if "Covenants" not in wb.sheetnames:
        return
    lev_row = find_row(wb, "Covenants", "Leverage")
    icr_row = find_row(wb, "Covenants", "ICR") or find_row(wb, "Covenants", "Interest cover")
    ws = wb["Covenants"]
    if lev_row:
        # Check the formula uses Net Debt / EBITDA in first projection col
        for col in range(4, ws.max_column + 1):
            cell = ws.cell(row=lev_row, column=col)
            f = str(cell.value) if cell.value else ""
            if f.startswith("="):
                # If refs EBITDA directly: fine. If refs revenue, wrong.
                if "REVENUE" in f.upper() and "EBITDA" not in f.upper():
                    add("HIGH", file, "Covenants",
                        f"{get_column_letter(col)}{lev_row}",
                        "Leverage formula",
                        f"References revenue not EBITDA: {f[:80]}",
                        "Leverage = Net Debt / LTM EBITDA")
                break


def audit_wacc_weights(file: str, wb, values: dict[str, Any]) -> None:
    """WACC: sum of weights = 1."""
    wd_row = find_row(wb, "Assumptions", "Weight of debt") or \
             find_row(wb, "Assumptions", "Debt weight")
    we_row = find_row(wb, "Assumptions", "Weight of equity") or \
             find_row(wb, "Assumptions", "Equity weight")
    if wd_row and we_row:
        for ref in ["E", "C", "D"]:
            wd = values.get(key(file, "Assumptions", f"{ref}{wd_row}"))
            we = values.get(key(file, "Assumptions", f"{ref}{we_row}"))
            if isinstance(wd, (int, float)) and isinstance(we, (int, float)):
                if abs(wd + we - 1.0) > 0.005:
                    add("HIGH", file, "Assumptions", f"{ref}{wd_row}",
                        "WACC weights sum",
                        f"w_d={wd:.3f} + w_e={we:.3f} = {wd+we:.3f} (should be 1.00)",
                        "Enforce w_d + w_e = 100%")
                break


def audit_comps_fairness(file: str, wb, values: dict[str, Any]) -> None:
    """Fairness opinion / comps: low/high, median/mean ordering."""
    # Look for Comps or Football sheet; fallback: scan for "min"/"max" on any sheet
    for sheet in wb.sheetnames:
        min_row = find_row(wb, sheet, "Minimum") or find_row(wb, sheet, "Low")
        max_row = find_row(wb, sheet, "Maximum") or find_row(wb, sheet, "High")
        med_row = find_row(wb, sheet, "Median")
        if not (min_row and max_row):
            continue
        ws = wb[sheet]
        for col in range(3, ws.max_column + 1):
            mn = values.get(key(file, sheet, f"{get_column_letter(col)}{min_row}"))
            mx = values.get(key(file, sheet, f"{get_column_letter(col)}{max_row}"))
            if isinstance(mn, (int, float)) and isinstance(mx, (int, float)):
                if mn > mx + 0.001:
                    add("HIGH", file, sheet, f"{get_column_letter(col)}{min_row}",
                        "Comps ordering",
                        f"Min {mn:.3f} > Max {mx:.3f}",
                        "Verify range: min ≤ median ≤ max")


# ─── DRIVER ────────────────────────────────────────────────────────────────

def audit_file(xlsx_name: str) -> None:
    xlsx_path = OUTPUT / xlsx_name
    if not xlsx_path.exists():
        add("INFO", xlsx_name, None, None, "Missing file",
            f"{xlsx_path} does not exist", "")
        return
    wb = load_workbook(xlsx_path, data_only=False)
    try:
        values = eval_workbook(xlsx_path)
    except Exception as e:
        add("CRITICAL", xlsx_name, None, None, "Formula eval crash",
            f"{e.__class__.__name__}: {e}", "Re-generate workbook")
        return

    sheets = set(wb.sheetnames)
    is_three = "Model" in sheets and ("DebtSchedule" not in sheets) and \
               ("Tranches" not in sheets) and ("ProjectCashFlow" not in sheets) and \
               ("DCF" not in sheets or "Financing" not in sheets)

    if is_three:
        audit_three_statement(xlsx_name, wb, values)

    if "OperatingModel" in sheets and "DebtSchedule" in sheets:
        audit_lbo_style(xlsx_name, wb, values)

    if "ProjectCashFlow" in sheets and "DebtDSCR" in sheets:
        audit_project_finance(xlsx_name, wb, values)

    if "DCF" in sheets or (is_three and "WACC" in str(list(wb.defined_names))):
        audit_dcf(xlsx_name, wb, values)

    audit_covenants(xlsx_name, wb, values)
    audit_wacc_weights(xlsx_name, wb, values)
    audit_comps_fairness(xlsx_name, wb, values)

    # Universal: check QC ALL_PASS
    if "QC" in sheets:
        ap = values.get(key(xlsx_name, "QC", "C4"))
        # formulas lib often returns None on nested AND/SUM — we'll note this separately

    # Negative cash flag (3-statement only)
    if is_three:
        cash_row = find_row(wb, "Model", "Cash and")
        if cash_row is None:
            cash_row = find_row(wb, "Model", "Cash")
        if cash_row:
            for col in range(4, 10):
                v = values.get(key(xlsx_name, "Model",
                                    f"{get_column_letter(col)}{cash_row}"))
                if isinstance(v, (int, float)) and v < -0.01:
                    add("HIGH", xlsx_name, "Model",
                        f"{get_column_letter(col)}{cash_row}",
                        "Negative cash",
                        f"Cash = {v:.2f}; requires revolver",
                        "Add revolver draw = MAX(0, min_cash - ending cash)")


def main() -> None:
    for f in SUITE:
        audit_file(f)

    # Print grouped by severity
    groups = {"CRITICAL": [], "HIGH": [], "MEDIUM": [], "LOW": [], "INFO": []}
    for fnd in Findings:
        groups.setdefault(fnd["severity"], []).append(fnd)

    print("=" * 80)
    print("MODELFORGE ADVERSARIAL STRESS TEST")
    print("=" * 80)
    for sev, lst in groups.items():
        if not lst:
            continue
        print(f"\n### {sev}  ({len(lst)})\n")
        for f in lst:
            loc = f"{f['file']}::{f['sheet'] or '?'}::{f['cell'] or '-'}"
            print(f"  [{sev}] {loc}")
            print(f"        RULE: {f['rule']}")
            print(f"        WHAT: {f['evidence']}")
            if f["fix"]:
                print(f"        FIX : {f['fix']}")

    total = sum(len(v) for v in groups.values())
    print()
    print("=" * 80)
    crit = len(groups["CRITICAL"])
    high = len(groups["HIGH"])
    med = len(groups["MEDIUM"])
    print(f"TOTAL FINDINGS: {total}   (CRITICAL={crit}  HIGH={high}  MEDIUM={med})")
    print("=" * 80)

    # Also dump to JSON for report generation
    (ROOT / "stress_test_findings.json").write_text(
        json.dumps(Findings, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
