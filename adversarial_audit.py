"""Final adversarial audit — detect specific error categories across all models."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent
OUTPUT = ROOT / "output"

FILES = [
    "unitranche_cdmo.xlsx",
    "minibond_logistics.xlsx",
    "credit_memo_cdmo.xlsx",
    "project_finance_solar.xlsx",
    "real_estate_pbsa.xlsx",
    "npl_mixed_portfolio.xlsx",
    "structured_credit_pmi.xlsx",
    "three_statement_cdmo.xlsx",
    "real_stevanato_3statement.xlsx",
    "real_enfinity_solar_pf.xlsx",
    "merger_tim_iliad.xlsx",
    "dcf_enel.xlsx",
    "fairness_amplifon.xlsx",
]

Findings: list[dict[str, Any]] = []


def add(sev: str, file: str, sheet: str | None, cell: str | None,
        rule: str, evidence: str, fix: str = "") -> None:
    Findings.append({
        "severity": sev, "file": file, "sheet": sheet, "cell": cell,
        "rule": rule, "evidence": evidence, "fix": fix,
    })


def cell_refs_in_formula(formula: str) -> list[tuple[str | None, str]]:
    """Parse (sheet, cell) refs from a formula string. Sheet may be None for same-sheet."""
    # Same-sheet: $A$1, A1, AA12
    # Cross-sheet: 'Sheet Name'!A1 or Sheet!A1
    refs = []
    # Cross-sheet with quotes
    for m in re.finditer(r"'([^']+)'\!\$?([A-Z]+)\$?(\d+)", formula):
        refs.append((m.group(1), f"{m.group(2)}{m.group(3)}"))
    # Cross-sheet no quotes
    for m in re.finditer(r"(?<!')([A-Za-z_][A-Za-z0-9_]*)\!\$?([A-Z]+)\$?(\d+)", formula):
        refs.append((m.group(1), f"{m.group(2)}{m.group(3)}"))
    # Same-sheet (avoid double-counting cross-sheet matches)
    remainder = re.sub(r"'[^']+'\!\$?[A-Z]+\$?\d+", "", formula)
    remainder = re.sub(r"[A-Za-z_][A-Za-z0-9_]*\!\$?[A-Z]+\$?\d+", "", remainder)
    for m in re.finditer(r"(?<![A-Za-z_])\$?([A-Z]+)\$?(\d+)(?!\s*\()", remainder):
        refs.append((None, f"{m.group(1)}{m.group(2)}"))
    return refs


def detect_same_period_circular(wb, sheet_name: str) -> list[tuple[str, str]]:
    """For each cell in the sheet, traverse dependencies to detect circular refs
    that stay within the same column (intra-period circular)."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    # Build graph: (row, col) → set of (sheet, row, col) it depends on
    graph: dict[tuple[int, int], set[tuple[int, int]]] = {}
    for row in ws.iter_rows():
        for cell in row:
            if not cell.value or not str(cell.value).startswith("="):
                continue
            refs = cell_refs_in_formula(str(cell.value))
            deps: set[tuple[int, int]] = set()
            for sheet, ref in refs:
                if sheet and sheet != sheet_name:
                    continue  # cross-sheet — skip for same-sheet circular
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                try:
                    col_str, row_num = coordinate_from_string(ref)
                    col_num = column_index_from_string(col_str)
                    deps.add((row_num, col_num))
                except Exception:
                    pass
            graph[(cell.row, cell.column)] = deps

    # DFS for cycles
    circulars: list[tuple[str, str]] = []
    visited: set[tuple[int, int]] = set()
    rec_stack: list[tuple[int, int]] = []

    def dfs(node: tuple[int, int]) -> bool:
        if node in rec_stack:
            # Found a cycle — capture it
            idx = rec_stack.index(node)
            cycle = rec_stack[idx:] + [node]
            addrs = [f"{get_column_letter(c)}{r}" for r, c in cycle]
            circulars.append((addrs[0], " → ".join(addrs)))
            return True
        if node in visited:
            return False
        visited.add(node)
        rec_stack.append(node)
        for dep in graph.get(node, set()):
            if dfs(dep):
                rec_stack.pop()
                return True
        rec_stack.pop()
        return False

    for node in graph:
        dfs(node)
    return circulars


def scan_hardcoded_magic_numbers(wb, sheet_name: str, excluded_rows: set[int] = None) -> list:
    """Find cells containing raw numbers with magnitude > 10 in non-Assumption sheets.
    These are candidate 'magic numbers' that should be named."""
    excluded_rows = excluded_rows or set()
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    hits = []
    for row in ws.iter_rows():
        if row[0].row in excluded_rows:
            continue
        for cell in row:
            if cell.column < 4:  # skip label cols
                continue
            v = cell.value
            if isinstance(v, (int, float)) and abs(v) >= 10:
                # Is it really a magic number, or is it OK (like historical data)?
                # We'll flag all and let human judgment apply
                hits.append((cell.coordinate, v))
    return hits


def audit_file(file: str) -> None:
    path = OUTPUT / file
    if not path.exists():
        add("INFO", file, None, None, "Missing", "File not found")
        return
    try:
        wb = load_workbook(path, data_only=False)
    except Exception as e:
        add("CRITICAL", file, None, None, "Load failure", str(e))
        return

    # 1) Iter calc setting
    iter_on = wb.calculation.iterate
    if iter_on is None or iter_on is False:
        # Only flag if a circular reference exists
        pass

    # 2) Same-period circular reference scan (only on debt-ish sheets to limit scope)
    for sh_name in wb.sheetnames:
        if sh_name in {"Cover", "Sources", "Assumptions", "QC", "Reproducibility",
                       "SensitivityAnalysis", "MonteCarlo"}:
            continue
        circ = detect_same_period_circular(wb, sh_name)
        if circ:
            for start, path in circ[:3]:  # cap at 3
                add("CRITICAL", file, sh_name, start,
                    "Circular reference (same-period)",
                    f"Cycle: {path}",
                    f"Enable iterative calc OR break by using prior-period reference. Iter currently: {iter_on}")

    # 3) Non-Assumption sheets: hardcoded numbers that look like parameters
    #    (not historical data rows). Heuristic: sheet is Valuation or DealStructure or FootballField
    #
    # v0.6: build a set of (sheet, cell) targets that are legitimately
    # raw-number inputs because they back a workbook-level DefinedName.
    # These cells are *intended* to hold a raw value; skip them.
    named_input_cells: set[tuple[str, str]] = set()
    for dn_name in wb.defined_names:
        dn = wb.defined_names[dn_name]
        for sheet_name, ref in dn.destinations:
            # ref like "$D$5" or "$D$5:$D$5"
            cell_ref = ref.replace("$", "").split(":")[0]
            named_input_cells.add((sheet_name, cell_ref))

    def _strip_cellrefs(formula: str) -> str:
        """Remove all cell refs from a formula so we can look for pure
        numeric literals. Cell refs are e.g. A1, $A$1, AB12, 'Sheet'!A1.
        This is a heuristic — good enough to avoid flagging row numbers
        like `16` in `$D$16`."""
        out = formula
        # Cross-sheet refs with quotes
        out = re.sub(r"'[^']+'\!\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?", "", out)
        # Cross-sheet without quotes
        out = re.sub(r"[A-Za-z_][A-Za-z0-9_]*\!\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?", "", out)
        # Same-sheet absolute and relative
        out = re.sub(r"\$?[A-Z]+\$?\d+(?::\$?[A-Z]+\$?\d+)?", "", out)
        return out

    for sh in ["Valuation", "DealStructure", "AccretionDilution", "FootballField",
               "ProForma"]:
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        # Scan formulas for embedded magic numbers (= not cell refs)
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                if (sh, cell.coordinate) in named_input_cells:
                    # Legitimate named-range input — skip
                    continue
                if isinstance(v, (int, float)) and abs(v) >= 10 and cell.column >= 4:
                    # Pure numeric value in a computation sheet
                    # Also skip if it's a historical-data label area (row has label of "Acquirer revenue", etc.)
                    label = ws.cell(row=cell.row, column=1).value
                    label_str = str(label or "").lower()
                    # Skip synergies/historical financials which are typical hardcoded
                    if any(k in label_str for k in ["synerg", "integration cost", "data tape",
                                                     "rating fees", "legal fees", "listing"]):
                        continue
                    add("MEDIUM", file, sh, cell.coordinate,
                        "Hardcoded magic number",
                        f"Value {v} in {sh}!{cell.coordinate} ({label})",
                        f"Replace with named range from Assumptions with source ID")
                elif isinstance(v, str) and v.startswith("="):
                    # Strip cell refs first — their row numbers (e.g. 16 in $D$16)
                    # must not be misread as magic numbers.
                    stripped = _strip_cellrefs(v)
                    for m in re.finditer(r"(?<![A-Z_])(\d{2,}(?:\.\d+)?)", stripped):
                        num = float(m.group(1))
                        if num >= 10:
                            label = ws.cell(row=cell.row, column=1).value
                            add("MEDIUM", file, sh, cell.coordinate,
                                "Magic number in formula",
                                f"{v[:70]} — embeds {num}",
                                "Replace embedded constant with named range from Assumptions")
                            break  # one per cell

    # 4) DCF-specific: mid-year convention
    if "Valuation" in wb.sheetnames and "FCFForecast" in wb.sheetnames:
        ws = wb["Valuation"]
        pv_row = None
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label and "explicit" in str(label).lower():
                pv_row = r
                break
        if pv_row:
            cell = ws.cell(row=pv_row, column=4)
            f = str(cell.value) if cell.value else ""
            if f.startswith("="):
                has_mid_year = ("0.5" in f) or ("mid" in f.lower()) or \
                               bool(re.search(r"\^\s*-\s*ROW.*0\.5", f))
                if not has_mid_year:
                    add("MEDIUM", file, "Valuation", cell.coordinate,
                        "No mid-year convention",
                        f"Explicit PV uses end-year discounting: {f[:80]}",
                        "Add mid-year toggle or use ^(-(t-0.5))")

    # 5) Fairness/FootballField: football field EV low/high should be derived, not hardcoded
    if "FootballField" in wb.sheetnames:
        ws = wb["FootballField"]
        for r in range(6, min(ws.max_row, 15) + 1):
            b = ws.cell(row=r, column=2).value  # EV low
            c = ws.cell(row=r, column=3).value  # EV high
            for col_name, v in [("B", b), ("C", c)]:
                if isinstance(v, (int, float)):
                    label = ws.cell(row=r, column=1).value
                    add("HIGH", file, "FootballField", f"{col_name}{r}",
                        "Football field EV hardcoded",
                        f"{label}: {col_name}{r}={v} — not derived from comps/DCF",
                        "Link to =MEDIAN(TradingComps!B6:B9) * EBITDA etc.")

    # 6) 3-statement: check debt is formulaic (roll-forward) not flat hardcoded
    if "Model" in wb.sheetnames:
        ws = wb["Model"]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label and str(label).lower().startswith("debt") and "service" not in str(label).lower():
                # All projection cols should be formulas
                flat_count = 0
                cell_addrs = []
                for c in range(4, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, (int, float)):
                        flat_count += 1
                        cell_addrs.append(ws.cell(row=r, column=c).coordinate)
                if flat_count >= 4:
                    add("MEDIUM", file, "Model", f"row {r}",
                        "Debt row is flat hardcoded",
                        f"{label}: {flat_count} cells with raw numbers ({cell_addrs[:4]}...)",
                        "Model debt roll-forward: BOP + draws - repay = EOP")

    # 7) Accretion/dilution: EPS should be formulaic
    if "AccretionDilution" in wb.sheetnames:
        ws = wb["AccretionDilution"]
        for r in range(1, ws.max_row + 1):
            label = ws.cell(row=r, column=1).value
            if label and "standalone" in str(label).lower() and "eps" in str(label).lower():
                hc = []
                for c in range(4, ws.max_column + 1):
                    v = ws.cell(row=r, column=c).value
                    if isinstance(v, (int, float)):
                        hc.append(ws.cell(row=r, column=c).coordinate)
                if len(hc) >= 2:
                    add("HIGH", file, "AccretionDilution", f"row {r}",
                        "Standalone EPS hardcoded",
                        f"{len(hc)} cells with raw EPS values",
                        "Compute EPS = NetIncome / Shares for transparency")

    # 8) Duplicate-column check (same formula in consecutive cols indicates off-by-one).
    #
    # v0.6: skip HISTORICAL columns (cols D-F on 3-statement, i.e. i < h).
    # In a flat-debt legacy 3-statement without a roll-forward schedule,
    # both historical interest cells legitimately reference the same BOP
    # named range. That pattern reads as a "duplicate" but isn't a bug
    # — it's an artefact of the model's flat-debt simplification (fixed
    # separately under US-075).
    for sh_name in ["AccretionDilution", "Model", "OperatingModel", "FCFForecast"]:
        if sh_name not in wb.sheetnames:
            continue
        ws = wb[sh_name]
        # Heuristic for 3-statement "Model" sheet: skip historical cols (D-F).
        skip_cols = {4, 5, 6} if sh_name == "Model" else set()
        for r in range(1, ws.max_row + 1):
            for c in range(5, ws.max_column):  # start at col E, check c vs c-1
                if c in skip_cols or (c - 1) in skip_cols:
                    continue
                v_curr = ws.cell(row=r, column=c).value
                v_prev = ws.cell(row=r, column=c - 1).value
                if (isinstance(v_curr, str) and isinstance(v_prev, str) and
                        v_curr.startswith("=") and v_prev.startswith("=")):
                    if v_curr == v_prev:
                        label = ws.cell(row=r, column=1).value
                        add("HIGH", file, sh_name,
                            f"{get_column_letter(c)}{r}",
                            "Duplicate formula (off-by-one candidate)",
                            f"{label}: formula identical to {get_column_letter(c-1)}{r}: {v_curr[:60]}",
                            "Verify this is intentional; usually indicates copy-paste error")


def main() -> None:
    for f in FILES:
        audit_file(f)

    groups = {"CRITICAL": [], "HIGH": [], "MEDIUM": [], "LOW": [], "INFO": []}
    for fnd in Findings:
        groups[fnd["severity"]].append(fnd)

    print("=" * 80)
    print("MODELFORGE — ADVERSARIAL AUDIT RESULTS")
    print("=" * 80)
    for sev in ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]:
        lst = groups[sev]
        if not lst:
            continue
        print(f"\n### {sev} ({len(lst)})\n")
        for f in lst:
            loc = f"{f['file']}::{f['sheet'] or '?'}!{f['cell'] or '-'}"
            print(f"  [{sev}] {loc}")
            print(f"        RULE: {f['rule']}")
            print(f"        WHAT: {f['evidence']}")
            if f["fix"]:
                print(f"        FIX : {f['fix']}")

    print()
    print("=" * 80)
    print(f"TOTAL: {sum(len(v) for v in groups.values())}  "
          f"(CRITICAL={len(groups['CRITICAL'])}  "
          f"HIGH={len(groups['HIGH'])}  "
          f"MEDIUM={len(groups['MEDIUM'])})")
    print("=" * 80)

    (ROOT / "adversarial_findings.json").write_text(
        json.dumps(Findings, indent=2), encoding="utf-8"
    )


if __name__ == "__main__":
    main()
