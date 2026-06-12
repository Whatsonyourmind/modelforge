# -*- coding: utf-8 -*-
"""
BM-DCF-001 run3 — IndustrialCo enterprise DCF, built from the brief alone.
Every derived cell is a live Excel formula. Wall Street color coding:
blue = hardcoded input, black = calculation, green = cross-sheet link.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\lukep\Desktop\Projects AI\ModelForge\benchmarks\artifacts\agent\BM-DCF-001_run3.xlsx"

# ---------------------------------------------------------------- styles
F = "Calibri"
BLUE = Font(color="0000FF", name=F, size=10)                  # inputs
BLUE_B = Font(color="0000FF", name=F, size=10, bold=True)
BLACK = Font(color="000000", name=F, size=10)                 # calcs
BLACK_B = Font(color="000000", name=F, size=10, bold=True)
GREEN = Font(color="006100", name=F, size=10)                 # sheet links
GREEN_B = Font(color="006100", name=F, size=10, bold=True)
LBL = Font(color="000000", name=F, size=10)
LBL_B = Font(color="000000", name=F, size=10, bold=True)
ITAL = Font(color="595959", name=F, size=9, italic=True)
TITLE_FONT = Font(color="FFFFFF", bold=True, name=F, size=12)
HDR_FONT = Font(color="FFFFFF", bold=True, name=F, size=10)
NAVY = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
GREY = PatternFill("solid", start_color="F2F2F2", end_color="F2F2F2")
YELL = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")   # headline outputs
THIN = Side(style="thin", color="BFBFBF")
TOPB = Border(top=Side(style="thin", color="000000"))
TOPDBL = Border(top=Side(style="double", color="000000"))
BOXB = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
RIGHT = Alignment(horizontal="right")
CENTER = Alignment(horizontal="center")

NUM = "#,##0.0;[Red](#,##0.0)"          # USD m, 1dp, red parens negatives
NUM2 = "#,##0.00;[Red](#,##0.00)"
PX = "$0.00"                            # per share
PCT1 = "0.0%"
PCT2 = "0.00%"
PCT3 = "0.000%"
DEC = "0.000000"                        # wacc as decimal
FACT = "0.0000"
X2 = "0.00\" x\""
INT = "0"

def put(ws, cell, value, font=BLACK, fmt=None, fill=None, align=None, border=None):
    c = ws[cell]
    c.value = value
    c.font = font
    if fmt:
        c.number_format = fmt
    if fill:
        c.fill = fill
    if align:
        c.alignment = align
    if border:
        c.border = border
    return c

def section(ws, row, text, last_col="H"):
    for col in range(1, ws[f"{last_col}1"].column + 1):
        cc = ws.cell(row=row, column=col)
        cc.fill = GREY
    put(ws, f"A{row}", text, LBL_B, fill=GREY)

def title(ws, text, last_col="H"):
    ws.merge_cells(f"A1:{last_col}1")
    put(ws, "A1", text, TITLE_FONT, fill=NAVY)
    for col in range(1, ws[f"{last_col}1"].column + 1):
        ws.cell(row=1, column=col).fill = NAVY
    put(ws, "A2", "Synthetic benchmark deal BM-DCF-001 - USD millions unless stated - "
                  "Blue = input | Black = calculation | Green = link", ITAL)

wb = Workbook()

# ================================================================ ASSUMPTIONS
wa = wb.active
wa.title = "Assumptions"
wa.sheet_properties.tabColor = "0070C0"
title(wa, "IndustrialCo - Enterprise DCF - Assumptions")

section(wa, 4, "Target & Capitalization")
put(wa, "A5", "Last FY revenue (FY0)");            put(wa, "B5", 1000.0, BLUE, NUM)
put(wa, "A6", "Last FY EBITDA (FY0)");             put(wa, "B6", 180.0, BLUE, NUM)
put(wa, "A7", "Net debt");                         put(wa, "B7", 250.0, BLUE, NUM)
put(wa, "A8", "Shares outstanding (m)");           put(wa, "B8", 150.0, BLUE, NUM)
put(wa, "A9", "Reference share price ($/sh)");     put(wa, "B9", 8.00, BLUE, PX)

section(wa, 11, "Operating Drivers (every projection year)")
put(wa, "A12", "D&A (% of revenue)");              put(wa, "B12", 0.06, BLUE, PCT1)
put(wa, "A13", "Capex (% of revenue)");            put(wa, "B13", 0.065, BLUE, PCT1)
put(wa, "A14", "Change in NWC (% of change in revenue)"); put(wa, "B14", 0.10, BLUE, PCT1)
put(wa, "A15", "Cash tax rate (on EBIT)");         put(wa, "B15", 0.25, BLUE, PCT1)

section(wa, 17, "Forecast Path (5 explicit years, revenue fade)")
for i in range(5):
    put(wa, f"{get_column_letter(3+i)}17", f"Year {i+1}", HDR_FONT, fill=NAVY, align=CENTER)
put(wa, "A18", "Revenue growth")
for i, gr in enumerate([0.04, 0.035, 0.03, 0.025, 0.02]):
    put(wa, f"{get_column_letter(3+i)}18", gr, BLUE, PCT1, align=RIGHT)
put(wa, "A19", "EBITDA margin")
for i in range(5):
    put(wa, f"{get_column_letter(3+i)}19", 0.18, BLUE, PCT1, align=RIGHT)

section(wa, 21, "WACC Inputs (plain CAPM - no extra premia)")
put(wa, "A22", "Risk-free rate (10Y UST)");        put(wa, "B22", 0.0425, BLUE, PCT2)
put(wa, "A23", "Equity risk premium");             put(wa, "B23", 0.05, BLUE, PCT2)
put(wa, "A24", "Beta (levered)");                  put(wa, "B24", 1.10, BLUE, NUM2)
put(wa, "A25", "Pre-tax cost of debt");            put(wa, "B25", 0.0575, BLUE, PCT2)
put(wa, "A26", "Tax rate (for WACC)");             put(wa, "B26", 0.25, BLUE, PCT1)
put(wa, "A27", "Target D/(D+E)");                  put(wa, "B27", 0.30, BLUE, PCT1)

section(wa, 29, "Terminal Value")
put(wa, "A30", "Terminal growth g (Gordon)");      put(wa, "B30", 0.02, BLUE, PCT1)
put(wa, "A31", "Note: capex = D&A in perpetuity; terminal NWC normalized at g", ITAL)

wa.column_dimensions["A"].width = 42
for col in "BCDEFGH":
    wa.column_dimensions[col].width = 12
wa.freeze_panes = "B4"
wa.print_area = "A1:H32"
wa.page_setup.orientation = "landscape"
wa.page_setup.fitToWidth = 1

# ================================================================ DCF
wd = wb.create_sheet("DCF")
wd.sheet_properties.tabColor = "1F4E79"
title(wd, "IndustrialCo - Enterprise DCF (end-of-year discounting, Gordon terminal)")

put(wd, "A3", "(USD m unless stated)", HDR_FONT, fill=NAVY)
put(wd, "B3", "FY0", HDR_FONT, fill=NAVY, align=CENTER)
for i in range(5):
    put(wd, f"{get_column_letter(3+i)}3", f"Year {i+1}", HDR_FONT, fill=NAVY, align=CENTER)
put(wd, "H3", "", HDR_FONT, fill=NAVY)

# ---- 1. WACC build
section(wd, 5, "1. WACC Build (CAPM)")
put(wd, "A6", "Risk-free rate (10Y UST)");   put(wd, "B6", "=Assumptions!B22", GREEN, PCT2)
put(wd, "A7", "Equity risk premium");        put(wd, "B7", "=Assumptions!B23", GREEN, PCT2)
put(wd, "A8", "Beta (levered)");             put(wd, "B8", "=Assumptions!B24", GREEN, NUM2)
put(wd, "A9", "Cost of equity (Ke = rf + beta x ERP)")
put(wd, "B9", "=B6+B8*B7", BLACK_B, PCT2)
put(wd, "A10", "Pre-tax cost of debt");      put(wd, "B10", "=Assumptions!B25", GREEN, PCT2)
put(wd, "A11", "Tax rate");                  put(wd, "B11", "=Assumptions!B26", GREEN, PCT1)
put(wd, "A12", "After-tax cost of debt (Kd x (1-t))")
put(wd, "B12", "=B10*(1-B11)", BLACK_B, PCT3)
put(wd, "A13", "Target D/(D+E)");            put(wd, "B13", "=Assumptions!B27", GREEN, PCT1)
put(wd, "A14", "Target E/(D+E)");            put(wd, "B14", "=1-B13", BLACK, PCT1)
put(wd, "A15", "WACC = E% x Ke + D% x after-tax Kd", LBL_B)
put(wd, "B15", "=B14*B9+B13*B12", BLACK_B, PCT3, fill=YELL, border=BOXB)

# ---- 2. FCFF forecast
section(wd, 17, "2. FCFF Forecast (revenue -> EBITDA -> EBIT -> NOPAT -> FCFF)")
put(wd, "A18", "Revenue growth")
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}18", f"=Assumptions!{col}18", GREEN, PCT1, align=RIGHT)
put(wd, "A19", "Revenue")
put(wd, "B19", "=Assumptions!B5", GREEN, NUM)
for i in range(5):
    col, prev = get_column_letter(3+i), get_column_letter(2+i)
    put(wd, f"{col}19", f"={prev}19*(1+{col}18)", BLACK, NUM)
put(wd, "A20", "EBITDA margin")
put(wd, "B20", "=Assumptions!B6/Assumptions!B5", BLACK, PCT1)
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}20", f"=Assumptions!{col}19", GREEN, PCT1, align=RIGHT)
put(wd, "A21", "EBITDA")
put(wd, "B21", "=Assumptions!B6", GREEN, NUM)
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}21", f"={col}19*{col}20", BLACK, NUM)
put(wd, "A22", "Less: D&A (6.0% of revenue)")
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}22", f"=-{col}19*Assumptions!$B$12", BLACK, NUM)
put(wd, "A23", "EBIT", LBL_B)
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}23", f"={col}21+{col}22", BLACK_B, NUM, border=TOPB)
put(wd, "A24", "Less: cash taxes (25% of EBIT)")
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}24", f"=-{col}23*Assumptions!$B$15", BLACK, NUM)
put(wd, "A25", "NOPAT", LBL_B)
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}25", f"={col}23+{col}24", BLACK_B, NUM, border=TOPB)
put(wd, "A26", "Plus: D&A")
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}26", f"=-{col}22", BLACK, NUM)
put(wd, "A27", "Less: capex (6.5% of revenue)")
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}27", f"=-{col}19*Assumptions!$B$13", BLACK, NUM)
put(wd, "A28", "Less: change in NWC (10% of change in revenue)")
for i in range(5):
    col, prev = get_column_letter(3+i), get_column_letter(2+i)
    put(wd, f"{col}28", f"=-({col}19-{prev}19)*Assumptions!$B$14", BLACK, NUM)
put(wd, "A29", "FCFF (unlevered free cash flow)", LBL_B)
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}29", f"=SUM({col}25:{col}28)", BLACK_B, NUM, border=TOPB)

put(wd, "A31", "Discount period (end-of-year)")
put(wd, "C31", 1, BLUE, INT, align=RIGHT)
for i in range(1, 5):
    col, prev = get_column_letter(3+i), get_column_letter(2+i)
    put(wd, f"{col}31", f"={prev}31+1", BLACK, INT, align=RIGHT)
put(wd, "A32", "Discount factor @ WACC")
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}32", f"=1/(1+$B$15)^{col}31", BLACK, FACT)
put(wd, "A33", "PV of FCFF", LBL_B)
for i in range(5):
    col = get_column_letter(3+i)
    put(wd, f"{col}33", f"={col}29*{col}32", BLACK_B, NUM, border=TOPB)

# ---- 3. Terminal value
section(wd, 35, "3. Terminal Value (Gordon growth on normalized FCF; capex = D&A in perpetuity)")
put(wd, "A36", "Terminal growth (g)");  put(wd, "B36", "=Assumptions!B30", GREEN, PCT1)
put(wd, "A37", "Year-5 NOPAT");         put(wd, "B37", "=G25", BLACK, NUM)
put(wd, "A38", "Less: terminal change in NWC (Yr-5 dNWC x (1+g))")
put(wd, "B38", "=-(G19-F19)*Assumptions!$B$14*(1+B36)", BLACK, NUM)
put(wd, "A39", "Normalized terminal FCF", LBL_B)
put(wd, "B39", "=B37+B38", BLACK_B, NUM, border=TOPB)
put(wd, "A40", "Terminal value at end of Year 5 = norm FCF x (1+g)/(WACC-g)")
put(wd, "B40", "=B39*(1+B36)/($B$15-B36)", BLACK_B, NUM)
put(wd, "A41", "PV of terminal value = TV/(1+WACC)^5")
put(wd, "B41", "=B40/(1+$B$15)^G31", BLACK_B, NUM)

# ---- 4. EV / equity bridge
section(wd, 43, "4. Enterprise & Equity Value")
put(wd, "A44", "PV of explicit FCFF (Years 1-5)")
put(wd, "B44", "=SUM(C33:G33)", BLACK, NUM)
put(wd, "A45", "PV of terminal value");  put(wd, "B45", "=B41", BLACK, NUM)
put(wd, "A46", "Enterprise value", LBL_B)
put(wd, "B46", "=B44+B45", BLACK_B, NUM, border=TOPB)
put(wd, "A47", "Less: net debt");        put(wd, "B47", "=-Assumptions!B7", GREEN, NUM)
put(wd, "A48", "Equity value", LBL_B)
put(wd, "B48", "=B46+B47", BLACK_B, NUM, border=TOPB)
put(wd, "A49", "Shares outstanding (m)")
put(wd, "B49", "=Assumptions!B8", GREEN, NUM)
put(wd, "A50", "Implied price per share ($)", LBL_B)
put(wd, "B50", "=B48/B49", BLACK_B, PX, border=TOPB)
put(wd, "A51", "Reference share price ($)")
put(wd, "B51", "=Assumptions!B9", GREEN, PX)
put(wd, "A52", "Implied premium / (discount) vs reference")
put(wd, "B52", "=B50/B51-1", BLACK, "0.0%;[Red](0.0%)")
put(wd, "A53", "Terminal value as % of enterprise value")
put(wd, "B53", "=B45/B46", BLACK, PCT1)

# ---- 5. Headline outputs (exact labels)
section(wd, 55, "5. Headline Outputs")
put(wd, "A56", "wacc", LBL_B);  put(wd, "B56", "=B15", BLACK_B, DEC, fill=YELL, border=BOXB)
put(wd, "C56", "(decimal)", ITAL)
put(wd, "A57", "enterprise_value", LBL_B)
put(wd, "B57", "=B46", BLACK_B, NUM, fill=YELL, border=BOXB);  put(wd, "C57", "($m)", ITAL)
put(wd, "A58", "equity_value", LBL_B)
put(wd, "B58", "=B48", BLACK_B, NUM, fill=YELL, border=BOXB);  put(wd, "C58", "($m)", ITAL)
put(wd, "A59", "implied_price_per_share", LBL_B)
put(wd, "B59", "=B50", BLACK_B, PX, fill=YELL, border=BOXB);   put(wd, "C59", "($/share)", ITAL)

# ---- 6. Checks
section(wd, 61, "6. Checks (all must be 0 / OK)")
put(wd, "A62", "EV bridge: EV - net debt - equity value")
put(wd, "B62", "=B46-Assumptions!B7-B48", BLACK, NUM2)
put(wd, "C62", '=IF(ABS(B62)<0.000001,"OK","FAIL")', BLACK_B)
put(wd, "A63", "Price x shares - equity value")
put(wd, "B63", "=B50*B49-B48", BLACK, NUM2)
put(wd, "C63", '=IF(ABS(B63)<0.000001,"OK","FAIL")', BLACK_B)
put(wd, "A64", "FCFF identity rebuild (Year 5): FCFF - [EBITx(1-t)+D&A-capex-dNWC]")
put(wd, "B64", "=G29-(G23*(1-Assumptions!B15)+G26-G19*Assumptions!B13-(G19-F19)*Assumptions!B14)",
    BLACK, NUM2)
put(wd, "C64", '=IF(ABS(B64)<0.000001,"OK","FAIL")', BLACK_B)
put(wd, "A65", "Gordon validity: WACC - g (must be > 0)")
put(wd, "B65", "=B15-B36", BLACK, PCT3)
put(wd, "C65", '=IF(B36<B15,"OK","FAIL")', BLACK_B)
put(wd, "A66", "Sensitivity center cell ties to base enterprise value")
put(wd, "B66", "=Sensitivity!E7-B46", BLACK, NUM2)
put(wd, "C66", '=IF(ABS(B66)<0.000001,"OK","FAIL")', BLACK_B)
put(wd, "A67", "WACC ties to brief (8.11875%)")
put(wd, "B67", "=B15-0.0811875", BLACK, PCT3)
put(wd, "C67", '=IF(ABS(B67)<0.0000001,"OK","FAIL")', BLACK_B)

wd.column_dimensions["A"].width = 52
for col in "BCDEFGH":
    wd.column_dimensions[col].width = 12
wd.freeze_panes = "B4"
wd.print_area = "A1:H68"
wd.page_setup.orientation = "landscape"
wd.page_setup.fitToWidth = 1

# ================================================================ SENSITIVITY
ws = wb.create_sheet("Sensitivity")
ws.sheet_properties.tabColor = "548235"
title(ws, "IndustrialCo - DCF Sensitivities (live formulas, full model re-derived per cell)")

# Table 1: enterprise value, WACC (rows) x terminal g (cols)
section(ws, 3, "Enterprise value ($m) - WACC vs terminal growth g")
put(ws, "B4", "WACC \\ g", LBL_B, fill=GREY, align=CENTER, border=BOXB)
# g headers: center = live link to model g, wings = +/- 50bp steps
put(ws, "E4", "=DCF!B36", GREEN_B, PCT1, fill=GREY, align=CENTER, border=BOXB)
put(ws, "D4", "=E4-0.005", BLACK_B, PCT1, fill=GREY, align=CENTER, border=BOXB)
put(ws, "C4", "=D4-0.005", BLACK_B, PCT1, fill=GREY, align=CENTER, border=BOXB)
put(ws, "F4", "=E4+0.005", BLACK_B, PCT1, fill=GREY, align=CENTER, border=BOXB)
put(ws, "G4", "=F4+0.005", BLACK_B, PCT1, fill=GREY, align=CENTER, border=BOXB)
# WACC rows: center = live link to model WACC, wings = +/- 50bp steps
put(ws, "B7", "=DCF!B15", GREEN_B, PCT3, fill=GREY, align=CENTER, border=BOXB)
put(ws, "B6", "=B7-0.005", BLACK_B, PCT3, fill=GREY, align=CENTER, border=BOXB)
put(ws, "B5", "=B6-0.005", BLACK_B, PCT3, fill=GREY, align=CENTER, border=BOXB)
put(ws, "B8", "=B7+0.005", BLACK_B, PCT3, fill=GREY, align=CENTER, border=BOXB)
put(ws, "B9", "=B8+0.005", BLACK_B, PCT3, fill=GREY, align=CENTER, border=BOXB)

# Each cell re-derives the full DCF: explicit-year PVs + normalized Gordon TV
EV_FORMULA = (
    "=DCF!$C$29/(1+$B{r})^1+DCF!$D$29/(1+$B{r})^2+DCF!$E$29/(1+$B{r})^3"
    "+DCF!$F$29/(1+$B{r})^4+DCF!$G$29/(1+$B{r})^5"
    "+(DCF!$G$25-(DCF!$G$19-DCF!$F$19)*Assumptions!$B$14*(1+{c}$4))"
    "*(1+{c}$4)/($B{r}-{c}$4)/(1+$B{r})^5"
)
for r in range(5, 10):
    for ci, c in enumerate("CDEFG"):
        is_center = (r == 7 and c == "E")
        put(ws, f"{c}{r}", EV_FORMULA.format(r=r, c=c),
            BLACK_B if is_center else BLACK, NUM,
            fill=YELL if is_center else None, border=BOXB)

# Table 2: implied price per share, derived from Table 1
section(ws, 12, "Implied price per share ($) - WACC vs terminal growth g")
put(ws, "B13", "WACC \\ g", LBL_B, fill=GREY, align=CENTER, border=BOXB)
for c in "CDEFG":
    put(ws, f"{c}13", f"={c}4", BLACK_B, PCT1, fill=GREY, align=CENTER, border=BOXB)
for i, r in enumerate(range(14, 19)):
    put(ws, f"B{r}", f"=B{5+i}", BLACK_B, PCT3, fill=GREY, align=CENTER, border=BOXB)
for i, r in enumerate(range(14, 19)):
    for c in "CDEFG":
        is_center = (r == 16 and c == "E")
        put(ws, f"{c}{r}", f"=({c}{5+i}-Assumptions!$B$7)/Assumptions!$B$8",
            BLACK_B if is_center else BLACK, PX,
            fill=YELL if is_center else None, border=BOXB)

put(ws, "A20", "Center row/column are live links to the model WACC (8.119%) and g (2.0%); "
               "steps of +/- 50bp. Every cell is a full live re-derivation (no data table).", ITAL)

ws.column_dimensions["A"].width = 14
ws.column_dimensions["B"].width = 12
for col in "CDEFGH":
    ws.column_dimensions[col].width = 11
ws.print_area = "A1:H21"
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1

wb.save(OUT)
print("Saved:", OUT)
