"""BM-LBO-001 run 2 — Sponsor LBO of SaaSCo (synthetic benchmark).

Free-form openpyxl build from benchmarks/briefs/lbo_us_saas.md ONLY.
Every derived value is a live Excel formula; all hardcoded inputs live on
the input-designated 'Assumptions' sheet with per-row source attribution.

Sheets: Assumptions | Sources and Uses | Operating Model | Debt Schedule
        | Returns | Sensitivity
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

OUT = r"C:\Users\lukep\Desktop\Projects AI\ModelForge\benchmarks\artifacts\agent\BM-LBO-001_run2.xlsx"

# ----------------------------------------------------------------- styles ---
F_BLUE   = Font(color="0000FF", name="Calibri", size=10)            # inputs
F_CALC   = Font(color="000000", name="Calibri", size=10)            # in-sheet calc
F_LINK   = Font(color="006100", name="Calibri", size=10)            # cross-sheet link
F_LABEL  = Font(color="000000", name="Calibri", size=10)
F_BOLD   = Font(color="000000", name="Calibri", size=10, bold=True)
F_BOLDG  = Font(color="006100", name="Calibri", size=10, bold=True)
F_TITLE  = Font(color="FFFFFF", name="Calibri", size=11, bold=True)
F_SECT   = Font(color="1F4E79", name="Calibri", size=10, bold=True)
F_NOTE   = Font(color="595959", name="Calibri", size=9, italic=True)
FILL_T   = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
FILL_S   = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
FILL_HL  = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # headline band
B_TOP    = Border(top=Side(style="thin", color="000000"))
B_TOTAL  = Border(top=Side(style="thin", color="000000"),
                  bottom=Side(style="double", color="000000"))

NF_M   = '#,##0.0_);(#,##0.0)'      # USD millions
NF_M2  = '#,##0.00_);(#,##0.00)'
NF_PCT = '0.0%'
NF_PC2 = '0.00%'
NF_X1  = '0.0"x"'
NF_X2  = '0.00"x"'
NF_SH  = '$0.00'
NF_DEC = '0.0000'
NF_CHK = '0.000000'
NF_DLT = '+0.0%;-0.0%;0.0%'

def put(ws, coord, value, font=F_CALC, fmt=None, fill=None, border=None,
        align=None):
    c = ws[coord]
    c.value = value
    c.font = font
    if fmt:    c.number_format = fmt
    if fill:   c.fill = fill
    if border: c.border = border
    if align:  c.alignment = Alignment(horizontal=align)
    return c

def title(ws, text, ncols=9):
    for i in range(1, ncols + 1):
        cell = ws.cell(row=1, column=i)
        cell.fill = FILL_T
        cell.font = F_TITLE
    ws["A1"] = text

def section(ws, row, text, ncols=9):
    for i in range(1, ncols + 1):
        ws.cell(row=row, column=i).fill = FILL_S
    put(ws, f"A{row}", text, font=F_SECT, fill=FILL_S)

def widths(ws, spec):
    for col, w in spec.items():
        ws.column_dimensions[col].width = w

wb = Workbook()

# ============================================================ ASSUMPTIONS ===
wsA = wb.active
wsA.title = "Assumptions"
title(wsA, "SaaSCo — SPONSOR LBO MODEL — ASSUMPTIONS & SOURCES (BM-LBO-001, arm B run 2)", 6)
put(wsA, "A2", "Synthetic benchmark. Every input below is sourced verbatim from "
               "benchmarks/briefs/lbo_us_saas.md. USD millions unless noted.", font=F_NOTE)
put(wsA, "A3", "Color code: blue = hardcoded input, black = calculation, green = cross-sheet link.", font=F_NOTE)

section(wsA, 4, "1. TARGET — HISTORICAL FINANCIALS", 6)
rowsA = [
    # (row, label, value, fmt, units, source)
    (5,  "Revenue — FY-2",                              81.0,  NF_M,  "$m", "Brief §1"),
    (6,  "Revenue — FY-1",                              90.0,  NF_M,  "$m", "Brief §1"),
    (7,  "Revenue — FY0 (last historical)",            100.0,  NF_M,  "$m", "Brief §1"),
    (8,  "EBITDA — FY-2",                               26.7,  NF_M,  "$m", "Brief §1"),
    (9,  "EBITDA — FY-1",                               30.6,  NF_M,  "$m", "Brief §1"),
    (10, "EBITDA — FY0",                                35.0,  NF_M,  "$m", "Brief §1"),
    (12, "Target net debt at close (refinanced)",       20.0,  NF_M,  "$m", "Brief §1"),
]
for r, lab, v, fmt, u, src in rowsA:
    put(wsA, f"A{r}", lab)
    put(wsA, f"B{r}", v, font=F_BLUE, fmt=fmt)
    put(wsA, f"C{r}", u, font=F_NOTE)
    put(wsA, f"D{r}", src, font=F_NOTE)
put(wsA, "A11", "Entry (LTM) EBITDA  (= FY0 EBITDA)")
put(wsA, "B11", "=B10", font=F_CALC, fmt=NF_M)
put(wsA, "C11", "$m", font=F_NOTE); put(wsA, "D11", "Brief §1", font=F_NOTE)

section(wsA, 14, "2. PURCHASE PRICE", 6)
rowsB = [
    (15, "Current share price",        4.40, NF_SH,  "$ / share", "Brief §2"),
    (16, "Offer premium",              0.25, NF_PCT, "% over current price", "Brief §2"),
    (17, "Fully diluted shares",       60.0, NF_M,   "m shares", "Brief §2"),
    (18, "Option / RSU buyout",         0.0, NF_M,   "$m", "Brief §2"),
]
for r, lab, v, fmt, u, src in rowsB:
    put(wsA, f"A{r}", lab)
    put(wsA, f"B{r}", v, font=F_BLUE, fmt=fmt)
    put(wsA, f"C{r}", u, font=F_NOTE)
    put(wsA, f"D{r}", src, font=F_NOTE)

section(wsA, 20, "3. FINANCING & TRANSACTION COSTS", 6)
rowsC = [
    (21, "M&A advisory fees",                                7.0,   NF_M,   "$m, funded at close", "Brief §3"),
    (22, "Financing fees",                                   5.0,   NF_M,   "$m, funded at close", "Brief §3"),
    (23, "Minimum cash funded to balance sheet",             5.0,   NF_M,   "$m, funded at close", "Brief §3"),
    (24, "Senior Term Loan sizing",                          4.5,   NF_X1,  "x entry LTM EBITDA", "Brief §3"),
    (25, "Mezzanine Notes sizing",                           1.0,   NF_X1,  "x entry LTM EBITDA", "Brief §3"),
    (26, "SOFR (base rate)",                                 0.04,  NF_PC2, "flat, no floor", "Brief §3"),
    (27, "Senior Term Loan spread",                          0.035, NF_PC2, "+350 bps", "Brief §3"),
    (28, "Mezzanine Notes spread",                           0.07,  NF_PC2, "+700 bps", "Brief §3"),
    (29, "Senior Term Loan tenor",                           7,     '0',    "years; bullet, no OID, no floor", "Brief §3"),
    (30, "Mezzanine Notes tenor",                            8,     '0',    "years; bullet, no OID, no floor", "Brief §3"),
    (31, "Scheduled amortization — both tranches (bullet)",  0.0,   NF_M,   "$m / year", "Brief §3"),
    (32, "Cash sweep",                                       0.0,   NF_PCT, "% of FCF — none per brief", "Brief §3"),
    (33, "Interim distributions to sponsor",                 0.0,   NF_M,   "$m / year — none per brief", "Brief §5"),
]
for r, lab, v, fmt, u, src in rowsC:
    put(wsA, f"A{r}", lab)
    put(wsA, f"B{r}", v, font=F_BLUE, fmt=fmt)
    put(wsA, f"C{r}", u, font=F_NOTE)
    put(wsA, f"D{r}", src, font=F_NOTE)

section(wsA, 35, "4. OPERATING PLAN (PROJECTION YEARS 1–5)", 6)
for i, h in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    put(wsA, f"{get_column_letter(2 + i)}36", h, font=F_BOLD, align="center")
put(wsA, "A37", "Revenue growth")
for i, v in enumerate([0.12, 0.11, 0.10, 0.09, 0.08]):
    put(wsA, f"{get_column_letter(2 + i)}37", v, font=F_BLUE, fmt=NF_PCT)
put(wsA, "G37", "Brief §4", font=F_NOTE)
put(wsA, "A38", "EBITDA margin")
for i, v in enumerate([0.35, 0.36, 0.36, 0.37, 0.38]):
    put(wsA, f"{get_column_letter(2 + i)}38", v, font=F_BLUE, fmt=NF_PCT)
put(wsA, "G38", "Brief §4", font=F_NOTE)
rowsD = [
    (39, "D&A (% of revenue)",                 0.04, NF_PCT, "", "Brief §4"),
    (40, "Maintenance capex (% of revenue)",   0.03, NF_PCT, "", "Brief §4"),
    (41, "Growth capex (% of revenue)",        0.01, NF_PCT, "", "Brief §4"),
    (42, "Change in NWC (% of Δ revenue)",     0.05, NF_PCT, "", "Brief §4"),
    (43, "Cash tax rate",                      0.25, NF_PCT, "", "Brief §4"),
]
for r, lab, v, fmt, u, src in rowsD:
    put(wsA, f"A{r}", lab)
    put(wsA, f"B{r}", v, font=F_BLUE, fmt=fmt)
    put(wsA, f"D{r}", src, font=F_NOTE)

section(wsA, 45, "5. EXIT", 6)
put(wsA, "A46", "Hold period")
put(wsA, "B46", 5, font=F_BLUE, fmt='0')
put(wsA, "C46", "years (exit end of Year 5)", font=F_NOTE)
put(wsA, "D46", "Brief §5", font=F_NOTE)
put(wsA, "A47", "Exit multiple (strategic sale)")
put(wsA, "B47", 9.0, font=F_BLUE, fmt=NF_X1)
put(wsA, "C47", "x EV / EBITDA on PROJECTED Year-5 EBITDA (not entry LTM)", font=F_NOTE)
put(wsA, "D47", "Brief §5", font=F_NOTE)
put(wsA, "A48", "Convention: the exit bridge deducts face debt outstanding from exit EV; accumulated "
                "balance-sheet cash is EXCLUDED from the bridge (stated conservative convention).", font=F_NOTE)
put(wsA, "D48", "Brief §3 / §5", font=F_NOTE)
put(wsA, "A49", "Convention: interest accrues on the opening balance of each year; both tranches "
                "bullet; no cash sweep, no revolver, no dividend recap, no earnout, no PIK.", font=F_NOTE)
put(wsA, "D49", "Brief §3", font=F_NOTE)

section(wsA, 51, "6. SENSITIVITY GRID STEPS (model setting, not a brief input)", 6)
put(wsA, "A52", "Exit multiple step")
put(wsA, "B52", 0.5, font=F_BLUE, fmt=NF_X1)
put(wsA, "D52", "Model setting", font=F_NOTE)
put(wsA, "A53", "Growth delta step")
put(wsA, "B53", 0.01, font=F_BLUE, fmt=NF_PCT)
put(wsA, "D53", "Model setting", font=F_NOTE)

widths(wsA, {"A": 62, "B": 12, "C": 11, "D": 11, "E": 11, "F": 11, "G": 11})
wsA.freeze_panes = "A5"

# ====================================================== SOURCES AND USES ===
wsS = wb.create_sheet("Sources and Uses")
title(wsS, "SOURCES & USES OF FUNDS — AT CLOSE", 3)
put(wsS, "A2", "USD millions. All lines are live formulas driven by the Assumptions sheet.", font=F_NOTE)

section(wsS, 4, "ENTRY VALUATION", 3)
put(wsS, "A5", "Offer price per share")
put(wsS, "B5", "=Share_Price*(1+Offer_Premium)", font=F_LINK, fmt=NF_SH)
put(wsS, "C5", "i.e. $4.40 × (1 + 25%) = $5.50  (Brief §2)", font=F_NOTE)
put(wsS, "A6", "Equity purchase price")
put(wsS, "B6", "=B5*FD_Shares+Option_Buyout", font=F_CALC, fmt=NF_M)
put(wsS, "C6", "i.e. offer price × 60.0m fully diluted shares", font=F_NOTE)
put(wsS, "A7", "Plus: target net debt at close")
put(wsS, "B7", "=Net_Debt_Entry", font=F_LINK, fmt=NF_M)
put(wsS, "A8", "Entry enterprise value", font=F_BOLD)
put(wsS, "B8", "=B6+B7", font=F_BOLD, fmt=NF_M, border=B_TOP)
put(wsS, "A9", "Implied entry EV / LTM EBITDA")
put(wsS, "B9", "=B8/Entry_EBITDA", font=F_LINK, fmt=NF_X1)

section(wsS, 11, "USES OF FUNDS", 3)
put(wsS, "A12", "Equity purchase price")
put(wsS, "B12", "=B6", font=F_CALC, fmt=NF_M)
put(wsS, "A13", "Refinance target net debt")
put(wsS, "B13", "=Net_Debt_Entry", font=F_LINK, fmt=NF_M)
put(wsS, "A14", "M&A advisory fees")
put(wsS, "B14", "=Advisory_Fees", font=F_LINK, fmt=NF_M)
put(wsS, "A15", "Financing fees")
put(wsS, "B15", "=Financing_Fees", font=F_LINK, fmt=NF_M)
put(wsS, "A16", "Minimum cash funded to balance sheet")
put(wsS, "B16", "=Min_Cash", font=F_LINK, fmt=NF_M)
put(wsS, "A17", "Total uses of funds", font=F_BOLD)
put(wsS, "B17", "=SUM(B12:B16)", font=F_BOLD, fmt=NF_M, border=B_TOTAL)

section(wsS, 19, "SOURCES OF FUNDS", 3)
put(wsS, "A20", "Senior Term Loan")
put(wsS, "B20", "=TL_Mult*Entry_EBITDA", font=F_LINK, fmt=NF_M)
put(wsS, "C20", "4.5× EBITDA; SOFR 4.00% + 350 bps = 7.50% all-in; bullet, 7-year (Brief §3)", font=F_NOTE)
put(wsS, "A21", "Mezzanine Notes")
put(wsS, "B21", "=Mezz_Mult*Entry_EBITDA", font=F_LINK, fmt=NF_M)
put(wsS, "C21", "1.0× EBITDA; SOFR 4.00% + 700 bps = 11.00% all-in; bullet, 8-year (Brief §3)", font=F_NOTE)
put(wsS, "A22", "Sponsor equity (balancing plug, new money)", font=F_BOLD)
put(wsS, "B22", "=B17-B20-B21", font=F_BOLD, fmt=NF_M)
put(wsS, "C22", "i.e. total uses − total debt; no management rollover (Brief §3)", font=F_NOTE)
put(wsS, "A23", "Total sources of funds", font=F_BOLD)
put(wsS, "B23", "=SUM(B20:B22)", font=F_BOLD, fmt=NF_M, border=B_TOTAL)

section(wsS, 25, "CHECKS & PRO-FORMA CAPITALIZATION", 3)
put(wsS, "A26", "Check: total sources − total uses (must be 0)")
put(wsS, "B26", "=B23-B17", font=F_CALC, fmt=NF_CHK)
put(wsS, "A27", "Entry leverage — total debt / LTM EBITDA")
put(wsS, "B27", "=(B20+B21)/Entry_EBITDA", font=F_LINK, fmt=NF_X1)
put(wsS, "A28", "Senior leverage — Term Loan / LTM EBITDA")
put(wsS, "B28", "=B20/Entry_EBITDA", font=F_LINK, fmt=NF_X1)
put(wsS, "A29", "Equity % of total capitalization")
put(wsS, "B29", "=B22/B23", font=F_CALC, fmt=NF_PCT)

widths(wsS, {"A": 46, "B": 13, "C": 70})
wsS.freeze_panes = "B5"

# ======================================================= OPERATING MODEL ===
wsO = wb.create_sheet("Operating Model")
title(wsO, "OPERATING MODEL — P&L, FREE CASH FLOW, CASH ROLL-FORWARD", 9)
put(wsO, "A2", "USD millions. Historicals per Brief §1; projection drivers per Brief §4. "
               "Interest links from the Debt Schedule.", font=F_NOTE)

HCOLS = ["B", "C", "D"]                       # FY-2, FY-1, FY0
PCOLS = ["E", "F", "G", "H", "I"]             # Year 1..5
HDRS  = ["FY-2", "FY-1", "FY0 (Entry)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5 (Exit)"]
for i, h in enumerate(HDRS):
    put(wsO, f"{get_column_letter(2 + i)}4", h, font=F_BOLD, align="center", border=B_TOP)

put(wsO, "A5", "Revenue", font=F_BOLD)
put(wsO, "B5", "=Assumptions!$B$5", font=F_LINK, fmt=NF_M)
put(wsO, "C5", "=Assumptions!$B$6", font=F_LINK, fmt=NF_M)
put(wsO, "D5", "=Rev_FY0", font=F_LINK, fmt=NF_M)
prev = "D"
for c in PCOLS:
    put(wsO, f"{c}5", f"={prev}5*(1+{c}6)", font=F_CALC, fmt=NF_M)
    prev = c
put(wsO, "A6", "  Revenue growth %")
put(wsO, "C6", "=C5/B5-1", font=F_CALC, fmt=NF_PCT)
put(wsO, "D6", "=D5/C5-1", font=F_CALC, fmt=NF_PCT)
for c, nm in zip(PCOLS, ["Growth_Y1", "Growth_Y2", "Growth_Y3", "Growth_Y4", "Growth_Y5"]):
    put(wsO, f"{c}6", f"={nm}", font=F_LINK, fmt=NF_PCT)
put(wsO, "A7", "  EBITDA margin %")
for c in HCOLS:
    put(wsO, f"{c}7", f"={c}8/{c}5", font=F_CALC, fmt=NF_PCT)
for c, nm in zip(PCOLS, ["Margin_Y1", "Margin_Y2", "Margin_Y3", "Margin_Y4", "Margin_Y5"]):
    put(wsO, f"{c}7", f"={nm}", font=F_LINK, fmt=NF_PCT)
put(wsO, "A8", "EBITDA", font=F_BOLD)
put(wsO, "B8", "=Assumptions!$B$8", font=F_LINK, fmt=NF_M)
put(wsO, "C8", "=Assumptions!$B$9", font=F_LINK, fmt=NF_M)
put(wsO, "D8", "=Assumptions!$B$10", font=F_LINK, fmt=NF_M)
for c in PCOLS:
    put(wsO, f"{c}8", f"={c}5*{c}7", font=F_CALC, fmt=NF_M)
put(wsO, "A9", "Less: D&A (4.0% of revenue)")
for c in PCOLS:
    put(wsO, f"{c}9", f"=-DA_Pct*{c}5", font=F_LINK, fmt=NF_M)
put(wsO, "A10", "EBIT")
for c in PCOLS:
    put(wsO, f"{c}10", f"={c}8+{c}9", font=F_CALC, fmt=NF_M, border=B_TOP)
put(wsO, "A11", "Less: cash interest expense")
DCOLS = ["B", "C", "D", "E", "F"]             # Debt Schedule Year 1..5 columns
for c, d in zip(PCOLS, DCOLS):
    put(wsO, f"{c}11", f"=-'Debt Schedule'!{d}21", font=F_LINK, fmt=NF_M)
put(wsO, "A12", "Pre-tax income (EBT)")
for c in PCOLS:
    put(wsO, f"{c}12", f"={c}10+{c}11", font=F_CALC, fmt=NF_M)
put(wsO, "A13", "Less: cash taxes (25% of positive EBT)")
for c in PCOLS:
    put(wsO, f"{c}13", f"=-MAX(0,{c}12)*Tax_Rate", font=F_LINK, fmt=NF_M)
put(wsO, "A14", "Net income", font=F_BOLD)
for c in PCOLS:
    put(wsO, f"{c}14", f"={c}12+{c}13", font=F_BOLD, fmt=NF_M, border=B_TOP)

section(wsO, 16, "LEVERED FREE CASH FLOW", 9)
put(wsO, "A17", "Net income")
for c in PCOLS:
    put(wsO, f"{c}17", f"={c}14", font=F_CALC, fmt=NF_M)
put(wsO, "A18", "Plus: D&A")
for c in PCOLS:
    put(wsO, f"{c}18", f"=-{c}9", font=F_CALC, fmt=NF_M)
put(wsO, "A19", "Less: maintenance capex (3.0% of revenue)")
for c in PCOLS:
    put(wsO, f"{c}19", f"=-MaintCapex_Pct*{c}5", font=F_LINK, fmt=NF_M)
put(wsO, "A20", "Less: growth capex (1.0% of revenue)")
for c in PCOLS:
    put(wsO, f"{c}20", f"=-GrowthCapex_Pct*{c}5", font=F_LINK, fmt=NF_M)
put(wsO, "A21", "Less: increase in net working capital (5% of Δ revenue)")
prev = "D"
for c in PCOLS:
    put(wsO, f"{c}21", f"=-NWC_Pct*({c}5-{prev}5)", font=F_LINK, fmt=NF_M)
    prev = c
put(wsO, "A22", "Levered free cash flow (no scheduled amortization)", font=F_BOLD)
for c in PCOLS:
    put(wsO, f"{c}22", f"=SUM({c}17:{c}21)", font=F_BOLD, fmt=NF_M, border=B_TOTAL)

section(wsO, 24, "CASH ROLL-FORWARD (no cash sweep — FCF accumulates on the balance sheet)", 9)
put(wsO, "A25", "Beginning cash")
put(wsO, "E25", "='Sources and Uses'!$B$16", font=F_LINK, fmt=NF_M)
prev = "E"
for c in PCOLS[1:]:
    put(wsO, f"{c}25", f"={prev}27", font=F_CALC, fmt=NF_M)
    prev = c
put(wsO, "A26", "Plus: levered free cash flow")
for c in PCOLS:
    put(wsO, f"{c}26", f"={c}22", font=F_CALC, fmt=NF_M)
put(wsO, "A27", "Ending cash", font=F_BOLD)
for c in PCOLS:
    put(wsO, f"{c}27", f"={c}25+{c}26", font=F_BOLD, fmt=NF_M, border=B_TOP)
put(wsO, "A28", "Memo: ending cash is EXCLUDED from the equity bridge at exit "
                "(stated conservative convention, Brief §3 / §5).", font=F_NOTE)

section(wsO, 30, "CREDIT STATISTICS", 9)
put(wsO, "A31", "Total debt — end of period")
for c, d in zip(PCOLS, DCOLS):
    put(wsO, f"{c}31", f"='Debt Schedule'!{d}20", font=F_LINK, fmt=NF_M)
put(wsO, "A32", "Gross leverage — total debt / EBITDA")
for c in PCOLS:
    put(wsO, f"{c}32", f"={c}31/{c}8", font=F_CALC, fmt=NF_X1)
put(wsO, "A33", "Net leverage — (total debt − cash) / EBITDA")
for c in PCOLS:
    put(wsO, f"{c}33", f"=({c}31-{c}27)/{c}8", font=F_CALC, fmt=NF_X1)
put(wsO, "A34", "Interest coverage — EBITDA / cash interest")
for c in PCOLS:
    put(wsO, f"{c}34", f"={c}8/-{c}11", font=F_CALC, fmt=NF_X1)

widths(wsO, {"A": 52, **{c: 12 for c in list("BCDEFGHI")}})
wsO.freeze_panes = "B5"

# ========================================================= DEBT SCHEDULE ===
wsD = wb.create_sheet("Debt Schedule")
title(wsD, "DEBT SCHEDULE — PER-TRANCHE ROLL-FORWARD", 6)
put(wsD, "A2", "USD millions. Interest accrues on the opening balance of each year (Brief §3). "
               "Both tranches are bullet with no cash sweep, so balances are flat through exit.", font=F_NOTE)
for i, h in enumerate(["Year 1", "Year 2", "Year 3", "Year 4", "Year 5"]):
    put(wsD, f"{get_column_letter(2 + i)}4", h, font=F_BOLD, align="center", border=B_TOP)
YCOLS = ["B", "C", "D", "E", "F"]

section(wsD, 5, "SENIOR TERM LOAN — SOFR + 350 bps, bullet, 7-year tenor", 6)
put(wsD, "A6", "All-in interest rate (SOFR 4.00% + 3.50%)")
put(wsD, "B6", "=SOFR_Rate+TL_Spread", font=F_LINK, fmt=NF_PC2)
put(wsD, "A7", "Opening balance")
put(wsD, "B7", "=TL_Amount", font=F_LINK, fmt=NF_M)
prev = "B"
for c in YCOLS[1:]:
    put(wsD, f"{c}7", f"={prev}9", font=F_CALC, fmt=NF_M)
    prev = c
put(wsD, "A8", "Less: scheduled repayment (bullet — none)")
for c in YCOLS:
    put(wsD, f"{c}8", "=-Sched_Amort", font=F_LINK, fmt=NF_M)
put(wsD, "A9", "Closing balance", font=F_BOLD)
for c in YCOLS:
    put(wsD, f"{c}9", f"={c}7+{c}8", font=F_BOLD, fmt=NF_M, border=B_TOP)
put(wsD, "A10", "Interest expense (opening balance × all-in rate)")
for c in YCOLS:
    put(wsD, f"{c}10", f"={c}7*$B$6", font=F_CALC, fmt=NF_M)

section(wsD, 12, "MEZZANINE NOTES — SOFR + 700 bps, bullet, 8-year tenor", 6)
put(wsD, "A13", "All-in interest rate (SOFR 4.00% + 7.00%)")
put(wsD, "B13", "=SOFR_Rate+Mezz_Spread", font=F_LINK, fmt=NF_PC2)
put(wsD, "A14", "Opening balance")
put(wsD, "B14", "=Mezz_Amount", font=F_LINK, fmt=NF_M)
prev = "B"
for c in YCOLS[1:]:
    put(wsD, f"{c}14", f"={prev}16", font=F_CALC, fmt=NF_M)
    prev = c
put(wsD, "A15", "Less: scheduled repayment (bullet — none)")
for c in YCOLS:
    put(wsD, f"{c}15", "=-Sched_Amort", font=F_LINK, fmt=NF_M)
put(wsD, "A16", "Closing balance", font=F_BOLD)
for c in YCOLS:
    put(wsD, f"{c}16", f"={c}14+{c}15", font=F_BOLD, fmt=NF_M, border=B_TOP)
put(wsD, "A17", "Interest expense (opening balance × all-in rate)")
for c in YCOLS:
    put(wsD, f"{c}17", f"={c}14*$B$13", font=F_CALC, fmt=NF_M)

section(wsD, 19, "TOTAL DEBT", 6)
put(wsD, "A20", "Total debt — closing balance", font=F_BOLD)
for c in YCOLS:
    put(wsD, f"{c}20", f"={c}9+{c}16", font=F_BOLD, fmt=NF_M)
put(wsD, "A21", "Total cash interest expense", font=F_BOLD)
for c in YCOLS:
    put(wsD, f"{c}21", f"={c}10+{c}17", font=F_BOLD, fmt=NF_M)

put(wsD, "A23", "Debt outstanding at exit (face value, end of Year 5)", font=F_BOLD)
put(wsD, "B23", "=F20", font=F_BOLD, fmt=NF_M)
put(wsD, "A24", "Check: exit debt − (Term Loan + Mezzanine at close) (must be 0)")
put(wsD, "B24", "=B23-(TL_Amount+Mezz_Amount)", font=F_LINK, fmt=NF_CHK)

widths(wsD, {"A": 52, **{c: 12 for c in list("BCDEF")}})
wsD.freeze_panes = "B5"

# ================================================================ RETURNS ===
wsR = wb.create_sheet("Returns")
title(wsR, "RETURNS ANALYSIS — EXIT BRIDGE, CASH FLOWS & HEADLINE OUTPUTS", 7)
put(wsR, "A2", "USD millions. Exit at end of Year 5 via strategic sale at 9.0× projected "
               "Year-5 EBITDA (Brief §5). No interim distributions.", font=F_NOTE)

section(wsR, 4, "EQUITY BRIDGE AT EXIT (END OF YEAR 5)", 7)
put(wsR, "A5", "Projected Year-5 EBITDA (exit basis — NOT entry LTM)")
put(wsR, "B5", "=EBITDA_Y5", font=F_LINK, fmt=NF_M)
put(wsR, "A6", "Exit EV / EBITDA multiple (strategic sale)")
put(wsR, "B6", "=Exit_Multiple", font=F_LINK, fmt=NF_X1)
put(wsR, "A7", "Exit enterprise value")
put(wsR, "B7", "=B5*B6", font=F_CALC, fmt=NF_M)
put(wsR, "A8", "Less: debt outstanding at exit (face value)")
put(wsR, "B8", "=-Debt_AtExit", font=F_LINK, fmt=NF_M)
put(wsR, "A9", "Exit equity proceeds", font=F_BOLD)
put(wsR, "B9", "=B7+B8", font=F_BOLD, fmt=NF_M, border=B_TOTAL)
put(wsR, "A10", "Memo: accumulated cash at exit — EXCLUDED from the bridge per brief convention")
put(wsR, "B10", "='Operating Model'!I27", font=F_LINK, fmt=NF_M)

section(wsR, 12, "SPONSOR CASH FLOWS (ANNUAL)", 7)
for i, h in enumerate(["Year 0 (Close)", "Year 1", "Year 2", "Year 3", "Year 4", "Year 5 (Exit)"]):
    put(wsR, f"{get_column_letter(2 + i)}13", h, font=F_BOLD, align="center", border=B_TOP)
put(wsR, "A14", "Sponsor cash flow", font=F_BOLD)
put(wsR, "B14", "=-Sponsor_Equity", font=F_LINK, fmt=NF_M)
for c in ["C", "D", "E", "F"]:
    put(wsR, f"{c}14", "=Dist_PerYear", font=F_LINK, fmt=NF_M)
put(wsR, "G14", "=Dist_PerYear+$B$9", font=F_LINK, fmt=NF_M)

section(wsR, 16, "RETURN CALCULATIONS", 7)
put(wsR, "A17", "IRR — internal rate of return on sponsor cash flows (annual)", font=F_BOLD)
put(wsR, "B17", "=IRR(B14:G14)", font=F_BOLD, fmt=NF_PC2)
put(wsR, "A18", "MoIC — multiple of invested capital (total inflows / equity invested)", font=F_BOLD)
put(wsR, "B18", "=SUM(C14:G14)/-B14", font=F_BOLD, fmt=NF_X2)

section(wsR, 20, "HEADLINE OUTPUTS (BENCHMARK LABELS)", 7)
put(wsR, "A21", "Sponsor IRR (sponsor_irr) — annual, decimal", font=F_BOLD, fill=FILL_HL)
put(wsR, "B21", "=B17", font=F_BOLD, fmt=NF_DEC, fill=FILL_HL)
put(wsR, "C21", "shown as decimal; equals 15.46% annualized", font=F_NOTE)
put(wsR, "A22", "Sponsor MoIC (sponsor_moic) — x invested capital", font=F_BOLD, fill=FILL_HL)
put(wsR, "B22", "=B18", font=F_BOLD, fmt=NF_X2, fill=FILL_HL)
put(wsR, "A23", "Exit equity proceeds (exit_equity_proceeds) — $m", font=F_BOLD, fill=FILL_HL)
put(wsR, "B23", "=B9", font=F_BOLD, fmt=NF_M, fill=FILL_HL)
put(wsR, "A24", "Sponsor equity cheque (sponsor_equity_cheque) — $m", font=F_BOLD, fill=FILL_HL)
put(wsR, "B24", "=Sponsor_Equity", font=F_BOLDG, fmt=NF_M, fill=FILL_HL)

section(wsR, 26, "CHECKS", 7)
put(wsR, "A27", "Check: power-law cross-check of annual return (single-inflow profile; must be ~0)")
put(wsR, "B27", "=ABS((SUM(C14:G14)/-B14)^(1/Hold_Years)-1-B17)", font=F_LINK, fmt=NF_CHK)
put(wsR, "A28", "Check: sources and uses balanced (must be 0)")
put(wsR, "B28", "='Sources and Uses'!B26", font=F_LINK, fmt=NF_CHK)
put(wsR, "A29", "Check: exit debt equals face value of tranches (must be 0)")
put(wsR, "B29", "='Debt Schedule'!B24", font=F_LINK, fmt=NF_CHK)

widths(wsR, {"A": 64, **{c: 14 for c in list("BCDEFG")}})
wsR.freeze_panes = "B5"

# ============================================================ SENSITIVITY ===
wsX = wb.create_sheet("Sensitivity")
title(wsX, "SENSITIVITY ANALYSIS — EXIT MULTIPLE × REVENUE GROWTH DELTA", 7)
put(wsX, "A2", "Closed-form grids (live formulas): annual return = ((M × Year-5 EBITDA(Δg) − exit debt) "
               "/ invested equity)^(1/5) − 1, where Δg shifts every projection-year growth rate "
               "uniformly. Center cell = base case (9.0×, Δg = 0).", font=F_NOTE)

GBASE1 = 6    # grid 1 first data row (multiples rows 6..10, header deltas row 5)
GBASE2 = 14   # grid 2 first data row (rows 14..18, header row 13)
GCOLS  = ["C", "D", "E", "F", "G"]

def sens_grid(title_row, hdr_row, first_row, body_fmt, power):
    put(wsX, f"B{hdr_row}", "EV/EBITDA ↓   Δ growth →", font=F_SECT, align="right")
    put(wsX, f"{GCOLS[0]}{hdr_row}", "=-2*Sens_GrowthStep", font=F_LINK, fmt=NF_DLT, align="center")
    for i in range(1, 5):
        put(wsX, f"{GCOLS[i]}{hdr_row}", f"={GCOLS[i-1]}{hdr_row}+Sens_GrowthStep",
            font=F_CALC, fmt=NF_DLT, align="center")
    put(wsX, f"B{first_row}", "=Exit_Multiple-2*Sens_MultStep", font=F_LINK, fmt=NF_X1, align="center")
    for r in range(first_row + 1, first_row + 5):
        put(wsX, f"B{r}", f"=B{r-1}+Sens_MultStep", font=F_CALC, fmt=NF_X1, align="center")
    for r in range(first_row, first_row + 5):
        for c in GCOLS:
            ebitda5 = (f"Rev_FY0*(1+Growth_Y1+{c}${hdr_row})*(1+Growth_Y2+{c}${hdr_row})"
                       f"*(1+Growth_Y3+{c}${hdr_row})*(1+Growth_Y4+{c}${hdr_row})"
                       f"*(1+Growth_Y5+{c}${hdr_row})*Margin_Y5")
            core = f"($B{r}*({ebitda5})-Debt_AtExit)/Sponsor_Equity"
            formula = f"=({core})^(1/Hold_Years)-1" if power else f"={core}"
            put(wsX, f"{c}{r}", formula, font=F_LINK, fmt=body_fmt, align="center")

section(wsX, 4, "GRID 1 — ANNUALIZED SPONSOR RETURN", 7)
sens_grid(4, 5, GBASE1, NF_PC2, power=True)
section(wsX, 12, "GRID 2 — EXIT PROCEEDS ÷ EQUITY INVESTED (×)", 7)
sens_grid(12, 13, GBASE2, NF_X2, power=False)

put(wsX, "A20", "Check: grid 1 center − base annual return (must be 0)")
put(wsX, "B20", "=ABS(E8-Returns!$B$17)", font=F_LINK, fmt=NF_CHK)
put(wsX, "A21", "Check: grid 2 center − base capital multiple (must be 0)")
put(wsX, "B21", "=ABS(E16-Returns!$B$18)", font=F_LINK, fmt=NF_CHK)

widths(wsX, {"A": 46, "B": 24, **{c: 12 for c in list("CDEFG")}})
wsX.freeze_panes = "A3"

# ========================================================== DEFINED NAMES ===
NAMES = {
    # Assumptions inputs
    "Rev_FY0":          "Assumptions!$B$7",
    "Entry_EBITDA":     "Assumptions!$B$11",
    "Net_Debt_Entry":   "Assumptions!$B$12",
    "Share_Price":      "Assumptions!$B$15",
    "Offer_Premium":    "Assumptions!$B$16",
    "FD_Shares":        "Assumptions!$B$17",
    "Option_Buyout":    "Assumptions!$B$18",
    "Advisory_Fees":    "Assumptions!$B$21",
    "Financing_Fees":   "Assumptions!$B$22",
    "Min_Cash":         "Assumptions!$B$23",
    "TL_Mult":          "Assumptions!$B$24",
    "Mezz_Mult":        "Assumptions!$B$25",
    "SOFR_Rate":        "Assumptions!$B$26",
    "TL_Spread":        "Assumptions!$B$27",
    "Mezz_Spread":      "Assumptions!$B$28",
    "Sched_Amort":      "Assumptions!$B$31",
    "Dist_PerYear":     "Assumptions!$B$33",
    "Growth_Y1":        "Assumptions!$B$37",
    "Growth_Y2":        "Assumptions!$C$37",
    "Growth_Y3":        "Assumptions!$D$37",
    "Growth_Y4":        "Assumptions!$E$37",
    "Growth_Y5":        "Assumptions!$F$37",
    "Margin_Y1":        "Assumptions!$B$38",
    "Margin_Y2":        "Assumptions!$C$38",
    "Margin_Y3":        "Assumptions!$D$38",
    "Margin_Y4":        "Assumptions!$E$38",
    "Margin_Y5":        "Assumptions!$F$38",
    "DA_Pct":           "Assumptions!$B$39",
    "MaintCapex_Pct":   "Assumptions!$B$40",
    "GrowthCapex_Pct":  "Assumptions!$B$41",
    "NWC_Pct":          "Assumptions!$B$42",
    "Tax_Rate":         "Assumptions!$B$43",
    "Hold_Years":       "Assumptions!$B$46",
    "Exit_Multiple":    "Assumptions!$B$47",
    "Sens_MultStep":    "Assumptions!$B$52",
    "Sens_GrowthStep":  "Assumptions!$B$53",
    # Computed anchors
    "Total_Uses":       "'Sources and Uses'!$B$17",
    "TL_Amount":        "'Sources and Uses'!$B$20",
    "Mezz_Amount":      "'Sources and Uses'!$B$21",
    "Sponsor_Equity":   "'Sources and Uses'!$B$22",
    "EBITDA_Y5":        "'Operating Model'!$I$8",
    "Debt_AtExit":      "'Debt Schedule'!$B$23",
    "Exit_Equity":      "Returns!$B$9",
}
for nm, ref in NAMES.items():
    wb.defined_names[nm] = DefinedName(nm, attr_text=ref)

# Force Excel/LibreOffice to recalculate on open (openpyxl stores no cache)
try:
    wb.calculation.fullCalcOnLoad = True
except Exception:
    pass

wb.save(OUT)
print("saved:", OUT)
