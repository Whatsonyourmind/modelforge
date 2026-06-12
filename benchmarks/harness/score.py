"""score.py — deterministic scorer for the ModelForge public benchmark.

Scores a directory of .xlsx artifacts named
    <brief_id>__arm<A|B>__run<k>.xlsx
against the pre-registered protocol (benchmarks/PROTOCOL.md) and the
clean-room ground truth (benchmarks/harness/ground_truth.py).

Metrics (exact definitions in PROTOCOL.md §4):
    m1 formula_error_cells   — third-party `formulas` recalc + cached scan
    m2 headline_accuracy     — extracted headlines vs ground truth (0.5% rel)
    m3 hardcode_rate         — numeric constants / (constants + formulas),
                               input-designated sheets excluded
    m4 reproducibility       — arm A: byte-identical double build;
                               arm B: cross-run headline range + structure
    m5 lineage               — sources/assumptions surface present
    m6 completeness          — per-brief checklist

The scorer imports NO modelforge code. Arm-A reproducibility shells out to
`python -m modelforge.cli build` (the product under test), which is the
measurement itself, not a grading dependency.

Usage:
    python benchmarks/harness/score.py --artifacts benchmarks/results/run1
    python benchmarks/harness/score.py --artifacts DIR --repro-arm-a
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import subprocess
import sys
import tempfile
from collections import defaultdict
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))
from ground_truth import compute_all  # noqa: E402  (clean-room, no modelforge)

REPO = Path(__file__).resolve().parents[2]
SPECS_DIR = REPO / "benchmarks" / "specs"
DEFAULT_OUT = REPO / "benchmarks" / "results"

ARTIFACT_RE = re.compile(
    r"^(?P<brief>[a-z0-9_]+)__arm(?P<arm>[AB])__run(?P<run>\d+)\.xlsx$")

# Excel error sigils — same convention as modelforge/qc/workbook_audit.py
# (clean-room re-implementation; the scorer does not import modelforge).
ERROR_RE = re.compile(r"^#(REF|DIV/0|VALUE|NAME|NUM|N/A|NULL|GETTING_DATA)")

# Sheets excluded from m3 (PROTOCOL §4 m3) — identical rule for both arms.
INPUT_SHEET_RE = re.compile(
    r"assumption|input|source|cover|readme|instruction|toc|index|reproducib"
    r"|redflags|compliance", re.I)

SOURCE_DATE_EPOCH = "1767225600"   # 2026-01-01T00:00:00Z — pinned in PROTOCOL

REL_TOL = 0.005          # m2: 0.5% relative tolerance
BALANCE_ABS_TOL = 0.01   # m2: |balance check| <= 0.01 ($m)

BALANCE_HEADLINES = {"final_balance_check"}
FINAL_YEAR_HEADLINES = {
    "final_net_income", "final_total_assets",
    "final_total_liabilities_equity", "final_balance_check",
    "final_cash", "final_debt",
}


# ─────────────────────────────────────────────────────────────────────────────
# Workbook loading / recalculation
# ─────────────────────────────────────────────────────────────────────────────

_CELL_REF_RE = re.compile(r"^[A-Z]{1,3}[0-9]+$")


def _unwrap_numpy_scalar(v):
    """numpy scalars (np.float64, np.str_, ...) and 0-d arrays have ndim == 0
    AND a .flatten attribute; flatten() yields the same scalar back, which made
    the flatten branches below recurse infinitely (DEVIATION D-001, 2026-06-12).
    Unwrap them to plain Python objects first."""
    if getattr(v, "ndim", None) == 0:
        try:
            return v.item()
        except Exception:
            return v
    return v


def _coerce_error(v):
    """Return the Excel error code if the engine value is an error, else None."""
    if v is None:
        return None
    v = _unwrap_numpy_scalar(v)
    flatten = getattr(v, "flatten", None)
    if flatten is not None:
        try:
            for item in v.flatten():
                code = _coerce_error(item)
                if code is not None:
                    return code
            return None
        except Exception:
            return None
    s = str(v)
    return s if ERROR_RE.match(s) else None


def _coerce_float(v):
    v = _unwrap_numpy_scalar(v)
    flatten = getattr(v, "flatten", None)
    if flatten is not None:
        try:
            flat = list(v.flatten())
            if len(flat) == 1:
                return _coerce_float(flat[0])
            return None
        except Exception:
            return None
    if isinstance(v, bool):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


class Workbook:
    """One artifact: raw cells, cached values, third-party recalc values."""

    def __init__(self, path: Path):
        self.path = path
        self.notes: list[str] = []
        self.wb_raw = load_workbook(path, data_only=False)
        try:
            self.wb_cached = load_workbook(path, data_only=True)
        except Exception as e:
            self.wb_cached = None
            self.notes.append(f"cached-value load failed: {e}")
        self.sheet_upper = {n.upper(): n for n in self.wb_raw.sheetnames}

        # Third-party recalculation (the authoritative value source).
        self.recalc_ok = False
        self.recalc_values: dict[tuple[str, str], float] = {}
        self.recalc_errors: set[tuple[str, str]] = set()
        try:
            import formulas
            sol = formulas.ExcelModel().loads(str(path)).finish().calculate()
            self.recalc_ok = True
        except Exception as e:
            self.notes.append(f"formulas recalc failed: {e}")
            sol = {}
        for key, val in sol.items():
            try:
                after_book = key.split("]", 1)[1]
                sheet_tok, cell = after_book.split("'!", 1)
            except (IndexError, ValueError):
                continue
            cell = cell.strip()
            if not _CELL_REF_RE.match(cell):
                continue
            sheet = sheet_tok.strip("'").upper()
            raw = val.value if hasattr(val, "value") else val
            code = _coerce_error(raw)
            if code is not None:
                self.recalc_errors.add((sheet, cell))
                continue
            f = _coerce_float(raw)
            if f is not None:
                self.recalc_values[(sheet, cell)] = f

    # value lookup: recalc > cached > raw constant -------------------------
    def value(self, sheet: str, coord: str):
        v = self.recalc_values.get((sheet.upper(), coord.upper()))
        if v is not None:
            return v
        if self.wb_cached is not None:
            real = self.sheet_upper.get(sheet.upper())
            if real is not None:
                cv = self.wb_cached[real][coord].value
                if isinstance(cv, (int, float)) and not isinstance(cv, bool):
                    return float(cv)
        real = self.sheet_upper.get(sheet.upper())
        if real is not None:
            rv = self.wb_raw[real][coord].value
            if isinstance(rv, (int, float)) and not isinstance(rv, bool):
                return float(rv)
        return None

    def row_numeric(self, sheet: str, row: int, first: bool,
                    min_col: int = 2, max_col: int = 60):
        """First/last numeric value in a row (cols B..BH by default)."""
        real = self.sheet_upper.get(sheet.upper())
        if real is None:
            return None
        cols = range(min_col, max_col + 1)
        if not first:
            cols = reversed(list(cols))
        from openpyxl.utils import get_column_letter
        for ci in cols:
            coord = f"{get_column_letter(ci)}{row}"
            v = self.value(real, coord)
            if v is not None:
                return v
        return None

    def iter_label_cells(self):
        """Yield (sheet, row, col_idx, text) for every static string cell."""
        for ws in self.wb_raw.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and not c.value.startswith("="):
                        yield ws.title, c.row, c.column, c.value


# ─────────────────────────────────────────────────────────────────────────────
# m1 — formula error cells
# ─────────────────────────────────────────────────────────────────────────────

def metric_formula_errors(wb: Workbook) -> dict:
    errors: set[str] = set()
    for sheet, cell in wb.recalc_errors:
        real = wb.sheet_upper.get(sheet, sheet)
        errors.add(f"{real}!{cell}")
    if wb.wb_cached is not None:
        for ws in wb.wb_cached.worksheets:
            for row in ws.iter_rows():
                for c in row:
                    if isinstance(c.value, str) and ERROR_RE.match(c.value):
                        errors.add(f"{ws.title}!{c.coordinate}")
    return {
        "formula_error_cells": len(errors),
        "error_cell_list": sorted(errors)[:50],
        "recalc_ran": wb.recalc_ok,
    }


# ─────────────────────────────────────────────────────────────────────────────
# m2 — headline extraction + accuracy
# ─────────────────────────────────────────────────────────────────────────────

def _find_label_rows(wb: Workbook, sheet: str, predicate):
    """Rows on `sheet` whose column A/B label satisfies predicate(text)."""
    out = []
    real = wb.sheet_upper.get(sheet.upper())
    if real is None:
        return out
    ws = wb.wb_raw[real]
    for row in ws.iter_rows(min_col=1, max_col=2):
        for c in row:
            if isinstance(c.value, str) and predicate(c.value):
                out.append(c.row)
                break
    return out


# ---- Arm A: fixed extraction against ModelForge's deterministic layout -----

def _arm_a_lbo(wb: Workbook) -> dict:
    out = {}
    su = "SourcesUses"
    rows = _find_label_rows(wb, su, lambda s: s.strip() == "Sponsor equity (new money)")
    if rows:
        out["sponsor_equity_cheque"] = wb.value(su, f"D{rows[0]}")
    exit_year = None
    rows = _find_label_rows(wb, su, lambda s: s.strip() == "Exit year")
    if rows:
        ey = wb.value(su, f"D{rows[0]}")
        exit_year = int(ey) if ey is not None else None
    # CF-series "Strategic sale" row = the one with >=3 numerics in D..M.
    strat_rows = _find_label_rows(wb, su, lambda s: s.strip() == "Strategic sale")
    from openpyxl.utils import get_column_letter
    cf_row = None
    for r in strat_rows:
        numerics = sum(
            1 for ci in range(4, 14)
            if wb.value(su, f"{get_column_letter(ci)}{r}") is not None)
        if numerics >= 3:
            cf_row = r
            break
    if cf_row is not None and exit_year is not None:
        col = get_column_letter(4 + exit_year)   # D = year 0
        out["exit_equity_proceeds"] = wb.value(su, f"{col}{cf_row}")
    # Returns summary: strategic block is written first → first IRR/MoIC row.
    rows = _find_label_rows(wb, su, lambda s: s.strip().startswith("IRR (on cash-flow"))
    if rows:
        out["sponsor_irr"] = wb.value(su, f"D{rows[0]}")
    rows = _find_label_rows(wb, su, lambda s: s.strip().startswith("MoIC (cash-on-cash"))
    if rows:
        out["sponsor_moic"] = wb.value(su, f"D{rows[0]}")
    return out


def _arm_a_dcf(wb: Workbook) -> dict:
    out = {}
    rows = _find_label_rows(wb, "WACCBuild", lambda s: s.strip() == "WACC")
    if rows:
        out["wacc"] = wb.value("WACCBuild", f"D{rows[0]}")
    for headline, label in [("enterprise_value", "Enterprise Value"),
                            ("equity_value", "Equity Value"),
                            ("implied_price_per_share", "Implied price per share")]:
        rows = _find_label_rows(wb, "Valuation", lambda s, L=label: s.strip() == L)
        if rows:
            out[headline] = wb.value("Valuation", f"D{rows[0]}")
    return out


def _arm_a_3s(wb: Workbook) -> dict:
    out = {}
    labels = {
        "final_net_income": "Net income",
        "final_cash": "Cash",
        "final_total_assets": "TOTAL ASSETS",
        "final_debt": "Debt",
        "final_total_liabilities_equity": "TOTAL L & E",
        "final_balance_check": "BS check (A - L - E)",
    }
    for headline, label in labels.items():
        rows = _find_label_rows(wb, "Model", lambda s, L=label: s.strip() == L)
        if rows:
            out[headline] = wb.row_numeric("Model", rows[0], first=False)
    return out


ARM_A_EXTRACTORS = {
    "lbo_us_saas": _arm_a_lbo,
    "dcf_industrial": _arm_a_dcf,
    "three_statement_mfg": _arm_a_3s,
}

# ---- Arm B: frozen label-search patterns (PROTOCOL §4 m2) -------------------

ARM_B_PATTERNS: dict[str, dict[str, dict]] = {
    "lbo_us_saas": {
        "sponsor_irr": {
            "include": r"\bIRR\b",
            "exclude": r"lender|hurdle|blended|unlevered|threshold",
            "mode": "first",
        },
        "sponsor_moic": {
            "include": r"\bMoIC\b|money multiple|multiple\s+of\s+(invested\s+)?capital|\bMOIC\b",
            "exclude": r"lender|blended",
            "mode": "first",
        },
        "exit_equity_proceeds": {
            "include": r"exit\s+equity|equity\s+proceeds|exit\s+proceeds|sponsor\s+proceeds",
            "exclude": r"per\s+share",
            "mode": "first",
        },
        "sponsor_equity_cheque": {
            "include": r"sponsor\s+equity|equity\s+cheque|equity\s+check\b|equity\s+contribution|equity\s+invested",
            "exclude": r"roll",
            "mode": "first",
        },
    },
    "dcf_industrial": {
        "wacc": {
            "include": r"^\s*WACC\s*$|weighted\s+average\s+cost\s+of\s+capital|discount\s+rate",
            "exclude": r"sensitiv",
            "mode": "first",
        },
        "enterprise_value": {
            "include": r"enterprise\s+value|^\s*EV\s*$",
            "exclude": r"sensitiv|/|x\b",
            "mode": "first",
        },
        "equity_value": {
            "include": r"equity\s+value",
            "exclude": r"per\s+share",
            "mode": "first",
        },
        "implied_price_per_share": {
            "include": r"per\s+share|implied\s+(share\s+)?price",
            "exclude": r"current|reference",
            "mode": "first",
        },
    },
    "three_statement_mfg": {
        "final_net_income": {
            "include": r"^\s*net\s+income\s*$",
            "exclude": r"margin|%",
            "mode": "last",
        },
        "final_total_assets": {
            "include": r"total\s+assets",
            "exclude": r"",
            "mode": "last",
        },
        "final_total_liabilities_equity": {
            "include": r"total\s+l\s*&\s*e|total\s+liab(ilities)?\s*(and|&|\+)?\s*(shareholders'?\s+)?equity",
            "exclude": r"",
            "mode": "last",
        },
        "final_balance_check": {
            "include": r"balance\s+check|bs\s+check|check\s*\(a\s*-\s*l\s*-\s*e\)|balance\s*\?|tie[\s-]*out",
            "exclude": r"",
            "mode": "last",
        },
        "final_cash": {
            "include": r"^\s*cash(\s*(&|and)\s*(cash\s+)?equivalents)?\s*$|ending\s+cash|closing\s+cash",
            "exclude": r"flow|change",
            "mode": "last",
        },
        "final_debt": {
            "include": r"^\s*(total\s+)?debt\s*$|debt\s*\(closing\)|closing\s+debt|debt\s+outstanding",
            "exclude": r"repay|issuance|net\s+debt|/",
            "mode": "last",
        },
    },
}


def _arm_b_extract(wb: Workbook, brief: str) -> dict:
    out = {}
    from openpyxl.utils import get_column_letter
    label_cells = list(wb.iter_label_cells())
    for headline, spec in ARM_B_PATTERNS[brief].items():
        inc = re.compile(spec["include"], re.I)
        exc = re.compile(spec["exclude"], re.I) if spec["exclude"] else None
        found = None
        for sheet, row, col, text in label_cells:
            if not inc.search(text):
                continue
            if exc is not None and exc.search(text):
                continue
            if headline == "wacc" and re.search(r"sensitiv", sheet, re.I):
                continue
            if spec["mode"] == "first":
                # nearest numeric to the right of the label (within 8 cols)
                for ci in range(col + 1, col + 9):
                    v = wb.value(sheet, f"{get_column_letter(ci)}{row}")
                    if v is not None:
                        found = v
                        break
            else:  # "last" — final-year column
                v = wb.row_numeric(sheet, row, first=False, min_col=col + 1)
                if v is not None:
                    found = v
            if found is not None:
                break
        out[headline] = found
    return out


def metric_headlines(wb: Workbook, brief: str, arm: str, truth: dict) -> dict:
    if arm == "A":
        extracted = ARM_A_EXTRACTORS[brief](wb)
    else:
        extracted = _arm_b_extract(wb, brief)
    expected = truth[brief]["headlines"]
    per = {}
    n_pass = 0
    for name, exp in expected.items():
        got = extracted.get(name)
        if got is None:
            per[name] = {"got": None, "expected": exp, "pass": False,
                         "note": "not located"}
            continue
        if name in BALANCE_HEADLINES:
            ok = abs(got) <= BALANCE_ABS_TOL
            rel = None
        else:
            rel = abs(got - exp) / abs(exp) if exp != 0 else abs(got)
            ok = rel <= REL_TOL
        per[name] = {"got": got, "expected": exp, "rel_err": rel, "pass": ok}
        n_pass += int(ok)
    return {
        "headline_accuracy": n_pass / len(expected) if expected else None,
        "headlines": per,
    }


# ─────────────────────────────────────────────────────────────────────────────
# m3 — hardcode rate
# ─────────────────────────────────────────────────────────────────────────────

def metric_hardcode_rate(wb: Workbook) -> dict:
    n_const = 0
    n_formula = 0
    for ws in wb.wb_raw.worksheets:
        if INPUT_SHEET_RE.search(ws.title):
            continue
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith("="):
                    n_formula += 1
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    if isinstance(v, int) or float(v).is_integer():
                        if 1900 <= float(v) <= 2100:
                            continue  # calendar-year header exclusion
                    n_const += 1
    denom = n_const + n_formula
    return {
        "hardcode_rate": (n_const / denom) if denom else None,
        "numeric_constant_cells": n_const,
        "formula_cells": n_formula,
    }


# ─────────────────────────────────────────────────────────────────────────────
# m5 — lineage
# ─────────────────────────────────────────────────────────────────────────────

def metric_lineage(wb: Workbook) -> dict:
    sheet_hit = False
    for ws in wb.wb_raw.worksheets:
        if re.search(r"source|lineage|assumption", ws.title, re.I):
            strings = sum(
                1 for row in ws.iter_rows() for c in row
                if isinstance(c.value, str) and c.value.strip())
            if strings >= 3:
                sheet_hit = True
                break
    n_comments = sum(
        1 for ws in wb.wb_raw.worksheets
        for row in ws.iter_rows() for c in row if c.comment is not None)
    return {"lineage": bool(sheet_hit or n_comments >= 5),
            "lineage_sheet": sheet_hit, "cell_comments": n_comments}


# ─────────────────────────────────────────────────────────────────────────────
# m6 — completeness
# ─────────────────────────────────────────────────────────────────────────────

COMPLETENESS_ITEMS = {
    "lbo_us_saas": {
        "sources_and_uses": r"sources\s*(&|and)\s*uses|uses\s+of\s+funds",
        "debt_schedule": r"debt\s*schedule|debt\s+roll|amorti[sz]ation\s+schedule",
        "returns_irr_moic": None,   # special: needs BOTH IRR and MoIC tokens
        "sensitivity": r"sensitivit|tornado",
    },
    "dcf_industrial": {
        "wacc_build": r"\bWACC\b|cost\s+of\s+capital",
        "terminal_value": r"terminal\s+value|gordon|perpetuity",
        "fcff_forecast": r"\bFCFF\b|free\s+cash\s+flow|\bNOPAT\b",
        "sensitivity": r"sensitivit|tornado",
    },
    "three_statement_mfg": {
        "income_statement": r"net\s+income",
        "balance_sheet": r"total\s+assets",
        "cash_flow_statement": r"cash\s+flow|\bCFO\b|operating\s+activities",
        "bs_tie": None,             # special: |final_balance_check| <= 0.01
    },
}


def metric_completeness(wb: Workbook, brief: str, headline_result: dict) -> dict:
    haystack = [t for _, _, _, t in wb.iter_label_cells()]
    haystack += list(wb.wb_raw.sheetnames)
    blob = "\n".join(haystack)
    items = {}
    for item, pattern in COMPLETENESS_ITEMS[brief].items():
        if item == "returns_irr_moic":
            ok = bool(re.search(r"\bIRR\b", blob, re.I)) and bool(
                re.search(r"MoIC|money\s+multiple|multiple.*capital", blob, re.I))
        elif item == "bs_tie":
            h = headline_result["headlines"].get("final_balance_check", {})
            got = h.get("got")
            ok = got is not None and abs(got) <= BALANCE_ABS_TOL
        elif item == "debt_schedule":
            ok = bool(re.search(pattern, blob, re.I)) or any(
                re.search(r"debt", n, re.I) for n in wb.wb_raw.sheetnames)
        else:
            ok = bool(re.search(pattern, blob, re.I))
        items[item] = ok
    return {"completeness": sum(items.values()) / len(items), "items": items}


# ─────────────────────────────────────────────────────────────────────────────
# m4 — reproducibility
# ─────────────────────────────────────────────────────────────────────────────

def repro_arm_a(brief: str) -> dict:
    """Build the arm-A spec twice with SOURCE_DATE_EPOCH pinned; compare bytes."""
    spec = SPECS_DIR / f"{brief}.yaml"
    if not spec.exists():
        return {"byte_identical": None, "note": f"spec not found: {spec}"}
    digests = []
    env = dict(os.environ, SOURCE_DATE_EPOCH=SOURCE_DATE_EPOCH)
    with tempfile.TemporaryDirectory() as td:
        for k in (1, 2):
            # SAME filename in separate subdirs: the Trust/Moat sheets echo
            # the output filename, so differing names would trivially break
            # byte-identity without measuring anything real.
            sub = Path(td) / f"build{k}"
            sub.mkdir()
            out = sub / f"{brief}.xlsx"
            p = subprocess.run(
                [sys.executable, "-m", "modelforge.cli", "build", str(spec),
                 "--out", str(out)],
                cwd=REPO, env=env, capture_output=True, text=True)
            if p.returncode != 0:
                return {"byte_identical": False,
                        "note": f"build {k} failed: {p.stderr[-300:]}"}
            digests.append(hashlib.sha256(out.read_bytes()).hexdigest())
    return {"byte_identical": digests[0] == digests[1],
            "sha256": digests}


def _structure_fingerprint(wb: Workbook):
    fp = []
    for ws in wb.wb_raw.worksheets:
        n_formula = sum(
            1 for row in ws.iter_rows() for c in row
            if isinstance(c.value, str) and c.value.startswith("="))
        fp.append((ws.title, ws.max_row, ws.max_column, n_formula))
    return sorted(fp)


def repro_arm_b(brief: str, run_results: list[dict],
                fingerprints: list) -> dict:
    """Cross-run stability for arm B (>=2 runs required)."""
    ranges = {}
    headline_names = list(run_results[0]["m2"]["headlines"].keys()) if run_results else []
    for name in headline_names:
        vals = [r["m2"]["headlines"][name]["got"] for r in run_results
                if r["m2"]["headlines"][name]["got"] is not None]
        if len(vals) >= 2:
            mean = sum(vals) / len(vals)
            ranges[name] = ((max(vals) - min(vals)) / abs(mean)) if mean else None
        else:
            ranges[name] = None
    structural = len(fingerprints) >= 2 and all(
        f == fingerprints[0] for f in fingerprints[1:])
    return {"headline_value_range": ranges,
            "structural_identical": structural if len(fingerprints) >= 2 else None,
            "n_runs": len(run_results)}


# ─────────────────────────────────────────────────────────────────────────────
# Orchestration
# ─────────────────────────────────────────────────────────────────────────────

def score_artifact(path: Path, brief: str, arm: str, truth: dict) -> tuple[dict, list]:
    wb = Workbook(path)
    m2 = metric_headlines(wb, brief, arm, truth)
    result = {
        "artifact": path.name,
        "brief": brief,
        "arm": arm,
        "m1": metric_formula_errors(wb),
        "m2": m2,
        "m3": metric_hardcode_rate(wb),
        "m5": metric_lineage(wb),
        "m6": metric_completeness(wb, brief, m2),
        "notes": wb.notes,
    }
    return result, _structure_fingerprint(wb)


def render_md(payload: dict) -> str:
    lines = ["# ModelForge Public Benchmark — Results",
             "",
             f"Scored {len(payload['artifacts'])} artifact(s). "
             "Protocol: benchmarks/PROTOCOL.md (pre-registered). "
             "Ground truth: benchmarks/harness/ground_truth.py.",
             ""]
    if payload.get("model_id_arm_b"):
        lines.append(f"Arm-B model: `{payload['model_id_arm_b']}`")
    lines.append(f"ModelForge version (arm A): "
                 f"`{payload.get('modelforge_version', 'unknown')}`")
    lines.append("")
    by_brief = defaultdict(list)
    for a in payload["artifacts"]:
        by_brief[a["brief"]].append(a)
    for brief in sorted(by_brief):
        lines.append(f"## {brief}")
        lines.append("")
        lines.append("| artifact | arm | m1 errors | m2 headline acc | "
                     "m3 hardcode | m5 lineage | m6 complete | recalc |")
        lines.append("|---|---|---|---|---|---|---|---|")
        for a in sorted(by_brief[brief], key=lambda x: x["artifact"]):
            m3v = a["m3"]["hardcode_rate"]
            lines.append(
                "| {f} | {arm} | {m1} | {m2:.0%} | {m3} | {m5} | {m6:.0%} | {rc} |".format(
                    f=a["artifact"], arm=a["arm"],
                    m1=a["m1"]["formula_error_cells"],
                    m2=a["m2"]["headline_accuracy"],
                    m3=("n/a" if m3v is None else f"{m3v:.1%}"),
                    m5="yes" if a["m5"]["lineage"] else "no",
                    m6=a["m6"]["completeness"],
                    rc="ok" if a["m1"]["recalc_ran"] else "FAILED"))
        lines.append("")
        # headline detail
        lines.append("<details><summary>headline detail</summary>")
        lines.append("")
        lines.append("| artifact | headline | got | expected | pass |")
        lines.append("|---|---|---|---|---|")
        for a in sorted(by_brief[brief], key=lambda x: x["artifact"]):
            for name, h in a["m2"]["headlines"].items():
                got = h["got"]
                lines.append("| {f} | {n} | {g} | {e:.6g} | {p} |".format(
                    f=a["artifact"], n=name,
                    g=("—" if got is None else f"{got:.6g}"),
                    e=h["expected"], p="PASS" if h["pass"] else "FAIL"))
        lines.append("")
        lines.append("</details>")
        lines.append("")
    lines.append("## m4 — reproducibility")
    lines.append("")
    for brief, rep in sorted(payload.get("m4", {}).items()):
        lines.append(f"### {brief}")
        a = rep.get("arm_a")
        if a is not None:
            lines.append(f"- arm A byte-identical double build: "
                         f"**{a['byte_identical']}**")
        b = rep.get("arm_b")
        if b is not None:
            lines.append(f"- arm B structural identical across {b['n_runs']} "
                         f"runs: **{b['structural_identical']}**")
            for name, rng in b["headline_value_range"].items():
                shown = "n/a" if rng is None else f"{rng:.2%}"
                lines.append(f"  - `{name}` cross-run range: {shown}")
        lines.append("")
    lines.append("---")
    lines.append("*Arm B is a raw frontier-agent + openpyxl baseline — NOT a "
                 "measurement of any commercial Excel product. See "
                 "PROTOCOL.md §3 for all honesty caveats.*")
    lines.append("")
    lines.append("*Conflict of interest: the arm-A specs were authored by the "
                 "ModelForge project, which also built and scored this "
                 "benchmark (PROTOCOL.md §3.4, §3.6). Arm-B model: "
                 "claude-opus-4-8 (PROTOCOL.md Deviations D-003). "
                 "three_statement arm-B m2 sub-100% scores are headline "
                 "LOCATABILITY under the frozen label-search, not wrong math "
                 "(D-004). Scorer numpy-coercion fix applied post-arms, "
                 "coercion-only (D-001).*")
    return "\n".join(lines) + "\n"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--artifacts", required=True,
                    help="directory of <brief>__arm<A|B>__run<k>.xlsx files")
    ap.add_argument("--out", default=str(DEFAULT_OUT),
                    help="output dir for results.json / RESULTS.md")
    ap.add_argument("--repro-arm-a", action="store_true",
                    help="run the arm-A double-build byte-identity check "
                         "(requires modelforge installed)")
    ap.add_argument("--model-id-b", default=None,
                    help="record the arm-B model id in the results")
    args = ap.parse_args()

    art_dir = Path(args.artifacts)
    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)
    truth = compute_all()

    artifacts = []
    fingerprints: dict[tuple[str, str], list] = defaultdict(list)
    grouped: dict[tuple[str, str], list[dict]] = defaultdict(list)
    skipped = []
    for f in sorted(art_dir.glob("*.xlsx")):
        m = ARTIFACT_RE.match(f.name)
        if not m:
            skipped.append(f.name)
            continue
        brief, arm = m.group("brief"), m.group("arm")
        if brief not in truth:
            skipped.append(f.name)
            continue
        print(f"scoring {f.name} ...")
        result, fp = score_artifact(f, brief, arm, truth)
        artifacts.append(result)
        grouped[(brief, arm)].append(result)
        fingerprints[(brief, arm)].append(fp)

    if not artifacts:
        print(f"no artifacts matching <brief>__arm<A|B>__run<k>.xlsx in "
              f"{art_dir}", file=sys.stderr)
        return 2

    m4: dict[str, dict] = defaultdict(dict)
    briefs_seen = sorted({a["brief"] for a in artifacts})
    if args.repro_arm_a:
        for brief in briefs_seen:
            print(f"m4 arm-A repro check: {brief} (two builds) ...")
            m4[brief]["arm_a"] = repro_arm_a(brief)
    for brief in briefs_seen:
        runs_b = grouped.get((brief, "B"), [])
        if len(runs_b) >= 2:
            m4[brief]["arm_b"] = repro_arm_b(
                brief, runs_b, fingerprints[(brief, "B")])

    # Dist metadata can be stale for editable installs; ask the interpreter
    # that builds arm A in a subprocess (the scorer itself still imports no
    # modelforge code). Fall back to dist metadata, then "unknown".
    mf_version = "unknown"
    try:
        p = subprocess.run(
            [sys.executable, "-c",
             "import modelforge; print(modelforge.__version__)"],
            capture_output=True, text=True, timeout=120)
        if p.returncode == 0 and p.stdout.strip():
            mf_version = p.stdout.strip()
    except Exception:
        pass
    if mf_version == "unknown":
        try:
            from importlib.metadata import version
            mf_version = version("modelforge-finance")
        except Exception:
            pass

    payload = {
        "protocol": "benchmarks/PROTOCOL.md v1",
        "modelforge_version": mf_version,
        "model_id_arm_b": args.model_id_b,
        "ground_truth": {k: v["headlines"] for k, v in truth.items()},
        "artifacts": artifacts,
        "m4": dict(m4),
        "skipped_files": skipped,
        "metrics_registered": [
            "m1_formula_error_cells", "m2_headline_accuracy",
            "m3_hardcode_rate", "m4_reproducibility", "m5_lineage",
            "m6_completeness",
        ],
    }
    results_json = out_dir / "results.json"
    results_json.write_text(json.dumps(payload, indent=2, sort_keys=False),
                            encoding="utf-8")
    results_md = out_dir / "RESULTS.md"
    results_md.write_text(render_md(payload), encoding="utf-8")
    print(f"\nwrote {results_json}")
    print(f"wrote {results_md}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
