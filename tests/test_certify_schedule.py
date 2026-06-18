"""Tests for the `modelforge certify --schedule` opt-in gate.

Proves the schedule auditor wires into certify as an opt-in gate: a certified
model still certifies clean WITH --schedule, and a model with a hardcode typed
over an interior formula cell PASSES plain certify (the defect is certify-blind)
but FAILS `certify --schedule`.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import yaml
from click.testing import CliRunner

from modelforge.cli import main, _load_spec_class
from modelforge.moat.classifier import classify_sheet
from modelforge.templates import build_model

ROOT = Path(__file__).resolve().parent.parent
_EXCLUDED = {"input", "reference", "audit"}


def _build(tmp_path: Path) -> Path:
    yml = ROOT / "examples" / "three_statement_cdmo.yaml"
    if not yml.exists():
        pytest.skip("example spec not present")
    raw = yaml.safe_load(yml.read_bytes())
    spec = _load_spec_class(raw["model_type"]).model_validate(raw)
    xlsx, _ = build_model(spec, tmp_path / "model.xlsx")
    return xlsx


def test_certify_schedule_clean_certifies(tmp_path):
    xlsx = _build(tmp_path)
    res = CliRunner().invoke(main, ["certify", str(xlsx), "--schedule"])
    assert res.exit_code == 0, res.output
    assert "SCHEDULE CLEAN" in res.output


def test_certify_schedule_gate_catches_certify_blind_hardcode(tmp_path):
    xlsx = _build(tmp_path)

    # Inject a hardcode over an interior cell of a contiguous formula run.
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    injected = False
    for name in wb.sheetnames:
        if classify_sheet(name) in _EXCLUDED:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            run: list = []
            prev = None
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    run = run + [c] if (prev is not None and c.column == prev + 1) else [c]
                    prev = c.column
                    if len(run) >= 3:
                        run[1].value = 424242.0  # interior of the run
                        injected = True
                        break
                else:
                    run = []
                    prev = None
            if injected:
                break
        if injected:
            break
    assert injected, "no contiguous 3+ formula run found to inject into"
    wb.save(xlsx)

    runner = CliRunner()
    # Plain certify is blind to it (still no formula errors) -> exit 0.
    plain = runner.invoke(main, ["certify", str(xlsx)])
    assert plain.exit_code == 0, plain.output
    # The opt-in schedule gate catches it -> exit 1.
    gated = runner.invoke(main, ["certify", str(xlsx), "--schedule"])
    assert gated.exit_code == 1, gated.output
    assert "SCHEDULE REVIEW" in gated.output
