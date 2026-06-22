"""Tests for the perturbation-replay conservation gate.

Proves the gate (a) catches a latent input-dependent conservation defect a
baseline check misses (a hardcoded total → breaks two-sided), (b) does NOT
false-fail a one-sided validity threshold (which breaks in only one direction)
nor a correct real model, and (c) fails-closed cleanly when there is no QC sheet.
"""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import Workbook

from modelforge.qc import audit_perturb_replay

ROOT = Path(__file__).resolve().parent.parent
SPEC = ROOT / "examples" / "sponsor_lbo_us_saas.yaml"


def _synth(tmp: Path, name: str, *, total_formula: str, check: str) -> Path:
    """Tiny model: leaf input A1, a part b=2·A1, a 'total', and one QC check."""
    wb = Workbook()
    a = wb.active
    a.title = "Assumptions"
    a["A1"] = 10.0  # leaf input driver
    m = wb.create_sheet("Model")
    m["B1"] = "=Assumptions!A1*2"   # part b (= 20 at baseline)
    m["B2"] = total_formula          # 'total'
    q = wb.create_sheet("QC")
    q["A4"] = "ALL CHECKS PASS"
    q["C4"] = "=IF(C5=1,1,0)"
    q["A5"] = "check"
    q["C5"] = check
    p = tmp / f"{name}.xlsx"
    wb.save(p)
    return p


class TestSynthetic:
    def test_clean_identity_passes(self, tmp_path):
        # total is a live formula (= b); the identity holds under any input.
        p = _synth(tmp_path, "clean", total_formula="=Model!B1",
                   check="=IF(ABS(Model!B2-Model!B1)<0.001,1,0)")
        rep = audit_perturb_replay(p)
        assert rep.status == "PASS", rep.summary()
        assert rep.drivers_active >= 1
        assert rep.n_findings == 0

    def test_hardcoded_total_defect_caught(self, tmp_path):
        # total HARDCODED to its baseline value (20): ties at baseline, breaks
        # both up AND down → a genuine input-dependent conservation defect.
        p = _synth(tmp_path, "defect", total_formula="20.0",
                   check="=IF(ABS(Model!B2-Model!B1)<0.001,1,0)")
        rep = audit_perturb_replay(p)
        assert rep.status == "FAIL", rep.summary()
        assert rep.n_findings >= 1
        assert any("check" in f.check_label for f in rep.findings)

    def test_one_sided_validity_threshold_not_flagged(self, tmp_path):
        # A validity threshold "b >= 39" sits just below baseline (b=20? no — set
        # near the edge): breaks DOWN only, holds UP → must NOT be flagged.
        # b = 20 at baseline; threshold 19.5 → up(21)>=19.5 holds, down(19)<19.5
        # breaks. One-sided → PASS.
        p = _synth(tmp_path, "onesided", total_formula="=Model!B1",
                   check="=IF(Model!B1>=19.5,1,0)")
        rep = audit_perturb_replay(p)
        assert rep.status == "PASS", rep.summary()
        assert rep.n_findings == 0

    def test_no_qc_sheet_is_indeterminate(self, tmp_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Model"
        ws["A1"] = 5.0
        ws["B1"] = "=A1*2"
        p = tmp_path / "noqc.xlsx"
        wb.save(p)
        rep = audit_perturb_replay(p)
        assert rep.status == "INDETERMINATE"
        assert not rep.passed


@pytest.fixture(scope="module")
def built_workbook(tmp_path_factory) -> Path:
    out_dir = tmp_path_factory.mktemp("perturb")
    from modelforge.cli import _inject_trust_moat_and_finish, _load_spec_class
    from modelforge.templates import build_model

    spec_bytes = SPEC.read_bytes()
    raw = yaml.safe_load(spec_bytes)
    spec = _load_spec_class(raw["model_type"]).model_validate(raw)
    xlsx_out = out_dir / "sponsor_lbo_us_saas.xlsx"
    xlsx, _ = build_model(
        spec, xlsx_out, spec_source_bytes=spec_bytes, spec_source_path=SPEC)
    _inject_trust_moat_and_finish(xlsx, spec, spec_bytes, SPEC, quiet=True)
    return Path(xlsx)


class TestRealModel:
    def test_certified_model_invariants_hold(self, built_workbook):
        rep = audit_perturb_replay(built_workbook)
        # A correct certified model: every QC invariant holds two-sided under
        # input shocks → no breaks (zero false-positives).
        assert rep.status == "PASS", [
            (f.driver, f.check_label) for f in rep.findings
        ]
        assert rep.drivers_active >= 1
        assert rep.n_findings == 0
