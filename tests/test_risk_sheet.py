"""Tests for modelforge.analytics.risk_sheet — integrated RiskAnalysis sheet."""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from modelforge.analytics.risk_sheet import append_risk_analysis_sheet
from modelforge.qc import run_qc
from modelforge.spec.risk_block import EquityMarketData, RiskAnalysisSpec
from modelforge.templates import build_model

ROOT = Path(__file__).resolve().parent.parent


def _load_credit_memo_with_risk(risk_block: RiskAnalysisSpec = None):
    """Load credit_memo spec; inject a risk_analysis block programmatically."""
    from modelforge.spec.credit_memo import CreditMemoSpec
    p = ROOT / "examples" / "credit_memo_cdmo.yaml"
    raw = yaml.safe_load(p.read_bytes())
    spec = CreditMemoSpec.model_validate(raw)
    if risk_block is not None:
        spec = spec.model_copy(update={"risk_analysis": risk_block})
    return spec, p


# ── Sheet presence ─────────────────────────────────────────────────────────


def test_risk_sheet_added_when_block_present(tmp_path):
    spec, p = _load_credit_memo_with_risk()
    out = tmp_path / "cm.xlsx"
    build_model(spec, out, spec_source_bytes=p.read_bytes(), spec_source_path=p)
    wb = load_workbook(out)
    # Credit memo example file has risk_analysis block → sheet must exist
    assert "RiskAnalysis" in wb.sheetnames


def test_risk_sheet_skipped_when_block_missing(tmp_path):
    """Stripping risk_analysis should skip the sheet."""
    spec, p = _load_credit_memo_with_risk()
    spec_no_risk = spec.model_copy(update={"risk_analysis": None})
    out = tmp_path / "cm_no_risk.xlsx"
    build_model(spec_no_risk, out, spec_source_bytes=p.read_bytes(),
                spec_source_path=p)
    wb = load_workbook(out)
    assert "RiskAnalysis" not in wb.sheetnames


def test_risk_sheet_has_native_chart(tmp_path):
    spec, p = _load_credit_memo_with_risk()
    out = tmp_path / "cm.xlsx"
    build_model(spec, out, spec_source_bytes=p.read_bytes(), spec_source_path=p)
    wb = load_workbook(out)
    ws = wb["RiskAnalysis"]
    assert len(ws._charts) >= 1


# ── QC preservation ─────────────────────────────────────────────────────────


def test_qc_still_passes_with_risk_sheet(tmp_path):
    spec, p = _load_credit_memo_with_risk()
    out = tmp_path / "cm.xlsx"
    build_model(spec, out, spec_source_bytes=p.read_bytes(), spec_source_path=p)
    report = run_qc(out)
    assert report.all_pass, [c.name for c in report.checks if not c.passed]


# ── Content correctness ────────────────────────────────────────────────────


def test_risk_sheet_shows_merton_and_kmv(tmp_path):
    spec, p = _load_credit_memo_with_risk()
    out = tmp_path / "cm.xlsx"
    build_model(spec, out, spec_source_bytes=p.read_bytes(), spec_source_path=p)
    wb = load_workbook(out)
    ws = wb["RiskAnalysis"]
    labels = [ws.cell(row=r, column=1).value
              for r in range(1, ws.max_row + 1)
              if ws.cell(row=r, column=1).value]
    # Must show Merton solve results + KMV calibration
    assert any("Merton PD" in str(l) for l in labels)
    assert any("KMV empirical PD" in str(l) for l in labels)
    assert any("12-month ECL" in str(l) for l in labels)
    assert any("Lifetime ECL" in str(l) for l in labels)


def test_standalone_helper_returns_none_when_no_block(tmp_path):
    """append_risk_analysis_sheet should no-op on a spec without block."""
    from modelforge.spec.credit_memo import CreditMemoSpec
    p = ROOT / "examples" / "credit_memo_cdmo.yaml"
    raw = yaml.safe_load(p.read_bytes())
    raw.pop("risk_analysis", None)
    spec = CreditMemoSpec.model_validate(raw)
    out = tmp_path / "cm_none.xlsx"
    build_model(spec, out, spec_source_bytes=p.read_bytes(),
                spec_source_path=p)
    result = append_risk_analysis_sheet(out, spec)
    assert result is None


# ── Cross-template compatibility ───────────────────────────────────────────


@pytest.mark.parametrize("yaml_file,cls_mod,cls_name", [
    ("unitranche_cdmo.yaml", "modelforge.spec.unitranche", "UnitrancheSpec"),
    ("credit_memo_cdmo.yaml", "modelforge.spec.credit_memo", "CreditMemoSpec"),
    ("npl_mixed_portfolio.yaml", "modelforge.spec.npl", "NPLSpec"),
    ("structured_credit_pmi.yaml", "modelforge.spec.structured_credit",
     "StructuredCreditSpec"),
])
def test_risk_block_accepted_on_all_credit_templates(yaml_file, cls_mod,
                                                      cls_name, tmp_path):
    """All four credit-adjacent templates accept a risk_analysis block."""
    SC = getattr(__import__(cls_mod, fromlist=[cls_name]), cls_name)
    p = ROOT / "examples" / yaml_file
    raw = yaml.safe_load(p.read_bytes())
    spec = SC.model_validate(raw)
    risk = RiskAnalysisSpec(
        equity_market=EquityMarketData(
            equity_value_eur_m=100.0, equity_volatility=0.35,
            debt_face_value_eur_m=60.0,
        ),
        maturity_years=5,
    )
    # Re-validate via model_copy to ensure the field is assignable
    spec_with_risk = spec.model_copy(update={"risk_analysis": risk})
    out = tmp_path / f"{cls_name}.xlsx"
    build_model(spec_with_risk, out, spec_source_bytes=p.read_bytes(),
                spec_source_path=p)
    wb = load_workbook(out)
    assert "RiskAnalysis" in wb.sheetnames
