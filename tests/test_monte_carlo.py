"""Tests for modelforge.analytics.monte_carlo (US-002)."""

from __future__ import annotations

import time
from pathlib import Path

import numpy as np
import pytest
import yaml
from openpyxl import load_workbook

from modelforge.analytics.factors import SensitivityFactor, default_factors_for
from modelforge.analytics.monte_carlo import (
    MCConfig,
    append_monte_carlo_sheet,
    run_monte_carlo,
)
from modelforge.qc import run_qc
from modelforge.templates import build_model

ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture(scope="module")
def built_unitranche(tmp_path_factory):
    from modelforge.spec.unitranche import UnitrancheSpec
    p = ROOT / "examples" / "unitranche_cdmo.yaml"
    spec = UnitrancheSpec.model_validate(yaml.safe_load(p.read_bytes()))
    out = tmp_path_factory.mktemp("mc") / "u.xlsx"
    build_model(spec, out, spec_source_bytes=p.read_bytes(), spec_source_path=p)
    return out


def test_sheet_and_chart_present(built_unitranche):
    wb = load_workbook(built_unitranche)
    assert "MonteCarlo" in wb.sheetnames
    ws = wb["MonteCarlo"]
    assert len(ws._charts) >= 1


def _find_sampled_stats_header(ws):
    """Row of the SAMPLED-block 'Statistic' header.

    The MonteCarlo sheet now leads with the LIVE parametric band (which also
    has a 'Statistic' header) and the SAMPLED snapshot block follows it. The
    sampled header is the one whose col B reads 'Δ output (frac)'. Locating it
    by label keeps the test robust to row-offset shifts from the parametric
    band rather than re-pinning brittle literal row numbers.
    """
    for row in ws.iter_rows(min_col=1, max_col=2):
        a, b = row[0].value, row[1].value
        if (str(a).strip() == "Statistic"
                and str(b).strip().startswith("Δ output")):
            return row[0].row
    raise AssertionError("SAMPLED stats header not found on MonteCarlo sheet")


def test_distribution_stats_populated(built_unitranche):
    wb = load_workbook(built_unitranche)
    ws = wb["MonteCarlo"]
    # The 9 SAMPLED stats (Mean/Std/P5/P25/P50/P75/P95/Min/Max) are static
    # float literals (honest build-time snapshot) directly below the sampled
    # 'Statistic' header, in col B. Locate the header by label, not a pinned
    # row, since the LIVE parametric band now precedes the sampled block.
    hdr = _find_sampled_stats_header(ws)
    for r in range(hdr + 1, hdr + 10):
        v = ws.cell(row=r, column=2).value
        assert v is not None and isinstance(v, float), (
            f"sampled stat at row {r} should be a static float snapshot, got {v!r}"
        )


def test_parametric_band_is_live(built_unitranche):
    """The parametric band must be LIVE Excel formulas off primary_output + mc_vol.

    A sampled MC cannot recompute in vanilla Excel, but the parametric
    (normal-approx) downside band can. This asserts the band's P5/P50/P95/mean
    cells are formulas referencing the live ``primary_output`` named range and
    the editable ``mc_vol`` input — i.e. they are NOT frozen literals. The
    upgrade over the prior all-static MC snapshot.
    """
    wb = load_workbook(built_unitranche)
    ws = wb["MonteCarlo"]
    # mc_vol named range must exist (the live volatility input).
    assert "mc_vol" in wb.defined_names, "mc_vol named range missing"

    # Collect the parametric stat formulas (labels end with '(LIVE)').
    live_formulas = []
    for row in ws.iter_rows(min_col=1, max_col=2):
        a, b = row[0].value, row[1].value
        if isinstance(a, str) and a.strip().endswith("(LIVE)") and a != "Parametric (LIVE)":
            live_formulas.append((a.strip(), b))
    # Expect Mean + P5 + P50 + P95.
    labels = {lab for lab, _ in live_formulas}
    assert {"Mean (LIVE)", "P5 (LIVE)", "P50 (LIVE)", "P95 (LIVE)"} <= labels, labels

    # Every parametric stat is a formula (starts with '='); the percentile ones
    # reference primary_output, NORM.S.INV and mc_vol (genuinely reactive).
    for lab, formula in live_formulas:
        assert isinstance(formula, str) and formula.startswith("="), (
            f"{lab} must be a live formula, got {formula!r}"
        )
        assert "primary_output" in formula, f"{lab} must reference primary_output"
        if lab != "Mean (LIVE)":
            assert "NORM.S.INV" in formula and "mc_vol" in formula, (
                f"{lab} must be a normal-approx band off mc_vol, got {formula!r}"
            )


def test_parametric_band_reacts_to_scenario(built_unitranche, tmp_path):
    """PROOF of reactivity: the parametric band CHANGES when scenario_index flips,
    while the SAMPLED snapshot stays frozen.

    Evaluates the workbook with the ``formulas`` engine at scenario_index=2,
    then re-evaluates at scenario_index=1, by editing the scenario control cell
    (Cover!C17) in-file. The parametric P5/P50/P95 must move (a frozen snapshot
    would not); the sampled stats must NOT move (they are an honest snapshot).
    """
    formulas = pytest.importorskip("formulas")
    import shutil
    import warnings

    def _scenario_cell(wb):
        dn = wb.defined_names.get("scenario_index")
        # attr like "'Cover'!$C$17"
        ref = dn.attr_text.split("!")[1].replace("$", "")
        sheet = dn.attr_text.split("!")[0].strip("'")
        return sheet, ref

    def _eval(scenario):
        src = built_unitranche
        tmp = tmp_path / f"mc_scn_{scenario}.xlsx"
        shutil.copy(src, tmp)
        wb = load_workbook(tmp)
        sheet, ref = _scenario_cell(wb)
        wb[sheet][ref] = scenario
        wb.save(tmp)
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            xl = formulas.ExcelModel().loads(str(tmp)).finish()
            sol = xl.calculate()
        out = {}

        def g(v):
            v = getattr(v, "value", v)
            try:
                return float(v[0][0])
            except Exception:
                return float(v)

        for k, v in sol.items():
            ku = k.upper()
            if "MONTECARLO" not in ku:
                continue
            # parametric P5/P50/P95 (col B rows 12/13/14) + sampled P5 (B20)
            for cell, name in (("!B12", "par_P5"), ("!B13", "par_P50"),
                               ("!B14", "par_P95"), ("!B20", "samp_P5")):
                if ku.endswith(cell):
                    out[name] = g(v)
        return out

    s2 = _eval(2)
    s1 = _eval(1)
    # Parametric band cells must CHANGE between scenarios (reactive).
    for k in ("par_P5", "par_P50", "par_P95"):
        assert k in s2 and k in s1, f"{k} not evaluated"
        assert abs(s1[k] - s2[k]) > 1e-9, (
            f"parametric {k} should react to scenario flip "
            f"(s2={s2[k]}, s1={s1[k]})"
        )
    # Sampled snapshot must STAY frozen (it is an honest build-time snapshot).
    if "samp_P5" in s2 and "samp_P5" in s1:
        assert abs(s1["samp_P5"] - s2["samp_P5"]) < 1e-12, (
            "sampled snapshot must not change on scenario flip"
        )


def test_deterministic_with_seed():
    factors = default_factors_for("unitranche")
    r1 = run_monte_carlo(factors, MCConfig(n_runs=500, seed=42))
    r2 = run_monte_carlo(factors, MCConfig(n_runs=500, seed=42))
    np.testing.assert_array_almost_equal(r1.samples, r2.samples)


def test_different_seeds_diverge():
    factors = default_factors_for("unitranche")
    r1 = run_monte_carlo(factors, MCConfig(n_runs=500, seed=1))
    r2 = run_monte_carlo(factors, MCConfig(n_runs=500, seed=2))
    assert not np.allclose(r1.samples, r2.samples)


def test_n_runs_respected():
    factors = default_factors_for("unitranche")
    r = run_monte_carlo(factors, MCConfig(n_runs=2500, seed=0))
    assert len(r.samples) == 2500


def test_triangular_respects_bounds():
    """With triangular(left=-0.2, mode=0, right=0.2), draws should stay in [-0.2, 0.2]."""
    factors = [SensitivityFactor(driver_name="revenue_growth_y1", label="t",
                                 low_shock=-0.2, high_shock=0.2)]
    r = run_monte_carlo(factors, MCConfig(n_runs=2000, distribution="triangular", seed=0))
    contrib = r.factor_contributions["revenue_growth_y1"]
    # elasticity for revenue_growth_y1 is 0.8 → contrib bounded by [-0.16, 0.16]
    assert contrib.min() >= -0.16 - 1e-9
    assert contrib.max() <= 0.16 + 1e-9


def test_normal_distribution_produces_near_symmetric_mean():
    factors = default_factors_for("unitranche")
    r = run_monte_carlo(factors, MCConfig(n_runs=5000, distribution="normal", seed=0))
    assert abs(r.mean) < 0.05  # near zero


def test_percentiles_ordered():
    factors = default_factors_for("unitranche")
    r = run_monte_carlo(factors, MCConfig(n_runs=1000, seed=0))
    assert r.percentile(5) < r.percentile(50) < r.percentile(95)


def test_performance_1000_runs_under_5s():
    """PRD US-002 AC: 1000 runs on a 5-year deal in < 5s."""
    from modelforge.spec.unitranche import UnitrancheSpec
    p = ROOT / "examples" / "unitranche_cdmo.yaml"
    spec = UnitrancheSpec.model_validate(yaml.safe_load(p.read_bytes()))
    factors = default_factors_for("unitranche")
    t0 = time.time()
    r = run_monte_carlo(factors, MCConfig(n_runs=1000, seed=0))
    elapsed = time.time() - t0
    assert elapsed < 5.0
    assert len(r.samples) == 1000


def test_mc_sheet_on_all_templates(tmp_path):
    """Every template gets a MonteCarlo sheet after build_model.

    Builds each example into a temp dir (self-contained) rather than reading
    gitignored prebuilt output/*.xlsx, so the test passes from a clean checkout.
    """
    from modelforge.cli import _load_spec_class

    specs = [
        ("unitranche_cdmo.yaml", "unitranche"),
        ("minibond_logistics.yaml", "minibond"),
        ("credit_memo_cdmo.yaml", "credit_memo"),
        ("project_finance_solar.yaml", "project_finance"),
        ("real_estate_pbsa.yaml", "real_estate"),
        ("npl_mixed_portfolio.yaml", "npl"),
        ("structured_credit_pmi.yaml", "structured_credit"),
        ("three_statement_cdmo.yaml", "three_statement"),
    ]
    for fname, mt in specs:
        spec_path = ROOT / "examples" / fname
        raw = yaml.safe_load(spec_path.read_bytes())
        spec = _load_spec_class(raw.get("model_type", mt)).model_validate(raw)
        out = tmp_path / f"{Path(fname).stem}.xlsx"
        build_model(spec, out, spec_source_bytes=spec_path.read_bytes(), spec_source_path=spec_path)
        wb = load_workbook(out)
        assert "MonteCarlo" in wb.sheetnames, f"MC missing on {mt}"


def test_qc_still_passes_with_mc(built_unitranche):
    report = run_qc(built_unitranche)
    assert report.all_pass, [c.name for c in report.checks if not c.passed]
