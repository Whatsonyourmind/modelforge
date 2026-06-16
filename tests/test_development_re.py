"""Ground-up development RE template (Template 17) — feature tests.

Covers the development_re template END TO END:
    - DevelopmentRESpec validates from the shipped example YAML
    - both revenue kinds (pbsa + generic) validate; kind-validators reject
      missing inputs
    - the template builds a non-stub workbook with the expected sheets
    - headline cells (IRR / MOIC / NPV) are LIVE Excel formulas, not constants
    - the delivered artifact certifies CERTIFIED (zero formula errors, zero
      styling gaps) via the same auditor the `certify` CLI uses
    - every in-workbook QC check evaluates to PASS (1) under recalculation,
      including the sources = uses identity (the grant-timing fix)
    - registry / sensitivity-factor / ingest registrations are in place
"""

from __future__ import annotations

import copy
from pathlib import Path

import pytest
import yaml

from modelforge.spec.development_re import DevelopmentRESpec
from modelforge.templates import REGISTRY, build_model
from modelforge.qc import audit_workbook


ROOT = Path(__file__).resolve().parent.parent
EXAMPLES_DIR = ROOT / "examples"
OUTPUT_DIR = ROOT / "output" / "test_development_re"
EXAMPLE_YAML = EXAMPLES_DIR / "development_pbsa_genericcity.yaml"


# ───────────────────────── fixtures / helpers ─────────────────────────────────


def _load_example_dict() -> dict:
    with open(EXAMPLE_YAML, encoding="utf-8") as f:
        return yaml.safe_load(f)


@pytest.fixture(scope="module")
def example_dict() -> dict:
    return _load_example_dict()


@pytest.fixture(scope="module")
def built_workbook(example_dict) -> Path:
    """Build the FULL delivered artifact (post-processors included) once."""
    spec = DevelopmentRESpec.model_validate(example_dict)
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "development_pbsa_smoke.xlsx"
    xlsx_path, _ = build_model(spec, out_path)
    assert xlsx_path.exists()
    return xlsx_path


def _recalc_values(path: Path) -> dict[str, float]:
    """Recompute every cell with the `formulas` engine; return {SHEET!CELL: v}.

    Keys are upper-cased ``SHEET!CELL`` (matching the auditor's engine). Returns
    an empty dict if the engine is unavailable (so dependent tests skip).
    """
    try:
        import formulas  # type: ignore
    except Exception:  # pragma: no cover - engine optional in some envs
        return {}
    xl = formulas.ExcelModel().loads(str(path)).finish()
    sol = xl.calculate()
    out: dict[str, float] = {}
    for k, v in sol.items():
        # keys look like "'[book.xlsx]SHEET'!CELL"
        try:
            after = k.split("]", 1)[1]
            sheet, cell = after.split("'!", 1)
        except (IndexError, ValueError):
            continue
        try:
            val = v.value[0, 0]
        except Exception:
            val = v.value
        out[f"{sheet.strip(chr(39)).upper()}!{cell.strip()}"] = val
    return out


def _row_of(ws, needle: str) -> int | None:
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v and needle.lower() in str(v).lower():
            return r
    return None


# ───────────────────────── spec validation ────────────────────────────────────


def test_example_yaml_exists():
    assert EXAMPLE_YAML.exists(), f"Missing example: {EXAMPLE_YAML}"


def test_spec_validates_from_example(example_dict):
    spec = DevelopmentRESpec.model_validate(example_dict)
    assert spec.model_type == "development_re"
    assert spec.revenue.kind == "pbsa"
    assert spec.timeline.delivery_month == (
        spec.timeline.permit_months + spec.timeline.construction_months
    )
    # all_assumptions surfaces the active-kind drivers + waterfall + discount
    names = {a.name for a in spec.all_assumptions()}
    assert "dev_hard_costs" in names
    assert "dev_exit_cap_rate" in names
    assert "dev_beds" in names  # pbsa kind
    assert "dev_discount_rate" in names


def _leak_denylist() -> list[str]:
    """Real-world identifiers that must never appear in the shipped example.

    Kept OUT of version control on purpose: hard-coding the strings in this
    public test would itself publish the exact identifiers the check exists to
    suppress. Supply them via a gitignored ``tests/_leak_denylist.txt`` (one
    term per line, ``#`` comments allowed) or the ``MODELFORGE_LEAK_DENYLIST``
    env var (comma-separated). When neither is present the positive
    synthetic-identifier assertions below still run.
    """
    import os

    terms: list[str] = []
    f = Path(__file__).resolve().parent / "_leak_denylist.txt"
    if f.exists():
        terms += [
            ln.strip()
            for ln in f.read_text(encoding="utf-8").splitlines()
            if ln.strip() and not ln.startswith("#")
        ]
    terms += [t.strip() for t in os.environ.get("MODELFORGE_LEAK_DENYLIST", "").split(",") if t.strip()]
    return terms


def test_example_is_fully_synthetic_and_generic(example_dict):
    """Safety: the shipped example uses only synthetic, generic identifiers."""
    spec = DevelopmentRESpec.model_validate(example_dict)
    assert spec.capital.grant_name == "Public development grant"
    assert "genericcity" in EXAMPLE_YAML.name
    raw = EXAMPLE_YAML.read_text(encoding="utf-8").lower()
    for term in _leak_denylist():
        assert term.lower() not in raw, (
            "A denylisted real-world identifier appears in the shipped example"
        )


def test_timeline_validator_rejects_exit_before_delivery(example_dict):
    bad = copy.deepcopy(example_dict)
    # delivery = 6 + 18 = 24; set hold below it → must raise
    bad["timeline"]["hold_total_months"] = 20
    with pytest.raises(Exception):
        DevelopmentRESpec.model_validate(bad)


def test_pbsa_kind_requires_beds(example_dict):
    bad = copy.deepcopy(example_dict)
    bad["revenue"].pop("beds")
    with pytest.raises(Exception):
        DevelopmentRESpec.model_validate(bad)


def test_generic_revenue_kind_validates():
    """The generic sqm-based revenue kind validates and surfaces its drivers."""
    data = _load_example_dict()
    data["revenue"] = {
        "kind": "generic",
        "lettable_sqm": {
            "id": "A-010", "name": "dev_lettable_sqm",
            "label": {"en": "Lettable area"}, "unit": "count",
            "base": 9000, "rationale": "Synthetic NLA.", "confidence": "M",
            "source_id": "S-001",
        },
        "rent_sqm_year": {
            "id": "A-011", "name": "dev_rent_sqm_year",
            "label": {"en": "Rent per sqm / yr"}, "unit": "eur",
            "base": 320, "rationale": "Synthetic rent.", "confidence": "M",
            "source_id": "S-002",
        },
        "vacancy_pct": {
            "id": "A-012", "name": "dev_vacancy_pct",
            "label": {"en": "Vacancy %"}, "unit": "pct",
            "base": 0.05, "rationale": "Synthetic vacancy.", "confidence": "M",
            "source_id": "S-002",
        },
        "opex_per_unit_year": {
            "id": "A-013", "name": "dev_opex_per_sqm_year",
            "label": {"en": "Opex per sqm / yr"}, "unit": "eur",
            "base": 60, "rationale": "Synthetic opex.", "confidence": "M",
            "source_id": "S-001",
        },
        "rev_growth_pct": {
            "id": "A-014", "name": "dev_rev_growth_pct",
            "label": {"en": "NOI growth"}, "unit": "pct",
            "base": 0.025, "rationale": "Synthetic growth.", "confidence": "M",
            "source_id": "S-002",
        },
    }
    spec = DevelopmentRESpec.model_validate(data)
    assert spec.revenue.kind == "generic"
    names = {a.name for a in spec.all_assumptions()}
    assert "dev_lettable_sqm" in names
    assert "dev_rent_sqm_year" in names
    assert "dev_beds" not in names  # pbsa-only driver excluded


# ───────────────────────── template build ─────────────────────────────────────


def test_template_builds_non_stub(built_workbook):
    assert built_workbook.stat().st_size > 5000


def test_workbook_has_expected_sheets(built_workbook):
    from openpyxl import load_workbook

    wb = load_workbook(built_workbook)
    for name in ("DevSchedule", "Returns", "QC", "ComplianceCheck"):
        assert name in wb.sheetnames, f"Missing sheet: {name}"


def test_headline_cells_are_formulas_not_constants(built_workbook):
    """IRR / MOIC / NPV headline cells must be LIVE formulas."""
    from openpyxl import load_workbook

    wb = load_workbook(built_workbook)
    ret = wb["Returns"]
    for label in ("Unlevered IRR", "Levered equity IRR",
                  "Levered equity MOIC", "Levered equity NPV"):
        r = _row_of(ret, label)
        assert r is not None, f"Row not found: {label}"
        v = ret.cell(row=r, column=4).value
        assert isinstance(v, str) and v.startswith("="), (
            f"{label} (Returns!D{r}) is not a formula: {v!r}"
        )
    # The IRR cells genuinely call IRR(...)
    irr_row = _row_of(ret, "Levered equity IRR")
    assert "IRR(" in ret.cell(row=irr_row, column=4).value


def test_schedule_outputs_are_formulas(built_workbook):
    """Key DevSchedule output rows are live formulas, not precomputed numbers."""
    from openpyxl import load_workbook

    wb = load_workbook(built_workbook)
    sched = wb["DevSchedule"]
    for label in ("Total development cost", "Stabilised annual NOI",
                  "Net exit proceeds", "Closing debt balance"):
        r = _row_of(sched, label)
        assert r is not None, f"Row not found: {label}"
        v = sched.cell(row=r, column=4).value
        assert isinstance(v, str) and v.startswith("="), (
            f"{label} (DevSchedule!D{r}) is not a formula: {v!r}"
        )


# ───────────────────────── certification ──────────────────────────────────────


def test_workbook_certifies_clean(built_workbook):
    """The delivered artifact must CERTIFY: zero formula errors, zero gaps."""
    report = audit_workbook(built_workbook)
    s = report.summary()
    assert s["error_cells"] == 0, (
        f"Formula errors present: "
        f"{[(e.sheet, e.cell, e.code) for e in report.error_cells]}"
    )
    assert s["style_gaps"] == 0, (
        f"Styling gaps present: "
        f"{[(g.sheet, g.cell) for g in report.style_gaps]}"
    )
    assert report.verdict == "CERTIFIED"


def test_all_qc_checks_pass_under_recalc(built_workbook):
    """Every in-workbook QC check (incl. sources=uses) must equal 1."""
    vals = _recalc_values(built_workbook)
    if not vals:
        pytest.skip("formulas engine unavailable")
    # ALL CHECKS PASS aggregator
    assert vals.get("QC!C4") == 1, "QC ALL-PASS aggregator is not 1"
    # individual checks C7.. (8 checks after adding debt-conservation + IDC)
    for r in range(7, 15):
        key = f"QC!C{r}"
        if key in vals:
            assert vals[key] == 1, f"QC check {key} failed (={vals[key]})"


def test_levered_irr_is_positive_and_sane(built_workbook):
    """Recalculated levered equity IRR is a positive, plausible number.

    Post-roll-forward-fix the tuned PBSA development genuinely earns a
    mid-to-high-teens levered equity IRR (the construction-interest
    capitalisation bug previously overstated this ~5x). Assert a plausible
    band, not a brittle point value.
    """
    from openpyxl import load_workbook

    vals = _recalc_values(built_workbook)
    if not vals:
        pytest.skip("formulas engine unavailable")
    ret = load_workbook(built_workbook)["Returns"]
    irr_row = _row_of(ret, "Levered equity IRR")
    irr = vals.get(f"RETURNS!D{irr_row}")
    assert irr is not None
    assert 0.0 < irr < 1.0, f"Levered IRR out of plausible range: {irr}"
    # Tuned ground-up PBSA: levered equity IRR lands in the mid-to-high teens.
    assert 0.10 < irr < 0.30, f"Levered IRR outside the expected band: {irr}"


def test_positive_leverage_levered_above_unlevered(built_workbook):
    """The tuned deal has POSITIVE leverage: levered equity IRR is strictly
    above the unlevered project IRR (yield-on-cost > debt cost). This would
    FAIL under the construction-interest-capitalisation bug, which collapsed
    levered IRR below unlevered (negative leverage)."""
    from openpyxl import load_workbook

    vals = _recalc_values(built_workbook)
    if not vals:
        pytest.skip("formulas engine unavailable")
    ret = load_workbook(built_workbook)["Returns"]
    lev = vals.get(f"RETURNS!D{_row_of(ret, 'Levered equity IRR')}")
    unlev = vals.get(f"RETURNS!D{_row_of(ret, 'Unlevered IRR')}")
    assert lev is not None and unlev is not None
    assert 0.0 < unlev < 1.0, f"Unlevered IRR implausible: {unlev}"
    assert lev > unlev, f"Negative leverage: lev={lev} <= unlev={unlev}"


def test_construction_interest_capitalises(built_workbook):
    """The senior facility must capitalise interest during construction (the
    sheet is titled 'IDC capitalised'): the opening balance grows year-on-year
    through the build, and cumulative IDC is strictly positive. The historical
    bug left opening debt at ~0 every year, so no interest ever capitalised."""
    from openpyxl import load_workbook

    vals = _recalc_values(built_workbook)
    if not vals:
        pytest.skip("formulas engine unavailable")
    sched = load_workbook(built_workbook)["DevSchedule"]
    open_row = _row_of(sched, "Opening debt balance")
    idc_row = _row_of(sched, "Interest capitalised")
    # year columns D.. on the schedule
    year_cols = []
    for c in range(4, sched.max_column + 1):
        if sched.cell(row=5, column=c).value:
            year_cols.append(sched.cell(row=5, column=c).column_letter)
    opens = [vals.get(f"DEVSCHEDULE!{col}{open_row}") for col in year_cols]
    opens = [o for o in opens if o is not None]
    # opening debt strictly increases across the early (construction) years
    assert opens[2] > opens[1] > 0, f"opening debt did not compound: {opens[:3]}"
    total_idc = sum(
        vals.get(f"DEVSCHEDULE!{col}{idc_row}") or 0.0 for col in year_cols
    )
    assert total_idc > 0.001, f"no construction interest capitalised: {total_idc}"


def test_debt_conservation_qc_present_and_passes(built_workbook):
    """The debt-conservation QC invariant (Σdraws+ΣIDC == Σprincipal repaid)
    is wired into the QC sheet and evaluates PASS. This is the check that
    auto-catches a broken roll-forward."""
    from openpyxl import load_workbook

    vals = _recalc_values(built_workbook)
    if not vals:
        pytest.skip("formulas engine unavailable")
    qc = load_workbook(built_workbook)["QC"]
    cons_row = _row_of(qc, "conserved")
    assert cons_row is not None, "debt-conservation QC row missing from QC sheet"
    assert vals.get(f"QC!C{cons_row}") == 1, "debt-conservation QC check failed"


def test_debt_repaid_to_zero_at_exit(built_workbook):
    """Closing senior debt at the exit column must be ~0."""
    from openpyxl import load_workbook

    vals = _recalc_values(built_workbook)
    if not vals:
        pytest.skip("formulas engine unavailable")
    sched = load_workbook(built_workbook)["DevSchedule"]
    cd_row = _row_of(sched, "Closing debt balance")
    # exit column is the last year column on the schedule
    last_col = None
    for c in range(4, sched.max_column + 1):
        if sched.cell(row=5, column=c).value:  # header row has t=i tags
            last_col = sched.cell(row=5, column=c).column_letter
    assert last_col is not None
    closing = vals.get(f"DEVSCHEDULE!{last_col}{cd_row}")
    assert closing is not None
    assert abs(closing) < 0.01, f"Debt not repaid at exit: {closing}"


# ───────────────────────── registrations ──────────────────────────────────────


def test_registered_in_template_registry():
    assert "development_re" in REGISTRY


def test_registered_in_spec_loader_map():
    from modelforge.cli import _load_spec_class

    cls = _load_spec_class("development_re")
    assert cls is DevelopmentRESpec


def test_has_curated_template_description():
    from modelforge.cli import _template_description

    desc = _template_description("development_re")
    assert isinstance(desc, str) and len(desc) > 10
    assert "development" in desc.lower()


def test_has_sensitivity_factors():
    from modelforge.analytics.factors import default_factors_for

    factors = default_factors_for("development_re")
    assert len(factors) >= 6
    names = {f.driver_name for f in factors}
    assert "dev_exit_cap_rate" in names


def test_registered_for_ingest():
    from modelforge.ingest.pipeline import TEMPLATE_SECTIONS

    assert "development_re" in TEMPLATE_SECTIONS
    spec_cls, sections = TEMPLATE_SECTIONS["development_re"]()
    assert spec_cls is DevelopmentRESpec
    assert sections[0][0] == "target"


def test_scaffold_seed_present():
    from modelforge.spec.scaffold import _SEED_EXAMPLE

    assert _SEED_EXAMPLE.get("development_re") == "development_pbsa_genericcity.yaml"
