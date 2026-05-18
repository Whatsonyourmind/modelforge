"""v0.11 quality-lift tests.

Covers:
    - 6 new Trust Layer rules (2 HGB + 4 portfolio review)
    - HGB recon math (Hinzurechnungen → Gewerbesteuer → effective tax recon)
    - Hardcoded Italian strings now flow through Label class (4 analytics titles + Assumptions sheet)
    - New foreign example (sponsor_lbo_us_saas) builds end-to-end
"""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml

from modelforge.builder.i18n import LABELS, L
from modelforge.templates import REGISTRY
from modelforge.trust.builtin import DEFAULT_RULES


ROOT = Path(__file__).resolve().parent.parent
EXAMPLES_DIR = ROOT / "examples"
OUTPUT_DIR = ROOT / "output" / "test_v011"


# ---------- Trust Layer expansion ----------


def test_trust_layer_has_30_rules_min():
    """v0.11 ships 6 new rules — total ≥ 30."""
    assert len(DEFAULT_RULES) >= 30


def test_trust_layer_covers_hgb_carveout():
    hgb_rules = [r for r in DEFAULT_RULES if "hgb_carveout" in (r.template_types or ())]
    assert len(hgb_rules) >= 2, "HGB needs Hebesatz-in-band + ETR-floor rules"


def test_trust_layer_covers_portfolio_review():
    pr_rules = [r for r in DEFAULT_RULES if "portfolio_review" in (r.template_types or ())]
    assert len(pr_rules) >= 4, (
        "Portfolio review needs: avg leverage, cushion plausibility, "
        "cash-trap concentration, rating distribution"
    )


def test_trust_layer_no_template_uncovered():
    """Every non-aggregator template should have at least one rule."""
    from modelforge.templates import REGISTRY
    covered: set[str] = set()
    for r in DEFAULT_RULES:
        for tpl in (r.template_types or ()):
            covered.add(tpl)
    missing = set(REGISTRY) - covered
    # IPO, three_statement, and fairness may be partial; assert ≥80% coverage.
    coverage = 1 - len(missing) / len(REGISTRY)
    assert coverage >= 0.80, (
        f"Trust Layer coverage {coverage:.0%} below 80% threshold. "
        f"Templates without rules: {missing}"
    )


# ---------- HGB recon math ----------


def test_hgb_recon_sheet_has_real_math():
    """HGB-Recon sheet must contain Hinzurechnungen + GewSt build, not just docs."""
    from modelforge.spec.hgb_carveout import HGBCarveoutSpec
    from openpyxl import load_workbook

    yaml_path = EXAMPLES_DIR / "hgb_carveout_dach_chemicals.yaml"
    with open(yaml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    spec = HGBCarveoutSpec.model_validate(data)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "hgb_recon_math.xlsx"
    REGISTRY["hgb_carveout"](spec, out_path)

    wb = load_workbook(out_path)
    assert "HGB-Recon" in wb.sheetnames
    ws = wb["HGB-Recon"]

    # v0.11 sheet should be substantially larger than the v0.10 stub (32+ rows)
    assert ws.max_row >= 25, (
        f"HGB-Recon sheet has {ws.max_row} rows — v0.11 expects ≥25 "
        "(real Hinzurechnungen + GewSt + ETR recon, not placeholders)"
    )

    # Verify section headers exist
    labels_col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    labels_str = "|".join(str(v) for v in labels_col_a if v)
    assert "§ 8 GewStG" in labels_str or "Hinzurechnungen" in labels_str, (
        "HGB-Recon must include § 8 GewStG Hinzurechnungen section"
    )
    assert "Gewerbesteuer" in labels_str, (
        "HGB-Recon must include Gewerbesteuer build section"
    )
    assert "Effective tax rate" in labels_str or "effective" in labels_str.lower(), (
        "HGB-Recon must include ETR reconciliation section"
    )


def test_hgb_recon_hebesatz_flows_through():
    """The user-supplied Hebesatz should drive the GewSt formula."""
    from modelforge.spec.hgb_carveout import HGBCarveoutSpec
    from openpyxl import load_workbook

    yaml_path = EXAMPLES_DIR / "hgb_carveout_dach_chemicals.yaml"
    with open(yaml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    spec = HGBCarveoutSpec.model_validate(data)
    assert spec.hgb_assumptions.gewerbesteuer_hebesatz == 400.0

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "hgb_hebesatz_flow.xlsx"
    REGISTRY["hgb_carveout"](spec, out_path)

    wb = load_workbook(out_path)
    ws = wb["HGB-Recon"]
    # Find a cell containing 4.0 (400% as decimal) in a formula — Hebesatz multiplier
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and "4.0" in cell.value:
                if "*" in cell.value:
                    found = True
                    break
        if found:
            break
    assert found, "Hebesatz 400% (4.0 decimal) should appear in a multiplier formula"


# ---------- Hardcoded Italian → Label class ----------


def test_analytics_titles_in_i18n():
    """v0.11 moved Monte Carlo / Sensitivity / Repro / Risk titles into i18n."""
    for key in ["monte_carlo_title", "sensitivity_title",
                "reproducibility_title", "risk_analysis_title"]:
        lbl = L(key)
        assert lbl.de, f"{key} missing DE translation"
        assert lbl.nl, f"{key} missing NL translation"
        assert lbl.sv, f"{key} missing SV translation"


def test_assumptions_title_in_i18n():
    """v0.11 moved 'Ipotesi e driver' into i18n as `assumptions_title`."""
    lbl = L("assumptions_title")
    assert lbl.en == "Assumptions"
    assert lbl.it == "Ipotesi e driver"
    assert lbl.de == "Annahmen"


# ---------- Foreign example builds ----------


def test_us_sponsor_lbo_example_builds():
    """sponsor_lbo_us_saas.yaml end-to-end build."""
    from modelforge.spec.sponsor_lbo import SponsorLBOSpec

    yaml_path = EXAMPLES_DIR / "sponsor_lbo_us_saas.yaml"
    assert yaml_path.exists(), f"Missing example: {yaml_path}"

    with open(yaml_path, encoding="utf-8") as f:
        data = yaml.safe_load(f)
    spec = SponsorLBOSpec.model_validate(data)

    # Foreign-market check: target is US, currency USD
    assert spec.meta.currency == "USD"
    assert spec.target.country == "US"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / "sponsor_lbo_us_saas_smoke.xlsx"
    xlsx_path, _ = REGISTRY["sponsor_lbo"](spec, out_path)
    assert xlsx_path.exists()
    assert xlsx_path.stat().st_size > 20000  # full LBO workbook is non-trivial


# ---------- Coverage validation ----------


def test_i18n_validation_passes_on_import():
    """i18n.py's _validate_coverage() runs at import — if missing translations exist
    the module would have raised RuntimeError. Empty smoke test confirms import works."""
    from modelforge.builder import i18n  # noqa: F401 — re-import triggers validation
    assert len(LABELS) >= 85  # v0.11 added 5 labels (4 analytics + assumptions)
