"""Loan-tape cash securitization (Template 19) — build, recalc, registration, wiring.

Verifies the model builds, recalculates with zero formula errors, the QC sheet
all-checks-pass aggregator evaluates to 1, every note balance rolls forward from
the prior period's CLOSING row (the off-by-one-prone construct), and the key
conservation invariants are present and economically correct (the residual
certificate absorbs exactly the net credit loss).
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pytest
import yaml

from modelforge.spec.loan_tape_securitization import LoanTapeSecuritizationSpec
from modelforge.templates import REGISTRY, build_model

_SPEC = Path("examples/clo_midmarket.yaml")
_REF_RE = re.compile(r"=\$?[A-Z]+\$?(\d+)$")


@pytest.fixture(scope="module")
def spec():
    return LoanTapeSecuritizationSpec.model_validate(yaml.safe_load(_SPEC.read_bytes()))


@pytest.fixture(scope="module")
def built(spec, tmp_path_factory):
    out = tmp_path_factory.mktemp("clo") / "clo.xlsx"
    build_model(spec, out, with_manifest=False)  # builder asserts fire here on drift
    return out


def _row(ws, prefix):
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if isinstance(c.value, str) and c.value.strip().startswith(prefix):
            return c.row
    return None


def _sol(path):
    formulas = pytest.importorskip("formulas")
    return formulas.ExcelModel().loads(str(path)).finish().calculate()


def _get(sol, fname, sheet, addr):
    want = f"'[{fname}]{sheet.upper()}'!{addr}"
    for k, v in sol.items():
        if k.upper() == want:
            try:
                return v.value[0, 0]
            except Exception:
                return getattr(v, "value", v)
    return None


# ── registration ─────────────────────────────────────────────────────────


def test_registered_in_registry():
    assert "loan_tape_securitization" in REGISTRY


def test_registered_in_spec_loader_map():
    from modelforge.cli import _spec_loader_map
    assert _spec_loader_map()["loan_tape_securitization"]() is LoanTapeSecuritizationSpec


def test_has_curated_template_description():
    from modelforge.cli import _template_description
    d = _template_description("loan_tape_securitization").lower()
    assert "securitization" in d or "clo" in d


def test_default_factors_cover_template():
    from modelforge.analytics.factors import default_factors_for
    assert len(default_factors_for("loan_tape_securitization")) >= 6


def test_ingest_sections_registered():
    from modelforge.ingest.pipeline import TEMPLATE_SECTIONS
    assert "loan_tape_securitization" in TEMPLATE_SECTIONS


# ── build + recalc ─────────────────────────────────────────────────────────


def test_builds_expected_sheets(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    for s in ("Cover", "Assumptions", "LoanTape", "Waterfall", "Notes",
              "QC", "ComplianceCheck"):
        assert s in wb.sheetnames, f"missing sheet {s}"


def test_recalc_no_errors_and_qc_passes(built):
    sol = _sol(built)
    fname = Path(built).name.upper()
    errs = []
    for k, v in sol.items():
        try:
            val = v.value[0, 0]
        except Exception:
            val = getattr(v, "value", v)
        if isinstance(val, str) and val.startswith("#") and val.endswith(("!", "?", "M", "L", "/", "A")):
            errs.append((k, val))
    assert not errs, f"formula errors: {errs[:10]}"
    assert _get(sol, fname, "QC", "C4") == 1, "QC ALL_PASS != 1"


# ── roll-forward correctness (symbolic prior-closing, not a literal offset) ──


def test_stratum_opening_rolls_from_prior_closing(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    lt = wb["LoanTape"]
    # First stratum block: opening (BOP) and closing (EOP) rows.
    open_row = _row(lt, "Performing balance — opening")
    close_row = _row(lt, "Performing balance — closing")
    assert open_row and close_row
    # Projection col F (t=2) opening must reference the prior column's CLOSING row.
    cell = lt[f"F{open_row}"].value
    assert isinstance(cell, str)
    m = _REF_RE.match(cell.strip())
    assert m and int(m.group(1)) == close_row, (
        f"stratum opening F{open_row}='{cell}' must reference closing row {close_row}"
    )


def test_note_balances_roll_from_prior_closing(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    wf = wb["Waterfall"]
    eop = _row(wf, "Outstanding — Class B")  # mezz: plain prior − paydown roll
    assert eop is not None
    # t=2 (col F) must reference the prior column's same EOP row.
    cell = wf[f"F{eop}"].value
    assert isinstance(cell, str) and f"E{eop}" in cell.replace("$", ""), (
        f"note EOP F{eop}='{cell}' must roll from prior column E{eop}"
    )


def test_qc_has_conservation_and_subordination_checks(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    labels = " | ".join(
        c.value for col in wb["QC"].iter_rows(min_col=1, max_col=1)
        for c in col if isinstance(c.value, str))
    assert "telescopes" in labels
    assert "Recovery conservation" in labels
    assert "Recovery timing" in labels
    assert "strictly sequential" in labels
    assert "Subordination" in labels
    assert "No excess spread to equity" in labels
    assert "Reserve returns to equity" in labels


# ── economic correctness ─────────────────────────────────────────────────


def test_residual_absorbs_net_credit_loss(built, spec):
    """The first-loss residual's writedown == pool defaults − recoveries, and
    senior + mezz are fully redeemed (the structure performs)."""
    sol = _sol(built)
    fname = Path(built).name.upper()
    wb = openpyxl.load_workbook(built, data_only=False)
    lt = wb["LoanTape"]
    last = chr(ord("D") + spec.horizon.periods)

    def lt_sum(prefix):
        r = _row(lt, prefix)
        return sum(_get(sol, fname, "LoanTape", f"{chr(ord('E') + i)}{r}")
                   for i in range(spec.horizon.periods))

    net_loss = lt_sum("Pool defaults") - lt_sum("Pool recoveries")

    wf = wb["Waterfall"]
    sr = _row(wf, "Outstanding — Class A")   # senior
    mr = _row(wf, "Outstanding — Class B")   # mezz
    senior_final = _get(sol, fname, "Waterfall", f"{last}{sr}")
    mezz_final = _get(sol, fname, "Waterfall", f"{last}{mr}")
    assert abs(senior_final) < 0.05, f"senior not redeemed ({senior_final})"
    assert abs(mezz_final) < 0.05, f"mezz not redeemed ({mezz_final})"

    # Residual writedown reported on the Notes analytics (loss column K).
    nt = wb["Notes"]
    res_row = None
    for row in nt.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if isinstance(c.value, str) and c.value.strip().startswith("Residual"):
            res_row = c.row
            break
    res_loss = _get(sol, fname, "Notes", f"K{res_row}")
    assert abs(res_loss - net_loss) < 0.05, (
        f"residual writedown {res_loss} != net credit loss {net_loss}"
    )


def test_turbo_breach_diverts_to_debt_not_equity(tmp_path):
    """The shipped example never breaches OC/IC (turbo == 0). Stress a
    high-default / tight-trigger variant so the turbo fires, and assert the
    core sequential-pay property the adversarial review hardened: while an
    OC/IC test is breached AND a debt note is still outstanding, NO excess
    spread leaks to the first-loss equity — it is diverted to a turbo paydown
    of the most-senior outstanding note. Also: no formula errors, no negative
    note balances, and residual interest never negative."""
    raw = yaml.safe_load(_SPEC.read_bytes())
    raw["tape"][0]["cdr_pct"]["base"] = 0.10
    raw["tape"][1]["cdr_pct"]["base"] = 0.22
    raw["pool"]["recovery_pct"]["base"] = 0.30
    raw["enhancement"]["oc_trigger_pct"]["base"] = 1.40
    raw["enhancement"]["ic_trigger_pct"]["base"] = 1.60
    spec = LoanTapeSecuritizationSpec.model_validate(raw)
    out = tmp_path / "stress.xlsx"
    REGISTRY["loan_tape_securitization"](spec, out)  # core sheets only
    sol = _sol(out)
    fname = out.name.upper()
    wb = openpyxl.load_workbook(out, data_only=False)
    wf = wb["Waterfall"]
    P = spec.horizon.periods
    cols = [chr(ord("D") + i) for i in range(P + 1)]

    errs = [k for k, v in sol.items()
            if isinstance(getattr(v, "value", None), str) and str(v.value).startswith("#")]
    assert not errs, f"formula errors under stress: {errs[:5]}"

    def gw(prefix, c):
        return _get(sol, fname, "Waterfall", f"{c}{_row(wf, prefix)}")

    # rows for every turbo leg and every note's EOP
    turbo_rows = [r[0].row for r in wf.iter_rows(min_col=1, max_col=1)
                  if isinstance(r[0].value, str) and r[0].value.strip().startswith("Turbo principal")]
    eop_rows = [r[0].row for r in wf.iter_rows(min_col=1, max_col=1)
                if isinstance(r[0].value, str) and r[0].value.strip().startswith("Outstanding —")]
    debt_eop_rows = eop_rows[:-1]  # last is the residual note

    def total_turbo(c):
        return sum(_get(sol, fname, "Waterfall", f"{c}{r}") for r in turbo_rows)

    assert any(total_turbo(c) > 0.01 for c in cols), "turbo never fired under stress"

    for i in range(1, P + 1):
        c = cols[i]
        oc_ok = gw("OC test pass", c)
        debt_out = sum(_get(sol, fname, "Waterfall", f"{c}{r}") for r in debt_eop_rows)
        rie = gw("Residual interest to equity", c)
        assert rie >= -0.01, f"residual interest to equity negative at t={i}: {rie}"
        # THE FIX: breach + debt still outstanding ⇒ no excess spread to equity
        if oc_ok == 0 and debt_out > 0.01:
            assert rie < 0.01, (
                f"excess spread leaked to equity at t={i} during an OC breach with "
                f"€{debt_out:.2f}m debt outstanding (rie={rie})"
            )

    for r in eop_rows:
        for c in cols:
            assert _get(sol, fname, "Waterfall", f"{c}{r}") >= -0.01, "negative note balance"


def test_no_negative_coupon_when_fees_exceed_yield(tmp_path):
    """Fee-heavy / low-yield stress (servicing fee > WAC) drives available
    interest negative. The interest waterfall must floor every note's coupon at
    zero — a note (the AAA in particular) can never pay cash INTO the SPV."""
    raw = yaml.safe_load(_SPEC.read_bytes())
    raw["tape"][0]["wac_pct"]["base"] = 0.005
    raw["tape"][1]["wac_pct"]["base"] = 0.005
    raw["pool"]["servicing_fee_pct"]["base"] = 0.02
    spec = LoanTapeSecuritizationSpec.model_validate(raw)
    out = tmp_path / "feedrag.xlsx"
    REGISTRY["loan_tape_securitization"](spec, out)
    sol = _sol(out)
    fname = out.name.upper()
    wf = openpyxl.load_workbook(out, data_only=False)["Waterfall"]
    P = spec.horizon.periods
    cols = [chr(ord("D") + i) for i in range(P + 1)]
    ip_rows = [r[0].row for r in wf.iter_rows(min_col=1, max_col=1)
               if isinstance(r[0].value, str) and r[0].value.strip().startswith("Interest paid —")]
    for r in ip_rows:
        for c in cols:
            v = _get(sol, fname, "Waterfall", f"{c}{r}")
            assert v >= -0.01, f"negative coupon paid at {c}{r}: {v}"
