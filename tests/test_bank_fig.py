"""Bank / FIG template (Template 18) — build, recalc, registration, wiring.

Verifies the model builds, recalculates with zero formula errors, the QC sheet
all-checks-pass aggregator evaluates to 1, the balance sheet balances every
period, and the two roll-forwards (common equity, allowance) reference the prior
period's CLOSING row symbolically (the off-by-one-prone construct).
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pytest
import yaml

from modelforge.spec.bank_fig import BankFigSpec
from modelforge.templates import REGISTRY, build_model

_SPEC = Path("examples/bank_fig_meridian.yaml")
_REF_RE = re.compile(r"=\$?[A-Z]+\$?(\d+)$")


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    raw = yaml.safe_load(_SPEC.read_bytes())
    spec = BankFigSpec.model_validate(raw)
    out = tmp_path_factory.mktemp("bank_fig") / "bank.xlsx"
    build_model(spec, out, with_manifest=False)  # asserts in builders fire here if drift
    return out


def _row(ws, prefix):
    for row in ws.iter_rows(min_col=1, max_col=1):
        c = row[0]
        if isinstance(c.value, str) and c.value.strip().startswith(prefix):
            return c.row
    return None


# ── registration ─────────────────────────────────────────────────────────


def test_registered_in_registry():
    assert "bank_fig" in REGISTRY


def test_registered_in_spec_loader_map():
    from modelforge.cli import _spec_loader_map
    assert _spec_loader_map()["bank_fig"]() is BankFigSpec


def test_has_curated_template_description():
    from modelforge.cli import _template_description
    assert "bank" in _template_description("bank_fig").lower()


def test_default_factors_cover_bank_fig():
    from modelforge.analytics.factors import default_factors_for
    assert len(default_factors_for("bank_fig")) >= 6


# ── build + recalc ─────────────────────────────────────────────────────────


def test_builds_expected_sheets(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    for s in ("Cover", "Assumptions", "NII", "P&L", "BalanceSheet",
              "Capital", "CapitalReturn", "QC", "ComplianceCheck"):
        assert s in wb.sheetnames, f"missing sheet {s}"


def test_recalc_has_no_formula_errors_and_qc_passes(built):
    formulas = pytest.importorskip("formulas")
    fname = Path(built).name.upper()
    xl = formulas.ExcelModel().loads(str(built)).finish()
    sol = xl.calculate()

    errs = []
    for k, v in sol.items():
        try:
            val = v.value[0, 0]
        except Exception:
            val = getattr(v, "value", v)
        if isinstance(val, str) and val.startswith("#") and val.endswith(("!", "?", "M", "L", "/", "A")):
            errs.append((k, val))
    assert not errs, f"formula errors in recalc: {errs[:10]}"

    # QC ALL_PASS aggregator at C4 must equal 1.
    want = f"'[{fname}]QC'!C4"
    allpass = None
    for k, v in sol.items():
        if k.upper() == want:
            allpass = v.value[0, 0] if hasattr(v.value, "__getitem__") else v.value
    assert allpass == 1, f"QC ALL_PASS != 1 (got {allpass})"


def test_balance_sheet_balances_every_period(built):
    formulas = pytest.importorskip("formulas")
    fname = Path(built).name.upper()
    wb = openpyxl.load_workbook(built, data_only=False)
    bs = wb["BalanceSheet"]
    chk = _row(bs, "Balance check")
    assert chk is not None
    xl = formulas.ExcelModel().loads(str(built)).finish()
    sol = xl.calculate()
    n = 1 + 5  # historical + projection
    for i in range(n):
        col = chr(ord("D") + i)
        want = f"'[{fname}]BALANCESHEET'!{col}{chk}"
        v = next((vv for kk, vv in sol.items() if kk.upper() == want), None)
        assert v is not None, f"no value for {want}"
        val = v.value[0, 0]
        assert abs(float(val)) < 0.01, f"balance check {col}{chk} = {val} (≠ 0)"


# ── roll-forward correctness (symbolic prior-closing, not a literal offset) ──


def test_equity_and_allowance_roll_from_prior_closing(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    bs = wb["BalanceSheet"]
    eq_open = _row(bs, "Common equity — opening")
    eq_close = _row(bs, "Common equity — closing")
    alw_open = _row(bs, "Loan-loss allowance — opening")
    alw_close = _row(bs, "Loan-loss allowance — closing")
    assert all((eq_open, eq_close, alw_open, alw_close))

    # Opening (projection col E = second column) must reference the prior
    # column's CLOSING row, never the repayment/charge row.
    for open_row, close_row, name in (
        (eq_open, eq_close, "equity"), (alw_open, alw_close, "allowance"),
    ):
        cell = bs[f"E{open_row}"].value
        assert isinstance(cell, str), f"{name} opening E{open_row} not a formula"
        m = _REF_RE.match(cell.strip())
        assert m and int(m.group(1)) == close_row, (
            f"{name} opening (E{open_row})='{cell}' must reference closing row "
            f"{close_row}, not {m.group(1) if m else '?'}"
        )


def test_qc_has_conservation_and_capital_checks(built):
    wb = openpyxl.load_workbook(built, data_only=False)
    qc = wb["QC"]
    labels = [c.value for col in qc.iter_rows(min_col=1, max_col=1) for c in col
              if isinstance(c.value, str)]
    joined = " | ".join(labels)
    assert "telescopes" in joined, "missing roll-forward telescoping check"
    assert "CET1 ratio" in joined and "requirement" in joined
    assert "MDA cap" in joined, "missing distribution (MDA) check"
