"""Tests for modelforge.analytics.sensitivity.

Verifies:
    * SensitivityAnalysis sheet is appended to every template
    * ≥ 6 factors per template (PRD US-001 acceptance criterion)
    * `primary_output` named range is registered and points to the
      expected Returns-type sheet
    * A native Excel BarChart is present on the sheet
    * All cell references on the sheet are formulas (no hardcoded
      output snapshots that would go stale on scenario flip)
    * Workbook still passes the 8-check QC gate after the sheet is
      appended
    * Unknown driver names are silently skipped (graceful degradation)
"""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from modelforge.analytics.factors import (
    DEFAULT_FACTORS_BY_TYPE,
    SensitivityFactor,
    default_factors_for,
)
from modelforge.analytics.sensitivity import (
    PrimaryOutputLoc,
    _ELASTICITY_REGISTRY,
    _elasticity_for,
    append_sensitivity_sheet,
)
from modelforge.qc import run_qc
from modelforge.templates import build_model


ROOT = Path(__file__).resolve().parent.parent

CASES = [
    ("unitranche_cdmo.yaml", "unitranche"),
    ("minibond_logistics.yaml", "minibond"),
    ("credit_memo_cdmo.yaml", "credit_memo"),
    ("project_finance_solar.yaml", "project_finance"),
    ("real_estate_pbsa.yaml", "real_estate"),
    ("npl_mixed_portfolio.yaml", "npl"),
    ("structured_credit_pmi.yaml", "structured_credit"),
    ("three_statement_cdmo.yaml", "three_statement"),
]


def _spec_class(model_type: str):
    # Lazy import so missing optional deps don't break collection
    if model_type == "unitranche":
        from modelforge.spec.unitranche import UnitrancheSpec
        return UnitrancheSpec
    if model_type == "minibond":
        from modelforge.spec.minibond import MinibondSpec
        return MinibondSpec
    if model_type == "credit_memo":
        from modelforge.spec.credit_memo import CreditMemoSpec
        return CreditMemoSpec
    if model_type == "project_finance":
        from modelforge.spec.project_finance import ProjectFinanceSpec
        return ProjectFinanceSpec
    if model_type == "real_estate":
        from modelforge.spec.real_estate import RealEstateSpec
        return RealEstateSpec
    if model_type == "npl":
        from modelforge.spec.npl import NPLSpec
        return NPLSpec
    if model_type == "structured_credit":
        from modelforge.spec.structured_credit import StructuredCreditSpec
        return StructuredCreditSpec
    if model_type == "three_statement":
        from modelforge.spec.three_statement import ThreeStatementSpec
        return ThreeStatementSpec
    raise ValueError(model_type)


@pytest.fixture(scope="module", params=CASES, ids=[c[1] for c in CASES])
def built_workbook(tmp_path_factory, request):
    """Build each template once, share across tests in the module."""
    fname, model_type = request.param
    SpecClass = _spec_class(model_type)
    raw = yaml.safe_load((ROOT / "examples" / fname).read_text())
    spec = SpecClass.model_validate(raw)
    out_dir = tmp_path_factory.mktemp("sens")
    xlsx_path = out_dir / f"{model_type}.xlsx"
    build_model(spec, xlsx_path)
    return model_type, xlsx_path, spec


def test_sensitivity_sheet_exists(built_workbook):
    model_type, xlsx_path, _ = built_workbook
    wb = load_workbook(xlsx_path)
    assert "SensitivityAnalysis" in wb.sheetnames, (
        f"SensitivityAnalysis sheet missing for {model_type}"
    )


def test_primary_output_named_range(built_workbook):
    model_type, xlsx_path, _ = built_workbook
    wb = load_workbook(xlsx_path)
    assert "primary_output" in wb.defined_names, (
        f"primary_output named range missing for {model_type}"
    )
    attr = wb.defined_names["primary_output"].attr_text
    assert "!" in attr, f"primary_output attr malformed: {attr}"


def test_factor_count_meets_ac(built_workbook):
    """US-001 AC: '6 factors swept at ±low / ±high'. Accept ≥ 6."""
    model_type, xlsx_path, _ = built_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["SensitivityAnalysis"]
    header_row = 9
    factor_rows = ws.max_row - header_row
    assert factor_rows >= 6, (
        f"{model_type} has only {factor_rows} factors on SensitivityAnalysis "
        f"(AC requires ≥ 6). Extend DEFAULT_FACTORS_BY_TYPE[{model_type!r}]."
    )


def test_tornado_chart_is_native(built_workbook):
    model_type, xlsx_path, _ = built_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["SensitivityAnalysis"]
    assert len(ws._charts) >= 1, f"No chart on SensitivityAnalysis for {model_type}"
    chart = ws._charts[0]
    # BarChart horizontal bars render the tornado look
    assert chart.type == "bar"


def test_factor_values_are_formulas(built_workbook):
    """Low/high values and Δ columns must be formulas (live), not static."""
    model_type, xlsx_path, _ = built_workbook
    wb = load_workbook(xlsx_path)
    ws = wb["SensitivityAnalysis"]
    header_row = 9
    # Check first data row
    for col in (4, 8, 9, 10, 11):  # Base, Low value, High value, Low Δ, High Δ
        v = ws.cell(row=header_row + 1, column=col).value
        assert isinstance(v, str) and v.startswith("="), (
            f"{model_type} SensitivityAnalysis row 10 col {col} is not a "
            f"formula (got {v!r})"
        )


def test_qc_still_passes(built_workbook):
    model_type, xlsx_path, _ = built_workbook
    report = run_qc(xlsx_path)
    assert report.all_pass, (
        f"{model_type} QC regressed after sensitivity append: "
        f"{[c.name for c in report.checks if not c.passed]}"
    )


def test_default_factors_cover_all_model_types():
    """Every model_type in the registry must have a default factor list."""
    from modelforge.templates import REGISTRY
    for mt in REGISTRY:
        factors = default_factors_for(mt)
        assert len(factors) >= 6, (
            f"Default factors for {mt} returns only {len(factors)} "
            f"(AC requires ≥ 6)"
        )


def test_elasticity_registry_has_coverage():
    """Every default factor driver should have an elasticity coefficient."""
    uncovered = []
    for mt, factors in DEFAULT_FACTORS_BY_TYPE.items():
        for f in factors:
            if f.driver_name not in _ELASTICITY_REGISTRY:
                uncovered.append(f"{mt}:{f.driver_name}")
    assert not uncovered, (
        f"{len(uncovered)} driver(s) missing elasticity coefficient:\n"
        f"{uncovered}"
    )


def test_unknown_driver_falls_back_to_default_elasticity():
    assert _elasticity_for("completely_unknown_driver") == 0.5


def test_shocked_values_computation():
    f = SensitivityFactor(
        driver_name="revenue_growth_y1",
        label="Test",
        low_shock=-0.20,
        high_shock=+0.20,
    )
    low, high = f.shocked_values(base=100.0)
    assert low == pytest.approx(80.0)
    assert high == pytest.approx(120.0)


def test_shocked_values_absolute_override():
    f = SensitivityFactor(
        driver_name="euribor_6m_rate",
        label="Test",
        low_shock=-0.50,
        high_shock=+0.50,
        absolute_low=0.02,
        absolute_high=0.05,
    )
    low, high = f.shocked_values(base=0.035)
    assert low == pytest.approx(0.02)
    assert high == pytest.approx(0.05)


def test_graceful_skip_on_unknown_driver(tmp_path):
    """If a provided factor's driver doesn't exist in the spec, it is
    silently skipped rather than raising."""
    from modelforge.spec.unitranche import UnitrancheSpec
    raw = yaml.safe_load((ROOT / "examples" / "unitranche_cdmo.yaml").read_text())
    spec = UnitrancheSpec.model_validate(raw)
    xlsx_path = tmp_path / "skip.xlsx"
    build_model(spec, xlsx_path, with_sensitivity=False)

    custom = [
        SensitivityFactor(driver_name="revenue_growth_y1", label="real",
                          low_shock=-0.1, high_shock=0.1),
        SensitivityFactor(driver_name="nonexistent_driver", label="fake",
                          low_shock=-0.1, high_shock=0.1),
    ]
    result = append_sensitivity_sheet(xlsx_path, spec, factors=custom)
    assert result is not None
    wb = load_workbook(xlsx_path)
    ws = wb["SensitivityAnalysis"]
    # Only 1 factor row present; the fake one was silently dropped
    assert ws.max_row == 10
    assert ws.cell(row=10, column=3).value == "revenue_growth_y1"
