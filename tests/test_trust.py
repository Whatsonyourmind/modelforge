"""Tests for the Trust Layer (modelforge.trust)."""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import pytest

from modelforge.trust import (
    DEFAULT_RULES,
    Severity,
    TrustEngine,
    TrustReport,
    TrustRule,
    TrustViolation,
    inject_red_flag_sheet,
)
from modelforge.trust.builtin import (
    _band_check,
    _dcf_market_cap_deviation,
    _dcf_terminal_growth_below_gdp,
    _dcf_terminal_growth_below_wacc,
    _dcf_wacc_above_rfr,
    _dcf_wacc_in_band,
    _npl_cumulative_recovery_under_100,
    _pf_dscr_above_one,
    _re_ltv_in_band,
    _source_freshness,
)


# ─── helpers ────────────────────────────────────────────────────────────────


class _StubProbe:
    """Minimal WorkbookProbe replacement for unit-testing rules."""
    def __init__(self, values: dict[str, float]) -> None:
        self._v = {k.lower(): v for k, v in values.items()}

    def get(self, key: str):
        return self._v.get(key.lower())

    def has(self, key: str) -> bool:
        return key.lower() in self._v


def _make_rule_stub(name="test_rule", severity: Severity = "warn"):
    rule = MagicMock()
    rule.name = name
    rule.severity = severity
    return rule


def _spec(model_type: str = "dcf", **extra) -> MagicMock:
    s = MagicMock()
    s.model_type = model_type
    for k, v in extra.items():
        setattr(s, k, v)
    return s


# ─── violations / report behavior ───────────────────────────────────────────


def test_violation_severity_rank_orders_correctly():
    fail = TrustViolation("r", "fail", "dcf", "msg")
    warn = TrustViolation("r", "warn", "dcf", "msg")
    info = TrustViolation("r", "info", "dcf", "msg")
    assert fail.severity_rank() > warn.severity_rank() > info.severity_rank()


def test_report_has_failures_only_when_fail_present():
    r = TrustReport(workbook="x", template="dcf")
    assert not r.has_failures()
    r.violations.append(TrustViolation("a", "warn", "dcf", "m"))
    assert not r.has_failures()
    r.violations.append(TrustViolation("b", "fail", "dcf", "m"))
    assert r.has_failures()


def test_report_summary_counts_by_severity():
    r = TrustReport(workbook="x", template="dcf")
    r.violations = [
        TrustViolation("a", "fail", "dcf", "m"),
        TrustViolation("b", "warn", "dcf", "m"),
        TrustViolation("c", "warn", "dcf", "m"),
        TrustViolation("d", "info", "dcf", "m"),
    ]
    s = r.summary()
    assert s["fail"] == 1
    assert s["warn"] == 2
    assert s["info"] == 1


# ─── band_check helper ──────────────────────────────────────────────────────


def test_band_check_inside_returns_true():
    assert _band_check(0.5, 0.0, 1.0)


def test_band_check_outside_low_returns_false():
    assert not _band_check(-0.1, 0.0, 1.0)


def test_band_check_outside_high_returns_false():
    assert not _band_check(1.1, 0.0, 1.0)


def test_band_check_none_returns_true():
    assert _band_check(None, 0.0, 1.0)


# ─── DCF rules ──────────────────────────────────────────────────────────────


def test_wacc_in_band_no_violation_when_normal():
    rule = _make_rule_stub()
    probe = _StubProbe({"wacc_rate": 0.085})
    out = list(_dcf_wacc_in_band(probe, _spec(), rule))
    assert out == []


def test_wacc_in_band_fires_when_too_low():
    rule = _make_rule_stub()
    probe = _StubProbe({"wacc_rate": 0.01})
    out = list(_dcf_wacc_in_band(probe, _spec(), rule))
    assert len(out) == 1
    assert "WACC" in out[0].message


def test_wacc_in_band_fires_when_too_high():
    rule = _make_rule_stub()
    probe = _StubProbe({"wacc_rate": 0.30})
    out = list(_dcf_wacc_in_band(probe, _spec(), rule))
    assert len(out) == 1


def test_wacc_above_rfr_fires_on_zero_beta_bug():
    """Catches the exact Enel bug — beta=0 → WACC=rfr → fail."""
    rule = _make_rule_stub()
    probe = _StubProbe({"wacc_rate": 0.0364, "risk_free_rate": 0.039})
    out = list(_dcf_wacc_above_rfr(probe, _spec(), rule))
    assert len(out) == 1
    assert "below risk-free" in out[0].message


def test_terminal_g_below_wacc_fires_when_close():
    rule = _make_rule_stub()
    # g ≥ WACC − 50bps fires
    probe = _StubProbe({"terminal_growth_pct": 0.07, "wacc_rate": 0.075})
    out = list(_dcf_terminal_growth_below_wacc(probe, _spec(), rule))
    assert len(out) == 1


def test_terminal_g_below_wacc_quiet_when_safe():
    rule = _make_rule_stub()
    probe = _StubProbe({"terminal_growth_pct": 0.02, "wacc_rate": 0.08})
    out = list(_dcf_terminal_growth_below_wacc(probe, _spec(), rule))
    assert out == []


def test_terminal_g_below_gdp_fires_above_3_5pct():
    rule = _make_rule_stub()
    probe = _StubProbe({"terminal_growth_pct": 0.05})
    out = list(_dcf_terminal_growth_below_gdp(probe, _spec(), rule))
    assert len(out) == 1


def test_market_cap_deviation_fires_above_100pct_premium():
    rule = _make_rule_stub()
    probe = _StubProbe({"Valuation!D26": 7.43})  # +743%, the original Enel bug
    out = list(_dcf_market_cap_deviation(probe, _spec(), rule))
    assert len(out) == 1
    assert "+743%" in out[0].message or "743" in out[0].message


# ─── PF / RE / NPL rules ────────────────────────────────────────────────────


def test_pf_dscr_below_one_fires():
    rule = _make_rule_stub()
    probe = _StubProbe({"min_dscr": 0.85})
    out = list(_pf_dscr_above_one(probe, _spec("project_finance"), rule))
    assert len(out) == 1
    assert "below 1.0" in out[0].message


def test_re_ltv_above_85_fires():
    rule = _make_rule_stub()
    probe = _StubProbe({"ltv": 0.95})
    out = list(_re_ltv_in_band(probe, _spec("real_estate"), rule))
    assert len(out) == 1


def test_npl_cumulative_recovery_above_100pct_fires():
    rule = _make_rule_stub()
    probe = _StubProbe({"cumulative_recovery_pct": 1.20})
    out = list(_npl_cumulative_recovery_under_100(probe, _spec("npl"), rule))
    assert len(out) == 1


# ─── Source freshness rule (universal) ──────────────────────────────────────


def test_source_freshness_fires_on_stale_source():
    rule = _make_rule_stub()
    spec = MagicMock()
    spec.model_type = "dcf"
    spec.meta.valuation_date = "2026-04-17"
    src = MagicMock()
    src.id = "S-001"
    src.date = "2024-01-01"  # >1y stale
    spec.sources = [src]
    out = list(_source_freshness(None, spec, rule))
    assert len(out) == 1


def test_source_freshness_quiet_on_fresh_source():
    rule = _make_rule_stub()
    spec = MagicMock()
    spec.model_type = "dcf"
    spec.meta.valuation_date = "2026-04-17"
    src = MagicMock()
    src.id = "S-001"
    src.date = "2026-02-15"
    spec.sources = [src]
    out = list(_source_freshness(None, spec, rule))
    assert out == []


# ─── Engine + applies_to filter ─────────────────────────────────────────────


class _ToyRule(TrustRule):
    name = "toy"
    template_types = ("dcf",)
    severity: Severity = "warn"

    def check(self, probe, spec):
        yield TrustViolation(self.name, self.severity, spec.model_type, "stub")


def test_engine_skips_rules_for_other_templates(tmp_path):
    rule = _ToyRule()
    engine = TrustEngine(rules=[rule])
    spec = _spec("npl")
    # Stub WorkbookProbe never gets opened because rule doesn't apply
    spec_with_path = tmp_path / "fake.xlsx"
    spec_with_path.write_bytes(b"")  # bogus, but engine doesn't read it
    rep = engine.evaluate(spec_with_path, spec)
    assert rep.rules_run == 0
    assert rep.rules_skipped == 1
    assert rep.violations == []


def test_default_rules_bundle_covers_all_templates():
    seen = set()
    for r in DEFAULT_RULES:
        seen.update(r.template_types)
    expected = {
        "dcf", "fairness", "merger",
        "sponsor_lbo", "project_finance", "real_estate",
        "npl", "credit_memo", "unitranche", "minibond",
        "structured_credit", "three_statement", "ipo", "restructuring",
    }
    missing = expected - seen
    assert not missing, f"DEFAULT_RULES missing coverage for: {missing}"


def test_default_rules_have_universal_freshness():
    universal = [r for r in DEFAULT_RULES if not r.template_types]
    assert any(r.name == "source_freshness" for r in universal)


# ─── Red flag sheet injection (round-trip) ─────────────────────────────────


def test_red_flag_sheet_injection_writes_violations(tmp_path):
    """End-to-end: build a tiny xlsx, inject a synthetic report, read it back."""
    from openpyxl import Workbook, load_workbook

    p = tmp_path / "tiny.xlsx"
    wb = Workbook()
    wb.active.title = "Cover"
    wb.active["A1"] = "Hello"
    wb.save(p)

    rep = TrustReport(workbook=str(p), template="dcf")
    rep.rules_run = 3
    rep.violations = [
        TrustViolation(
            rule_name="dcf_test_fail",
            severity="fail",
            template="dcf",
            message="Synthetic FAIL for round-trip test",
            cell="WACCBuild!D18",
            actual=0.01,
            expected_low=0.03,
            expected_high=0.25,
            recommendation="Fix it.",
        ),
        TrustViolation(
            rule_name="dcf_test_warn",
            severity="warn",
            template="dcf",
            message="Synthetic WARN",
        ),
    ]
    inject_red_flag_sheet(p, rep)

    wb2 = load_workbook(p)
    assert "RedFlags" in wb2.sheetnames
    assert wb2.sheetnames[0] == "RedFlags"  # surfaced at the front
    ws = wb2["RedFlags"]
    # Headers row
    assert ws.cell(row=7, column=2).value == "Severity"
    # First violation row
    assert ws.cell(row=8, column=2).value == "FAIL"
    assert ws.cell(row=8, column=3).value == "dcf_test_fail"
    assert ws.cell(row=8, column=4).value == "WACCBuild!D18"


def test_red_flag_sheet_clean_workbook_says_all_clear(tmp_path):
    from openpyxl import Workbook, load_workbook
    p = tmp_path / "clean.xlsx"
    wb = Workbook()
    wb.active.title = "Cover"
    wb.save(p)
    rep = TrustReport(workbook=str(p), template="dcf", rules_run=5)
    inject_red_flag_sheet(p, rep)
    wb2 = load_workbook(p)
    ws = wb2["RedFlags"]
    cell_value = (ws.cell(row=8, column=2).value or "")
    assert "ALL CLEAR" in cell_value or "clear" in cell_value.lower()
