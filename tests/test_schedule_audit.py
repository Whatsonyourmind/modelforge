"""Tests for the schedule auditor (interior-hardcode detection).

Proves the auditor (a) catches a non-innocuous number wedged between formula
cells in a period series — the certify-blind "typed a number over a formula"
defect — including on a REAL built model, and (b) does NOT false-positive on
legitimate edge inputs, innocuous literals, or input/reference/audit sheets.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from modelforge.moat.classifier import classify_sheet
from modelforge.qc import audit_schedule

ROOT = Path(__file__).resolve().parent.parent
_EXCLUDED = {"input", "reference", "audit"}


def _wb(tmp_path: Path, sheet_name: str, row: list) -> Path:
    """Build a one-row workbook; row values placed from column A rightward."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for cidx, val in enumerate(row, start=1):
        if val is not None:
            ws.cell(row=1, column=cidx, value=val)
    p = tmp_path / f"{sheet_name}.xlsx"
    wb.save(p)
    return p


def test_interior_hardcode_flagged(tmp_path):
    # DebtSchedule (core_output). Cols: A=label, B,C=formula, D=hardcode, E,F=formula.
    p = _wb(tmp_path, "DebtSchedule",
            ["Debt closing", "=C1+1", "=D1+1", 12345.0, "=F1+1", "=G1+1"])
    rep = audit_schedule(p)
    assert rep.n_findings == 1
    f = rep.findings[0]
    assert f.value == 12345.0
    assert f.sheet == "DebtSchedule"
    assert f.cell == "D1"  # 4th column
    assert not rep.passed
    assert rep.verdict == "REVIEW"


def test_clean_formula_series_not_flagged(tmp_path):
    p = _wb(tmp_path, "DebtSchedule",
            ["Debt closing", "=C1+1", "=D1+1", "=E1+1", "=F1+1"])
    rep = audit_schedule(p)
    assert rep.passed
    assert rep.n_findings == 0
    assert rep.verdict == "CLEAN"


def test_edge_period_inputs_not_flagged(tmp_path):
    # Leading hardcode (a period-0 opening input) — at the band edge -> clean.
    p_lead = _wb(tmp_path, "DebtSchedule",
                 ["Opening", 5000.0, "=C1+1", "=D1+1", "=E1+1"])
    assert audit_schedule(p_lead).passed
    # Trailing hardcode — also at the edge -> clean.
    p_trail = _wb(tmp_path, "DebtSchedule",
                  ["Series", "=C1+1", "=D1+1", "=E1+1", 9999.0])
    assert audit_schedule(p_trail).passed


def test_innocuous_interior_number_not_flagged(tmp_path):
    # 1 is innocuous (toggles/counters) -> not flagged even when interior.
    p = _wb(tmp_path, "DebtSchedule",
            ["Toggle", "=C1+1", "=D1+1", 1, "=F1+1", "=G1+1"])
    assert audit_schedule(p).passed


def test_input_sheet_excluded(tmp_path):
    # The exact interior-hardcode pattern, but on an input sheet -> excluded.
    p = _wb(tmp_path, "Assumptions",
            ["x", "=C1+1", "=D1+1", 12345.0, "=F1+1", "=G1+1"])
    rep = audit_schedule(p)
    assert rep.passed
    assert "Assumptions" in rep.sheets_skipped


def test_short_series_not_flagged(tmp_path):
    # Fewer than 3 contiguous numeric/formula cells is not a "series".
    p = _wb(tmp_path, "DebtSchedule", ["label", "=C1+1", 777.0])
    assert audit_schedule(p).passed


def _wb_grid(tmp_path: Path, sheet_name: str, rows: list[list]) -> Path:
    """Build a multi-row workbook; rows[r][c] placed at row r+1, col c+1."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    for r, row in enumerate(rows, start=1):
        for c, val in enumerate(row, start=1):
            if val is not None:
                ws.cell(row=r, column=c, value=val)
    p = tmp_path / f"{sheet_name}_grid.xlsx"
    wb.save(p)
    return p


def test_rollforward_correct_not_flagged(tmp_path):
    # Real layout: A=label, B=unit spacer, C/D/E=periods. opening[t] references
    # the closing row at the correct prior period column.
    rows = [
        ["Opening debt", None, "=0", "=C2", "=D2"],
        ["Closing debt", None, "=C1+1", "=D1+1", "=E1+1"],
    ]
    p = _wb_grid(tmp_path, "DebtSchedule", rows)
    rep = audit_schedule(p)
    assert not any(f.kind == "rollforward" for f in rep.findings)


def test_rollforward_wrong_period_flagged(tmp_path):
    # E1 (period 3) wrongly references C2 (period-1 closing) instead of D2.
    rows = [
        ["Opening debt", None, "=0", "=C2", "=C2"],   # last cell -> wrong period
        ["Closing debt", None, "=C1+1", "=D1+1", "=E1+1"],
    ]
    p = _wb_grid(tmp_path, "DebtSchedule", rows)
    rep = audit_schedule(p)
    rf = [f for f in rep.findings if f.kind == "rollforward"]
    assert len(rf) == 1
    assert rf[0].cell == "E1"
    assert "wrong period" in rf[0].detail


def test_rollforward_named_range_opening_not_flagged(tmp_path):
    # opening references a named range (no coordinate ref to closing) -> skip, no FP.
    rows = [
        ["Opening debt", "=0", "=prior_close", "=prior_close"],
        ["Closing debt", "=C1+1", "=D1+1", "=E1+1"],
    ]
    p = _wb_grid(tmp_path, "DebtSchedule", rows)
    rep = audit_schedule(p)
    assert not any(f.kind == "rollforward" for f in rep.findings)


def test_rollforward_cross_sheet_ref_not_flagged(tmp_path):
    # A cross-sheet ref whose row number coincides with the closing row must
    # NOT be treated as a same-sheet roll-forward (no false positive).
    rows = [
        ["Opening debt", "=0", "=OtherSheet!$C$2", "=OtherSheet!$D$2"],
        ["Closing debt", "=C1+1", "=D1+1", "=E1+1"],
    ]
    p = _wb_grid(tmp_path, "DebtSchedule", rows)
    rep = audit_schedule(p)
    assert not any(f.kind == "rollforward" for f in rep.findings)


def test_catches_injected_rollforward_drift_in_real_model(tmp_path):
    """Take a certified model, repoint an opening cell at the WRONG prior period,
    and confirm the auditor flags the roll-forward drift certify cannot see."""
    import yaml
    from modelforge.cli import _load_spec_class
    from modelforge.templates import build_model
    from modelforge.qc.schedule_audit import (
        _OPENING_RE, _CLOSING_RE, _row_label, _period_cols, _refs_to_row,
    )
    from openpyxl.utils import get_column_letter

    yml = ROOT / "examples" / "credit_memo_cdmo.yaml"
    if not yml.exists():
        pytest.skip("example spec not present")
    raw = yaml.safe_load(yml.read_bytes())
    spec = _load_spec_class(raw["model_type"]).model_validate(raw)
    xlsx, _ = build_model(spec, tmp_path / "model.xlsx")
    assert audit_schedule(xlsx).passed  # clean baseline

    wb = openpyxl.load_workbook(xlsx, data_only=False)
    injected = None
    for name in wb.sheetnames:
        if classify_sheet(name) in _EXCLUDED:
            continue
        ws = wb[name]
        openings = [r for r in range(1, ws.max_row + 1) if _OPENING_RE.search(_row_label(ws, r))]
        closings = [r for r in range(1, ws.max_row + 1) if _CLOSING_RE.search(_row_label(ws, r))]
        for r_o in openings:
            r_c = next((rc for rc in closings if r_o < rc <= r_o + 12), None)
            if r_c is None:
                continue
            pcols = _period_cols(ws, r_c)
            # find an opening cell (t>=3) that correctly refs the prior closing,
            # then repoint it two periods back (a genuine wrong-period drift).
            for i in range(2, len(pcols)):
                ocell = ws.cell(row=r_o, column=pcols[i])
                if isinstance(ocell.value, str) and (pcols[i - 1] in _refs_to_row(ocell.value, r_c)):
                    ocell.value = f"=${get_column_letter(pcols[i - 2])}${r_c}"  # wrong period
                    injected = (name, ocell.coordinate)
                    break
            if injected:
                break
        if injected:
            break

    if injected is None:
        pytest.skip("no correct prior-period roll-forward cell found to corrupt")
    wb.save(xlsx)
    rep = audit_schedule(xlsx)
    assert not rep.passed
    assert any(f.kind == "rollforward" and f.sheet == injected[0] and f.cell == injected[1]
               for f in rep.findings)


def test_catches_injected_hardcode_in_real_model(tmp_path):
    """Strongest evidence: a clean certified model becomes a REVIEW the moment a
    number is typed over an interior formula cell — the defect certify misses."""
    import yaml
    from modelforge.cli import _load_spec_class
    from modelforge.templates import build_model

    yml = ROOT / "examples" / "three_statement_cdmo.yaml"
    if not yml.exists():
        pytest.skip("example spec not present")
    raw = yaml.safe_load(yml.read_bytes())
    spec = _load_spec_class(raw["model_type"]).model_validate(raw)
    xlsx, _ = build_model(spec, tmp_path / "model.xlsx")

    # Baseline: the certified model is clean.
    assert audit_schedule(xlsx).passed

    # Inject a hardcode over an interior cell of a contiguous formula run.
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    target = None
    for name in wb.sheetnames:
        if classify_sheet(name) in _EXCLUDED:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            run: list = []
            prev_col = None
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    if prev_col is not None and c.column == prev_col + 1:
                        run.append(c)
                    else:
                        run = [c]
                    prev_col = c.column
                    if len(run) >= 3:
                        target = (name, run[1])  # interior of the run
                        break
                else:
                    run = []
                    prev_col = None
            if target:
                break
        if target:
            break

    assert target is not None, "no contiguous 3+ formula run found to inject into"
    sheet_name, cell = target
    cell.value = 424242.0
    wb.save(xlsx)

    rep = audit_schedule(xlsx)
    assert not rep.passed
    assert any(f.sheet == sheet_name and f.value == 424242.0 for f in rep.findings)
