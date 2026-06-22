"""Tests for the conservation gate (`audit_conservation` + `certify --conservation`).

Proves the gate promotes recomputed in-workbook conservation to a PASS/FAIL
certificate: a clean model passes, and a QC check repointed to recompute 0 (a
valid formula, NOT an Excel error) PASSES plain certify (the defect is
certify-blind) but FAILS `certify --conservation`. A workbook with no QC sheet is
INDETERMINATE (a gate failure under --strict).
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
import yaml
from click.testing import CliRunner

from modelforge.cli import main, _load_spec_class
from modelforge.qc import audit_conservation
from modelforge.qc.conservation_audit import _ALL_PASS_LABELS
from modelforge.templates import build_model

ROOT = Path(__file__).resolve().parent.parent


def _build(tmp_path: Path) -> Path:
    yml = ROOT / "examples" / "three_statement_cdmo.yaml"
    if not yml.exists():
        pytest.skip("example spec not present")
    raw = yaml.safe_load(yml.read_bytes())
    spec = _load_spec_class(raw["model_type"]).model_validate(raw)
    xlsx, _ = build_model(spec, tmp_path / "model.xlsx")
    return xlsx


def _find_qc(wb):
    """Return (sheet, anchor_row, anchor_col) of the ALL CHECKS PASS aggregator."""
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.strip().upper() in _ALL_PASS_LABELS:
                    return ws, c.row, c.column
    return None, None, None


def test_conservation_clean_passes(tmp_path):
    xlsx = _build(tmp_path)
    rep = audit_conservation(xlsx)
    assert rep.recalc_ran is True
    assert rep.status == "PASS", (rep.status, rep.notes, [f.ref for f in rep.findings])
    assert rep.passed is True
    assert rep.n_checks >= 1
    assert rep.all_pass_value is not None

    res = CliRunner().invoke(main, ["certify", str(xlsx), "--conservation"])
    assert res.exit_code == 0, res.output
    assert "CONSERVATION PASS" in res.output


def test_conservation_gate_catches_certify_blind_failing_check(tmp_path):
    xlsx = _build(tmp_path)

    # Repoint the first per-check result cell to recompute 0. This is a perfectly
    # valid formula (no #REF!), so certify is blind to it; ALL_PASS then drops.
    wb = openpyxl.load_workbook(xlsx, data_only=False)
    ws, anchor_row, anchor_col = _find_qc(wb)
    assert ws is not None, "QC sheet not found"
    result_col = anchor_col + 2
    injected_ref = None
    for r in range(anchor_row + 1, ws.max_row + 1):
        label = ws.cell(row=r, column=anchor_col).value
        formula = ws.cell(row=r, column=result_col).value
        if isinstance(label, str) and label.strip() and isinstance(formula, str) and formula.startswith("="):
            ws.cell(row=r, column=result_col).value = "=0"
            injected_ref = f"{ws.title}!{ws.cell(row=r, column=result_col).coordinate}"
            break
    assert injected_ref is not None, "no per-check formula cell found to break"
    wb.save(xlsx)

    # The auditor catches it: that check (and ALL_PASS) recompute to 0.
    rep = audit_conservation(xlsx)
    assert rep.status == "FAIL", (rep.status, rep.notes)
    assert rep.passed is False
    assert any(f.ref == injected_ref for f in rep.findings)

    runner = CliRunner()
    # Plain certify is blind to it (still zero formula errors) -> exit 0.
    plain = runner.invoke(main, ["certify", str(xlsx)])
    assert plain.exit_code == 0, plain.output
    # The opt-in conservation gate catches it -> exit 1.
    gated = runner.invoke(main, ["certify", str(xlsx), "--conservation"])
    assert gated.exit_code == 1, gated.output
    assert "CONSERVATION FAIL" in gated.output


def test_conservation_indeterminate_without_qc(tmp_path):
    # A workbook with no QC sheet -> INDETERMINATE (a gate failure), never a
    # silent pass.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Model"
    ws["A1"] = 2
    ws["A2"] = 3
    ws["A3"] = "=A1+A2"
    p = tmp_path / "no_qc.xlsx"
    wb.save(p)

    rep = audit_conservation(p)
    assert rep.status == "INDETERMINATE"
    assert rep.passed is False
    assert rep.qc_sheet is None
