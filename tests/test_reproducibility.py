"""Tests for modelforge.analytics.reproducibility (US-011)."""

from __future__ import annotations

from pathlib import Path

import pytest
import yaml
from openpyxl import load_workbook

from modelforge.analytics.reproducibility import (
    NAME_PYTHON,
    NAME_SHA,
    NAME_TIMESTAMP,
    NAME_VERSION,
    SHEET_NAME,
    compute_spec_hash,
    read_reproducibility,
    verify_spec_hash,
)
from modelforge.qc import run_qc
from modelforge.templates import build_model

ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture(scope="module")
def built(tmp_path_factory):
    from modelforge.spec.unitranche import UnitrancheSpec
    spec_path = ROOT / "examples" / "unitranche_cdmo.yaml"
    raw_bytes = spec_path.read_bytes()
    spec = UnitrancheSpec.model_validate(yaml.safe_load(raw_bytes))
    out = tmp_path_factory.mktemp("repro") / "u.xlsx"
    build_model(spec, out, spec_source_bytes=raw_bytes, spec_source_path=spec_path)
    return out, raw_bytes


def test_sheet_appended(built):
    xlsx, _ = built
    wb = load_workbook(xlsx)
    assert SHEET_NAME in wb.sheetnames


def test_named_ranges_registered(built):
    xlsx, _ = built
    wb = load_workbook(xlsx)
    for name in (NAME_SHA, NAME_VERSION, NAME_PYTHON, NAME_TIMESTAMP):
        assert name in wb.defined_names, f"Missing named range: {name}"


def test_hash_deterministic(built):
    """Same bytes → same hash, regardless of platform / run."""
    _, raw_bytes = built
    sha1, _ = compute_spec_hash(None, raw_bytes)  # spec unused when bytes given
    sha2, _ = compute_spec_hash(None, raw_bytes)
    assert sha1 == sha2
    assert len(sha1) == 64  # SHA-256 hex


def test_verify_matches_original_spec(built):
    xlsx, raw_bytes = built
    match, stored, recomputed = verify_spec_hash(xlsx, raw_bytes)
    assert match, f"Stored {stored} != recomputed {recomputed}"


def test_verify_detects_tamper(built):
    xlsx, raw_bytes = built
    tampered = raw_bytes.replace(b"eur", b"USD", 1)
    match, stored, recomputed = verify_spec_hash(xlsx, tampered)
    assert not match
    assert stored != recomputed


def test_metadata_readable(built):
    xlsx, _ = built
    meta = read_reproducibility(xlsx)
    assert meta.get("Spec SHA-256") and len(meta["Spec SHA-256"]) == 64
    assert meta.get("ModelForge version")
    assert meta.get("Build timestamp (UTC)", "").startswith("20")


def test_qc_still_passes(built):
    xlsx, _ = built
    report = run_qc(xlsx)
    assert report.all_pass, [c.name for c in report.checks if not c.passed]


def test_canonical_json_fallback_when_no_bytes():
    from modelforge.spec.unitranche import UnitrancheSpec
    spec_path = ROOT / "examples" / "unitranche_cdmo.yaml"
    spec = UnitrancheSpec.model_validate(yaml.safe_load(spec_path.read_bytes()))
    sha, descriptor = compute_spec_hash(spec)
    assert len(sha) == 64
    assert descriptor == "canonical_json"
