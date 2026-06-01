"""Tests for D6: comment threads + 4-eyes review state machine + workbook diff."""
from __future__ import annotations

from pathlib import Path

import pytest

from modelforge.collab import (
    ApprovalStatus,
    Comment,
    CommentStore,
    InvalidStateTransition,
    ReviewerRole,
    ReviewState,
    diff_workbooks,
    reject,
    request_review,
    sign_off,
)


# ── CommentStore ──────────────────────────────────────────────────────────


def test_create_comment_returns_comment_with_thread_eq_id(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c = store.create(anchor="wacc_rate", author="md@example.com", body="Why 8.5%?")
    assert isinstance(c, Comment)
    assert c.id == c.thread
    assert c.anchor == "wacc_rate"
    assert c.event == "create"
    assert c.resolved is False


def test_reply_inherits_anchor_from_thread_root(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c1 = store.create(anchor="wacc_rate", author="md@example.com", body="Q?")
    c2 = store.reply(thread=c1.id, author="jr@example.com", body="A.")
    assert c2.thread == c1.id
    assert c2.anchor == "wacc_rate"
    assert c2.event == "reply"


def test_reply_to_unknown_thread_raises(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    with pytest.raises(KeyError):
        store.reply(thread="nope", author="x", body="y")


def test_resolve_marks_event_resolved_true(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c1 = store.create(anchor="x", author="md@x", body="Q?")
    store.reply(thread=c1.id, author="jr@x", body="A")
    res = store.resolve(thread=c1.id, author="md@x", note="ok")
    assert res.event == "resolve"
    assert res.resolved is True


def test_open_threads_excludes_resolved(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c1 = store.create(anchor="a", author="md@x", body="Q1")
    c2 = store.create(anchor="b", author="md@x", body="Q2")
    store.resolve(thread=c1.id, author="md@x")
    open_ = store.open_threads()
    assert len(open_) == 1
    assert open_[0].id == c2.id


def test_open_threads_includes_reopened(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c1 = store.create(anchor="a", author="md@x", body="Q1")
    store.resolve(thread=c1.id, author="md@x")
    store.reopen(thread=c1.id, author="md@x", note="actually no")
    open_ = store.open_threads()
    assert any(t.id == c1.id for t in open_)


def test_threads_for_anchor_returns_only_roots(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c1 = store.create(anchor="wacc_rate", author="a", body="Q1")
    store.reply(thread=c1.id, author="b", body="A1")
    c2 = store.create(anchor="wacc_rate", author="a", body="Q2")
    threads = store.threads_for_anchor("wacc_rate")
    assert {t.id for t in threads} == {c1.id, c2.id}
    assert all(t.event == "create" for t in threads)


def test_events_for_thread_orders_chronologically(tmp_path):
    store = CommentStore(tmp_path / "c.db")
    c1 = store.create(anchor="x", author="a", body="0")
    store.reply(thread=c1.id, author="b", body="1")
    store.reply(thread=c1.id, author="c", body="2")
    events = store.events_for_thread(c1.id)
    assert [e.body for e in events] == ["0", "1", "2"]


# ── Review state machine ──────────────────────────────────────────────────


def _new_state(author="alice", workbook="deal.xlsx") -> ReviewState:
    return ReviewState(workbook=workbook, author=author)


def test_initial_state_is_draft():
    s = _new_state()
    assert s.status == ApprovalStatus.DRAFT


def test_request_review_promotes_to_under_review():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    assert s.status == ApprovalStatus.UNDER_REVIEW
    assert s.required_reviewers == ["bob"]
    assert s.events[-1].action == "request_review"


def test_request_review_rejects_author_as_reviewer():
    s = _new_state(author="alice")
    with pytest.raises(InvalidStateTransition):
        request_review(s, reviewers=["alice", "bob"], actor="alice")


def test_request_review_requires_reviewers_list():
    s = _new_state()
    with pytest.raises(InvalidStateTransition):
        request_review(s, reviewers=[], actor="alice")


def test_sign_off_with_single_reviewer_approves():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    sign_off(s, reviewer="bob")
    assert s.status == ApprovalStatus.APPROVED
    assert s.signed_off_by == ["bob"]


def test_sign_off_with_two_reviewers_needs_both():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob", "carol"], actor="alice")
    sign_off(s, reviewer="bob")
    assert s.status == ApprovalStatus.UNDER_REVIEW
    sign_off(s, reviewer="carol")
    assert s.status == ApprovalStatus.APPROVED


def test_sign_off_rejects_author():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    with pytest.raises(InvalidStateTransition):
        sign_off(s, reviewer="alice")


def test_sign_off_rejects_unknown_reviewer():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    with pytest.raises(InvalidStateTransition):
        sign_off(s, reviewer="dave")


def test_sign_off_rejects_double_sign():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob", "carol"], actor="alice")
    sign_off(s, reviewer="bob")
    with pytest.raises(InvalidStateTransition):
        sign_off(s, reviewer="bob")


def test_reject_sends_back_to_rejected():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    reject(s, reviewer="bob", reason="WACC too low")
    assert s.status == ApprovalStatus.REJECTED
    assert s.rejected_by == "bob"
    assert s.rejection_note == "WACC too low"


def test_reject_then_re_request_review_loops():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    reject(s, reviewer="bob", reason="WACC too low")
    request_review(s, reviewers=["bob"], actor="alice")
    assert s.status == ApprovalStatus.UNDER_REVIEW
    assert s.signed_off_by == []


def test_reject_rejects_author():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    with pytest.raises(InvalidStateTransition):
        reject(s, reviewer="alice", reason="cheat")


def test_state_serializes_and_deserializes_via_json():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    sign_off(s, reviewer="bob")
    blob = s.to_json()
    s2 = ReviewState.from_json(blob)
    assert s2.status == ApprovalStatus.APPROVED
    assert s2.signed_off_by == ["bob"]
    assert len(s2.events) == len(s.events)


def test_status_change_logged_on_final_sign_off():
    s = _new_state(author="alice")
    request_review(s, reviewers=["bob"], actor="alice")
    sign_off(s, reviewer="bob")
    actions = [e.action for e in s.events]
    assert "status_change" in actions


# ── Workbook diff ─────────────────────────────────────────────────────────


def _build_workbook(tmp_path) -> Path:
    """Build a real workbook from the canonical Enel example."""
    import yaml
    from modelforge.spec.dcf import DCFSpec
    from modelforge.templates import build_model
    spec = DCFSpec.model_validate(
        yaml.safe_load(Path("examples/dcf_enel.yaml").read_text(encoding="utf-8"))
    )
    out = tmp_path / "v.xlsx"
    build_model(spec, out, with_manifest=False)
    return out


def test_diff_identical_workbooks_has_no_cell_changes(tmp_path):
    a = _build_workbook(tmp_path)
    b = tmp_path / "v_copy.xlsx"
    b.write_bytes(a.read_bytes())
    d = diff_workbooks(a, b, ignore_sheets={"Reproducibility"})
    # Bytes-identical workbooks should diff to zero cell-level changes
    assert d.cell_changes == []
    assert d.sheets_added == []
    assert d.sheets_removed == []


def test_diff_detects_added_sheet(tmp_path):
    from openpyxl import load_workbook
    a = _build_workbook(tmp_path)
    b_path = tmp_path / "v_extra.xlsx"
    b_path.write_bytes(a.read_bytes())
    wb = load_workbook(b_path)
    wb.create_sheet("ExtraSheet")
    wb.save(b_path)
    d = diff_workbooks(a, b_path, ignore_sheets={"Reproducibility"})
    assert "ExtraSheet" in d.sheets_added
    assert d.sheets_removed == []


def test_diff_summary_string_contains_counts(tmp_path):
    a = _build_workbook(tmp_path)
    b = tmp_path / "v.xlsx"
    b.write_bytes(a.read_bytes())
    d = diff_workbooks(a, b, ignore_sheets={"Reproducibility"})
    assert "changed=" in d.summary()
    assert "added_cells=" in d.summary()
