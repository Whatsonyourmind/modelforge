"""Regression: debt-schedule opening balance must roll forward from the prior
period's CLOSING balance — never the repayment row — under every amortization
profile, including ``mandatory_1pct``.

``mandatory_1pct`` inserts an extra "Mandatory amortization rate" input row
between Drawdown and Scheduled amortization, shifting Closing down by one. A
historical off-by-one baked a literal cursor offset (``r+3``) into the opening
formula, so once that input row appeared the opening balance pointed at the
REPAYMENT row instead of CLOSING — silently breaking the debt roll-forward
(interest understated ~84%, propagating into tax/NI/FCF/Returns/covenants).

The mis-wire was certify-INVISIBLE: it is a valid in-range reference, so
formula-integrity certification passes. A structural assertion on the emitted
formula is therefore the correct guard. This is the test that would have caught
the bug; it also covers the two other amortization profiles as controls.
"""
from __future__ import annotations

import re
from pathlib import Path

import openpyxl
import pytest
import yaml

from modelforge.spec.unitranche import UnitrancheSpec
from modelforge.templates import build_model

_REF_RE = re.compile(r"=\$[A-Z]+\$(\d+)$")
_SPEC = Path("examples/unitranche_cdmo.yaml")


def _build(amortization: str, out_dir: Path) -> Path:
    raw = yaml.safe_load(_SPEC.read_bytes())
    raw["debt"]["tranches"][0]["amortization"] = amortization
    spec = UnitrancheSpec.model_validate(raw)
    out = out_dir / f"unitranche_{amortization}.xlsx"
    build_model(spec, out, with_manifest=False)
    return out


def _first_tranche_rows(ws) -> dict[str, int]:
    """Locate the first tranche's Opening / Scheduled-amortization / Closing rows."""
    rows: dict[str, int] = {}
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if not isinstance(cell.value, str):
            continue
        label = cell.value.strip()
        if label == "Opening debt" and "opening" not in rows:
            rows["opening"] = cell.row
        elif label == "Scheduled amortization" and "repayment" not in rows:
            rows["repayment"] = cell.row
        elif label == "Closing debt" and "closing" not in rows:
            rows["closing"] = cell.row
        if {"opening", "repayment", "closing"} <= rows.keys():
            break
    return rows


def _opening_rollforward_target_rows(ws, opening_row: int) -> set[int]:
    """Row numbers referenced by the opening-balance roll-forward formulas."""
    targets: set[int] = set()
    for cell in ws[opening_row]:
        if isinstance(cell.value, str):
            m = _REF_RE.match(cell.value.strip())
            if m:
                targets.add(int(m.group(1)))
    return targets


@pytest.mark.parametrize("amortization", ["mandatory_1pct", "bullet", "linear"])
def test_debt_opening_rolls_from_closing_not_repayment(tmp_path, amortization):
    out = _build(amortization, tmp_path)
    ws = openpyxl.load_workbook(out, data_only=False)["DebtSchedule"]

    rows = _first_tranche_rows(ws)
    assert {"opening", "repayment", "closing"} <= rows.keys(), (
        f"could not locate first-tranche rows for {amortization}: {rows}"
    )

    targets = _opening_rollforward_target_rows(ws, rows["opening"])
    assert targets, f"no opening-balance roll-forward formulas found ({amortization})"

    # Every opening roll-forward must point at the CLOSING row...
    assert targets == {rows["closing"]}, (
        f"opening[{amortization}] references rows {sorted(targets)}, "
        f"expected only closing row {rows['closing']} "
        f"(repayment row is {rows['repayment']})"
    )
    # ...and must NOT point at the repayment row (the historical off-by-one).
    assert rows["repayment"] not in targets, (
        f"opening[{amortization}] references the REPAYMENT row "
        f"{rows['repayment']} — the debt roll-forward off-by-one has regressed"
    )


def test_mandatory_1pct_inserts_rate_row_above_repayment(tmp_path):
    """Guards the precondition: mandatory_1pct really does shift Closing down by
    one (vs bullet), which is what makes the literal-offset form dangerous."""
    bullet = openpyxl.load_workbook(_build("bullet", tmp_path),
                                    data_only=False)["DebtSchedule"]
    mand = openpyxl.load_workbook(_build("mandatory_1pct", tmp_path),
                                  data_only=False)["DebtSchedule"]
    b, m = _first_tranche_rows(bullet), _first_tranche_rows(mand)
    # Same opening row, but the inserted rate row pushes closing one lower.
    assert m["closing"] == b["closing"] + 1, (
        f"expected mandatory_1pct closing ({m['closing']}) one below bullet "
        f"({b['closing']}) due to the inserted rate row"
    )
