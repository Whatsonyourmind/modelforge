"""Quick formula/cell census for a built ModelForge workbook."""
import sys
from pathlib import Path
from openpyxl import load_workbook

path = Path(sys.argv[1])
wb = load_workbook(path, data_only=False)
print(f"Workbook: {path}")
print(f"Sheets: {len(wb.sheetnames)}")
total_f, total_c, total_cells = 0, 0, 0
for s in wb.sheetnames:
    ws = wb[s]
    formulas = constants = 0
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            if isinstance(c.value, str) and c.value.startswith("="):
                formulas += 1
            else:
                constants += 1
    cells = formulas + constants
    total_f += formulas
    total_c += constants
    total_cells += cells
    print(f"  {s:<22}  cells={cells:>5}  formulas={formulas:>4}  hardcodes={constants:>4}")
print(f"TOTAL cells: {total_cells}  formulas: {total_f}  hardcodes: {total_c}")
if total_cells:
    print(f"Live-formula ratio: {total_f/total_cells*100:.1f}%")
print(f"Named ranges: {len(wb.defined_names)}")
print(f"File size: {path.stat().st_size:,} bytes")
